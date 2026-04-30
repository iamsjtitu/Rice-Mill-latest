from models import round_amount
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import io
from typing import Optional
from datetime import datetime, timezone
from database import db
from pydantic import BaseModel, Field, ConfigDict
from utils.date_format import fmt_date
import uuid

router = APIRouter()

# ============ MODELS ============

class StaffMember(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    salary_type: str  # "weekly" or "monthly"
    salary_amount: float  # per day for weekly, per month for monthly
    active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class AttendanceEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    staff_id: str
    staff_name: str = ""
    date: str
    status: str  # "present", "absent", "half_day", "holiday"
    kms_year: str = ""
    season: str = ""

class StaffAdvance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    staff_id: str
    staff_name: str = ""
    amount: float
    date: str
    description: str = ""
    kms_year: str = ""
    season: str = ""
    account: str = "cash"     # cash / bank / owner
    bank_name: str = ""
    owner_name: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class StaffPayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    staff_id: str
    staff_name: str = ""
    salary_type: str = ""
    salary_amount: float = 0
    period_from: str = ""
    period_to: str = ""
    total_days: int = 0
    days_worked: float = 0
    holidays: float = 0
    half_days: float = 0
    absents: int = 0
    gross_salary: float = 0
    advance_balance: float = 0
    advance_deducted: float = 0
    net_payment: float = 0
    date: str = ""
    kms_year: str = ""
    season: str = ""
    round_off: float = 0
    account: str = "cash"     # cash / bank / owner
    bank_name: str = ""
    owner_name: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


# ============ STAFF CRUD ============

@router.get("/staff")
async def get_staff(active_only: bool = True):
    query = {"active": True} if active_only else {}
    staff = await db.staff.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return staff

@router.post("/staff")
async def add_staff(s: StaffMember):
    doc = s.model_dump()
    await db.staff.insert_one(doc)
    doc.pop("_id", None)
    return doc

@router.put("/staff/{staff_id}")
async def update_staff(staff_id: str, data: dict):
    existing = await db.staff.find_one({"id": staff_id})
    if not existing:
        raise HTTPException(404, "Staff not found")
    data.pop("id", None)
    data.pop("_id", None)
    await db.staff.update_one({"id": staff_id}, {"$set": data})
    updated = await db.staff.find_one({"id": staff_id}, {"_id": 0})
    return updated

@router.delete("/staff/{staff_id}")
async def delete_staff(staff_id: str):
    await db.staff.update_one({"id": staff_id}, {"$set": {"active": False}})
    return {"message": "Staff deactivated"}


# ============ ATTENDANCE ============

@router.get("/staff/attendance")
async def get_attendance(staff_id: Optional[str] = None, date: Optional[str] = None,
                         date_from: Optional[str] = None, date_to: Optional[str] = None,
                         kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if staff_id: query["staff_id"] = staff_id
    if date: query["date"] = date
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        query["date"] = date_q
    return await db.staff_attendance.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(10000)

@router.post("/staff/attendance")
async def mark_attendance(entry: AttendanceEntry):
    # Upsert - one entry per staff per date
    existing = await db.staff_attendance.find_one({"staff_id": entry.staff_id, "date": entry.date})
    doc = entry.model_dump()
    if existing:
        await db.staff_attendance.update_one(
            {"staff_id": entry.staff_id, "date": entry.date},
            {"$set": {"status": doc["status"], "kms_year": doc["kms_year"], "season": doc["season"]}}
        )
    else:
        await db.staff_attendance.insert_one(doc)
    doc.pop("_id", None)
    return doc

@router.post("/staff/attendance/bulk")
async def bulk_mark_attendance(data: dict):
    date = data.get("date", "")
    records = data.get("records", [])
    kms_year = data.get("kms_year", "")
    season = data.get("season", "")
    for r in records:
        existing = await db.staff_attendance.find_one({"staff_id": r["staff_id"], "date": date})
        if existing:
            await db.staff_attendance.update_one(
                {"staff_id": r["staff_id"], "date": date},
                {"$set": {"status": r["status"], "kms_year": kms_year, "season": season}}
            )
        else:
            doc = {
                "id": str(uuid.uuid4()), "staff_id": r["staff_id"],
                "staff_name": r.get("staff_name", ""), "date": date,
                "status": r["status"], "kms_year": kms_year, "season": season
            }
            await db.staff_attendance.insert_one(doc)
    return {"message": f"{len(records)} attendance records saved"}


# ============ ADVANCE ============

@router.get("/staff/advance")
async def get_advances(staff_id: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if staff_id: query["staff_id"] = staff_id
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    return await db.staff_advance.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)

@router.post("/staff/advance")
async def add_advance(adv: StaffAdvance, username: str = ""):
    doc = adv.model_dump()
    await db.staff_advance.insert_one(doc)
    doc.pop("_id", None)
    # Auto-create Cash Book Nikasi entry
    if doc["amount"] > 0:
        staff_name = doc.get("staff_name", "Staff")
        pay_account = (doc.get("account") or "cash").lower()
        if pay_account not in ("cash", "bank", "owner"):
            pay_account = "cash"
        cb_entry = {
            "id": str(uuid.uuid4()),
            "date": doc["date"],
            "account": pay_account,
            "bank_name": doc.get("bank_name", "") if pay_account == "bank" else "",
            "owner_name": doc.get("owner_name", "") if pay_account == "owner" else "",
            "txn_type": "nikasi",
            "category": staff_name,
            "party_type": "Staff",
            "description": f"Staff Advance: {staff_name} - {doc.get('description', '')}",
            "amount": round_amount(doc["amount"]),
            "reference": f"staff_advance:{doc['id']}",
            "kms_year": doc.get("kms_year", ""),
            "season": doc.get("season", ""),
            "created_by": username or "system",
            "linked_payment_id": doc["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(cb_entry)
        cb_entry.pop("_id", None)
        # Auto-create Ledger Jama entry (staff owes us the advance)
        ledger_entry = {
            "id": str(uuid.uuid4()),
            "date": doc["date"],
            "account": "ledger",
            "txn_type": "jama",
            "category": staff_name,
            "party_type": "Staff",
            "description": f"Staff Advance: {staff_name} - {doc.get('description', '')}",
            "amount": round_amount(doc["amount"]),
            "reference": f"staff_advance_ledger:{doc['id']}",
            "kms_year": doc.get("kms_year", ""),
            "season": doc.get("season", ""),
            "created_by": username or "system",
            "linked_payment_id": doc["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(ledger_entry)
        ledger_entry.pop("_id", None)
    return doc

@router.delete("/staff/advance/{adv_id}")
async def delete_advance(adv_id: str):
    result = await db.staff_advance.delete_one({"id": adv_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Advance not found")
    # Delete linked cash book entry and ledger entry
    await db.cash_transactions.delete_many({"linked_payment_id": adv_id})
    await db.cash_transactions.delete_many({"reference": f"staff_advance:{adv_id}"})
    await db.cash_transactions.delete_many({"reference": f"staff_advance_ledger:{adv_id}"})
    return {"message": "Advance deleted"}

@router.get("/staff/advance-balance/{staff_id}")
async def get_advance_balance(staff_id: str, kms_year: Optional[str] = None, season: Optional[str] = None):
    # Total advances given in current FY
    q = {"staff_id": staff_id}
    if kms_year: q["kms_year"] = kms_year
    if season: q["season"] = season
    advances = await db.staff_advance.find(q, {"_id": 0}).to_list(5000)
    total_advance = sum(a.get("amount", 0) for a in advances)

    # Total advance deducted from payments in current FY
    pq = {"staff_id": staff_id}
    if kms_year: pq["kms_year"] = kms_year
    if season: pq["season"] = season
    payments = await db.staff_payments.find(pq, {"_id": 0}).to_list(5000)
    total_deducted = sum(p.get("advance_deducted", 0) for p in payments)

    # Compute opening balance from previous FY
    opening_balance = 0.0
    if kms_year:
        parts = kms_year.split('-')
        if len(parts) == 2:
            try:
                prev_fy = f"{int(parts[0])-1}-{int(parts[1])-1}"
                prev_adv_q = {"staff_id": staff_id, "kms_year": prev_fy}
                if season: prev_adv_q["season"] = season
                prev_advances = await db.staff_advance.find(prev_adv_q, {"_id": 0}).to_list(5000)
                prev_total_adv = sum(a.get("amount", 0) for a in prev_advances)
                prev_pay_q = {"staff_id": staff_id, "kms_year": prev_fy}
                if season: prev_pay_q["season"] = season
                prev_payments = await db.staff_payments.find(prev_pay_q, {"_id": 0}).to_list(5000)
                prev_total_ded = sum(p.get("advance_deducted", 0) for p in prev_payments)
                opening_balance = round(prev_total_adv - prev_total_ded, 2)
            except (ValueError, IndexError):
                pass

    return {"opening_balance": opening_balance, "total_advance": round(total_advance, 2), "total_deducted": round(total_deducted, 2),
            "balance": round(opening_balance + total_advance - total_deducted, 2)}


# ============ SALARY CALCULATION ============

@router.get("/staff/salary-calculate")
async def calculate_salary(staff_id: str, period_from: str, period_to: str,
                           kms_year: Optional[str] = None, season: Optional[str] = None):
    staff = await db.staff.find_one({"id": staff_id}, {"_id": 0})
    if not staff:
        raise HTTPException(404, "Staff not found")

    # Get attendance for period
    att_q = {"staff_id": staff_id, "date": {"$gte": period_from, "$lte": period_to}}
    attendance = await db.staff_attendance.find(att_q, {"_id": 0}).to_list(1000)

    present_days = sum(1 for a in attendance if a.get("status") == "present")
    half_days = sum(1 for a in attendance if a.get("status") == "half_day")
    holidays = sum(1 for a in attendance if a.get("status") == "holiday")
    absents = sum(1 for a in attendance if a.get("status") == "absent")
    days_worked = present_days + (half_days * 0.5) + holidays  # Holiday = paid leave

    # Calculate total days in period
    from datetime import datetime as dt
    d1 = dt.strptime(period_from, "%Y-%m-%d")
    d2 = dt.strptime(period_to, "%Y-%m-%d")
    total_days = (d2 - d1).days + 1

    # Calculate salary
    if staff["salary_type"] == "weekly":
        # Per day rate
        per_day = staff["salary_amount"]
        gross_salary = round(days_worked * per_day, 2)
    else:
        # Monthly: always /30
        per_day = staff["salary_amount"] / 30
        gross_salary = round(days_worked * per_day, 2)

    # Get advance balance (with FY carry-forward)
    adv_q = {"staff_id": staff_id}
    if kms_year: adv_q["kms_year"] = kms_year
    if season: adv_q["season"] = season
    advances = await db.staff_advance.find(adv_q, {"_id": 0}).to_list(5000)
    total_advance = sum(a.get("amount", 0) for a in advances)
    pq = {"staff_id": staff_id}
    if kms_year: pq["kms_year"] = kms_year
    if season: pq["season"] = season
    payments = await db.staff_payments.find(pq, {"_id": 0}).to_list(5000)
    total_deducted = sum(p.get("advance_deducted", 0) for p in payments)
    
    # Opening advance balance from previous FY
    adv_opening = 0.0
    if kms_year:
        fy_parts = kms_year.split('-')
        if len(fy_parts) == 2:
            try:
                prev_fy = f"{int(fy_parts[0])-1}-{int(fy_parts[1])-1}"
                prev_adv_q = {"staff_id": staff_id, "kms_year": prev_fy}
                if season: prev_adv_q["season"] = season
                prev_adv = await db.staff_advance.find(prev_adv_q, {"_id": 0}).to_list(5000)
                prev_total_adv = sum(a.get("amount", 0) for a in prev_adv)
                prev_pay_q = {"staff_id": staff_id, "kms_year": prev_fy}
                if season: prev_pay_q["season"] = season
                prev_pay = await db.staff_payments.find(prev_pay_q, {"_id": 0}).to_list(5000)
                prev_total_ded = sum(p.get("advance_deducted", 0) for p in prev_pay)
                adv_opening = round(prev_total_adv - prev_total_ded, 2)
            except (ValueError, IndexError):
                pass
    advance_balance = round(adv_opening + total_advance - total_deducted, 2)

    return {
        "staff": staff,
        "period_from": period_from, "period_to": period_to,
        "total_days": total_days,
        "present_days": present_days, "half_days": half_days,
        "holidays": holidays, "absents": absents,
        "days_worked": days_worked,
        "per_day_rate": round(per_day, 2),
        "gross_salary": gross_salary,
        "advance_balance": advance_balance,
        "attendance_details": sorted(attendance, key=lambda x: x.get("date", ""))
    }


# ============ PAYMENT (SETTLE SALARY) ============

@router.get("/staff/payments")
async def get_payments(staff_id: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if staff_id: query["staff_id"] = staff_id
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    return await db.staff_payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

@router.post("/staff/payments")
async def settle_salary(pay: StaffPayment):
    doc = pay.model_dump()
    doc["net_payment"] = round(doc["gross_salary"] - doc["advance_deducted"], 2)
    await db.staff_payments.insert_one(doc)
    doc.pop("_id", None)

    # Auto-create Cash Book Nikasi entry
    if doc["net_payment"] > 0:
        pay_account = (doc.get("account") or "cash").lower()
        if pay_account not in ("cash", "bank", "owner"):
            pay_account = "cash"
        cb_entry = {
            "id": str(uuid.uuid4()),
            "date": doc["date"],
            "account": pay_account,
            "bank_name": doc.get("bank_name", "") if pay_account == "bank" else "",
            "owner_name": doc.get("owner_name", "") if pay_account == "owner" else "",
            "txn_type": "nikasi",
            "category": "Staff Salary",
            "description": f"Salary: {doc['staff_name']} ({doc['period_from']} to {doc['period_to']})",
            "amount": round_amount(doc["net_payment"]),
            "reference": f"staff_payment:{doc['id']}",
            "kms_year": doc.get("kms_year", ""),
            "season": doc.get("season", ""),
            "created_by": "system",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(cb_entry)
        cb_entry.pop("_id", None)
        doc["cash_book_entry"] = cb_entry

    return doc

@router.delete("/staff/payments/{payment_id}")
async def delete_payment(payment_id: str, username: str = "", role: str = ""):
    from services.edit_lock import check_edit_lock
    payment = await db.staff_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(404, "Payment not found")
    can_edit, message = await check_edit_lock(payment, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    # Delete cash book entry too
    await db.cash_transactions.delete_one({"reference": f"staff_payment:{payment_id}"})
    await db.staff_payments.delete_one({"id": payment_id})
    return {"message": "Payment deleted and cash book entry removed"}



# ============ EXPORT: ATTENDANCE REPORT ============

@router.get("/staff/export/attendance")
async def export_attendance(date_from: str, date_to: str, fmt: str = "excel",
                            kms_year: Optional[str] = None, season: Optional[str] = None):
    staff_list = await db.staff.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(500)
    att_q = {"date": {"$gte": date_from, "$lte": date_to}}
    attendance = await db.staff_attendance.find(att_q, {"_id": 0}).to_list(10000)

    # Build date range
    from datetime import datetime as dt, timedelta
    d1 = dt.strptime(date_from, "%Y-%m-%d")
    d2 = dt.strptime(date_to, "%Y-%m-%d")
    dates = []
    cur = d1
    while cur <= d2:
        dates.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)

    # Map: staff_id -> date -> status
    att_map = {}
    for a in attendance:
        att_map.setdefault(a["staff_id"], {})[a["date"]] = a["status"]

    status_short = {"present": "P", "absent": "A", "half_day": "H", "holiday": "CH"}

    # Build monthly summary data
    from collections import defaultdict
    monthly_data = {}  # {month_key: {staff_id: {P:0, A:0, H:0, CH:0}}}
    month_names = {
        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr", "05": "May", "06": "Jun",
        "07": "Jul", "08": "Aug", "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec"
    }
    for d in dates:
        mk = d[:7]  # YYYY-MM
        if mk not in monthly_data:
            monthly_data[mk] = {s["id"]: {"P": 0, "A": 0, "H": 0, "CH": 0} for s in staff_list}
        for s in staff_list:
            st = att_map.get(s["id"], {}).get(d, "-")
            val = status_short.get(st, "-")
            if val in monthly_data[mk][s["id"]]:
                monthly_data[mk][s["id"]][val] += 1
    sorted_months = sorted(monthly_data.keys())

    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer, PageBreak
        from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=10, rightMargin=10, topMargin=10, bottomMargin=10)
        styles = get_pdf_styles()
        elements = []

        from utils.branding_helper import get_pdf_company_header_from_db
        elements.extend(await get_pdf_company_header_from_db())
        elements.append(Paragraph(f"Staff Attendance: {date_from} to {date_to}", ParagraphStyle('t', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#1a365d'), spaceAfter=2, fontName='FreeSansBold')))

        # Column-wise: dates as rows, staff as columns
        headers = ["Date"] + [s["name"] for s in staff_list]
        rows = [headers]

        bg_map = {"P": colors.HexColor('#bbf7d0'), "A": colors.HexColor('#fecaca'),
                  "H": colors.HexColor('#fde68a'), "CH": colors.HexColor('#bfdbfe')}
        tx_map = {"P": colors.HexColor('#14532d'), "A": colors.HexColor('#7f1d1d'),
                  "H": colors.HexColor('#78350f'), "CH": colors.HexColor('#1e3a8a')}

        staff_totals = {s["id"]: {"P": 0, "H": 0, "CH": 0, "A": 0} for s in staff_list}

        for d in dates:
            row = [d[-5:]]
            for s in staff_list:
                st = att_map.get(s["id"], {}).get(d, "-")
                val = status_short.get(st, "-")
                row.append(val)
                if val in staff_totals[s["id"]]: staff_totals[s["id"]][val] += 1
            rows.append(row)

        for label in ["P", "H", "CH", "A", "Total"]:
            row = [label]
            for s in staff_list:
                if label == "Total":
                    t = staff_totals[s["id"]]
                    row.append(str(t["P"] + t["CH"] + t["H"] * 0.5))
                else:
                    row.append(str(staff_totals[s["id"]].get(label, 0)))
            rows.append(row)

        n_cols = len(headers)
        available_w = 800
        name_col_w = max(45, min(70, (available_w - 45) // max(n_cols - 1, 1)))
        col_widths = [45] + [name_col_w] * (n_cols - 1)

        total_rows = len(rows)
        row_h = min(14, max(10, 540 // total_rows))

        t = RTable(rows, colWidths=col_widths, repeatRows=1, rowHeights=[row_h] * total_rows)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
            ('FONTSIZE', (0, 0), (-1, -1), 5.5),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('FONTNAME', (0, 1), (0, -1), 'FreeSansBold'),
        ]

        n_date_rows = len(dates)
        for ri in range(1, 1 + n_date_rows):
            for ci in range(1, n_cols):
                val = rows[ri][ci]
                if val in bg_map:
                    style_cmds.append(('BACKGROUND', (ci, ri), (ci, ri), bg_map[val]))
                    style_cmds.append(('TEXTCOLOR', (ci, ri), (ci, ri), tx_map[val]))
                    style_cmds.append(('FONTNAME', (ci, ri), (ci, ri), 'FreeSansBold'))

        total_start = 1 + n_date_rows
        style_cmds.append(('BACKGROUND', (0, total_start), (-1, -1), colors.HexColor('#e0e7ff')))
        style_cmds.append(('FONTNAME', (0, total_start), (-1, -1), 'FreeSansBold'))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        # ---- PAGE 2: MONTHLY SUMMARY ----
        elements.append(PageBreak())
        elements.append(Paragraph("Monthly Summary / Masik Saransh", ParagraphStyle('ms', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#1a365d'), spaceAfter=8, fontName='FreeSansBold')))

        ms_headers = ["Staff", "Sal.Type", "Rate"] + [f"{month_names.get(m[5:7], m[5:7])} {m[:4]}" for m in sorted_months] + ["Total Days", "Est. Salary"]
        ms_rows = [ms_headers]

        # Fetch salary info for each staff
        staff_salary_map = {s["id"]: s for s in staff_list}

        for s in staff_list:
            info = staff_salary_map[s["id"]]
            sal_type = "Monthly" if info.get("salary_type") == "monthly" else "Daily"
            sal_amt = info.get("salary_amount", 0)
            per_day = sal_amt if info.get("salary_type") != "monthly" else sal_amt / 30
            row = [s["name"], sal_type, f"{sal_amt:,.0f}"]
            grand = 0
            for mk in sorted_months:
                md = monthly_data[mk][s["id"]]
                worked = md["P"] + md["CH"] + md["H"] * 0.5
                grand += worked
                row.append(f'{worked:.1f}')
            est_salary = round(grand * per_day, 0)
            row.append(f'{grand:.1f}')
            row.append(f'Rs.{est_salary:,.0f}')
            ms_rows.append(row)

        # Detail breakdown rows per month
        ms_rows.append([""] * len(ms_headers))  # blank row
        ms_rows.append(["Breakdown (P/A/H/CH)"] + [""] * (len(ms_headers) - 1))
        for s in staff_list:
            info = staff_salary_map[s["id"]]
            per_day = info.get("salary_amount", 0) if info.get("salary_type") != "monthly" else info.get("salary_amount", 0) / 30
            row = [s["name"], "", ""]
            grand_sal = 0
            for mk in sorted_months:
                md = monthly_data[mk][s["id"]]
                row.append(f'{md["P"]}/{md["A"]}/{md["H"]}/{md["CH"]}')
                worked = md["P"] + md["CH"] + md["H"] * 0.5
                grand_sal += worked * per_day
            row.append("")
            row.append(f'Rs.{round(grand_sal, 0):,.0f}')
            ms_rows.append(row)

        # Per-month salary estimate row
        ms_rows.append([""] * len(ms_headers))
        ms_rows.append(["Month-wise Est. Salary"] + [""] * (len(ms_headers) - 1))
        for s in staff_list:
            info = staff_salary_map[s["id"]]
            per_day = info.get("salary_amount", 0) if info.get("salary_type") != "monthly" else info.get("salary_amount", 0) / 30
            row = [s["name"], "", ""]
            grand_sal = 0
            for mk in sorted_months:
                md = monthly_data[mk][s["id"]]
                worked = md["P"] + md["CH"] + md["H"] * 0.5
                m_sal = round(worked * per_day, 0)
                grand_sal += m_sal
                row.append(f'Rs.{m_sal:,.0f}')
            row.append("")
            row.append(f'Rs.{round(grand_sal, 0):,.0f}')
            ms_rows.append(row)

        n_ms_cols = len(ms_headers)
        ms_col_w = max(50, min(80, 800 // max(n_ms_cols, 1)))
        ms_col_widths = [70, 45, 45] + [ms_col_w] * (n_ms_cols - 5) + [50, 65]
        # Adjust to fit 800px
        total_w = sum(ms_col_widths)
        if total_w > 800:
            scale = 800 / total_w
            ms_col_widths = [w * scale for w in ms_col_widths]

        ms_table = RTable(ms_rows, colWidths=ms_col_widths, repeatRows=1)
        ms_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#065f46')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6.5),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('FONTNAME', (0, 1), (0, -1), 'FreeSansBold'),
        ]
        # Color worked days & salary columns
        for ri in range(1, 1 + len(staff_list)):
            ms_style.append(('BACKGROUND', (-2, ri), (-2, ri), colors.HexColor('#d1fae5')))
            ms_style.append(('FONTNAME', (-2, ri), (-2, ri), 'FreeSansBold'))
            ms_style.append(('BACKGROUND', (-1, ri), (-1, ri), colors.HexColor('#fef3c7')))
            ms_style.append(('FONTNAME', (-1, ri), (-1, ri), 'FreeSansBold'))
            ms_style.append(('TEXTCOLOR', (-1, ri), (-1, ri), colors.HexColor('#92400e')))
            if ri % 2 == 0:
                ms_style.append(('BACKGROUND', (0, ri), (-3, ri), colors.HexColor('#f0fdf4')))

        # Breakdown section header
        breakdown_start = 2 + len(staff_list)
        if breakdown_start < len(ms_rows):
            ms_style.append(('BACKGROUND', (0, breakdown_start), (-1, breakdown_start), colors.HexColor('#fef3c7')))
            ms_style.append(('FONTNAME', (0, breakdown_start), (-1, breakdown_start), 'FreeSansBold'))
            ms_style.append(('SPAN', (0, breakdown_start), (-1, breakdown_start)))

        # Month-wise salary section header
        salary_section_start = breakdown_start + 1 + len(staff_list) + 1
        if salary_section_start < len(ms_rows):
            ms_style.append(('BACKGROUND', (0, salary_section_start), (-1, salary_section_start), colors.HexColor('#dbeafe')))
            ms_style.append(('FONTNAME', (0, salary_section_start), (-1, salary_section_start), 'FreeSansBold'))
            ms_style.append(('SPAN', (0, salary_section_start), (-1, salary_section_start)))
            # Salary rows styling
            for ri in range(salary_section_start + 1, salary_section_start + 1 + len(staff_list)):
                if ri < len(ms_rows):
                    ms_style.append(('BACKGROUND', (-1, ri), (-1, ri), colors.HexColor('#fef3c7')))
                    ms_style.append(('FONTNAME', (-1, ri), (-1, ri), 'FreeSansBold'))

        ms_table.setStyle(TableStyle(ms_style))
        elements.append(ms_table)

        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=staff_attendance_{date_from}_to_{date_to}.pdf"})

    else:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        hdr_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF', size=9)
        tb = Border(left=Side(style='thin', color='cbd5e1'), right=Side(style='thin', color='cbd5e1'),
                    top=Side(style='thin', color='cbd5e1'), bottom=Side(style='thin', color='cbd5e1'))
        fill_p = PatternFill(start_color='bbf7d0', end_color='bbf7d0', fill_type='solid')
        fill_a = PatternFill(start_color='fecaca', end_color='fecaca', fill_type='solid')
        fill_h = PatternFill(start_color='fde68a', end_color='fde68a', fill_type='solid')
        fill_ch = PatternFill(start_color='bfdbfe', end_color='bfdbfe', fill_type='solid')
        font_p = Font(color='14532d', size=9, bold=True)
        font_a = Font(color='7f1d1d', size=9, bold=True)
        font_h = Font(color='78350f', size=9, bold=True)
        font_ch = Font(color='1e3a8a', size=9, bold=True)
        fill_map = {"P": (fill_p, font_p), "A": (fill_a, font_a), "H": (fill_h, font_h), "CH": (fill_ch, font_ch)}
        total_fill = PatternFill(start_color='e0e7ff', end_color='e0e7ff', fill_type='solid')

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1+len(staff_list))
        ws['A1'] = f"Staff Attendance: {date_from} to {date_to}"
        ws['A1'].font = Font(bold=True, size=12, color='1a365d')

        # Column-wise: Date col + staff name cols
        headers = ["Date"] + [s["name"] for s in staff_list]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.fill = hdr_fill; c.font = hdr_font; c.border = tb; c.alignment = Alignment(horizontal='center')

        staff_totals = {s["id"]: {"P": 0, "H": 0, "CH": 0, "A": 0} for s in staff_list}

        row_num = 4
        for d in dates:
            ws.cell(row=row_num, column=1, value=d[-5:]).border = tb
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=9)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            for si, s in enumerate(staff_list):
                st = att_map.get(s["id"], {}).get(d, "-")
                val = status_short.get(st, "-")
                c = ws.cell(row=row_num, column=2+si, value=val)
                c.border = tb; c.alignment = Alignment(horizontal='center')
                if val in fill_map:
                    c.fill, c.font = fill_map[val]
                if val in staff_totals[s["id"]]: staff_totals[s["id"]][val] += 1
            row_num += 1

        for label in ["P", "H", "CH", "A", "Total"]:
            c = ws.cell(row=row_num, column=1, value=label)
            c.font = Font(bold=True, size=9); c.fill = total_fill; c.border = tb; c.alignment = Alignment(horizontal='center')
            for si, s in enumerate(staff_list):
                if label == "Total":
                    t = staff_totals[s["id"]]
                    v = t["P"] + t["CH"] + t["H"] * 0.5
                else:
                    v = staff_totals[s["id"]].get(label, 0)
                c = ws.cell(row=row_num, column=2+si, value=v)
                c.font = Font(bold=True, size=9); c.fill = total_fill; c.border = tb; c.alignment = Alignment(horizontal='center')
            row_num += 1

        ws.column_dimensions['A'].width = 8
        for i in range(len(staff_list)):
            col_letter = chr(66 + i) if i < 25 else 'A' + chr(65 + i - 25)
            ws.column_dimensions[col_letter].width = 10

        # Set print area to fit on one page
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr = None

        # ---- SHEET 2: MONTHLY SUMMARY ----
        ws2 = wb.create_sheet("Monthly Summary")
        ms_hdr_fill = PatternFill(start_color='065f46', end_color='065f46', fill_type='solid')
        ms_hdr_font = Font(bold=True, color='FFFFFF', size=9)
        green_fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
        yellow_fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')

        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(sorted_months))
        ws2['A1'] = f"Monthly Summary / Masik Saransh ({date_from} to {date_to})"
        ws2['A1'].font = Font(bold=True, size=12, color='065f46')

        # Headers: Staff | Salary Type | Rate | Month1 | Month2 ... | Total Days | Est. Salary
        ms_hdrs = ["Staff", "Sal.Type", "Rate"] + [f"{month_names.get(m[5:7], m[5:7])} {m[:4]}" for m in sorted_months] + ["Total Days", "Est. Salary"]
        for i, h in enumerate(ms_hdrs, 1):
            c = ws2.cell(row=3, column=i, value=h)
            c.fill = ms_hdr_fill; c.font = ms_hdr_font; c.border = tb; c.alignment = Alignment(horizontal='center')

        sal_fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
        sal_font = Font(bold=True, size=9, color='92400e')
        blue_fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')

        r = 4
        for s in staff_list:
            sal_type = "Monthly" if s.get("salary_type") == "monthly" else "Daily"
            sal_amt = s.get("salary_amount", 0)
            per_day = sal_amt if s.get("salary_type") != "monthly" else sal_amt / 30
            ws2.cell(row=r, column=1, value=s["name"]).font = Font(bold=True, size=9)
            ws2.cell(row=r, column=1).border = tb
            ws2.cell(row=r, column=2, value=sal_type).border = tb
            ws2.cell(row=r, column=2).alignment = Alignment(horizontal='center')
            ws2.cell(row=r, column=3, value=sal_amt).border = tb
            ws2.cell(row=r, column=3).alignment = Alignment(horizontal='center')
            ws2.cell(row=r, column=3).font = Font(size=9)
            grand = 0
            for mi, mk in enumerate(sorted_months):
                md = monthly_data[mk][s["id"]]
                worked = md["P"] + md["CH"] + md["H"] * 0.5
                grand += worked
                c = ws2.cell(row=r, column=4 + mi, value=worked)
                c.border = tb; c.alignment = Alignment(horizontal='center'); c.font = Font(size=9)
            # Total Days column
            gt_cell = ws2.cell(row=r, column=4 + len(sorted_months), value=grand)
            gt_cell.fill = green_fill; gt_cell.font = Font(bold=True, size=9); gt_cell.border = tb
            gt_cell.alignment = Alignment(horizontal='center')
            # Est. Salary column
            est_salary = round(grand * per_day, 0)
            es_cell = ws2.cell(row=r, column=5 + len(sorted_months), value=est_salary)
            es_cell.fill = sal_fill; es_cell.font = sal_font; es_cell.border = tb
            es_cell.alignment = Alignment(horizontal='center')
            es_cell.number_format = '#,##0'
            r += 1

        # Blank row + Breakdown header
        r += 1
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(ms_hdrs))
        c = ws2.cell(row=r, column=1, value="Breakdown (P / A / H / CH)")
        c.fill = yellow_fill; c.font = Font(bold=True, size=10, color='78350f'); c.border = tb
        r += 1

        for s in staff_list:
            ws2.cell(row=r, column=1, value=s["name"]).font = Font(bold=True, size=9)
            ws2.cell(row=r, column=1).border = tb
            for mi, mk in enumerate(sorted_months):
                md = monthly_data[mk][s["id"]]
                c = ws2.cell(row=r, column=4 + mi, value=f'{md["P"]} / {md["A"]} / {md["H"]} / {md["CH"]}')
                c.border = tb; c.alignment = Alignment(horizontal='center'); c.font = Font(size=8)
            r += 1

        # Month-wise Estimated Salary section
        r += 1
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(ms_hdrs))
        c = ws2.cell(row=r, column=1, value="Month-wise Estimated Salary / Mahine Ka Anumanit Vetan")
        c.fill = blue_fill; c.font = Font(bold=True, size=10, color='1e3a8a'); c.border = tb
        r += 1

        for s in staff_list:
            per_day = s.get("salary_amount", 0) if s.get("salary_type") != "monthly" else s.get("salary_amount", 0) / 30
            ws2.cell(row=r, column=1, value=s["name"]).font = Font(bold=True, size=9)
            ws2.cell(row=r, column=1).border = tb
            grand_sal = 0
            for mi, mk in enumerate(sorted_months):
                md = monthly_data[mk][s["id"]]
                worked = md["P"] + md["CH"] + md["H"] * 0.5
                m_sal = round(worked * per_day, 0)
                grand_sal += m_sal
                c = ws2.cell(row=r, column=4 + mi, value=m_sal)
                c.border = tb; c.alignment = Alignment(horizontal='center'); c.font = Font(size=9)
                c.number_format = '#,##0'
            # Grand total salary
            es_cell = ws2.cell(row=r, column=5 + len(sorted_months), value=round(grand_sal, 0))
            es_cell.fill = sal_fill; es_cell.font = sal_font; es_cell.border = tb
            es_cell.alignment = Alignment(horizontal='center')
            es_cell.number_format = '#,##0'
            r += 1

        # Set column widths for summary sheet
        ws2.column_dimensions['A'].width = 14
        ws2.column_dimensions['B'].width = 8
        ws2.column_dimensions['C'].width = 8
        for i in range(len(sorted_months) + 2):  # +2 for Total Days and Est. Salary
            col_letter = chr(68 + i) if (68 + i) <= 90 else 'A' + chr(65 + (68 + i - 91))
            ws2.column_dimensions[col_letter].width = 12

        ws2.page_setup.orientation = 'landscape'
        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = 1

        buf = io.BytesIO()
        # 🎯 v104.44.9 — Apply consolidated multi-record polish
        from utils.export_helpers import apply_consolidated_excel_polish
        apply_consolidated_excel_polish(ws)
        wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=staff_attendance_{date_from}_to_{date_to}.xlsx"})


# ============ EXPORT: PAYMENT REPORT ============

@router.get("/staff/export/payments")
async def export_payments(fmt: str = "excel", kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.staff_payments.find(query, {"_id": 0}).sort([("date", 1), ("created_at", 1)]).to_list(5000)

    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer
        from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=15, bottomMargin=15)
        styles = get_pdf_styles()
        elements = []

        from utils.branding_helper import get_pdf_company_header_from_db
        elements.extend(await get_pdf_company_header_from_db())
        elements.append(Paragraph(f"Staff Payment Report", ParagraphStyle('t', parent=styles['Title'], fontSize=14, textColor=colors.HexColor('#1a365d'))))
        elements.append(Spacer(1, 8))

        headers = ['Date', 'Staff', 'Period', 'Days Worked', 'Gross', 'Adv Deduct', 'Net Paid']
        rows = [headers]
        total_gross = total_adv = total_net = 0
        for p in payments:
            rows.append([fmt_date(p.get("date","")), p.get("staff_name",""),
                f"{p.get('period_from','')} to {p.get('period_to','')}",
                str(p.get("days_worked",0)),
                f"Rs.{p.get('gross_salary',0):,.0f}",
                f"Rs.{p.get('advance_deducted',0):,.0f}",
                f"Rs.{p.get('net_payment',0):,.0f}"])
            total_gross += p.get("gross_salary", 0)
            total_adv += p.get("advance_deducted", 0)
            total_net += p.get("net_payment", 0)
        rows.append(["", "", "TOTAL", "", f"Rs.{total_gross:,.0f}", f"Rs.{total_adv:,.0f}", f"Rs.{total_net:,.0f}"])

        t = RTable(rows, colWidths=[65, 80, 130, 55, 75, 75, 75], repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e7ff')),
            ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
        ]
        for i in range(1, len(rows)-1):
            if i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f5f5f5')))
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=staff_payments.pdf"})

    else:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active; ws.title = "Staff Payments"
        hdr_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF', size=9)
        tb = Border(left=Side(style='thin', color='cbd5e1'), right=Side(style='thin', color='cbd5e1'),
                    top=Side(style='thin', color='cbd5e1'), bottom=Side(style='thin', color='cbd5e1'))

        ws.merge_cells('A1:G1')
        ws['A1'] = "Staff Payment Report"
        ws['A1'].font = Font(bold=True, size=12, color='1a365d')

        headers = ['Date', 'Staff', 'Period', 'Days', 'Gross', 'Adv Deducted', 'Net Paid']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.fill = hdr_fill; c.font = hdr_font; c.border = tb; c.alignment = Alignment(horizontal='center')

        row_n = 4
        total_gross = total_adv = total_net = 0
        for p in payments:
            vals = [fmt_date(p.get("date","")), p.get("staff_name",""),
                f"{p.get('period_from','')} to {p.get('period_to','')}",
                p.get("days_worked",0), p.get("gross_salary",0),
                p.get("advance_deducted",0), p.get("net_payment",0)]
            for i, v in enumerate(vals, 1):
                c = ws.cell(row=row_n, column=i, value=v)
                c.border = tb; c.font = Font(size=9)
            total_gross += p.get("gross_salary", 0)
            total_adv += p.get("advance_deducted", 0)
            total_net += p.get("net_payment", 0)
            row_n += 1
        # Total row
        ws.cell(row=row_n, column=3, value="TOTAL").font = Font(bold=True, size=9)
        for ci, v in enumerate([total_gross, total_adv, total_net], 5):
            c = ws.cell(row=row_n, column=ci, value=v)
            c.font = Font(bold=True, size=9); c.border = tb

        for w, col in [(12,'A'),(16,'B'),(24,'C'),(8,'D'),(12,'E'),(12,'F'),(12,'G')]:
            ws.column_dimensions[col].width = w

        # 🎯 v104.44.9 — Apply consolidated multi-record polish
        from utils.export_helpers import apply_consolidated_excel_polish
        apply_consolidated_excel_polish(ws)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=staff_payments.xlsx"})


# ============= ADVANCE LEDGER EXCEL EXPORT =============
class LedgerItem(BaseModel):
    date: str = ""
    staff_name: str = ""
    description: str = ""
    debit: float = 0
    credit: float = 0
    balance: float = 0

class LedgerExportRequest(BaseModel):
    ledger: list[LedgerItem] = []
    staff_name: str = "All Staff"
    kms_year: str = ""
    season: str = ""

@router.post("/staff/advance-ledger/export")
async def export_advance_ledger(req: LedgerExportRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = "Advance Ledger"
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:G1')
    ws['A1'] = f"Advance Ledger - {req.staff_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"{req.kms_year} {req.season}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')

    headers = ['#', 'Date', 'Staff', 'Description', 'Debit (Rs.)', 'Credit (Rs.)', 'Balance (Rs.)']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = tb; c.alignment = Alignment(horizontal='center')

    def fmt_date(d):
        if not d: return ''
        parts = d.split('-')
        return f"{parts[2]}-{parts[1]}-{parts[0]}" if len(parts) == 3 else d

    for i, item in enumerate(req.ledger):
        row_n = i + 4
        ws.cell(row=row_n, column=1, value=i+1).border = tb
        ws.cell(row=row_n, column=2, value=fmt_date(item.date)).border = tb
        ws.cell(row=row_n, column=3, value=item.staff_name).border = tb
        ws.cell(row=row_n, column=4, value=item.description).border = tb
        c = ws.cell(row=row_n, column=5, value=item.debit if item.debit > 0 else None)
        c.border = tb; c.font = Font(color="dc2626") if item.debit > 0 else Font()
        c = ws.cell(row=row_n, column=6, value=item.credit if item.credit > 0 else None)
        c.border = tb; c.font = Font(color="059669") if item.credit > 0 else Font()
        c = ws.cell(row=row_n, column=7, value=item.balance)
        c.border = tb; c.font = Font(bold=True)

    # Totals
    row_n = len(req.ledger) + 4
    ws.cell(row=row_n, column=1, value="TOTAL").font = Font(bold=True)
    total_d = sum(l.debit for l in req.ledger)
    total_c = sum(l.credit for l in req.ledger)
    ws.cell(row=row_n, column=5, value=total_d).font = Font(bold=True, color="dc2626")
    ws.cell(row=row_n, column=6, value=total_c).font = Font(bold=True, color="059669")
    ws.cell(row=row_n, column=7, value=round(total_d - total_c, 2)).font = Font(bold=True)

    for w, col in [(5,'A'),(12,'B'),(16,'C'),(30,'D'),(14,'E'),(14,'F'),(14,'G')]:
        ws.column_dimensions[col].width = w

    # 🎯 v104.44.9 — Apply consolidated multi-record polish
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=advance_ledger_{req.staff_name}.xlsx"})
