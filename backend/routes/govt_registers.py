from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone, timedelta
from database import db
from utils.date_format import fmt_date
import uuid, io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors

router = APIRouter()

# ============ HELPER FUNCTIONS ============

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
data_font = Font(name="Calibri", size=10)
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
total_font = Font(name="Calibri", bold=True, size=10)


async def get_company_name():
    branding = await db.branding.find_one({}, {"_id": 0})
    if branding:
        return branding.get("company_name", "NAVKAR AGRO"), branding.get("tagline", "")
    return "NAVKAR AGRO", ""


def style_govt_excel(ws, title, headers, data_rows, col_widths, company_name="", subtitle=""):
    """Apply professional government register styling to Excel worksheet."""
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=company_name or title)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub_cell = ws.cell(row=2, column=1, value=subtitle or title)
    sub_cell.font = Font(name="Calibri", bold=True, size=12, color="4472C4")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Header row (row 3)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[3].height = 30

    # Data rows
    for ri, row_data in enumerate(data_rows, 4):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")

    # Column widths
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    return ws


# ============ FORM A: Paddy Received from OSCSC ============

@router.get("/govt-registers/form-a")
async def get_form_a(kms_year: Optional[str] = None, season: Optional[str] = None,
                     date_from: Optional[str] = None, date_to: Optional[str] = None,
                     group_by: Optional[str] = "daily"):
    """Form A - Paddy stock register (from OSCSC/State Procuring Agency).
    group_by: daily (default) or weekly."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    date_q = {}
    if date_from:
        date_q["$gte"] = date_from
    if date_to:
        date_q["$lte"] = date_to
    if date_q:
        query["date"] = date_q

    # Get all mill entries grouped by date
    entries = await db.mill_entries.find(query, {"_id": 0, "date": 1, "final_w": 1, "kg": 1, "agent_name": 1, "mandi_name": 1, "bag": 1}).sort("date", 1).to_list(50000)

    # Get milling entries for the same period
    mill_query = {}
    if kms_year:
        mill_query["kms_year"] = kms_year
    if season:
        mill_query["season"] = season
    if date_q:
        mill_query["date"] = date_q
    milling_entries = await db.milling_entries.find(mill_query, {"_id": 0, "date": 1, "paddy_input_qntl": 1}).sort("date", 1).to_list(50000)

    # Group by date
    daily_received = {}
    for e in entries:
        d = e.get("date", "")
        if not d:
            continue
        if d not in daily_received:
            daily_received[d] = {"received_qntl": 0, "bags": 0, "count": 0}
        final_w = float(e.get("final_w", 0) or 0) / 100  # final_w is stored in KG, convert to QNTL
        if final_w == 0:
            final_w = float(e.get("kg", 0) or 0) / 100
        daily_received[d]["received_qntl"] += final_w
        daily_received[d]["bags"] += int(e.get("bag", 0) or 0)
        daily_received[d]["count"] += 1

    daily_milled = {}
    for m in milling_entries:
        d = m.get("date", "")
        if not d:
            continue
        if d not in daily_milled:
            daily_milled[d] = 0
        daily_milled[d] += float(m.get("paddy_input_qntl", 0) or 0)

    # Build daily register with running balance
    all_dates = sorted(set(list(daily_received.keys()) + list(daily_milled.keys())))
    rows = []
    opening_balance = 0
    total_received = 0
    total_milled = 0

    for d in all_dates:
        received = round(daily_received.get(d, {}).get("received_qntl", 0), 2)
        bags = daily_received.get(d, {}).get("bags", 0)
        count = daily_received.get(d, {}).get("count", 0)
        milled = round(daily_milled.get(d, 0), 2)
        total_paddy = round(opening_balance + received, 2)
        closing_balance = round(total_paddy - milled, 2)

        total_received += received
        total_milled += milled

        rows.append({
            "date": d,
            "opening_balance": round(opening_balance, 2),
            "received_qntl": received,
            "bags": bags,
            "entries_count": count,
            "total_paddy": total_paddy,
            "milled_qntl": milled,
            "closing_balance": closing_balance,
        })
        opening_balance = closing_balance

    # Weekly grouping if requested
    if group_by == "weekly" and rows:
        from datetime import datetime as _dt
        weekly_rows = []
        week_data = None
        for r in rows:
            try:
                dt = _dt.strptime(r["date"], "%Y-%m-%d")
                # Monday-based week
                week_start = dt - timedelta(days=dt.weekday())
                week_key = week_start.strftime("%Y-%m-%d")
            except:
                week_key = r["date"]
            if week_data is None or week_data["_week_key"] != week_key:
                if week_data:
                    weekly_rows.append(week_data)
                week_end = (week_start + timedelta(days=6)).strftime("%Y-%m-%d")
                week_data = {
                    "_week_key": week_key,
                    "date": f"{fmt_date(week_key)} to {fmt_date(week_end)}",
                    "date_from": week_key,
                    "date_to": week_end,
                    "opening_balance": r["opening_balance"],
                    "received_qntl": 0, "bags": 0, "entries_count": 0,
                    "total_paddy": 0, "milled_qntl": 0, "closing_balance": 0,
                }
            week_data["received_qntl"] = round(week_data["received_qntl"] + r["received_qntl"], 2)
            week_data["bags"] += r["bags"]
            week_data["entries_count"] += r["entries_count"]
            week_data["milled_qntl"] = round(week_data["milled_qntl"] + r["milled_qntl"], 2)
            week_data["total_paddy"] = round(week_data["opening_balance"] + week_data["received_qntl"], 2)
            week_data["closing_balance"] = round(week_data["total_paddy"] - week_data["milled_qntl"], 2)
        if week_data:
            weekly_rows.append(week_data)
        # Clean internal keys
        for wr in weekly_rows:
            wr.pop("_week_key", None)
        rows = weekly_rows

    return {
        "rows": rows,
        "summary": {
            "total_received": round(total_received, 2),
            "total_milled": round(total_milled, 2),
            "final_balance": round(opening_balance, 2),
            "total_days": len(rows)
        }
    }


@router.get("/govt-registers/form-a/excel")
async def export_form_a_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                               date_from: Optional[str] = None, date_to: Optional[str] = None,
                               group_by: Optional[str] = "daily"):
    data = await get_form_a(kms_year, season, date_from, date_to, group_by)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "Form A"

    headers = ["Date", "Opening Bal (Qtl)", "Paddy Received (Qtl)", "Bags", "Total Paddy (Qtl)", "Paddy Milled (Qtl)", "Closing Bal (Qtl)"]
    col_widths = [28 if group_by == "weekly" else 14, 18, 20, 10, 18, 20, 18]
    data_rows = []
    for r in data["rows"]:
        date_val = r["date"] if group_by == "weekly" else fmt_date(r["date"])
        data_rows.append([
            date_val, r["opening_balance"], r["received_qntl"],
            r["bags"], r["total_paddy"], r["milled_qntl"], r["closing_balance"]
        ])

    # Total row
    s = data["summary"]
    data_rows.append(["TOTAL", "", s["total_received"], "", "", s["total_milled"], s["final_balance"]])

    style_govt_excel(ws, "Form A - Paddy Stock Register", headers, data_rows, col_widths,
                     company, f"Form A - Paddy Received from State Procuring Agency | {kms_year or 'All'} {season or ''}")

    # Bold total row
    total_row_num = len(data_rows) + 3
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_num, column=ci)
        cell.font = total_font
        cell.fill = total_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Form_A_Paddy_Register_{kms_year or 'all'}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/govt-registers/form-a/pdf")
async def export_form_a_pdf(kms_year: Optional[str] = None, season: Optional[str] = None,
                             date_from: Optional[str] = None, date_to: Optional[str] = None,
                             group_by: Optional[str] = "daily"):
    data = await get_form_a(kms_year, season, date_from, date_to, group_by)
    company, _ = await get_company_name()
    rows = data["rows"]
    summary = data["summary"]

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    w, h = landscape(A4)

    # Watermark settings
    wm_doc = await db.app_settings.find_one({"setting_id": "watermark"}, {"_id": 0})
    wm = wm_doc or {}

    def draw_tiled_watermark():
        if not wm.get("enabled") or wm.get("type") != "text" or not wm.get("text"): return
        c.saveState()
        c.setFont("Helvetica", 28)
        c.setFillAlpha(wm.get("opacity", 0.06))
        c.setFillColor(colors.HexColor("#94A3B8"))
        for yi in range(0, int(h) + 80, 80):
            for xi in range(0, int(w) + 200, 200):
                c.saveState(); c.translate(xi, yi); c.rotate(30); c.drawString(0, 0, wm["text"]); c.restoreState()
        c.restoreState()

    def draw_page_header(pg_num=1):
        draw_tiled_watermark()
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(w / 2, h - 30, company)
        c.setFont("Helvetica", 10)
        title = f"Form A - Paddy Received from State Procuring Agency | {kms_year or 'All'}"
        if season: title += f" | {season}"
        c.drawCentredString(w / 2, h - 46, title)
        c.setFont("Helvetica", 7)
        c.drawRightString(w - 25, h - 46, f"Page {pg_num}")

    headers = ["Date", "Opening Bal (Q)", "Paddy Recd (Q)", "Bags", "Total Paddy (Q)", "Paddy Milled (Q)", "Closing Bal (Q)"]
    col_widths = [90 if group_by == "weekly" else 70, 85, 90, 50, 90, 90, 85]
    table_start_x = 40
    y_start = h - 70

    def draw_table_header(y):
        c.setFillColor(colors.HexColor("#1F4E79"))
        c.rect(table_start_x, y - 14, sum(col_widths), 18, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 7.5)
        x = table_start_x + 4
        for i, hdr in enumerate(headers):
            c.drawString(x, y - 10, hdr)
            x += col_widths[i]
        return y - 18

    pg = 1
    draw_page_header(pg)
    y = draw_table_header(y_start)

    for idx, row in enumerate(rows):
        if y < 50:
            c.showPage(); pg += 1; draw_page_header(pg); y = draw_table_header(y_start)
        if idx % 2 == 0:
            c.setFillColor(colors.HexColor("#F8FAFC"))
            c.rect(table_start_x, y - 12, sum(col_widths), 15, fill=True, stroke=False)
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 7)
        x = table_start_x + 4
        vals = [row.get("date", ""), f"{row.get('opening_balance', 0):.2f}", f"{row.get('received_qntl', 0):.2f}",
                str(row.get("bags", 0)), f"{row.get('total_paddy', 0):.2f}", f"{row.get('milled_qntl', 0):.2f}", f"{row.get('closing_balance', 0):.2f}"]
        for i, val in enumerate(vals):
            c.drawString(x, y - 8, str(val))
            x += col_widths[i]
        y -= 15

    # Totals row
    if y < 50: c.showPage(); pg += 1; draw_page_header(pg); y = draw_table_header(y_start)
    c.setFillColor(colors.HexColor("#1F4E79"))
    c.rect(table_start_x, y - 12, sum(col_widths), 16, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 7.5)
    x = table_start_x + 4
    totals = ["TOTAL", "", f"{summary.get('total_received', 0):.2f}", "", "", f"{summary.get('total_milled', 0):.2f}", f"{summary.get('final_balance', 0):.2f}"]
    for i, val in enumerate(totals):
        c.drawString(x, y - 8, val)
        x += col_widths[i]

    c.save()
    buf.seek(0)
    fname = f"Form_A_Paddy_Register_{kms_year or 'all'}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


# ============ FORM B: CMR Produced and Delivered ============

@router.get("/govt-registers/form-b")
async def get_form_b(kms_year: Optional[str] = None, season: Optional[str] = None,
                     date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Form B - CMR produced and delivered to OSCSC/State Agency."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    date_q = {}
    if date_from:
        date_q["$gte"] = date_from
    if date_to:
        date_q["$lte"] = date_to
    if date_q:
        query["date"] = date_q

    # Milling entries = CMR production
    milling = await db.milling_entries.find(query, {"_id": 0, "date": 1, "cmr_delivery_qntl": 1, "rice_qntl": 1,
                                                     "rice_type": 1, "outturn_ratio": 1, "paddy_input_qntl": 1}).sort("date", 1).to_list(50000)

    # Sale book = CMR delivery
    sale_entries = await db.salebook.find(query, {"_id": 0, "date": 1, "items": 1, "party_name": 1}).sort("date", 1).to_list(50000)

    daily_produced = {}
    for m in milling:
        d = m.get("date", "")
        if not d:
            continue
        if d not in daily_produced:
            daily_produced[d] = {"cmr_qntl": 0, "paddy_qntl": 0}
        daily_produced[d]["cmr_qntl"] += float(m.get("cmr_delivery_qntl", 0) or m.get("rice_qntl", 0) or 0)
        daily_produced[d]["paddy_qntl"] += float(m.get("paddy_input_qntl", 0) or 0)

    daily_delivered = {}
    for s in sale_entries:
        d = s.get("date", "")
        if not d:
            continue
        if d not in daily_delivered:
            daily_delivered[d] = {"delivered_qntl": 0, "parties": []}
        total_qty = sum(float(it.get("quantity", 0) or 0) for it in s.get("items", []))
        daily_delivered[d]["delivered_qntl"] += total_qty / 100  # KG to QNTL
        party = s.get("party_name", "")
        if party and party not in daily_delivered[d]["parties"]:
            daily_delivered[d]["parties"].append(party)

    all_dates = sorted(set(list(daily_produced.keys()) + list(daily_delivered.keys())))
    rows = []
    opening_balance = 0
    total_produced = 0
    total_delivered = 0

    for d in all_dates:
        produced = round(daily_produced.get(d, {}).get("cmr_qntl", 0), 2)
        delivered = round(daily_delivered.get(d, {}).get("delivered_qntl", 0), 2)
        parties = daily_delivered.get(d, {}).get("parties", [])
        total_rice = round(opening_balance + produced, 2)
        closing = round(total_rice - delivered, 2)

        total_produced += produced
        total_delivered += delivered

        rows.append({
            "date": d,
            "opening_balance": round(opening_balance, 2),
            "cmr_produced": produced,
            "total_rice": total_rice,
            "cmr_delivered": delivered,
            "closing_balance": closing,
            "delivered_to": ", ".join(parties) if parties else "-",
        })
        opening_balance = closing

    return {
        "rows": rows,
        "summary": {
            "total_produced": round(total_produced, 2),
            "total_delivered": round(total_delivered, 2),
            "final_balance": round(opening_balance, 2),
        }
    }


@router.get("/govt-registers/form-b/excel")
async def export_form_b_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                               date_from: Optional[str] = None, date_to: Optional[str] = None):
    data = await get_form_b(kms_year, season, date_from, date_to)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "Form B"

    headers = ["Date", "Opening Bal (Qtl)", "CMR Produced (Qtl)", "Total Rice (Qtl)", "CMR Delivered (Qtl)", "Closing Bal (Qtl)", "Delivered To"]
    col_widths = [14, 18, 20, 18, 20, 18, 30]
    data_rows = []
    for r in data["rows"]:
        data_rows.append([
            fmt_date(r["date"]), r["opening_balance"], r["cmr_produced"],
            r["total_rice"], r["cmr_delivered"], r["closing_balance"], r["delivered_to"]
        ])
    s = data["summary"]
    data_rows.append(["TOTAL", "", s["total_produced"], "", s["total_delivered"], s["final_balance"], ""])

    style_govt_excel(ws, "Form B - CMR Register", headers, data_rows, col_widths,
                     company, f"Form B - Custom Milled Rice Produced & Delivered | {kms_year or 'All'} {season or ''}")

    total_row_num = len(data_rows) + 3
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_num, column=ci)
        cell.font = total_font
        cell.fill = total_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Form_B_CMR_Register_{kms_year or 'all'}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


# ============ FORM E: Miller's Own Paddy ============

@router.get("/govt-registers/form-e")
async def get_form_e(kms_year: Optional[str] = None, season: Optional[str] = None,
                     date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Form E - Miller's own paddy purchases, milling and stock."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    date_q = {}
    if date_from:
        date_q["$gte"] = date_from
    if date_to:
        date_q["$lte"] = date_to
    if date_q:
        query["date"] = date_q

    # Private paddy purchases
    purchases = await db.private_paddy.find(query, {"_id": 0, "date": 1, "party_name": 1,
                                                     "kg": 1, "bag": 1, "rate_per_qntl": 1, "amount": 1}).sort("date", 1).to_list(50000)

    daily_data = {}
    for p in purchases:
        d = p.get("date", "")
        if not d:
            continue
        if d not in daily_data:
            daily_data[d] = {"purchased_qntl": 0, "bags": 0, "parties": [], "amount": 0}
        kg = float(p.get("kg", 0) or 0)
        daily_data[d]["purchased_qntl"] += kg / 100
        daily_data[d]["bags"] += int(p.get("bag", 0) or 0)
        daily_data[d]["amount"] += float(p.get("amount", 0) or 0)
        party = p.get("party_name", "")
        if party and party not in daily_data[d]["parties"]:
            daily_data[d]["parties"].append(party)

    all_dates = sorted(daily_data.keys())
    rows = []
    opening_balance = 0
    total_purchased = 0

    for d in all_dates:
        purchased = round(daily_data[d]["purchased_qntl"], 2)
        bags = daily_data[d]["bags"]
        parties = daily_data[d]["parties"]
        amount = round(daily_data[d]["amount"], 2)
        total = round(opening_balance + purchased, 2)

        total_purchased += purchased
        rows.append({
            "date": d,
            "opening_balance": round(opening_balance, 2),
            "purchased_qntl": purchased,
            "bags": bags,
            "total": total,
            "closing_balance": total,
            "parties": ", ".join(parties) if parties else "-",
            "amount": amount,
        })
        opening_balance = total

    return {
        "rows": rows,
        "summary": {
            "total_purchased": round(total_purchased, 2),
            "final_balance": round(opening_balance, 2),
        }
    }


@router.get("/govt-registers/form-e/excel")
async def export_form_e_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                               date_from: Optional[str] = None, date_to: Optional[str] = None):
    data = await get_form_e(kms_year, season, date_from, date_to)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "Form E"

    headers = ["Date", "Opening Bal (Qtl)", "Paddy Purchased (Qtl)", "Bags", "Total (Qtl)", "Closing Bal (Qtl)", "Party Name", "Amount (Rs)"]
    col_widths = [14, 18, 22, 10, 15, 18, 28, 16]
    data_rows = []
    for r in data["rows"]:
        data_rows.append([
            fmt_date(r["date"]), r["opening_balance"], r["purchased_qntl"],
            r["bags"], r["total"], r["closing_balance"], r["parties"], r["amount"]
        ])
    s = data["summary"]
    data_rows.append(["TOTAL", "", s["total_purchased"], "", "", s["final_balance"], "", ""])

    style_govt_excel(ws, "Form E - Miller's Own Paddy", headers, data_rows, col_widths,
                     company, f"Form E - Miller's Own Paddy Purchase & Stock | {kms_year or 'All'} {season or ''}")

    total_row_num = len(data_rows) + 3
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_num, column=ci)
        cell.font = total_font
        cell.fill = total_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=Form_E_Miller_Paddy_{kms_year or 'all'}.xlsx"})


# ============ FORM F: Miller's Own Rice Sale ============

@router.get("/govt-registers/form-f")
async def get_form_f(kms_year: Optional[str] = None, season: Optional[str] = None,
                     date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Form F - Miller's own rice produced and sold."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    date_q = {}
    if date_from:
        date_q["$gte"] = date_from
    if date_to:
        date_q["$lte"] = date_to
    if date_q:
        query["date"] = date_q

    # Private rice sales
    sales = await db.private_rice_sales.find(query, {"_id": 0, "date": 1, "party_name": 1,
                                                      "quantity_qntl": 1, "rate": 1, "amount": 1}).sort("date", 1).to_list(50000)

    # Also get from salebook (private sales)
    sale_vouchers = await db.salebook.find(query, {"_id": 0, "date": 1, "party_name": 1, "items": 1, "total": 1}).sort("date", 1).to_list(50000)

    daily_data = {}
    for s in sales:
        d = s.get("date", "")
        if not d:
            continue
        if d not in daily_data:
            daily_data[d] = {"sold_qntl": 0, "parties": [], "amount": 0}
        daily_data[d]["sold_qntl"] += float(s.get("quantity_qntl", 0) or 0)
        daily_data[d]["amount"] += float(s.get("amount", 0) or 0)
        party = s.get("party_name", "")
        if party and party not in daily_data[d]["parties"]:
            daily_data[d]["parties"].append(party)

    for sv in sale_vouchers:
        d = sv.get("date", "")
        if not d:
            continue
        if d not in daily_data:
            daily_data[d] = {"sold_qntl": 0, "parties": [], "amount": 0}
        total_qty = sum(float(it.get("quantity", 0) or 0) for it in sv.get("items", []))
        daily_data[d]["sold_qntl"] += total_qty / 100
        daily_data[d]["amount"] += float(sv.get("total", 0) or 0)
        party = sv.get("party_name", "")
        if party and party not in daily_data[d]["parties"]:
            daily_data[d]["parties"].append(party)

    all_dates = sorted(daily_data.keys())
    rows = []
    total_sold = 0

    for d in all_dates:
        sold = round(daily_data[d]["sold_qntl"], 2)
        parties = daily_data[d]["parties"]
        amount = round(daily_data[d]["amount"], 2)
        total_sold += sold
        rows.append({
            "date": d,
            "sold_qntl": sold,
            "parties": ", ".join(parties) if parties else "-",
            "amount": amount,
        })

    return {
        "rows": rows,
        "summary": {"total_sold": round(total_sold, 2)}
    }


@router.get("/govt-registers/form-f/excel")
async def export_form_f_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                               date_from: Optional[str] = None, date_to: Optional[str] = None):
    data = await get_form_f(kms_year, season, date_from, date_to)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "Form F"

    headers = ["Date", "Rice Sold (Qtl)", "Party Name", "Amount (Rs)"]
    col_widths = [14, 20, 30, 18]
    data_rows = []
    for r in data["rows"]:
        data_rows.append([fmt_date(r["date"]), r["sold_qntl"], r["parties"], r["amount"]])
    data_rows.append(["TOTAL", data["summary"]["total_sold"], "", ""])

    style_govt_excel(ws, "Form F - Miller's Own Rice Sale", headers, data_rows, col_widths,
                     company, f"Form F - Miller's Own Rice Produced & Sold | {kms_year or 'All'} {season or ''}")

    total_row_num = len(data_rows) + 3
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_num, column=ci)
        cell.font = total_font
        cell.fill = total_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=Form_F_Miller_Rice_{kms_year or 'all'}.xlsx"})


# ============ FRK BLENDING REGISTER ============

@router.get("/govt-registers/frk")
async def get_frk_entries(kms_year: Optional[str] = None, season: Optional[str] = None,
                          date_from: Optional[str] = None, date_to: Optional[str] = None):
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    date_q = {}
    if date_from:
        date_q["$gte"] = date_from
    if date_to:
        date_q["$lte"] = date_to
    if date_q:
        query["date"] = date_q
    entries = await db.frk_register.find(query, {"_id": 0}).sort("date", 1).to_list(50000)
    return entries


@router.post("/govt-registers/frk")
async def create_frk_entry(data: dict, username: str = ""):
    doc = {
        "id": str(uuid.uuid4()),
        "date": data.get("date", ""),
        "kms_year": data.get("kms_year", ""),
        "season": data.get("season", ""),
        "batch_no": data.get("batch_no", ""),
        "supplier": data.get("supplier", ""),
        "opening_balance": float(data.get("opening_balance", 0) or 0),
        "received_qty": float(data.get("received_qty", 0) or 0),
        "issued_for_blending": float(data.get("issued_for_blending", 0) or 0),
        "closing_balance": float(data.get("opening_balance", 0) or 0) + float(data.get("received_qty", 0) or 0) - float(data.get("issued_for_blending", 0) or 0),
        "rice_blended_qty": float(data.get("rice_blended_qty", 0) or 0),
        "blend_ratio": data.get("blend_ratio", "1:100"),
        "remark": data.get("remark", ""),
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    doc["total"] = round(doc["opening_balance"] + doc["received_qty"], 2)
    doc["closing_balance"] = round(doc["total"] - doc["issued_for_blending"], 2)
    await db.frk_register.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/govt-registers/frk/{entry_id}")
async def update_frk_entry(entry_id: str, data: dict, username: str = ""):
    existing = await db.frk_register.find_one({"id": entry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="FRK entry not found")
    update_data = {
        "date": data.get("date", existing.get("date", "")),
        "batch_no": data.get("batch_no", existing.get("batch_no", "")),
        "supplier": data.get("supplier", existing.get("supplier", "")),
        "opening_balance": float(data.get("opening_balance", existing.get("opening_balance", 0)) or 0),
        "received_qty": float(data.get("received_qty", existing.get("received_qty", 0)) or 0),
        "issued_for_blending": float(data.get("issued_for_blending", existing.get("issued_for_blending", 0)) or 0),
        "rice_blended_qty": float(data.get("rice_blended_qty", existing.get("rice_blended_qty", 0)) or 0),
        "blend_ratio": data.get("blend_ratio", existing.get("blend_ratio", "1:100")),
        "remark": data.get("remark", existing.get("remark", "")),
        "updated_by": username,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    update_data["total"] = round(update_data["opening_balance"] + update_data["received_qty"], 2)
    update_data["closing_balance"] = round(update_data["total"] - update_data["issued_for_blending"], 2)
    await db.frk_register.update_one({"id": entry_id}, {"$set": update_data})
    return {"success": True}


@router.delete("/govt-registers/frk/{entry_id}")
async def delete_frk_entry(entry_id: str, username: str = "", role: str = ""):
    result = await db.frk_register.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="FRK entry not found")
    return {"success": True}


@router.get("/govt-registers/frk/excel")
async def export_frk_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                            date_from: Optional[str] = None, date_to: Optional[str] = None):
    entries = await get_frk_entries(kms_year, season, date_from, date_to)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "FRK Register"

    headers = ["Date", "Batch No", "Supplier", "Opening Bal (Kg)", "Received (Kg)", "Total (Kg)", "Issued for Blending (Kg)", "Closing Bal (Kg)", "Rice Blended (Qtl)", "Ratio", "Remarks"]
    col_widths = [14, 16, 22, 18, 15, 15, 22, 18, 18, 10, 20]
    data_rows = []
    for e in entries:
        data_rows.append([
            fmt_date(e.get("date", "")), e.get("batch_no", ""), e.get("supplier", ""),
            e.get("opening_balance", 0), e.get("received_qty", 0), e.get("total", 0),
            e.get("issued_for_blending", 0), e.get("closing_balance", 0),
            e.get("rice_blended_qty", 0), e.get("blend_ratio", ""), e.get("remark", "")
        ])

    style_govt_excel(ws, "FRK Blending Register", headers, data_rows, col_widths,
                     company, f"Fortified Rice Kernel (FRK) Blending Register | {kms_year or 'All'} {season or ''}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=FRK_Register_{kms_year or 'all'}.xlsx"})


# ============ GUNNY BAG REGISTER ============

@router.get("/govt-registers/gunny-bags")
async def get_gunny_bag_entries(kms_year: Optional[str] = None, season: Optional[str] = None,
                                 date_from: Optional[str] = None, date_to: Optional[str] = None):
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    date_q = {}
    if date_from:
        date_q["$gte"] = date_from
    if date_to:
        date_q["$lte"] = date_to
    if date_q:
        query["date"] = date_q
    entries = await db.gunny_bag_register.find(query, {"_id": 0}).sort("date", 1).to_list(50000)
    return entries


@router.post("/govt-registers/gunny-bags")
async def create_gunny_bag_entry(data: dict, username: str = ""):
    doc = {
        "id": str(uuid.uuid4()),
        "date": data.get("date", ""),
        "kms_year": data.get("kms_year", ""),
        "season": data.get("season", ""),
        "bag_type": data.get("bag_type", "new"),  # new, old, plastic
        "source": data.get("source", ""),  # OSCSC, Purchase, Return
        "opening_balance": int(data.get("opening_balance", 0) or 0),
        "received": int(data.get("received", 0) or 0),
        "used_for_rice": int(data.get("used_for_rice", 0) or 0),
        "used_for_paddy": int(data.get("used_for_paddy", 0) or 0),
        "damaged": int(data.get("damaged", 0) or 0),
        "returned": int(data.get("returned", 0) or 0),
        "remark": data.get("remark", ""),
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    total_in = doc["opening_balance"] + doc["received"]
    total_out = doc["used_for_rice"] + doc["used_for_paddy"] + doc["damaged"] + doc["returned"]
    doc["closing_balance"] = total_in - total_out
    await db.gunny_bag_register.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/govt-registers/gunny-bags/{entry_id}")
async def update_gunny_bag_entry(entry_id: str, data: dict, username: str = ""):
    existing = await db.gunny_bag_register.find_one({"id": entry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Gunny bag entry not found")
    update_data = {
        "date": data.get("date", existing.get("date", "")),
        "bag_type": data.get("bag_type", existing.get("bag_type", "new")),
        "source": data.get("source", existing.get("source", "")),
        "opening_balance": int(data.get("opening_balance", existing.get("opening_balance", 0)) or 0),
        "received": int(data.get("received", existing.get("received", 0)) or 0),
        "used_for_rice": int(data.get("used_for_rice", existing.get("used_for_rice", 0)) or 0),
        "used_for_paddy": int(data.get("used_for_paddy", existing.get("used_for_paddy", 0)) or 0),
        "damaged": int(data.get("damaged", existing.get("damaged", 0)) or 0),
        "returned": int(data.get("returned", existing.get("returned", 0)) or 0),
        "remark": data.get("remark", existing.get("remark", "")),
        "updated_by": username,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    total_in = update_data["opening_balance"] + update_data["received"]
    total_out = update_data["used_for_rice"] + update_data["used_for_paddy"] + update_data["damaged"] + update_data["returned"]
    update_data["closing_balance"] = total_in - total_out
    await db.gunny_bag_register.update_one({"id": entry_id}, {"$set": update_data})
    return {"success": True}


@router.delete("/govt-registers/gunny-bags/{entry_id}")
async def delete_gunny_bag_entry(entry_id: str, username: str = "", role: str = ""):
    result = await db.gunny_bag_register.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Gunny bag entry not found")
    return {"success": True}


@router.get("/govt-registers/gunny-bags/excel")
async def export_gunny_bags_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                                    date_from: Optional[str] = None, date_to: Optional[str] = None):
    entries = await get_gunny_bag_entries(kms_year, season, date_from, date_to)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "Gunny Bag Register"

    headers = ["Date", "Bag Type", "Source", "Opening Bal", "Received", "Used (Rice)", "Used (Paddy)", "Damaged", "Returned", "Closing Bal", "Remarks"]
    col_widths = [14, 14, 18, 14, 12, 14, 14, 12, 12, 14, 20]
    data_rows = []
    for e in entries:
        data_rows.append([
            fmt_date(e.get("date", "")), e.get("bag_type", ""), e.get("source", ""),
            e.get("opening_balance", 0), e.get("received", 0),
            e.get("used_for_rice", 0), e.get("used_for_paddy", 0),
            e.get("damaged", 0), e.get("returned", 0),
            e.get("closing_balance", 0), e.get("remark", "")
        ])

    style_govt_excel(ws, "Gunny Bag Stock Register", headers, data_rows, col_widths,
                     company, f"Gunny Bag Stock Register | {kms_year or 'All'} {season or ''}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=Gunny_Bag_Register_{kms_year or 'all'}.xlsx"})


# ============ TRANSIT PASS REGISTER (Auto-generated from Mill Entries) ============

@router.get("/govt-registers/transit-pass")
async def get_transit_pass_register(kms_year: Optional[str] = None, season: Optional[str] = None,
                                     date_from: Optional[str] = None, date_to: Optional[str] = None,
                                     mandi_name: Optional[str] = None, agent_name: Optional[str] = None):
    """Transit Pass Register with mandi/agent filters."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if mandi_name:
        query["mandi_name"] = {"$regex": f"^{mandi_name}$", "$options": "i"}
    if agent_name:
        query["agent_name"] = {"$regex": f"^{agent_name}$", "$options": "i"}
    date_q = {}
    if date_from:
        date_q["$gte"] = date_from
    if date_to:
        date_q["$lte"] = date_to
    if date_q:
        query["date"] = date_q

    # Only entries with TP number
    query["tp_no"] = {"$ne": "", "$exists": True}

    entries = await db.mill_entries.find(query, {
        "_id": 0, "date": 1, "tp_no": 1, "truck_no": 1, "agent_name": 1,
        "mandi_name": 1, "kg": 1, "bag": 1, "final_w": 1, "rst_no": 1,
        "tp_weight": 1, "remark": 1, "created_at": 1
    }).sort("date", 1).to_list(50000)

    rows = []
    total_qty = 0
    total_bags = 0
    total_tp_weight = 0
    mandis = set()
    agents = set()
    for e in entries:
        tp_no = str(e.get("tp_no", "")).strip()
        if not tp_no:
            continue
        final_w = float(e.get("final_w", 0) or 0) / 100  # final_w stored in KG, convert to QNTL
        if final_w == 0:
            final_w = float(e.get("kg", 0) or 0) / 100
        bags = int(e.get("bag", 0) or 0)
        tp_wt = round(float(e.get("tp_weight", 0) or 0), 2)
        total_qty += final_w
        total_bags += bags
        total_tp_weight += tp_wt
        m_name = e.get("mandi_name", "")
        a_name = e.get("agent_name", "")
        if m_name:
            mandis.add(m_name)
        if a_name:
            agents.add(a_name)
        rows.append({
            "date": e.get("date", ""),
            "tp_no": tp_no,
            "rst_no": str(e.get("rst_no", "")),
            "truck_no": e.get("truck_no", ""),
            "agent_name": a_name,
            "mandi_name": m_name,
            "qty_qntl": round(final_w, 2),
            "tp_weight": tp_wt,
            "bags": bags,
            "status": "Accepted",
            "remark": e.get("remark", ""),
        })
    rows.sort(key=lambda x: x.get("date", ""))

    return {
        "rows": rows,
        "summary": {
            "total_entries": len(rows),
            "total_qty": round(total_qty, 2),
            "total_tp_weight": round(total_tp_weight, 2),
            "total_bags": total_bags,
        },
        "filter_options": {
            "mandis": sorted(mandis),
            "agents": sorted(agents),
        }
    }


@router.get("/govt-registers/transit-pass/excel")
async def export_transit_pass_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                                     date_from: Optional[str] = None, date_to: Optional[str] = None,
                                     mandi_name: Optional[str] = None, agent_name: Optional[str] = None):
    data = await get_transit_pass_register(kms_year, season, date_from, date_to, mandi_name, agent_name)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "Transit Pass"

    filter_text = ""
    if mandi_name:
        filter_text += f" | Mandi: {mandi_name}"
    if agent_name:
        filter_text += f" | Agent: {agent_name}"

    headers = ["Date", "TP No.", "RST No.", "Vehicle No.", "Agent/Society", "Mandi/PPC", "Qty (Qtl)", "TP Weight", "Bags", "Status", "Remarks"]
    col_widths = [14, 14, 12, 16, 22, 20, 14, 14, 10, 12, 20]
    data_rows = []
    for r in data["rows"]:
        data_rows.append([
            fmt_date(r["date"]), r["tp_no"], r["rst_no"], r["truck_no"],
            r["agent_name"], r["mandi_name"], r["qty_qntl"], r["tp_weight"],
            r["bags"], r["status"], r["remark"]
        ])
    s = data["summary"]
    data_rows.append(["TOTAL", f'{s["total_entries"]} entries', "", "", "", "", s["total_qty"], s["total_tp_weight"], s["total_bags"], "", ""])

    style_govt_excel(ws, "Transit Pass Register", headers, data_rows, col_widths,
                     company, f"Transit Pass-cum-Acceptance Register | {kms_year or 'All'} {season or ''}{filter_text}")

    total_row_num = len(data_rows) + 3
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_num, column=ci)
        cell.font = total_font
        cell.fill = total_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=Transit_Pass_Register_{kms_year or 'all'}.xlsx"})


@router.get("/govt-registers/transit-pass/pdf")
async def export_transit_pass_pdf(kms_year: Optional[str] = None, season: Optional[str] = None,
                                   date_from: Optional[str] = None, date_to: Optional[str] = None,
                                   mandi_name: Optional[str] = None, agent_name: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    data = await get_transit_pass_register(kms_year, season, date_from, date_to, mandi_name, agent_name)
    company, _ = await get_company_name()
    filter_text = ""
    if mandi_name:
        filter_text += f" | Mandi: {mandi_name}"
    if agent_name:
        filter_text += f" | Agent: {agent_name}"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(f"<b>{company}</b>", styles['Title']))
    elements.append(Paragraph(f"Transit Pass Register | {kms_year or 'All'} {season or ''}{filter_text}", styles['Normal']))
    elements.append(Spacer(1, 8*mm))

    headers = ["Date", "TP No.", "RST", "Vehicle", "Agent", "Mandi", "Qty(Q)", "TP Wt", "Bags", "Status"]
    table_data = [headers]
    for r in data["rows"]:
        table_data.append([
            fmt_date(r["date"]), r["tp_no"], r["rst_no"], r["truck_no"],
            r["agent_name"][:15], r["mandi_name"][:15], r["qty_qntl"], r["tp_weight"],
            r["bags"], "Accepted"
        ])
    s = data["summary"]
    table_data.append(["TOTAL", f'{s["total_entries"]}', "", "", "", "", s["total_qty"], s["total_tp_weight"], s["total_bags"], ""])

    col_widths = [22*mm, 20*mm, 18*mm, 22*mm, 30*mm, 30*mm, 20*mm, 20*mm, 15*mm, 20*mm]
    t = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (6, 0), (8, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F0F4F8')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D6E4F0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=Transit_Pass_{kms_year or 'all'}.pdf"})


# ============ TP WEIGHT STOCK ============

@router.get("/govt-registers/tp-weight-stock")
async def get_tp_weight_stock(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Total paddy from TP Weight (vehicle_weights entries)"""
    query = {"tp_weight": {"$gt": 0}}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.vehicle_weights.find(query, {"_id": 0, "tp_weight": 1}).to_list(50000)
    total = round(sum(e.get("tp_weight", 0) or 0 for e in entries), 2)
    return {"total_tp_weight": total, "count": len(entries)}


# ============ MILLING REGISTER (Paddy/Rice Daily Ledger) ============

@router.get("/govt-registers/milling-register")
async def get_milling_register(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Auto-compute daily Milling Register from mill_entries + milling_entries + dc_deliveries"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    # Milling Register: OB = 0, only paddy_release feeds "Rcvd from CM A/c"
    ob_paddy = 0
    ob_rice = 0

    # 1. Paddy released daily (from paddy_release) = "Received from CM A/c"
    paddy_releases = await db.paddy_release.find(query, {"_id": 0, "date": 1, "qty_qtl": 1}).to_list(50000)
    daily_paddy_rcvd = {}
    for e in paddy_releases:
        d = e.get("date", "")
        if not d: continue
        daily_paddy_rcvd[d] = daily_paddy_rcvd.get(d, 0) + (e.get("qty_qtl", 0) or 0)

    # 2. Paddy milled + Rice produced (from milling_entries)
    milling = await db.milling_entries.find(query, {"_id": 0, "date": 1, "paddy_input_qntl": 1, "rice_qntl": 1, "cmr_delivery_qntl": 1}).to_list(50000)
    daily_milled = {}
    daily_rice_produced = {}
    for m in milling:
        d = m.get("date", "")
        if not d: continue
        daily_milled[d] = daily_milled.get(d, 0) + (m.get("paddy_input_qntl", 0) or 0)
        rice = m.get("cmr_delivery_qntl", 0) or m.get("rice_qntl", 0) or 0
        daily_rice_produced[d] = daily_rice_produced.get(d, 0) + rice

    # 3. Rice delivered (from dc_deliveries) — classification via parent dc_entries.delivery_to (FCI/RRC)
    del_query = {}
    if kms_year: del_query["kms_year"] = kms_year
    if season: del_query["season"] = season
    deliveries = await db.dc_deliveries.find(del_query, {"_id": 0, "date": 1, "quantity_qntl": 1, "godown_name": 1, "dc_id": 1}).to_list(50000)
    # Build dc_id -> delivery_to map from dc_entries
    dc_ids = list({d.get("dc_id") for d in deliveries if d.get("dc_id")})
    dc_map = {}
    if dc_ids:
        dc_docs = await db.dc_entries.find({"id": {"$in": dc_ids}}, {"_id": 0, "id": 1, "delivery_to": 1}).to_list(10000)
        dc_map = {d.get("id"): (d.get("delivery_to") or "").strip().upper() for d in dc_docs}
    daily_delivery_rrc = {}
    daily_delivery_fci = {}
    for dlv in deliveries:
        d = dlv.get("date", "")
        if not d: continue
        qty = dlv.get("quantity_qntl", 0) or 0
        # Prefer parent DC entry's delivery_to (authoritative). Fallback: godown_name substring.
        delivery_to = dc_map.get(dlv.get("dc_id"), "")
        if not delivery_to:
            godown = (dlv.get("godown_name", "") or "").lower()
            delivery_to = "FCI" if "fci" in godown else "RRC"
        if delivery_to == "FCI":
            daily_delivery_fci[d] = daily_delivery_fci.get(d, 0) + qty
        else:
            daily_delivery_rrc[d] = daily_delivery_rrc.get(d, 0) + qty

    # Collect all dates
    all_dates = sorted(set(list(daily_paddy_rcvd.keys()) + list(daily_milled.keys()) + list(daily_rice_produced.keys()) + list(daily_delivery_rrc.keys()) + list(daily_delivery_fci.keys())))

    # Build register rows with running balances
    rows = []
    prog_paddy_rcvd = 0
    prog_paddy_milled = 0
    prog_rice_milled = 0
    prog_rice_delivered = 0
    cb_paddy = ob_paddy
    cb_rice = ob_rice
    initial_ob_paddy = ob_paddy
    initial_ob_rice = ob_rice

    for date in all_dates:
        rcvd = round(daily_paddy_rcvd.get(date, 0), 2)
        milled = round(daily_milled.get(date, 0), 2)
        rice_prod = round(daily_rice_produced.get(date, 0), 2)
        del_rrc = round(daily_delivery_rrc.get(date, 0), 2)
        del_fci = round(daily_delivery_fci.get(date, 0), 2)

        ob_paddy = cb_paddy
        total_paddy = round(ob_paddy + rcvd, 2)
        cb_paddy = round(total_paddy - milled, 2)

        prog_paddy_rcvd = round(prog_paddy_rcvd + rcvd, 2)
        prog_paddy_milled = round(prog_paddy_milled + milled, 2)

        ob_rice = cb_rice
        total_rice = round(ob_rice + rice_prod, 2)
        total_del = round(del_rrc + del_fci, 2)
        cb_rice = round(total_rice - total_del, 2)

        prog_rice_milled = round(prog_rice_milled + rice_prod, 2)
        prog_rice_delivered = round(prog_rice_delivered + total_del, 2)

        try:
            from datetime import datetime as dt
            month = dt.strptime(date, "%Y-%m-%d").strftime("%B")
        except: month = ""

        rows.append({
            "date": date, "month": month,
            "ob_paddy": ob_paddy, "rcvd_from_cm": rcvd, "total_paddy": total_paddy,
            "issue_for_milling": milled,
            "prog_rcpt_paddy": prog_paddy_rcvd, "prog_milling_paddy": prog_paddy_milled,
            "cb_paddy": cb_paddy,
            "ob_rice": ob_rice, "rice_from_milling": rice_prod, "total_rice": total_rice,
            "delivery_rrc": del_rrc, "delivery_fci": del_fci,
            "prog_rice_milling": prog_rice_milled, "prog_rice_delivered": prog_rice_delivered,
            "cb_rice": cb_rice,
        })

    return {
        "rows": rows,
        "opening_stock": {"paddy": initial_ob_paddy, "rice": initial_ob_rice},
        "summary": {
            "total_paddy_received": prog_paddy_rcvd,
            "total_paddy_milled": prog_paddy_milled,
            "cb_paddy": cb_paddy,
            "total_rice_produced": prog_rice_milled,
            "total_rice_delivered": prog_rice_delivered,
            "cb_rice": cb_rice,
            "ob_paddy": initial_ob_paddy,
            "ob_rice": initial_ob_rice,
        }
    }


def _fmt_date_short(d):
    if not d: return ""
    try:
        if "T" in str(d): d = str(d).split("T")[0]
        p = str(d).split("-")
        if len(p) == 3: return f"{p[2]}/{p[1]}/{p[0]}"
    except: pass
    return str(d)


@router.get("/govt-registers/verification-report")
async def get_verification_report(
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    from_date: Optional[str] = None,  # previous verification date (YYYY-MM-DD)
    to_date: Optional[str] = None,    # current verification date (YYYY-MM-DD)
    last_meter_reading: float = 0,    # 4b. Metre Reading at last verification
    units_per_qtl: float = 6.0,       # config: electricity units per quintal
    rice_recovery: float = 0.67       # rice output ratio (67% FCI standard)
):
    """Compute FCI Weekly Verification Report from Milling Register data.
    - 4c. Present Metre Reading = 4b + 4d
    - 4d. Total units Consumed = Paddy Milled × units_per_qtl
    - Paddy / Rice / Delivery progressives derived from milling + paddy_release + dc_deliveries
    """
    q = {}
    if kms_year: q["kms_year"] = kms_year
    if season: q["season"] = season

    def _in_range(d):
        if not d: return (False, False)
        ds = str(d).split("T")[0]
        before = (from_date and ds <= from_date)
        current_week = (from_date and to_date and ds > from_date and ds <= to_date)
        return (before, current_week)

    # Aggregate (before = till previous verification, this_week = this reporting period)
    def agg(docs, field):
        before = 0.0; this_week = 0.0
        for x in docs:
            d = x.get("date", "")
            bf, tw = _in_range(d)
            v = x.get(field, 0) or 0
            if bf: before += v
            if tw: this_week += v
        return before, this_week

    milling = await db.milling_entries.find(q, {"_id": 0, "date": 1, "paddy_input_qntl": 1, "rice_qntl": 1, "cmr_delivery_qntl": 1}).to_list(50000)
    releases = await db.paddy_release.find(q, {"_id": 0, "date": 1, "qty_qtl": 1}).to_list(50000)
    deliveries = await db.dc_deliveries.find(q, {"_id": 0, "date": 1, "quantity_qntl": 1}).to_list(50000)

    # Normalize rice field for milling
    for m in milling:
        m["rice_out"] = m.get("cmr_delivery_qntl") or m.get("rice_qntl") or 0

    paddy_prev, paddy_week = agg(releases, "qty_qtl")
    milled_prev, milled_week = agg(milling, "paddy_input_qntl")
    rice_prev, rice_week = agg(milling, "rice_out")
    deliv_prev, deliv_week = agg(deliveries, "quantity_qntl")

    paddy_prog = paddy_prev + paddy_week
    milled_prog = milled_prev + milled_week
    rice_prog_milling = rice_prev + rice_week
    deliv_prog = deliv_prev + deliv_week

    # Electricity meter calculations (4c / 4d)
    units_consumed = round(milled_week * units_per_qtl, 2)
    present_meter = round((last_meter_reading or 0) + units_consumed, 2)

    # Book balance calculations
    book_balance_paddy = round(paddy_prog - milled_prog, 2)
    # Rice delivered progressive = deliv_prog (sum of all dc_deliveries till to_date)
    book_balance_rice = round(rice_prog_milling - deliv_prog, 2)

    # Rice recovery expected (formula-based cross-check)
    expected_rice_from_milling_week = round(milled_week * rice_recovery, 2)
    expected_rice_from_milling_prog = round(milled_prog * rice_recovery, 2)

    return {
        "period": {
            "from_date": from_date or "",
            "to_date": to_date or "",
            "kms_year": kms_year or "",
            "season": season or "",
        },
        "meter": {
            "last_reading": round(last_meter_reading or 0, 2),      # 4b
            "present_reading": present_meter,                        # 4c
            "units_consumed": units_consumed,                        # 4d
            "units_per_qtl": units_per_qtl,
        },
        "weekly": {
            "paddy_received": round(paddy_week, 2),
            "paddy_milled": round(milled_week, 2),
            "rice_produced": round(rice_week, 2),
            "rice_delivered": round(deliv_week, 2),
            "expected_rice": expected_rice_from_milling_week,
        },
        "progressive": {
            "paddy_received": round(paddy_prog, 2),
            "paddy_milled": round(milled_prog, 2),
            "rice_produced": round(rice_prog_milling, 2),
            "rice_delivered": round(deliv_prog, 2),
            "expected_rice": expected_rice_from_milling_prog,
        },
        "book_balance": {
            "paddy": book_balance_paddy,
            "rice": book_balance_rice,
        },
        "settings": {
            "units_per_qtl": units_per_qtl,
            "rice_recovery": rice_recovery,
        }
    }


# ============ ANNEXURE-1 FULL VERIFICATION REPORT ============
def _agency_key(agent_name: str, agency_field: str = "") -> str:
    """Map free-form agent/agency name to standard agency key."""
    if agency_field:
        k = agency_field.strip().upper()
        if k in ("OSCSC_OWN", "OSCSC_KORAPUT", "NAFED", "TDCC", "LEVY"):
            return k
    a = (agent_name or "").strip().lower()
    if not a: return "OSCSC_OWN"
    if "koraput" in a: return "OSCSC_KORAPUT"
    if "nafed" in a: return "NAFED"
    if "tdcc" in a: return "TDCC"
    if "levy" in a: return "LEVY"
    return "OSCSC_OWN"


@router.get("/govt-registers/verification-report/full")
async def get_verification_report_full(
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    from_date: Optional[str] = None,     # previous verification date (exclusive lower bound)
    to_date: Optional[str] = None,       # current verification date (inclusive upper bound)
    last_meter_reading: float = 0,
    units_per_qtl: float = 6.0,
    rice_recovery: float = 0.67,
    variety: str = "Boiled",
):
    """Annexure-1 Verification Report of Authorized Officer (full FCI format).
    Agency-wise split for Paddy (I-VI); RRC/FCI/RRC_FRK/FCI_FRK split for Rice (VII-XIV).
    """
    AGENCIES = ["OSCSC_OWN", "OSCSC_KORAPUT", "NAFED", "TDCC", "LEVY"]
    RICE_COLS = ["RRC", "FCI", "RRC_FRK", "FCI_FRK"]
    zero_agencies = lambda: {a: 0.0 for a in AGENCIES}
    zero_rice = lambda: {c: 0.0 for c in RICE_COLS}

    q = {}
    if kms_year: q["kms_year"] = kms_year
    if season: q["season"] = season

    def in_week(d):
        if not d: return False
        ds = str(d).split("T")[0]
        return (not from_date or ds > from_date) and (not to_date or ds <= to_date)

    def till_to(d):
        if not d: return False
        ds = str(d).split("T")[0]
        return (not to_date or ds <= to_date)

    # ===== I & II: Paddy Procured (from mill_entries.tp_weight grouped by agent_name) =====
    mill_docs = await db.mill_entries.find(q, {"_id": 0, "date": 1, "tp_weight": 1, "agent_name": 1}).to_list(100000)
    i_week = zero_agencies(); ii_prog = zero_agencies()
    for m in mill_docs:
        tpw = float(m.get("tp_weight", 0) or 0)
        if tpw <= 0: continue
        d = m.get("date", "")
        k = _agency_key(m.get("agent_name", ""))
        if till_to(d): ii_prog[k] += tpw
        if in_week(d): i_week[k] += tpw

    # ===== III & IV: Paddy Milled (milling_entries grouped by paddy_release agency ratio) =====
    milling = await db.milling_entries.find(q, {"_id": 0, "date": 1, "paddy_input_qntl": 1, "rice_qntl": 1, "cmr_delivery_qntl": 1}).to_list(100000)
    releases = await db.paddy_release.find(q, {"_id": 0, "date": 1, "qty_qtl": 1, "agency": 1}).to_list(100000)

    # Agency ratio of paddy_release till to_date
    rel_by_agency = zero_agencies()
    for r in releases:
        if till_to(r.get("date", "")):
            k = _agency_key("", r.get("agency", ""))
            rel_by_agency[k] += float(r.get("qty_qtl", 0) or 0)
    rel_total = sum(rel_by_agency.values()) or 0

    iii_week_total = 0.0; iv_prog_total = 0.0
    rice_vii_week = 0.0; rice_viii_prog = 0.0
    for m in milling:
        d = m.get("date", "")
        paddy = float(m.get("paddy_input_qntl", 0) or 0)
        rice = float(m.get("cmr_delivery_qntl", 0) or 0) or float(m.get("rice_qntl", 0) or 0)
        if till_to(d):
            iv_prog_total += paddy
            rice_viii_prog += rice
        if in_week(d):
            iii_week_total += paddy
            rice_vii_week += rice

    def split_by_ratio(total):
        out = zero_agencies()
        if rel_total > 0 and total > 0:
            for k in AGENCIES:
                out[k] = round(total * rel_by_agency[k] / rel_total, 2)
        elif total > 0:
            out["OSCSC_OWN"] = round(total, 2)
        return out

    iii_week = split_by_ratio(iii_week_total)
    iv_prog = split_by_ratio(iv_prog_total)

    # ===== V & VI: Book Balance of Paddy = II - IV =====
    v_book_bal = {k: round(ii_prog[k] - iv_prog[k], 2) for k in AGENCIES}
    vi_verified = dict(v_book_bal)  # same unless manually adjusted

    # ===== IX: Rice delivered during week against DC (split RRC/FCI/RRC_FRK/FCI_FRK) =====
    deliveries = await db.dc_deliveries.find(q, {"_id": 0, "date": 1, "quantity_qntl": 1, "dc_id": 1, "godown_name": 1}).to_list(100000)
    dc_ids = list({d.get("dc_id") for d in deliveries if d.get("dc_id")})
    dc_map = {}
    if dc_ids:
        dc_docs = await db.dc_entries.find({"id": {"$in": dc_ids}}, {"_id": 0, "id": 1, "delivery_to": 1}).to_list(10000)
        dc_map = {d.get("id"): (d.get("delivery_to") or "").strip().upper() for d in dc_docs}

    ix_week = zero_rice(); xi_prog_delivered = zero_rice()
    for d in deliveries:
        qty = float(d.get("quantity_qntl", 0) or 0)
        to = dc_map.get(d.get("dc_id"), "")
        if not to:
            godown = (d.get("godown_name", "") or "").lower()
            to = "FCI" if "fci" in godown else "RRC"
        col = to if to in ("FCI", "RRC") else "RRC"
        # TODO: FRK support (all non-FRK for now)
        if till_to(d.get("date", "")): xi_prog_delivered[col] += qty
        if in_week(d.get("date", "")): ix_week[col] += qty

    # ===== X: Progressive DC issued till verification =====
    dc_all = await db.dc_entries.find(q, {"_id": 0, "date": 1, "quantity_qntl": 1, "delivery_to": 1}).to_list(100000)
    x_prog_issued = zero_rice()
    for d in dc_all:
        if till_to(d.get("date", "")):
            to = (d.get("delivery_to") or "").strip().upper()
            col = to if to in ("FCI", "RRC") else "RRC"
            x_prog_issued[col] += float(d.get("quantity_qntl", 0) or 0)

    # XII: Balance undelivered against DC = X - XI
    xii_undelivered = {c: round(x_prog_issued[c] - xi_prog_delivered[c], 2) for c in RICE_COLS}
    # XIII: Book balance of rice = VIII - XI_prog (total delivered progressive)
    xiii_book_rice_total = round(rice_viii_prog - sum(xi_prog_delivered.values()), 2)
    xiv_verified_rice_total = xiii_book_rice_total

    def sum_a(d): return round(sum(d.values()), 2)
    def sum_r(d): return round(sum(d.values()), 2)

    # Meter calculations
    units_consumed = round(iii_week_total * units_per_qtl, 2)
    present_meter = round((last_meter_reading or 0) + units_consumed, 2)

    # Load miller details from branding + settings
    branding = await db.branding.find_one({}, {"_id": 0}) or {}
    settings = await db.app_settings.find_one({"setting_id": "verification_meter"}, {"_id": 0}) or {}

    def _round_dict(d): return {k: round(v, 2) for k, v in d.items()}

    return {
        "header": {
            "miller_name": branding.get("company_name", ""),
            "miller_code": branding.get("mill_code", ""),
            "address": branding.get("tagline", ""),
            "milling_capacity_mt": float(settings.get("milling_capacity_mt", 0) or 0),
            "kms_year": kms_year or "",
            "variety": variety or settings.get("variety", "Boiled"),
            "last_verification_date": from_date or "",
            "present_verification_date": to_date or "",
            "electricity_kw": float(settings.get("electricity_kw", 0) or 0),
            "electricity_kv": float(settings.get("electricity_kv", 0) or 0),
            "meter": {
                "last_reading": round(last_meter_reading or 0, 2),
                "present_reading": present_meter,
                "units_consumed": units_consumed,
            },
        },
        "agencies": AGENCIES,
        "rice_cols": RICE_COLS,
        "paddy": {
            "I_week":      {"by_agency": _round_dict(i_week),      "total": sum_a(i_week)},
            "II_prog":     {"by_agency": _round_dict(ii_prog),     "total": sum_a(ii_prog)},
            "III_week":    {"by_agency": _round_dict(iii_week),    "total": sum_a(iii_week)},
            "IV_prog":     {"by_agency": _round_dict(iv_prog),     "total": sum_a(iv_prog)},
            "V_book":      {"by_agency": _round_dict(v_book_bal),  "total": sum_a(v_book_bal)},
            "VI_verified": {"by_agency": _round_dict(vi_verified), "total": sum_a(vi_verified)},
        },
        "rice": {
            "VII_week":       {"total": round(rice_vii_week, 2)},
            "VIII_prog":      {"total": round(rice_viii_prog, 2)},
            "IX_week":        {"by_col": _round_dict(ix_week),          "total": sum_r(ix_week)},
            "X_prog_issued":  {"by_col": _round_dict(x_prog_issued),    "total": sum_r(x_prog_issued)},
            "XI_prog_delivered": {"by_col": _round_dict(xi_prog_delivered), "total": sum_r(xi_prog_delivered)},
            "XII_undelivered":{"by_col": _round_dict(xii_undelivered),  "total": sum_r(xii_undelivered)},
            "XIII_book":      {"total": xiii_book_rice_total},
            "XIV_verified":   {"total": xiv_verified_rice_total},
        },
        "settings": {
            "units_per_qtl": units_per_qtl,
            "rice_recovery": rice_recovery,
        }
    }


# ============ ANNEXURE-1 PDF EXPORT ============
@router.get("/govt-registers/verification-report/pdf")
async def export_verification_report_pdf(
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    last_meter_reading: float = 0,
    units_per_qtl: float = 6.0,
    rice_recovery: float = 0.67,
    variety: str = "Boiled",
):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors as rlcolors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from fastapi.responses import Response

    vr = await get_verification_report_full(
        kms_year=kms_year, season=season, from_date=from_date, to_date=to_date,
        last_meter_reading=last_meter_reading, units_per_qtl=units_per_qtl,
        rice_recovery=rice_recovery, variety=variety,
    )
    h = vr["header"]
    AGENCIES = vr["agencies"]; RICE_COLS = vr["rice_cols"]
    P = vr["paddy"]; R = vr["rice"]
    AGENCY_LABEL = {"OSCSC_OWN":"OSCSC(OWN)","OSCSC_KORAPUT":"OSCSC(Koraput)","NAFED":"NAFED","TDCC":"TDCC","LEVY":"Levy A/c"}
    RICE_LABEL = {"RRC":"RRC","FCI":"FCI","RRC_FRK":"RRC FRK","FCI_FRK":"FCI FRK"}

    def fmt_d(d):
        if not d: return ""
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0][2:]}" if len(p) == 3 else d
    def num(v):
        try: return f"{float(v or 0):.2f}"
        except: return "0.00"

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('t', parent=styles['Heading1'], fontSize=12, alignment=TA_CENTER, spaceAfter=6)
    foot_s = ParagraphStyle('f', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT, spaceBefore=2)

    story = []
    story.append(Paragraph("<u><b>Verification Report of Authorized Officer</b></u> &nbsp;&nbsp;&nbsp;&nbsp;<font size=7>Annexure-1</font>", title_s))

    header_tbl = [
        ["1a. Miller Name:", h["miller_name"], "1b. Address:", h["address"], "4a. Electricity Contract (KW)", str(h["electricity_kw"]), f"{h['electricity_kv']} KV"],
        ["1c. Miller Code:", h["miller_code"], "1d. Milling Capacity:", f"{h['milling_capacity_mt']} MT", "4b. Metre Reading at last verification", num(h["meter"]["last_reading"]), ""],
        ["2a. KMS:", h["kms_year"], "2b. Variety:", h["variety"], "4c. Present Metre Reading", num(h["meter"]["present_reading"]), ""],
        ["3a. Last Verification Date:", fmt_d(h["last_verification_date"]), "3b. Present Verification Date:", fmt_d(h["present_verification_date"]), "4d. Total units Consumed", num(h["meter"]["units_consumed"]), ""],
    ]
    ht = RLTable(header_tbl, colWidths=[32*mm, 28*mm, 30*mm, 26*mm, 42*mm, 18*mm, 14*mm])
    ht.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 7.5),
        ('GRID', (0,0), (-1,-1), 0.5, rlcolors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTNAME', (4,0), (4,-1), 'Helvetica-Bold'),
        ('ALIGN', (5,0), (6,-1), 'RIGHT'),
        ('TEXTCOLOR', (5,2), (5,3), rlcolors.red),
        ('FONTNAME', (5,2), (5,3), 'Helvetica-Bold'),
    ]))
    story.append(ht); story.append(Spacer(1, 4))

    # Unified Paddy + Rice Table (Annexure-1 nested format)
    # Columns: [Group] [Sl] [Particulars] [RRC] [FCI] [RRC FRK] [FCI FRK] [NAFED] [TDCC] [Levy] [TOTAL]
    # OSCSC(OWN) = RRC+FCI (colSpan=2), OSCSC(Koraput) = RRC FRK+FCI FRK (colSpan=2)
    main_header = ["", "Sl No", "", "OSCSC(OWN)", "", "OSCSC(Koraput)", "", "NAFED", "TDCC", "Levy A/c", "TOTAL"]
    data = [main_header]

    def paddy_row(sl, label, key):
        row = P.get(key, {"by_agency":{}, "total":0})
        # OSCSC_OWN value shown in RRC col with merge across FCI; OSCSC_KORAPUT same in RRC FRK col
        return ["", sl, label, num(row["by_agency"].get("OSCSC_OWN", 0)), "",
                num(row["by_agency"].get("OSCSC_KORAPUT", 0)), "",
                num(row["by_agency"].get("NAFED", 0)),
                num(row["by_agency"].get("TDCC", 0)),
                num(row["by_agency"].get("LEVY", 0)),
                num(row["total"])]

    paddy_red_rows = []
    paddy_def = [
        ("I", "Paddy Procured/Received during the week", "I_week", False),
        ("II", "Prog Paddy Procured/Recived till verification date", "II_prog", True),
        ("III", "Paddy Milled during the week", "III_week", False),
        ("IV", "Progressive paddy milled till verification date", "IV_prog", True),
        ("V", "Book Balance of Paddy Stock(sl No II-IV)", "V_book", True),
        ("VI", "Verified balance of paddy", "VI_verified", True),
    ]
    for (sl, label, key, colored) in paddy_def:
        data.append(paddy_row(sl, label, key))
        if colored: paddy_red_rows.append(len(data) - 1)

    PADDY_START = 1
    PADDY_END = len(data) - 1
    RICE_START = len(data)

    # VII, VIII: rice received (total under OSCSC_OWN colSpan=2)
    for (sl, label, key, colored) in [("VII", "Rice received from the milling during the week", "VII_week", True), ("VIII", "Progressive rice received from milling till date", "VIII_prog", True)]:
        total = R[key]["total"]
        data.append(["", sl, label, num(total), "", num(0), "", num(0), num(0), num(0), num(total)])
        if colored: paddy_red_rows.append(len(data) - 1)

    # Sub-header: RRC | FCI | RRC FRK | FCI FRK
    data.append(["", "", "", "RRC", "FCI", "RRC FRK", "FCI FRK", "", "", "", ""])
    SUBHEADER_ROW = len(data) - 1

    # IX-XII: rice delivery with RRC/FCI under OSCSC_OWN, RRC_FRK/FCI_FRK under OSCSC_KORAPUT
    rice2_def = [
        ("IX", "Rice delivered during the week against DC", "IX_week", False),
        ("X", "Progressive DC issued till verification", "X_prog_issued", False),
        ("XI", "Prog. Rice delivered against total DC issued", "XI_prog_delivered", True),
        ("XII", "Balance of rice remain undelivered against DC (Sl no x-xi)", "XII_undelivered", True),
    ]
    for (sl, label, key, colored) in rice2_def:
        row = R.get(key, {"by_col":{}, "total":0})
        data.append(["", sl, label,
                     num(row["by_col"].get("RRC", 0)),
                     num(row["by_col"].get("FCI", 0)),
                     num(row["by_col"].get("RRC_FRK", 0)),
                     num(row["by_col"].get("FCI_FRK", 0)),
                     "", "", "",
                     num(row["total"])])
        if colored: paddy_red_rows.append(len(data) - 1)

    # XIII, XIV: total only at RRC position
    for (sl, label, key) in [("XIII", "Book balannce of rice (Sl no viii-ix)", "XIII_book"), ("XIV", "Verified balance of rice", "XIV_verified")]:
        t = R[key]["total"]
        data.append(["", sl, label, num(t), "", "", "", "", "", "", num(t)])
        paddy_red_rows.append(len(data) - 1)
    RICE_END = len(data) - 1

    # Column widths (A4 portrait ~190mm usable): total = 190mm
    col_widths = [7*mm, 8*mm, 40*mm, 13*mm, 13*mm, 14*mm, 14*mm, 17*mm, 15*mm, 17*mm, 22*mm]
    tbl = RLTable(data, colWidths=col_widths, repeatRows=1)
    t_style = [
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 7.5),
        ('GRID', (0,0), (-1,-1), 0.5, rlcolors.black),
        ('BACKGROUND', (0,0), (-1,0), rlcolors.lightgrey),
        ('BACKGROUND', (3, SUBHEADER_ROW), (6, SUBHEADER_ROW), rlcolors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (3, SUBHEADER_ROW), (6, SUBHEADER_ROW), 'Helvetica-Bold'),
        # Header row: SPAN OSCSC(OWN) over cols 3-4, OSCSC(Koraput) over 5-6
        ('SPAN', (3, 0), (4, 0)),
        ('SPAN', (5, 0), (6, 0)),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('ALIGN', (1,1), (1,-1), 'CENTER'),
        ('ALIGN', (3,1), (-1,-1), 'RIGHT'),
        ('ALIGN', (3, SUBHEADER_ROW), (6, SUBHEADER_ROW), 'CENTER'),
        ('FONTNAME', (-1,1), (-1,-1), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        # Row-group labels
        ('SPAN', (0, PADDY_START), (0, PADDY_END)),
        ('SPAN', (0, RICE_START), (0, RICE_END)),
        ('BACKGROUND', (0, PADDY_START), (0, PADDY_END), rlcolors.HexColor('#f5f5f5')),
        ('BACKGROUND', (0, RICE_START), (0, RICE_END), rlcolors.HexColor('#f5f5f5')),
        ('FONTNAME', (0, PADDY_START), (0, RICE_END), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
    ]
    # Paddy rows I-VI: merge OSCSC_OWN (cols 3-4) and OSCSC_KORAPUT (cols 5-6) for each row
    for r in range(PADDY_START, PADDY_END + 1):
        t_style.append(('SPAN', (3, r), (4, r)))
        t_style.append(('SPAN', (5, r), (6, r)))
    # Rice rows VII-VIII: same (cols 3-4 merge + cols 5-6 merge)
    t_style.append(('SPAN', (3, RICE_START), (4, RICE_START)))
    t_style.append(('SPAN', (5, RICE_START), (6, RICE_START)))
    t_style.append(('SPAN', (3, RICE_START + 1), (4, RICE_START + 1)))
    t_style.append(('SPAN', (5, RICE_START + 1), (6, RICE_START + 1)))
    # Rice rows XIII-XIV: merge cols 3-6 to show single value
    xiii_row = RICE_END - 1; xiv_row = RICE_END
    t_style.append(('SPAN', (3, xiii_row), (6, xiii_row)))
    t_style.append(('SPAN', (3, xiv_row), (6, xiv_row)))
    t_style.append(('SPAN', (7, xiii_row), (9, xiii_row)))
    t_style.append(('SPAN', (7, xiv_row), (9, xiv_row)))
    # Red text for progressive rows
    for rnum in paddy_red_rows:
        t_style.append(('TEXTCOLOR', (3, rnum), (-1, rnum), rlcolors.red))
    tbl.setStyle(TableStyle(t_style))
    data[PADDY_START][0] = "Paddy"
    data[RICE_START][0] = "Rice"

    story.append(tbl); story.append(Spacer(1, 6))

    story.append(Paragraph("Total qty of CMB delivered as per M-reporting by the miller &nbsp;&nbsp; qtls", foot_s))
    story.append(Paragraph("It is certified that there is no missappropriation/diversion by the miller and paddy/rice available has been stored safely", foot_s))
    story.append(Spacer(1, 18))

    sig_tbl = RLTable([
        ["Name and Signature of Miller Agent/Authorised Representative", "Signature of Authorised Officer"],
        ["Copy Submitted to CSO cum District Manager, Kalahandi/Concerned Miller", "(With Name & Designation)"],
        ["*Milling Capacity per shift of 8 hrs", ""],
    ], colWidths=[95*mm, 95*mm])
    sig_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
    ]))
    story.append(sig_tbl)

    doc.build(story)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=Verification_Report_{to_date or 'current'}.pdf"})


# ============ ANNEXURE-1 EXCEL EXPORT ============
@router.get("/govt-registers/verification-report/excel")
async def export_verification_report_excel(
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    last_meter_reading: float = 0,
    units_per_qtl: float = 6.0,
    rice_recovery: float = 0.67,
    variety: str = "Boiled",
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response

    vr = await get_verification_report_full(
        kms_year=kms_year, season=season, from_date=from_date, to_date=to_date,
        last_meter_reading=last_meter_reading, units_per_qtl=units_per_qtl,
        rice_recovery=rice_recovery, variety=variety,
    )
    h = vr["header"]; P = vr["paddy"]; R = vr["rice"]

    def fmt_d(d):
        if not d: return ""
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0][2:]}" if len(p) == 3 else d

    wb = Workbook(); ws = wb.active; ws.title = "Verification Report"
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    subhdr_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    red_font = Font(color="C00000", bold=True, size=10)
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Cols: A=Group, B=Sl, C=Particulars, D=RRC, E=FCI, F=RRC FRK, G=FCI FRK, H=NAFED, I=TDCC, J=Levy, K=TOTAL
    col_widths = {'A': 6, 'B': 6, 'C': 48, 'D': 11, 'E': 11, 'F': 12, 'G': 12, 'H': 14, 'I': 10, 'J': 12, 'K': 15}
    for col, w in col_widths.items(): ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "Verification Report of Authorized Officer"
    ws['A1'].font = Font(bold=True, size=14, underline='single')
    ws['A1'].alignment = center
    ws['K1'] = "Annexure-1"
    ws['K1'].alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[1].height = 24

    # Miller Details Table (rows 2-5) - 7 cols
    miller_rows = [
        [("1a. Miller Name:", True), (h["miller_name"], False), ("1b. Address:", True), (h["address"], False), ("4a. Electricity Contract (KW)", True), (str(h["electricity_kw"]), False), (f"{h['electricity_kv']} KV", False)],
        [("1c. Miller Code:", True), (h["miller_code"], False), ("1d. Milling Capacity*:", True), (f"{h['milling_capacity_mt']} MT", False), ("4b. Metre Reading at last verification", True), (f"{h['meter']['last_reading']:.2f}", False), ("", False)],
        [("2a. KMS:", True), (h["kms_year"], False), ("2b. Variety:", True), (h["variety"], False), ("4c. Present Metre Reading", True), (f"{h['meter']['present_reading']:.2f}", "red"), ("", False)],
        [("3a. Last Verification Date:", True), (fmt_d(h["last_verification_date"]), False), ("3b. Present Verification Date:", True), (fmt_d(h["present_verification_date"]), False), ("4d. Total units Consumed", True), (f"{h['meter']['units_consumed']:.2f}", "red"), ("", False)],
    ]
    # Allocate cols A-C for miller info (left), D-F (mid), G-I for electricity, J-K (extra)
    for i, row in enumerate(miller_rows, start=2):
        # 7 cols: A-B (label+val), C-D (label+val), E (label), F (val), G (val_extra)
        positions = ['A', 'B', 'C', 'D', 'E', 'F', 'G']  # we'll merge
        # Simpler: merge A:B for label, C:D for value (pair 1); E:F for label pair 2; G:H for val2; I for label3; J for val3; K for unit
        # Actually simpler plain: use cols A, B, D, F, H, J, K (7 positions)
        ws.merge_cells(f'A{i}:B{i}'); ws[f'A{i}'] = row[0][0]; ws[f'A{i}'].font = bold_font if row[0][1] else normal_font; ws[f'A{i}'].border = border; ws[f'A{i}'].alignment = left
        ws.merge_cells(f'C{i}:D{i}'); ws[f'C{i}'] = row[1][0]; ws[f'C{i}'].font = normal_font; ws[f'C{i}'].border = border; ws[f'C{i}'].alignment = left
        ws.merge_cells(f'E{i}:F{i}'); ws[f'E{i}'] = row[2][0]; ws[f'E{i}'].font = bold_font; ws[f'E{i}'].border = border; ws[f'E{i}'].alignment = left
        ws.merge_cells(f'G{i}:H{i}'); ws[f'G{i}'] = row[3][0]; ws[f'G{i}'].font = normal_font; ws[f'G{i}'].border = border; ws[f'G{i}'].alignment = left
        ws[f'I{i}'] = row[4][0]; ws[f'I{i}'].font = bold_font; ws[f'I{i}'].border = border; ws[f'I{i}'].alignment = left
        ws[f'J{i}'] = row[5][0]; ws[f'J{i}'].font = red_font if row[5][1] == "red" else normal_font; ws[f'J{i}'].border = border; ws[f'J{i}'].alignment = right
        ws[f'K{i}'] = row[6][0]; ws[f'K{i}'].font = normal_font; ws[f'K{i}'].border = border; ws[f'K{i}'].alignment = right
        ws.row_dimensions[i].height = 22

    # Blank row
    hdr_row = 7

    # Main table header (row 7-8)
    # Row 7: [blank] [Sl No] [blank] [OSCSC(OWN) colspan=2] [OSCSC(Koraput) colspan=2] [NAFED] [TDCC] [Levy A/c] [TOTAL]
    ws[f'A{hdr_row}'] = ""; ws[f'B{hdr_row}'] = "Sl No"; ws[f'C{hdr_row}'] = ""
    ws.merge_cells(f'D{hdr_row}:E{hdr_row}'); ws[f'D{hdr_row}'] = "OSCSC(OWN)"
    ws.merge_cells(f'F{hdr_row}:G{hdr_row}'); ws[f'F{hdr_row}'] = "OSCSC(Koraput)"
    ws[f'H{hdr_row}'] = "NAFED"; ws[f'I{hdr_row}'] = "TDCC"; ws[f'J{hdr_row}'] = "Levy A/c"; ws[f'K{hdr_row}'] = "TOTAL"
    for col in "ABCDEFGHIJK":
        cell = ws[f'{col}{hdr_row}']; cell.font = bold_font; cell.alignment = center; cell.fill = hdr_fill; cell.border = border
    ws.row_dimensions[hdr_row].height = 20

    # Data rows
    cur = hdr_row + 1

    def put_paddy(sl, label, key, colored):
        nonlocal cur
        row = P.get(key, {"by_agency":{}, "total":0})
        own_v = row["by_agency"].get("OSCSC_OWN", 0)
        kor_v = row["by_agency"].get("OSCSC_KORAPUT", 0)
        nafed_v = row["by_agency"].get("NAFED", 0)
        tdcc_v = row["by_agency"].get("TDCC", 0)
        levy_v = row["by_agency"].get("LEVY", 0)
        tot_v = row["total"]
        ws[f'B{cur}'] = sl; ws[f'B{cur}'].font = bold_font; ws[f'B{cur}'].alignment = center; ws[f'B{cur}'].border = border
        ws[f'C{cur}'] = label; ws[f'C{cur}'].font = normal_font; ws[f'C{cur}'].alignment = left; ws[f'C{cur}'].border = border
        ws.merge_cells(f'D{cur}:E{cur}'); ws[f'D{cur}'] = own_v; ws[f'D{cur}'].font = red_font if colored else normal_font; ws[f'D{cur}'].alignment = right; ws[f'D{cur}'].border = border; ws[f'D{cur}'].number_format = '0.00'
        ws.merge_cells(f'F{cur}:G{cur}'); ws[f'F{cur}'] = kor_v; ws[f'F{cur}'].font = red_font if colored else normal_font; ws[f'F{cur}'].alignment = right; ws[f'F{cur}'].border = border; ws[f'F{cur}'].number_format = '0.00'
        for col, val in [('H', nafed_v), ('I', tdcc_v), ('J', levy_v)]:
            ws[f'{col}{cur}'] = val; ws[f'{col}{cur}'].font = red_font if colored else normal_font; ws[f'{col}{cur}'].alignment = right; ws[f'{col}{cur}'].border = border; ws[f'{col}{cur}'].number_format = '0.00'
        ws[f'K{cur}'] = tot_v; ws[f'K{cur}'].font = Font(bold=True, color="C00000") if colored else bold_font; ws[f'K{cur}'].alignment = right; ws[f'K{cur}'].border = border; ws[f'K{cur}'].number_format = '0.00'
        ws.row_dimensions[cur].height = 22
        cur += 1

    paddy_start_row = cur
    paddy_def = [
        ("I", "Paddy Procured/Received during the week", "I_week", False),
        ("II", "Prog Paddy Procured/Recived till verification date", "II_prog", True),
        ("III", "Paddy Milled during the week", "III_week", False),
        ("IV", "Progressive paddy milled till verification date", "IV_prog", True),
        ("V", "Book Balance of Paddy Stock(sl No II-IV)", "V_book", True),
        ("VI", "Verified balance of paddy", "VI_verified", True),
    ]
    for (sl, label, key, colored) in paddy_def: put_paddy(sl, label, key, colored)
    paddy_end_row = cur - 1
    # Merge A column for "Paddy" label
    ws.merge_cells(f'A{paddy_start_row}:A{paddy_end_row}')
    ws[f'A{paddy_start_row}'] = "Paddy"; ws[f'A{paddy_start_row}'].font = bold_font; ws[f'A{paddy_start_row}'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=90); ws[f'A{paddy_start_row}'].fill = subhdr_fill
    for r in range(paddy_start_row, paddy_end_row + 1): ws[f'A{r}'].border = border

    # VII, VIII: rice received (agency cols)
    rice_start_row = cur
    for (sl, label, key) in [("VII", "Rice received from the milling during the week", "VII_week"), ("VIII", "Progressive rice received from milling till date", "VIII_prog")]:
        total = R[key]["total"]
        ws[f'B{cur}'] = sl; ws[f'B{cur}'].font = bold_font; ws[f'B{cur}'].alignment = center; ws[f'B{cur}'].border = border
        ws[f'C{cur}'] = label; ws[f'C{cur}'].font = normal_font; ws[f'C{cur}'].alignment = left; ws[f'C{cur}'].border = border
        ws.merge_cells(f'D{cur}:E{cur}'); ws[f'D{cur}'] = total; ws[f'D{cur}'].font = red_font; ws[f'D{cur}'].alignment = right; ws[f'D{cur}'].border = border; ws[f'D{cur}'].number_format = '0.00'
        ws.merge_cells(f'F{cur}:G{cur}'); ws[f'F{cur}'] = 0; ws[f'F{cur}'].alignment = right; ws[f'F{cur}'].border = border; ws[f'F{cur}'].number_format = '0.00'
        for col in ['H', 'I', 'J']:
            ws[f'{col}{cur}'] = 0; ws[f'{col}{cur}'].alignment = right; ws[f'{col}{cur}'].border = border; ws[f'{col}{cur}'].number_format = '0.00'
        ws[f'K{cur}'] = total; ws[f'K{cur}'].font = Font(bold=True, color="C00000"); ws[f'K{cur}'].alignment = right; ws[f'K{cur}'].border = border; ws[f'K{cur}'].number_format = '0.00'
        ws.row_dimensions[cur].height = 22
        cur += 1

    # Sub-header row: RRC | FCI | RRC FRK | FCI FRK
    ws[f'B{cur}'] = ""; ws[f'C{cur}'] = ""
    for col, lbl in [('D', 'RRC'), ('E', 'FCI'), ('F', 'RRC FRK'), ('G', 'FCI FRK')]:
        ws[f'{col}{cur}'] = lbl; ws[f'{col}{cur}'].font = bold_font; ws[f'{col}{cur}'].alignment = center; ws[f'{col}{cur}'].fill = subhdr_fill; ws[f'{col}{cur}'].border = border
    for col in ['A','B','C','H','I','J','K']:
        ws[f'{col}{cur}'].border = border
    ws.row_dimensions[cur].height = 20
    cur += 1

    # IX-XII: rice delivery
    rice2_def = [
        ("IX", "Rice delivered during the week against DC", "IX_week", False),
        ("X", "Progressive DC issued till verification", "X_prog_issued", False),
        ("XI", "Prog. Rice delivered against total DC issued", "XI_prog_delivered", True),
        ("XII", "Balance of rice remain undelivered against DC (Sl no x-xi)", "XII_undelivered", True),
    ]
    for (sl, label, key, colored) in rice2_def:
        row = R.get(key, {"by_col":{}, "total":0})
        ws[f'B{cur}'] = sl; ws[f'B{cur}'].font = bold_font; ws[f'B{cur}'].alignment = center; ws[f'B{cur}'].border = border
        ws[f'C{cur}'] = label; ws[f'C{cur}'].font = normal_font; ws[f'C{cur}'].alignment = left; ws[f'C{cur}'].border = border
        for col, ck in [('D', 'RRC'), ('E', 'FCI'), ('F', 'RRC_FRK'), ('G', 'FCI_FRK')]:
            ws[f'{col}{cur}'] = row["by_col"].get(ck, 0); ws[f'{col}{cur}'].font = red_font if colored else normal_font; ws[f'{col}{cur}'].alignment = right; ws[f'{col}{cur}'].border = border; ws[f'{col}{cur}'].number_format = '0.00'
        for col in ['H', 'I', 'J']:
            ws[f'{col}{cur}'].border = border
        ws[f'K{cur}'] = row["total"]; ws[f'K{cur}'].font = Font(bold=True, color="C00000") if colored else bold_font; ws[f'K{cur}'].alignment = right; ws[f'K{cur}'].border = border; ws[f'K{cur}'].number_format = '0.00'
        ws.row_dimensions[cur].height = 30 if sl == "XII" else 22
        cur += 1

    # XIII, XIV
    for (sl, label, key) in [("XIII", "Book balannce of rice (Sl no viii-ix)", "XIII_book"), ("XIV", "Verified balance of rice", "XIV_verified")]:
        t = R[key]["total"]
        ws[f'B{cur}'] = sl; ws[f'B{cur}'].font = bold_font; ws[f'B{cur}'].alignment = center; ws[f'B{cur}'].border = border
        ws[f'C{cur}'] = label; ws[f'C{cur}'].font = normal_font; ws[f'C{cur}'].alignment = left; ws[f'C{cur}'].border = border
        ws[f'D{cur}'] = t; ws[f'D{cur}'].font = red_font; ws[f'D{cur}'].alignment = right; ws[f'D{cur}'].border = border; ws[f'D{cur}'].number_format = '0.00'
        for col in ['E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{col}{cur}'].border = border
        ws[f'K{cur}'] = t; ws[f'K{cur}'].font = Font(bold=True, color="C00000"); ws[f'K{cur}'].alignment = right; ws[f'K{cur}'].border = border; ws[f'K{cur}'].number_format = '0.00'
        ws.row_dimensions[cur].height = 22
        cur += 1
    rice_end_row = cur - 1

    # Merge A column for "Rice" label
    ws.merge_cells(f'A{rice_start_row}:A{rice_end_row}')
    ws[f'A{rice_start_row}'] = "Rice"; ws[f'A{rice_start_row}'].font = bold_font; ws[f'A{rice_start_row}'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=90); ws[f'A{rice_start_row}'].fill = subhdr_fill
    for r in range(rice_start_row, rice_end_row + 1): ws[f'A{r}'].border = border

    # Footer text
    cur += 1
    ws[f'A{cur}'] = "Total qty of CMB delivered as per M-reporting by the miller   qtls"; ws[f'A{cur}'].font = normal_font
    cur += 1
    ws[f'A{cur}'] = "It is certified that there is no missappropriation /diversion by the miller and paddy/rice available has been stored safely"; ws[f'A{cur}'].font = normal_font
    cur += 2
    ws[f'A{cur}'] = "Name and Signature of Miller Agent/Authorised Representative"; ws[f'A{cur}'].font = bold_font
    ws[f'I{cur}'] = "Signature of Authorised Officer"; ws[f'I{cur}'].font = bold_font
    cur += 1
    ws[f'A{cur}'] = "Copy Submitted to CSO cum District Manager, Kalahandi/Concerned Miller"; ws[f'A{cur}'].font = normal_font
    ws[f'I{cur}'] = "(With Name & Desgination)"; ws[f'I{cur}'].font = normal_font
    cur += 1
    ws[f'A{cur}'] = "*Milling Capacity per shift of 8 hrs"; ws[f'A{cur}'].font = Font(size=9, italic=True, color="666666")

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=Verification_Report_{to_date or 'current'}.xlsx"})



@router.get("/govt-registers/milling-register/excel")
async def export_milling_register_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response

    reg_data = await get_milling_register(kms_year, season)
    rows = reg_data["rows"]
    branding = await db.branding.find_one({}, {"_id": 0}) or {}
    company = branding.get("company_name", "Rice Mill")
    tagline = branding.get("tagline", "")
    custom_fields = branding.get("custom_fields", [])

    wb = Workbook(); ws = wb.active; ws.title = "Milling Register"
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    green_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Row 1: Company Name
    ws.merge_cells('A1:R1')
    c1 = ws.cell(row=1, column=1, value=company.upper())
    c1.font = Font(bold=True, size=14, color="1F4E79"); c1.alignment = Alignment(horizontal='center')

    # Row 2: Tagline + custom fields
    info_parts = [tagline] if tagline else []
    for cf in custom_fields:
        info_parts.append(f"{cf.get('label','')}: {cf.get('value','')}")
    if info_parts:
        ws.merge_cells('A2:R2')
        c2 = ws.cell(row=2, column=1, value="  |  ".join(info_parts))
        c2.font = Font(size=9, color="666666"); c2.alignment = Alignment(horizontal='center')

    # Row 3: Title
    title = "MILLING REGISTER"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    ws.merge_cells('A3:Q3')
    c3 = ws.cell(row=3, column=1, value=title)
    c3.font = Font(bold=True, size=12, color="FFFFFF"); c3.fill = header_fill; c3.alignment = Alignment(horizontal='center')

    # Row 5: Column headers (single color, no partition)
    headers = ['Date', 'Milling Month', 'OB Paddy', 'Rcvd from CM A/c', 'Total Paddy',
        'Issue For Milling', 'Prog Rcpt of Paddy', 'Prog Milling of Paddy', 'CB of Paddy',
        'OB Rice', 'Rice Rcpt from Milling', 'Total Rice',
        'Rice Delivery RRC', 'Rice Delivery FCI', 'Prog Rice Milling', 'Prog Rice Delivered', 'CB of Rice']
    widths = [11, 10, 10, 14, 10, 12, 14, 14, 10, 8, 14, 10, 12, 12, 12, 12, 10]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = Font(bold=True, size=8, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = border

    def _v(val):
        """Return '-' for empty/zero values"""
        if val is None or val == "" or val == 0:
            return "-"
        return val

    # Data rows
    for idx, r in enumerate(rows):
        row_num = 6 + idx
        fill = alt_fill if idx % 2 == 0 else None
        vals = [_fmt_date_short(r.get("date","")), r.get("month","") or "-",
            _v(r.get("ob_paddy",0)), _v(r.get("rcvd_from_cm",0)), _v(r.get("total_paddy",0)),
            _v(r.get("issue_for_milling",0)), _v(r.get("prog_rcpt_paddy",0)), _v(r.get("prog_milling_paddy",0)), _v(r.get("cb_paddy",0)),
            _v(r.get("ob_rice",0)), _v(r.get("rice_from_milling",0)), _v(r.get("total_rice",0)),
            _v(r.get("delivery_rrc",0)), _v(r.get("delivery_fci",0)), _v(r.get("prog_rice_milling",0)), _v(r.get("prog_rice_delivered",0)), _v(r.get("cb_rice",0))]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font = Font(size=9); cell.border = border
            if ci >= 3: cell.alignment = Alignment(horizontal='right')
            if fill: cell.fill = fill
            if ci in (9, 17): cell.font = Font(size=9, bold=True)

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.page_setup.orientation = 'landscape'; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    fn = f"milling_register_{kms_year or 'all'}.xlsx"
    return Response(content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/govt-registers/milling-register/pdf")
async def export_milling_register_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from fastapi.responses import Response

    reg_data = await get_milling_register(kms_year, season)
    rows = reg_data["rows"]
    summary = reg_data["summary"]
    branding = await db.branding.find_one({}, {"_id": 0}) or {}
    company = branding.get("company_name", "Rice Mill")
    tagline = branding.get("tagline", "")
    custom_fields = branding.get("custom_fields", [])

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10, rightMargin=10, topMargin=12, bottomMargin=12)
    elements = []
    styles = getSampleStyleSheet()

    # Header
    co_style = ParagraphStyle('Co', parent=styles['Title'], fontSize=13, textColor=colors.HexColor('#1F4E79'), spaceAfter=1, alignment=1)
    elements.append(Paragraph(company.upper(), co_style))
    info_parts = [tagline] if tagline else []
    for cf in custom_fields:
        info_parts.append(f"{cf.get('label','')}: {cf.get('value','')}")
    if info_parts:
        addr_style = ParagraphStyle('Addr', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#666666'), spaceAfter=2, alignment=1)
        elements.append(Paragraph("  |  ".join(info_parts), addr_style))

    title = "MILLING REGISTER"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    t_style = ParagraphStyle('T', parent=styles['Heading2'], fontSize=10, textColor=colors.white,
        backColor=colors.HexColor('#2E75B6'), spaceAfter=4, alignment=1, borderPadding=(2,2,2,2))
    elements.append(Paragraph(title, t_style))
    elements.append(Spacer(1, 2))

    headers = ['Date', 'Month', 'OB\nPaddy', 'Rcvd from\nCM A/c', 'Total\nPaddy',
        'Issue For\nMilling', 'Prog Rcpt\nPaddy', 'Prog Mill\nPaddy', 'CB\nPaddy',
        'OB\nRice', 'Rice Rcpt\nMilling', 'Total\nRice',
        'Delivery\nRRC', 'Delivery\nFCI', 'Prog Rice\nMilling', 'Prog Rice\nDelivered', 'CB\nRice']
    col_widths = [42, 28, 36, 48, 38, 42, 45, 45, 38, 32, 45, 38, 40, 40, 45, 45, 38]

    # Auto-fit
    usable = 818
    total_w = sum(col_widths)
    if total_w > usable:
        scale = usable / total_w
        col_widths = [round(w * scale) for w in col_widths]

    def _pv(val):
        if val is None or val == "" or val == 0:
            return "-"
        return val

    data = [headers]
    for r in rows:
        data.append([_fmt_date_short(r.get("date","")), r.get("month","") or "-",
            _pv(r.get("ob_paddy",0)), _pv(r.get("rcvd_from_cm",0)), _pv(r.get("total_paddy",0)),
            _pv(r.get("issue_for_milling",0)), _pv(r.get("prog_rcpt_paddy",0)), _pv(r.get("prog_milling_paddy",0)), _pv(r.get("cb_paddy",0)),
            _pv(r.get("ob_rice",0)), _pv(r.get("rice_from_milling",0)), _pv(r.get("total_rice",0)),
            _pv(r.get("delivery_rrc",0)), _pv(r.get("delivery_fci",0)), _pv(r.get("prog_rice_milling",0)), _pv(r.get("prog_rice_delivered",0)), _pv(r.get("cb_rice",0))])

    nrows = len(data)
    table = RLTable(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 5.5),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF1F8')]),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]
    # Bold CB columns (10th=paddy CB, 18th=rice CB)
    for ri in range(1, nrows):
        style_cmds.append(('FONTNAME', (9, ri), (9, ri), 'Helvetica-Bold'))
        style_cmds.append(('FONTNAME', (17, ri), (17, ri), 'Helvetica-Bold'))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    # Summary
    elements.append(Spacer(1, 4))
    s = summary
    sum_style = ParagraphStyle('Sum', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#1F4E79'))
    elements.append(Paragraph(
        f"<b>Summary:</b> Paddy Received: {s.get('total_paddy_received',0)} Q  |  Milled: {s.get('total_paddy_milled',0)} Q  |  CB Paddy: <b>{s.get('cb_paddy',0)}</b> Q  ||  "
        f"Rice Produced: {s.get('total_rice_produced',0)} Q  |  Delivered: {s.get('total_rice_delivered',0)} Q  |  CB Rice: <b>{s.get('cb_rice',0)}</b> Q", sum_style))

    from datetime import datetime
    elements.append(Spacer(1, 3))
    gen_style = ParagraphStyle('Gen', parent=styles['Normal'], fontSize=6, textColor=colors.HexColor('#999999'))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", gen_style))

    doc.build(elements); buffer.seek(0)
    fn = f"milling_register_{kms_year or 'all'}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


# ============ CMR DELIVERY TRACKER WITH OTR ============

@router.get("/govt-registers/cmr-delivery")
async def get_cmr_deliveries(kms_year: Optional[str] = None, season: Optional[str] = None,
                              date_from: Optional[str] = None, date_to: Optional[str] = None):
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    date_q = {}
    if date_from:
        date_q["$gte"] = date_from
    if date_to:
        date_q["$lte"] = date_to
    if date_q:
        query["date"] = date_q

    entries = await db.cmr_deliveries.find(query, {"_id": 0}).sort("date", 1).to_list(50000)

    # Get total paddy received for OTR calculation
    paddy_query = {}
    if kms_year:
        paddy_query["kms_year"] = kms_year
    if season:
        paddy_query["season"] = season
    paddy_pipeline = [
        {"$match": paddy_query},
        {"$group": {"_id": None, "total_paddy": {"$sum": "$final_w"}}}
    ]
    paddy_result = await db.mill_entries.aggregate(paddy_pipeline).to_list(1)
    total_paddy = (paddy_result[0]["total_paddy"] if paddy_result else 0) / 100  # final_w is KG, convert to QNTL

    total_cmr = sum(float(e.get("cmr_qty", 0) or 0) for e in entries)
    otr = round((total_cmr / total_paddy * 100), 2) if total_paddy > 0 else 0

    return {
        "entries": entries,
        "summary": {
            "total_cmr_delivered": round(total_cmr, 2),
            "total_paddy_received": round(total_paddy, 2),
            "outturn_ratio": otr,
            "total_deliveries": len(entries),
            "total_bags": sum(int(e.get("bags", 0) or 0) for e in entries),
        }
    }


@router.post("/govt-registers/cmr-delivery")
async def create_cmr_delivery(data: dict, username: str = ""):
    doc = {
        "id": str(uuid.uuid4()),
        "date": data.get("date", ""),
        "kms_year": data.get("kms_year", ""),
        "season": data.get("season", ""),
        "delivery_no": data.get("delivery_no", ""),
        "rrc_depot": data.get("rrc_depot", ""),
        "rice_type": data.get("rice_type", "Parboiled"),
        "cmr_qty": float(data.get("cmr_qty", 0) or 0),
        "bags": int(data.get("bags", 0) or 0),
        "vehicle_no": data.get("vehicle_no", ""),
        "driver_name": data.get("driver_name", ""),
        "fortified": data.get("fortified", True),
        "gate_pass_no": data.get("gate_pass_no", ""),
        "quality_grade": data.get("quality_grade", "FAQ"),
        "remark": data.get("remark", ""),
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.cmr_deliveries.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/govt-registers/cmr-delivery/{entry_id}")
async def update_cmr_delivery(entry_id: str, data: dict, username: str = ""):
    existing = await db.cmr_deliveries.find_one({"id": entry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="CMR delivery not found")
    update_data = {
        "date": data.get("date", existing.get("date", "")),
        "delivery_no": data.get("delivery_no", existing.get("delivery_no", "")),
        "rrc_depot": data.get("rrc_depot", existing.get("rrc_depot", "")),
        "rice_type": data.get("rice_type", existing.get("rice_type", "")),
        "cmr_qty": float(data.get("cmr_qty", existing.get("cmr_qty", 0)) or 0),
        "bags": int(data.get("bags", existing.get("bags", 0)) or 0),
        "vehicle_no": data.get("vehicle_no", existing.get("vehicle_no", "")),
        "driver_name": data.get("driver_name", existing.get("driver_name", "")),
        "fortified": data.get("fortified", existing.get("fortified", True)),
        "gate_pass_no": data.get("gate_pass_no", existing.get("gate_pass_no", "")),
        "quality_grade": data.get("quality_grade", existing.get("quality_grade", "")),
        "remark": data.get("remark", existing.get("remark", "")),
        "updated_by": username,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.cmr_deliveries.update_one({"id": entry_id}, {"$set": update_data})
    return {"success": True}


@router.delete("/govt-registers/cmr-delivery/{entry_id}")
async def delete_cmr_delivery(entry_id: str, username: str = "", role: str = ""):
    result = await db.cmr_deliveries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="CMR delivery not found")
    return {"success": True}


@router.get("/govt-registers/cmr-delivery/excel")
async def export_cmr_delivery_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                                     date_from: Optional[str] = None, date_to: Optional[str] = None):
    data = await get_cmr_deliveries(kms_year, season, date_from, date_to)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "CMR Delivery"

    headers = ["Date", "Delivery No.", "RRC/Depot", "Rice Type", "CMR Qty (Qtl)", "Bags", "Vehicle No.", "Fortified", "Grade", "Remarks"]
    col_widths = [14, 16, 22, 16, 18, 10, 16, 12, 10, 20]
    data_rows = []
    for e in data["entries"]:
        data_rows.append([
            fmt_date(e.get("date", "")), e.get("delivery_no", ""), e.get("rrc_depot", ""),
            e.get("rice_type", ""), e.get("cmr_qty", 0), e.get("bags", 0),
            e.get("vehicle_no", ""), "Yes (+F)" if e.get("fortified") else "No",
            e.get("quality_grade", ""), e.get("remark", "")
        ])
    s = data["summary"]
    data_rows.append(["TOTAL", f'{s["total_deliveries"]} deliveries', "", "", s["total_cmr_delivered"], s["total_bags"], "", "", f'OTR: {s["outturn_ratio"]}%', ""])

    style_govt_excel(ws, "CMR Delivery Tracker", headers, data_rows, col_widths,
                     company, f"CMR Delivery Register with OTR | {kms_year or 'All'} {season or ''} | OTR: {s['outturn_ratio']}%")

    total_row_num = len(data_rows) + 3
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_num, column=ci)
        cell.font = total_font
        cell.fill = total_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=CMR_Delivery_{kms_year or 'all'}.xlsx"})


# ============ SECURITY DEPOSIT MANAGEMENT ============

@router.get("/govt-registers/security-deposit")
async def get_security_deposits(kms_year: Optional[str] = None):
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    entries = await db.security_deposits.find(query, {"_id": 0}).sort("issue_date", -1).to_list(50000)
    # Auto-check expiry status
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for e in entries:
        if e.get("status") == "active" and e.get("expiry_date", "") < today and e.get("expiry_date", ""):
            e["status"] = "expired"
    total_amount = sum(float(e.get("amount", 0) or 0) for e in entries if e.get("status") in ("active", None))
    return {
        "entries": entries,
        "summary": {
            "total_deposits": len(entries),
            "active_count": sum(1 for e in entries if e.get("status") == "active"),
            "total_active_amount": round(total_amount, 2),
            "released_count": sum(1 for e in entries if e.get("status") == "released"),
            "expired_count": sum(1 for e in entries if e.get("status") == "expired"),
        }
    }


@router.post("/govt-registers/security-deposit")
async def create_security_deposit(data: dict, username: str = ""):
    doc = {
        "id": str(uuid.uuid4()),
        "kms_year": data.get("kms_year", ""),
        "bg_number": data.get("bg_number", ""),
        "bank_name": data.get("bank_name", ""),
        "amount": float(data.get("amount", 0) or 0),
        "sd_ratio": data.get("sd_ratio", "1:6"),
        "milling_capacity_mt": float(data.get("milling_capacity_mt", 0) or 0),
        "issue_date": data.get("issue_date", ""),
        "expiry_date": data.get("expiry_date", ""),
        "status": data.get("status", "active"),
        "miller_type": data.get("miller_type", "regular"),
        "remark": data.get("remark", ""),
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.security_deposits.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/govt-registers/security-deposit/{entry_id}")
async def update_security_deposit(entry_id: str, data: dict, username: str = ""):
    existing = await db.security_deposits.find_one({"id": entry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Security deposit not found")
    update_data = {
        "bg_number": data.get("bg_number", existing.get("bg_number", "")),
        "bank_name": data.get("bank_name", existing.get("bank_name", "")),
        "amount": float(data.get("amount", existing.get("amount", 0)) or 0),
        "sd_ratio": data.get("sd_ratio", existing.get("sd_ratio", "")),
        "milling_capacity_mt": float(data.get("milling_capacity_mt", existing.get("milling_capacity_mt", 0)) or 0),
        "issue_date": data.get("issue_date", existing.get("issue_date", "")),
        "expiry_date": data.get("expiry_date", existing.get("expiry_date", "")),
        "status": data.get("status", existing.get("status", "")),
        "miller_type": data.get("miller_type", existing.get("miller_type", "")),
        "remark": data.get("remark", existing.get("remark", "")),
        "updated_by": username,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.security_deposits.update_one({"id": entry_id}, {"$set": update_data})
    return {"success": True}


@router.delete("/govt-registers/security-deposit/{entry_id}")
async def delete_security_deposit(entry_id: str, username: str = "", role: str = ""):
    result = await db.security_deposits.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Security deposit not found")
    return {"success": True}


@router.get("/govt-registers/security-deposit/excel")
async def export_security_deposit_excel(kms_year: Optional[str] = None):
    data = await get_security_deposits(kms_year)
    company, _ = await get_company_name()
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Deposit"

    headers = ["BG Number", "Bank Name", "Amount (Rs)", "SD Ratio", "Capacity (MT)", "Issue Date", "Expiry Date", "Status", "Miller Type", "Remarks"]
    col_widths = [18, 24, 18, 12, 16, 14, 14, 14, 16, 20]
    data_rows = []
    for e in data["entries"]:
        data_rows.append([
            e.get("bg_number", ""), e.get("bank_name", ""), e.get("amount", 0),
            e.get("sd_ratio", ""), e.get("milling_capacity_mt", 0),
            fmt_date(e.get("issue_date", "")), fmt_date(e.get("expiry_date", "")),
            (e.get("status", "") or "").upper(), e.get("miller_type", ""), e.get("remark", "")
        ])
    s = data["summary"]
    data_rows.append([f'{s["total_deposits"]} total', "", s["total_active_amount"], "", "", "", "", f'Active: {s["active_count"]}', "", ""])

    style_govt_excel(ws, "Security Deposit Register", headers, data_rows, col_widths,
                     company, f"Security Deposit (Bank Guarantee) Register | {kms_year or 'All'}")

    total_row_num = len(data_rows) + 3
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_num, column=ci)
        cell.font = total_font
        cell.fill = total_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=Security_Deposit_{kms_year or 'all'}.xlsx"})
