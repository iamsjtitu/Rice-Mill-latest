from models import round_amount
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime
from database import db
from utils.report_helper import get_columns, fmt_val, get_pdf_headers, get_pdf_widths_mm, get_excel_headers, get_entry_row
from utils.date_format import fmt_date
import io

router = APIRouter()

@router.get("/reports/daily")
async def get_daily_report(date: str, kms_year: Optional[str] = None, season: Optional[str] = None, mode: str = "normal"):
    q_date = {"date": date}
    q_fy = {}
    if kms_year: q_fy["kms_year"] = kms_year
    if season: q_fy["season"] = season
    q = {**q_date, **q_fy}

    # Paddy Entries
    entries = await db.mill_entries.find(q, {"_id": 0}).to_list(500)
    total_paddy_kg = sum(e.get("kg", 0) for e in entries)
    total_paddy_bags = sum(e.get("bag", 0) for e in entries)
    total_final_w = sum(e.get("final_w", 0) for e in entries)

    # Private Paddy
    pvt_paddy = await db.private_paddy.find(q, {"_id": 0}).to_list(500)
    pvt_paddy_qntl = round(sum(e.get("final_qntl", 0) or e.get("qntl", 0) or e.get("quantity_qntl", 0) or 0 for e in pvt_paddy), 2)
    pvt_paddy_amount = sum(e.get("total_amount", 0) for e in pvt_paddy)

    # Sale Vouchers
    sale_vouchers = await db.sale_vouchers.find(q, {"_id": 0}).to_list(500)
    sv_total_amount = sum(sv.get("total", 0) or sv.get("subtotal", 0) for sv in sale_vouchers)
    sv_total_items = sum(len(sv.get("items", [])) for sv in sale_vouchers)

    # Purchase Vouchers
    purchase_vouchers_daily = await db.purchase_vouchers.find(q, {"_id": 0}).to_list(500)
    pv_total_amount = sum(pv.get("total", 0) or pv.get("subtotal", 0) for pv in purchase_vouchers_daily)
    pv_total_items = sum(len(pv.get("items", [])) for pv in purchase_vouchers_daily)

    # Rice Sales
    rice_sales = await db.rice_sales.find(q, {"_id": 0}).to_list(500)
    rice_sale_qntl = sum(s.get("quantity_qntl", 0) for s in rice_sales)
    rice_sale_amount = sum(s.get("total_amount", 0) for s in rice_sales)

    # Milling
    milling = await db.milling_entries.find(q, {"_id": 0}).to_list(500)
    milling_paddy_input = sum(m.get("paddy_input_qntl", 0) for m in milling)
    milling_rice_output = sum(m.get("rice_qntl", 0) for m in milling)
    milling_frk_used = sum(m.get("frk_used_qntl", 0) for m in milling)

    # DC Deliveries
    dc_deliveries = await db.dc_deliveries.find(q_date, {"_id": 0}).to_list(500)
    dc_delivery_qntl = sum(d.get("quantity_qntl", 0) for d in dc_deliveries)

    # Diesel / Pump Account
    diesel_txns = await db.diesel_accounts.find(q, {"_id": 0}).to_list(500)
    diesel_total_amount = sum(t.get("amount", 0) for t in diesel_txns if t.get("txn_type") in ("diesel", "debit"))
    diesel_acc_paid = sum(t.get("amount", 0) for t in diesel_txns if t.get("txn_type") in ("payment", "credit"))

    # Cash Book
    cash_txns = await db.cash_transactions.find(q, {"_id": 0}).to_list(500)
    # Also check ledger nikasi for diesel payments (handles manual cashbook payments)
    diesel_ledger_paid = sum(t.get("amount", 0) for t in cash_txns
        if t.get("account") == "ledger" and t.get("txn_type") == "nikasi"
        and ((t.get("party_type") or "") == "Diesel" or (t.get("reference") or "").startswith("diesel_pay")))
    diesel_total_paid = max(diesel_acc_paid, diesel_ledger_paid)
    cash_jama = sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "jama" and t.get("account") == "cash")
    cash_nikasi = sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "nikasi" and t.get("account") == "cash")
    bank_jama = sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "jama" and t.get("account") == "bank")
    bank_nikasi = sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "nikasi" and t.get("account") == "bank")

    # MSP Payments
    msp = await db.msp_payments.find(q, {"_id": 0}).to_list(500)
    msp_amount = sum(p.get("amount", 0) for p in msp)

    # Private Payments
    pvt_payments = await db.private_payments.find(q_date, {"_id": 0}).to_list(500)
    pvt_paid = sum(p.get("amount", 0) for p in pvt_payments if p.get("ref_type") == "paddy_purchase")
    pvt_received = sum(p.get("amount", 0) for p in pvt_payments if p.get("ref_type") == "rice_sale")

    # By-product Sales (from bp_sale_register)
    bp_sales = await db.bp_sale_register.find(q, {"_id": 0}).to_list(500)
    bp_amount = sum(s.get("total", 0) or s.get("amount", 0) for s in bp_sales)

    # FRK Purchases
    frk = await db.frk_purchases.find(q, {"_id": 0}).to_list(500)
    frk_qntl = sum(f.get("quantity_qntl", 0) for f in frk)
    frk_amount = sum(f.get("total_amount", 0) for f in frk)

    # Mill Parts Stock
    parts_txns = await db.mill_parts_stock.find(q_date, {"_id": 0}).to_list(500)
    parts_in = [t for t in parts_txns if t.get("txn_type") == "in"]
    parts_used = [t for t in parts_txns if t.get("txn_type") == "used"]
    parts_in_amount = sum(t.get("total_amount", 0) for t in parts_in)

    # Hemali Payments
    hemali_q = {"date": date}
    if kms_year:
        hemali_q["kms_year"] = kms_year
    if season:
        hemali_q["season"] = season
    hemali_payments = await db.hemali_payments.find(hemali_q, {"_id": 0}).to_list(500)

    # Paddy Cutting (Chalna) - today's entries + cumulative totals
    cutting_q = {"date": date}
    if kms_year: cutting_q["kms_year"] = kms_year
    if season: cutting_q["season"] = season
    paddy_cutting = await db.paddy_cutting.find(cutting_q, {"_id": 0}).to_list(500)
    cutting_bags = sum(int(c.get("bags_cut", 0) or 0) for c in paddy_cutting)
    # Cumulative: all cutting till date + total bags from entries
    cum_cut_q = {}
    if kms_year: cum_cut_q["kms_year"] = kms_year
    if season: cum_cut_q["season"] = season
    all_cutting = await db.paddy_cutting.find(cum_cut_q, {"_id": 0, "bags_cut": 1}).to_list(50000)
    cum_total_cut = sum(int(c.get("bags_cut", 0) or 0) for c in all_cutting)
    cum_mill_q = {}
    if kms_year: cum_mill_q["kms_year"] = kms_year
    if season: cum_mill_q["season"] = season
    cum_mill_entries = await db.mill_entries.find(cum_mill_q, {"_id": 0, "bag": 1, "plastic_bag": 1}).to_list(50000)
    cum_total_received = sum(int(e.get("bag", 0) or 0) for e in cum_mill_entries) + sum(int(e.get("plastic_bag", 0) or 0) for e in cum_mill_entries)
    cum_remaining = cum_total_received - cum_total_cut
    hemali_paid = [h for h in hemali_payments if h.get("status") == "paid"]
    hemali_unpaid = [h for h in hemali_payments if h.get("status") != "paid"]
    hemali_total_paid = sum(h.get("amount_paid", 0) for h in hemali_paid)
    hemali_total_work = sum(h.get("total", 0) for h in hemali_paid)

    # Staff Attendance - always show all active staff
    staff_att = await db.staff_attendance.find(q_date, {"_id": 0}).to_list(500)
    all_staff = await db.staff.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(500)
    att_map_local = {a["staff_id"]: a["status"] for a in staff_att}
    staff_details = []
    present_c = absent_c = half_c = holiday_c = not_marked_c = 0
    for s in all_staff:
        status = att_map_local.get(s["id"], "not_marked")
        staff_details.append({"name": s["name"], "status": status})
        if status == "present": present_c += 1
        elif status == "absent": absent_c += 1
        elif status == "half_day": half_c += 1
        elif status == "holiday": holiday_c += 1
        else: not_marked_c += 1

    # ══ v104.44.18 — P0 NEW SECTIONS ══
    # Vehicle Weight (Auto Vehicle Weight) — Sale + Purchase trips with Bag Type + Bhada
    vw_entries = await db.vehicle_weights.find(q_date, {"_id": 0}).to_list(500)
    vw_sale = [v for v in vw_entries if (v.get("trans_type") or "").lower() in ("dispatch", "dispatch(sale)", "sale")]
    vw_purchase = [v for v in vw_entries if (v.get("trans_type") or "").lower() in ("receive", "receive(purchase)", "purchase")]
    vw_sale_bhada_total = sum(float(v.get("bhada", 0) or 0) for v in vw_sale)
    vw_sale_bags = sum(int(v.get("bags", 0) or 0) for v in vw_sale)
    vw_sale_net = sum(float(v.get("net_weight", 0) or 0) for v in vw_sale)
    vw_purchase_bhada_total = sum(float(v.get("bhada", 0) or 0) for v in vw_purchase)
    vw_purchase_bags = sum(int(v.get("bags", 0) or 0) for v in vw_purchase)
    vw_purchase_net = sum(float(v.get("net_weight", 0) or 0) for v in vw_purchase)

    # Truck Owner Per-Trip Bhada — trips with bhada from VW (today's date), settled/partial/pending by truck
    # Simple view: aggregate today's VW trips grouped by vehicle_no, along with bhada amounts
    pertrip_by_truck = {}
    for v in vw_entries:
        vn = (v.get("vehicle_no") or "").strip().upper()
        if not vn:
            continue
        bhada = float(v.get("bhada", 0) or 0)
        if bhada <= 0:
            continue
        row = pertrip_by_truck.setdefault(vn, {"vehicle_no": vn, "trips": 0, "bhada": 0.0})
        row["trips"] += 1
        row["bhada"] += bhada
    # Payments made to truck_owner type parties today (any truck) as a total
    truck_owner_paid = sum(float(t.get("amount", 0) or 0) for t in cash_txns
        if (t.get("party_type") or "").lower() == "truck" and t.get("txn_type") == "nikasi")
    per_trip_trucks_list = list(pertrip_by_truck.values())
    per_trip_trucks_list.sort(key=lambda x: x["bhada"], reverse=True)

    # Truck / Agent / LocalParty payment summaries — derived from cash_txns
    def _party_txn_summary(party_type_label: str):
        filt = [t for t in cash_txns if (t.get("party_type") or "").lower() == party_type_label.lower()]
        jama = sum(float(t.get("amount", 0) or 0) for t in filt if t.get("txn_type") == "jama")
        nikasi = sum(float(t.get("amount", 0) or 0) for t in filt if t.get("txn_type") == "nikasi")
        return {
            "count": len(filt),
            "jama": round(jama, 2), "nikasi": round(nikasi, 2),
            "net": round(jama - nikasi, 2),
            "details": [{
                "party": t.get("category", "") or t.get("party_name", ""),
                "txn_type": t.get("txn_type", ""),
                "amount": round(float(t.get("amount", 0) or 0), 2),
                "account": t.get("account", ""),
                "description": t.get("description", ""),
            } for t in filt],
        }
    truck_payments_summary = _party_txn_summary("Truck")
    agent_payments_summary = _party_txn_summary("Agent")
    localparty_payments_summary = _party_txn_summary("LocalParty")

    # ══ v104.44.19 — P1 NEW SECTIONS ══
    # Leased Truck Payments (today's `truck_lease_payments`)
    lease_payments_today = await db.truck_lease_payments.find(q_date, {"_id": 0}).to_list(500)
    # Enrich with truck_no from active leases
    active_leases = await db.truck_leases.find({}, {"_id": 0}).to_list(500)
    lease_map = {l.get("id", ""): l for l in active_leases}
    lease_total_paid = sum(float(p.get("amount", 0) or 0) for p in lease_payments_today)

    # Oil Premium (Lab Test Report) — today's entries
    oil_premium_today = await db.oil_premium.find(q_date, {"_id": 0}).to_list(500)
    oil_prem_total = sum(float(op.get("premium_amount", 0) or 0) for op in oil_premium_today)
    oil_prem_pos = sum(1 for op in oil_premium_today if float(op.get("premium_amount", 0) or 0) > 0)
    oil_prem_neg = sum(1 for op in oil_premium_today if float(op.get("premium_amount", 0) or 0) < 0)

    is_detail = mode == "detail"

    # Build entry_id -> mandi_name map for diesel mandi lookup
    _entry_mandi_map = {e.get("id", ""): e.get("mandi_name", "") for e in entries}

    result = {
        "date": date, "mode": mode,
        "paddy_entries": {
            "count": len(entries), "total_kg": round(total_paddy_kg, 2),
            "total_bags": total_paddy_bags, "total_final_w": round(total_final_w, 2),
            "total_tp_weight": round(sum(float(e.get("tp_weight", 0) or 0) for e in entries), 2),
            "total_mill_w": round(sum(e.get("mill_w", 0) for e in entries), 2),
            "total_g_deposite": sum(e.get("g_deposite", 0) for e in entries),
            "total_g_issued": sum(e.get("g_issued", 0) for e in entries),
            "total_cash_paid": round(sum(e.get("cash_paid", 0) for e in entries), 2),
            "total_diesel_paid": round(sum(e.get("diesel_paid", 0) for e in entries), 2),
            "details": [{"truck_no": e.get("truck_no", ""), "agent": e.get("agent_name", ""),
                "mandi": e.get("mandi_name", ""), "rst_no": e.get("rst_no", ""),
                "tp_no": e.get("tp_no", ""), "tp_weight": float(e.get("tp_weight", 0) or 0),
                "season": e.get("season", ""),
                "kg": e.get("kg", 0), "qntl": e.get("qntl", 0), "bags": e.get("bag", 0),
                "g_deposite": e.get("g_deposite", 0), "gbw_cut": e.get("gbw_cut", 0),
                "mill_w": e.get("mill_w", 0),
                "moisture": e.get("moisture", 0),
                "moisture_cut": e.get("moisture_cut", 0),
                "cutting_percent": e.get("cutting_percent", 0),
                "disc_dust_poll": e.get("disc_dust_poll", 0),
                "final_w": e.get("final_w", 0),
                "plastic_bag": e.get("plastic_bag", 0),
                "p_pkt_cut": e.get("p_pkt_cut", 0),
                "g_issued": e.get("g_issued", 0),
                "cash_paid": e.get("cash_paid", 0),
                "diesel_paid": e.get("diesel_paid", 0)} for e in entries]
        },
        "pvt_paddy": {
            "count": len(pvt_paddy), "total_qntl": round(pvt_paddy_qntl, 2),
            "total_amount": round(pvt_paddy_amount, 2),
            "details": [{"party": p.get("party_name", ""), "mandi": p.get("mandi_name", ""),
                "truck_no": p.get("truck_no", ""), "qntl": round(p.get("final_qntl", 0) or p.get("qntl", 0) or p.get("quantity_qntl", 0) or 0, 2),
                "rate": p.get("rate_per_qntl", 0) or p.get("rate", 0),
                "amount": p.get("total_amount", 0), "bag": p.get("bag", 0),
                "cash_paid": p.get("cash_paid", 0), "diesel_paid": p.get("diesel_paid", 0)} for p in pvt_paddy] if is_detail else
                [{"party": p.get("party_name", ""), "mandi": p.get("mandi_name", ""),
                "qntl": round(p.get("final_qntl", 0) or p.get("qntl", 0) or p.get("quantity_qntl", 0) or 0, 2), "amount": p.get("total_amount", 0)} for p in pvt_paddy]
        },
        "rice_sales": {
            "count": len(rice_sales), "total_qntl": round(rice_sale_qntl, 2),
            "total_amount": round(rice_sale_amount, 2),
            "details": [{"party": s.get("party_name", ""), "qntl": s.get("quantity_qntl", 0),
                "type": s.get("rice_type", ""), "rate": s.get("rate", 0),
                "amount": s.get("total_amount", 0), "vehicle": s.get("vehicle_no", "")} for s in rice_sales] if is_detail else
                [{"party": s.get("party_name", ""), "qntl": s.get("quantity_qntl", 0),
                "type": s.get("rice_type", ""), "amount": s.get("total_amount", 0)} for s in rice_sales]
        },
        "milling": {
            "count": len(milling), "paddy_input_qntl": round(milling_paddy_input, 2),
            "rice_output_qntl": round(milling_rice_output, 2),
            "frk_used_qntl": round(milling_frk_used, 2),
            "details": [{"paddy_in": m.get("paddy_input_qntl", 0), "rice_out": m.get("rice_qntl", 0),
                "type": m.get("rice_type", ""), "frk": m.get("frk_used_qntl", 0),
                "cmr_ready": m.get("cmr_delivery_qntl", 0), "outturn": m.get("outturn_pct", 0)} for m in milling] if is_detail else
                [{"paddy_in": m.get("paddy_input_qntl", 0), "rice_out": m.get("rice_qntl", 0),
                "type": m.get("rice_type", "")} for m in milling]
        },
        "dc_deliveries": {
            "count": len(dc_deliveries), "total_qntl": round(dc_delivery_qntl, 2),
            "details": [{"dc_no": d.get("dc_no", ""), "godown": d.get("godown", ""),
                "vehicle": d.get("vehicle_no", ""), "qntl": d.get("quantity_qntl", 0),
                "bags": d.get("bags", 0)} for d in dc_deliveries] if is_detail else []
        },
        "cash_flow": {
            "cash_jama": round(cash_jama, 2), "cash_nikasi": round(cash_nikasi, 2),
            "bank_jama": round(bank_jama, 2), "bank_nikasi": round(bank_nikasi, 2),
            "net_cash": round(cash_jama - cash_nikasi, 2),
            "net_bank": round(bank_jama - bank_nikasi, 2),
            "details": [{"desc": t.get("description", ""), "type": t.get("txn_type", ""),
                "account": t.get("account", ""), "category": t.get("category", ""),
                "amount": t.get("amount", 0), "party": t.get("party_name", "")} for t in cash_txns]
        },
        "payments": {
            "msp_received": round(msp_amount, 2),
            "pvt_paddy_paid": round(pvt_paid, 2),
            "rice_sale_received": round(pvt_received, 2),
            "msp_details": [{"dc_no": p.get("dc_number", ""), "amount": p.get("amount", 0),
                "qntl": p.get("quantity_qntl", 0), "rate": p.get("rate_per_qntl", 0),
                "mode": p.get("payment_mode", "")} for p in msp] if is_detail else [],
            "pvt_payment_details": [{"party": p.get("party_name", ""), "amount": p.get("amount", 0),
                "ref_type": p.get("ref_type", ""), "mode": p.get("payment_mode", "")} for p in pvt_payments] if is_detail else []
        },
        "byproducts": {
            "count": len(bp_sales), "total_amount": round(bp_amount, 2),
            "details": [{"product": s.get("product", ""), "voucher_no": s.get("voucher_no", ""),
                "bill_number": s.get("bill_number", ""), "billing_date": s.get("billing_date", ""),
                "rst_no": s.get("rst_no", ""), "vehicle_no": s.get("vehicle_no", ""),
                "bill_from": s.get("bill_from", ""), "party_name": s.get("party_name", ""),
                "destination": s.get("destination", ""),
                "net_weight_kg": s.get("net_weight_kg", 0), "bags": s.get("bags", 0),
                "rate_per_qtl": s.get("rate_per_qtl", 0), "amount": s.get("amount", 0),
                "total": s.get("total", 0) or s.get("amount", 0)} for s in bp_sales] if is_detail else []
        },
        "frk": {
            "count": len(frk), "total_qntl": round(frk_qntl, 2), "total_amount": round(frk_amount, 2),
            "details": [{"party": f.get("party_name", ""), "qntl": f.get("quantity_qntl", 0),
                "rate": f.get("rate", 0), "amount": f.get("total_amount", 0)} for f in frk] if is_detail else []
        },
        "sale_vouchers": {
            "count": len(sale_vouchers), "total_amount": round(sv_total_amount, 2),
            "total_items": sv_total_items,
            "details": [{
                "voucher_no": sv.get("voucher_no", ""),
                "party": sv.get("party_name", ""),
                "truck_no": sv.get("truck_no", ""),
                "items": [{
                    "name": it.get("item_name", ""),
                    "qty": round(it.get("quantity", 0), 2),
                    "rate": it.get("rate", 0),
                    "amount": it.get("amount", 0)
                } for it in sv.get("items", [])],
                "total": sv.get("total", 0) or sv.get("subtotal", 0),
                "advance": sv.get("advance", 0) or sv.get("paid_amount", 0),
                "balance": sv.get("balance", 0)
            } for sv in sale_vouchers]
        },
        "purchase_vouchers": {
            "count": len(purchase_vouchers_daily), "total_amount": round(pv_total_amount, 2),
            "total_items": pv_total_items,
            "details": [{
                "voucher_no": pv.get("voucher_no", ""),
                "party": pv.get("party_name", ""),
                "truck_no": pv.get("truck_no", ""),
                "items": [{
                    "name": it.get("item_name", ""),
                    "qty": round(it.get("quantity", 0), 2),
                    "rate": it.get("rate", 0),
                    "amount": it.get("amount", 0)
                } for it in pv.get("items", [])],
                "total": pv.get("total", 0) or pv.get("subtotal", 0),
                "advance": pv.get("advance", 0) or pv.get("paid_amount", 0),
                "balance": pv.get("balance", 0)
            } for pv in purchase_vouchers_daily]
        },
        "mill_parts": {
            "in_count": len(parts_in), "used_count": len(parts_used),
            "in_amount": round(parts_in_amount, 2),
            "in_details": [{"part": t.get("part_name", ""), "qty": t.get("quantity", 0),
                "rate": t.get("rate", 0), "party": t.get("party_name", ""),
                "bill_no": t.get("bill_no", ""), "amount": t.get("total_amount", 0),
                "store_room": t.get("store_room_name", "")} for t in parts_in],
            "used_details": [{"part": t.get("part_name", ""), "qty": t.get("quantity", 0),
                "remark": t.get("remark", ""), "store_room": t.get("store_room_name", "")} for t in parts_used]
        },
        "staff_attendance": {
            "total": len(all_staff),
            "present": present_c, "absent": absent_c,
            "half_day": half_c, "holiday": holiday_c, "not_marked": not_marked_c,
            "details": staff_details
        },
        "hemali_payments": {
            "count": len(hemali_payments),
            "paid_count": len(hemali_paid),
            "unpaid_count": len(hemali_unpaid),
            "total_work": round(hemali_total_work, 2),
            "total_paid": round(hemali_total_paid, 2),
            "details": [{
                "sardar": h.get("sardar_name", ""),
                "items": ", ".join(f"{i.get('item_name','')} x{i.get('quantity',0)}" for i in h.get("items", [])),
                "total": h.get("total", 0),
                "advance_deducted": h.get("advance_deducted", 0),
                "amount_paid": h.get("amount_paid", 0),
                "new_advance": h.get("new_advance", 0),
                "status": h.get("status", ""),
            } for h in hemali_payments],
        },
        "pump_account": {
            "total_diesel": round(diesel_total_amount, 2),
            "total_paid": round(diesel_total_paid, 2),
            "balance": round(diesel_total_amount - diesel_total_paid, 2),
            "details": [{"pump": t.get("pump_name", ""), "txn_type": t.get("txn_type", ""),
                "amount": t.get("amount", 0), "truck_no": t.get("truck_no", ""),
                "mandi": t.get("mandi_name", "") or _entry_mandi_map.get(t.get("linked_entry_id", ""), "") or (t.get("description", "").split("Mandi ")[-1] if "Mandi " in t.get("description", "") else ""),
                "desc": t.get("description", "")} for t in diesel_txns]
        },
        "paddy_cutting": {
            "count": len(paddy_cutting), "total_bags_cut": cutting_bags,
            "cum_total_received": cum_total_received, "cum_total_cut": cum_total_cut, "cum_remaining": cum_remaining,
            "details": [{"bags_cut": c.get("bags_cut", 0), "remark": c.get("remark", "")} for c in paddy_cutting]
        },
        "cash_transactions": {
            "count": len([t for t in cash_txns if t.get("account") == "cash"]),
            "total_jama": round(sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "jama" and t.get("account") == "cash"), 2),
            "total_nikasi": round(sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "nikasi" and t.get("account") == "cash"), 2),
            "details": [{
                "date": t.get("date", date),
                "party_name": t.get("category", ""),
                "party_type": t.get("party_type", ""),
                "txn_type": t.get("txn_type", ""),
                "amount": round_amount(t.get("amount", 0)),
                "description": t.get("description", ""),
                "payment_mode": "Cash"
            } for t in cash_txns if t.get("account") == "cash"]
        },
        # ══ v104.44.18 — P0 New Sections ══
        "vehicle_weight": {
            "sale_count": len(vw_sale),
            "sale_bags": vw_sale_bags,
            "sale_net_qntl": round(vw_sale_net / 100, 2),
            "sale_bhada_total": round(vw_sale_bhada_total, 2),
            "purchase_count": len(vw_purchase),
            "purchase_bags": vw_purchase_bags,
            "purchase_net_qntl": round(vw_purchase_net / 100, 2),
            "purchase_bhada_total": round(vw_purchase_bhada_total, 2),
            "sale_details": [{
                "rst_no": v.get("rst_no", ""),
                "vehicle_no": v.get("vehicle_no", ""),
                "party": v.get("party_name", ""),
                "destination": v.get("destination", "") or v.get("mandi_name", ""),
                "product": v.get("product", ""),
                "bags": v.get("bags", 0),
                "bag_type": v.get("bag_type", ""),
                "net_wt": v.get("net_weight", 0),
                "bhada": float(v.get("bhada", 0) or 0),
                "remark": v.get("remark", ""),
            } for v in vw_sale],
            "purchase_details": [{
                "rst_no": v.get("rst_no", ""),
                "vehicle_no": v.get("vehicle_no", ""),
                "party": v.get("party_name", ""),
                "mandi": v.get("mandi_name", ""),
                "product": v.get("product", ""),
                "bags": v.get("bags", 0),
                "net_wt": v.get("net_weight", 0),
                "bhada": float(v.get("bhada", 0) or 0),
                "remark": v.get("remark", ""),
            } for v in vw_purchase] if is_detail else [],
        },
        "per_trip_bhada": {
            "truck_count": len(per_trip_trucks_list),
            "trip_count": sum(x["trips"] for x in per_trip_trucks_list),
            "bhada_total": round(sum(x["bhada"] for x in per_trip_trucks_list), 2),
            "paid_today": round(truck_owner_paid, 2),
            "pending_today": round(sum(x["bhada"] for x in per_trip_trucks_list) - truck_owner_paid, 2),
            "details": [{
                "vehicle_no": x["vehicle_no"],
                "trips": x["trips"],
                "bhada": round(x["bhada"], 2),
            } for x in per_trip_trucks_list],
        },
        "truck_payments": truck_payments_summary,
        "agent_payments": agent_payments_summary,
        "local_party_payments": localparty_payments_summary,
        # ══ v104.44.19 — P1 New Sections ══
        "leased_truck": {
            "count": len(lease_payments_today),
            "total_paid": round(lease_total_paid, 2),
            "details": [{
                "truck_no": (lease_map.get(p.get("lease_id", ""), {}) or {}).get("truck_no", "") or p.get("truck_no", ""),
                "owner": (lease_map.get(p.get("lease_id", ""), {}) or {}).get("owner_name", "") or p.get("owner_name", ""),
                "amount": round(float(p.get("amount", 0) or 0), 2),
                "payment_type": p.get("payment_type", ""),
                "mode": p.get("mode", "") or p.get("payment_mode", ""),
                "remark": p.get("remark", "") or p.get("description", ""),
            } for p in lease_payments_today],
        },
        "oil_premium": {
            "count": len(oil_premium_today),
            "total_premium": round(oil_prem_total, 2),
            "positive_count": oil_prem_pos,
            "negative_count": oil_prem_neg,
            "details": [{
                "voucher_no": op.get("voucher_no", ""),
                "rst_no": op.get("rst_no", ""),
                "party": op.get("party_name", "") or op.get("buyer_name", ""),
                "qty_qntl": op.get("qty_qtl", 0) or op.get("qty_qntl", 0) or op.get("quantity_qntl", 0),
                "rate": op.get("rate", 0) or op.get("sauda_amount", 0),
                "diff_pct": op.get("difference_pct", 0) or op.get("diff_pct", 0) or op.get("diff_percent", 0),
                "premium_amount": round(float(op.get("premium_amount", 0) or 0), 2),
                "remark": op.get("remark", ""),
            } for op in oil_premium_today],
        },
    }
    return result


def _fmt_amt(val):
    if val == 0: return "0"
    return f"{val:,.0f}"


def _fmt_date(d):
    if not d: return ''
    parts = str(d).split('-')
    return f"{parts[2]}-{parts[1]}-{parts[0]}" if len(parts) == 3 else d


@router.get("/reports/daily/pdf")
async def export_daily_pdf(date: str, kms_year: Optional[str] = None, season: Optional[str] = None, mode: str = "normal"):
    data = await get_daily_report(date, kms_year, season, mode)

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer, HRFlowable
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm

    is_detail = mode == "detail"
    buf = io.BytesIO()
    if is_detail:
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15, rightMargin=15, topMargin=20, bottomMargin=20)
    else:
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=25, rightMargin=25, topMargin=20, bottomMargin=20)
    styles = get_pdf_styles()
    elements = []

    from utils.export_helpers import get_pdf_table_style
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())

    # Custom styles
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1a365d'), spaceAfter=4)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=9, textColor=colors.grey, spaceAfter=8)
    section_style = ParagraphStyle('SectionHead', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#1a365d'),
        spaceBefore=12, spaceAfter=4, borderWidth=0, leftIndent=0)

    from utils.export_helpers import get_pdf_table_style

    hdr_bg = colors.HexColor('#1a365d')
    hdr_font_color = colors.white
    green = colors.HexColor('#166534')
    red = colors.HexColor('#991b1b')
    border_color = colors.HexColor('#cbd5e1')

    def make_table(headers, rows, col_widths=None, font_size=7):
        from reportlab.platypus import Paragraph as P
        from reportlab.lib.styles import ParagraphStyle as PS
        hdr_ps = PS('TblHdr', fontName='FreeSansBold', fontSize=font_size, textColor=colors.white,
                     alignment=1, leading=font_size + 2)
        cell_ps = PS('TblCell', fontName='FreeSans', fontSize=font_size, leading=font_size + 2)
        # Wrap headers and cells in Paragraphs so text wraps properly
        hdr_row = [P(str(h), hdr_ps) for h in headers]
        wrapped_rows = []
        for r in rows:
            wrapped_rows.append([P(str(v), cell_ps) for v in r])
        data_rows = [hdr_row] + wrapped_rows
        t = RTable(data_rows, colWidths=col_widths, repeatRows=1)
        style_cmds = get_pdf_table_style(len(data_rows))
        style_cmds.append(('FONTSIZE', (0, 0), (-1, -1), font_size))
        style_cmds.append(('FONTSIZE', (0, 0), (-1, 0), font_size))
        style_cmds.append(('TOPPADDING', (0, 0), (-1, -1), 2))
        style_cmds.append(('BOTTOMPADDING', (0, 0), (-1, -1), 2))
        style_cmds.append(('LEFTPADDING', (0, 0), (-1, -1), 2))
        style_cmds.append(('RIGHTPADDING', (0, 0), (-1, -1), 2))
        t.setStyle(TableStyle(style_cmds))
        return t

    def summary_row(label, value, color_val=None):
        return [Paragraph(f"<b>{label}</b>", styles['Normal']),
                Paragraph(f"<b>Rs. {_fmt_amt(value)}</b>", ParagraphStyle('val', parent=styles['Normal'],
                    textColor=color_val or colors.black, alignment=2))]

    mode_label = "DETAILED" if is_detail else "SUMMARY"

    # Title
    elements.append(Paragraph(f"Daily Report - {_fmt_date(date)}", title_style))
    elements.append(Paragraph(f"Mode: {mode_label} | FY: {kms_year or 'All'} | Season: {season or 'All'}", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
    elements.append(Spacer(1, 6))

    # ===== PADDY ENTRIES =====
    p = data["paddy_entries"]
    elements.append(Paragraph(f"1. Paddy Entries ({p['count']})", section_style))
    summary_data = [
        ['Total Mill W (QNTL)', 'Total BAG', 'Final W. QNTL (Auto)', 'TP Wt (Q)', 'Bag Deposite', 'Bag Issued'],
        [f"{p.get('total_mill_w', 0)/100:.2f}", str(p['total_bags']), f"{p['total_final_w']/100:.2f}",
         f"{p.get('total_tp_weight', 0):.2f}",
         str(p.get('total_g_deposite', 0)), str(p.get('total_g_issued', 0))]
    ]
    st = RTable(summary_data, colWidths=[100, 80, 100, 80, 80, 80])
    st.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0f2fe')),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(st)
    # Cash/Diesel totals
    cash_diesel_row = [
        ['Total Cash Paid', 'Total Diesel Paid'],
        [f"Rs.{_fmt_amt(p.get('total_cash_paid', 0))}", f"Rs.{_fmt_amt(p.get('total_diesel_paid', 0))}"]
    ]
    cd_t = RTable(cash_diesel_row, colWidths=[250, 250])
    cd_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dcfce7')),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(cd_t)
    if p["details"]:
        col_key = "detail_mode_columns" if is_detail else "summary_mode_columns"
        daily_cols = get_columns("daily_paddy_entries_report", col_key)
        pdf_hdrs = get_pdf_headers(daily_cols)
        if is_detail:
            pdf_widths = [w * mm for w in get_pdf_widths_mm(daily_cols)]
        else:
            pdf_widths = [w * mm for w in get_pdf_widths_mm(daily_cols)]
        pdf_rows = [[str(fmt_val(d.get(c["field"], 0), c["type"])) for c in daily_cols] for d in p["details"]]
        elements.append(make_table(pdf_hdrs, pdf_rows, pdf_widths, font_size=5 if is_detail else 7))
    elements.append(Spacer(1, 4))

    # ===== MILLING =====
    ml = data["milling"]
    if ml["count"]:
        elements.append(Paragraph(f"2. Milling ({ml['count']})", section_style))
        sm = [['Paddy In (Q)', 'Rice Out (Q)', 'FRK Used (Q)'],
              [str(ml['paddy_input_qntl']), str(ml['rice_output_qntl']), str(ml['frk_used_qntl'])]]
        st2 = RTable(sm, colWidths=[170, 170, 170])
        st2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef3c7')),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(st2)
        if is_detail and ml["details"]:
            elements.append(make_table(
                ['Paddy In (Q)', 'Rice Out (Q)', 'Type', 'FRK (Q)', 'CMR Ready (Q)', 'Outturn%'],
                [[str(d.get("paddy_in",0)), str(d.get("rice_out",0)), d.get("type",""),
                  str(d.get("frk",0)), str(d.get("cmr_ready",0)), str(d.get("outturn",0))] for d in ml["details"]],
                [90, 85, 80, 75, 85, 75]
            ))
        elements.append(Spacer(1, 4))

    # ===== PRIVATE TRADING =====
    pp = data["pvt_paddy"]
    rs = data["rice_sales"]
    if pp["count"] or rs["count"]:
        elements.append(Paragraph("3. Private Trading", section_style))
        if pp["count"]:
            elements.append(Paragraph(f"<b>Paddy Purchase ({pp['count']}): {pp['total_qntl']} Qntl | Rs. {_fmt_amt(pp['total_amount'])}</b>",
                ParagraphStyle('sub', parent=styles['Normal'], fontSize=8, spaceAfter=2)))
            if is_detail and pp["details"]:
                elements.append(make_table(
                    ['Party', 'Mandi', 'Truck', 'Qntl', 'Rate/Q', 'Amount', 'Cash', 'Diesel'],
                    [[d.get("party",""), d.get("mandi",""), d.get("truck_no",""), str(d.get("qntl",0)),
                      str(d.get("rate",0)), f"Rs.{_fmt_amt(d.get('amount',0))}",
                      f"Rs.{_fmt_amt(d.get('cash_paid',0))}", f"Rs.{_fmt_amt(d.get('diesel_paid',0))}"] for d in pp["details"]],
                    [80, 65, 65, 50, 55, 75, 55, 55]
                ))
            elif pp["details"]:
                elements.append(make_table(
                    ['Party', 'Mandi', 'Qntl', 'Amount'],
                    [[d["party"], d.get("mandi",""), str(d.get("qntl",0)), f"Rs.{_fmt_amt(d['amount'])}"] for d in pp["details"]],
                    [160, 120, 100, 130]
                ))
        if rs["count"]:
            elements.append(Spacer(1, 4))
            elements.append(Paragraph(f"<b>Rice Sales ({rs['count']}): {rs['total_qntl']} Q | Rs. {_fmt_amt(rs['total_amount'])}</b>",
                ParagraphStyle('sub', parent=styles['Normal'], fontSize=8, spaceAfter=2)))
            if is_detail and rs["details"]:
                elements.append(make_table(
                    ['Party', 'Qntl', 'Type', 'Rate', 'Amount', 'Vehicle'],
                    [[d.get("party",""), str(d.get("qntl",0)), d.get("type",""), str(d.get("rate",0)),
                      f"Rs.{_fmt_amt(d.get('amount',0))}", d.get("vehicle","")] for d in rs["details"]],
                    [110, 65, 80, 65, 90, 80]
                ))
            elif rs["details"]:
                elements.append(make_table(
                    ['Party', 'Qntl', 'Type', 'Amount'],
                    [[d["party"], str(d["qntl"]), d["type"], f"Rs.{_fmt_amt(d['amount'])}"] for d in rs["details"]],
                    [165, 95, 100, 110]
                ))
        elements.append(Spacer(1, 4))

    # ===== CASH FLOW =====
    cf = data["cash_flow"]
    elements.append(Paragraph("4. Cash Flow", section_style))
    cf_sum = [
        ['', 'Jama (Cr)', 'Nikasi (Dr)', 'Net'],
        ['Cash', f"Rs.{_fmt_amt(cf['cash_jama'])}", f"Rs.{_fmt_amt(cf['cash_nikasi'])}", f"Rs.{_fmt_amt(cf['net_cash'])}"],
        ['Bank', f"Rs.{_fmt_amt(cf['bank_jama'])}", f"Rs.{_fmt_amt(cf['bank_nikasi'])}", f"Rs.{_fmt_amt(cf['net_bank'])}"],
    ]
    cft = RTable(cf_sum, colWidths=[80, 130, 130, 130])
    cf_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dcfce7')),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (0, -1), 'FreeSansBold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    cft.setStyle(TableStyle(cf_style))
    elements.append(cft)
    if cf["details"]:
        if is_detail:
            elements.append(make_table(
                ['Description', 'Party', 'Category', 'Type', 'Account', 'Amount'],
                [[d.get("desc",""), d.get("party",""), d.get("category",""), d.get("type","").upper(),
                  d.get("account","").upper(), f"Rs.{_fmt_amt(d.get('amount',0))}"] for d in cf["details"]],
                [150, 85, 70, 55, 55, 85]
            ))
        else:
            elements.append(make_table(
                ['Description', 'Type', 'Account', 'Amount'],
                [[d["desc"], d["type"].upper(), d["account"].upper(), f"Rs.{_fmt_amt(d['amount'])}"] for d in cf["details"]],
                [240, 90, 90, 110]
            ))
    elements.append(Spacer(1, 4))

    # ===== PAYMENTS =====
    pay = data["payments"]
    has_payments = (pay.get('msp_received', 0) or 0) > 0 or (pay.get('pvt_paddy_paid', 0) or 0) > 0 or (pay.get('rice_sale_received', 0) or 0) > 0
    if has_payments:
        elements.append(Paragraph("5. Payments Summary", section_style))
        pay_data = [
            ['MSP Received', 'Pvt Paddy Paid', 'Rice Sale Received'],
            [f"Rs.{_fmt_amt(pay['msp_received'])}", f"Rs.{_fmt_amt(pay['pvt_paddy_paid'])}", f"Rs.{_fmt_amt(pay['rice_sale_received'])}"]
        ]
        pt = RTable(pay_data, colWidths=[170, 170, 170])
        pt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e7ff')),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(pt)
        if is_detail:
            if pay.get("msp_details"):
                elements.append(Paragraph("<b>MSP Payment Details:</b>", ParagraphStyle('sub', parent=styles['Normal'], fontSize=7, spaceBefore=4, spaceAfter=2)))
                elements.append(make_table(['DC No', 'Qntl', 'Rate/Q', 'Amount', 'Mode'],
                    [[d.get("dc_no",""), str(d.get("qntl",0)), str(d.get("rate",0)),
                      f"Rs.{_fmt_amt(d.get('amount',0))}", d.get("mode","")] for d in pay["msp_details"]],
                    [80, 90, 80, 130, 100]))
            if pay.get("pvt_payment_details"):
                elements.append(Paragraph("<b>Private Payment Details:</b>", ParagraphStyle('sub', parent=styles['Normal'], fontSize=7, spaceBefore=4, spaceAfter=2)))
                elements.append(make_table(['Party', 'Type', 'Mode', 'Amount'],
                    [[d.get("party",""), d.get("ref_type",""), d.get("mode",""), f"Rs.{_fmt_amt(d.get('amount',0))}"] for d in pay["pvt_payment_details"]],
                    [170, 120, 100, 110]))
    elements.append(Spacer(1, 4))

    # ===== PUMP ACCOUNT =====
    pa = data.get("pump_account", {})
    if pa.get("details"):
        elements.append(Paragraph("6. Pump Account / Diesel", section_style))
        pa_sum = [
            ['Total Diesel', 'Total Paid', 'Balance'],
            [f"Rs.{_fmt_amt(pa.get('total_diesel', 0))}", f"Rs.{_fmt_amt(pa.get('total_paid', 0))}", f"Rs.{_fmt_amt(pa.get('balance', 0))}"]
        ]
        pat = RTable(pa_sum, colWidths=[170, 170, 170])
        pat.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fff7ed')),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(pat)
        elements.append(make_table(
            ['Pump', 'Type', 'Truck', 'Mandi', 'Description', 'Amount'],
            [[d.get("pump",""), "PAID" if d.get("txn_type") in ("payment","credit") else "DIESEL",
              d.get("truck_no",""), d.get("mandi",""), d.get("desc",""),
              f"Rs.{_fmt_amt(d.get('amount',0))}"] for d in pa["details"]],
            [75, 50, 70, 70, 190, 75]
        ))
        elements.append(Spacer(1, 4))

    # ===== DC DELIVERIES =====
    dc = data["dc_deliveries"]
    if dc["count"]:
        elements.append(Paragraph(f"7. DC Deliveries ({dc['count']}) - {dc['total_qntl']} Q", section_style))
        if is_detail and dc.get("details"):
            elements.append(make_table(
                ['DC No', 'Godown', 'Vehicle', 'Qntl', 'Bags'],
                [[d.get("dc_no",""), d.get("godown",""), d.get("vehicle",""),
                  str(d.get("qntl",0)), str(d.get("bags",0))] for d in dc["details"]],
                [100, 120, 110, 95, 80]
            ))

    # ===== BY-PRODUCTS =====
    bp = data["byproducts"]
    if bp["count"]:
        elements.append(Paragraph(f"8. By-Product Sales ({bp['count']}) - Rs. {_fmt_amt(bp['total_amount'])}", section_style))
        if is_detail and bp.get("details"):
            # Dynamic columns - only show columns with data
            bp_dets = bp["details"]
            has_v = any(d.get("voucher_no") for d in bp_dets)
            has_bill = any(d.get("bill_number") for d in bp_dets)
            has_bdate = any(d.get("billing_date") for d in bp_dets)
            has_rst = any(d.get("rst_no") for d in bp_dets)
            has_veh = any(d.get("vehicle_no") for d in bp_dets)
            has_bf = any(d.get("bill_from") for d in bp_dets)
            has_dest = any(d.get("destination") for d in bp_dets)
            has_bags = any(d.get("bags") for d in bp_dets)

            bp_headers = ['Product']
            bp_widths = [55]
            if has_v: bp_headers.append('Voucher'); bp_widths.append(35)
            if has_bill: bp_headers.append('Bill No'); bp_widths.append(40)
            if has_bdate: bp_headers.append('Bill Dt'); bp_widths.append(38)
            if has_rst: bp_headers.append('RST'); bp_widths.append(30)
            if has_veh: bp_headers.append('Vehicle'); bp_widths.append(45)
            if has_bf: bp_headers.append('Bill From'); bp_widths.append(50)
            bp_headers.append('Party'); bp_widths.append(60)
            if has_dest: bp_headers.append('Destination'); bp_widths.append(48)
            bp_headers.extend(['N/W(Kg)', 'Rate/Q', 'Total'])
            bp_widths.extend([38, 35, 45])
            if has_bags: bp_headers.insert(-3, 'Bags'); bp_widths.insert(-3, 28)

            bp_rows = []
            for d in bp_dets:
                row = [d.get("product", "")[:12]]
                if has_v: row.append(d.get("voucher_no", ""))
                if has_bill: row.append(d.get("bill_number", ""))
                if has_bdate: row.append(d.get("billing_date", "")[-5:] if d.get("billing_date") else "")
                if has_rst: row.append(d.get("rst_no", ""))
                if has_veh: row.append(d.get("vehicle_no", ""))
                if has_bf: row.append((d.get("bill_from", "") or "")[:10])
                row.append((d.get("party_name", "") or "")[:14])
                if has_dest: row.append((d.get("destination", "") or "")[:10])
                if has_bags: row.append(str(d.get("bags", 0) or ""))
                row.extend([str(d.get("net_weight_kg", 0)), str(d.get("rate_per_qtl", 0)), f"Rs.{_fmt_amt(d.get('total', 0))}"])
                bp_rows.append(row)

            # Auto-fit widths
            total_w = sum(bp_widths)
            if total_w > 500:
                scale = 500 / total_w
                bp_widths = [round(w * scale) for w in bp_widths]

            elements.append(make_table(bp_headers, bp_rows, bp_widths))

    # ===== FRK =====
    fk = data["frk"]
    if fk["count"]:
        elements.append(Paragraph(f"9. FRK Purchase ({fk['count']}) - {fk['total_qntl']} Q | Rs. {_fmt_amt(fk['total_amount'])}", section_style))
        if is_detail and fk.get("details"):
            elements.append(make_table(
                ['Party', 'Qntl', 'Rate', 'Amount'],
                [[d.get("party",""), str(d.get("qntl",0)), str(d.get("rate",0)),
                  f"Rs.{_fmt_amt(d.get('amount',0))}"] for d in fk["details"]],
                [170, 100, 100, 130]
            ))

    # ===== MILL PARTS STOCK =====
    mp = data["mill_parts"]
    if mp["in_count"] or mp["used_count"]:
        elements.append(Paragraph(f"10. Mill Parts Stock (In: {mp['in_count']} | Used: {mp['used_count']}) | Purchase: Rs. {_fmt_amt(mp.get('in_amount',0))}", section_style))

    # ===== SALE VOUCHERS =====
    sv = data.get("sale_vouchers", {})
    if sv.get("count", 0):
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"11. Sale Vouchers ({sv['count']}) - Rs. {_fmt_amt(sv['total_amount'])}", section_style))
        if sv.get("details"):
            sv_rows = []
            for d in sv["details"]:
                items_str = ", ".join(f"{it['name']} ({it['qty']}Q @ Rs.{it['rate']})" for it in d.get("items", []))
                sv_rows.append([d.get("voucher_no",""), d.get("party",""), d.get("truck_no",""),
                    items_str[:60], f"Rs.{_fmt_amt(d.get('total',0))}", f"Rs.{_fmt_amt(d.get('balance',0))}"])
            elements.append(make_table(
                ['V.No', 'Party', 'Truck', 'Items', 'Total', 'Balance'],
                sv_rows, [45, 85, 65, 160, 75, 70], font_size=6
            ))

    # ===== PURCHASE VOUCHERS =====
    pvr = data.get("purchase_vouchers", {})
    if pvr.get("count", 0):
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"12. Purchase Vouchers ({pvr['count']}) - Rs. {_fmt_amt(pvr['total_amount'])}", section_style))
        if pvr.get("details"):
            pv_rows = []
            for d in pvr["details"]:
                items_str = ", ".join(f"{it['name']} ({it['qty']}Q @ Rs.{it['rate']})" for it in d.get("items", []))
                pv_rows.append([d.get("voucher_no",""), d.get("party",""), d.get("truck_no",""),
                    items_str[:60], f"Rs.{_fmt_amt(d.get('total',0))}", f"Rs.{_fmt_amt(d.get('balance',0))}"])
            elements.append(make_table(
                ['V.No', 'Party', 'Truck', 'Items', 'Total', 'Balance'],
                pv_rows, [45, 85, 65, 160, 75, 70], font_size=6
            ))
        if mp["in_details"]:
            elements.append(Paragraph("<b>Parts Purchased:</b>", ParagraphStyle('sub', parent=styles['Normal'], fontSize=7, spaceBefore=2, spaceAfter=2)))
            elements.append(make_table(
                ['Part', 'Qty', 'Rate', 'Party', 'Bill No', 'Store Room', 'Amount'],
                [[d.get("part",""), str(d.get("qty",0)), str(d.get("rate",0)),
                  d.get("party",""), d.get("bill_no",""), d.get("store_room",""), f"Rs.{_fmt_amt(d.get('amount',0))}"] for d in mp["in_details"]],
                [90, 50, 55, 85, 70, 70, 75]
            ))
        if mp["used_details"]:
            elements.append(Spacer(1, 3))
            elements.append(Paragraph("<b>Parts Used:</b>", ParagraphStyle('sub', parent=styles['Normal'], fontSize=7, spaceBefore=2, spaceAfter=2)))
            elements.append(make_table(
                ['Part', 'Qty', 'Store Room', 'Remark'],
                [[d.get("part",""), str(d.get("qty",0)), d.get("store_room",""), d.get("remark","")] for d in mp["used_details"]],
                [150, 80, 100, 170]
            ))

    # ===== STAFF ATTENDANCE =====
    sa = data.get("staff_attendance", {})
    if sa.get("total", 0):
        elements.append(Paragraph(f"11. Staff Attendance ({sa['total']})", section_style))
        sa_sum = [
            ['Present', 'Half Day', 'Holiday', 'Absent', 'Not Marked'],
            [str(sa.get('present', 0)), str(sa.get('half_day', 0)), str(sa.get('holiday', 0)), str(sa.get('absent', 0)), str(sa.get('not_marked', 0))]
        ]
        sat = RTable(sa_sum, colWidths=[95, 95, 95, 95, 95])
        sat.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(sat)
        if sa.get("details"):
            status_map = {"present": "P", "absent": "A", "half_day": "H", "holiday": "CH", "not_marked": "-"}
            elements.append(make_table(
                ['Staff Name', 'Status'],
                [[d.get("name",""), status_map.get(d.get("status",""), d.get("status",""))] for d in sa["details"]],
                [300, 150]
            ))

    # ===== HEMALI PAYMENTS =====
    hp = data.get("hemali_payments", {})
    if hp.get("count", 0):
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Hemali Payments ({hp['count']}) - Paid: {hp['paid_count']} | Work: Rs. {_fmt_amt(hp['total_work'])} | Paid: Rs. {_fmt_amt(hp['total_paid'])}", section_style))
        if hp.get("details"):
            elements.append(make_table(
                ['Sardar', 'Items', 'Total', 'Adv Deduct', 'Paid', 'New Adv', 'Status'],
                [[d.get("sardar",""), d.get("items","")[:40], f"Rs.{_fmt_amt(d.get('total',0))}",
                  f"Rs.{_fmt_amt(d.get('advance_deducted',0))}" if d.get('advance_deducted',0) else "-",
                  f"Rs.{_fmt_amt(d.get('amount_paid',0))}", f"Rs.{_fmt_amt(d.get('new_advance',0))}" if d.get('new_advance',0) else "-",
                  d.get("status","").upper()] for d in hp["details"]],
                [70, 130, 65, 65, 65, 60, 45], font_size=6
            ))

    # ===== PADDY CUTTING (CHALNA) =====
    pc = data.get("paddy_cutting", {})
    if pc.get("count", 0):
        elements.append(Spacer(1, 4))
        today_cut = pc.get("total_bags_cut", 0)
        elements.append(Paragraph(f"Paddy Chalna / Cutting - Aaj: {today_cut} Bags", section_style))
        # Summary table
        pc_summary = [
            ['Total Paddy Bags', 'Total Cut (All)', 'Remaining', 'Aaj Cut'],
            [str(pc.get("cum_total_received", 0)), str(pc.get("cum_total_cut", 0)),
             str(pc.get("cum_remaining", 0)), str(today_cut)]
        ]
        pct = RTable(pc_summary, colWidths=[130, 130, 130, 100])
        pct.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef3c7')),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTNAME', (0, 1), (-1, 1), 'FreeSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 7), ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D5DD')),
            ('ALIGN', (0, 1), (-1, 1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(pct)
        if is_detail and pc.get("details"):
            elements.append(make_table(
                ['Bags Cut', 'Remark'],
                [[str(d.get("bags_cut", 0)), d.get("remark", "") or "-"] for d in pc["details"]],
                [100, 400]
            ))

    # ===== CASH TRANSACTIONS =====
    ct = data.get("cash_transactions", {})
    if ct.get("count", 0) > 0:
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Cash Transactions ({ct['count']})", section_style))
        ct_sum = [
            ['Total Jama', 'Total Nikasi', 'Balance'],
            [f"Rs.{_fmt_amt(ct.get('total_jama', 0))}", f"Rs.{_fmt_amt(ct.get('total_nikasi', 0))}",
             f"Rs.{_fmt_amt(ct.get('total_jama', 0) - ct.get('total_nikasi', 0))}"]
        ]
        ctt = RTable(ct_sum, colWidths=[170, 170, 170])
        ctt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef3c7')),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(ctt)
        if ct.get("details"):
            ct_headers = ['Date', 'Party Name', 'Type', 'Amount (Rs.)', 'Description'] if is_detail else ['Date', 'Party Name', 'Type', 'Amount (Rs.)']
            ct_rows = []
            for d in ct["details"]:
                txn_label = "JAMA" if d.get("txn_type") == "jama" else "NIKASI"
                row = [fmt_date(d.get("date", "")[:10]), d.get("party_name", ""), txn_label, f"Rs.{_fmt_amt(d.get('amount', 0))}"]
                if is_detail:
                    row.append(d.get("description", ""))
                ct_rows.append(row)
            ct_widths = [55, 100, 50, 75, 190] if is_detail else [80, 200, 70, 120]
            elements.append(make_table(ct_headers, ct_rows, ct_widths))
        elements.append(Spacer(1, 4))

    # ══ v104.44.19 — P0/P1 NEW SECTIONS (PDF) ══
    # Vehicle Weight
    vw = data.get("vehicle_weight", {})
    if vw.get("sale_count", 0) or vw.get("purchase_count", 0):
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Vehicle Weight (Auto) — Sale: {vw.get('sale_count',0)} trips | Purchase: {vw.get('purchase_count',0)} trips", section_style))
        vw_sum = [
            ['Sale Trips', 'Sale Net(Q)', 'Sale Bhada', 'Sale Bags', 'Purchase Trips', 'Purchase Net(Q)', 'Purchase Bhada', 'Purchase Bags'],
            [str(vw.get('sale_count',0)), str(vw.get('sale_net_qntl',0)), f"Rs.{_fmt_amt(vw.get('sale_bhada_total',0))}", str(vw.get('sale_bags',0)),
             str(vw.get('purchase_count',0)), str(vw.get('purchase_net_qntl',0)), f"Rs.{_fmt_amt(vw.get('purchase_bhada_total',0))}", str(vw.get('purchase_bags',0))]
        ]
        vwt = RTable(vw_sum, colWidths=[60]*8)
        vwt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0f2fe')),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(vwt)
        if vw.get("sale_details"):
            elements.append(Paragraph("<b>Sale / Dispatch:</b>", ParagraphStyle('sub', parent=styles['Normal'], fontSize=7, spaceBefore=4, spaceAfter=2)))
            elements.append(make_table(
                ['RST', 'Vehicle', 'Party', 'Destination', 'Product', 'Bags', 'Bag Type', 'Net Wt', 'Bhada'],
                [[d.get("rst_no","") or "-", d.get("vehicle_no",""), d.get("party",""),
                  d.get("destination","") or "-", d.get("product",""), str(d.get("bags",0)),
                  d.get("bag_type","") or "-", str(d.get("net_wt",0)),
                  f"Rs.{_fmt_amt(d.get('bhada',0))}"] for d in vw["sale_details"]],
                [45, 60, 75, 65, 60, 30, 55, 40, 55], font_size=6
            ))
        if is_detail and vw.get("purchase_details"):
            elements.append(Paragraph("<b>Purchase / Receive:</b>", ParagraphStyle('sub', parent=styles['Normal'], fontSize=7, spaceBefore=4, spaceAfter=2)))
            elements.append(make_table(
                ['RST', 'Vehicle', 'Party', 'Mandi', 'Product', 'Bags', 'Net Wt', 'Bhada'],
                [[d.get("rst_no","") or "-", d.get("vehicle_no",""), d.get("party",""),
                  d.get("mandi","") or "-", d.get("product",""), str(d.get("bags",0)),
                  str(d.get("net_wt",0)), f"Rs.{_fmt_amt(d.get('bhada',0))}"] for d in vw["purchase_details"]],
                [45, 60, 80, 70, 60, 40, 50, 65], font_size=6
            ))

    # Per-Trip Bhada
    ptb = data.get("per_trip_bhada", {})
    if ptb.get("truck_count", 0) > 0:
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Per-Trip Bhada — {ptb['truck_count']} trucks · {ptb['trip_count']} trips", section_style))
        ptb_sum = [
            ['Trucks', 'Trips', 'Bhada Total', 'Paid Today', 'Pending'],
            [str(ptb.get('truck_count',0)), str(ptb.get('trip_count',0)),
             f"Rs.{_fmt_amt(ptb.get('bhada_total',0))}", f"Rs.{_fmt_amt(ptb.get('paid_today',0))}",
             f"Rs.{_fmt_amt(ptb.get('pending_today',0))}"]
        ]
        ptbt = RTable(ptb_sum, colWidths=[100]*5)
        ptbt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fff7ed')),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(ptbt)
        if ptb.get("details"):
            elements.append(make_table(
                ['Vehicle', 'Trips', 'Bhada Total'],
                [[d.get("vehicle_no",""), str(d.get("trips",0)), f"Rs.{_fmt_amt(d.get('bhada',0))}"] for d in ptb["details"]],
                [200, 100, 200]
            ))

    # Party Payments Breakdown — Truck / Agent / LocalParty (from cash_txns)
    for party_key, party_label, party_color in [
        ("truck_payments", "Truck Owner Payments", '#dbeafe'),
        ("agent_payments", "Agent Payments", '#fef3c7'),
        ("local_party_payments", "Local Party Payments", '#ccfbf1'),
    ]:
        ps = data.get(party_key, {})
        if ps.get("count", 0) > 0:
            elements.append(Spacer(1, 4))
            elements.append(Paragraph(f"{party_label} ({ps['count']}) — Jama: Rs.{_fmt_amt(ps.get('jama',0))} | Nikasi: Rs.{_fmt_amt(ps.get('nikasi',0))} | Net: Rs.{_fmt_amt(ps.get('net',0))}", section_style))
            if is_detail and ps.get("details"):
                elements.append(make_table(
                    ['Party', 'Type', 'Account', 'Amount', 'Description'],
                    [[d.get("party","") or "-", d.get("txn_type","").upper(),
                      d.get("account","").upper(), f"Rs.{_fmt_amt(d.get('amount',0))}",
                      d.get("description","") or "-"] for d in ps["details"]],
                    [120, 50, 60, 80, 190], font_size=6
                ))

    # Leased Truck
    lt = data.get("leased_truck", {})
    if lt.get("count", 0):
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Leased Truck Payments ({lt['count']}) — Total Paid: Rs.{_fmt_amt(lt.get('total_paid',0))}", section_style))
        if lt.get("details"):
            elements.append(make_table(
                ['Truck No', 'Owner', 'Payment Type', 'Mode', 'Amount', 'Remark'],
                [[d.get("truck_no","") or "-", d.get("owner","") or "-",
                  d.get("payment_type","") or "-", d.get("mode","") or "-",
                  f"Rs.{_fmt_amt(d.get('amount',0))}", d.get("remark","") or "-"] for d in lt["details"]],
                [80, 90, 80, 55, 70, 125], font_size=6
            ))

    # Oil Premium / Lab Test
    op = data.get("oil_premium", {})
    if op.get("count", 0):
        elements.append(Spacer(1, 4))
        elements.append(Paragraph(f"Lab Test / Oil Premium ({op['count']}) — Positive: {op.get('positive_count',0)} | Negative: {op.get('negative_count',0)} | Net: Rs.{_fmt_amt(op.get('total_premium',0))}", section_style))
        if op.get("details"):
            elements.append(make_table(
                ['V.No', 'RST', 'Party', 'Qty(Q)', 'Sauda Amt', 'Diff %', 'Premium'],
                [[d.get("voucher_no","") or "-", d.get("rst_no","") or "-",
                  d.get("party","") or "-", str(d.get("qty_qntl",0)),
                  f"Rs.{d.get('rate',0)}", f"{d.get('diff_pct',0):+.2f}%",
                  f"Rs.{_fmt_amt(d.get('premium_amount',0))}"] for d in op["details"]],
                [50, 45, 110, 50, 65, 55, 75], font_size=6
            ))

    # Build
    doc.build(elements)
    buf.seek(0)
    fn = f"daily_report_{mode}_{date}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/reports/daily/excel")
async def export_daily_excel(date: str, kms_year: Optional[str] = None, season: Optional[str] = None, mode: str = "normal"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)

    data = await get_daily_report(date, kms_year, season, mode)
    is_detail = mode == "detail"
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daily Report {_fmt_date(date)}"
    mode_label = "DETAILED" if is_detail else "SUMMARY"
    ncols = 8

    title = f"Daily Report / दैनिक रिपोर्ट - {_fmt_date(date)} ({mode_label})"
    subtitle = f"FY: {kms_year or 'All'} | Season: {season or 'All'}"
    style_excel_title(ws, title, ncols, subtitle)
    row = 4

    section_font = Font(bold=True, size=11, color=COLORS['title_text'])
    sub_font = Font(bold=True, size=9, color='475569')
    # Track max width needed per column across the entire sheet
    col_max_widths = {}

    def write_section(title_text):
        nonlocal row
        ws.cell(row=row, column=1, value=title_text).font = section_font
        row += 1

    def write_headers(headers_list):
        nonlocal row
        for i, h in enumerate(headers_list, 1):
            ws.cell(row=row, column=i, value=h)
        style_excel_header_row(ws, row, len(headers_list))
        row += 1

    def write_data_rows(data_rows, header_count):
        nonlocal row
        start = row
        for vals in data_rows:
            for i, v in enumerate(vals, 1):
                ws.cell(row=row, column=i, value=v)
            row += 1
        if data_rows:
            style_excel_data_rows(ws, start, row - 1, header_count)

    def write_sub(text):
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = sub_font
        row += 1

    def write_row(values, bold_row=False):
        nonlocal row
        for i, v in enumerate(values, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.border = BORDER_THIN; c.font = Font(bold=bold_row, size=9)
            c.alignment = Alignment(vertical='center')
        row += 1

    def set_col_widths(width_list):
        """Set explicit column widths for current section."""
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(width_list, 1):
            letter = get_column_letter(i)
            current = col_max_widths.get(letter, 0)
            col_max_widths[letter] = max(current, w)

    # Paddy Entries
    p = data["paddy_entries"]
    write_section(f"1. Paddy Entries ({p['count']})")
    write_sub(f"Total Mill W(Q): {p.get('total_mill_w',0)/100:.2f} | Bags: {p['total_bags']} | Final W(Q): {p['total_final_w']/100:.2f} | TP Wt(Q): {p.get('total_tp_weight',0):.2f}")
    write_sub(f"Bag Dep: {p.get('total_g_deposite',0)} | Bag Issued: {p.get('total_g_issued',0)} | Cash: Rs.{p.get('total_cash_paid',0):,.0f} | Diesel: Rs.{p.get('total_diesel_paid',0):,.0f}")
    if p["details"]:
        col_key = "detail_mode_columns" if is_detail else "summary_mode_columns"
        daily_cols = get_columns("daily_paddy_entries_report", col_key)
        write_headers(get_excel_headers(daily_cols))
        for d in p["details"]:
            write_row([fmt_val(d.get(c["field"], 0), c["type"]) for c in daily_cols])
        set_col_widths([c.get("width_excel", 12) for c in daily_cols])
    row += 1

    # Milling
    ml = data["milling"]
    if ml["count"]:
        write_section(f"2. Milling ({ml['count']})")
        write_sub(f"Paddy In: {ml['paddy_input_qntl']}Q | Rice Out: {ml['rice_output_qntl']}Q | FRK: {ml['frk_used_qntl']}Q")
        if is_detail and ml["details"]:
            write_headers(['Paddy In(Q)', 'Rice Out(Q)', 'Type', 'FRK(Q)', 'CMR Ready(Q)', 'Outturn%'])
            for d in ml["details"]:
                write_row([d.get("paddy_in",0), d.get("rice_out",0), d.get("type",""),
                    d.get("frk",0), d.get("cmr_ready",0), d.get("outturn",0)])
        row += 1

    # Private Trading
    pp = data["pvt_paddy"]
    rs = data["rice_sales"]
    if pp["count"] or rs["count"]:
        write_section("3. Private Trading")
        if pp["count"]:
            write_sub(f"Paddy Purchase ({pp['count']}): {pp['total_qntl']} Qntl | Rs. {pp['total_amount']:,.0f}")
            if pp["details"]:
                if is_detail:
                    write_headers(['Party', 'Mandi', 'Qntl', 'Rate', 'Amount', 'Truck'])
                    for d in pp["details"]:
                        write_row([d.get("party",""), d.get("mandi",""), d.get("qntl",0),
                            d.get("rate",0), d.get("amount",0), d.get("truck_no","")])
                else:
                    write_headers(['Party', 'Qntl', 'Amount'])
                    for d in pp["details"]:
                        write_row([d.get("party",""), d.get("qntl",0), d.get("amount",0)])
        if rs["count"]:
            write_sub(f"Rice Sales ({rs['count']}): {rs['total_qntl']}Q | Rs. {rs['total_amount']:,.0f}")
            if rs["details"]:
                if is_detail:
                    write_headers(['Party', 'Qntl', 'Type', 'Rate', 'Amount', 'Vehicle'])
                    for d in rs["details"]:
                        write_row([d.get("party",""), d.get("qntl",0), d.get("type",""),
                            d.get("rate",0), d.get("amount",0), d.get("vehicle","")])
                else:
                    write_headers(['Party', 'Qntl', 'Type', 'Amount'])
                    for d in rs["details"]:
                        write_row([d["party"], d["qntl"], d["type"], d["amount"]])
        row += 1

    # Cash Flow
    cf = data["cash_flow"]
    write_section("4. Cash Flow")
    write_headers(['', 'Jama (Cr)', 'Nikasi (Dr)', 'Net'])
    write_row(['Cash', cf['cash_jama'], cf['cash_nikasi'], cf['net_cash']])
    write_row(['Bank', cf['bank_jama'], cf['bank_nikasi'], cf['net_bank']])
    if cf["details"]:
        row += 1
        if is_detail:
            write_headers(['Description', 'Party', 'Category', 'Type', 'Account', 'Amount'])
            for d in cf["details"]:
                write_row([d.get("desc",""), d.get("party",""), d.get("category",""),
                    d["type"].upper(), d["account"].upper(), d["amount"]])
            set_col_widths([35, 18, 15, 10, 10, 14])
        else:
            write_headers(['Description', 'Type', 'Account', 'Amount'])
            for d in cf["details"]:
                write_row([d["desc"], d["type"].upper(), d["account"].upper(), d["amount"]])
            set_col_widths([40, 10, 10, 14])
    row += 1

    # Payments
    pay = data["payments"]
    write_section("5. Payments")
    write_headers(['MSP Received', 'Pvt Paddy Paid', 'Rice Sale Received'])
    write_row([pay['msp_received'], pay['pvt_paddy_paid'], pay['rice_sale_received']])
    if is_detail:
        if pay.get("msp_details"):
            write_sub("MSP Details:")
            write_headers(['Agent', 'Mandi', 'Amount'])
            for d in pay["msp_details"]:
                write_row([d.get("agent",""), d.get("mandi",""), d.get("amount",0)])
        if pay.get("pvt_payment_details"):
            write_sub("Private Payment Details:")
            write_headers(['Party', 'Type', 'Mode', 'Amount'])
            for d in pay["pvt_payment_details"]:
                write_row([d.get("party",""), d.get("ref_type",""), d.get("mode",""), d.get("amount",0)])
    row += 1

    # Pump Account
    pa = data.get("pump_account", {})
    if pa.get("details"):
        write_section("6. Pump Account / Diesel")
        write_sub(f"Total Diesel: Rs.{pa.get('total_diesel',0):,.0f} | Paid: Rs.{pa.get('total_paid',0):,.0f} | Balance: Rs.{pa.get('balance',0):,.0f}")
        write_headers(['Pump', 'Type', 'Truck', 'Mandi', 'Description', 'Amount'])
        for d in pa["details"]:
            write_row([d.get("pump",""), "PAID" if d.get("txn_type") in ("payment","credit") else "DIESEL",
                d.get("truck_no",""), d.get("mandi",""), d.get("desc",""), d.get("amount",0)])
        set_col_widths([16, 10, 14, 14, 30, 14])
        row += 1

    # Mill Parts Stock
    mp = data["mill_parts"]
    if mp["in_count"] or mp["used_count"]:
        write_section(f"7. Mill Parts Stock (In: {mp['in_count']} | Used: {mp['used_count']})")
        if mp["in_details"]:
            write_sub(f"Parts Purchased - Total: Rs. {mp.get('in_amount',0):,.0f}")
            write_headers(['Part', 'Qty', 'Rate', 'Party', 'Bill No', 'Store Room', 'Amount'])
            for d in mp["in_details"]:
                write_row([d.get("part",""), d.get("qty",0), d.get("rate",0),
                    d.get("party",""), d.get("bill_no",""), d.get("store_room",""), d.get("amount",0)])
        if mp["used_details"]:
            write_sub("Parts Used:")
            write_headers(['Part', 'Qty', 'Store Room', 'Remark'])
            for d in mp["used_details"]:
                write_row([d.get("part",""), d.get("qty",0), d.get("store_room",""), d.get("remark","")])
    row += 1

    # By-products & FRK
    bp = data["byproducts"]
    fk = data["frk"]
    if bp["count"] or fk["count"]:
        write_section("8. Others")
        if bp["count"]:
            write_sub(f"By-Product Sales ({bp['count']}): Rs. {bp['total_amount']:,.0f}")
            if is_detail and bp.get("details"):
                bp_dets = bp["details"]
                has_v = any(d.get("voucher_no") for d in bp_dets)
                has_bill = any(d.get("bill_number") for d in bp_dets)
                has_bdate = any(d.get("billing_date") for d in bp_dets)
                has_rst = any(d.get("rst_no") for d in bp_dets)
                has_veh = any(d.get("vehicle_no") for d in bp_dets)
                has_bf = any(d.get("bill_from") for d in bp_dets)
                has_dest = any(d.get("destination") for d in bp_dets)
                has_bags = any(d.get("bags") for d in bp_dets)

                hdrs = ['Product']
                if has_v: hdrs.append('Voucher')
                if has_bill: hdrs.append('Bill No')
                if has_bdate: hdrs.append('Bill Date')
                if has_rst: hdrs.append('RST')
                if has_veh: hdrs.append('Vehicle')
                if has_bf: hdrs.append('Bill From')
                hdrs.append('Party')
                if has_dest: hdrs.append('Destination')
                hdrs.append('N/W(Kg)')
                if has_bags: hdrs.append('Bags')
                hdrs.extend(['Rate/Q', 'Amount', 'Total'])
                write_headers(hdrs)
                for d in bp_dets:
                    r = [d.get("product", "")]
                    if has_v: r.append(d.get("voucher_no", ""))
                    if has_bill: r.append(d.get("bill_number", ""))
                    if has_bdate: r.append(d.get("billing_date", ""))
                    if has_rst: r.append(d.get("rst_no", ""))
                    if has_veh: r.append(d.get("vehicle_no", ""))
                    if has_bf: r.append(d.get("bill_from", ""))
                    r.append(d.get("party_name", ""))
                    if has_dest: r.append(d.get("destination", ""))
                    r.append(d.get("net_weight_kg", 0))
                    if has_bags: r.append(d.get("bags", 0))
                    r.extend([d.get("rate_per_qtl", 0), d.get("amount", 0), d.get("total", 0)])
                    write_row(r)
        if fk["count"]:
            write_sub(f"FRK Purchase ({fk['count']}): {fk['total_qntl']}Q | Rs. {fk['total_amount']:,.0f}")
            if is_detail and fk.get("details"):
                write_headers(['Party', 'Qntl', 'Rate', 'Amount'])
                for d in fk["details"]:
                    write_row([d.get("party",""), d.get("qntl",0), d.get("rate",0), d.get("amount",0)])

    # Staff Attendance
    sa = data.get("staff_attendance", {})
    if sa.get("total", 0):
        row += 1
        write_section(f"9. Staff Attendance ({sa['total']})")
        write_headers(['Present', 'Half Day', 'Holiday', 'Absent', 'Not Marked'])
        write_row([sa.get('present', 0), sa.get('half_day', 0), sa.get('holiday', 0), sa.get('absent', 0), sa.get('not_marked', 0)])
        if sa.get("details"):
            row += 1
            status_map = {"present": "P", "absent": "A", "half_day": "H", "holiday": "CH", "not_marked": "-"}
            write_headers(['Staff Name', 'Status'])
            for d in sa["details"]:
                write_row([d.get("name",""), status_map.get(d.get("status",""), d.get("status",""))])

    # Hemali Payments
    hp = data.get("hemali_payments", {})
    if hp.get("count", 0):
        row += 1
        write_section(f"Hemali Payments ({hp['count']})")
        write_sub(f"Paid: {hp['paid_count']} | Unpaid: {hp['unpaid_count']} | Work: Rs.{hp['total_work']:,.0f} | Paid: Rs.{hp['total_paid']:,.0f}")
        if hp.get("details"):
            write_headers(['Sardar', 'Items', 'Total', 'Adv Deducted', 'Paid', 'New Advance', 'Status'])
            for d in hp["details"]:
                write_row([d.get("sardar",""), d.get("items",""), d.get("total",0),
                    d.get("advance_deducted",0), d.get("amount_paid",0), d.get("new_advance",0), d.get("status","").upper()])

    # Paddy Cutting (Chalna)
    pc = data.get("paddy_cutting", {})
    if pc.get("count", 0):
        row += 1
        write_section(f"Paddy Chalna / Cutting - Aaj: {pc.get('total_bags_cut', 0)} Bags")
        write_headers(['Total Paddy Bags', 'Total Cut (All)', 'Remaining', 'Aaj Cut'])
        write_row([pc.get('cum_total_received', 0), pc.get('cum_total_cut', 0), pc.get('cum_remaining', 0), pc.get('total_bags_cut', 0)])
        if is_detail and pc.get("details"):
            write_headers(['Bags Cut', 'Remark'])
            for d in pc["details"]:
                write_row([d.get("bags_cut", 0), d.get("remark", "") or "-"])

    # Cash Transactions
    ct = data.get("cash_transactions", {})
    if ct.get("count", 0) > 0:
        row += 1
        write_section(f"Cash Transactions ({ct['count']})")
        write_sub(f"Jama: Rs.{ct.get('total_jama',0):,.0f} | Nikasi: Rs.{ct.get('total_nikasi',0):,.0f} | Balance: Rs.{(ct.get('total_jama',0) - ct.get('total_nikasi',0)):,.0f}")
        if ct.get("details"):
            if is_detail:
                write_headers(['Date', 'Party Name', 'Type (Jama/Nikasi)', 'Amount (Rs.)', 'Description', 'Payment Mode'])
                set_col_widths([12, 22, 12, 14, 35, 10])
            else:
                write_headers(['Date', 'Party Name', 'Type (Jama/Nikasi)', 'Amount (Rs.)', 'Payment Mode'])
                set_col_widths([12, 22, 12, 14, 10])
            for d in ct["details"]:
                txn_label = "Jama" if d.get("txn_type") == "jama" else "Nikasi"
                if is_detail:
                    write_row([fmt_date(d.get("date","")), d.get("party_name",""), txn_label, round(d.get("amount",0), 2), d.get("description",""), d.get("payment_mode","")])
                else:
                    write_row([fmt_date(d.get("date","")), d.get("party_name",""), txn_label, round(d.get("amount",0), 2), d.get("payment_mode","")])

    # ══ v104.44.19 — P0/P1 NEW SECTIONS (EXCEL) ══
    # Vehicle Weight
    vw = data.get("vehicle_weight", {})
    if vw.get("sale_count", 0) or vw.get("purchase_count", 0):
        row += 1
        write_section(f"Vehicle Weight (Auto) — Sale: {vw.get('sale_count',0)} | Purchase: {vw.get('purchase_count',0)}")
        write_sub(f"Sale: Net {vw.get('sale_net_qntl',0)}Q, Bhada Rs.{vw.get('sale_bhada_total',0):,.0f}, Bags {vw.get('sale_bags',0)} | Purchase: Net {vw.get('purchase_net_qntl',0)}Q, Bhada Rs.{vw.get('purchase_bhada_total',0):,.0f}, Bags {vw.get('purchase_bags',0)}")
        if vw.get("sale_details"):
            write_sub("Sale / Dispatch:")
            write_headers(['RST', 'Vehicle', 'Party', 'Destination', 'Product', 'Bags', 'Bag Type', 'Net Wt', 'Bhada'])
            for d in vw["sale_details"]:
                write_row([d.get("rst_no","") or "-", d.get("vehicle_no",""), d.get("party",""),
                    d.get("destination","") or "-", d.get("product",""), d.get("bags",0),
                    d.get("bag_type","") or "-", d.get("net_wt",0), d.get("bhada",0)])
            set_col_widths([10, 12, 20, 15, 15, 8, 14, 10, 12])
        if is_detail and vw.get("purchase_details"):
            write_sub("Purchase / Receive:")
            write_headers(['RST', 'Vehicle', 'Party', 'Mandi', 'Product', 'Bags', 'Net Wt', 'Bhada'])
            for d in vw["purchase_details"]:
                write_row([d.get("rst_no","") or "-", d.get("vehicle_no",""), d.get("party",""),
                    d.get("mandi","") or "-", d.get("product",""), d.get("bags",0),
                    d.get("net_wt",0), d.get("bhada",0)])

    # Per-Trip Bhada
    ptb = data.get("per_trip_bhada", {})
    if ptb.get("truck_count", 0) > 0:
        row += 1
        write_section(f"Per-Trip Bhada — {ptb['truck_count']} trucks · {ptb['trip_count']} trips")
        write_headers(['Trucks', 'Trips', 'Bhada Total', 'Paid Today', 'Pending'])
        write_row([ptb.get('truck_count',0), ptb.get('trip_count',0),
                   ptb.get('bhada_total',0), ptb.get('paid_today',0), ptb.get('pending_today',0)])
        if ptb.get("details"):
            write_headers(['Vehicle', 'Trips', 'Bhada Total'])
            for d in ptb["details"]:
                write_row([d.get("vehicle_no",""), d.get("trips",0), d.get("bhada",0)])
            set_col_widths([15, 10, 15])

    # Party Payments Breakdown
    for party_key, party_label in [
        ("truck_payments", "Truck Owner Payments"),
        ("agent_payments", "Agent Payments"),
        ("local_party_payments", "Local Party Payments"),
    ]:
        ps = data.get(party_key, {})
        if ps.get("count", 0) > 0:
            row += 1
            write_section(f"{party_label} ({ps['count']})")
            write_sub(f"Jama: Rs.{ps.get('jama',0):,.0f} | Nikasi: Rs.{ps.get('nikasi',0):,.0f} | Net: Rs.{ps.get('net',0):,.0f}")
            if is_detail and ps.get("details"):
                write_headers(['Party', 'Type', 'Account', 'Amount', 'Description'])
                for d in ps["details"]:
                    write_row([d.get("party","") or "-", d.get("txn_type","").upper(),
                        d.get("account","").upper(), round(d.get("amount",0), 2),
                        d.get("description","") or "-"])
                set_col_widths([22, 10, 10, 14, 35])

    # Leased Truck
    lt = data.get("leased_truck", {})
    if lt.get("count", 0):
        row += 1
        write_section(f"Leased Truck Payments ({lt['count']})")
        write_sub(f"Total Paid: Rs.{lt.get('total_paid',0):,.0f}")
        if lt.get("details"):
            write_headers(['Truck No', 'Owner', 'Payment Type', 'Mode', 'Amount', 'Remark'])
            for d in lt["details"]:
                write_row([d.get("truck_no","") or "-", d.get("owner","") or "-",
                    d.get("payment_type","") or "-", d.get("mode","") or "-",
                    round(d.get("amount",0), 2), d.get("remark","") or "-"])
            set_col_widths([12, 18, 14, 10, 14, 25])

    # Oil Premium / Lab Test
    op = data.get("oil_premium", {})
    if op.get("count", 0):
        row += 1
        write_section(f"Lab Test / Oil Premium ({op['count']})")
        write_sub(f"Positive: {op.get('positive_count',0)} | Negative: {op.get('negative_count',0)} | Net: Rs.{op.get('total_premium',0):,.0f}")
        if op.get("details"):
            write_headers(['V.No', 'RST', 'Party', 'Qty(Q)', 'Sauda Amt', 'Diff %', 'Premium', 'Remark'])
            for d in op["details"]:
                write_row([d.get("voucher_no","") or "-", d.get("rst_no","") or "-",
                    d.get("party","") or "-", d.get("qty_qntl",0),
                    d.get("rate",0), f"{d.get('diff_pct',0):+.2f}%",
                    round(d.get("premium_amount",0), 2), d.get("remark","") or "-"])
            set_col_widths([10, 10, 20, 10, 12, 10, 14, 20])

    # Apply collected column widths + smart auto-fit
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        # If an explicit section width was set, use it
        if col_letter in col_max_widths:
            ws.column_dimensions[col_letter].width = col_max_widths[col_letter]
        else:
            # Auto-fit based on content
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    buf = io.BytesIO()
    # 🎯 v104.44.9 — Apply consolidated multi-record polish
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)
    wb.save(buf)
    buf.seek(0)
    fn = f"daily_report_{mode}_{date}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})
