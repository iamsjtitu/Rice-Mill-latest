from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import round_amount
from pydantic import BaseModel, ConfigDict, Field
from utils.date_format import fmt_date
import uuid, io, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

# ============ DC (DELIVERY CHALLAN) MANAGEMENT ============

class DCEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    dc_number: str
    date: str
    quantity_qntl: float = 0
    rice_type: str = "parboiled"  # parboiled / raw
    godown_name: str = ""
    depot_name: str = ""
    depot_code: str = ""
    delivery_to: str = "FCI"  # FCI / RRC
    no_of_lots: str = ""
    deadline: str = ""
    notes: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class DCDelivery(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    dc_id: str
    date: str
    quantity_qntl: float = 0
    vehicle_no: str = ""
    driver_name: str = ""
    slip_no: str = ""
    godown_name: str = ""
    invoice_no: str = ""
    rst_no: str = ""
    bags_used: int = 0  # Govt bags used in this delivery (minus from stock)
    cash_paid: float = 0  # Cash paid to driver → cash book auto entry
    diesel_paid: float = 0  # Diesel paid → truck payment auto entry
    depot_expenses: float = 0  # Depot expenses → cash book auto nikasi
    contract_no: str = ""
    fci_lot_no: str = ""
    eway_bill_no: str = ""
    cgst_amount: float = 0
    sgst_amount: float = 0
    notes: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


@router.post("/dc-entries")
async def add_dc_entry(dc: DCEntry, username: str = ""):
    d = dc.model_dump()
    d['created_by'] = username
    d['quantity_qntl'] = round(d['quantity_qntl'], 2)
    await db.dc_entries.insert_one(d)
    d.pop('_id', None)
    return d


@router.get("/dc-entries")
async def get_dc_entries(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.dc_entries.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(1000)
    # Attach delivery summary to each DC
    for e in entries:
        deliveries = await db.dc_deliveries.find({"dc_id": e["id"]}, {"_id": 0}).to_list(500)
        delivered = round(sum(d.get("quantity_qntl", 0) for d in deliveries), 2)
        e["delivered_qntl"] = delivered
        e["pending_qntl"] = round(e["quantity_qntl"] - delivered, 2)
        e["delivery_count"] = len(deliveries)
        e["status"] = "completed" if delivered >= e["quantity_qntl"] else ("partial" if delivered > 0 else "pending")
        e["deliveries"] = deliveries
    return entries


@router.put("/dc-entries/{dc_id}")
async def update_dc_entry(dc_id: str, dc: DCEntry):
    d = dc.model_dump()
    d['quantity_qntl'] = round(d['quantity_qntl'], 2)
    d.pop('id', None)
    result = await db.dc_entries.update_one({"id": dc_id}, {"$set": d})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="DC not found")
    updated = await db.dc_entries.find_one({"id": dc_id}, {"_id": 0})
    return updated


@router.delete("/dc-entries/{dc_id}")
async def delete_dc_entry(dc_id: str):
    result = await db.dc_entries.delete_one({"id": dc_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="DC not found")
    await db.dc_deliveries.delete_many({"dc_id": dc_id})
    return {"message": "DC and its deliveries deleted", "id": dc_id}


@router.post("/dc-deliveries")
async def add_dc_delivery(delivery: DCDelivery, username: str = ""):
    d = delivery.model_dump()
    d['created_by'] = username
    d['quantity_qntl'] = round(d['quantity_qntl'], 2)
    dc = await db.dc_entries.find_one({"id": d["dc_id"]}, {"_id": 0})
    if not dc:
        raise HTTPException(status_code=404, detail="DC not found")
    await db.dc_deliveries.insert_one(d)
    d.pop('_id', None)

    now_iso = datetime.now(timezone.utc).isoformat()
    base = {"kms_year": d.get("kms_year", ""), "season": d.get("season", ""), "created_by": username, "created_at": now_iso, "updated_at": now_iso}
    dc_num = dc.get("dc_number", "")
    vehicle = d.get("vehicle_no", "")

    # Auto-entry: Cash Paid → Cash Book (Truck payment) + Truck Ledger
    cash_paid = d.get("cash_paid", 0) or 0
    if cash_paid > 0:
        cash_desc = f"DC Delivery Cash - {dc_num} | {vehicle}"
        # Cash Book nikasi
        cash_entry = {
            "id": str(uuid.uuid4()), "date": d["date"], "account": "cash", "txn_type": "nikasi",
            "category": vehicle or f"Truck-{dc_num}", "party_type": "Truck",
            "description": cash_desc,
            "amount": round_amount(cash_paid), "reference": f"delivery:{d['id'][:8]}",
            "bank_name": "", "linked_entry_id": d["id"], **base
        }
        await db.cash_transactions.insert_one(cash_entry)
        cash_entry.pop("_id", None)
        # Truck Ledger nikasi (so it shows in Truck Payments)
        if vehicle:
            ledger_entry = {
                "id": str(uuid.uuid4()), "date": d["date"], "account": "ledger", "txn_type": "nikasi",
                "category": vehicle, "party_type": "Truck",
                "description": cash_desc,
                "amount": round_amount(cash_paid), "reference": f"delivery_tcash:{d['id'][:8]}",
                "bank_name": "", "linked_entry_id": d["id"], **base
            }
            await db.cash_transactions.insert_one(ledger_entry)

    # Auto-entry: Diesel Paid → Truck Ledger + Diesel Account (NOT Cash Book - diesel is not cash)
    diesel_paid = d.get("diesel_paid", 0) or 0
    if diesel_paid > 0:
        diesel_desc = f"DC Delivery Diesel - {dc_num} | {vehicle}"
        # Truck Ledger nikasi (so it shows in Truck Payments)
        if vehicle:
            ledger_diesel = {
                "id": str(uuid.uuid4()), "date": d["date"], "account": "ledger", "txn_type": "nikasi",
                "category": vehicle, "party_type": "Truck",
                "description": diesel_desc,
                "amount": round_amount(diesel_paid), "reference": f"delivery_tdiesel:{d['id'][:8]}",
                "bank_name": "", "linked_entry_id": d["id"], **base
            }
            await db.cash_transactions.insert_one(ledger_diesel)
        # Diesel Account entry
        default_pump = await db.diesel_pumps.find_one({"is_default": True}, {"_id": 0})
        pump_name = default_pump["name"] if default_pump else "Default Pump"
        pump_id = default_pump["id"] if default_pump else "default"
        await db.diesel_accounts.insert_one({
            "id": str(uuid.uuid4()), "date": d["date"],
            "pump_id": pump_id, "pump_name": pump_name,
            "truck_no": vehicle, "agent_name": "",
            "mandi_name": "", "amount": round_amount(diesel_paid), "txn_type": "debit",
            "description": diesel_desc, "linked_entry_id": d["id"],
            **base
        })
        # Pump Ledger Jama (credit to pump - we owe them for diesel)
        pump_jama = {
            "id": str(uuid.uuid4()), "date": d["date"], "account": "ledger", "txn_type": "jama",
            "category": pump_name, "party_type": "Diesel",
            "description": f"Diesel Fill: Truck {vehicle} - {pump_name} - Rs.{diesel_paid}",
            "amount": round_amount(diesel_paid), "reference": f"delivery_dfill:{d['id'][:8]}",
            "bank_name": "", "linked_entry_id": d["id"], **base
        }
        await db.cash_transactions.insert_one(pump_jama)

    # Auto-entry: Depot Expenses → Cash Book Nikasi
    depot_expenses = d.get("depot_expenses", 0) or 0
    if depot_expenses > 0:
        depot_desc = f"DC Delivery Depot Expenses - {dc_num} | {vehicle}" if vehicle else f"DC Delivery Depot Expenses - {dc_num}"
        depot_entry = {
            "id": str(uuid.uuid4()), "date": d["date"], "account": "cash", "txn_type": "nikasi",
            "category": "Depot", "party_type": "Depot",
            "description": depot_desc,
            "amount": round_amount(depot_expenses), "reference": f"delivery_depot:{d['id'][:8]}",
            "bank_name": "", "linked_entry_id": d["id"], **base
        }
        await db.cash_transactions.insert_one(depot_entry)

    # Auto-entry: Bags Used → Govt Bags stock minus
    bags_used = d.get("bags_used", 0) or 0
    if bags_used > 0:
        bag_entry = {
            "id": str(uuid.uuid4()), "date": d["date"], "bag_type": "new", "txn_type": "out",
            "quantity": bags_used, "source": f"DC Delivery - {dc_num}",
            "party_name": "", "rate": 0, "amount": 0, "invoice_no": "", "truck_no": vehicle,
            "rst_no": "", "gst_type": "none", "cgst_percent": 0, "sgst_percent": 0,
            "gst_percent": 0, "gst_amount": 0, "cgst_amount": 0, "sgst_amount": 0,
            "subtotal": 0, "total": 0, "advance": 0, "reference": f"delivery:{d['id'][:8]}",
            "notes": f"Auto: DC delivery bags used", **base
        }
        await db.gunny_bags.insert_one(bag_entry)

    return d


@router.get("/dc-deliveries")
async def get_dc_deliveries(dc_id: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if dc_id: query["dc_id"] = dc_id
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    return await db.dc_deliveries.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(2000)


@router.delete("/dc-deliveries/{delivery_id}")
async def delete_dc_delivery(delivery_id: str):
    # Also clean up auto-created entries (cash, ledger, diesel, gunny)
    ref_prefix = delivery_id[:8]
    await db.cash_transactions.delete_many({"reference": {"$in": [
        f"delivery:{ref_prefix}", f"delivery_diesel:{ref_prefix}",
        f"delivery_tcash:{ref_prefix}", f"delivery_tdiesel:{ref_prefix}",
        f"delivery_dfill:{ref_prefix}", f"delivery_jama:{ref_prefix}",
        f"delivery_depot:{ref_prefix}"
    ]}})
    await db.gunny_bags.delete_many({"reference": f"delivery:{ref_prefix}"})
    await db.diesel_accounts.delete_many({"linked_entry_id": delivery_id})
    result = await db.dc_deliveries.delete_one({"id": delivery_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Delivery not found")
    return {"message": "Delivery deleted", "id": delivery_id}


@router.get("/dc-deliveries/invoice/{delivery_id}")
async def get_delivery_invoice(delivery_id: str):
    delivery = await db.dc_deliveries.find_one({"id": delivery_id}, {"_id": 0})
    if not delivery:
        raise HTTPException(status_code=404, detail="Delivery not found")
    dc = await db.dc_entries.find_one({"id": delivery.get("dc_id", "")}, {"_id": 0})
    settings = await db.settings.find_one({}, {"_id": 0}) or {}
    mill_name = settings.get("mill_name", "NAVKAR AGRO")
    mill_address = settings.get("mill_address", "JOLKO, KESINGA")
    dc_number = dc.get("dc_number", "") if dc else ""
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>Delivery Invoice</title>
    <style>body{{font-family:Arial;margin:20px}}table{{width:100%;border-collapse:collapse}}
    td,th{{border:1px solid #333;padding:6px 10px;text-align:left}}th{{background:#1a365d;color:#fff}}
    .header{{text-align:center;margin-bottom:15px}}.header h1{{margin:0;font-size:22px}}
    .header p{{margin:2px 0;color:#555;font-size:12px}}.info-grid{{display:flex;flex-wrap:wrap;gap:8px;margin:10px 0}}
    .info-item{{flex:1;min-width:140px;background:#f7f7f7;padding:6px 10px;border-radius:4px}}
    .info-item label{{font-size:10px;color:#666;display:block}}.info-item span{{font-size:13px;font-weight:bold}}
    .total-row{{background:#f0f0f0;font-weight:bold}}
    @media print{{body{{margin:0}}button{{display:none}}}}</style></head><body>
    <div class="header"><h1>{mill_name}</h1><p>{mill_address} - Delivery Challan</p></div>
    <div class="info-grid">
      <div class="info-item"><label>DC Number</label><span>{dc_number}</span></div>
      <div class="info-item"><label>Date</label><span>{delivery.get('date','')}</span></div>
      <div class="info-item"><label>Invoice No</label><span>{delivery.get('invoice_no','')}</span></div>
      <div class="info-item"><label>RST No</label><span>{delivery.get('rst_no','')}</span></div>
      <div class="info-item"><label>E-Way Bill</label><span>{delivery.get('eway_bill_no','')}</span></div>
      <div class="info-item"><label>Vehicle No</label><span>{delivery.get('vehicle_no','')}</span></div>
      <div class="info-item"><label>Driver</label><span>{delivery.get('driver_name','')}</span></div>
      <div class="info-item"><label>Slip No</label><span>{delivery.get('slip_no','')}</span></div>
      <div class="info-item"><label>Godown</label><span>{delivery.get('godown_name','')}</span></div>
    </div>
    <table><tr><th>Item</th><th style="text-align:right">Details</th></tr>
      <tr><td>Quantity</td><td style="text-align:right">{delivery.get('quantity_qntl',0)} Quintals</td></tr>
      <tr><td>Bags Used (Govt)</td><td style="text-align:right">{delivery.get('bags_used',0)}</td></tr>
      <tr><td>Cash Paid</td><td style="text-align:right">Rs.{delivery.get('cash_paid',0):,.2f}</td></tr>
      <tr><td>Diesel Paid</td><td style="text-align:right">Rs.{delivery.get('diesel_paid',0):,.2f}</td></tr>
      <tr><td>CGST</td><td style="text-align:right">Rs.{delivery.get('cgst_amount',0):,.2f}</td></tr>
      <tr><td>SGST</td><td style="text-align:right">Rs.{delivery.get('sgst_amount',0):,.2f}</td></tr>
      <tr class="total-row"><td>Total Payment</td><td style="text-align:right">Rs.{(delivery.get('cash_paid',0)+delivery.get('diesel_paid',0)):,.2f}</td></tr>
    </table>
    {f'<p style="margin-top:10px;font-size:12px;color:#555">Notes: {delivery.get("notes","")}</p>' if delivery.get('notes') else ''}
    <div style="margin-top:30px;display:flex;justify-content:space-between"><div style="border-top:1px solid #333;width:150px;text-align:center;padding-top:5px;font-size:11px">Signature</div></div>
    <button onclick="window.print()" style="margin-top:15px;padding:8px 24px;background:#1a365d;color:white;border:none;border-radius:4px;cursor:pointer">Print</button>
    </body></html>"""
    return Response(content=html, media_type="text/html")


@router.get("/dc-summary")
async def get_dc_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    dcs = await db.dc_entries.find(query, {"_id": 0}).to_list(1000)
    total_allotted = round(sum(d.get("quantity_qntl", 0) for d in dcs), 2)
    all_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    total_delivered = round(sum(d.get("quantity_qntl", 0) for d in all_deliveries), 2)
    completed = 0; partial = 0; pending_count = 0
    for dc in dcs:
        deld = sum(d.get("quantity_qntl", 0) for d in all_deliveries if d.get("dc_id") == dc["id"])
        if deld >= dc["quantity_qntl"]: completed += 1
        elif deld > 0: partial += 1
        else: pending_count += 1
    return {
        "total_dc": len(dcs), "total_allotted_qntl": total_allotted,
        "total_delivered_qntl": total_delivered, "total_pending_qntl": round(total_allotted - total_delivered, 2),
        "completed": completed, "partial": partial, "pending": pending_count,
        "total_deliveries": len(all_deliveries)
    }


@router.get("/dc-entries/excel")
async def export_dc_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    dcs = await db.dc_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    all_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    wb = Workbook(); ws = wb.active; ws.title = "DC Register"
    ncols = 9
    title = "DC Register / डीसी रजिस्टर"
    if kms_year: title += f" | FY {kms_year}"
    style_excel_title(ws, title, ncols)
    
    headers = ['DC No', 'Date', 'Rice Type', 'Allotted (Q)', 'Delivered (Q)', 'Pending (Q)', 'Status', 'Deadline', 'Godown']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_excel_header_row(ws, 4, ncols)
    
    data_start = 5; row = data_start
    for dc in dcs:
        deld = round(sum(d.get("quantity_qntl", 0) for d in all_deliveries if d.get("dc_id") == dc["id"]), 2)
        pend = round(dc["quantity_qntl"] - deld, 2)
        status = "Completed" if deld >= dc["quantity_qntl"] else ("Partial" if deld > 0 else "Pending")
        for col, v in enumerate([dc.get("dc_number",""), fmt_date(dc.get("date","")), (dc.get("rice_type","")).capitalize(), dc["quantity_qntl"], deld, pend, status, fmt_date(dc.get("deadline","")), dc.get("godown_name","")], 1):
            ws.cell(row=row, column=col, value=v)
            if col in [4,5,6]: ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        row += 1
    if dcs:
        style_excel_data_rows(ws, data_start, row - 1, ncols, headers)
    
    total_allot = round(sum(d["quantity_qntl"] for d in dcs), 2)
    total_del = round(sum(d.get("quantity_qntl",0) for d in all_deliveries), 2)
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=4, value=total_allot)
    ws.cell(row=row, column=5, value=total_del)
    ws.cell(row=row, column=6, value=round(total_allot-total_del, 2))
    style_excel_total_row(ws, row, ncols)
    
    # Delivery detail sheet
    ws2 = wb.create_sheet("Deliveries")
    dheaders = ['DC No', 'Date', 'Invoice No', 'RST No', 'E-Way Bill', 'Qty (Q)', 'Vehicle', 'Driver', 'Bags', 'Cash Paid', 'Diesel Paid', 'CGST', 'SGST', 'Godown', 'Note']
    for col, h in enumerate(dheaders, 1):
        ws2.cell(row=1, column=col, value=h)
    style_excel_header_row(ws2, 1, len(dheaders))
    dc_map = {d["id"]: d.get("dc_number","") for d in dcs}
    for i, dl in enumerate(sorted(all_deliveries, key=lambda x: x.get("date","")), 2):
        vals = [dc_map.get(dl.get("dc_id",""),""), fmt_date(dl.get("date","")), dl.get("invoice_no",""), dl.get("rst_no",""),
                dl.get("eway_bill_no",""), dl.get("quantity_qntl",0), dl.get("vehicle_no",""), dl.get("driver_name",""),
                dl.get("bags_used",0), dl.get("cash_paid",0), dl.get("diesel_paid",0),
                dl.get("cgst_amount",0), dl.get("sgst_amount",0), dl.get("godown_name",""), dl.get("notes","")]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=i, column=col, value=v)
    if all_deliveries:
        style_excel_data_rows(ws2, 2, len(all_deliveries) + 1, len(dheaders), dheaders)
    
    for letter in 'ABCDEFGHIJKLMNO':
        ws.column_dimensions[letter].width = 16; ws2.column_dimensions[letter].width = 16
    ws.page_setup.orientation = 'landscape'; ws.page_setup.fitToWidth = 1
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dc_register_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/dc-entries/pdf")
async def export_dc_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib import colors
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    dcs = await db.dc_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    all_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    from utils.export_helpers import get_pdf_table_style
    from utils.branding_helper import get_pdf_company_header_from_db
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = get_pdf_styles()
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph("DC Register / डीसी रजिस्टर", styles['Title'])); elements.append(Spacer(1, 12))
    data = [['DC No','Date','Type','Allotted(Q)','Delivered(Q)','Pending(Q)','Status','Deadline','Godown']]
    ta = td = 0
    for dc in dcs:
        deld = round(sum(d.get("quantity_qntl",0) for d in all_deliveries if d.get("dc_id")==dc["id"]),2)
        pend = round(dc["quantity_qntl"]-deld,2); ta += dc["quantity_qntl"]; td += deld
        status = "Done" if deld >= dc["quantity_qntl"] else ("Partial" if deld > 0 else "Pending")
        data.append([dc.get("dc_number",""), fmt_date(dc.get("date","")), (dc.get("rice_type","")).capitalize()[:5], dc["quantity_qntl"], deld, pend, status, fmt_date(dc.get("deadline","")), dc.get("godown_name","")[:12]])
    data.append(['TOTAL','','', round(ta,2), round(td,2), round(ta-td,2), '','',''])
    table = RLTable(data, colWidths=[55,55,40,55,55,50,40,55,60], repeatRows=1)
    style_cmds = get_pdf_table_style(len(data))
    style_cmds.append(('ALIGN',(3,0),(5,-1),'RIGHT'))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=dc_register_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ MSP PAYMENT TRACKING ============

class MSPPayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    dc_id: str = ""  # optional link to DC
    amount: float = 0
    quantity_qntl: float = 0
    rate_per_qntl: float = 0
    payment_mode: str = ""  # NEFT/RTGS/Cheque/Cash
    reference: str = ""  # UTR/Cheque number
    bank_name: str = ""
    notes: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


@router.post("/msp-payments")
async def add_msp_payment(pay: MSPPayment, username: str = ""):
    d = pay.model_dump()
    d['created_by'] = username
    d['amount'] = round(d['amount'], 2)
    d['quantity_qntl'] = round(d['quantity_qntl'], 2)
    d['rate_per_qntl'] = round(d['rate_per_qntl'], 2)
    await db.msp_payments.insert_one(d)
    d.pop('_id', None)
    # Auto-create Cash Book Jama entry (MSP payment received from govt)
    if d['amount'] > 0:
        cb_entry = {
            "id": str(uuid.uuid4()),
            "date": d["date"],
            "account": "bank",
            "txn_type": "jama",
            "category": "MSP Payment",
            "party_type": "MSP",
            "description": f"MSP Payment: {d.get('quantity_qntl', 0)}Q @ Rs.{d.get('rate_per_qntl', 0)}/Q",
            "amount": round_amount(d['amount']),
            "bank_name": d.get("bank_name", ""),
            "reference": f"msp:{d['id'][:8]}",
            "kms_year": d.get("kms_year", ""),
            "season": d.get("season", ""),
            "created_by": username or "system",
            "linked_payment_id": f"msp:{d['id']}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(cb_entry)
        cb_entry.pop("_id", None)
    return d


@router.get("/msp-payments")
async def get_msp_payments(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.msp_payments.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(2000)
    # Attach DC number
    dc_ids = list(set(p.get("dc_id","") for p in payments if p.get("dc_id")))
    dcs = {}
    if dc_ids:
        dc_docs = await db.dc_entries.find({"id": {"$in": dc_ids}}, {"_id": 0, "id": 1, "dc_number": 1}).to_list(500)
        dcs = {d["id"]: d.get("dc_number","") for d in dc_docs}
    for p in payments:
        p["dc_number"] = dcs.get(p.get("dc_id",""), "")
    return payments


@router.delete("/msp-payments/{payment_id}")
async def delete_msp_payment(payment_id: str):
    result = await db.msp_payments.delete_one({"id": payment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    # Delete linked cash book entry
    await db.cash_transactions.delete_many({"linked_payment_id": f"msp:{payment_id}"})
    return {"message": "Payment deleted", "id": payment_id}


@router.get("/msp-payments/summary")
async def get_msp_payment_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.msp_payments.find(query, {"_id": 0}).to_list(5000)
    dcs = await db.dc_entries.find(query, {"_id": 0}).to_list(1000)
    all_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    total_delivered = round(sum(d.get("quantity_qntl",0) for d in all_deliveries), 2)
    total_paid_amount = round(sum(p.get("amount",0) for p in payments), 2)
    total_paid_qty = round(sum(p.get("quantity_qntl",0) for p in payments), 2)
    avg_rate = round(total_paid_amount / total_paid_qty, 2) if total_paid_qty > 0 else 0
    return {
        "total_payments": len(payments), "total_paid_amount": total_paid_amount,
        "total_paid_qty": total_paid_qty, "avg_rate": avg_rate,
        "total_delivered_qntl": total_delivered,
        "pending_payment_qty": round(total_delivered - total_paid_qty, 2),
    }


@router.get("/msp-payments/excel")
async def export_msp_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.msp_payments.find(query, {"_id": 0}).sort("date", 1).to_list(5000)
    dc_ids = list(set(p.get("dc_id","") for p in payments if p.get("dc_id")))
    dcs = {}
    if dc_ids:
        dc_docs = await db.dc_entries.find({"id": {"$in": dc_ids}}, {"_id": 0}).to_list(500)
        dcs = {d["id"]: d.get("dc_number","") for d in dc_docs}
    wb = Workbook(); ws = wb.active; ws.title = "MSP Payments"
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    ncols = 7
    title = "MSP Payment Register / एमएसपी भुगतान"
    style_excel_title(ws, title, ncols)
    
    headers = ['Date','DC No','Qty (Q)','Rate (Rs/Q)','Amount (Rs)','Mode','Bank']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_excel_header_row(ws, 4, ncols)
    
    data_start = 5; row = data_start
    for p in payments:
        for col, v in enumerate([fmt_date(p.get("date","")), dcs.get(p.get("dc_id",""),""), p.get("quantity_qntl",0), p.get("rate_per_qntl",0), p.get("amount",0), p.get("payment_mode",""), p.get("bank_name","")], 1):
            ws.cell(row=row, column=col, value=v)
            if col in [3,4,5]: ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        row += 1
    if payments:
        style_excel_data_rows(ws, data_start, row - 1, ncols, headers)
    
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=3, value=round(sum(p.get("quantity_qntl",0) for p in payments),2))
    ws.cell(row=row, column=5, value=round(sum(p.get("amount",0) for p in payments),2))
    style_excel_total_row(ws, row, ncols)
    for letter in ['A','B','C','D','E','F','G']: ws.column_dimensions[letter].width = 18
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=msp_payments_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/msp-payments/pdf")
async def export_msp_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib import colors
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.msp_payments.find(query, {"_id": 0}).sort("date", 1).to_list(5000)
    dc_ids = list(set(p.get("dc_id","") for p in payments if p.get("dc_id")))
    dcs = {}
    if dc_ids:
        dc_docs = await db.dc_entries.find({"id": {"$in": dc_ids}}, {"_id": 0}).to_list(500)
        dcs = {d["id"]: d.get("dc_number","") for d in dc_docs}
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = get_pdf_styles()
    from utils.export_helpers import get_pdf_table_style
    from utils.branding_helper import get_pdf_company_header_from_db
    
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph("MSP Payment Register / एमएसपी भुगतान", styles['Title'])); elements.append(Spacer(1, 12))
    data = [['Date','DC No','Qty(Q)','Rate(Rs/Q)','Amount(Rs)','Mode','Bank']]
    tq = ta = 0
    for p in payments:
        tq += p.get("quantity_qntl",0); ta += p.get("amount",0)
        data.append([fmt_date(p.get("date","")), dcs.get(p.get("dc_id",""),""), p.get("quantity_qntl",0), p.get("rate_per_qntl",0), p.get("amount",0), p.get("payment_mode",""), p.get("bank_name","")])
    data.append(['TOTAL','',round(tq,2),'',round(ta,2),'',''])
    table = RLTable(data, colWidths=[60,55,45,55,60,40,100], repeatRows=1)
    style_cmds = get_pdf_table_style(len(data))
    style_cmds.append(('ALIGN',(2,0),(4,-1),'RIGHT'))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=msp_payments_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ GUNNY BAG TRACKING ============

class GunnyBagEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    bag_type: str = "new"  # new (govt free) / old (market purchase)
    txn_type: str = "in"   # in / out
    quantity: int = 0
    source: str = ""       # where from / where to (backward compat)
    party_name: str = ""   # explicit party name for purchases
    rate: float = 0        # rate per bag (for old/purchased)
    amount: float = 0
    invoice_no: str = ""
    truck_no: str = ""
    rst_no: str = ""
    gst_type: str = "none"  # none / cgst_sgst / igst
    gst_percent: float = 0
    cgst_percent: float = 0
    sgst_percent: float = 0
    gst_amount: float = 0
    cgst_amount: float = 0
    sgst_amount: float = 0
    subtotal: float = 0
    total: float = 0
    advance: float = 0
    reference: str = ""
    notes: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


async def _create_gunny_accounting_entries(d, doc_id, username):
    """Create accounting entries for gunny bag purchase"""
    party = (d.get('party_name') or d.get('source') or '').strip()
    if not party:
        return
    total = d.get('total', 0) or d.get('amount', 0) or 0
    advance = d.get('advance', 0) or 0
    truck = (d.get('truck_no') or '').strip()
    inv = d.get('invoice_no', '')
    now_iso = datetime.now(timezone.utc).isoformat()
    base = {"kms_year": d.get('kms_year', ''), "season": d.get('season', ''), "created_by": username, "created_at": now_iso, "updated_at": now_iso}
    desc_suffix = f" | Inv:{inv}" if inv else ""

    entries = []

    # 1. Party Ledger JAMA: total purchase amount (we owe the party)
    if total > 0:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "jama",
            "amount": total, "category": party, "party_type": "Gunny Bag",
            "description": f"Gunny Bags x{d.get('quantity', 0)} @ Rs.{d.get('rate', 0)}{desc_suffix}",
            "reference": f"gunny_purchase:{doc_id}", **base
        })

    # 2. Advance paid → NIKASI in party ledger
    if advance > 0:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": advance, "category": party, "party_type": "Gunny Bag",
            "description": f"Advance paid - Gunny Bags{desc_suffix}",
            "reference": f"gunny_advance:{doc_id}", **base
        })

    # 3. Cash NIKASI for advance (cash going out)
    if advance > 0:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "cash", "txn_type": "nikasi",
            "amount": advance, "category": party, "party_type": "Gunny Bag",
            "description": f"Gunny Bags advance - {party}{desc_suffix}",
            "reference": f"gunny_cash:{doc_id}", **base
        })

    for entry in entries:
        await db.cash_transactions.insert_one(entry)

    # Create local_party_accounts entry for gunny bag purchase (we owe = debit)
    if total > 0:
        lp = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "party_name": party, "txn_type": "debit",
            "amount": total, "description": f"Gunny Bags x{d.get('quantity', 0)} @ Rs.{d.get('rate', 0)}{desc_suffix}",
            "source_type": "gunny_bag", "reference": f"gunny_purchase:{doc_id}",
            "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
            "created_by": username, "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.local_party_accounts.insert_one(lp)
    # If advance paid, add payment entry in local_party
    if advance > 0:
        lp_adv = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "party_name": party, "txn_type": "payment",
            "amount": advance, "description": f"Advance paid - Gunny Bags{desc_suffix}",
            "source_type": "gunny_bag_advance", "reference": f"gunny_advance:{doc_id}",
            "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
            "created_by": username, "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.local_party_accounts.insert_one(lp_adv)


@router.post("/gunny-bags")
async def add_gunny_bag_entry(entry: GunnyBagEntry, username: str = ""):
    d = entry.model_dump()
    d['created_by'] = username

    # Calculate amounts
    subtotal = round(d.get('quantity', 0) * d.get('rate', 0), 2)
    d['subtotal'] = subtotal
    d['amount'] = subtotal  # backward compat

    # GST calculation
    gst_amount = 0
    cgst_amount = 0
    sgst_amount = 0
    if d.get('gst_type') == 'cgst_sgst':
        cgst_pct = d.get('cgst_percent', 0) or d.get('gst_percent', 0) or 0
        sgst_pct = d.get('sgst_percent', 0) or d.get('gst_percent', 0) or 0
        cgst_amount = round(subtotal * cgst_pct / 100, 2)
        sgst_amount = round(subtotal * sgst_pct / 100, 2)
        gst_amount = cgst_amount + sgst_amount
        d['cgst_percent'] = cgst_pct
        d['sgst_percent'] = sgst_pct
    elif d.get('gst_type') == 'igst':
        gst_amount = round(subtotal * d.get('gst_percent', 0) / 100, 2)
    d['cgst_amount'] = cgst_amount
    d['sgst_amount'] = sgst_amount
    d['gst_amount'] = round(gst_amount, 2)
    d['total'] = round(subtotal + gst_amount, 2)

    # Use party_name if provided, else fall back to source
    if d.get('party_name') and not d.get('source'):
        d['source'] = d['party_name']
    elif d.get('source') and not d.get('party_name'):
        d['party_name'] = d['source']

    await db.gunny_bags.insert_one(d)
    d.pop('_id', None)

    # Create accounting entries for purchases
    if d.get("txn_type") == "in" and d.get("party_name") and d.get("total", 0) > 0:
        await _create_gunny_accounting_entries(d, d['id'], username)

    return d


@router.get("/gunny-bags")
async def get_gunny_bag_entries(kms_year: Optional[str] = None, season: Optional[str] = None, bag_type: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if bag_type: query["bag_type"] = bag_type
    entries = await db.gunny_bags.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    
    # Get ledger-based paid amounts per party (includes Cash Book manual payments)
    party_names = list(set(e.get("party_name", "") for e in entries if e.get("party_name") and e.get("txn_type") == "in"))
    ledger_paid_map = {}
    if party_names:
        ledger_query = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": party_names}}
        if kms_year: ledger_query["kms_year"] = kms_year
        if season: ledger_query["season"] = season
        ledger_txns = await db.cash_transactions.find(ledger_query, {"_id": 0}).to_list(50000)
        for lt in ledger_txns:
            pn = lt.get("category", "")
            ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lt.get("amount", 0)
    
    # Add ledger_paid and ledger_balance to each purchase entry
    for e in entries:
        if e.get("txn_type") == "in" and e.get("party_name"):
            total_paid = round(ledger_paid_map.get(e["party_name"], 0), 2)
            e["ledger_paid"] = total_paid
            e["ledger_balance"] = round((e.get("total", 0) or e.get("amount", 0)) - total_paid, 2)
    
    return entries


@router.delete("/gunny-bags/{entry_id}")
async def delete_gunny_bag_entry(entry_id: str):
    # Remove linked local party entry + accounting entries
    await db.local_party_accounts.delete_many({"linked_gunny_id": entry_id})
    await db.local_party_accounts.delete_many({"reference": {"$regex": f"gunny_.*:{entry_id}"}})
    await db.cash_transactions.delete_many({"reference": {"$regex": f"gunny_.*:{entry_id}"}})
    await db.cash_transactions.delete_many({"reference": f"lp_gunny:{entry_id[:8]}"})
    result = await db.gunny_bags.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Deleted", "id": entry_id}


@router.put("/gunny-bags/{entry_id}")
async def update_gunny_bag_entry(entry_id: str, entry: GunnyBagEntry, username: str = ""):
    existing = await db.gunny_bags.find_one({"id": entry_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")
    d = entry.model_dump()
    d["id"] = entry_id  # preserve original id
    d["updated_at"] = datetime.now(timezone.utc).isoformat()

    # Calculate amounts
    subtotal = round(d.get("quantity", 0) * d.get("rate", 0), 2)
    d['subtotal'] = subtotal
    d['amount'] = subtotal

    # GST calculation
    gst_amount = 0
    cgst_amount = 0
    sgst_amount = 0
    if d.get('gst_type') == 'cgst_sgst':
        cgst_pct = d.get('cgst_percent', 0) or d.get('gst_percent', 0) or 0
        sgst_pct = d.get('sgst_percent', 0) or d.get('gst_percent', 0) or 0
        cgst_amount = round(subtotal * cgst_pct / 100, 2)
        sgst_amount = round(subtotal * sgst_pct / 100, 2)
        gst_amount = cgst_amount + sgst_amount
        d['cgst_percent'] = cgst_pct
        d['sgst_percent'] = sgst_pct
    elif d.get('gst_type') == 'igst':
        gst_amount = round(subtotal * d.get('gst_percent', 0) / 100, 2)
    d['cgst_amount'] = cgst_amount
    d['sgst_amount'] = sgst_amount
    d['gst_amount'] = round(gst_amount, 2)
    d['total'] = round(subtotal + gst_amount, 2)

    # Sync party_name <-> source
    if d.get('party_name') and not d.get('source'):
        d['source'] = d['party_name']
    elif d.get('source') and not d.get('party_name'):
        d['party_name'] = d['source']

    await db.gunny_bags.update_one({"id": entry_id}, {"$set": d})

    # Delete old accounting entries and recreate
    await db.local_party_accounts.delete_many({"linked_gunny_id": entry_id})
    await db.local_party_accounts.delete_many({"reference": {"$regex": f"gunny_.*:{entry_id}"}})
    await db.cash_transactions.delete_many({"reference": {"$regex": f"gunny_.*:{entry_id}"}})
    await db.cash_transactions.delete_many({"reference": f"lp_gunny:{entry_id[:8]}"})

    if d.get("txn_type") == "in" and d.get("party_name") and d.get("total", 0) > 0:
        await _create_gunny_accounting_entries(d, entry_id, username)

    updated = await db.gunny_bags.find_one({"id": entry_id}, {"_id": 0})
    return updated


@router.get("/gunny-bags/summary")
async def get_gunny_bag_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.gunny_bags.find(query, {"_id": 0}).to_list(10000)

    # Fetch opening stock for gunny bags
    ob_query = {"kms_year": kms_year} if kms_year else {}
    ob_doc = await db.opening_stock.find_one(ob_query, {"_id": 0}) if ob_query else None
    ob_stocks = (ob_doc or {}).get("stocks", {})
    ob_old = float(ob_stocks.get("gunny_old", 0) or 0)
    ob_govt = float(ob_stocks.get("gunny_govt", 0) or 0)
    ob_bran = float(ob_stocks.get("gunny_bran_ppkt", 0) or 0)
    ob_broken = float(ob_stocks.get("gunny_broken_ppkt", 0) or 0)

    # Separate manual vs auto entries
    manual_entries = [e for e in entries if not e.get("linked_entry_id")]
    auto_entries = [e for e in entries if e.get("linked_entry_id")]

    result = {}
    # New (Govt) bags - manual only
    new_items = [e for e in manual_entries if e.get("bag_type") == "new"]
    new_in = sum(e.get("quantity",0) for e in new_items if e.get("txn_type") == "in")
    new_out = sum(e.get("quantity",0) for e in new_items if e.get("txn_type") == "out")
    result["new"] = {"total_in": new_in, "total_out": new_out, "balance": ob_govt + new_in - new_out, "total_cost": 0, "opening": ob_govt}

    # Old (Market) bags - manual only (purchased from market)
    old_items = [e for e in manual_entries if e.get("bag_type") == "old"]
    old_in = sum(e.get("quantity",0) for e in old_items if e.get("txn_type") == "in")
    old_out = sum(e.get("quantity",0) for e in old_items if e.get("txn_type") == "out")
    old_cost = round(sum(e.get("amount",0) for e in old_items if e.get("txn_type") == "in"), 2)
    result["old"] = {"total_in": old_in, "total_out": old_out, "balance": ob_old + old_in - old_out, "total_cost": old_cost, "opening": ob_old}

    # Bran Plastic Pkt
    bran_items = [e for e in entries if e.get("bag_type") == "bran_plastic"]
    bran_in = sum(e.get("quantity",0) for e in bran_items if e.get("txn_type") == "in")
    bran_out = sum(e.get("quantity",0) for e in bran_items if e.get("txn_type") == "out")
    bran_cost = round(sum(e.get("amount",0) for e in bran_items if e.get("txn_type") == "in"), 2)
    result["bran_plastic"] = {"total_in": bran_in, "total_out": bran_out, "balance": ob_bran + bran_in - bran_out, "total_cost": bran_cost, "opening": ob_bran}

    # Broken Plastic Pkt
    broken_items = [e for e in entries if e.get("bag_type") == "broken_plastic"]
    broken_in = sum(e.get("quantity",0) for e in broken_items if e.get("txn_type") == "in")
    broken_out = sum(e.get("quantity",0) for e in broken_items if e.get("txn_type") == "out")
    broken_cost = round(sum(e.get("amount",0) for e in broken_items if e.get("txn_type") == "in"), 2)
    result["broken_plastic"] = {"total_in": broken_in, "total_out": broken_out, "balance": ob_broken + broken_in - broken_out, "total_cost": broken_cost, "opening": ob_broken}

    # Auto entries from mill entries (bag=IN, g_issued=OUT)
    auto_in = sum(e.get("quantity",0) for e in auto_entries if e.get("txn_type") == "in")
    auto_out = sum(e.get("quantity",0) for e in auto_entries if e.get("txn_type") == "out")
    result["auto_mill"] = {"total_in": auto_in, "total_out": auto_out, "balance": auto_in - auto_out}

    # Paddy-received bags from truck entries
    paddy_entries = await db.mill_entries.find(query, {"_id": 0, "bag": 1, "plastic_bag": 1, "g_issued": 1}).to_list(10000)
    paddy_bags = sum(e.get("bag", 0) for e in paddy_entries)
    paddy_ppkt = sum(e.get("plastic_bag", 0) for e in paddy_entries)
    result["paddy_bags"] = {"total": paddy_bags, "label": "Paddy Receive Bags"}
    result["ppkt"] = {"total": paddy_ppkt, "label": "P.Pkt (Plastic Bags)"}

    # Grand total (Excl Govt): ALL old bag entries (manual + auto) IN - OUT + opening
    all_old = [e for e in entries if e.get("bag_type") == "old"]
    all_old_in = sum(e.get("quantity",0) for e in all_old if e.get("txn_type") == "in")
    all_old_out = sum(e.get("quantity",0) for e in all_old if e.get("txn_type") == "out")
    result["grand_total"] = ob_old + all_old_in - all_old_out
    result["g_issued_total"] = sum(e.get("quantity",0) for e in entries if e.get("txn_type") == "out" and e.get("bag_type") == "old")
    return result


@router.get("/gunny-bags/excel")
async def export_gunny_bags_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                                   bag_filter: Optional[str] = None, txn_filter: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.gunny_bags.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    summary = await get_gunny_bag_summary(kms_year=kms_year, season=season)

    # Apply filters
    def apply_filters(items):
        result = items
        if bag_filter == "mill": result = [e for e in result if e.get("linked_entry_id")]
        elif bag_filter == "market": result = [e for e in result if e.get("bag_type") == "old" and not e.get("linked_entry_id")]
        elif bag_filter == "govt": result = [e for e in result if e.get("bag_type") == "new"]
        elif bag_filter == "bran_plastic": result = [e for e in result if e.get("bag_type") == "bran_plastic"]
        elif bag_filter == "broken_plastic": result = [e for e in result if e.get("bag_type") == "broken_plastic"]
        if txn_filter == "in": result = [e for e in result if e.get("txn_type") == "in"]
        elif txn_filter == "out": result = [e for e in result if e.get("txn_type") == "out"]
        return result
    filtered = apply_filters(entries)

    wb = Workbook(); ws = wb.active; ws.title = "Gunny Bags"
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    ncols = 12
    title = "Gunny Bag Register / बोरी रजिस्टर"
    style_excel_title(ws, title, ncols)
    filter_txt = f"FY: {kms_year or 'All'} | Season: {season or 'All'}"
    if bag_filter and bag_filter != 'all': filter_txt += f" | Type: {bag_filter}"
    if txn_filter and txn_filter != 'all': filter_txt += f" | Txn: {txn_filter.upper()}"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    ws.cell(row=3, column=1, value=filter_txt).font = Font(size=9, italic=True)

    # Summary
    ws.cell(row=5, column=1, value="Summary").font = Font(bold=True, size=11, color=COLORS['title_text'])
    sum_headers = ['Category', 'In', 'Out', 'Balance', 'Cost (Rs.)']
    for col, h in enumerate(sum_headers, 1):
        ws.cell(row=6, column=col, value=h)
    style_excel_header_row(ws, 6, 5)
    row = 7; sum_start = row
    am = summary.get("auto_mill", {})
    for label, tin, tout, bal, cost in [
        ("Bag Received (Mill)", am.get("total_in",0), am.get("total_out",0), am.get("balance",0), "-"),
        ("Old Bags (Market)", summary.get("old",{}).get("total_in",0), summary.get("old",{}).get("total_out",0), summary.get("old",{}).get("balance",0), summary.get("old",{}).get("total_cost",0)),
        ("Govt Bags (Free)", summary.get("new",{}).get("total_in",0), summary.get("new",{}).get("total_out",0), summary.get("new",{}).get("balance",0), "-"),
    ]:
        for col, v in enumerate([label, tin, tout, bal, cost], 1):
            ws.cell(row=row, column=col, value=v)
        row += 1
    style_excel_data_rows(ws, sum_start, row - 1, 5, sum_headers)
    ws.cell(row=row, column=1, value="Total G.Issued"); ws.cell(row=row, column=4, value=summary.get("g_issued_total",0))
    style_excel_total_row(ws, row, 5)
    row += 1
    ws.cell(row=row, column=1, value="Total (Excl Govt)"); ws.cell(row=row, column=4, value=summary.get("grand_total",0))
    style_excel_total_row(ws, row, 5)
    row += 1
    ws.cell(row=row, column=1, value="P.Pkt (Plastic)"); ws.cell(row=row, column=4, value=summary.get("ppkt",{}).get("total",0))
    row += 2

    # Transactions
    ws.cell(row=row, column=1, value="Transactions").font = Font(bold=True, size=11, color=COLORS['title_text']); row += 1
    txn_headers = ['Date','Bag Type','In/Out','Qty','Source/To','Rate','Amount (Rs.)','Used For','Damaged','Return','Type','Remark']
    for col, h in enumerate(txn_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_excel_header_row(ws, row, ncols)
    row += 1; txn_start = row
    for e in filtered:
        bt = "New (Govt)" if e.get("bag_type")=="new" else "Bran P.Pkt" if e.get("bag_type")=="bran_plastic" else "Broken P.Pkt" if e.get("bag_type")=="broken_plastic" else "Old (Market)"
        src = (e.get("party_name","") or e.get("source","")) + (" [Auto]" if e.get("linked_entry_id") else "")
        for col, v in enumerate([fmt_date(e.get("date","")), bt, "In" if e.get("txn_type")=="in" else "Out",
            e.get("quantity",0), src, e.get("rate",0), e.get("amount",0),
            e.get("used_for_bp","") or "-", e.get("damaged",0) or "-", e.get("returned",0) or "-",
            bt, e.get("notes","") or "-"], 1):
            ws.cell(row=row, column=col, value=v)
        row += 1
    if filtered:
        style_excel_data_rows(ws, txn_start, row - 1, ncols, txn_headers)
    # Totals
    total_in = sum(e.get("quantity",0) for e in filtered if e.get("txn_type") == "in")
    total_out = sum(e.get("quantity",0) for e in filtered if e.get("txn_type") == "out")
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=3, value=f"In: {total_in} | Out: {total_out}")
    ws.cell(row=row, column=4, value=total_in - total_out)
    style_excel_total_row(ws, row, ncols)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 20
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gunny_bags_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/gunny-bags/pdf")
async def export_gunny_bags_pdf(kms_year: Optional[str] = None, season: Optional[str] = None,
                                 bag_filter: Optional[str] = None, txn_filter: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib import colors
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.gunny_bags.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    summary = await get_gunny_bag_summary(kms_year=kms_year, season=season)

    # Apply filters
    filtered = entries
    if bag_filter == "mill": filtered = [e for e in filtered if e.get("linked_entry_id")]
    elif bag_filter == "market": filtered = [e for e in filtered if e.get("bag_type") == "old" and not e.get("linked_entry_id")]
    elif bag_filter == "govt": filtered = [e for e in filtered if e.get("bag_type") == "new"]
    elif bag_filter == "bran_plastic": filtered = [e for e in filtered if e.get("bag_type") == "bran_plastic"]
    elif bag_filter == "broken_plastic": filtered = [e for e in filtered if e.get("bag_type") == "broken_plastic"]
    if txn_filter == "in": filtered = [e for e in filtered if e.get("txn_type") == "in"]
    elif txn_filter == "out": filtered = [e for e in filtered if e.get("txn_type") == "out"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = get_pdf_styles()
    src_style = ParagraphStyle('src', fontName='FreeSans', fontSize=7, leading=8.5, alignment=TA_LEFT)

    from utils.branding_helper import get_pdf_company_header_from_db
    title = "Gunny Bag Register"
    filter_txt = f"FY: {kms_year or 'All'} | Season: {season or 'All'}"
    if bag_filter and bag_filter != 'all': filter_txt += f" | Type: {bag_filter}"
    if txn_filter and txn_filter != 'all': filter_txt += f" | Txn: {txn_filter.upper()}"
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Paragraph(filter_txt, styles['Normal'])); elements.append(Spacer(1, 8))

    # Summary table
    am = summary.get("auto_mill", {})
    sdata = [['Category','In','Out','Balance','Cost(Rs.)']]
    sdata.append(["Bag Received (Mill)", am.get("total_in",0), am.get("total_out",0), am.get("balance",0), "-"])
    sdata.append(["Old Bags (Market)", summary.get("old",{}).get("total_in",0), summary.get("old",{}).get("total_out",0), summary.get("old",{}).get("balance",0), summary.get("old",{}).get("total_cost",0)])
    sdata.append(["Govt Bags (Free)", summary.get("new",{}).get("total_in",0), summary.get("new",{}).get("total_out",0), summary.get("new",{}).get("balance",0), "-"])
    sdata.append(["Total G.Issued", "", "", summary.get("g_issued_total",0), ""])
    sdata.append(["Total (Excl Govt)", "", "", summary.get("grand_total",0), ""])
    sdata.append(["P.Pkt (Plastic)", "", "", summary.get("ppkt",{}).get("total",0), ""])
    from utils.export_helpers import get_pdf_table_style
    
    st = RLTable(sdata, colWidths=[120,50,50,55,65])
    style_cmds = get_pdf_table_style(len(sdata))
    style_cmds.extend([('ALIGN',(1,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),
        ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)])
    st.setStyle(TableStyle(style_cmds))
    elements.append(st); elements.append(Spacer(1, 12))

    # Transactions
    elements.append(Paragraph("Transactions", styles['Heading2'])); elements.append(Spacer(1, 6))
    data = [['Date','Bag Type','In/Out','Qty','Source/To','Rate','Amt(Rs.)','Used For','Dmg','Ret','Remark']]
    for e in filtered:
        bt = "New(Govt)" if e.get("bag_type")=="new" else "Bran P.Pkt" if e.get("bag_type")=="bran_plastic" else "Broken P.Pkt" if e.get("bag_type")=="broken_plastic" else "Old(Mkt)"
        src = (e.get("party_name","") or e.get("source",""))
        if e.get("linked_entry_id"): src += " [Auto]"
        data.append([fmt_date(e.get("date","")), bt, "In" if e.get("txn_type")=="in" else "Out",
            e.get("quantity",0), Paragraph(src, src_style), e.get("rate",0), e.get("amount",0),
            e.get("used_for_bp","") or "-", e.get("damaged",0) or "-", e.get("returned",0) or "-",
            e.get("notes","") or "-"])
    total_in = sum(e.get("quantity",0) for e in filtered if e.get("txn_type") == "in")
    total_out = sum(e.get("quantity",0) for e in filtered if e.get("txn_type") == "out")
    data.append(['TOTAL', '', f'In:{total_in} Out:{total_out}', total_in - total_out, '', '', '', '', '', '', ''])

    table = RLTable(data, colWidths=[45,48,30,32,120,32,48,55,28,28,55], repeatRows=1)
    txn_style = get_pdf_table_style(len(data))
    txn_style.extend([('ALIGN',(3,0),(6,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),
        ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)])
    table.setStyle(TableStyle(txn_style))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=gunny_bags_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ GUNNY BAG PURCHASE REPORT (PARTY-WISE + GST BREAKUP) ============

@router.get("/gunny-bags/purchase-report")
async def get_gunny_purchase_report(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get party-wise gunny bag purchase summary with GST breakup"""
    query = {"txn_type": "in"}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.gunny_bags.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    # Only include manual purchases (not auto mill entries)
    purchases = [e for e in entries if not e.get("linked_entry_id") and (e.get("total", 0) > 0 or e.get("amount", 0) > 0)]

    # Group by party
    party_map = {}
    for e in purchases:
        party = e.get("party_name") or e.get("source") or "Unknown"
        if party not in party_map:
            party_map[party] = {"entries": [], "total_qty": 0, "subtotal": 0, "cgst": 0, "sgst": 0, "igst": 0, "gst_total": 0, "grand_total": 0, "advance": 0}
        p = party_map[party]
        p["entries"].append(e)
        p["total_qty"] += e.get("quantity", 0)
        p["subtotal"] += e.get("subtotal", 0) or e.get("amount", 0) or 0
        p["cgst"] += e.get("cgst_amount", 0) or 0
        p["sgst"] += e.get("sgst_amount", 0) or 0
        p["igst"] += (e.get("gst_amount", 0) or 0) if e.get("gst_type") == "igst" else 0
        p["gst_total"] += e.get("gst_amount", 0) or 0
        p["grand_total"] += e.get("total", 0) or e.get("amount", 0) or 0
        p["advance"] += e.get("advance", 0) or 0

    # Use LEDGER as source of truth for total_paid (includes advance + Cash Book manual payments + voucher payments)
    all_party_names = list(party_map.keys())
    ledger_query = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": all_party_names}}
    if kms_year: ledger_query["kms_year"] = kms_year
    if season: ledger_query["season"] = season
    ledger_payments = await db.cash_transactions.find(ledger_query, {"_id": 0}).to_list(50000)
    
    ledger_paid_map = {}
    for lp in ledger_payments:
        pn = lp.get("category", "")
        ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lp.get("amount", 0)
    
    for party, data in party_map.items():
        data["total_paid"] = round(ledger_paid_map.get(party, 0), 2)
        data["balance"] = round(data["grand_total"] - data["total_paid"], 2)

    # Round everything
    for data in party_map.values():
        for k in ["subtotal", "cgst", "sgst", "igst", "gst_total", "grand_total", "advance", "total_paid", "balance"]:
            data[k] = round(data[k], 2)

    totals = {
        "total_qty": sum(d["total_qty"] for d in party_map.values()),
        "subtotal": round(sum(d["subtotal"] for d in party_map.values()), 2),
        "cgst": round(sum(d["cgst"] for d in party_map.values()), 2),
        "sgst": round(sum(d["sgst"] for d in party_map.values()), 2),
        "igst": round(sum(d["igst"] for d in party_map.values()), 2),
        "gst_total": round(sum(d["gst_total"] for d in party_map.values()), 2),
        "grand_total": round(sum(d["grand_total"] for d in party_map.values()), 2),
        "advance": round(sum(d["advance"] for d in party_map.values()), 2),
        "total_paid": round(sum(d["total_paid"] for d in party_map.values()), 2),
        "balance": round(sum(d["balance"] for d in party_map.values()), 2),
    }

    # Strip entries from response (keep summary only)
    parties = []
    for party, data in sorted(party_map.items()):
        parties.append({
            "party_name": party, "entry_count": len(data["entries"]),
            "total_qty": data["total_qty"], "subtotal": data["subtotal"],
            "cgst": data["cgst"], "sgst": data["sgst"], "igst": data["igst"],
            "gst_total": data["gst_total"], "grand_total": data["grand_total"],
            "advance": data["advance"], "total_paid": data["total_paid"], "balance": data["balance"],
        })

    return {"parties": parties, "totals": totals}


@router.get("/gunny-bags/purchase-report/excel")
async def export_gunny_purchase_report_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    report = await get_gunny_purchase_report(kms_year=kms_year, season=season)
    parties = report["parties"]
    totals = report["totals"]

    wb = Workbook(); ws = wb.active; ws.title = "Purchase Report"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    bf = Font(bold=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:L1')
    ws['A1'] = "Gunny Bag Purchase Report (Party-wise / GST Breakup)"
    ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:L2')
    ws['A2'] = f"FY: {kms_year or 'All'} | Season: {season or 'All'}"
    ws['A2'].font = Font(size=9, italic=True); ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Party Name', 'Entries', 'Total Qty', 'Subtotal', 'CGST', 'SGST', 'IGST', 'GST Total', 'Grand Total', 'Advance', 'Paid', 'Balance']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb; c.alignment = Alignment(horizontal='right' if col > 1 else 'left')

    row = 5
    for p in parties:
        vals = [p["party_name"], p["entry_count"], p["total_qty"], p["subtotal"], p["cgst"], p["sgst"], p["igst"], p["gst_total"], p["grand_total"], p["advance"], p["total_paid"], p["balance"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v); c.border = tb
            if col > 1: c.alignment = Alignment(horizontal='right')
            if col == 12 and v > 0: c.font = Font(color="CC0000", bold=True)
        row += 1

    # Totals row
    tot_vals = ['TOTAL', len(parties), totals["total_qty"], totals["subtotal"], totals["cgst"], totals["sgst"], totals["igst"], totals["gst_total"], totals["grand_total"], totals["advance"], totals["total_paid"], totals["balance"]]
    for col, v in enumerate(tot_vals, 1):
        c = ws.cell(row=row, column=col, value=v); c.border = tb; c.font = bf
        if col > 1: c.alignment = Alignment(horizontal='right')

    for col, w in enumerate([30, 8, 10, 12, 12, 12, 12, 12, 14, 12, 12, 14], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gunny_purchase_report_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/gunny-bags/purchase-report/pdf")
async def export_gunny_purchase_report_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib import colors
    from io import BytesIO

    report = await get_gunny_purchase_report(kms_year=kms_year, season=season)
    parties = report["parties"]
    totals = report["totals"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=25, bottomMargin=25)
    elements = []; styles = get_pdf_styles()

    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph("Gunny Bag Purchase Report (Party-wise / GST Breakup)", styles['Title']))
    elements.append(Paragraph(f"FY: {kms_year or 'All'} | Season: {season or 'All'}", styles['Normal']))
    elements.append(Spacer(1, 10))

    data = [['Party', '#', 'Qty', 'Subtotal', 'CGST', 'SGST', 'IGST', 'GST', 'Total', 'Advance', 'Paid', 'Balance']]
    for p in parties:
        data.append([p["party_name"], p["entry_count"], p["total_qty"],
            f"Rs.{p['subtotal']:,.0f}", f"Rs.{p['cgst']:,.0f}", f"Rs.{p['sgst']:,.0f}", f"Rs.{p['igst']:,.0f}",
            f"Rs.{p['gst_total']:,.0f}", f"Rs.{p['grand_total']:,.0f}", f"Rs.{p['advance']:,.0f}",
            f"Rs.{p['total_paid']:,.0f}", f"Rs.{p['balance']:,.0f}"])
    data.append(['TOTAL', len(parties), totals["total_qty"],
        f"Rs.{totals['subtotal']:,.0f}", f"Rs.{totals['cgst']:,.0f}", f"Rs.{totals['sgst']:,.0f}", f"Rs.{totals['igst']:,.0f}",
        f"Rs.{totals['gst_total']:,.0f}", f"Rs.{totals['grand_total']:,.0f}", f"Rs.{totals['advance']:,.0f}",
        f"Rs.{totals['total_paid']:,.0f}", f"Rs.{totals['balance']:,.0f}"])

    cw = [110, 25, 35, 60, 50, 50, 50, 50, 65, 55, 55, 65]
    table = RLTable(data, colWidths=cw, repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5), ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
        ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'), ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'), ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    # Highlight balance > 0 in red
    for i, p in enumerate(parties, 1):
        if p["balance"] > 0:
            style.append(('TEXTCOLOR', (11, i), (11, i), colors.red))
    table.setStyle(TableStyle(style))
    elements.append(table)
    doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=gunny_purchase_report_{datetime.now().strftime('%Y%m%d')}.pdf"})



# ===== DC STACKS =====
@router.get("/dc-stacks")
async def get_dc_stacks(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    stacks = await db.dc_stacks.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for s in stacks:
        lots = await db.dc_stack_lots.find({"stack_id": s["id"]}, {"_id": 0}).sort("lot_number", 1).to_list(100)
        s["lots"] = lots
        s["lots_total"] = len(lots)
        delivered = [l for l in lots if l.get("status") == "delivered"]
        s["lots_delivered"] = len(delivered)
        s["last_delivered_date"] = max((l.get("date", "") for l in delivered), default=None) if delivered else None
    return stacks

@router.post("/dc-stacks")
async def create_dc_stack(data: dict):
    data["id"] = str(uuid.uuid4())
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.dc_stacks.insert_one(data)
    data.pop("_id", None)
    return data

@router.put("/dc-stacks/{stack_id}")
async def update_dc_stack(stack_id: str, data: dict):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.dc_stacks.update_one({"id": stack_id}, {"$set": data})
    updated = await db.dc_stacks.find_one({"id": stack_id}, {"_id": 0})
    return updated

@router.delete("/dc-stacks/{stack_id}")
async def delete_dc_stack(stack_id: str):
    await db.dc_stacks.delete_one({"id": stack_id})
    await db.dc_stack_lots.delete_many({"stack_id": stack_id})
    return {"success": True}

# ===== DC STACK LOTS =====
@router.get("/dc-stacks/{stack_id}/lots")
async def get_stack_lots(stack_id: str):
    lots = await db.dc_stack_lots.find({"stack_id": stack_id}, {"_id": 0}).sort("lot_number", 1).to_list(100)
    return lots

@router.post("/dc-stacks/{stack_id}/lots")
async def create_stack_lot(stack_id: str, data: dict):
    existing = await db.dc_stack_lots.count_documents({"stack_id": stack_id})
    data["id"] = str(uuid.uuid4())
    data["stack_id"] = stack_id
    data["lot_number"] = existing + 1
    data["status"] = data.get("status", "pending")
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.dc_stack_lots.insert_one(data)
    data.pop("_id", None)
    return data

@router.put("/dc-stacks/{stack_id}/lots/{lot_id}")
async def update_stack_lot(stack_id: str, lot_id: str, data: dict):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.dc_stack_lots.update_one({"id": lot_id}, {"$set": data})
    updated = await db.dc_stack_lots.find_one({"id": lot_id}, {"_id": 0})
    return updated

@router.delete("/dc-stacks/{stack_id}/lots/{lot_id}")
async def delete_stack_lot(stack_id: str, lot_id: str):
    await db.dc_stack_lots.delete_one({"id": lot_id})
    return {"success": True}

