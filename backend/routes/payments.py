from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import *
import uuid, io, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

async def get_company_name():
    settings = await db.settings.find_one({"type": "company"}, {"_id": 0})
    if settings:
        return settings.get("company_name", "Mill Entry System"), settings.get("tagline", "")
    return "Mill Entry System", ""

async def _find_truck_entry(entry_id):
    """Find entry from mill_entries, private_paddy, rice_sales, sale_vouchers, or dc_deliveries - return (entry, source, final_qntl)"""
    entry = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if entry:
        fq = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
        return entry, "cmr", fq
    entry = await db.private_paddy.find_one({"id": entry_id}, {"_id": 0})
    if entry:
        return entry, "pvt", round(entry.get("final_qntl", 0), 2)
    entry = await db.rice_sales.find_one({"id": entry_id}, {"_id": 0})
    if entry:
        return entry, "rice_sale", round(entry.get("quantity_qntl", 0), 2)
    entry = await db.sale_vouchers.find_one({"id": entry_id}, {"_id": 0})
    if entry:
        total_qty = sum(i.get("quantity", 0) for i in entry.get("items", []))
        return entry, "sale_book", round(total_qty, 2)
    entry = await db.purchase_vouchers.find_one({"id": entry_id}, {"_id": 0})
    if entry:
        total_qty = sum(i.get("quantity", 0) for i in entry.get("items", []))
        return entry, "purchase_voucher", round(total_qty, 2)
    entry = await db.dc_deliveries.find_one({"id": entry_id}, {"_id": 0})
    if entry:
        return entry, "dc_delivery", round(entry.get("quantity_qntl", 0), 2)
    return None, None, 0

# ============ TRUCK PAYMENT ENDPOINTS ============

@router.get("/truck-payments", response_model=List[TruckPaymentStatus])
async def get_truck_payments(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get all truck payments with their status - uses ledger as source of truth for paid amounts"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(1000)
    
    # Bulk fetch all ledger nikasi entries for all trucks (for paid calculation)
    all_truck_nos = list(set(e.get("truck_no", "") for e in entries if e.get("truck_no")))
    ledger_nikasi = []
    if all_truck_nos:
        ledger_nikasi = await db.cash_transactions.find({
            "account": "ledger", "txn_type": "nikasi",
            "category": {"$in": all_truck_nos}
        }, {"_id": 0}).to_list(50000)
    
    # Group ledger nikasi by truck_no
    truck_ledger_map = {}
    for txn in ledger_nikasi:
        cat = txn.get("category", "")
        truck_ledger_map.setdefault(cat, []).append(txn)
    
    # Deduction reference prefixes (already counted in deductions field)
    DEDUCTION_PREFIXES = ("truck_cash_ded:", "truck_diesel_ded:", "entry_cash:")
    
    payments = []
    # Group entries by truck to handle FIFO distribution of manual payments
    truck_entries_map = {}
    for entry in entries:
        truck_no = entry.get("truck_no", "")
        truck_entries_map.setdefault(truck_no, []).append(entry)
    
    for truck_no, truck_entries in truck_entries_map.items():
        # Sort entries by date ascending (oldest first) for FIFO payment distribution
        truck_entries_sorted = sorted(truck_entries, key=lambda e: (e.get("date", ""), e.get("created_at", "")))
        
        truck_ledger = truck_ledger_map.get(truck_no, [])
        
        # Separate entry-specific payments from manual (unattributed) payments
        entry_specific_paid = {}  # entry_id -> amount
        manual_payments_total = 0
        
        for txn in truck_ledger:
            ref = txn.get("reference", "")
            amount = txn.get("amount", 0)
            
            # Skip deduction entries (already counted as deductions)
            if any(ref.startswith(p) for p in DEDUCTION_PREFIXES):
                continue
            
            # Check if this payment is linked to a specific entry
            attributed = False
            for entry in truck_entries_sorted:
                eid = entry.get("id", "")
                eid_short = eid[:8]
                if (ref.startswith(f"truck_pay_ledger:{eid_short}") or
                    ref.startswith(f"truck_markpaid_ledger:{eid}") or
                    ref == f"truck_pay_ledger:{eid_short}" or
                    ref == f"truck_markpaid_ledger:{eid}"
                ):
                    entry_specific_paid[eid] = entry_specific_paid.get(eid, 0) + amount
                    attributed = True
                    break
            
            if not attributed:
                # Manual payment (auto_ledger or other) - will be distributed FIFO
                manual_payments_total += amount
        
        # Distribute manual payments FIFO (oldest entry first)
        remaining_manual = manual_payments_total
        entry_manual_paid = {}
        for entry in truck_entries_sorted:
            if remaining_manual <= 0:
                break
            eid = entry.get("id", "")
            payment_doc = await db.truck_payments.find_one({"entry_id": eid}, {"_id": 0})
            rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
            final_qntl = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
            cash_taken = float(entry.get("cash_paid", 0) or 0)
            diesel_taken = float(entry.get("diesel_paid", 0) or 0)
            net = round((final_qntl * rate) - cash_taken - diesel_taken, 2)
            already_paid = entry_specific_paid.get(eid, 0)
            remaining_for_entry = max(0, net - already_paid)
            manual_alloc = min(remaining_manual, remaining_for_entry)
            if manual_alloc > 0:
                entry_manual_paid[eid] = manual_alloc
                remaining_manual -= manual_alloc
        
        # Build payment status for each entry
        for entry in truck_entries_sorted:
            entry_id = entry.get("id")
            payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
            rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
            
            final_qntl = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
            cash_taken = float(entry.get("cash_paid", 0) or 0)
            diesel_taken = float(entry.get("diesel_paid", 0) or 0)
            
            gross_amount = round(final_qntl * rate, 2)
            deductions = round(cash_taken + diesel_taken, 2)
            net_amount = round(gross_amount - deductions, 2)
            
            # paid_amount from ledger: entry-specific + manual FIFO allocation
            paid_amount = round(
                entry_specific_paid.get(entry_id, 0) + entry_manual_paid.get(entry_id, 0), 2
            )
            balance = round(net_amount - paid_amount, 2)
            
            if paid_amount <= 0:
                status = "pending"
            elif balance < 0.10:
                status = "paid"
            else:
                status = "partial"
            
            payments.append(TruckPaymentStatus(
                entry_id=entry_id,
                truck_no=truck_no,
                date=entry.get("date", ""),
                total_qntl=round(entry.get("qntl", 0), 2),
                total_bag=entry.get("bag", 0),
                final_qntl=final_qntl,
                cash_taken=cash_taken,
                diesel_taken=diesel_taken,
                rate_per_qntl=rate,
                gross_amount=gross_amount,
                deductions=deductions,
                net_amount=net_amount,
                paid_amount=paid_amount,
                balance_amount=max(0, balance),
                status=status,
                kms_year=entry.get("kms_year", ""),
                season=entry.get("season", ""),
                agent_name=entry.get("agent_name", ""),
                mandi_name=entry.get("mandi_name", "")
            ))
    
    # Also include Pvt Paddy entries with truck_no (cash + diesel go to truck)
    # Exclude "agent_extra" entries (moved to paddy purchase from reports) - they are not truck transport entries
    pvt_query = dict(query)
    pvt_query["source"] = {"$ne": "agent_extra"}
    pvt_paddy = await db.private_paddy.find(pvt_query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(1000)
    for p in pvt_paddy:
        truck_no = p.get("truck_no", "")
        if not truck_no:
            continue
        cash_paid = float(p.get("cash_paid", 0) or 0)
        diesel_paid = float(p.get("diesel_paid", 0) or 0)
        deductions = round(cash_paid + diesel_paid, 2)
        party = p.get("party_name", "")
        mandi = p.get("mandi_name", "")
        party_label = f"{party} - {mandi}" if party and mandi else party
        final_qntl = round(p.get("final_qntl", 0), 2)
        # Check if rate/payment exists in truck_payments
        tp = await db.truck_payments.find_one({"entry_id": p["id"]}, {"_id": 0})
        rate = tp.get("rate_per_qntl", 0) if tp else 0
        extra_paid = tp.get("paid_amount", 0) if tp else 0
        tp_status = tp.get("status", "") if tp else ""
        gross = round(final_qntl * rate, 2) if rate > 0 else 0
        net = round(gross - deductions, 2) if gross > 0 else 0
        total_paid = round(deductions + extra_paid, 2)
        balance = round(net - extra_paid, 2) if gross > 0 else 0
        status = tp_status if tp_status else ("paid" if (gross > 0 and balance <= 0) else ("partial" if extra_paid > 0 else "pending"))
        payments.append(TruckPaymentStatus(
            entry_id=p.get("id", ""),
            truck_no=truck_no,
            date=p.get("date", ""),
            total_qntl=final_qntl,
            total_bag=int(p.get("bag", 0)),
            final_qntl=final_qntl,
            cash_taken=cash_paid,
            diesel_taken=diesel_paid,
            rate_per_qntl=rate,
            gross_amount=gross,
            deductions=deductions,
            net_amount=net,
            paid_amount=total_paid,
            balance_amount=balance,
            status=status,
            kms_year=p.get("kms_year", ""),
            season=p.get("season", ""),
            agent_name=p.get("agent_name", ""),
            mandi_name=party_label,
            source="Pvt Paddy"
        ))

    # Also include Rice Sale entries with truck_no (cash + diesel go to truck)
    rice_query = dict(query)
    rice_sales = await db.rice_sales.find(rice_query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(1000)
    for r in rice_sales:
        truck_no = r.get("truck_no", "")
        if not truck_no:
            continue
        cash_paid = float(r.get("cash_paid", 0) or 0)
        diesel_paid = float(r.get("diesel_paid", 0) or 0)
        if cash_paid == 0 and diesel_paid == 0:
            continue
        deductions = round(cash_paid + diesel_paid, 2)
        party = r.get("party_name", "")
        qty = r.get("quantity_qntl", 0) or 0
        tp = await db.truck_payments.find_one({"entry_id": r["id"]}, {"_id": 0})
        rate = tp.get("rate_per_qntl", 0) if tp else 0
        extra_paid = tp.get("paid_amount", 0) if tp else 0
        tp_status = tp.get("status", "") if tp else ""
        gross = round(qty * rate, 2) if rate > 0 else 0
        net = round(gross - deductions, 2) if gross > 0 else 0
        balance = round(net - extra_paid, 2) if gross > 0 else 0
        status = tp_status if tp_status else ("paid" if (gross > 0 and balance <= 0) else ("partial" if extra_paid > 0 else "pending"))
        payments.append(TruckPaymentStatus(
            entry_id=r.get("id", ""),
            truck_no=truck_no,
            date=r.get("date", ""),
            total_qntl=qty,
            total_bag=int(r.get("bags", 0)),
            final_qntl=qty,
            cash_taken=cash_paid,
            diesel_taken=diesel_paid,
            rate_per_qntl=rate,
            gross_amount=gross,
            deductions=deductions,
            net_amount=net,
            paid_amount=round(deductions + extra_paid, 2),
            balance_amount=max(0, balance),
            status=status,
            kms_year=r.get("kms_year", ""),
            season=r.get("season", ""),
            agent_name="",
            mandi_name=party,
            source="Rice Sale"
        ))

    # Also include Sale Book vouchers with truck_no
    sb_query = dict(query)
    sale_vouchers = await db.sale_vouchers.find(sb_query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(1000)
    for sv in sale_vouchers:
        truck_no = sv.get("truck_no", "")
        if not truck_no:
            continue
        cash_paid = float(sv.get("cash_paid", 0) or 0)
        diesel_paid = float(sv.get("diesel_paid", 0) or 0)
        if cash_paid == 0 and diesel_paid == 0:
            continue
        deductions = round(cash_paid + diesel_paid, 2)
        party = sv.get("party_name", "")
        inv = sv.get("invoice_no", "")
        label = f"Sale #{sv.get('voucher_no', '')} - {party}"
        if inv: label += f" ({inv})"
        # Read actual rate, paid from truck_payments
        tp = await db.truck_payments.find_one({"entry_id": sv["id"]}, {"_id": 0})
        rate = tp.get("rate_per_qntl", 0) if tp else 0
        extra_paid = tp.get("paid_amount", 0) if tp else 0
        tp_status = tp.get("status", "") if tp else ""
        qty = round(sum(i.get("quantity", 0) for i in sv.get("items", [])), 2)
        gross = round(qty * rate, 2) if rate > 0 else 0
        net = round(gross - deductions, 2) if gross > 0 else 0
        balance = round(net - extra_paid, 2) if gross > 0 else 0
        status = tp_status if tp_status else ("paid" if (gross > 0 and balance <= 0) else ("partial" if extra_paid > 0 else "pending"))
        payments.append(TruckPaymentStatus(
            entry_id=sv.get("id", ""),
            truck_no=truck_no,
            date=sv.get("date", ""),
            total_qntl=qty,
            total_bag=0,
            final_qntl=qty,
            cash_taken=cash_paid,
            diesel_taken=diesel_paid,
            rate_per_qntl=rate,
            gross_amount=gross,
            deductions=deductions,
            net_amount=net,
            paid_amount=round(deductions + extra_paid, 2),
            balance_amount=max(0, balance),
            status=status,
            kms_year=sv.get("kms_year", ""),
            season=sv.get("season", ""),
            agent_name="",
            mandi_name=label,
            source="Sale Book"
        ))

    # Also include Purchase Vouchers with truck_no
    pv_query = dict(query)
    purchase_vouchers = await db.purchase_vouchers.find(pv_query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(1000)
    for pv in purchase_vouchers:
        truck_no = pv.get("truck_no", "")
        if not truck_no:
            continue
        cash_paid = float(pv.get("cash_paid", 0) or 0)
        diesel_paid = float(pv.get("diesel_paid", 0) or 0)
        if cash_paid == 0 and diesel_paid == 0:
            continue
        deductions = round(cash_paid + diesel_paid, 2)
        party = pv.get("party_name", "")
        inv = pv.get("invoice_no", "")
        label = f"Purchase #{pv.get('voucher_no', '')} - {party}"
        if inv: label += f" ({inv})"
        tp = await db.truck_payments.find_one({"entry_id": pv["id"]}, {"_id": 0})
        rate = tp.get("rate_per_qntl", 0) if tp else 0
        extra_paid = tp.get("paid_amount", 0) if tp else 0
        tp_status = tp.get("status", "") if tp else ""
        qty = round(sum(i.get("quantity", 0) for i in pv.get("items", [])), 2)
        gross = round(qty * rate, 2) if rate > 0 else 0
        net = round(gross - deductions, 2) if gross > 0 else 0
        balance = round(net - extra_paid, 2) if gross > 0 else 0
        status = tp_status if tp_status else ("paid" if (gross > 0 and balance <= 0) else ("partial" if extra_paid > 0 else "pending"))
        payments.append(TruckPaymentStatus(
            entry_id=pv.get("id", ""),
            truck_no=truck_no,
            date=pv.get("date", ""),
            total_qntl=qty,
            total_bag=0,
            final_qntl=qty,
            cash_taken=cash_paid,
            diesel_taken=diesel_paid,
            rate_per_qntl=rate,
            gross_amount=gross,
            deductions=deductions,
            net_amount=net,
            paid_amount=round(deductions + extra_paid, 2),
            balance_amount=max(0, balance),
            status=status,
            kms_year=pv.get("kms_year", ""),
            season=pv.get("season", ""),
            agent_name="",
            mandi_name=label,
            source="Purchase Voucher"
        ))

    # Also include DC Deliveries with vehicle_no (cash + diesel go to truck)
    dc_del_query = {}
    if kms_year: dc_del_query["kms_year"] = kms_year
    if season: dc_del_query["season"] = season
    dc_deliveries = await db.dc_deliveries.find(dc_del_query, {"_id": 0}).sort([("date", -1)]).to_list(1000)
    for dd in dc_deliveries:
        truck_no = dd.get("vehicle_no", "")
        if not truck_no:
            continue
        cash_paid = float(dd.get("cash_paid", 0) or 0)
        diesel_paid = float(dd.get("diesel_paid", 0) or 0)
        if cash_paid == 0 and diesel_paid == 0:
            continue
        deductions = round(cash_paid + diesel_paid, 2)
        dc_id = dd.get("dc_id", "")
        dc_entry = await db.dc_entries.find_one({"id": dc_id}, {"_id": 0})
        dc_num = dc_entry.get("dc_number", "") if dc_entry else ""
        label = f"DC Delivery - {dc_num}" if dc_num else "DC Delivery"
        inv = dd.get("invoice_no", "")
        if inv: label += f" ({inv})"
        qty = round(dd.get("quantity_qntl", 0), 2)
        tp = await db.truck_payments.find_one({"entry_id": dd["id"]}, {"_id": 0})
        rate = tp.get("rate_per_qntl", 0) if tp else 0
        extra_paid = tp.get("paid_amount", 0) if tp else 0
        tp_status = tp.get("status", "") if tp else ""
        gross = round(qty * rate, 2) if rate > 0 else 0
        net = round(gross - deductions, 2) if gross > 0 else 0
        balance = round(net - extra_paid, 2) if gross > 0 else 0
        status = tp_status if tp_status else ("paid" if (gross > 0 and balance <= 0) else ("partial" if extra_paid > 0 else "pending"))
        payments.append(TruckPaymentStatus(
            entry_id=dd.get("id", ""),
            truck_no=truck_no,
            date=dd.get("date", ""),
            total_qntl=qty,
            total_bag=0,
            final_qntl=qty,
            cash_taken=cash_paid,
            diesel_taken=diesel_paid,
            rate_per_qntl=rate,
            gross_amount=gross,
            deductions=deductions,
            net_amount=net,
            paid_amount=round(deductions + extra_paid, 2),
            balance_amount=max(0, balance),
            status=status,
            kms_year=dd.get("kms_year", ""),
            season=dd.get("season", ""),
            agent_name=dd.get("driver_name", ""),
            mandi_name=label,
            source="DC Delivery"
        ))

    # Sort by date descending (newest first)
    payments.sort(key=lambda x: x.date, reverse=True)
    return payments


@router.put("/truck-payments/{entry_id}/rate")
async def set_truck_rate(entry_id: str, request: SetRateRequest, username: str = "", role: str = ""):
    """Set rate for a specific truck entry - auto-updates all entries with same truck_no + mandi_name"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin rate set kar sakta hai")
    
    entry, source, final_qntl = await _find_truck_entry(entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    truck_no = entry.get("truck_no", "") or entry.get("vehicle_no", "")
    mandi_name = entry.get("mandi_name", "")
    updated_count = 0
    
    if source == "pvt":
        # For pvt paddy, set rate and create/update Jama (credit) ledger entry for truck
        pvt_truck = entry.get("truck_no", "")
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {"entry_id": entry_id, "rate_per_qntl": request.rate_per_qntl, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        if final_qntl > 0 and pvt_truck:
            new_gross = round(final_qntl * request.rate_per_qntl, 2)
            base_fields = {
                "kms_year": entry.get("kms_year", ""), "season": entry.get("season", ""),
                "created_by": username or "system", "updated_at": datetime.now(timezone.utc).isoformat()
            }
            existing_jama = await db.cash_transactions.find_one(
                {"linked_entry_id": entry_id, "reference": {"$regex": "^pvt_truck_jama:"}}, {"_id": 0}
            )
            party = entry.get("party_name", "")
            mandi_n = entry.get("mandi_name", "")
            party_label = f"{party} - {mandi_n}" if party and mandi_n else party
            if existing_jama:
                await db.cash_transactions.update_one(
                    {"id": existing_jama["id"]},
                    {"$set": {
                        "amount": new_gross,
                        "description": f"Pvt Paddy Truck: {pvt_truck} - {party_label} - {final_qntl}Q @ Rs.{request.rate_per_qntl}",
                        **base_fields
                    }}
                )
            else:
                await db.cash_transactions.insert_one({
                    "id": str(uuid.uuid4()), "date": entry.get("date", ""),
                    "account": "ledger", "txn_type": "jama",
                    "category": pvt_truck, "party_type": "Truck",
                    "description": f"Pvt Paddy Truck: {pvt_truck} - {party_label} - {final_qntl}Q @ Rs.{request.rate_per_qntl}",
                    "amount": new_gross, "bank_name": "",
                    "reference": f"pvt_truck_jama:{entry_id[:8]}",
                    "linked_entry_id": entry_id,
                    "created_at": datetime.now(timezone.utc).isoformat(), **base_fields
                })
        updated_count = 1
    elif source == "sale_book":
        # For sale book entries, set rate and create/update Jama (credit) ledger entry
        sb_truck = entry.get("truck_no", "")
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {"entry_id": entry_id, "rate_per_qntl": request.rate_per_qntl, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        if final_qntl > 0 and sb_truck:
            new_gross = round(final_qntl * request.rate_per_qntl, 2)
            cash_taken = float(entry.get("cash_paid", 0) or 0)
            diesel_taken = float(entry.get("diesel_paid", 0) or 0)
            deductions = cash_taken + diesel_taken
            net = round(new_gross - deductions, 2)
            # Upsert Jama entry in ledger
            existing_jama = await db.cash_transactions.find_one(
                {"linked_entry_id": entry_id, "reference": {"$regex": "^sale_truck_jama:"}}, {"_id": 0}
            )
            base_fields = {
                "kms_year": entry.get("kms_year", ""), "season": entry.get("season", ""),
                "created_by": username or "system", "updated_at": datetime.now(timezone.utc).isoformat()
            }
            if existing_jama:
                await db.cash_transactions.update_one(
                    {"id": existing_jama["id"]},
                    {"$set": {
                        "amount": new_gross,
                        "description": f"Sale Truck: {sb_truck} - {final_qntl}Q @ Rs.{request.rate_per_qntl}",
                        **base_fields
                    }}
                )
            else:
                jama_entry = {
                    "id": str(uuid.uuid4()), "date": entry.get("date", ""),
                    "account": "ledger", "txn_type": "jama",
                    "category": sb_truck, "party_type": "Truck",
                    "description": f"Sale Truck: {sb_truck} - {final_qntl}Q @ Rs.{request.rate_per_qntl}",
                    "amount": new_gross, "bank_name": "",
                    "reference": f"sale_truck_jama:{entry_id[:8]}",
                    "linked_entry_id": entry_id,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    **base_fields
                }
                await db.cash_transactions.insert_one(jama_entry)
        updated_count = 1
    elif source == "purchase_voucher":
        # Same logic as sale_book
        pv_truck = entry.get("truck_no", "")
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {"entry_id": entry_id, "rate_per_qntl": request.rate_per_qntl, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        if final_qntl > 0 and pv_truck:
            new_gross = round(final_qntl * request.rate_per_qntl, 2)
            cash_taken = float(entry.get("cash_paid", 0) or 0)
            diesel_taken = float(entry.get("diesel_paid", 0) or 0)
            deductions = cash_taken + diesel_taken
            base_fields = {
                "kms_year": entry.get("kms_year", ""), "season": entry.get("season", ""),
                "created_by": username or "system", "updated_at": datetime.now(timezone.utc).isoformat()
            }
            existing_jama = await db.cash_transactions.find_one(
                {"linked_entry_id": entry_id, "reference": {"$regex": "^purchase_truck_jama:"}}, {"_id": 0}
            )
            if existing_jama:
                await db.cash_transactions.update_one(
                    {"id": existing_jama["id"]},
                    {"$set": {"amount": new_gross, "description": f"Purchase Truck: {pv_truck} - {final_qntl}Q @ Rs.{request.rate_per_qntl}", **base_fields}}
                )
            else:
                await db.cash_transactions.insert_one({
                    "id": str(uuid.uuid4()), "date": entry.get("date", ""),
                    "account": "ledger", "txn_type": "jama",
                    "category": pv_truck, "party_type": "Truck",
                    "description": f"Purchase Truck: {pv_truck} - {final_qntl}Q @ Rs.{request.rate_per_qntl}",
                    "amount": new_gross, "bank_name": "",
                    "reference": f"purchase_truck_jama:{entry_id[:8]}",
                    "linked_entry_id": entry_id,
                    "created_at": datetime.now(timezone.utc).isoformat(), **base_fields
                })
        updated_count = 1
    elif source == "dc_delivery":
        # For DC delivery, set rate and create/update Jama (credit) ledger entry
        dc_truck = entry.get("vehicle_no", "")
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {"entry_id": entry_id, "rate_per_qntl": request.rate_per_qntl, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        if final_qntl > 0 and dc_truck:
            new_gross = round(final_qntl * request.rate_per_qntl, 2)
            cash_taken = float(entry.get("cash_paid", 0) or 0)
            diesel_taken = float(entry.get("diesel_paid", 0) or 0)
            deductions = cash_taken + diesel_taken
            net = round(new_gross - deductions, 2)
            # Upsert Jama entry in ledger
            existing_jama = await db.cash_transactions.find_one(
                {"linked_entry_id": entry_id, "reference": {"$regex": "^delivery_jama:"}}, {"_id": 0}
            )
            base_fields = {
                "kms_year": entry.get("kms_year", ""), "season": entry.get("season", ""),
                "created_by": username or "system", "updated_at": datetime.now(timezone.utc).isoformat()
            }
            if existing_jama:
                await db.cash_transactions.update_one(
                    {"id": existing_jama["id"]},
                    {"$set": {
                        "amount": new_gross,
                        "description": f"DC Delivery: {dc_truck} - {final_qntl}Q @ Rs.{request.rate_per_qntl}" + (f" (Ded: Rs.{deductions})" if deductions > 0 else ""),
                        **base_fields
                    }}
                )
            else:
                jama_entry = {
                    "id": str(uuid.uuid4()), "date": entry.get("date", ""),
                    "account": "ledger", "txn_type": "jama",
                    "category": dc_truck, "party_type": "Truck",
                    "description": f"DC Delivery: {dc_truck} - {final_qntl}Q @ Rs.{request.rate_per_qntl}" + (f" (Ded: Rs.{deductions})" if deductions > 0 else ""),
                    "amount": new_gross, "bank_name": "",
                    "reference": f"delivery_jama:{entry_id[:8]}",
                    "linked_entry_id": entry_id,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    **base_fields
                }
                await db.cash_transactions.insert_one(jama_entry)
        updated_count = 1
    elif truck_no and mandi_name:
        # Find all entries with same truck_no + mandi_name
        matching = await db.mill_entries.find(
            {"truck_no": truck_no, "mandi_name": {"$regex": f"^{mandi_name}$", "$options": "i"}}, {"_id": 0}
        ).to_list(None)
        for m in matching:
            await db.truck_payments.update_one(
                {"entry_id": m["id"]},
                {"$set": {"entry_id": m["id"], "rate_per_qntl": request.rate_per_qntl, "updated_at": datetime.now(timezone.utc).isoformat()}},
                upsert=True
            )
            # Update Jama ledger entry with new rate
            final_qntl = round(m.get("qntl", 0) - m.get("bag", 0) / 100, 2)
            if final_qntl > 0:
                new_gross = round(final_qntl * request.rate_per_qntl, 2)
                cash_taken = float(m.get("cash_paid", 0) or 0)
                diesel_taken = float(m.get("diesel_paid", 0) or 0)
                deductions = cash_taken + diesel_taken
                await db.cash_transactions.update_one(
                    {"linked_entry_id": m["id"], "reference": {"$regex": "^truck_entry:"}},
                    {"$set": {
                        "amount": new_gross,
                        "description": f"Truck Entry: {truck_no} - {final_qntl}Q @ Rs.{request.rate_per_qntl}" + (f" (Ded: Rs.{deductions})" if deductions > 0 else ""),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
        updated_count = len(matching)
    else:
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {"entry_id": entry_id, "rate_per_qntl": request.rate_per_qntl, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
    
    return {"success": True, "message": f"Rate ₹{request.rate_per_qntl}/QNTL set for {updated_count} entries", "updated_count": updated_count, "truck_no": truck_no, "mandi_name": mandi_name}


@router.post("/truck-payments/{entry_id}/pay")
async def make_truck_payment(entry_id: str, request: MakePaymentRequest, username: str = "", role: str = ""):
    """Record a payment for truck (partial or full)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin payment kar sakta hai")
    
    entry, source, final_qntl = await _find_truck_entry(entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Get or create payment record
    payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
    current_paid = payment_doc.get("paid_amount", 0) if payment_doc else 0
    payments_history = payment_doc.get("payments_history", []) if payment_doc else []
    
    new_paid = current_paid + request.amount + (request.round_off or 0)
    payments_history.append({
        "amount": request.amount + (request.round_off or 0),
        "date": datetime.now(timezone.utc).isoformat(),
        "note": request.note,
        "by": username
    })
    
    await db.truck_payments.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "entry_id": entry_id,
            "paid_amount": new_paid,
            "payments_history": payments_history,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    # Auto-create Cash Book Nikasi entry for truck payment
    round_off_amt = request.round_off or 0
    total_settled = round(request.amount + round_off_amt, 2)
    if request.amount > 0:
        truck_no = entry.get("truck_no", "") or entry.get("vehicle_no", "")
        kms_year = entry.get("kms_year", "")
        season = entry.get("season", "")
        pay_id = str(uuid.uuid4())
        cb_entry = {
            "id": pay_id,
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "account": "cash",
            "txn_type": "nikasi",
            "category": truck_no,
            "party_type": "Truck",
            "description": f"Truck Payment: {truck_no} - Rs.{request.amount}",
            "amount": round_amount(request.amount),
            "reference": f"truck_pay:{entry_id[:8]}",
            "kms_year": kms_year,
            "season": season,
            "created_by": username or "system",
            "linked_payment_id": f"truck:{entry_id}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(cb_entry)

        # Ledger Nikasi - reduce truck outstanding (includes round off)
        ledger_entry = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "account": "ledger",
            "txn_type": "nikasi",
            "category": truck_no,
            "party_type": "Truck",
            "description": f"Truck Payment: {truck_no} - Rs.{total_settled}" + (f" (Cash: {request.amount}, Round Off: {round_off_amt})" if round_off_amt else ""),
            "amount": round_amount(total_settled),
            "reference": f"truck_pay_ledger:{entry_id[:8]}",
            "kms_year": kms_year,
            "season": season,
            "created_by": username or "system",
            "linked_payment_id": f"truck_ledger:{entry_id}:{uuid.uuid4().hex[:6]}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(ledger_entry)
    
    return {"success": True, "message": f"Rs.{request.amount} payment recorded", "total_paid": new_paid}


@router.post("/truck-payments/{entry_id}/mark-paid")
async def mark_truck_paid(entry_id: str, username: str = "", role: str = ""):
    """Mark truck payment as fully paid"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin paid mark kar sakta hai")
    
    entry, source, final_qntl = await _find_truck_entry(entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
    rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
    
    cash_taken = float(entry.get("cash_paid", 0) or 0)
    diesel_taken = float(entry.get("diesel_paid", 0) or 0)
    net_amount = round((final_qntl * rate) - cash_taken - diesel_taken, 2)
    
    payments_history = payment_doc.get("payments_history", []) if payment_doc else []
    payments_history.append({
        "amount": net_amount,
        "date": datetime.now(timezone.utc).isoformat(),
        "note": "Full payment - marked as paid",
        "by": username
    })
    
    await db.truck_payments.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "entry_id": entry_id,
            "paid_amount": net_amount,
            "payments_history": payments_history,
            "status": "paid",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    # Auto-create Cash Book Nikasi entry
    if net_amount > 0:
        kms_year = entry.get("kms_year", "")
        season = entry.get("season", "")
        truck_no = entry.get("truck_no", "") or entry.get("vehicle_no", "")
        cb_entry = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "account": "cash",
            "txn_type": "nikasi",
            "category": truck_no,
            "party_type": "Truck",
            "description": f"Truck Payment: {truck_no} (Full - Mark Paid)",
            "amount": round_amount(net_amount),
            "reference": f"truck_markpaid:{entry_id}",
            "kms_year": kms_year,
            "season": season,
            "created_by": username or "system",
            "linked_payment_id": f"truck:{entry_id}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(cb_entry)

        # Ledger Nikasi - reduce truck outstanding
        ledger_entry = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "account": "ledger",
            "txn_type": "nikasi",
            "category": truck_no,
            "party_type": "Truck",
            "description": f"Truck Payment: {truck_no} (Full - Mark Paid)",
            "amount": round_amount(net_amount),
            "reference": f"truck_markpaid_ledger:{entry_id}",
            "kms_year": kms_year,
            "season": season,
            "created_by": username or "system",
            "linked_payment_id": f"truck_ledger_markpaid:{entry_id}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(ledger_entry)
    
    return {"success": True, "message": "Truck payment cleared"}


# ============ AGENT PAYMENT ENDPOINTS ============

@router.get("/agent-payments", response_model=List[AgentPaymentStatus])
async def get_agent_payments(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get all agent payments based on mandi targets (not achieved)"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    # Get all mandi targets
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    
    payments = []
    for target in targets:
        mandi_name = target["mandi_name"]
        target_qntl = target["target_qntl"]
        cutting_percent = target["cutting_percent"]
        cutting_qntl = round(target_qntl * cutting_percent / 100, 2)
        expected_total = target["expected_total"]
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        
        # Calculate amounts
        target_amount = round(target_qntl * base_rate, 2)
        cutting_amount = round(cutting_qntl * cutting_rate, 2)
        total_amount = round(target_amount + cutting_amount, 2)
        
        # Get achieved for this mandi (case-insensitive)
        entry_query = {
            "mandi_name": {"$regex": f"^{mandi_name}$", "$options": "i"},
            "kms_year": target["kms_year"],
            "season": target["season"]
        }
        pipeline = [
            {"$match": entry_query},
            {"$group": {
                "_id": None, 
                "total_final_w": {"$sum": "$final_w"},
                "total_tp_weight": {"$sum": "$tp_weight"},
                "agent_name": {"$first": "$agent_name"}
            }}
        ]
        result = await db.mill_entries.aggregate(pipeline).to_list(1)
        achieved_kg = result[0]["total_final_w"] if result else 0
        achieved_qntl = round(achieved_kg / 100, 2)
        tp_weight_qntl = round(result[0]["total_tp_weight"], 2) if result else 0  # tp_weight already in QNTL
        excess_weight = round(achieved_qntl - (target_qntl + cutting_qntl), 2)  # Achieved - (Target + Cutting)
        agent_name = result[0]["agent_name"] if result else mandi_name
        
        is_target_complete = achieved_qntl >= expected_total
        
        # Get payment: use ledger as source of truth (includes manual Cash Book payments)
        ledger_query_agent = {
            "account": "ledger", "txn_type": "nikasi",
            "category": {"$regex": f"^{mandi_name}$", "$options": "i"}
        }
        if target.get("kms_year"): ledger_query_agent["kms_year"] = target["kms_year"]
        if target.get("season"): ledger_query_agent["season"] = target["season"]
        agent_ledger_txns = await db.cash_transactions.find(ledger_query_agent, {"_id": 0}).to_list(50000)
        paid_amount = round(sum(t.get("amount", 0) for t in agent_ledger_txns), 2)
        
        balance = round(total_amount - paid_amount, 2)
        status = "paid" if balance <= 0 else ("partial" if paid_amount > 0 else "pending")
        
        payments.append(AgentPaymentStatus(
            mandi_name=mandi_name,
            agent_name=agent_name,
            target_qntl=target_qntl,
            cutting_percent=cutting_percent,
            cutting_qntl=cutting_qntl,
            base_rate=base_rate,
            cutting_rate=cutting_rate,
            target_amount=target_amount,
            cutting_amount=cutting_amount,
            total_amount=total_amount,
            tp_weight_qntl=tp_weight_qntl,
            achieved_qntl=achieved_qntl,
            excess_weight=excess_weight,
            is_target_complete=is_target_complete,
            paid_amount=paid_amount,
            balance_amount=max(0, balance),
            status=status,
            kms_year=target["kms_year"],
            season=target["season"]
        ))
    
    return payments


@router.post("/agent-payments/{mandi_name}/pay")
async def make_agent_payment(mandi_name: str, request: MakePaymentRequest, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Record a payment for agent/mandi (partial or full)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin payment kar sakta hai")
    
    # Get or create payment record
    payment_doc = await db.agent_payments.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    
    current_paid = payment_doc.get("paid_amount", 0) if payment_doc else 0
    payments_history = payment_doc.get("payments_history", []) if payment_doc else []
    
    new_paid = current_paid + request.amount + (request.round_off or 0)
    payments_history.append({
        "amount": request.amount + (request.round_off or 0),
        "date": datetime.now(timezone.utc).isoformat(),
        "note": request.note,
        "by": username
    })
    
    await db.agent_payments.update_one(
        {"mandi_name": mandi_name, "kms_year": kms_year, "season": season},
        {"$set": {
            "mandi_name": mandi_name,
            "kms_year": kms_year,
            "season": season,
            "paid_amount": new_paid,
            "payments_history": payments_history,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    # Auto-create Cash Book entries
    if request.amount > 0:
        # JAMA (Ledger) - Agent Commission (what we owe) - create once, update if needed
        # Calculate total_amount from mandi_targets
        target = await db.mandi_targets.find_one({"mandi_name": mandi_name, "kms_year": kms_year, "season": season}, {"_id": 0})
        if target:
            base_rate = target.get("base_rate", 0)
            cutting_rate = target.get("cutting_rate", 5)
            entries_for_mandi = await db.mill_entries.find(
                {"mandi_name": {"$regex": f"^{mandi_name}$", "$options": "i"}, "kms_year": kms_year},
                {"_id": 0, "qntl": 1, "bag": 1}
            ).to_list(None)
            achieved_qntl = round(sum(e.get("qntl", 0) - e.get("bag", 0) / 100 for e in entries_for_mandi), 2)
            target_qntl = target.get("target_qntl", 0)
            cutting_percent = target.get("cutting_percent", 0)
            cutting_qntl = round(achieved_qntl * cutting_percent / 100, 2) if cutting_percent > 0 else 0
            total_amount = round((target_qntl * base_rate) + (cutting_qntl * cutting_rate), 2)
            
            linked_id = f"agent_jama:{mandi_name}:{kms_year}:{season}"
            existing_jama = await db.cash_transactions.find_one({"linked_payment_id": linked_id}, {"_id": 0})
            if not existing_jama and total_amount > 0:
                jama_entry = {
                    "id": str(uuid.uuid4()),
                    "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "account": "ledger", "txn_type": "jama",
                    "category": mandi_name, "party_type": "Agent",
                    "description": f"Agent Commission: {mandi_name} @ Rs.{base_rate}",
                    "amount": round_amount(total_amount),
                    "reference": f"agent_comm:{mandi_name[:10]}",
                    "kms_year": kms_year, "season": season,
                    "created_by": username or "system",
                    "linked_payment_id": linked_id,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
                await db.cash_transactions.insert_one(jama_entry)
            elif existing_jama and total_amount > 0:
                await db.cash_transactions.update_one(
                    {"linked_payment_id": linked_id},
                    {"$set": {"amount": round_amount(total_amount),
                              "description": f"Agent Commission: {mandi_name} @ Rs.{base_rate}",
                              "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
        
        # NIKASI (Cash) - Agent Payment
        cb_entry = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "account": "cash",
            "txn_type": "nikasi",
            "category": mandi_name,
            "party_type": "Agent",
            "description": f"Agent Payment: {mandi_name} - Rs.{request.amount}",
            "amount": round_amount(request.amount),
            "reference": f"agent_pay:{mandi_name[:10]}",
            "kms_year": kms_year,
            "season": season,
            "created_by": username or "system",
            "linked_payment_id": f"agent:{mandi_name}:{kms_year}:{season}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(cb_entry)

        # NIKASI (Ledger) - Reduce agent outstanding in party ledger (includes round off)
        agent_ro = request.round_off or 0
        agent_total = round(request.amount + agent_ro, 2)
        ledger_nikasi = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "account": "ledger",
            "txn_type": "nikasi",
            "category": mandi_name,
            "party_type": "Agent",
            "description": f"Agent Payment: {mandi_name} - Rs.{agent_total}" + (f" (Cash: {request.amount}, Round Off: {agent_ro})" if agent_ro else ""),
            "amount": agent_total,
            "reference": f"agent_pay_ledger:{mandi_name[:10]}",
            "kms_year": kms_year,
            "season": season,
            "created_by": username or "system",
            "linked_payment_id": f"agent_ledger_pay:{mandi_name}:{kms_year}:{season}:{uuid.uuid4().hex[:6]}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(ledger_nikasi)
    
    return {"success": True, "message": f"Rs.{request.amount} payment recorded", "total_paid": new_paid}


@router.post("/agent-payments/{mandi_name}/mark-paid")
async def mark_agent_paid(mandi_name: str, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Mark agent/mandi payment as fully paid"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin paid mark kar sakta hai")
    
    # Get target for this mandi
    target = await db.mandi_targets.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    
    if not target:
        raise HTTPException(status_code=404, detail="Mandi target not found")
    
    # Calculate total amount based on target
    target_qntl = target["target_qntl"]
    cutting_qntl = target_qntl * target["cutting_percent"] / 100
    base_rate = target.get("base_rate", 10)
    cutting_rate = target.get("cutting_rate", 5)
    total_amount = (target_qntl * base_rate) + (cutting_qntl * cutting_rate)
    
    payment_doc = await db.agent_payments.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    payments_history = payment_doc.get("payments_history", []) if payment_doc else []
    payments_history.append({
        "amount": total_amount,
        "date": datetime.now(timezone.utc).isoformat(),
        "note": "Full payment - marked as paid",
        "by": username
    })
    
    await db.agent_payments.update_one(
        {"mandi_name": mandi_name, "kms_year": kms_year, "season": season},
        {"$set": {
            "mandi_name": mandi_name,
            "kms_year": kms_year,
            "season": season,
            "paid_amount": total_amount,
            "payments_history": payments_history,
            "status": "paid",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    # Auto-create Cash Book entries for agent mark-paid
    if total_amount > 0:
        # JAMA (Ledger) - Agent Commission entry
        entries_for_mandi = await db.mill_entries.find(
            {"mandi_name": {"$regex": f"^{mandi_name}$", "$options": "i"}, "kms_year": kms_year},
            {"_id": 0, "qntl": 1, "bag": 1}
        ).to_list(None)
        achieved_qntl = round(sum(e.get("qntl", 0) - e.get("bag", 0) / 100 for e in entries_for_mandi), 2)
        cutting_qntl = round(achieved_qntl * target.get("cutting_percent", 0) / 100, 2) if target.get("cutting_percent", 0) > 0 else 0

        linked_jama_id = f"agent_jama:{mandi_name}:{kms_year}:{season}"
        existing_jama = await db.cash_transactions.find_one({"linked_payment_id": linked_jama_id}, {"_id": 0})
        if not existing_jama:
            jama_entry = {
                "id": str(uuid.uuid4()),
                "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "account": "ledger", "txn_type": "jama",
                "category": mandi_name, "party_type": "Agent",
                "description": f"Agent Commission: {mandi_name} @ Rs.{base_rate}",
                "amount": round_amount(total_amount),
                "reference": f"agent_comm:{mandi_name[:10]}",
                "kms_year": kms_year, "season": season,
                "created_by": username or "system",
                "linked_payment_id": linked_jama_id,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.cash_transactions.insert_one(jama_entry)
        elif existing_jama:
            await db.cash_transactions.update_one(
                {"linked_payment_id": linked_jama_id},
                {"$set": {"amount": round_amount(total_amount),
                          "description": f"Agent Commission: {mandi_name} @ Rs.{base_rate}",
                          "updated_at": datetime.now(timezone.utc).isoformat()}}
            )

        # NIKASI (Cash) - Agent Payment
        cb_entry = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "account": "cash",
            "txn_type": "nikasi",
            "category": mandi_name,
            "party_type": "Agent",
            "description": f"Agent Payment: {mandi_name} (Full - Mark Paid)",
            "amount": round_amount(total_amount),
            "reference": f"agent_markpaid:{mandi_name[:10]}",
            "kms_year": kms_year,
            "season": season,
            "created_by": username or "system",
            "linked_payment_id": f"agent:{mandi_name}:{kms_year}:{season}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(cb_entry)

        # NIKASI (Ledger) - Reduce agent outstanding in party ledger
        ledger_nikasi = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "account": "ledger",
            "txn_type": "nikasi",
            "category": mandi_name,
            "party_type": "Agent",
            "description": f"Agent Payment: {mandi_name} (Full - Mark Paid)",
            "amount": round_amount(total_amount),
            "reference": f"agent_markpaid_ledger:{mandi_name[:10]}",
            "kms_year": kms_year,
            "season": season,
            "created_by": username or "system",
            "linked_payment_id": f"agent_ledger_markpaid:{mandi_name}:{kms_year}:{season}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(ledger_nikasi)
    
    return {"success": True, "message": "Agent/Mandi payment cleared"}


@router.post("/truck-payments/{entry_id}/undo-paid")
async def undo_truck_paid(entry_id: str, username: str = "", role: str = ""):
    """Undo paid status - reset payment to 0 (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin undo kar sakta hai")
    
    # Find the entry first to get truck_no
    entry, source, _ = await _find_truck_entry(entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    truck_no = entry.get("truck_no", "") or entry.get("vehicle_no", "")
    
    # Reset truck_payments if exists
    payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
    if payment_doc:
        payments_history = payment_doc.get("payments_history", [])
        payments_history.append({
            "amount": -payment_doc.get("paid_amount", 0),
            "date": datetime.now(timezone.utc).isoformat(),
            "note": "UNDO - Payment reversed",
            "by": username
        })
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {
                "paid_amount": 0,
                "payments_history": payments_history,
                "status": "pending",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # Delete ALL payment-related cash book entries for this truck entry
    # 1. Cash entries from Pay/Mark Paid buttons
    await db.cash_transactions.delete_many({"linked_payment_id": f"truck:{entry_id}"})
    # 2. Ledger entries from Pay button
    await db.cash_transactions.delete_many({"linked_payment_id": {"$regex": f"^truck_ledger:{entry_id}"}})
    # 3. Ledger entries from Mark Paid button
    await db.cash_transactions.delete_many({"linked_payment_id": f"truck_ledger_markpaid:{entry_id}"})
    # 4. Owner-level cash entries
    if truck_no:
        await db.cash_transactions.delete_many({
            "linked_payment_id": {"$regex": f"^truck_owner:{truck_no}:"}
        })
    
    # 5. Delete manual Cash Book entries for this truck that are NOT auto-deductions
    #    (entries with auto_ledger reference or manual entries linked to this truck via linked_entry_id)
    eid_short = entry_id[:8]
    deduction_refs = [f"truck_cash_ded:{eid_short}", f"truck_diesel_ded:{eid_short}", f"entry_cash:{eid_short}", f"truck_entry:{eid_short}",
                      f"delivery:{eid_short}", f"delivery_tcash:{eid_short}", f"delivery_tdiesel:{eid_short}", f"delivery_diesel:{eid_short}", f"delivery_jama:{eid_short}",
                      f"sale_truck_jama:{eid_short}", f"sale_truck_cash:{eid_short}", f"sale_truck_diesel:{eid_short}",
                      f"purchase_truck_jama:{eid_short}", f"purchase_truck_cash:{eid_short}", f"purchase_truck_diesel:{eid_short}",
                      f"pvt_truck_jama:{eid_short}", f"pvt_paddy_tcash:{eid_short}", f"pvt_paddy_tdiesel:{eid_short}", f"pvt_paddy_cash:{eid_short}"]
    
    # Find and delete non-deduction ledger nikasi entries for this truck that are NOT auto-generated from entry
    all_ledger = await db.cash_transactions.find({
        "account": "ledger", "txn_type": "nikasi", "category": truck_no
    }, {"_id": 0, "id": 1, "reference": 1}).to_list(50000)
    
    ids_to_delete = []
    for txn in all_ledger:
        ref = txn.get("reference", "")
        if not any(ref.startswith(dp) for dp in deduction_refs):
            ids_to_delete.append(txn["id"])
    
    if ids_to_delete:
        await db.cash_transactions.delete_many({"id": {"$in": ids_to_delete}})
    
    # Also delete corresponding cash entries (manual Cash Book payments for the truck)
    all_cash = await db.cash_transactions.find({
        "account": "cash", "txn_type": "nikasi", "category": truck_no
    }, {"_id": 0, "id": 1, "reference": 1}).to_list(50000)
    
    cash_ids_to_delete = []
    for txn in all_cash:
        ref = txn.get("reference", "")
        if not any(ref.startswith(dp) for dp in deduction_refs):
            cash_ids_to_delete.append(txn["id"])
    
    if cash_ids_to_delete:
        await db.cash_transactions.delete_many({"id": {"$in": cash_ids_to_delete}})
    
    return {"success": True, "message": "Payment undo ho gaya - sab entries delete ho gayi"}


@router.post("/agent-payments/{mandi_name}/undo-paid")
async def undo_agent_paid(mandi_name: str, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Undo paid status - reset agent payment to 0 (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin undo kar sakta hai")
    
    payment_doc = await db.agent_payments.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    
    if not payment_doc:
        raise HTTPException(status_code=404, detail="Payment record not found")
    
    payments_history = payment_doc.get("payments_history", [])
    payments_history.append({
        "amount": -payment_doc.get("paid_amount", 0),
        "date": datetime.now(timezone.utc).isoformat(),
        "note": "UNDO - Payment reversed",
        "by": username
    })
    
    await db.agent_payments.update_one(
        {"mandi_name": mandi_name, "kms_year": kms_year, "season": season},
        {"$set": {
            "paid_amount": 0,
            "payments_history": payments_history,
            "status": "pending",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Delete linked cash book entries for this agent (both nikasi and jama)
    await db.cash_transactions.delete_many({"linked_payment_id": f"agent:{mandi_name}:{kms_year}:{season}"})
    await db.cash_transactions.delete_many({"linked_payment_id": f"agent_jama:{mandi_name}:{kms_year}:{season}"})
    # Delete linked LEDGER entries (from Pay button and Mark Paid button)
    await db.cash_transactions.delete_many({"linked_payment_id": {"$regex": f"^agent_ledger_pay:{mandi_name}:{kms_year}:{season}"}})
    await db.cash_transactions.delete_many({"linked_payment_id": f"agent_ledger_markpaid:{mandi_name}:{kms_year}:{season}"})
    
    return {"success": True, "message": "Payment undo ho gaya - status reset to pending"}


@router.get("/truck-payments/{entry_id}/history")
async def get_truck_payment_history(entry_id: str):
    """Get payment history for a truck entry - includes both Pay button and manual Cash Book payments"""
    # Get truck_payments history (from Pay/Mark Paid buttons)
    payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
    button_history = payment_doc.get("payments_history", []) if payment_doc else []
    
    # Also get ledger-based payment history from cash_transactions
    entry = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        return {"history": button_history, "total_paid": payment_doc.get("paid_amount", 0) if payment_doc else 0}
    
    truck_no = entry.get("truck_no", "")
    eid_short = entry_id[:8]
    DEDUCTION_PREFIXES = [f"truck_cash_ded:{eid_short}", f"truck_diesel_ded:{eid_short}", f"entry_cash:{eid_short}"]
    
    # Get all ledger nikasi entries for this truck (payments, not deductions)
    ledger_payments = await db.cash_transactions.find({
        "account": "ledger", "txn_type": "nikasi", "category": truck_no
    }, {"_id": 0}).to_list(50000)
    
    # Build combined history from ledger entries
    ledger_history = []
    for txn in ledger_payments:
        ref = txn.get("reference", "")
        if any(ref.startswith(p) for p in DEDUCTION_PREFIXES):
            continue
        ledger_history.append({
            "amount": txn.get("amount", 0),
            "date": txn.get("created_at") or txn.get("date", ""),
            "note": txn.get("description", ""),
            "by": txn.get("created_by", "system"),
            "source": "ledger"
        })
    
    # Use ledger history as the source of truth (it includes ALL payments)
    # Sort by date
    all_history = sorted(ledger_history, key=lambda h: h.get("date", ""), reverse=True)
    total_paid = round(sum(h.get("amount", 0) for h in all_history), 2)
    
    return {"history": all_history, "total_paid": total_paid}


@router.get("/agent-payments/{mandi_name}/history")
async def get_agent_payment_history(mandi_name: str, kms_year: str = "", season: str = ""):
    """Get payment history for an agent/mandi - includes both Pay button and manual Cash Book payments"""
    # Get agent_payments history (from Pay/Mark Paid buttons)
    payment_doc = await db.agent_payments.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    
    # Get all ledger nikasi entries for this agent (payments)
    ledger_payments = await db.cash_transactions.find({
        "account": "ledger", "txn_type": "nikasi", "category": mandi_name
    }, {"_id": 0}).to_list(50000)
    
    ledger_history = []
    for txn in ledger_payments:
        if kms_year and txn.get("kms_year") != kms_year:
            continue
        if season and txn.get("season") != season:
            continue
        ledger_history.append({
            "amount": txn.get("amount", 0),
            "date": txn.get("created_at") or txn.get("date", ""),
            "note": txn.get("description", ""),
            "by": txn.get("created_by", "system"),
            "source": "ledger"
        })
    
    all_history = sorted(ledger_history, key=lambda h: h.get("date", ""), reverse=True)
    total_paid = round(sum(h.get("amount", 0) for h in all_history), 2)
    
    return {"history": all_history, "total_paid": total_paid}



# ===== TRUCK OWNER CONSOLIDATED PAYMENT ENDPOINTS =====

@router.post("/truck-owner/{truck_no}/pay")
async def pay_truck_owner(truck_no: str, request: Request, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Make partial payment to truck owner - distributes across all trips"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin payment kar sakta hai")
    
    body = await request.json()
    amount = float(body.get("amount", 0))
    note = body.get("note", "")
    payment_mode = body.get("payment_mode", "cash")
    
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount 0 se zyada hona chahiye")
    
    # Get all entries for this truck
    query = {"truck_no": truck_no}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(None)
    if not entries:
        raise HTTPException(status_code=404, detail="Is truck ke entries nahi mile")
    
    # Distribute payment across unpaid trips (oldest first)
    remaining = amount
    for entry in sorted(entries, key=lambda e: e.get("date", "")):
        if remaining <= 0:
            break
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        rate = payment_doc.get("rate_per_qntl", 0) if payment_doc else 0
        paid_so_far = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
        cash_taken = float(entry.get("cash_paid", 0) or 0)
        diesel_taken = float(entry.get("diesel_paid", 0) or 0)
        gross = round(final_qntl * rate, 2)
        net = round(gross - cash_taken - diesel_taken, 2)
        trip_balance = max(0, round(net - paid_so_far, 2))
        
        if trip_balance <= 0:
            continue
        
        allot = min(remaining, trip_balance)
        new_paid = round(paid_so_far + allot, 2)
        new_balance = max(0, round(net - new_paid, 2))
        
        history = payment_doc.get("payments_history", []) if payment_doc else []
        history.append({
            "amount": allot, "date": datetime.now(timezone.utc).isoformat(),
            "note": f"Owner Payment: {note}" if note else "Owner Payment",
            "by": username, "payment_mode": payment_mode
        })
        
        if new_balance < 0.10:
            status = "paid"
        elif new_paid > 0:
            status = "partial"
        else:
            status = "pending"
        
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {
                "entry_id": entry_id, "paid_amount": new_paid,
                "payments_history": history, "status": status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        remaining = round(remaining - allot, 2)
    
    # Create cash book entry (Cash/Bank Nikasi)
    round_off = float(body.get("round_off", 0))
    owner_total = round(amount + round_off, 2)
    txn_id = f"txn_{datetime.now().strftime('%Y%m%d%H%M%S')}_{truck_no}"
    cash_txn = {
        "id": txn_id,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "account": payment_mode,
        "txn_type": "nikasi",
        "category": truck_no,
        "party_type": "Truck",
        "description": f"Truck Owner Payment: {truck_no}" + (f" - {note}" if note else ""),
        "amount": amount,
        "reference": f"truck_owner:{truck_no}",
        "linked_payment_id": f"truck_owner:{truck_no}:{kms_year}:{season}",
        "kms_year": kms_year,
        "season": season,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(cash_txn)

    # Ledger Nikasi - reduce truck outstanding (includes round off)
    ledger_txn = {
        "id": f"txn_ledger_{datetime.now().strftime('%Y%m%d%H%M%S')}_{truck_no}",
        "date": datetime.now().strftime("%Y-%m-%d"),
        "account": "ledger",
        "txn_type": "nikasi",
        "category": truck_no,
        "party_type": "Truck",
        "description": f"Truck Owner Payment: {truck_no} - Rs.{owner_total}" + (f" (Cash: {amount}, Round Off: {round_off})" if round_off else "") + (f" - {note}" if note else ""),
        "amount": owner_total,
        "reference": f"truck_owner_ledger:{truck_no}",
        "linked_payment_id": f"truck_owner_ledger:{truck_no}:{kms_year}:{season}:{uuid.uuid4().hex[:6]}",
        "kms_year": kms_year,
        "season": season,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(ledger_txn)
    
    # Store owner payment history
    owner_doc = await db.truck_owner_payments.find_one({"truck_no": truck_no, "kms_year": kms_year, "season": season})
    owner_history = owner_doc.get("payments_history", []) if owner_doc else []
    owner_history.append({
        "amount": amount, "date": datetime.now(timezone.utc).isoformat(),
        "note": note, "by": username, "payment_mode": payment_mode
    })
    await db.truck_owner_payments.update_one(
        {"truck_no": truck_no, "kms_year": kms_year, "season": season},
        {"$set": {"payments_history": owner_history, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"success": True, "message": f"₹{amount:,.0f} payment ho gaya! ({round(amount - remaining, 0)} distributed)"}


@router.post("/truck-owner/{truck_no}/mark-paid")
async def mark_truck_owner_paid(truck_no: str, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Mark all trips for a truck as fully paid"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin mark paid kar sakta hai")
    
    query = {"truck_no": truck_no}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(None)
    
    # Also include dc_deliveries for this truck
    dc_query = {"vehicle_no": truck_no}
    if kms_year: dc_query["kms_year"] = kms_year
    if season: dc_query["season"] = season
    dc_entries = await db.dc_deliveries.find(dc_query, {"_id": 0}).to_list(None)
    entries = entries + dc_entries
    
    if not entries:
        raise HTTPException(status_code=404, detail="Entries nahi mile")
    
    total_marked = 0
    for entry in entries:
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        rate = payment_doc.get("rate_per_qntl", 0) if payment_doc else 0
        paid_so_far = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
        gross = round(final_qntl * rate, 2)
        deductions = float(entry.get("cash_paid", 0) or 0) + float(entry.get("diesel_paid", 0) or 0)
        net = round(gross - deductions, 2)
        
        if paid_so_far >= net and net > 0:
            continue  # Already paid
        
        trip_balance = max(0, round(net - paid_so_far, 2))
        total_marked += trip_balance
        
        history = payment_doc.get("payments_history", []) if payment_doc else []
        history.append({
            "amount": trip_balance, "date": datetime.now(timezone.utc).isoformat(),
            "note": "Owner Mark Paid (Full)", "by": username
        })
        
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {
                "entry_id": entry_id, "paid_amount": net,
                "payments_history": history, "status": "paid",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
    
    if total_marked > 0:
        # Cash book nikasi
        txn_id = f"txn_{datetime.now().strftime('%Y%m%d%H%M%S')}_{truck_no}_full"
        await db.cash_transactions.insert_one({
            "id": txn_id, "date": datetime.now().strftime("%Y-%m-%d"),
            "account": "cash", "txn_type": "nikasi",
            "category": truck_no, "party_type": "Truck",
            "description": f"Truck Owner Full Payment: {truck_no}",
            "amount": total_marked,
            "reference": f"truck_owner:{truck_no}",
            "linked_payment_id": f"truck_owner:{truck_no}:{kms_year}:{season}",
            "kms_year": kms_year, "season": season,
            "created_at": datetime.now(timezone.utc).isoformat()
        })

        # Ledger Nikasi - reduce truck outstanding
        await db.cash_transactions.insert_one({
            "id": f"txn_ledger_{datetime.now().strftime('%Y%m%d%H%M%S')}_{truck_no}_full",
            "date": datetime.now().strftime("%Y-%m-%d"),
            "account": "ledger", "txn_type": "nikasi",
            "category": truck_no, "party_type": "Truck",
            "description": f"Truck Owner Full Payment: {truck_no}",
            "amount": total_marked,
            "reference": f"truck_owner_ledger:{truck_no}",
            "linked_payment_id": f"truck_owner_ledger_markpaid:{truck_no}:{kms_year}:{season}",
            "kms_year": kms_year, "season": season,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    return {"success": True, "message": f"Sab trips paid! ₹{total_marked:,.0f} mark paid kiya"}


@router.post("/truck-owner/{truck_no}/undo-paid")
async def undo_truck_owner_paid(truck_no: str, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Undo all payments for a truck owner - resets all trips"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin undo kar sakta hai")
    
    query = {"truck_no": truck_no}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(None)
    
    # Also include dc_deliveries for this truck
    dc_query = {"vehicle_no": truck_no}
    if kms_year: dc_query["kms_year"] = kms_year
    if season: dc_query["season"] = season
    dc_dels = await db.dc_deliveries.find(dc_query, {"_id": 0}).to_list(None)
    
    # Also include pvt_paddy and rice_sales (exclude agent_extra entries)
    pvt_query = {"truck_no": truck_no, "source": {"$ne": "agent_extra"}}
    if kms_year: pvt_query["kms_year"] = kms_year
    if season: pvt_query["season"] = season
    pvt_entries = await db.private_paddy.find(pvt_query, {"_id": 0}).to_list(None)
    rice_entries = await db.rice_sales.find(pvt_query, {"_id": 0}).to_list(None)
    
    all_entries = entries + dc_dels + pvt_entries + rice_entries
    
    for entry in all_entries:
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        if payment_doc:
            history = payment_doc.get("payments_history", [])
            history.append({
                "amount": -payment_doc.get("paid_amount", 0),
                "date": datetime.now(timezone.utc).isoformat(),
                "note": "UNDO - Owner payment reversed", "by": username
            })
            await db.truck_payments.update_one(
                {"entry_id": entry_id},
                {"$set": {"paid_amount": 0, "payments_history": history, "status": "pending", "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
        # Also delete individual linked entries
        await db.cash_transactions.delete_many({"linked_payment_id": f"truck:{entry_id}"})
    
    # Delete owner-level cash transactions
    await db.cash_transactions.delete_many({"linked_payment_id": f"truck_owner:{truck_no}:{kms_year}:{season}"})
    
    return {"success": True, "message": f"{truck_no} ke saare payments undo ho gaye"}


@router.get("/truck-owner/{truck_no}/history")
async def get_truck_owner_history(truck_no: str, kms_year: str = "", season: str = ""):
    """Get consolidated payment history for a truck owner - uses ledger as source of truth"""
    # Get all ledger nikasi entries for this truck (payments, not deductions)
    ledger_query = {"account": "ledger", "txn_type": "nikasi", "category": truck_no}
    if kms_year: ledger_query["kms_year"] = kms_year
    if season: ledger_query["season"] = season
    ledger_payments = await db.cash_transactions.find(ledger_query, {"_id": 0}).to_list(50000)
    
    # Get entry IDs for this truck to identify deduction prefixes
    query = {"truck_no": truck_no}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.mill_entries.find(query, {"_id": 0, "id": 1}).to_list(None)
    # Also include dc_deliveries
    dc_query = {"vehicle_no": truck_no}
    if kms_year: dc_query["kms_year"] = kms_year
    if season: dc_query["season"] = season
    dc_entries = await db.dc_deliveries.find(dc_query, {"_id": 0, "id": 1}).to_list(None)
    entry_short_ids = [e["id"][:8] for e in entries] + [e["id"][:8] for e in dc_entries]
    
    all_history = []
    for txn in ledger_payments:
        ref = txn.get("reference", "")
        # Skip deduction entries (auto-created from entries/deliveries)
        is_deduction = False
        for eid in entry_short_ids:
            if (ref.startswith(f"truck_cash_ded:{eid}") or ref.startswith(f"truck_diesel_ded:{eid}") or 
                ref.startswith(f"entry_cash:{eid}") or ref.startswith(f"delivery_tcash:{eid}") or 
                ref.startswith(f"delivery_tdiesel:{eid}") or ref.startswith(f"delivery:{eid}")):
                is_deduction = True
                break
        if is_deduction:
            continue
        all_history.append({
            "amount": txn.get("amount", 0),
            "date": txn.get("created_at") or txn.get("date", ""),
            "note": txn.get("description", ""),
            "by": txn.get("created_by", "system"),
            "source": "ledger"
        })
    
    all_history.sort(key=lambda x: x.get("date", ""), reverse=True)
    return {"history": all_history}


@router.get("/export/agent-payments-excel")
async def export_agent_payments_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Export agent/mandi payments to styled Excel file"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    
    payments_data = []
    total_amount_sum = 0
    total_paid_sum = 0
    total_balance_sum = 0
    
    for target in targets:
        mandi_name = target["mandi_name"]
        target_qntl = target["target_qntl"]
        cutting_qntl = round(target_qntl * target["cutting_percent"] / 100, 2)
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        
        target_amount = round(target_qntl * base_rate, 2)
        cutting_amount = round(cutting_qntl * cutting_rate, 2)
        total_amount = round(target_amount + cutting_amount, 2)
        
        # Get achieved + TP Weight
        entry_query = {"mandi_name": mandi_name, "kms_year": target["kms_year"], "season": target["season"]}
        pipeline = [{"$match": entry_query}, {"$group": {"_id": None, "total_final_w": {"$sum": "$final_w"}, "total_tp_weight": {"$sum": "$tp_weight"}, "agent_name": {"$first": "$agent_name"}}}]
        result = await db.mill_entries.aggregate(pipeline).to_list(1)
        achieved_qntl = round(result[0]["total_final_w"] / 100, 2) if result else 0
        tp_weight_qntl = round(result[0]["total_tp_weight"], 2) if result else 0  # already QNTL
        excess_weight = round(achieved_qntl - (target_qntl + cutting_qntl), 2)  # Achieved - (Target + Cutting)
        agent_name = result[0]["agent_name"] if result else mandi_name
        
        # Get payment
        payment_doc = await db.agent_payments.find_one({"mandi_name": mandi_name, "kms_year": target["kms_year"], "season": target["season"]}, {"_id": 0})
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        balance = round(max(0, total_amount - paid_amount), 2)
        status = "Paid" if balance <= 0 else ("Partial" if paid_amount > 0 else "Pending")
        
        total_amount_sum += total_amount
        total_paid_sum += paid_amount
        total_balance_sum += balance
        
        payments_data.append({
            "mandi_name": mandi_name,
            "agent_name": agent_name,
            "target_qntl": target_qntl,
            "cutting_qntl": cutting_qntl,
            "base_rate": base_rate,
            "cutting_rate": cutting_rate,
            "target_amount": target_amount,
            "cutting_amount": cutting_amount,
            "total_amount": total_amount,
            "tp_weight_qntl": tp_weight_qntl,
            "achieved_qntl": achieved_qntl,
            "excess_weight": excess_weight,
            "paid": paid_amount,
            "balance": balance,
            "status": status
        })
    
    # Create Excel
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Agent Payments"
    ncols = 15
    
    company_name, tagline = await get_company_name()
    title = f"Agent/Mandi Payments - {company_name}"
    subtitle = f"FY: {kms_year or 'All'} | {season or 'All'}"
    style_excel_title(ws, title, ncols, subtitle)
    
    headers = ["Mandi", "Agent", "Target QNTL", "Cutting QNTL", "Base Rate", "Cut Rate", "Target Amt", "Cut Amt", "Total Amt", "TP Weight", "Achieved", "Excess Wt", "Paid", "Balance", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_excel_header_row(ws, 4, ncols)
    
    data_start = 5
    for row_idx, p in enumerate(payments_data, data_start):
        ws.cell(row=row_idx, column=1, value=p["mandi_name"])
        ws.cell(row=row_idx, column=2, value=p["agent_name"])
        ws.cell(row=row_idx, column=3, value=p["target_qntl"])
        ws.cell(row=row_idx, column=4, value=p["cutting_qntl"])
        ws.cell(row=row_idx, column=5, value=p["base_rate"])
        ws.cell(row=row_idx, column=6, value=p["cutting_rate"])
        ws.cell(row=row_idx, column=7, value=p["target_amount"])
        ws.cell(row=row_idx, column=8, value=p["cutting_amount"])
        ws.cell(row=row_idx, column=9, value=p["total_amount"])
        ws.cell(row=row_idx, column=10, value=p["tp_weight_qntl"])
        ws.cell(row=row_idx, column=11, value=p["achieved_qntl"])
        ws.cell(row=row_idx, column=12, value=p["excess_weight"])
        ws.cell(row=row_idx, column=13, value=p["paid"])
        ws.cell(row=row_idx, column=14, value=p["balance"])
        ws.cell(row=row_idx, column=15, value=p["status"])
    
    if payments_data:
        style_excel_data_rows(ws, data_start, data_start + len(payments_data) - 1, ncols, headers)
    
    total_row = data_start + len(payments_data)
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=9, value=round(total_amount_sum, 2))
    ws.cell(row=total_row, column=13, value=round(total_paid_sum, 2))
    ws.cell(row=total_row, column=14, value=round(total_balance_sum, 2))
    style_excel_total_row(ws, total_row, ncols)
    
    col_widths = [14, 12, 12, 12, 10, 10, 12, 10, 12, 12, 12, 12, 10, 12, 10]
    from openpyxl.utils import get_column_letter
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"agent_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/agent-payments-pdf")
async def export_agent_payments_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Export agent/mandi payments to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.enums import TA_CENTER
    
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    
    payments_data = []
    total_amount_sum = 0
    total_paid_sum = 0
    total_balance_sum = 0
    
    for target in targets:
        mandi_name = target["mandi_name"]
        target_qntl = target["target_qntl"]
        cutting_qntl = round(target_qntl * target["cutting_percent"] / 100, 2)
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        
        target_amount = round(target_qntl * base_rate, 2)
        cutting_amount = round(cutting_qntl * cutting_rate, 2)
        total_amount = round(target_amount + cutting_amount, 2)
        
        entry_query = {"mandi_name": mandi_name, "kms_year": target["kms_year"], "season": target["season"]}
        pipeline = [{"$match": entry_query}, {"$group": {"_id": None, "total_final_w": {"$sum": "$final_w"}, "total_tp_weight": {"$sum": "$tp_weight"}, "agent_name": {"$first": "$agent_name"}}}]
        result = await db.mill_entries.aggregate(pipeline).to_list(1)
        achieved_qntl = round(result[0]["total_final_w"] / 100, 2) if result else 0
        tp_weight_qntl = round(result[0]["total_tp_weight"], 2) if result else 0  # already QNTL
        excess_weight = round(achieved_qntl - (target_qntl + cutting_qntl), 2)  # Achieved - (Target + Cutting)
        
        payment_doc = await db.agent_payments.find_one({"mandi_name": mandi_name, "kms_year": target["kms_year"], "season": target["season"]}, {"_id": 0})
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        balance = round(max(0, total_amount - paid_amount), 2)
        status = "Paid" if balance <= 0 else ("Partial" if paid_amount > 0 else "Pending")
        
        total_amount_sum += total_amount
        total_paid_sum += paid_amount
        total_balance_sum += balance
        
        payments_data.append([
            mandi_name[:12],
            f"{target_qntl}",
            f"{cutting_qntl}",
            f"Rs.{base_rate}+Rs.{cutting_rate}",
            f"Rs.{total_amount}",
            f"{tp_weight_qntl}",
            f"{achieved_qntl}",
            f"{excess_weight}",
            f"Rs.{paid_amount}",
            f"Rs.{balance}",
            status
        ])
    
    buffer = io.BytesIO()
    page_width, page_height = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    elements = []
    styles = get_pdf_styles()
    
    from utils.export_helpers import get_pdf_table_style
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.white, alignment=TA_CENTER)
    
    company_name, tagline = await get_company_name()
    title_data = [[Paragraph(f"<b>AGENT/MANDI PAYMENTS - {company_name} | FY: {kms_year or 'All'} | {season or 'All'}</b>", title_style)]]
    title_table = Table(title_data, colWidths=[page_width - 20*mm])
    title_table.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), 
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1B4F72')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(title_table)
    elements.append(Table([[""]], colWidths=[page_width], rowHeights=[5*mm]))
    
    headers = ["Mandi", "Target", "Cutting", "Rates", "Total Amt", "TP Wt", "Achieved", "Excess", "Paid", "Balance", "Status"]
    table_data = [headers] + payments_data
    table_data.append(["TOTAL", "", "", "", f"Rs.{round(total_amount_sum, 2)}", "", "", "", f"Rs.{round(total_paid_sum, 2)}", f"Rs.{round(total_balance_sum, 2)}", ""])
    
    col_widths = [26*mm, 18*mm, 16*mm, 24*mm, 22*mm, 18*mm, 18*mm, 18*mm, 20*mm, 20*mm, 16*mm]
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    cols_info = [{'header': h} for h in headers]
    style_commands = get_pdf_table_style(len(table_data), cols_info)
    style_commands.append(('ALIGN', (1, 1), (-1, -1), 'RIGHT'))
    style_commands.append(('ALIGN', (0, 0), (0, -1), 'LEFT'))
    style_commands.append(('TOPPADDING', (0, 0), (-1, -1), 4))
    style_commands.append(('BOTTOMPADDING', (0, 0), (-1, -1), 4))
    
    main_table.setStyle(TableStyle(style_commands))
    elements.append(main_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"agent_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


