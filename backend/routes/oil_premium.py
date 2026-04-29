from fastapi import APIRouter, HTTPException
from datetime import datetime, timezone
from database import db
import uuid

router = APIRouter()

STANDARD_OIL = {"Raw": 22, "Boiled": 25}


async def _sync_oil_premium_ledger(op: dict, username: str = "system"):
    """Sync local_party_accounts entry for oil-premium adjustment.

    - premium_amount > 0  → Party owes us MORE → debit entry
    - premium_amount < 0  → Party owes us LESS → payment entry (reduces balance)
    - premium_amount == 0 → remove any existing entry

    Reference: `oil_premium:{op_id}` (one entry per oil_premium record, idempotent)
    """
    op_id = op.get("id")
    if not op_id:
        return
    ref = f"oil_premium:{op_id}"
    party = (op.get("party_name") or "").strip()
    premium = float(op.get("premium_amount") or 0)
    voucher_no = op.get("voucher_no") or op.get("rst_no") or ""

    # Always remove existing first (idempotent), then add fresh if non-zero
    await db.local_party_accounts.delete_many({"reference": ref})

    if not party or premium == 0:
        return

    now_iso = datetime.now(timezone.utc).isoformat()
    if premium > 0:
        # Bonus due to better quality → party owes more
        txn_type = "debit"
        desc = f"Lab Test Bonus (+{op.get('difference_pct', 0)}%) - Voucher #{voucher_no}"
        amount = round(premium, 2)
    else:
        # Penalty due to lower quality → party owes less (we credit them)
        txn_type = "payment"
        desc = f"Lab Test Penalty ({op.get('difference_pct', 0)}%) - Voucher #{voucher_no}"
        amount = round(abs(premium), 2)

    await db.local_party_accounts.insert_one({
        "id": str(uuid.uuid4()),
        "date": op.get("date") or now_iso.split("T")[0],
        "party_name": f"{party} (Ka)",
        "txn_type": txn_type,
        "amount": amount,
        "description": desc,
        "source_type": "bp_sale_ka_oil_premium",
        "reference": ref,
        "kms_year": op.get("kms_year", ""),
        "season": op.get("season", ""),
        "created_by": username or "system",
        "created_at": now_iso,
    })


@router.get("/oil-premium")
async def get_oil_premiums(kms_year: str = "", season: str = "", bran_type: str = ""):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if bran_type: query["bran_type"] = bran_type
    items = await db.oil_premium.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    return items


@router.post("/oil-premium")
async def create_oil_premium(data: dict, username: str = "", role: str = ""):
    data["id"] = str(uuid.uuid4())[:12]
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_at"] = data["created_at"]
    data["created_by"] = username

    bran_type = data.get("bran_type", "Boiled")
    standard = STANDARD_OIL.get(bran_type, 25)
    actual = float(data.get("actual_oil_pct", 0) or 0)
    rate = float(data.get("rate", 0) or 0)
    qty = float(data.get("qty_qtl", 0) or 0)

    data["standard_oil_pct"] = standard
    data["difference_pct"] = round(actual - standard, 4)
    data["premium_amount"] = round(rate * (actual - standard) * qty / standard, 2) if standard else 0

    await db.oil_premium.insert_one({**data})
    data.pop("_id", None)
    await _sync_oil_premium_ledger(data, username)
    return data


@router.put("/oil-premium/{item_id}")
async def update_oil_premium(item_id: str, data: dict, username: str = "", role: str = ""):
    existing = await db.oil_premium.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")

    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = username

    bran_type = data.get("bran_type", "Boiled")
    standard = STANDARD_OIL.get(bran_type, 25)
    actual = float(data.get("actual_oil_pct", 0) or 0)
    rate = float(data.get("rate", 0) or 0)
    qty = float(data.get("qty_qtl", 0) or 0)

    data["standard_oil_pct"] = standard
    data["difference_pct"] = round(actual - standard, 4)
    data["premium_amount"] = round(rate * (actual - standard) * qty / standard, 2) if standard else 0

    data.pop("id", None)
    data.pop("_id", None)
    await db.oil_premium.update_one({"id": item_id}, {"$set": data})
    # Re-sync ledger entry with fresh values
    fresh = await db.oil_premium.find_one({"id": item_id}, {"_id": 0})
    if fresh:
        await _sync_oil_premium_ledger(fresh, username)
    return {"success": True}


@router.delete("/oil-premium/{item_id}")
async def delete_oil_premium(item_id: str, username: str = "", role: str = ""):
    # Cleanup linked ledger entry first (idempotent — works even if missing)
    await db.local_party_accounts.delete_many({"reference": f"oil_premium:{item_id}"})
    result = await db.oil_premium.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}


@router.get("/oil-premium/lookup-sale")
async def lookup_sale(voucher_no: str = "", rst_no: str = "", kms_year: str = ""):
    """Lookup a Rice Bran sale by voucher_no or rst_no to auto-fill Oil Premium form."""
    if not voucher_no and not rst_no:
        raise HTTPException(status_code=400, detail="voucher_no or rst_no required")

    query = {"product": "Rice Bran"}
    if kms_year: query["kms_year"] = kms_year

    if voucher_no:
        query["voucher_no"] = voucher_no
    elif rst_no:
        query["rst_no"] = rst_no

    sale = await db.bp_sale_register.find_one(query, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    return sale


def fmt_date(d):
    if not d: return ""
    try:
        if "T" in str(d): d = str(d).split("T")[0]
        parts = str(d).split("-")
        if len(parts) == 3: return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except: pass
    return str(d)


@router.get("/oil-premium/export/excel")
async def export_oil_premium_excel(kms_year: str = "", season: str = "", bran_type: str = "",
    date_from: str = "", date_to: str = "", party_name: str = ""):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if bran_type: query["bran_type"] = bran_type
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    items = await db.oil_premium.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    branding = await db.branding.find_one({}, {"_id": 0}) or {}
    company = branding.get("company_name", "Rice Mill")
    address = branding.get("address", "")
    phone = branding.get("phone", "")

    wb = Workbook(); ws = wb.active
    ws.title = "Oil Premium"

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    total_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    title = "Oil Premium Register"
    if kms_year: title += f" - FY {kms_year}"
    if bran_type: title += f" ({bran_type})"

    # Dynamic columns
    has_voucher = any(i.get('voucher_no') for i in items)
    has_rst = any(i.get('rst_no') for i in items)
    has_remark = any(i.get('remark') for i in items)

    cols = [('S.No', 5, 'sno'), ('Date', 10, 'date')]
    if has_voucher: cols.append(('Voucher No', 10, 'voucher_no'))
    if has_rst: cols.append(('RST No', 8, 'rst_no'))
    cols.append(('Type', 8, 'bran_type'))
    cols.append(('Party Name', 18, 'party_name'))
    cols.append(('Rate', 10, 'rate'))
    cols.append(('Qty (Qtl)', 10, 'qty_qtl'))
    cols.append(('Standard %', 9, 'standard_oil_pct'))
    cols.append(('Actual %', 9, 'actual_oil_pct'))
    cols.append(('Diff %', 9, 'difference_pct'))
    cols.append(('Premium', 14, 'premium_amount'))
    if has_remark: cols.append(('Remark', 16, 'remark'))

    headers = [c[0] for c in cols]
    widths = [c[1] for c in cols]
    keys = [c[2] for c in cols]
    ncols = len(headers)

    last_col = get_column_letter(ncols)
    ws.merge_cells(f'A1:{last_col}1')
    c1 = ws.cell(row=1, column=1, value=company.upper())
    c1.font = Font(bold=True, size=14, color="1F4E79"); c1.alignment = Alignment(horizontal='center')
    if address:
        ws.merge_cells(f'A2:{last_col}2')
        c2 = ws.cell(row=2, column=1, value=f"{address}  |  {phone}")
        c2.font = Font(size=9, color="666666"); c2.alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A3:{last_col}3')
    c3 = ws.cell(row=3, column=1, value=title)
    c3.font = Font(bold=True, size=12, color="FFFFFF")
    c3.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    c3.alignment = Alignment(horizontal='center')

    row = 5
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = header_fill; cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = border

    t_qty = t_premium = 0
    for idx, item in enumerate(items):
        r = row + 1 + idx
        fill = alt_fill if idx % 2 == 0 else None
        t_qty += item.get('qty_qtl', 0)
        t_premium += item.get('premium_amount', 0)
        for col_idx, key in enumerate(keys, 1):
            if key == 'sno': val = idx + 1
            elif key == 'date': val = fmt_date(item.get('date', ''))
            elif key == 'difference_pct': val = round(item.get(key, 0), 2)
            elif key == 'premium_amount': val = round(item.get(key, 0), 2)
            elif key == 'qty_qtl': val = round(item.get(key, 0), 2)
            elif key in ('rate', 'standard_oil_pct', 'actual_oil_pct'): val = item.get(key, 0)
            else: val = item.get(key, '')
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = Font(size=9); cell.border = border
            if fill: cell.fill = fill
            if key in ('rate', 'qty_qtl', 'standard_oil_pct', 'actual_oil_pct', 'difference_pct', 'premium_amount'):
                cell.alignment = Alignment(horizontal='right')
            if key == 'premium_amount':
                cell.number_format = '#,##0.00'
                if (item.get(key, 0) or 0) < 0:
                    cell.font = Font(size=9, color="FF0000")

    tr = row + 1 + len(items)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=tr, column=col_idx)
        cell.border = border; cell.fill = total_fill; cell.font = Font(bold=True, size=9, color="FFFFFF")
    for col_idx, key in enumerate(keys, 1):
        if key == 'date': ws.cell(row=tr, column=col_idx, value="TOTAL")
        elif key == 'qty_qtl':
            c = ws.cell(row=tr, column=col_idx, value=round(t_qty, 2))
            c.alignment = Alignment(horizontal='right')
        elif key == 'premium_amount':
            c = ws.cell(row=tr, column=col_idx, value=round(t_premium, 2))
            c.alignment = Alignment(horizontal='right')
            c.number_format = '#,##0.00'

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'landscape'; ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    fn = f"oil_premium_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/oil-premium/export/pdf")
async def export_oil_premium_pdf(kms_year: str = "", season: str = "", bran_type: str = "",
    date_from: str = "", date_to: str = "", party_name: str = ""):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from fastapi.responses import Response

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if bran_type: query["bran_type"] = bran_type
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    items = await db.oil_premium.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    branding = await db.branding.find_one({}, {"_id": 0}) or {}
    company = branding.get("company_name", "Rice Mill")
    address = branding.get("address", "")
    phone = branding.get("phone", "")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=15, rightMargin=15, topMargin=15, bottomMargin=15)
    elements = []
    styles = getSampleStyleSheet()

    # Use shared branded header (company name + address + phone + custom_fields like proprietor, GST, etc.)
    from utils.export_helpers import get_pdf_company_header
    elements.extend(get_pdf_company_header(branding))

    title = "Oil Premium Register"
    if kms_year: title += f" - FY {kms_year}"
    if bran_type: title += f" ({bran_type})"
    title_style = ParagraphStyle('RegTitle', parent=styles['Heading2'], fontSize=9,
        textColor=colors.white, backColor=colors.HexColor('#2E75B6'), spaceAfter=4, alignment=1,
        borderPadding=(2, 2, 2, 2))
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 3))

    # Dynamic columns
    has_voucher = any(i.get('voucher_no') for i in items)
    has_rst = any(i.get('rst_no') for i in items)

    pdf_cols = [('S.No', 22, 'sno'), ('Date', 45, 'date')]
    if has_voucher: pdf_cols.append(('Voucher', 45, 'voucher_no'))
    if has_rst: pdf_cols.append(('RST', 30, 'rst_no'))
    pdf_cols.append(('Type', 35, 'bran_type'))
    pdf_cols.append(('Party', 80, 'party_name'))
    pdf_cols.append(('Rate', 42, 'rate'))
    pdf_cols.append(('Qty(Q)', 40, 'qty_qtl'))
    pdf_cols.append(('Std%', 30, 'standard_oil_pct'))
    pdf_cols.append(('Actual%', 38, 'actual_oil_pct'))
    pdf_cols.append(('Diff%', 35, 'difference_pct'))
    pdf_cols.append(('Premium', 60, 'premium_amount'))

    pdf_headers = [c[0] for c in pdf_cols]
    col_widths = [c[1] for c in pdf_cols]
    col_keys = [c[2] for c in pdf_cols]

    # Auto-fit to A4 landscape
    usable_width = 812
    total_w = sum(col_widths)
    if total_w > usable_width:
        scale = usable_width / total_w
        col_widths = [round(w * scale) for w in col_widths]

    data = [pdf_headers]
    t_qty = t_premium = 0
    for idx, item in enumerate(items):
        t_qty += item.get('qty_qtl', 0)
        t_premium += item.get('premium_amount', 0)
        row_data = []
        for key in col_keys:
            if key == 'sno': row_data.append(idx + 1)
            elif key == 'date': row_data.append(fmt_date(item.get('date', '')))
            elif key == 'party_name': row_data.append((item.get('party_name', '') or '')[:18])
            elif key == 'difference_pct':
                d = item.get(key, 0)
                row_data.append(f"{'+' if d > 0 else ''}{d:.2f}%")
            elif key == 'premium_amount': row_data.append(f"{item.get(key, 0):,.0f}")
            elif key == 'qty_qtl': row_data.append(f"{item.get(key, 0):.2f}")
            else: row_data.append(item.get(key, '') or '')
        data.append(row_data)

    total_row = []
    for key in col_keys:
        if key == 'date': total_row.append('TOTAL')
        elif key == 'qty_qtl': total_row.append(f"{t_qty:.2f}")
        elif key == 'premium_amount': total_row.append(f"{t_premium:,.0f}")
        else: total_row.append('')
    data.append(total_row)

    table = RLTable(data, colWidths=col_widths, repeatRows=1)
    first_num = next((i for i, k in enumerate(col_keys) if k in ('rate', 'qty_qtl', 'standard_oil_pct', 'actual_oil_pct', 'difference_pct', 'premium_amount')), len(col_keys))
    nrows = len(data)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (first_num, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EBF1F8')]),
        ('BACKGROUND', (0, nrows - 1), (-1, nrows - 1), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, nrows - 1), (-1, nrows - 1), colors.white),
        ('FONTNAME', (0, nrows - 1), (-1, nrows - 1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    # Summary
    pos = sum(1 for i in items if (i.get('premium_amount', 0) or 0) > 0)
    neg = sum(1 for i in items if (i.get('premium_amount', 0) or 0) < 0)
    pos_amt = sum(i.get('premium_amount', 0) for i in items if (i.get('premium_amount', 0) or 0) > 0)
    neg_amt = sum(i.get('premium_amount', 0) for i in items if (i.get('premium_amount', 0) or 0) < 0)
    elements.append(Spacer(1, 8))
    summary_parts = []
    if pos: summary_parts.append(f"Premium ({pos}): <font color='green'>{pos_amt:,.0f}</font>")
    if neg: summary_parts.append(f"Deduction ({neg}): <font color='red'>{neg_amt:,.0f}</font>")
    summary_parts.append(f"<b>Net: <font color='{'green' if t_premium >= 0 else 'red'}'>{t_premium:,.0f}</font></b>")
    pay_style = ParagraphStyle('PaySummary', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#1F4E79'))
    elements.append(Paragraph(f"<b>Summary:</b>  {'  |  '.join(summary_parts)}", pay_style))

    elements.append(Spacer(1, 4))
    gen_style = ParagraphStyle('Gen', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#999999'))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", gen_style))

    doc.build(elements)
    buffer.seek(0)
    fn = f"oil_premium_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"})
