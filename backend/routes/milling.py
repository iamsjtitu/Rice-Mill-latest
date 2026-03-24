from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import *
import uuid
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

# ============ MILLING ENTRY CRUD APIs ============

def calculate_milling_fields(data: dict) -> dict:
    """Auto-calculate QNTL values from percentages and paddy input. FRK from stock."""
    paddy = data.get('paddy_input_qntl', 0) or 0
    
    rice_pct = data.get('rice_percent', 0) or 0
    bran_pct = data.get('bran_percent', 0) or 0
    kunda_pct = data.get('kunda_percent', 0) or 0
    broken_pct = data.get('broken_percent', 0) or 0
    kanki_pct = data.get('kanki_percent', 0) or 0
    
    used_pct = rice_pct + bran_pct + kunda_pct + broken_pct + kanki_pct
    husk_pct = max(0, round(100 - used_pct, 2))
    
    data['husk_percent'] = husk_pct
    data['rice_qntl'] = round(paddy * rice_pct / 100, 2)
    data['bran_qntl'] = round(paddy * bran_pct / 100, 2)
    data['kunda_qntl'] = round(paddy * kunda_pct / 100, 2)
    data['broken_qntl'] = round(paddy * broken_pct / 100, 2)
    data['kanki_qntl'] = round(paddy * kanki_pct / 100, 2)
    data['husk_qntl'] = round(paddy * husk_pct / 100, 2)
    
    frk_used = data.get('frk_used_qntl', 0) or 0
    data['cmr_delivery_qntl'] = round(data['rice_qntl'] + frk_used, 2)
    data['outturn_ratio'] = round(data['cmr_delivery_qntl'] / paddy * 100, 2) if paddy > 0 else 0
    
    return data


@router.post("/milling-entries")
async def create_milling_entry(input: MillingEntryCreate, username: str = "", role: str = ""):
    entry_dict = input.model_dump()
    entry_dict = calculate_milling_fields(entry_dict)
    entry_dict['created_by'] = username
    entry_obj = MillingEntry(**entry_dict)
    doc = entry_obj.model_dump()
    await db.milling_entries.insert_one(doc)
    doc.pop('_id', None)
    return doc


@router.get("/milling-entries")
async def get_milling_entries(
    rice_type: Optional[str] = None, kms_year: Optional[str] = None,
    season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None
):
    query = {}
    if rice_type: query["rice_type"] = rice_type
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to: dq["$lte"] = date_to
        if dq: query["date"] = dq
    return await db.milling_entries.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.get("/milling-entries/{entry_id}")
async def get_milling_entry(entry_id: str):
    entry = await db.milling_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry: raise HTTPException(status_code=404, detail="Milling entry not found")
    return entry


@router.put("/milling-entries/{entry_id}")
async def update_milling_entry(entry_id: str, input: MillingEntryCreate, username: str = "", role: str = ""):
    existing = await db.milling_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Milling entry not found")
    update_dict = input.model_dump()
    update_dict = calculate_milling_fields(update_dict)
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.milling_entries.update_one({"id": entry_id}, {"$set": update_dict})
    return await db.milling_entries.find_one({"id": entry_id}, {"_id": 0})


@router.delete("/milling-entries/{entry_id}")
async def delete_milling_entry(entry_id: str, username: str = "", role: str = ""):
    existing = await db.milling_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Milling entry not found")
    await db.milling_entries.delete_one({"id": entry_id})
    return {"message": "Milling entry deleted", "id": entry_id}


@router.get("/paddy-stock")
async def get_paddy_stock(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    # CMR paddy: QNTL - BAG - P.Cut
    mill_entries = await db.mill_entries.find(query, {"qntl": 1, "bag": 1, "p_pkt_cut": 1, "_id": 0}).to_list(10000)
    cmr_paddy_in = round(sum(e.get('qntl', 0) - e.get('bag', 0) / 100 - e.get('p_pkt_cut', 0) / 100 for e in mill_entries), 2)
    # Private paddy purchases (NOT agent_extra - those are already counted in CMR)
    pvt_query = dict(query)
    pvt_query["source"] = {"$ne": "agent_extra"}
    pvt_entries = await db.private_paddy.find(pvt_query, {"qntl": 1, "bag": 1, "_id": 0}).to_list(10000)
    pvt_paddy_in = round(sum(e.get('qntl', 0) - e.get('bag', 0) / 100 for e in pvt_entries), 2)
    # Purchase Voucher paddy
    purchase_vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).to_list(10000)
    pv_paddy = 0
    for pv in purchase_vouchers:
        for item in pv.get('items', []):
            if item.get('item_name', '') == 'Paddy':
                pv_paddy += item.get('quantity', 0) or 0
    pv_paddy = round(pv_paddy, 2)
    total_paddy_in = round(cmr_paddy_in + pvt_paddy_in + pv_paddy, 2)
    milling_entries = await db.milling_entries.find(query, {"paddy_input_qntl": 1, "_id": 0}).to_list(10000)
    total_paddy_used = round(sum(e.get('paddy_input_qntl', 0) for e in milling_entries), 2)
    return {"total_paddy_in_qntl": total_paddy_in, "total_paddy_used_qntl": total_paddy_used,
        "available_paddy_qntl": round(total_paddy_in - total_paddy_used, 2),
        "cmr_paddy_in_qntl": cmr_paddy_in, "pvt_paddy_in_qntl": pvt_paddy_in,
        "pv_paddy_in_qntl": pv_paddy}


@router.get("/milling-summary")
async def get_milling_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.milling_entries.find(query, {"_id": 0}).to_list(1000)
    
    total_paddy = sum(e.get('paddy_input_qntl', 0) for e in entries)
    total_rice = sum(e.get('rice_qntl', 0) for e in entries)
    total_frk = sum(e.get('frk_used_qntl', 0) for e in entries)
    total_bran = sum(e.get('bran_qntl', 0) for e in entries)
    total_kunda = sum(e.get('kunda_qntl', 0) for e in entries)
    total_broken = sum(e.get('broken_qntl', 0) for e in entries)
    total_kanki = sum(e.get('kanki_qntl', 0) for e in entries)
    total_husk = sum(e.get('husk_qntl', 0) for e in entries)
    total_cmr = sum(e.get('cmr_delivery_qntl', 0) for e in entries)
    avg_outturn = round(total_cmr / total_paddy * 100, 2) if total_paddy > 0 else 0
    
    def type_summary(elist):
        tp = sum(e.get('paddy_input_qntl', 0) for e in elist)
        tr = sum(e.get('rice_qntl', 0) for e in elist)
        tf = sum(e.get('frk_used_qntl', 0) for e in elist)
        tc = sum(e.get('cmr_delivery_qntl', 0) for e in elist)
        return {"count": len(elist), "total_paddy_qntl": round(tp, 2), "total_rice_qntl": round(tr, 2),
            "total_frk_qntl": round(tf, 2), "total_cmr_qntl": round(tc, 2),
            "avg_outturn": round(tc / tp * 100, 2) if tp > 0 else 0}
    
    return {"total_entries": len(entries), "total_paddy_qntl": round(total_paddy, 2),
        "total_rice_qntl": round(total_rice, 2), "total_frk_qntl": round(total_frk, 2),
        "total_bran_qntl": round(total_bran, 2), "total_kunda_qntl": round(total_kunda, 2),
        "total_broken_qntl": round(total_broken, 2), "total_kanki_qntl": round(total_kanki, 2),
        "total_husk_qntl": round(total_husk, 2), "total_cmr_qntl": round(total_cmr, 2),
        "avg_outturn_ratio": avg_outturn,
        "parboiled": type_summary([e for e in entries if e.get('rice_type') == 'parboiled']),
        "raw": type_summary([e for e in entries if e.get('rice_type') == 'raw'])}


# ============ FRK PURCHASE APIs ============

@router.post("/frk-purchases")
async def create_frk_purchase(input: FrkPurchaseCreate, username: str = "", role: str = ""):
    d = input.model_dump()
    d['total_amount'] = round((d.get('quantity_qntl', 0) or 0) * (d.get('rate_per_qntl', 0) or 0), 2)
    d['created_by'] = username
    obj = FrkPurchase(**d)
    doc = obj.model_dump()
    await db.frk_purchases.insert_one(doc)
    doc.pop('_id', None)
    return doc


@router.get("/frk-purchases")
async def get_frk_purchases(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    return await db.frk_purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.delete("/frk-purchases/{purchase_id}")
async def delete_frk_purchase(purchase_id: str, username: str = "", role: str = ""):
    existing = await db.frk_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="FRK purchase not found")
    await db.frk_purchases.delete_one({"id": purchase_id})
    return {"message": "FRK purchase deleted", "id": purchase_id}


@router.get("/frk-stock")
async def get_frk_stock(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(1000)
    total_purchased = round(sum(p.get('quantity_qntl', 0) for p in purchases), 2)
    total_cost = round(sum(p.get('total_amount', 0) for p in purchases), 2)
    milling_entries = await db.milling_entries.find(query, {"frk_used_qntl": 1, "_id": 0}).to_list(1000)
    total_used = round(sum(e.get('frk_used_qntl', 0) for e in milling_entries), 2)
    return {"total_purchased_qntl": total_purchased, "total_used_qntl": total_used,
        "available_qntl": round(total_purchased - total_used, 2), "total_cost": total_cost}



@router.get("/rice-stock")
async def get_rice_stock(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Rice Stock = Milling se produced + Purchase Voucher se kharida - Govt ko diya (DC) - Pvt mein becha - Sale Book mein becha"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    
    # Rice produced from milling
    milling_entries = await db.milling_entries.find(query, {"_id": 0, "rice_qntl": 1, "rice_type": 1}).to_list(10000)
    total_produced = round(sum(e.get('rice_qntl', 0) for e in milling_entries), 2)
    parboiled_produced = round(sum(e.get('rice_qntl', 0) for e in milling_entries if e.get('rice_type') == 'parboiled'), 2)
    raw_produced = round(sum(e.get('rice_qntl', 0) for e in milling_entries if e.get('rice_type') == 'raw'), 2)
    
    # Rice purchased from Purchase Vouchers
    purchase_vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).to_list(10000)
    pv_bought = {}
    for pv in purchase_vouchers:
        for item in pv.get('items', []):
            name = item.get('item_name', '')
            pv_bought[name] = pv_bought.get(name, 0) + (item.get('quantity', 0) or 0)
    
    pv_rice_usna = round(pv_bought.get("Rice (Usna)", 0), 2)
    pv_rice_raw = round(pv_bought.get("Rice (Raw)", 0), 2)
    pv_rice_total = round(pv_rice_usna + pv_rice_raw, 2)
    
    # Rice delivered to Govt (DC deliveries) - type-wise
    dc_query = {}
    if kms_year: dc_query["kms_year"] = kms_year
    if season: dc_query["season"] = season
    dc_deliveries = await db.dc_deliveries.find(dc_query, {"_id": 0, "quantity_qntl": 1, "dc_id": 1}).to_list(10000)
    # Get DC rice_type mapping
    dc_entries = await db.dc_entries.find({}, {"_id": 0, "id": 1, "rice_type": 1}).to_list(10000)
    dc_type_map = {d["id"]: d.get("rice_type", "") for d in dc_entries}
    total_govt_delivered = round(sum(d.get('quantity_qntl', 0) for d in dc_deliveries), 2)
    parboiled_govt = round(sum(d.get('quantity_qntl', 0) for d in dc_deliveries if dc_type_map.get(d.get('dc_id')) == 'parboiled'), 2)
    raw_govt = round(sum(d.get('quantity_qntl', 0) for d in dc_deliveries if dc_type_map.get(d.get('dc_id')) == 'raw'), 2)
    
    # Rice sold privately - type-wise
    pvt_sales = await db.rice_sales.find(query, {"_id": 0, "quantity_qntl": 1, "rice_type": 1}).to_list(10000)
    total_pvt_sold = round(sum(s.get('quantity_qntl', 0) for s in pvt_sales), 2)
    parboiled_pvt = round(sum(s.get('quantity_qntl', 0) for s in pvt_sales if s.get('rice_type') in ('Usna', 'parboiled')), 2)
    raw_pvt = round(sum(s.get('quantity_qntl', 0) for s in pvt_sales if s.get('rice_type') in ('Raw', 'raw')), 2)
    
    # Sale Book sales
    sale_vouchers = await db.sale_vouchers.find(query, {"_id": 0}).to_list(10000)
    sb_rice_usna = 0
    sb_rice_raw = 0
    for sv in sale_vouchers:
        for item in sv.get('items', []):
            name = item.get('item_name', '')
            qty = item.get('quantity', 0) or 0
            if name == "Rice (Usna)": sb_rice_usna += qty
            elif name == "Rice (Raw)": sb_rice_raw += qty
    sb_rice_total = round(sb_rice_usna + sb_rice_raw, 2)
    
    available = round(total_produced + pv_rice_total - total_govt_delivered - total_pvt_sold - sb_rice_total, 2)
    parboiled_available = round(parboiled_produced + pv_rice_usna - parboiled_govt - parboiled_pvt - sb_rice_usna, 2)
    raw_available = round(raw_produced + pv_rice_raw - raw_govt - raw_pvt - sb_rice_raw, 2)
    
    return {
        "total_produced_qntl": total_produced,
        "parboiled_produced_qntl": parboiled_produced,
        "raw_produced_qntl": raw_produced,
        "purchased_qntl": pv_rice_total,
        "purchased_usna_qntl": pv_rice_usna,
        "purchased_raw_qntl": pv_rice_raw,
        "govt_delivered_qntl": total_govt_delivered,
        "pvt_sold_qntl": total_pvt_sold,
        "sb_sold_qntl": sb_rice_total,
        "available_qntl": available,
        "parboiled_available_qntl": parboiled_available,
        "raw_available_qntl": raw_available,
        "milling_count": len(milling_entries),
        "dc_delivery_count": len(dc_deliveries),
        "pvt_sale_count": len(pvt_sales)
    }



# ============ BY-PRODUCT STOCK & SALE APIs ============

@router.get("/byproduct-stock")
async def get_byproduct_stock(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    milling_entries = await db.milling_entries.find(query, {"_id": 0}).to_list(1000)
    sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(1000)
    # Also count Sale Book sales for by-products
    sale_vouchers = await db.sale_vouchers.find(query, {"_id": 0}).to_list(10000)
    sb_sold = {}
    for sv in sale_vouchers:
        for item in sv.get('items', []):
            name = item.get('item_name', '').lower()
            sb_sold[name] = sb_sold.get(name, 0) + (item.get('quantity', 0) or 0)
    # Purchase Voucher bought quantities
    purchase_vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).to_list(10000)
    pv_bought = {}
    for pv in purchase_vouchers:
        for item in pv.get('items', []):
            name = item.get('item_name', '').lower()
            pv_bought[name] = pv_bought.get(name, 0) + (item.get('quantity', 0) or 0)
    products = ["bran", "kunda", "broken", "kanki", "husk"]
    stock = {}
    for p in products:
        produced = round(sum(e.get(f'{p}_qntl', 0) for e in milling_entries), 2)
        purchased = round(pv_bought.get(p, 0), 2)
        sold = round(sum(s.get('quantity_qntl', 0) for s in sales if s.get('product') == p), 2)
        sold_sb = round(sb_sold.get(p, 0), 2)
        revenue = round(sum(s.get('total_amount', 0) for s in sales if s.get('product') == p), 2)
        stock[p] = {"produced_qntl": produced, "purchased_qntl": purchased, "sold_qntl": round(sold + sold_sb, 2), "available_qntl": round(produced + purchased - sold - sold_sb, 2), "total_revenue": revenue}
    return stock


@router.post("/byproduct-sales")
async def create_byproduct_sale(input: ByProductSaleCreate, username: str = "", role: str = ""):
    d = input.model_dump()
    d['total_amount'] = round((d.get('quantity_qntl', 0) or 0) * (d.get('rate_per_qntl', 0) or 0), 2)
    d['created_by'] = username
    obj = ByProductSale(**d)
    doc = obj.model_dump()
    await db.byproduct_sales.insert_one(doc)
    doc.pop('_id', None)
    
    # Auto-create party_ledger entry if buyer_name is provided
    buyer = (d.get('buyer_name') or '').strip()
    if buyer and d['total_amount'] > 0:
        import uuid as _uuid
        from datetime import datetime as _dt, timezone as _tz
        ledger_entry = {
            "id": str(_uuid.uuid4()),
            "date": d.get('date', ''),
            "account": "ledger",
            "txn_type": "jama",
            "amount": d['total_amount'],
            "category": buyer,
            "party_type": "By-Product Sale",
            "description": f"{d.get('product','').title()} sale - {d.get('quantity_qntl',0)} Qntl @ Rs.{d.get('rate_per_qntl',0)}/Q",
            "reference": f"byproduct:{doc.get('id','')}",
            "kms_year": d.get('kms_year', ''),
            "season": d.get('season', ''),
            "created_by": username,
            "created_at": _dt.now(_tz.utc).isoformat(),
            "updated_at": _dt.now(_tz.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(ledger_entry)
        ledger_entry.pop('_id', None)
    
    return doc


@router.get("/byproduct-sales")
async def get_byproduct_sales(product: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if product: query["product"] = product
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    return await db.byproduct_sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.delete("/byproduct-sales/{sale_id}")
async def delete_byproduct_sale(sale_id: str, username: str = "", role: str = ""):
    existing = await db.byproduct_sales.find_one({"id": sale_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Sale entry not found")
    await db.byproduct_sales.delete_one({"id": sale_id})
    # Also delete linked party_ledger entry
    await db.cash_transactions.delete_many({"reference": f"byproduct:{sale_id}"})
    return {"message": "Sale entry deleted", "id": sale_id}


# ============ PADDY CUSTODY MAINTENANCE REGISTER ============

@router.get("/paddy-custody-register")
async def get_paddy_custody_register(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Paddy custody register - all movements: received (mill entries) and released (milling entries)"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    
    # Paddy received from mill entries
    mill_entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    # Paddy released for milling
    milling_entries = await db.milling_entries.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    
    # Build register rows
    rows = []
    for e in mill_entries:
        rows.append({
            "date": e.get('date', ''),
            "type": "received",
            "description": f"Truck: {e.get('truck_no', '')} | Agent: {e.get('agent_name', '')} | Mandi: {e.get('mandi_name', '')}",
            "received_qntl": round(e.get('qntl', 0) - e.get('bag', 0) / 100, 2),
            "issued_qntl": 0,
            "source_id": e.get('id', '')
        })
    for e in milling_entries:
        rows.append({
            "date": e.get('date', ''),
            "type": "issued",
            "description": f"Milling ({e.get('rice_type', 'parboiled').title()}) | Rice: {e.get('rice_qntl', 0)}Q",
            "received_qntl": 0,
            "issued_qntl": e.get('paddy_input_qntl', 0),
            "source_id": e.get('id', '')
        })
    
    # Sort by date
    rows.sort(key=lambda x: x['date'])
    
    # Add running balance
    balance = 0
    for r in rows:
        balance += r['received_qntl'] - r['issued_qntl']
        r['balance_qntl'] = round(balance, 2)
    
    return {"rows": rows, "total_received": round(sum(r['received_qntl'] for r in rows), 2),
        "total_issued": round(sum(r['issued_qntl'] for r in rows), 2),
        "final_balance": round(balance, 2)}


# ============ MILLING REPORT EXPORT ============

@router.get("/milling-report/excel")
async def export_milling_report_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS)
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.milling_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    wb = Workbook(); ws = wb.active; ws.title = "Milling Report"
    ncols = 12
    title = "Milling Report / मिलिंग रिपोर्ट"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    style_excel_title(ws, title, ncols, "Mill Entry System")
    
    headers = ['Date', 'Type', 'Paddy (Q)', 'Rice %', 'Rice (Q)', 'FRK Used (Q)', 'CMR (Q)', 'Outturn %', 'Bran (Q)', 'Kunda (Q)', 'Husk %', 'Note']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_excel_header_row(ws, 4, ncols)
    
    data_start = 5
    for idx, e in enumerate(entries):
        row = idx + data_start
        vals = [e.get('date',''), e.get('rice_type','').title(), e.get('paddy_input_qntl',0), e.get('rice_percent',0),
            e.get('rice_qntl',0), e.get('frk_used_qntl',0), e.get('cmr_delivery_qntl',0), e.get('outturn_ratio',0),
            e.get('bran_qntl',0), e.get('kunda_qntl',0), e.get('husk_percent',0), e.get('note','')]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
            if col >= 3: ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    
    if entries:
        style_excel_data_rows(ws, data_start, data_start + len(entries) - 1, ncols, headers)
    
    tr = data_start + len(entries)
    ws.cell(row=tr, column=1, value="TOTAL")
    if entries:
        for col, key in [(3,'paddy_input_qntl'),(5,'rice_qntl'),(6,'frk_used_qntl'),(7,'cmr_delivery_qntl'),(9,'bran_qntl'),(10,'kunda_qntl')]:
            ws.cell(row=tr, column=col, value=round(sum(e.get(key,0) for e in entries),2))
    style_excel_total_row(ws, tr, ncols)
    
    for i in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    fn = f"milling_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/milling-report/pdf")
async def export_milling_report_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.milling_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    title = "Milling Report"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    
    headers = ['Date', 'Type', 'Paddy(Q)', 'Rice%', 'Rice(Q)', 'FRK(Q)', 'CMR(Q)', 'Outturn%', 'Bran(Q)', 'Kunda(Q)', 'Husk%']
    data = [headers]
    tp = tr = tf = tc = tb = tk = 0
    for e in entries:
        tp += e.get('paddy_input_qntl',0); tr += e.get('rice_qntl',0); tf += e.get('frk_used_qntl',0)
        tc += e.get('cmr_delivery_qntl',0); tb += e.get('bran_qntl',0); tk += e.get('kunda_qntl',0)
        data.append([e.get('date',''), e.get('rice_type','').title()[:3], e.get('paddy_input_qntl',0),
            f"{e.get('rice_percent',0)}%", e.get('rice_qntl',0), e.get('frk_used_qntl',0),
            e.get('cmr_delivery_qntl',0), f"{e.get('outturn_ratio',0)}%", e.get('bran_qntl',0), e.get('kunda_qntl',0), f"{e.get('husk_percent',0)}%"])
    data.append(['TOTAL', '', round(tp,2), '', round(tr,2), round(tf,2), round(tc,2), '', round(tb,2), round(tk,2), ''])
    
    from utils.export_helpers import get_pdf_table_style
    
    col_widths = [65, 35, 55, 40, 50, 45, 50, 55, 45, 50, 40]
    table = RLTable(data, colWidths=col_widths, repeatRows=1)
    style_cmds = get_pdf_table_style(len(data))
    style_cmds.append(('ALIGN', (2,0), (-1,-1), 'RIGHT'))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    fn = f"milling_report_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/paddy-custody-register/excel")
async def export_paddy_custody_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from io import BytesIO
    
    register = await get_paddy_custody_register(kms_year=kms_year, season=season)
    rows = register['rows']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Paddy Custody Register"
    
    from openpyxl.utils import get_column_letter
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    title = "Paddy Custody Register / धान कस्टडी"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    ncols = 5
    style_excel_title(ws, title, ncols, "Mill Entry System")
    
    headers = ['Date', 'Description', 'Received (QNTL)', 'Released (QNTL)', 'Balance (QNTL)']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_excel_header_row(ws, 4, ncols)
    
    data_start = 5
    for idx, r in enumerate(rows):
        row_num = idx + data_start
        vals = [r['date'], r['description'], r['received_qntl'] if r['received_qntl'] > 0 else '',
            r['issued_qntl'] if r['issued_qntl'] > 0 else '', r['balance_qntl']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            if col >= 3: cell.alignment = Alignment(horizontal='right')
    
    if rows:
        style_excel_data_rows(ws, data_start, data_start + len(rows) - 1, ncols, headers)
    
    tr = data_start + len(rows)
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=3, value=register['total_received'])
    ws.cell(row=tr, column=4, value=register['total_issued'])
    ws.cell(row=tr, column=5, value=register['final_balance'])
    style_excel_total_row(ws, tr, ncols)
    
    ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 16; ws.column_dimensions['D'].width = 16; ws.column_dimensions['E'].width = 16
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fn = f"paddy_custody_register_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/paddy-custody-register/pdf")
async def export_paddy_custody_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    
    register = await get_paddy_custody_register(kms_year=kms_year, season=season)
    rows = register['rows']
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    title = "Paddy Custody Maintenance Register"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    
    from utils.export_helpers import get_pdf_table_style
    
    data = [['Date', 'Description', 'Received (Q)', 'Released (Q)', 'Balance (Q)']]
    for r in rows:
        data.append([r['date'], r['description'][:60], r['received_qntl'] if r['received_qntl'] > 0 else '-',
            r['issued_qntl'] if r['issued_qntl'] > 0 else '-', r['balance_qntl']])
    data.append(['TOTAL', '', register['total_received'], register['total_issued'], register['final_balance']])
    
    table = RLTable(data, colWidths=[65, 300, 70, 70, 70], repeatRows=1)
    style_cmds = get_pdf_table_style(len(data))
    style_cmds.append(('ALIGN', (2,0), (-1,-1), 'RIGHT'))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=paddy_custody_{datetime.now().strftime('%Y%m%d')}.pdf"})


@router.get("/frk-purchases/excel")
async def export_frk_purchases_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    purchases = await db.frk_purchases.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    wb = Workbook(); ws = wb.active; ws.title = "FRK Purchases"
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    ncols = 6
    title = "FRK Purchase Register / एफआरके खरीद"
    if kms_year: title += f" - KMS {kms_year}"
    style_excel_title(ws, title, ncols, "Mill Entry System")
    
    headers = ['Date', 'Party Name', 'Qty (QNTL)', 'Rate (Rs/Q)', 'Amount (Rs)', 'Note']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_excel_header_row(ws, 4, ncols)
    
    data_start = 5
    for idx, p in enumerate(purchases):
        row = idx + data_start
        for col, v in enumerate([p.get('date',''), p.get('party_name',''), p.get('quantity_qntl',0), p.get('rate_per_qntl',0), p.get('total_amount',0), p.get('note','')], 1):
            ws.cell(row=row, column=col, value=v)
            if col >= 3: ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
    
    if purchases:
        style_excel_data_rows(ws, data_start, data_start + len(purchases) - 1, ncols, headers)
    
    tr = data_start + len(purchases)
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=3, value=round(sum(p.get('quantity_qntl',0) for p in purchases),2))
    ws.cell(row=tr, column=5, value=round(sum(p.get('total_amount',0) for p in purchases),2))
    style_excel_total_row(ws, tr, ncols)
    for letter in ['A','B','C','D','E','F']: ws.column_dimensions[letter].width = 16
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=frk_purchases_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/frk-purchases/pdf")
async def export_frk_purchases_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    purchases = await db.frk_purchases.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    title = "FRK Purchase Register"
    if kms_year: title += f" - KMS {kms_year}"
    elements.append(Paragraph(title, styles['Title'])); elements.append(Spacer(1, 12))
    
    from utils.export_helpers import get_pdf_table_style
    
    data = [['Date', 'Party', 'Qty(Q)', 'Rate(Rs)', 'Amount(Rs)', 'Note']]
    tq = ta = 0
    for p in purchases:
        tq += p.get('quantity_qntl',0); ta += p.get('total_amount',0)
        data.append([p.get('date',''), p.get('party_name','')[:25], p.get('quantity_qntl',0), p.get('rate_per_qntl',0), p.get('total_amount',0), p.get('note','')[:20]])
    data.append(['TOTAL', '', round(tq,2), '', round(ta,2), ''])
    
    table = RLTable(data, colWidths=[60, 120, 55, 55, 70, 80], repeatRows=1)
    style_cmds = get_pdf_table_style(len(data))
    style_cmds.append(('ALIGN', (2,0), (-1,-1), 'RIGHT'))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=frk_purchases_{datetime.now().strftime('%Y%m%d')}.pdf"})


@router.get("/byproduct-sales/excel")
async def export_byproduct_sales_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    sales = await db.byproduct_sales.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    stock_data = await get_byproduct_stock(kms_year=kms_year, season=season)
    
    wb = Workbook(); ws = wb.active; ws.title = "By-Product Sales"
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    ncols = 7
    title = "By-Product Stock & Sales / उप-उत्पाद बिक्री"
    if kms_year: title += f" - KMS {kms_year}"
    style_excel_title(ws, title, ncols, "Mill Entry System")
    
    # Stock summary section
    ws.cell(row=4, column=1, value="Stock Summary").font = Font(bold=True, size=11, color=COLORS['title_text'])
    stock_headers = ['Product', 'Produced (Q)', 'Sold (Q)', 'Available (Q)', 'Revenue (Rs)']
    for col, h in enumerate(stock_headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_excel_header_row(ws, 5, 5)
    row = 6
    stock_start = row
    for prod, label in [('bran','Bran'), ('kunda','Kunda'), ('broken','Broken'), ('kanki','Kanki'), ('husk','Husk')]:
        s = stock_data.get(prod, {})
        for col, v in enumerate([label, s.get('produced_qntl',0), s.get('sold_qntl',0), s.get('available_qntl',0), s.get('total_revenue',0)], 1):
            ws.cell(row=row, column=col, value=v)
            if col >= 2: ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        row += 1
    style_excel_data_rows(ws, stock_start, row - 1, 5, stock_headers)
    
    # Sales detail section
    row += 1
    ws.cell(row=row, column=1, value="Sales Detail").font = Font(bold=True, size=11, color=COLORS['title_text'])
    row += 1
    sale_headers = ['Date', 'Product', 'Qty (Q)', 'Rate (Rs/Q)', 'Amount (Rs)', 'Buyer', 'Note']
    for col, h in enumerate(sale_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_excel_header_row(ws, row, ncols)
    row += 1
    sale_start = row
    for s in sales:
        for col, v in enumerate([s.get('date',''), s.get('product','').title(), s.get('quantity_qntl',0), s.get('rate_per_qntl',0), s.get('total_amount',0), s.get('buyer_name',''), s.get('note','')], 1):
            ws.cell(row=row, column=col, value=v)
            if col >= 3 and col <= 5: ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        row += 1
    if sales:
        style_excel_data_rows(ws, sale_start, row - 1, ncols, sale_headers)
    
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=3, value=round(sum(s.get('quantity_qntl',0) for s in sales),2))
    ws.cell(row=row, column=5, value=round(sum(s.get('total_amount',0) for s in sales),2))
    style_excel_total_row(ws, row, ncols)
    for letter in ['A','B','C','D','E','F','G']: ws.column_dimensions[letter].width = 16
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=byproduct_sales_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/byproduct-sales/pdf")
async def export_byproduct_sales_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    sales = await db.byproduct_sales.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    stock_data = await get_byproduct_stock(kms_year=kms_year, season=season)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    title = "By-Product Stock & Sales Report"
    if kms_year: title += f" - KMS {kms_year}"
    elements.append(Paragraph(title, styles['Title'])); elements.append(Spacer(1, 12))
    
    from utils.export_helpers import get_pdf_table_style
    
    # Stock summary table
    elements.append(Paragraph("Stock Summary", styles['Heading2'])); elements.append(Spacer(1, 6))
    sdata = [['Product', 'Produced(Q)', 'Sold(Q)', 'Available(Q)', 'Revenue(Rs)']]
    for prod, label in [('bran','Bran'), ('kunda','Kunda'), ('broken','Broken'), ('kanki','Kanki'), ('husk','Husk')]:
        s = stock_data.get(prod, {})
        sdata.append([label, s.get('produced_qntl',0), s.get('sold_qntl',0), s.get('available_qntl',0), s.get('total_revenue',0)])
    st = RLTable(sdata, colWidths=[70, 70, 60, 70, 70])
    style_cmds = get_pdf_table_style(len(sdata))
    style_cmds.append(('ALIGN', (1,0), (-1,-1), 'RIGHT'))
    st.setStyle(TableStyle(style_cmds))
    elements.append(st); elements.append(Spacer(1, 15))
    
    # Sales table
    elements.append(Paragraph("Sales Detail", styles['Heading2'])); elements.append(Spacer(1, 6))
    data = [['Date', 'Product', 'Qty(Q)', 'Rate(Rs)', 'Amount(Rs)', 'Buyer']]
    tq = ta = 0
    for s in sales:
        tq += s.get('quantity_qntl',0); ta += s.get('total_amount',0)
        data.append([s.get('date',''), s.get('product','').title(), s.get('quantity_qntl',0), s.get('rate_per_qntl',0), s.get('total_amount',0), s.get('buyer_name','')[:20]])
    data.append(['TOTAL', '', round(tq,2), '', round(ta,2), ''])
    
    table = RLTable(data, colWidths=[55, 55, 45, 50, 60, 90], repeatRows=1)
    style_cmds2 = get_pdf_table_style(len(data))
    style_cmds2.append(('ALIGN', (2,0), (-1,-1), 'RIGHT'))
    table.setStyle(TableStyle(style_cmds2))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=byproduct_sales_{datetime.now().strftime('%Y%m%d')}.pdf"})


