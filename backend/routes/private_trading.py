from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import *
import uuid
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.report_helper import get_columns, get_entry_row, get_total_row, get_excel_headers, get_pdf_headers, get_excel_widths, get_pdf_widths_mm, col_count


def _fmt_detail(qntl, rate):
    """Format qty @ rate with clean numbers (no trailing .0)"""
    q = int(qntl) if qntl == int(qntl) else qntl
    r = int(rate) if rate == int(rate) else round(rate, 2)
    return f"{q} Qntl @ Rs.{r}"

router = APIRouter()

# ============ PRIVATE TRADING: Paddy Purchase & Rice Sale ============

@router.post("/private-paddy")
async def create_private_paddy(data: dict, username: str = "", role: str = ""):
    def _f(v): return float(v) if v not in (None, "") else 0
    def _i(v): return int(float(v)) if v not in (None, "") else 0
    doc = {
        "id": str(uuid.uuid4()), "date": data.get("date", ""),
        "kms_year": data.get("kms_year", ""), "season": data.get("season", ""),
        "party_name": data.get("party_name", ""), "truck_no": data.get("truck_no", ""),
        "rst_no": data.get("rst_no", ""), "agent_name": data.get("agent_name", ""),
        "mandi_name": data.get("mandi_name", ""),
        "kg": _f(data.get("kg", 0)), "bag": _i(data.get("bag", 0)),
        "rate_per_qntl": _f(data.get("rate_per_qntl", 0)),
        "g_deposite": _f(data.get("g_deposite", 0)),
        "gbw_cut": _f(data.get("gbw_cut", 0)),
        "plastic_bag": _i(data.get("plastic_bag", 0)),
        "moisture": _f(data.get("moisture", 0)),
        "cutting_percent": _f(data.get("cutting_percent", 0)),
        "disc_dust_poll": _f(data.get("disc_dust_poll", 0)),
        "g_issued": _i(data.get("g_issued", 0)),
        "cash_paid": _f(data.get("cash_paid", 0)),
        "diesel_paid": _f(data.get("diesel_paid", 0)),
        "remark": data.get("remark", ""),
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    # Auto calculations
    doc["qntl"] = round(doc["kg"] / 100, 2) if doc["kg"] else 0
    doc["gbw_cut"] = _f(data.get("gbw_cut", 0)) or round(doc["bag"] * 1.0, 2)
    doc["mill_w"] = round(doc["kg"] - doc["gbw_cut"], 2)
    doc["p_pkt_cut"] = round(doc["plastic_bag"] * 0.5, 2)
    moist_pct = max(0, doc["moisture"] - 17) if doc["moisture"] > 17 else 0
    doc["moisture_cut_percent"] = round(moist_pct, 2)
    doc["moisture_cut"] = round(doc["mill_w"] * moist_pct / 100, 2)
    after_moisture = doc["mill_w"] - doc["moisture_cut"]
    doc["cutting"] = round(after_moisture * doc["cutting_percent"] / 100, 2)
    doc["final_w"] = round(after_moisture - doc["cutting"] - doc["p_pkt_cut"] - doc["disc_dust_poll"], 2)
    doc["final_qntl"] = round(doc["final_w"] / 100, 2)
    doc["total_amount"] = round(doc["final_qntl"] * doc["rate_per_qntl"], 2)
    doc["paid_amount"] = float(data.get("paid_amount", 0))
    doc["balance"] = round(doc["total_amount"] - doc["paid_amount"], 2)
    await db.private_paddy.insert_one(doc)
    doc.pop("_id", None)
    # Auto gunny bag entries
    await _create_gunny_entries_for_pvt_paddy(doc, username)
    # Auto cash book + diesel entries
    await _create_cashbook_diesel_for_pvt_paddy(doc, username)
    return doc


async def _create_gunny_entries_for_pvt_paddy(doc, username=""):
    """Auto-create gunny bag entries for bag (IN) and g_issued (OUT) in a pvt paddy entry."""
    entry_id = doc["id"]
    party = doc.get("party_name", "")
    agent = doc.get("agent_name", "")
    mandi = doc.get("mandi_name", "")
    source = f"Pvt: {party}" if party else "Pvt Paddy"
    if agent and mandi:
        source = f"Pvt: {agent} - {mandi}"
    truck = doc.get("truck_no", "")
    base = {
        "bag_type": "old", "rate": 0, "amount": 0, "notes": "Auto from Pvt Paddy",
        "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
        "created_by": username or "system", "linked_entry_id": entry_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    bag_in = int(float(doc.get("bag", 0) or 0))
    if bag_in > 0:
        entry = {**base, "id": str(uuid.uuid4()), "date": doc.get("date", ""),
                 "txn_type": "in", "quantity": bag_in, "source": source, "reference": truck}
        await db.gunny_bags.insert_one(entry)
    g_issued = int(float(doc.get("g_issued", 0) or 0))
    if g_issued > 0:
        entry = {**base, "id": str(uuid.uuid4()), "date": doc.get("date", ""),
                 "txn_type": "out", "quantity": g_issued, "source": source, "reference": truck}
        await db.gunny_bags.insert_one(entry)


async def _create_cashbook_diesel_for_pvt_paddy(doc, username=""):
    """Auto-create truck jama, truck cash/diesel nikasi, party jama (paddy purchase debt), and advance nikasi."""
    entry_id = doc["id"]
    party = doc.get("party_name", "")
    mandi = doc.get("mandi_name", "")
    party_label = party or "Pvt Paddy"
    truck_no = doc.get("truck_no", "")
    date = doc.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    qntl = doc.get("final_qntl", 0) or doc.get("qntl", 0) or 0
    rate = doc.get("rate_per_qntl", 0) or doc.get("rate", 0) or 0
    if not rate and qntl:
        rate = round(float(doc.get("total_amount", 0) or 0) / float(qntl), 2)
    detail = _fmt_detail(qntl, rate) if qntl and rate else ""
    base_fields = {
        "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
        "created_by": username or "system", "linked_entry_id": entry_id,
        "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # --- Party Jama --- what we owe the party for paddy purchase (shows in Cash Transactions)
    total_amount = float(doc.get("total_amount", 0) or 0)
    if total_amount > 0:
        party_jama_desc = f"Paddy Purchase: {party_label} - {qntl}Q @ Rs.{rate}/Q = Rs.{total_amount}"
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "cash", "txn_type": "jama",
            "category": party_label, "party_type": "Pvt Paddy Purchase",
            "description": party_jama_desc,
            "amount": round(total_amount, 2), "bank_name": "",
            "reference": f"pvt_party_jama:{entry_id[:8]}",
            **base_fields
        })
        # Also create ledger entry for Party Ledger view
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "ledger", "txn_type": "jama",
            "category": party_label, "party_type": "Pvt Paddy Purchase",
            "description": party_jama_desc,
            "amount": round(total_amount, 2), "bank_name": "",
            "reference": f"pvt_party_jama_ledger:{entry_id[:8]}",
            **base_fields
        })
    
    # --- Truck Jama (Ledger) --- what we owe the truck for transport
    final_qntl = round(doc.get("final_qntl", 0) or doc.get("quantity_qntl", 0) or 0, 2)
    if final_qntl > 0 and truck_no:
        existing_rate_doc = await db.truck_payments.find_one(
            {"entry_id": {"$in": [e["id"] async for e in db.mill_entries.find({"truck_no": truck_no}, {"_id": 0, "id": 1})]}},
            {"_id": 0, "rate_per_qntl": 1}
        )
        if not existing_rate_doc:
            existing_rate_doc = await db.truck_payments.find_one(
                {"entry_id": {"$in": [e["id"] async for e in db.private_paddy.find({"truck_no": truck_no, "id": {"$ne": entry_id}}, {"_id": 0, "id": 1})]}},
                {"_id": 0, "rate_per_qntl": 1}
            )
        truck_rate = existing_rate_doc.get("rate_per_qntl", 32) if existing_rate_doc else 32
        gross_amount = round(final_qntl * truck_rate, 2)
        jama_desc = f"Pvt Paddy Truck: {truck_no} - {party_label} - {final_qntl}Q @ Rs.{truck_rate}"
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "ledger", "txn_type": "jama",
            "category": truck_no, "party_type": "Truck",
            "description": jama_desc,
            "amount": round(gross_amount, 2), "bank_name": "",
            "reference": f"pvt_truck_jama:{entry_id[:8]}",
            **base_fields
        })
    
    cash_paid = float(doc.get("cash_paid", 0) or 0)
    if cash_paid > 0 and truck_no:
        cash_desc = f"{party_label} - {detail}" if detail else f"{party_label} - Rs.{cash_paid}"
        # Cash Book nikasi (under truck)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "cash", "txn_type": "nikasi",
            "category": truck_no, "party_type": "Truck",
            "description": cash_desc,
            "amount": round(cash_paid, 2), "reference": f"pvt_paddy_cash:{entry_id[:8]}",
            **base_fields
        })
        # Truck Ledger nikasi (so it shows in truck payment)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "ledger", "txn_type": "nikasi",
            "category": truck_no, "party_type": "Truck",
            "description": cash_desc,
            "amount": round(cash_paid, 2), "reference": f"pvt_paddy_tcash:{entry_id[:8]}",
            **base_fields
        })
    diesel_paid = float(doc.get("diesel_paid", 0) or 0)
    if diesel_paid > 0:
        diesel_desc = f"{party_label} - {detail}" if detail else f"{party_label} - Rs.{diesel_paid}"
        # Diesel Account entry
        default_pump = await db.diesel_pumps.find_one({"is_default": True}, {"_id": 0})
        pump_name = default_pump["name"] if default_pump else "Default Pump"
        pump_id = default_pump["id"] if default_pump else "default"
        await db.diesel_accounts.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "pump_id": pump_id, "pump_name": pump_name,
            "truck_no": truck_no, "agent_name": doc.get("agent_name", ""),
            "mandi_name": mandi, "amount": round(diesel_paid, 2), "txn_type": "debit",
            "description": diesel_desc,
            **base_fields
        })
        if truck_no:
            # Truck Ledger nikasi for diesel
            await db.cash_transactions.insert_one({
                "id": str(uuid.uuid4()), "date": date,
                "account": "ledger", "txn_type": "nikasi",
                "category": truck_no, "party_type": "Truck",
                "description": diesel_desc,
                "amount": round(diesel_paid, 2), "reference": f"pvt_paddy_tdiesel:{entry_id[:8]}",
                **base_fields
            })
    advance_paid = float(doc.get("paid_amount", 0) or 0)
    if advance_paid > 0:
        adv_desc = f"Advance - {detail}" if detail else f"Advance - {party_label} - Rs.{advance_paid}"
        # Cash Book nikasi for advance (under party)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "cash", "txn_type": "nikasi",
            "category": party_label, "party_type": "Pvt Paddy Purchase",
            "description": adv_desc,
            "amount": round(advance_paid, 2), "reference": f"pvt_paddy_adv:{entry_id[:8]}",
            **base_fields
        })
        # Ledger nikasi for advance (so it shows in party ledger)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "ledger", "txn_type": "nikasi",
            "category": party_label, "party_type": "Pvt Paddy Purchase",
            "description": adv_desc,
            "amount": round(advance_paid, 2), "reference": f"pvt_paddy_advl:{entry_id[:8]}",
            **base_fields
        })



@router.post("/private-paddy/migrate-cashbook")
async def migrate_pvt_paddy_cashbook():
    """One-time migration: create cash book + diesel entries for existing pvt paddy entries that don't have them."""
    entries = await db.private_paddy.find({}, {"_id": 0}).to_list(10000)
    migrated = 0
    for doc in entries:
        cash_paid = float(doc.get("cash_paid", 0) or 0)
        diesel_paid = float(doc.get("diesel_paid", 0) or 0)
        advance_paid = float(doc.get("paid_amount", 0) or 0)
        if cash_paid <= 0 and diesel_paid <= 0 and advance_paid <= 0:
            continue
        existing = await db.cash_transactions.find_one({"linked_entry_id": doc["id"], "reference": {"$regex": "^(pvt_paddy|pvt_party_jama:|pvt_truck_jama:)"}})
        if existing:
            continue
        await _create_cashbook_diesel_for_pvt_paddy(doc, "migration")
        migrated += 1
    return {"message": f"Migrated {migrated} entries", "total_checked": len(entries)}


@router.get("/private-paddy")
async def get_private_paddy(kms_year: Optional[str] = None, season: Optional[str] = None, party_name: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    items = await db.private_paddy.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    return items

@router.put("/private-paddy/{item_id}")
async def update_private_paddy(item_id: str, data: dict, username: str = ""):
    existing = await db.private_paddy.find_one({"id": item_id})
    if not existing: raise HTTPException(status_code=404, detail="Not found")
    update_data = {k: v for k, v in data.items() if v is not None}
    for f in ["kg", "rate_per_qntl", "g_deposite", "gbw_cut", "moisture", "cutting_percent", "disc_dust_poll", "paid_amount", "cash_paid", "diesel_paid", "advance_paid"]:
        if f in update_data: update_data[f] = float(update_data[f] or 0)
    for f in ["bag", "plastic_bag", "g_issued"]:
        if f in update_data: update_data[f] = int(float(update_data[f] or 0))
    merged = {**existing, **update_data}
    merged["qntl"] = round(merged["kg"] / 100, 2) if merged["kg"] else 0
    merged["mill_w"] = round(merged["kg"] - merged["gbw_cut"], 2)
    merged["p_pkt_cut"] = round(merged["plastic_bag"] * 0.5, 2)
    moist_pct = max(0, merged["moisture"] - 17) if merged["moisture"] > 17 else 0
    merged["moisture_cut_percent"] = round(moist_pct, 2)
    merged["moisture_cut"] = round(merged["mill_w"] * moist_pct / 100, 2)
    after_moisture = merged["mill_w"] - merged["moisture_cut"]
    merged["cutting"] = round(after_moisture * merged["cutting_percent"] / 100, 2)
    merged["final_w"] = round(after_moisture - merged["cutting"] - merged["p_pkt_cut"] - merged["disc_dust_poll"], 2)
    merged["final_qntl"] = round(merged["final_w"] / 100, 2)
    merged["total_amount"] = round(merged["final_qntl"] * merged["rate_per_qntl"], 2)
    merged["balance"] = round(merged["total_amount"] - merged.get("paid_amount", 0), 2)
    merged["updated_at"] = datetime.now(timezone.utc).isoformat()
    merged.pop("_id", None)
    await db.private_paddy.update_one({"id": item_id}, {"$set": merged})
    # Re-create gunny bag entries
    await db.gunny_bags.delete_many({"linked_entry_id": item_id})
    await _create_gunny_entries_for_pvt_paddy(merged, username)
    # Re-create cash book + diesel entries (delete both pvt_paddy*, pvt_party_jama:*, and pvt_truck_jama:* refs)
    await db.cash_transactions.delete_many({"linked_entry_id": item_id, "reference": {"$regex": "^(pvt_paddy|pvt_party_jama:|pvt_truck_jama:)"}})
    await db.diesel_accounts.delete_many({"linked_entry_id": item_id})
    await _create_cashbook_diesel_for_pvt_paddy(merged, username)
    return merged

@router.delete("/private-paddy/{item_id}")
async def delete_private_paddy(item_id: str):
    result = await db.private_paddy.delete_one({"id": item_id})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Not found")
    await db.gunny_bags.delete_many({"linked_entry_id": item_id})
    await db.cash_transactions.delete_many({"linked_entry_id": item_id, "reference": {"$regex": "^(pvt_paddy|pvt_party_jama:|pvt_truck_jama:)"}})
    await db.diesel_accounts.delete_many({"linked_entry_id": item_id})
    await db.truck_payments.delete_many({"entry_id": item_id})
    return {"message": "Deleted", "id": item_id}

# --- Rice Sale ---
@router.post("/rice-sales")
async def create_rice_sale(data: dict, username: str = "", role: str = ""):
    qty = float(data.get("quantity_qntl", 0) or 0)
    rate = float(data.get("rate_per_qntl", 0) or 0)
    total = round(qty * rate, 2)
    paid = float(data.get("paid_amount", 0) or 0)
    doc = {
        "id": str(uuid.uuid4()), "date": data.get("date", ""),
        "kms_year": data.get("kms_year", ""), "season": data.get("season", ""),
        "party_name": data.get("party_name", ""), "rice_type": data.get("rice_type", ""),
        "rst_no": data.get("rst_no", ""),
        "quantity_qntl": qty, "rate_per_qntl": rate, "total_amount": total,
        "bags": int(data.get("bags", 0) or 0), "truck_no": data.get("truck_no", ""),
        "cash_paid": float(data.get("cash_paid", 0) or 0),
        "diesel_paid": float(data.get("diesel_paid", 0) or 0),
        "paid_amount": paid, "balance": round(total - paid, 2),
        "remark": data.get("remark", ""),
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.rice_sales.insert_one(doc)
    doc.pop("_id", None)
    await _create_cashbook_for_rice_sale(doc, username)
    return doc


async def _create_cashbook_for_rice_sale(doc, username=""):
    """Auto-create truck payment entries for cash/diesel + party ledger jama for rice sale."""
    entry_id = doc["id"]
    party = doc.get("party_name", "")
    truck_no = doc.get("truck_no", "")
    date = doc.get("date", "")
    qty = doc.get("quantity_qntl", 0) or 0
    rate = doc.get("rate_per_qntl", 0) or 0
    detail = _fmt_detail(qty, rate) if qty and rate else ""
    total = float(doc.get("total_amount", 0) or 0)
    base = {
        "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
        "created_by": username or "system", "linked_entry_id": entry_id,
        "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
    }
    # 1. Party Ledger: jama (sale amount receivable from buyer)
    if total > 0:
        sale_desc = f"Rice Sale: {party} - {detail}" if detail else f"Rice Sale: {party} - Rs.{total}"
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "ledger", "txn_type": "jama",
            "category": party, "party_type": "Rice Sale",
            "description": sale_desc,
            "amount": round(total, 2), "reference": f"rice_sale_jama:{entry_id[:8]}",
            **base
        })
    # 2. Cash paid → truck payment
    cash_paid = float(doc.get("cash_paid", 0) or 0)
    if cash_paid > 0 and truck_no:
        cash_desc = f"Rice Sale: {party} - {detail}" if detail else f"Rice Sale: {party} - Rs.{cash_paid}"
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "cash", "txn_type": "nikasi",
            "category": truck_no, "party_type": "Truck",
            "description": cash_desc,
            "amount": round(cash_paid, 2), "reference": f"rice_sale_cash:{entry_id[:8]}",
            **base
        })
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "ledger", "txn_type": "nikasi",
            "category": truck_no, "party_type": "Truck",
            "description": cash_desc,
            "amount": round(cash_paid, 2), "reference": f"rice_sale_tcash:{entry_id[:8]}",
            **base
        })
    # 3. Diesel paid → diesel account + truck ledger
    diesel_paid = float(doc.get("diesel_paid", 0) or 0)
    if diesel_paid > 0:
        diesel_desc = f"Rice Sale: {party} - {detail}" if detail else f"Rice Sale: {party} - Rs.{diesel_paid}"
        default_pump = await db.diesel_pumps.find_one({"is_default": True}, {"_id": 0})
        pump_name = default_pump["name"] if default_pump else "Default Pump"
        pump_id = default_pump["id"] if default_pump else "default"
        await db.diesel_accounts.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "pump_id": pump_id, "pump_name": pump_name,
            "truck_no": truck_no, "agent_name": "",
            "mandi_name": "", "amount": round(diesel_paid, 2), "txn_type": "debit",
            "description": diesel_desc,
            **base
        })
        if truck_no:
            await db.cash_transactions.insert_one({
                "id": str(uuid.uuid4()), "date": date,
                "account": "ledger", "txn_type": "nikasi",
                "category": truck_no, "party_type": "Truck",
                "description": diesel_desc,
                "amount": round(diesel_paid, 2), "reference": f"rice_sale_tdiesel:{entry_id[:8]}",
                **base
            })
    # 4. Advance received → cash jama + ledger nikasi
    advance = float(doc.get("paid_amount", 0) or 0)
    if advance > 0:
        adv_desc = f"Rice Advance: {party} - {detail}" if detail else f"Rice Advance: {party} - Rs.{advance}"
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "cash", "txn_type": "jama",
            "category": party, "party_type": "Rice Sale",
            "description": adv_desc,
            "amount": round(advance, 2), "reference": f"rice_sale_adv:{entry_id[:8]}",
            **base
        })
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "date": date,
            "account": "ledger", "txn_type": "nikasi",
            "category": party, "party_type": "Rice Sale",
            "description": adv_desc,
            "amount": round(advance, 2), "reference": f"rice_sale_adv_ledger:{entry_id[:8]}",
            **base
        })

@router.get("/rice-sales")
async def get_rice_sales(kms_year: Optional[str] = None, season: Optional[str] = None, party_name: Optional[str] = None, search: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"rst_no": {"$regex": search, "$options": "i"}},
            {"truck_no": {"$regex": search, "$options": "i"}},
            {"rice_type": {"$regex": search, "$options": "i"}},
        ]
    items = await db.rice_sales.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    return items

@router.put("/rice-sales/{item_id}")
async def update_rice_sale(item_id: str, data: dict):
    existing = await db.rice_sales.find_one({"id": item_id})
    if not existing: raise HTTPException(status_code=404, detail="Not found")
    update_data = {k: v for k, v in data.items() if v is not None}
    for f in ["quantity_qntl", "rate_per_qntl", "paid_amount", "cash_paid", "diesel_paid"]:
        if f in update_data: update_data[f] = float(update_data[f]) if update_data[f] != "" else 0
    if "bags" in update_data: update_data["bags"] = int(update_data["bags"]) if update_data["bags"] != "" else 0
    merged = {**existing, **update_data}
    merged["total_amount"] = round(merged["quantity_qntl"] * merged["rate_per_qntl"], 2)
    merged["balance"] = round(merged["total_amount"] - merged.get("paid_amount", 0), 2)
    merged["updated_at"] = datetime.now(timezone.utc).isoformat()
    merged.pop("_id", None)
    await db.rice_sales.update_one({"id": item_id}, {"$set": merged})
    # Re-create cash/diesel entries
    await db.cash_transactions.delete_many({"linked_entry_id": item_id, "reference": {"$regex": "^rice_sale_"}})
    await db.diesel_accounts.delete_many({"linked_entry_id": item_id})
    await _create_cashbook_for_rice_sale(merged, merged.get("created_by", ""))
    return merged

@router.delete("/rice-sales/{item_id}")
async def delete_rice_sale(item_id: str):
    result = await db.rice_sales.delete_one({"id": item_id})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Not found")
    # Cascade delete linked entries
    await db.cash_transactions.delete_many({"linked_entry_id": item_id})
    await db.diesel_accounts.delete_many({"linked_entry_id": item_id})
    # Delete linked payments and their cash entries
    payments = await db.private_payments.find({"ref_id": item_id, "ref_type": "rice_sale"}, {"_id": 0, "id": 1}).to_list(1000)
    for p in payments:
        await db.cash_transactions.delete_many({"linked_payment_id": p["id"]})
    await db.private_payments.delete_many({"ref_id": item_id, "ref_type": "rice_sale"})
    return {"message": "Deleted", "id": item_id}

# --- Private Payments (for both paddy purchase & rice sale parties) ---
@router.post("/private-payments")
async def create_private_payment(data: dict, username: str = "", role: str = ""):
    doc = {
        "id": str(uuid.uuid4()), "date": data.get("date", ""),
        "kms_year": data.get("kms_year", ""), "season": data.get("season", ""),
        "party_name": data.get("party_name", ""), "payment_type": data.get("payment_type", ""),
        "ref_type": data.get("ref_type", ""),  # "paddy_purchase" or "rice_sale"
        "ref_id": data.get("ref_id", ""),
        "amount": float(data.get("amount", 0)),
        "mode": data.get("mode", "cash"),  # cash/bank/cheque
        "reference": data.get("reference", ""),
        "remark": data.get("remark", ""),
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.private_payments.insert_one(doc)
    # Update balance on the referenced entry
    if doc["ref_type"] == "paddy_purchase" and doc["ref_id"]:
        entry = await db.private_paddy.find_one({"id": doc["ref_id"]})
        if entry:
            new_paid = round(entry.get("paid_amount", 0) + doc["amount"], 2)
            await db.private_paddy.update_one({"id": doc["ref_id"]}, {"$set": {"paid_amount": new_paid, "balance": round(entry.get("total_amount", 0) - new_paid, 2)}})
    elif doc["ref_type"] == "rice_sale" and doc["ref_id"]:
        entry = await db.rice_sales.find_one({"id": doc["ref_id"]})
        if entry:
            new_paid = round(entry.get("paid_amount", 0) + doc["amount"], 2)
            await db.rice_sales.update_one({"id": doc["ref_id"]}, {"$set": {"paid_amount": new_paid, "balance": round(entry.get("total_amount", 0) - new_paid, 2)}})
    # Auto-create Cash Book + Ledger entries
    account = "bank" if doc["mode"] == "bank" else "cash"
    base_cb = {
        "date": doc["date"], "kms_year": doc["kms_year"], "season": doc["season"],
        "created_by": username, "linked_payment_id": doc["id"],
        "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if doc["ref_type"] == "paddy_purchase":
        # Build party label from entry data
        ref_entry = await db.private_paddy.find_one({"id": doc["ref_id"]}, {"_id": 0})
        party = doc["party_name"]
        mandi = ref_entry.get("mandi_name", "") if ref_entry else ""
        party_label = f"{party} - {mandi}" if party and mandi else party
        qntl = ref_entry.get("qntl", 0) if ref_entry else 0
        rate = ref_entry.get("rate", 0) if ref_entry else 0
        if not rate and qntl and ref_entry:
            rate = round(float(ref_entry.get("total_amount", 0) or 0) / float(qntl), 2)
        detail = _fmt_detail(qntl, rate) if qntl and rate else f"Rs.{doc['amount']}"
        pay_desc = f"{party_label} - {detail}"
        # Cash Book nikasi
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "account": account, "txn_type": "nikasi",
            "category": party_label, "party_type": "Pvt Paddy Purchase",
            "description": pay_desc,
            "amount": doc["amount"], "reference": doc["reference"] or f"pvt_pay:{doc['id'][:8]}",
            **base_cb
        })
        # Party Ledger entry (credit = nikasi in party's account)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "account": "ledger", "txn_type": "nikasi",
            "category": party_label, "party_type": "Pvt Paddy Purchase",
            "description": pay_desc,
            "amount": doc["amount"], "reference": doc["reference"] or f"pvt_pay_ledger:{doc['id'][:8]}",
            **base_cb
        })
    else:
        # Rice Sale - party label
        ref_entry = await db.rice_sales.find_one({"id": doc["ref_id"]}, {"_id": 0})
        party = doc["party_name"]
        party_label = party
        # Cash Book jama (money received)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "account": account, "txn_type": "jama",
            "category": party_label, "party_type": "Rice Sale",
            "description": f"Rice Sale Payment Received: {party_label} - Rs.{doc['amount']}",
            "amount": doc["amount"], "reference": doc["reference"] or f"rice_pay:{doc['id'][:8]}",
            **base_cb
        })
        # Ledger nikasi (buyer's debt reduced)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "account": "ledger", "txn_type": "nikasi",
            "category": party_label, "party_type": "Rice Sale",
            "description": f"Rice Sale Payment Received: {party_label} - Rs.{doc['amount']}",
            "amount": doc["amount"], "reference": doc["reference"] or f"rice_pay_ledger:{doc['id'][:8]}",
            **base_cb
        })
    doc.pop("_id", None)
    return doc

@router.get("/private-payments")
async def get_private_payments(party_name: Optional[str] = None, ref_type: Optional[str] = None, ref_id: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if ref_type: query["ref_type"] = ref_type
    if ref_id: query["ref_id"] = ref_id
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    items = await db.private_payments.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    return items

@router.delete("/private-payments/{pay_id}")
async def delete_private_payment(pay_id: str):
    pay = await db.private_payments.find_one({"id": pay_id})
    if not pay: raise HTTPException(status_code=404, detail="Not found")
    # Reverse the payment from the referenced entry
    if pay.get("ref_type") == "paddy_purchase" and pay.get("ref_id"):
        entry = await db.private_paddy.find_one({"id": pay["ref_id"]})
        if entry:
            new_paid = round(max(0, entry.get("paid_amount", 0) - pay["amount"]), 2)
            await db.private_paddy.update_one({"id": pay["ref_id"]}, {"$set": {"paid_amount": new_paid, "balance": round(entry.get("total_amount", 0) - new_paid, 2)}})
    elif pay.get("ref_type") == "rice_sale" and pay.get("ref_id"):
        entry = await db.rice_sales.find_one({"id": pay["ref_id"]})
        if entry:
            new_paid = round(max(0, entry.get("paid_amount", 0) - pay["amount"]), 2)
            await db.rice_sales.update_one({"id": pay["ref_id"]}, {"$set": {"paid_amount": new_paid, "balance": round(entry.get("total_amount", 0) - new_paid, 2)}})
    # Delete linked Cash Book entry
    await db.cash_transactions.delete_many({"linked_payment_id": pay_id})
    await db.private_payments.delete_one({"id": pay_id})
    return {"message": "Deleted", "id": pay_id}


# ============ MARK PAID / UNDO / HISTORY ============

@router.post("/private-paddy/{entry_id}/mark-paid")
async def mark_pvt_paddy_paid(entry_id: str, username: str = "", role: str = ""):
    """Mark pvt paddy entry as fully paid"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin paid mark kar sakta hai")
    entry = await db.private_paddy.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    total = float(entry.get("total_amount", 0) or 0)
    already_paid = float(entry.get("paid_amount", 0) or 0)
    remaining = round(total - already_paid, 2)
    party = entry.get("party_name", "")
    mandi = entry.get("mandi_name", "")
    party_label = f"{party} - {mandi}" if party and mandi else party
    qntl = entry.get("qntl", 0) or 0
    rate = entry.get("rate", 0) or 0
    if not rate and qntl:
        rate = round(total / float(qntl), 2)
    detail = _fmt_detail(qntl, rate) if qntl and rate else f"Rs.{total}"
    # Update entry
    await db.private_paddy.update_one({"id": entry_id}, {"$set": {
        "paid_amount": total, "balance": 0, "payment_status": "paid",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }})
    # Create cash book + ledger entries for remaining balance
    if remaining > 0:
        mark_id = f"mark_paid:{entry_id[:8]}"
        base = {
            "date": entry.get("date", ""), "kms_year": entry.get("kms_year", ""),
            "season": entry.get("season", ""), "created_by": username,
            "linked_payment_id": mark_id,
            "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        desc = f"{party_label} - {detail} (Mark Paid)"
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "account": "cash", "txn_type": "nikasi",
            "category": party_label, "party_type": "Pvt Paddy Purchase",
            "description": desc, "amount": remaining, "reference": mark_id, **base
        })
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "account": "ledger", "txn_type": "nikasi",
            "category": party_label, "party_type": "Pvt Paddy Purchase",
            "description": desc, "amount": remaining, "reference": f"mark_paid_ledger:{entry_id[:8]}", **base
        })
    return {"success": True, "message": f"Marked paid - Rs.{remaining} cleared"}


@router.post("/private-paddy/{entry_id}/undo-paid")
async def undo_pvt_paddy_paid(entry_id: str, username: str = "", role: str = ""):
    """Undo paid - reset all payments to 0"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin undo kar sakta hai")
    entry = await db.private_paddy.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    total = float(entry.get("total_amount", 0) or 0)
    # Reset entry
    await db.private_paddy.update_one({"id": entry_id}, {"$set": {
        "paid_amount": 0, "balance": total, "payment_status": "pending",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }})
    # Delete all linked private_payments
    await db.private_payments.delete_many({"ref_id": entry_id, "ref_type": "paddy_purchase"})
    # Delete all linked cash book entries (from payments and mark-paid)
    await db.cash_transactions.delete_many({"linked_payment_id": {"$regex": f"mark_paid:{entry_id[:8]}|mark_paid_ledger:{entry_id[:8]}"}})
    # Also delete payment-linked cash entries
    payments = await db.private_payments.find({"ref_id": entry_id}, {"_id": 0, "id": 1}).to_list(1000)
    for p in payments:
        await db.cash_transactions.delete_many({"linked_payment_id": p["id"]})
    # Delete advance entry
    await db.cash_transactions.delete_many({"linked_entry_id": entry_id, "reference": {"$regex": "pvt_paddy_adv"}})
    return {"success": True, "message": "Payment undo - sab reset ho gaya"}


@router.get("/private-paddy/{entry_id}/history")
async def get_pvt_paddy_history(entry_id: str):
    """Get payment history for a pvt paddy entry"""
    payments = await db.private_payments.find(
        {"ref_id": entry_id, "ref_type": "paddy_purchase"},
        {"_id": 0}
    ).sort([("created_at", -1)]).to_list(1000)
    entry = await db.private_paddy.find_one({"id": entry_id}, {"_id": 0})
    total_paid = float(entry.get("paid_amount", 0)) if entry else 0
    return {"history": payments, "total_paid": total_paid}


# ============ RICE SALE: MARK PAID / UNDO / HISTORY ============

@router.post("/rice-sales/{entry_id}/mark-paid")
async def mark_rice_sale_paid(entry_id: str, username: str = "", role: str = ""):
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin paid mark kar sakta hai")
    entry = await db.rice_sales.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    total = float(entry.get("total_amount", 0) or 0)
    already_paid = float(entry.get("paid_amount", 0) or 0)
    remaining = round(total - already_paid, 2)
    party = entry.get("party_name", "")
    qty = entry.get("quantity_qntl", 0) or 0
    rate = entry.get("rate_per_qntl", 0) or 0
    detail = _fmt_detail(qty, rate) if qty and rate else f"Rs.{total}"
    await db.rice_sales.update_one({"id": entry_id}, {"$set": {
        "paid_amount": total, "balance": 0, "payment_status": "paid",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }})
    if remaining > 0:
        mark_id = f"mark_paid_rice:{entry_id[:8]}"
        base = {
            "date": entry.get("date", ""), "kms_year": entry.get("kms_year", ""),
            "season": entry.get("season", ""), "created_by": username,
            "linked_payment_id": mark_id,
            "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        desc = f"Rice Sale: {party} - {detail} (Mark Paid)"
        # Cash jama (money received in hand)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "account": "cash", "txn_type": "jama",
            "category": party, "party_type": "Rice Sale",
            "description": desc, "amount": remaining, "reference": mark_id, **base
        })
        # Ledger nikasi (buyer's debt reduced)
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()), "account": "ledger", "txn_type": "nikasi",
            "category": party, "party_type": "Rice Sale",
            "description": desc, "amount": remaining, "reference": f"mark_paid_rice_ledger:{entry_id[:8]}", **base
        })
    return {"success": True, "message": f"Marked paid - Rs.{remaining} cleared"}


@router.post("/rice-sales/{entry_id}/undo-paid")
async def undo_rice_sale_paid(entry_id: str, username: str = "", role: str = ""):
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin undo kar sakta hai")
    entry = await db.rice_sales.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    total = float(entry.get("total_amount", 0) or 0)
    await db.rice_sales.update_one({"id": entry_id}, {"$set": {
        "paid_amount": 0, "balance": total, "payment_status": "pending",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }})
    # Delete mark-paid entries
    await db.cash_transactions.delete_many({"linked_payment_id": {"$regex": f"mark_paid_rice:{entry_id[:8]}|mark_paid_rice_ledger:{entry_id[:8]}"}})
    # Delete all linked private_payments and their cash entries
    payments = await db.private_payments.find({"ref_id": entry_id, "ref_type": "rice_sale"}, {"_id": 0, "id": 1}).to_list(1000)
    for p in payments:
        await db.cash_transactions.delete_many({"linked_payment_id": p["id"]})
    await db.private_payments.delete_many({"ref_id": entry_id, "ref_type": "rice_sale"})
    return {"success": True, "message": "Payment undo - sab reset ho gaya"}


@router.get("/rice-sales/{entry_id}/history")
async def get_rice_sale_history(entry_id: str):
    payments = await db.private_payments.find(
        {"ref_id": entry_id, "ref_type": "rice_sale"},
        {"_id": 0}
    ).sort([("created_at", -1)]).to_list(1000)
    entry = await db.rice_sales.find_one({"id": entry_id}, {"_id": 0})
    total_paid = float(entry.get("paid_amount", 0)) if entry else 0
    return {"history": payments, "total_paid": total_paid}


@router.get("/private-payments/fix-old-entries")
async def fix_old_payment_cashbook_entries():
    """Fix ALL pvt paddy related cash_transactions descriptions to use 'qty @ Rs.rate' format."""
    fixed = 0
    # Fix entries linked to private_paddy (via linked_entry_id) - advance, cash, diesel
    linked_entries = await db.cash_transactions.find(
        {"linked_entry_id": {"$exists": True, "$ne": ""}},
        {"_id": 0}
    ).to_list(10000)
    for entry in linked_entries:
        eid = entry.get("linked_entry_id", "")
        if not eid:
            continue
        ref = await db.private_paddy.find_one({"id": eid}, {"_id": 0})
        if not ref:
            continue
        party = ref.get("party_name", "")
        mandi = ref.get("mandi_name", "")
        party_label = f"{party} - {mandi}" if party and mandi else party
        qntl = ref.get("qntl", 0) or 0
        rate = ref.get("rate", 0) or 0
        if not rate and qntl:
            rate = round(float(ref.get("total_amount", 0) or 0) / float(qntl), 2)
        detail = _fmt_detail(qntl, rate) if qntl and rate else ""
        ref_str = entry.get("reference", "")
        if "adv" in ref_str:
            desc = f"Advance - {detail}" if detail else f"Advance - {party_label}"
        else:
            desc = f"{party_label} - {detail}" if detail else party_label
        await db.cash_transactions.update_one({"id": entry["id"]}, {"$set": {"description": desc}})
        fixed += 1
    # Fix entries linked to private_payments (via linked_payment_id) - ₹ button payments
    pay_entries = await db.cash_transactions.find(
        {"linked_payment_id": {"$exists": True, "$ne": ""}},
        {"_id": 0}
    ).to_list(10000)
    for entry in pay_entries:
        pay_id = entry.get("linked_payment_id", "")
        if not pay_id:
            continue
        pay = await db.private_payments.find_one({"id": pay_id}, {"_id": 0})
        if not pay:
            continue
        party = pay.get("party_name", "")
        if not party:
            continue
        mandi = ""
        ref = None
        if pay.get("ref_type") == "paddy_purchase" and pay.get("ref_id"):
            ref = await db.private_paddy.find_one({"id": pay["ref_id"]}, {"_id": 0})
            if ref:
                mandi = ref.get("mandi_name", "")
        party_label = f"{party} - {mandi}" if party and mandi else party
        is_paddy = pay.get("ref_type") == "paddy_purchase"
        party_type = "Pvt Paddy Purchase" if is_paddy else "Rice Sale"
        qntl = ref.get("qntl", 0) if ref else 0
        rate = ref.get("rate", 0) if ref else 0
        if not rate and qntl and ref:
            rate = round(float(ref.get("total_amount", 0) or 0) / float(qntl), 2)
        detail = _fmt_detail(qntl, rate) if qntl and rate else f"Rs.{pay['amount']}"
        desc = f"{party_label} - {detail}"
        await db.cash_transactions.update_one(
            {"id": entry["id"]},
            {"$set": {"category": party_label, "party_type": party_type, "description": desc}}
        )
        fixed += 1
    # Also fix diesel_accounts descriptions
    diesel_entries = await db.diesel_accounts.find(
        {"linked_entry_id": {"$exists": True, "$ne": ""}},
        {"_id": 0}
    ).to_list(10000)
    for entry in diesel_entries:
        eid = entry.get("linked_entry_id", "")
        if not eid:
            continue
        ref = await db.private_paddy.find_one({"id": eid}, {"_id": 0})
        if not ref:
            continue
        party = ref.get("party_name", "")
        mandi = ref.get("mandi_name", "")
        party_label = f"{party} - {mandi}" if party and mandi else party
        qntl = ref.get("qntl", 0) or 0
        rate = ref.get("rate", 0) or 0
        if not rate and qntl:
            rate = round(float(ref.get("total_amount", 0) or 0) / float(qntl), 2)
        detail = _fmt_detail(qntl, rate) if qntl and rate else ""
        desc = f"{party_label} - {detail}" if detail else party_label
        await db.diesel_accounts.update_one({"id": entry["id"]}, {"$set": {"description": desc}})
        fixed += 1
    return {"message": f"Fixed {fixed} entries with new description format"}



@router.get("/migrate/fix-missing-ledger-nikasi")
async def fix_missing_ledger_nikasi():
    """Migration: Find all cash nikasi entries that don't have corresponding ledger nikasi entries and create them.
    This fixes the bug where payments didn't create ledger entries, causing Party Summary to show wrong balances."""
    fixed = 0
    skipped = 0
    details = []
    
    # Get ALL cash/bank nikasi entries (these are payments made)
    cash_payments = await db.cash_transactions.find({
        "account": {"$in": ["cash", "bank"]},
        "txn_type": "nikasi"
    }, {"_id": 0}).to_list(None)
    
    for payment in cash_payments:
        cat = payment.get("category", "").strip()
        amt = payment.get("amount", 0)
        pt = payment.get("party_type", "")
        desc = payment.get("description", "")
        
        if not cat or amt <= 0:
            continue
        
        # Check if a ledger nikasi with same description and amount exists
        existing = await db.cash_transactions.find_one({
            "account": "ledger",
            "txn_type": "nikasi",
            "category": cat,
            "amount": amt,
            "description": desc
        })
        
        if existing:
            skipped += 1
            continue
        
        # Create missing ledger nikasi
        import uuid as _uuid
        entry = {
            "id": str(_uuid.uuid4()),
            "date": payment.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "account": "ledger",
            "txn_type": "nikasi",
            "category": cat,
            "party_type": pt,
            "description": desc,
            "amount": amt,
            "reference": f"migration_ledger:{cat[:10]}",
            "kms_year": payment.get("kms_year", ""),
            "season": payment.get("season", ""),
            "created_by": "migration",
            "linked_payment_id": f"migration:{cat}:{_uuid.uuid4().hex[:6]}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.cash_transactions.insert_one(entry)
        fixed += 1
        details.append(f"{cat} ({pt}) - Rs.{amt} - {desc[:50]}")
    
    return {
        "success": True,
        "message": f"Migration complete: {fixed} ledger entries created, {skipped} already existed",
        "fixed_count": fixed,
        "skipped_count": skipped,
        "details": details
    }

# ============ EXPORT: Private Paddy PDF/Excel ============

@router.get("/private-paddy/excel")
async def export_private_paddy_excel(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None):
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    items = await db.private_paddy.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    if search:
        s = search.lower()
        items = [i for i in items if s in (i.get("party_name","")).lower() or s in (i.get("mandi_name","")).lower() or s in (i.get("agent_name","")).lower()]
    # Normalize fields for agent_extra entries
    for item in items:
        if not item.get("final_qntl") and item.get("quantity_qntl"):
            item["final_qntl"] = item["quantity_qntl"]
        if not item.get("balance"):
            item["balance"] = round((item.get("total_amount", 0) or 0) - (item.get("paid_amount", 0) or 0), 2)

    cols = get_columns("private_paddy_report")
    ncols = col_count(cols)
    headers = get_excel_headers(cols)
    widths = get_excel_widths(cols)

    wb = Workbook(); ws = wb.active; ws.title = "Pvt Paddy Purchase"
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)

    title = "Private Paddy Purchase / निजी धान खरीद"
    if kms_year: title += f" | KMS: {kms_year}"
    if season: title += f" | {season}"
    style_excel_title(ws, title, ncols)

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=h)
    style_excel_header_row(ws, 4, ncols)

    data_start = 5; row = data_start
    totals = {"total_kg": 0, "total_final_qntl": 0, "total_amount": 0, "total_paid": 0, "total_balance": 0, "total_g_issued": 0, "total_cash": 0, "total_diesel": 0}
    for item in items:
        vals = get_entry_row(item, cols)
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            if cols[col_idx-1]["align"] == "right": c.alignment = Alignment(horizontal='right')
        totals["total_kg"] += item.get("kg", 0) or 0
        totals["total_final_qntl"] += item.get("final_qntl", 0) or 0
        totals["total_amount"] += item.get("total_amount", 0) or 0
        totals["total_paid"] += item.get("paid_amount", 0) or 0
        totals["total_balance"] += item.get("balance", 0) or 0
        totals["total_g_issued"] += item.get("g_issued", 0) or 0
        totals["total_cash"] += item.get("cash_paid", 0) or 0
        totals["total_diesel"] += item.get("diesel_paid", 0) or 0
        row += 1

    if items:
        style_excel_data_rows(ws, data_start, row - 1, ncols, headers)

    for k in totals: totals[k] = round(totals[k], 2)
    total_vals = get_total_row(totals, cols)
    ws.cell(row=row, column=1, value="TOTAL / कुल")
    for col_idx, val in enumerate(total_vals, 1):
        if val is not None:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.alignment = Alignment(horizontal='right')
    style_excel_total_row(ws, row, ncols)

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pvt_paddy_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/private-paddy/pdf")
async def export_private_paddy_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    items = await db.private_paddy.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    if search:
        s = search.lower()
        items = [i for i in items if s in (i.get("party_name","")).lower() or s in (i.get("mandi_name","")).lower() or s in (i.get("agent_name","")).lower()]
    for item in items:
        if not item.get("final_qntl") and item.get("quantity_qntl"):
            item["final_qntl"] = item["quantity_qntl"]
        if not item.get("balance"):
            item["balance"] = round((item.get("total_amount", 0) or 0) - (item.get("paid_amount", 0) or 0), 2)

    cols = get_columns("private_paddy_report")
    ncols = col_count(cols)
    headers = get_pdf_headers(cols)
    col_widths = [w*mm for w in get_pdf_widths_mm(cols)]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=8*mm, rightMargin=8*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []; styles = get_pdf_styles()

    from utils.export_helpers import get_pdf_table_style, get_pdf_company_header
    elements.extend(get_pdf_company_header())
    title = "Private Paddy Purchase"
    if kms_year: title += f" | KMS: {kms_year}"
    if season: title += f" | {season}"
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#1B4F72'), alignment=TA_CENTER)
    elements.append(Paragraph(title, title_style)); elements.append(Spacer(1, 8))

    table_data = [headers]
    totals = {"total_kg": 0, "total_final_qntl": 0, "total_amount": 0, "total_paid": 0, "total_balance": 0, "total_g_issued": 0, "total_cash": 0, "total_diesel": 0}
    for item in items:
        table_data.append([str(v) for v in get_entry_row(item, cols)])
        totals["total_kg"] += item.get("kg", 0) or 0
        totals["total_final_qntl"] += item.get("final_qntl", 0) or 0
        totals["total_amount"] += item.get("total_amount", 0) or 0
        totals["total_paid"] += item.get("paid_amount", 0) or 0
        totals["total_balance"] += item.get("balance", 0) or 0
        totals["total_g_issued"] += item.get("g_issued", 0) or 0
        totals["total_cash"] += item.get("cash_paid", 0) or 0
        totals["total_diesel"] += item.get("diesel_paid", 0) or 0
    for k in totals: totals[k] = round(totals[k], 2)
    total_vals = get_total_row(totals, cols)
    total_row = []
    for i, val in enumerate(total_vals):
        if i == 0: total_row.append("TOTAL / कुल")
        elif val is not None: total_row.append(str(val))
        else: total_row.append("")
    table_data.append(total_row)

    first_right = next((i for i, c in enumerate(cols) if c["align"] == "right"), 2)
    tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    cols_info = [{'header': h} for h in headers]
    style_cmds = get_pdf_table_style(len(table_data), cols_info)
    style_cmds.append(('ALIGN', (first_right, 1), (-1, -1), 'RIGHT'))
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)
    doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=pvt_paddy_{datetime.now().strftime('%Y%m%d')}.pdf"})


@router.get("/rice-sales/excel")
async def export_rice_sales_excel(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None):
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    items = await db.rice_sales.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    if search:
        s = search.lower()
        items = [i for i in items if s in (i.get("party_name","")).lower()]
    for item in items:
        if not item.get("balance"):
            item["balance"] = round((item.get("total_amount", 0) or 0) - (item.get("paid_amount", 0) or 0), 2)

    cols = get_columns("rice_sales_report")
    ncols = col_count(cols)
    headers = get_excel_headers(cols)
    widths = get_excel_widths(cols)

    wb = Workbook(); ws = wb.active; ws.title = "Rice Sales"
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS)

    title = "Rice Sales Report / चावल बिक्री"
    if kms_year: title += f" | KMS: {kms_year}"
    if season: title += f" | {season}"
    style_excel_title(ws, title, ncols)

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=h)
    style_excel_header_row(ws, 4, ncols)

    data_start = 5; row = data_start
    totals = {"total_qntl": 0, "total_amount": 0, "total_paid": 0, "total_balance": 0}
    for item in items:
        vals = get_entry_row(item, cols)
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            if cols[col_idx-1]["align"] == "right": c.alignment = Alignment(horizontal='right')
        totals["total_qntl"] += item.get("quantity_qntl", 0) or 0
        totals["total_amount"] += item.get("total_amount", 0) or 0
        totals["total_paid"] += item.get("paid_amount", 0) or 0
        totals["total_balance"] += item.get("balance", 0) or 0
        row += 1

    if items:
        style_excel_data_rows(ws, data_start, row - 1, ncols, headers)

    for k in totals: totals[k] = round(totals[k], 2)
    total_vals = get_total_row(totals, cols)
    ws.cell(row=row, column=1, value="TOTAL / कुल")
    for col_idx, val in enumerate(total_vals, 1):
        if val is not None:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.alignment = Alignment(horizontal='right')
    style_excel_total_row(ws, row, ncols)

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rice_sales_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/rice-sales/pdf")
async def export_rice_sales_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    items = await db.rice_sales.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    if search:
        s = search.lower()
        items = [i for i in items if s in (i.get("party_name","")).lower()]
    for item in items:
        if not item.get("balance"):
            item["balance"] = round((item.get("total_amount", 0) or 0) - (item.get("paid_amount", 0) or 0), 2)

    cols = get_columns("rice_sales_report")
    ncols = col_count(cols)
    headers = get_pdf_headers(cols)
    col_widths = [w*mm for w in get_pdf_widths_mm(cols)]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=8*mm, rightMargin=8*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []; styles = get_pdf_styles()

    from utils.export_helpers import get_pdf_table_style, get_pdf_company_header
    elements.extend(get_pdf_company_header())
    title = "Rice Sales Report"
    if kms_year: title += f" | KMS: {kms_year}"
    if season: title += f" | {season}"
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#1B4F72'), alignment=TA_CENTER)
    elements.append(Paragraph(title, title_style)); elements.append(Spacer(1, 8))

    table_data = [headers]
    totals = {"total_qntl": 0, "total_amount": 0, "total_paid": 0, "total_balance": 0}
    for item in items:
        table_data.append([str(v) for v in get_entry_row(item, cols)])
        totals["total_qntl"] += item.get("quantity_qntl", 0) or 0
        totals["total_amount"] += item.get("total_amount", 0) or 0
        totals["total_paid"] += item.get("paid_amount", 0) or 0
        totals["total_balance"] += item.get("balance", 0) or 0
    for k in totals: totals[k] = round(totals[k], 2)
    total_vals = get_total_row(totals, cols)
    total_row = []
    for i, val in enumerate(total_vals):
        if i == 0: total_row.append("TOTAL / कुल")
        elif val is not None: total_row.append(str(val))
        else: total_row.append("")
    table_data.append(total_row)

    first_right = next((i for i, c in enumerate(cols) if c["align"] == "right"), 2)
    tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    cols_info = [{'header': h} for h in headers]
    style_cmds = get_pdf_table_style(len(table_data), cols_info)
    style_cmds.append(('ALIGN', (first_right, 1), (-1, -1), 'RIGHT'))
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)
    doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rice_sales_{datetime.now().strftime('%Y%m%d')}.pdf"})




# ============ PARTY SUMMARY ============

async def _get_party_summary(kms_year=None, season=None, date_from=None, date_to=None, search=None, view_type=None):
    """Aggregate party-wise summary separately for paddy_purchase, sale_vouchers, and purchase_vouchers"""
    base_q = {}
    if kms_year: base_q["kms_year"] = kms_year
    if season: base_q["season"] = season
    if date_from or date_to:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to: dq["$lte"] = date_to
        base_q["date"] = dq

    paddy_items = await db.private_paddy.find(base_q, {"_id": 0}).to_list(5000)
    sale_voucher_items = await db.sale_vouchers.find(base_q, {"_id": 0}).to_list(5000)
    purchase_voucher_items = await db.purchase_vouchers.find(base_q, {"_id": 0}).to_list(5000)

    # Build ledger-based paid maps for sale and purchase vouchers (source of truth)
    ledger_q = {"account": "ledger", "txn_type": "nikasi"}
    if kms_year: ledger_q["kms_year"] = kms_year
    if season: ledger_q["season"] = season

    sale_party_names = list(set(v.get("party_name", "") for v in sale_voucher_items if v.get("party_name")))
    sale_ledger_paid = {}
    if sale_party_names:
        sale_ledger_txns = await db.cash_transactions.find(
            {**ledger_q, "category": {"$in": sale_party_names}, "party_type": "Sale Book"}, {"_id": 0}
        ).to_list(50000)
        for lt in sale_ledger_txns:
            pn = lt.get("category", "")
            sale_ledger_paid[pn] = sale_ledger_paid.get(pn, 0) + lt.get("amount", 0)

    purchase_party_names = list(set(v.get("party_name", "") for v in purchase_voucher_items if v.get("party_name")))
    purchase_ledger_paid = {}
    if purchase_party_names:
        purchase_ledger_txns = await db.cash_transactions.find(
            {**ledger_q, "category": {"$in": purchase_party_names}, "party_type": "Purchase Voucher"}, {"_id": 0}
        ).to_list(50000)
        for lt in purchase_ledger_txns:
            pn = lt.get("category", "")
            purchase_ledger_paid[pn] = purchase_ledger_paid.get(pn, 0) + lt.get("amount", 0)

    def _agg(items, amt_key, paid_key, label, ledger_paid_map=None):
        pmap = {}
        for item in items:
            name = item.get("party_name", "Unknown")
            if name not in pmap:
                pmap[name] = {"party_name": name, "amount": 0, "paid": 0, "balance": 0,
                              "mandi_name": item.get("mandi_name", ""), "agent_name": item.get("agent_name", ""),
                              "entries": 0}
            pm = pmap[name]
            pm["amount"] += item.get(amt_key, 0) or 0
            if ledger_paid_map is None:
                pm["paid"] += item.get(paid_key, 0) or 0
            pm["entries"] += 1
            if not pm["mandi_name"] and item.get("mandi_name"):
                pm["mandi_name"] = item["mandi_name"]
            if not pm["agent_name"] and item.get("agent_name"):
                pm["agent_name"] = item["agent_name"]
        # Override paid from ledger if available
        if ledger_paid_map is not None:
            for name, pm in pmap.items():
                pm["paid"] = round(ledger_paid_map.get(name, 0), 2)
        result = []
        for pm in pmap.values():
            pm["amount"] = round(pm["amount"], 2)
            pm["paid"] = round(pm["paid"], 2)
            pm["balance"] = round(pm["amount"] - pm["paid"], 2)
            result.append(pm)
        if search:
            s = search.lower()
            result = [r for r in result if s in r["party_name"].lower() or s in r.get("mandi_name", "").lower() or s in r.get("agent_name", "").lower()]
        result.sort(key=lambda x: abs(x["balance"]), reverse=True)
        total_amt = round(sum(r["amount"] for r in result), 2)
        total_paid = round(sum(r["paid"] for r in result), 2)
        total_bal = round(sum(r["balance"] for r in result), 2)
        return {"parties": result, "total_amount": total_amt, "total_paid": total_paid, "total_balance": total_bal}

    paddy_summary = _agg(paddy_items, "total_amount", "paid_amount", "Paddy Purchase")
    sale_summary = _agg(sale_voucher_items, "total", "paid_amount", "Sale Voucher", ledger_paid_map=sale_ledger_paid)
    purchase_summary = _agg(purchase_voucher_items, "total", "paid_amount", "Purchase Voucher", ledger_paid_map=purchase_ledger_paid)

    # Combined totals for top cards
    all_purchase = paddy_summary["total_amount"] + purchase_summary["total_amount"]
    all_purchase_paid = paddy_summary["total_paid"] + purchase_summary["total_paid"]
    all_sale = sale_summary["total_amount"]
    all_sale_rcvd = sale_summary["total_paid"]
    total_parties = len(set(
        [p["party_name"] for p in paddy_summary["parties"]] +
        [p["party_name"] for p in sale_summary["parties"]] +
        [p["party_name"] for p in purchase_summary["parties"]]
    ))

    totals = {
        "total_parties": total_parties,
        "total_purchase": round(all_purchase, 2),
        "total_purchase_paid": round(all_purchase_paid, 2),
        "total_purchase_balance": round(all_purchase - all_purchase_paid, 2),
        "total_sale": round(all_sale, 2),
        "total_sale_received": round(all_sale_rcvd, 2),
        "total_sale_balance": round(all_sale - all_sale_rcvd, 2),
        "total_net_balance": round((all_purchase - all_purchase_paid) - (all_sale - all_sale_rcvd), 2),
    }

    return {
        "paddy_purchase": paddy_summary,
        "sale_vouchers": sale_summary,
        "purchase_vouchers": purchase_summary,
        "totals": totals,
        "parties": [],  # backward compat
    }


@router.get("/private-trading/party-summary")
async def get_party_summary(kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, search: Optional[str] = None, view_type: Optional[str] = None):
    return await _get_party_summary(kms_year, season, date_from, date_to, search, view_type)


@router.get("/private-trading/party-summary/excel")
async def export_party_summary_excel(kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, search: Optional[str] = None, view_type: Optional[str] = None):
    from io import BytesIO
    data = await _get_party_summary(kms_year, season, date_from, date_to, search, view_type)

    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")

    wb = Workbook(); ws = wb.active; ws.title = "Party Summary"
    thin_s = Side(style='thin', color='CBD5E1')
    tb = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    hfill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=9)

    ws.merge_cells('A1:E1')
    ws['A1'] = f"{company} - Party Summary"
    ws['A1'].font = Font(bold=True, size=14, color="D97706")
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ["Party", "Entries", "Amount", "Paid/Received", "Balance"]
    widths_list = [25, 10, 18, 18, 18]
    sections = [
        ("Sale Vouchers", data.get("sale_vouchers", {}), PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"), "10B981"),
        ("Purchase Vouchers", data.get("purchase_vouchers", {}), PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"), "8B5CF6"),
        ("Paddy Purchase", data.get("paddy_purchase", {}), PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"), "D97706"),
    ]

    row = 3
    for sec_name, sec_data, sec_fill, sec_color in sections:
        parties = sec_data.get("parties", [])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=f"{sec_name} ({len(parties)} parties)")
        c.font = Font(bold=True, size=11, color=sec_color); c.fill = sec_fill; c.border = tb
        row += 1
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal='center'); c.border = tb
        row += 1
        for p in parties:
            vals = [p["party_name"], p.get("entries", 0), p.get("amount", 0), p.get("paid", 0), p.get("balance", 0)]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v); c.border = tb
                if ci >= 3: c.alignment = Alignment(horizontal='right')
                if ci == 5: c.font = Font(bold=True, color="DC2626" if (v or 0) > 0 else "059669")
            row += 1
        tf = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = tf; ws.cell(row=row, column=1).border = tb
        ws.cell(row=row, column=2, value=len(parties)).fill = tf; ws.cell(row=row, column=2).border = tb
        for ci, key in [(3, "total_amount"), (4, "total_paid"), (5, "total_balance")]:
            c = ws.cell(row=row, column=ci, value=sec_data.get(key, 0))
            c.font = Font(bold=True); c.fill = tf; c.border = tb; c.alignment = Alignment(horizontal='right')
        row += 2

    for i, w in enumerate(widths_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=party_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/private-trading/party-summary/pdf")
async def export_party_summary_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, search: Optional[str] = None, view_type: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from io import BytesIO

    data = await _get_party_summary(kms_year, season, date_from, date_to, search, view_type)
    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=10*mm)
    elements = []; styles = get_pdf_styles()

    from utils.export_helpers import get_pdf_company_header
    elements.extend(get_pdf_company_header())
    title = f"Party Summary"
    if kms_year: title += f" | {kms_year}"
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#D97706'), alignment=TA_CENTER, spaceAfter=8)
    elements.append(Paragraph(title, title_style))

    cell_s = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10)
    cell_r = ParagraphStyle('CellR', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_RIGHT)
    cell_b = ParagraphStyle('CellB', parent=styles['Normal'], fontSize=9, leading=11, alignment=TA_RIGHT)
    col_widths = [55*mm, 18*mm, 35*mm, 35*mm, 35*mm]

    sections = [
        ("Sale Vouchers", data.get("sale_vouchers", {}), colors.HexColor('#10B981')),
        ("Purchase Vouchers", data.get("purchase_vouchers", {}), colors.HexColor('#8B5CF6')),
        ("Paddy Purchase", data.get("paddy_purchase", {}), colors.HexColor('#D97706')),
    ]

    for sec_name, sec_data, sec_color in sections:
        parties = sec_data.get("parties", [])
        sec_style = ParagraphStyle('SecTitle', parent=styles['Normal'], fontSize=11, textColor=sec_color, spaceBefore=10, spaceAfter=4)
        elements.append(Paragraph(f"<b>{sec_name}</b> ({len(parties)} parties)", sec_style))
        table_data = [["Party", "Entries", "Amount", "Paid/Received", "Balance"]]
        for p in parties:
            table_data.append([
                Paragraph(f"<b>{p['party_name']}</b>", cell_s),
                Paragraph(str(p.get('entries', 0)), cell_r),
                Paragraph(f"Rs.{(p.get('amount', 0)):,.0f}", cell_r),
                Paragraph(f"Rs.{(p.get('paid', 0)):,.0f}", cell_r),
                Paragraph(f"<b>Rs.{(p.get('balance', 0)):,.0f}</b>", cell_b),
            ])
        table_data.append([
            Paragraph("<b>TOTAL</b>", cell_s), Paragraph(str(len(parties)), cell_r),
            Paragraph(f"<b>Rs.{sec_data.get('total_amount', 0):,.0f}</b>", cell_b),
            Paragraph(f"<b>Rs.{sec_data.get('total_paid', 0):,.0f}</b>", cell_b),
            Paragraph(f"<b>Rs.{sec_data.get('total_balance', 0):,.0f}</b>", cell_b),
        ])
        tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
        ]
        for i in range(1, len(table_data) - 1):
            if i % 2 == 0: style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
        tbl.setStyle(TableStyle(style_cmds))
        elements.append(tbl); elements.append(Spacer(1, 6))

    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#999'), alignment=TA_CENTER, spaceBefore=12)
    elements.append(Paragraph(f"{company} | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", footer_style))
    doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=party_summary_{datetime.now().strftime('%Y%m%d')}.pdf"})
