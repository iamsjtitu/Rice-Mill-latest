from fastapi import APIRouter, HTTPException
from datetime import datetime, timezone
from typing import Optional
from database import db
import uuid

router = APIRouter()

def fmt_date(d):
    if not d: return ""
    try:
        if "T" in str(d): d = str(d).split("T")[0]
        parts = str(d).split("-")
        if len(parts) == 3: return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except: pass
    return str(d)


async def _get_default_pump():
    pump = await db.diesel_accounts.find_one({}, {"_id": 0, "pump_name": 1}, sort=[("_id", -1)])
    return pump.get("pump_name", "Diesel Pump") if pump else "Diesel Pump"


async def _create_bp_ledger_entries(d, doc_id, username):
    """Create all accounting entries for a by-product sale"""
    party = (d.get('party_name') or '').strip()
    cash = d.get('cash_paid', 0) or 0
    diesel = d.get('diesel_paid', 0) or 0
    advance = d.get('advance', 0) or 0
    total = d.get('total', 0) or 0
    vehicle = (d.get('vehicle_no') or '').strip()
    product = d.get('product', 'By-Product')
    vno = d.get('voucher_no', '') or doc_id[:8]
    now_iso = datetime.now(timezone.utc).isoformat()
    base = {"kms_year": d.get('kms_year', ''), "season": d.get('season', ''), "created_by": username, "created_at": now_iso, "updated_at": now_iso}
    entries = []

    # 1. Party Ledger NIKASI: sale amount (humne maal diya - party owes us)
    # Split-billing: 2 alag ledger entries — `{party} (PKA)` + `{party} (KCA)`
    # Non-split: single ledger under `{party}`
    is_split = bool(d.get("split_billing"))
    if party and total > 0:
        if is_split:
            billed_amt = float(d.get("billed_amount", 0) or 0)
            tax_amt = float(d.get("tax_amount", 0) or 0)
            kaccha_amt = float(d.get("kaccha_amount", 0) or 0)
            pakka_total = round(billed_amt + tax_amt, 2)
            if pakka_total > 0:
                entries.append({
                    "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
                    "amount": pakka_total, "category": f"{party} (PKA)", "party_type": "BP Sale",
                    "description": f"{product} Sale #{vno} - PKA (GST Bill)",
                    "reference": f"bp_sale_pka:{doc_id}", **base
                })
            if kaccha_amt > 0:
                entries.append({
                    "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
                    "amount": round(kaccha_amt, 2), "category": f"{party} (KCA)", "party_type": "BP Sale",
                    "description": f"{product} Sale #{vno} - KCA (Slip)",
                    "reference": f"bp_sale_ka:{doc_id}", **base
                })
        else:
            entries.append({
                "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
                "amount": total, "category": party, "party_type": "BP Sale",
                "description": f"{product} Sale #{vno}",
                "reference": f"bp_sale:{doc_id}", **base
            })

    # 2. Advance from party: Party Ledger NIKASI (party ka baki kam hua) + Cash JAMA
    # Split-billing me advance hamesha Kaccha sub-ledger me jata hai (cash advance)
    if advance > 0 and party:
        adv_party = f"{party} (KCA)" if is_split else party
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": advance, "category": adv_party, "party_type": "BP Sale",
            "description": f"Advance received - {product} #{vno}",
            "reference": f"bp_sale_adv:{doc_id}", **base
        })
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "cash", "txn_type": "jama",
            "amount": advance, "category": adv_party, "party_type": "BP Sale",
            "description": f"Advance received - {product} #{vno}",
            "reference": f"bp_sale_adv_cash:{doc_id}", **base
        })

    # 3. Cash paid to truck → Cash NIKASI
    if cash > 0:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "cash", "txn_type": "nikasi",
            "amount": cash, "category": vehicle or party, "party_type": "Truck" if vehicle else "BP Sale",
            "description": f"Truck cash - {product} #{vno}",
            "reference": f"bp_sale_cash:{doc_id}", **base
        })

    # 4. Diesel → Diesel Pump Ledger JAMA (humne pump se kharida) + diesel_accounts
    if diesel > 0:
        pump_name = await _get_default_pump()
        pump_doc = await db.diesel_accounts.find_one({"pump_name": pump_name}, {"_id": 0, "pump_id": 1})
        pump_id = pump_doc.get("pump_id", "") if pump_doc else ""
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "jama",
            "amount": diesel, "category": pump_name, "party_type": "Diesel",
            "description": f"Diesel for truck - {product} #{vno} - {party}",
            "reference": f"bp_sale_diesel:{doc_id}", **base
        })
        diesel_entry = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "pump_id": pump_id, "pump_name": pump_name,
            "truck_no": vehicle, "agent_name": party,
            "amount": diesel, "txn_type": "debit",
            "description": f"Diesel for {product} #{vno} - {party}",
            "reference": f"bp_sale_diesel:{doc_id}", **base
        }
        await db.diesel_accounts.insert_one({**diesel_entry})

    # 5. Truck cash+diesel → Truck Ledger NIKASI
    if cash > 0 and vehicle:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": cash, "category": vehicle, "party_type": "Truck",
            "description": f"Truck cash deduction - {product} #{vno}",
            "reference": f"bp_truck_cash:{doc_id}", **base
        })
    if diesel > 0 and vehicle:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": diesel, "category": vehicle, "party_type": "Truck",
            "description": f"Truck diesel deduction - {product} #{vno}",
            "reference": f"bp_truck_diesel:{doc_id}", **base
        })

    # Truck payments entry
    truck_total = cash + diesel
    if truck_total > 0 and vehicle:
        truck_entry = {
            "entry_id": doc_id, "truck_no": vehicle, "date": d.get('date', ''),
            "cash_taken": cash, "diesel_taken": diesel,
            "gross_amount": 0, "deductions": truck_total,
            "net_amount": 0, "paid_amount": 0,
            "balance_amount": 0, "status": "pending",
            "source": "BP Sale", "description": f"{product} #{vno} - {party}",
            "reference": f"bp_sale_truck:{doc_id}", **base
        }
        await db.truck_payments.insert_one({**truck_entry})

    for entry in entries:
        await db.cash_transactions.insert_one({**entry})

    # Local party accounts: debit (party owes us)
    # Split-billing: 2 separate sub-ledgers — `{party} (PKA)` + `{party} (KCA)`
    # Non-split: single ledger under `{party}`
    if party and total > 0:
        if is_split:
            billed_amt = float(d.get("billed_amount", 0) or 0)
            tax_amt = float(d.get("tax_amount", 0) or 0)
            kaccha_amt = float(d.get("kaccha_amount", 0) or 0)
            pakka_total = round(billed_amt + tax_amt, 2)
            if pakka_total > 0:
                await db.local_party_accounts.insert_one({
                    "id": str(uuid.uuid4()), "date": d.get('date', ''),
                    "party_name": f"{party} (PKA)", "txn_type": "debit",
                    "amount": pakka_total,
                    "description": f"{product} Sale #{vno} - PKA (GST Bill)",
                    "source_type": "bp_sale_pka", "reference": f"bp_sale_pka:{doc_id}",
                    "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
                    "created_by": username, "created_at": now_iso
                })
            if kaccha_amt > 0:
                await db.local_party_accounts.insert_one({
                    "id": str(uuid.uuid4()), "date": d.get('date', ''),
                    "party_name": f"{party} (KCA)", "txn_type": "debit",
                    "amount": round(kaccha_amt, 2),
                    "description": f"{product} Sale #{vno} - KCA (Slip)",
                    "source_type": "bp_sale_ka", "reference": f"bp_sale_ka:{doc_id}",
                    "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
                    "created_by": username, "created_at": now_iso
                })
        else:
            await db.local_party_accounts.insert_one({
                "id": str(uuid.uuid4()), "date": d.get('date', ''),
                "party_name": party, "txn_type": "debit",
                "amount": total, "description": f"{product} Sale #{vno}",
                "source_type": "bp_sale", "reference": f"bp_sale:{doc_id}",
                "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
                "created_by": username, "created_at": now_iso
            })
    if advance > 0 and party:
        # Advance always goes to Kaccha sub-ledger when split (cash advance), else to main
        adv_party = f"{party} (KCA)" if is_split else party
        lp_adv = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "party_name": adv_party, "txn_type": "payment",
            "amount": advance, "description": f"Advance received - {product} #{vno}",
            "source_type": "bp_sale_advance", "reference": f"bp_sale_adv:{doc_id}",
            "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
            "created_by": username, "created_at": now_iso
        }
        await db.local_party_accounts.insert_one({**lp_adv})


async def _delete_bp_ledger_entries(doc_id):
    """Remove all accounting entries linked to a bp sale"""
    await db.cash_transactions.delete_many({"reference": {"$regex": f"bp_sale.*:{doc_id}"}})
    await db.cash_transactions.delete_many({"reference": {"$regex": f"bp_truck.*:{doc_id}"}})
    await db.truck_payments.delete_many({"reference": {"$regex": f"bp_sale.*:{doc_id}"}})
    await db.diesel_accounts.delete_many({"reference": {"$regex": f"bp_sale.*:{doc_id}"}})
    await db.local_party_accounts.delete_many({"reference": {"$regex": f"bp_sale.*:{doc_id}"}})
    await db.local_party_accounts.delete_many({"reference": {"$regex": f"bp_sale_adv:{doc_id}"}})


def _safe_num(v):
    try: return float(v or 0)
    except (ValueError, TypeError): return 0.0


# v104.44.56 — Payment fetching + FIFO allocation per sale (Option A + B)
async def _fetch_party_payments(party: str, kms_year: str = "", season: str = "") -> list:
    """Fetch all 'payment' txns for a party (PKA/KCA aware). Returns chronological list."""
    if not party: return []
    q = {"party_name": party, "txn_type": "payment"}
    if kms_year: q["kms_year"] = kms_year
    if season: q["season"] = season
    items = await db.local_party_accounts.find(q, {"_id": 0}).sort("date", 1).to_list(10000)
    return items


async def _enrich_sales_with_payments_fifo(sales: list, gst_filter: str = "") -> list:
    """For each sale, attach FIFO-allocated payments. Mutates and returns sales list.
    Each sale gets: payments_alloc[], total_received, last_payment_date, final_balance.
    For split-billing: PKA payments allocated to PKA portion (billed+tax), KCA payments to kaccha portion.
    """
    # Group sales by (party_with_suffix, kms_year, season)
    # For ALL view: each sale contributes BOTH PKA and KCA debits separately (need 2 buckets)
    # For PKA/KCA view: only one bucket per sale
    # We'll iterate per party-bucket and FIFO-allocate payments.
    from collections import defaultdict
    buckets = defaultdict(list)  # key=(party_suffixed, kms, season), val=list of (sale, debit_amount, bucket_type)

    for s in sales:
        party = s.get('party_name', '') or ''
        kms = s.get('kms_year', '') or ''
        ssn = s.get('season', '') or ''
        if not party: continue
        s.setdefault('payments_alloc', [])
        s['_pka_alloc'] = []; s['_kca_alloc'] = []
        view_mode = s.get('_view_mode', '')
        is_split = _safe_num(s.get('billed_amount')) > 0 and _safe_num(s.get('kaccha_amount')) > 0

        if view_mode == 'PKA':
            buckets[(f"{party} (PKA)", kms, ssn)].append((s, _safe_num(s.get('total')), 'pka'))
        elif view_mode == 'KCA':
            buckets[(f"{party} (KCA)", kms, ssn)].append((s, _safe_num(s.get('total')), 'kca'))
        else:
            # ALL view
            if is_split:
                pka_debit = _safe_num(s.get('billed_amount')) + _safe_num(s.get('tax_amount'))
                kca_debit = _safe_num(s.get('kaccha_amount'))
                if pka_debit > 0:
                    buckets[(f"{party} (PKA)", kms, ssn)].append((s, pka_debit, 'pka'))
                if kca_debit > 0:
                    buckets[(f"{party} (KCA)", kms, ssn)].append((s, kca_debit, 'kca'))
            else:
                # Non-split: party stored without suffix
                buckets[(party, kms, ssn)].append((s, _safe_num(s.get('total')), 'all'))

    # Allocate payments via FIFO per bucket
    for (party_key, kms, ssn), entries in buckets.items():
        # Sort entries by sale date (FIFO)
        entries.sort(key=lambda x: (x[0].get('date', '') or '', x[0].get('created_at', '') or ''))
        payments = await _fetch_party_payments(party_key, kms, ssn)
        # v104.44.57 — Skip premium-related ledger entries from Received calculation
        # (premium is already adjusted via Balance column; don't double-count)
        payments = [p for p in payments if not any(k in (p.get('description', '') or '').lower() for k in ('lab test premium', 'oil premium'))]
        # Each payment is allocated to oldest unpaid sale first
        remaining = [{"sale": s, "debit": debit, "remaining": debit, "btype": bt} for (s, debit, bt) in entries]
        for p in payments:
            amt = _safe_num(p.get('amount'))
            pdate = p.get('date', '') or ''
            pdesc = p.get('description', '') or ''
            for r in remaining:
                if amt <= 0: break
                if r['remaining'] <= 0: continue
                take = min(amt, r['remaining'])
                r['remaining'] = round(r['remaining'] - take, 2)
                amt = round(amt - take, 2)
                alloc_entry = {"date": pdate, "amount": take, "description": pdesc, "type": r['btype']}
                if r['btype'] == 'pka': r['sale']['_pka_alloc'].append(alloc_entry)
                elif r['btype'] == 'kca': r['sale']['_kca_alloc'].append(alloc_entry)
                else: r['sale']['payments_alloc'].append(alloc_entry)

    # Aggregate per sale
    for s in sales:
        all_pmts = list(s.get('payments_alloc', [])) + list(s.get('_pka_alloc', [])) + list(s.get('_kca_alloc', []))
        all_pmts.sort(key=lambda x: x.get('date', '') or '')
        s['payments_alloc'] = all_pmts
        s['total_received'] = round(sum(_safe_num(p.get('amount')) for p in all_pmts), 2)
        s['last_payment_date'] = all_pmts[-1]['date'] if all_pmts else ''
        # final_balance = total - received - (advance baked into balance already)
        # Use existing balance field as the source-of-truth opening, then subtract additional received
        # Actually balance was computed as total - cash_paid - diesel - advance at sale entry time
        # We'll compute net_pending = max(0, balance - subsequent received)
        existing_balance = _safe_num(s.get('balance'))
        net_pending = round(existing_balance - s['total_received'], 2)
        s['pending_balance'] = max(0, net_pending) if net_pending >= 0 else net_pending  # allow negative (overpayment)
        s.pop('_pka_alloc', None); s.pop('_kca_alloc', None)
    return sales


def _project_pakka_view(s: dict) -> dict:
    """v104.44.44 — Return a row-level pakka-only projection of entry.
    Zero out kaccha fields, recompute total = billed + tax.
    v104.44.53 — Balance for PKA view = billed + tax (no advance — advance is kaccha side)."""
    s2 = dict(s)
    billed = _safe_num(s.get("billed_amount"))
    tax = _safe_num(s.get("tax_amount"))
    s2["kaccha_weight_kg"] = 0
    s2["kaccha_weight_qtl"] = 0
    s2["kaccha_weight_qtl_display"] = ""
    s2["kaccha_amount"] = 0
    s2["kaccha_rate_per_qtl"] = 0
    # Net weight + amounts → only pakka portion
    s2["net_weight_kg"] = _safe_num(s.get("billed_weight_kg"))
    s2["net_weight_qtl"] = _safe_num(s.get("billed_weight_qtl"))
    s2["net_weight_qtl_display"] = s.get("billed_weight_qtl_display", "")
    s2["amount"] = billed
    s2["total"] = billed + tax
    s2["balance"] = billed + tax
    s2["advance"] = 0
    s2["_view_mode"] = "PKA"
    return s2


def _project_kaccha_view(s: dict) -> dict:
    """v104.44.44 — Return a row-level kaccha-only projection of entry.
    Zero out pakka/GST fields, recompute total = kaccha amount.
    v104.44.53 — Balance for KCA view = kaccha - advance (advance always on kaccha side)."""
    s2 = dict(s)
    kac = _safe_num(s.get("kaccha_amount"))
    adv = _safe_num(s.get("advance"))
    s2["billed_weight_kg"] = 0
    s2["billed_weight_qtl"] = 0
    s2["billed_weight_qtl_display"] = ""
    s2["billed_amount"] = 0
    s2["gst_type"] = "none"
    s2["gst_percent"] = 0
    s2["tax_amount"] = 0
    # Use kaccha rate as primary rate
    if _safe_num(s.get("kaccha_rate_per_qtl")) > 0:
        s2["rate_per_qtl"] = s.get("kaccha_rate_per_qtl")
    s2["net_weight_kg"] = _safe_num(s.get("kaccha_weight_kg"))
    s2["net_weight_qtl"] = _safe_num(s.get("kaccha_weight_qtl"))
    s2["net_weight_qtl_display"] = s.get("kaccha_weight_qtl_display", "")
    s2["amount"] = kac
    s2["total"] = kac
    s2["balance"] = max(0, kac - adv)
    s2["_view_mode"] = "KCA"
    return s2


@router.get("/bp-sale-register")
async def get_bp_sales(product: str = "", kms_year: str = "", season: str = "",
                       gst_filter: Optional[str] = None):
    query = {}
    if product: query["product"] = product
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    sales = await db.bp_sale_register.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    # v104.44.44 — Row-level split: PKA shows only pakka portion, KCA only kaccha portion
    if gst_filter == "PKA":
        # Include entries with any pakka portion (full pakka OR split). Mask kaccha fields.
        sales = [_project_pakka_view(s) for s in sales if (
            _safe_num(s.get("billed_amount")) > 0 or _safe_num(s.get("gst_percent")) > 0
        )]
    elif gst_filter == "KCA":
        # Include entries with any kaccha portion (pure kaccha OR split). Mask pakka fields.
        sales = [_project_kaccha_view(s) for s in sales if (
            _safe_num(s.get("kaccha_amount")) > 0
            or (_safe_num(s.get("billed_amount")) == 0 and _safe_num(s.get("gst_percent")) == 0)
        )]
    return sales


# v104.44.56 — Sales enriched with payment allocation (Option A + B)
@router.get("/bp-sale-register/with-payments")
async def get_bp_sales_with_payments(product: str = "", kms_year: str = "", season: str = "",
                                     gst_filter: Optional[str] = None):
    sales = await get_bp_sales(product=product, kms_year=kms_year, season=season, gst_filter=gst_filter)
    sales = await _enrich_sales_with_payments_fifo(sales, gst_filter or "")
    return sales


# v104.44.56 — Party Statement: chronological ledger (Option C)
@router.get("/bp-sale-register/party-statement")
async def get_bp_party_statement(party: str, kms_year: str = "", season: str = "",
                                  gst_filter: Optional[str] = None):
    """Returns chronological ledger entries for a party: sales (debit) + payments + premium adjustments.
    gst_filter: 'PKA' → only `{party} (PKA)` ledger; 'KCA' → `(KCA)`; else → all (PKA + KCA + non-split combined).
    """
    if not party: return {"party": "", "entries": [], "summary": {}}
    # Determine which party_name suffixes to include
    if gst_filter == "PKA":
        party_keys = [f"{party} (PKA)"]
    elif gst_filter == "KCA":
        party_keys = [f"{party} (KCA)"]
    else:
        party_keys = [f"{party} (PKA)", f"{party} (KCA)", party]

    q = {"party_name": {"$in": party_keys}}
    if kms_year: q["kms_year"] = kms_year
    if season: q["season"] = season
    raw = await db.local_party_accounts.find(q, {"_id": 0}).to_list(20000)
    raw.sort(key=lambda x: (x.get('date', '') or '', x.get('created_at', '') or ''))
    # Note: Oil premium adjustments are already auto-created as payment entries in
    # local_party_accounts by the oil_premium flow (description "Lab Test Premium..."),
    # so we do NOT re-fetch them here to avoid double-counting.

    # Compute running balance
    balance = 0.0
    entries = []
    for r in raw:
        amt = _safe_num(r.get('amount'))
        ttype = r.get('txn_type', '') or ''
        # Normalize description text to use PKA/KCA only (legacy entries used Pakka/Kaccha)
        desc = (r.get('description', '') or '').replace('Pakka (GST Bill)', 'PKA (GST Bill)').replace('Kaccha (Slip)', 'KCA (Slip)').replace(' - Pakka', ' - PKA').replace(' - Kaccha', ' - KCA')
        # debit increases balance (party owes), payment decreases
        if ttype == 'debit':
            balance += amt; flow = 'Dr'
        elif ttype == 'payment':
            balance -= amt; flow = 'Cr'
        else:
            flow = ttype.upper()
        entries.append({
            "date": r.get('date', '') or '',
            "party_name": r.get('party_name', '') or '',
            "txn_type": ttype, "flow": flow,
            "amount": round(amt, 2),
            "description": desc,
            "reference": r.get('reference', '') or '',
            "running_balance": round(balance, 2),
        })

    summary = {
        "party": party,
        "total_debit": round(sum(_safe_num(e.get('amount')) for e in entries if e.get('flow') == 'Dr'), 2),
        "total_credit": round(sum(_safe_num(e.get('amount')) for e in entries if e.get('flow') == 'Cr'), 2),
        "closing_balance": round(balance, 2),
        "entry_count": len(entries),
    }
    return {"party": party, "gst_filter": gst_filter or "ALL", "kms_year": kms_year, "season": season,
            "entries": entries, "summary": summary}


def _compute_amounts_and_tax(data: dict) -> None:
    """Mutates data with computed fields. Matches desktop/local-server split-billing logic exactly.
    Note: `sauda_amount` is informational only — never used in any calculation.
    """
    rate = float(data.get("rate_per_qtl", 0) or 0)
    raw_kaccha_rate = data.get("kaccha_rate_per_qtl")
    kaccha_rate = (float(raw_kaccha_rate) if raw_kaccha_rate not in (None, "", 0, "0") else rate)
    is_split = bool(data.get("split_billing"))

    if is_split:
        billed_kg = float(data.get("billed_weight_kg", 0) or 0)
        kaccha_kg = float(data.get("kaccha_weight_kg", 0) or 0)
        billed_qtl = round(billed_kg / 100, 4)
        kaccha_qtl = round(kaccha_kg / 100, 4)
        billed_amt = round(billed_qtl * rate, 2)
        kaccha_amt = round(kaccha_qtl * kaccha_rate, 2)
        data["net_weight_kg"] = round(billed_kg + kaccha_kg, 3)  # sum for physical dispatch
        data["net_weight_qtl"] = round(billed_qtl + kaccha_qtl, 4)
        data["billed_weight_qtl"] = billed_qtl
        data["kaccha_weight_qtl"] = kaccha_qtl
        data["billed_amount"] = billed_amt
        data["kaccha_amount"] = kaccha_amt
        data["kaccha_rate_per_qtl"] = kaccha_rate
        data["amount"] = billed_amt  # GST-taxable portion (kept under same key for register compat)
        gst_pct = float(data.get("gst_percent") or 0)
        tax_amt = round(billed_amt * gst_pct / 100, 2) if gst_pct else 0
        data["tax_amount"] = tax_amt
        data["total"] = round(billed_amt + tax_amt + kaccha_amt, 2)
    else:
        nw = float(data.get("net_weight_kg", 0) or 0)
        nw_qtl = round(nw / 100, 4)
        amount = round(nw_qtl * rate, 2)
        data["net_weight_qtl"] = nw_qtl
        data["amount"] = amount
        # Clear split fields if toggled off (or never on)
        data["billed_weight_kg"] = 0
        data["billed_weight_qtl"] = 0
        data["billed_amount"] = 0
        data["kaccha_weight_kg"] = 0
        data["kaccha_weight_qtl"] = 0
        data["kaccha_amount"] = 0
        data["kaccha_rate_per_qtl"] = 0
        gst_pct = float(data.get("gst_percent") or 0)
        tax_amt = round(amount * gst_pct / 100, 2) if gst_pct else 0
        data["tax_amount"] = tax_amt
        data["total"] = round(amount + tax_amt, 2)


@router.post("/bp-sale-register")
async def create_bp_sale(data: dict, username: str = "", role: str = ""):
    data["id"] = str(uuid.uuid4())[:12]
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_at"] = data["created_at"]
    data["created_by"] = username

    # Auto-generate voucher_no if empty (format: S-001, S-002, ...).
    # User-entered voucher_no preserved as-is.
    if not (data.get("voucher_no") or "").strip():
        import re
        max_n = 0
        cursor = db.bp_sale_register.find({"voucher_no": {"$regex": r"^S-\d+$"}}, {"_id": 0, "voucher_no": 1})
        async for doc in cursor:
            m = re.match(r"^S-(\d+)$", doc.get("voucher_no") or "")
            if m:
                n = int(m.group(1))
                if n > max_n:
                    max_n = n
        data["voucher_no"] = f"S-{max_n + 1:03d}"

    _compute_amounts_and_tax(data)

    cash = float(data.get("cash_paid", 0) or 0)
    diesel = float(data.get("diesel_paid", 0) or 0)
    advance = float(data.get("advance", 0) or 0)
    data["cash_paid"] = cash
    data["diesel_paid"] = diesel
    data["advance"] = advance
    data["balance"] = round(data["total"] - advance, 2)

    await db.bp_sale_register.insert_one({**data})
    data.pop("_id", None)
    await _create_bp_ledger_entries(data, data["id"], username)
    return data


@router.put("/bp-sale-register/{sale_id}")
async def update_bp_sale(sale_id: str, data: dict, username: str = "", role: str = ""):
    from services.edit_lock import check_edit_lock
    existing = await db.bp_sale_register.find_one({"id": sale_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Sale not found")
    existing_clean = {k: v for k, v in existing.items() if k != "_id"}
    can_edit, message = await check_edit_lock(existing_clean, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)

    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = username

    _compute_amounts_and_tax(data)

    cash = float(data.get("cash_paid", 0) or 0)
    diesel = float(data.get("diesel_paid", 0) or 0)
    advance = float(data.get("advance", 0) or 0)
    data["cash_paid"] = cash
    data["diesel_paid"] = diesel
    data["advance"] = advance
    data["balance"] = round(data["total"] - advance, 2)

    data.pop("id", None)
    data.pop("_id", None)
    await db.bp_sale_register.update_one({"id": sale_id}, {"$set": data})
    # Re-create accounting entries
    await _delete_bp_ledger_entries(sale_id)
    data["id"] = sale_id
    await _create_bp_ledger_entries(data, sale_id, username)
    return {"success": True}


@router.delete("/bp-sale-register/{sale_id}")
async def delete_bp_sale(sale_id: str, username: str = "", role: str = ""):
    from services.edit_lock import check_edit_lock
    existing = await db.bp_sale_register.find_one({"id": sale_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Sale not found")
    can_edit, message = await check_edit_lock(existing, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    await db.bp_sale_register.delete_one({"id": sale_id})
    await _delete_bp_ledger_entries(sale_id)
    return {"success": True}


@router.get("/bp-sale-register/suggestions/bill-from")
async def get_bill_from_suggestions():
    pipeline = [{"$group": {"_id": "$bill_from"}}, {"$sort": {"_id": 1}}]
    results = await db.bp_sale_register.aggregate(pipeline).to_list(500)
    return [r["_id"] for r in results if r["_id"]]


@router.get("/bp-sale-register/next-voucher-no")
async def next_bp_voucher_no():
    """Generate next sequential voucher number in `S-001` format.
    Scans all existing BP sale voucher_no values matching `^S-\\d+$`,
    takes the max numeric suffix, returns MAX+1 zero-padded to 3 digits.
    Non-S-prefixed custom voucher numbers are preserved and ignored.
    """
    import re
    cursor = db.bp_sale_register.find({"voucher_no": {"$regex": r"^S-\d+$"}}, {"_id": 0, "voucher_no": 1})
    max_n = 0
    async for doc in cursor:
        m = re.match(r"^S-(\d+)$", doc.get("voucher_no") or "")
        if m:
            n = int(m.group(1))
            if n > max_n:
                max_n = n
    return {"voucher_no": f"S-{max_n + 1:03d}"}


@router.get("/bp-sale-register/suggestions/party-name")
async def get_party_suggestions():
    pipeline = [{"$group": {"_id": "$party_name"}}, {"$sort": {"_id": 1}}]
    results = await db.bp_sale_register.aggregate(pipeline).to_list(500)
    return [r["_id"] for r in results if r["_id"]]


@router.get("/bp-sale-register/suggestions/destination")
async def get_destination_suggestions():
    pipeline = [{"$group": {"_id": "$destination"}}, {"$sort": {"_id": 1}}]
    results = await db.bp_sale_register.aggregate(pipeline).to_list(500)
    return [r["_id"] for r in results if r["_id"]]


@router.get("/bp-sale-register/export/excel")
async def export_bp_sales_excel(product: str = "", kms_year: str = "", season: str = "",
    date_from: str = "", date_to: str = "", billing_date_from: str = "", billing_date_to: str = "",
    rst_no: str = "", vehicle_no: str = "", bill_from: str = "", party_name: str = "", destination: str = "",
    gst_filter: Optional[str] = None):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response
    from utils.export_helpers import COLORS

    query = {}
    if product: query["product"] = product
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if rst_no: query["rst_no"] = {"$regex": rst_no, "$options": "i"}
    if vehicle_no: query["vehicle_no"] = {"$regex": vehicle_no, "$options": "i"}
    if bill_from: query["bill_from"] = {"$regex": bill_from, "$options": "i"}
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if destination: query["destination"] = {"$regex": destination, "$options": "i"}
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    if billing_date_from or billing_date_to:
        query["billing_date"] = {}
        if billing_date_from: query["billing_date"]["$gte"] = billing_date_from
        if billing_date_to: query["billing_date"]["$lte"] = billing_date_to
    sales = await db.bp_sale_register.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    # v104.44.44 — Row-level PKA/KCA: split entries appear with only relevant portion
    if gst_filter == "PKA":
        sales = [_project_pakka_view(s) for s in sales if (_safe_num(s.get("billed_amount")) > 0 or _safe_num(s.get("gst_percent")) > 0)]
    elif gst_filter == "KCA":
        sales = [_project_kaccha_view(s) for s in sales if (_safe_num(s.get("kaccha_amount")) > 0 or (_safe_num(s.get("billed_amount")) == 0 and _safe_num(s.get("gst_percent")) == 0))]

    # v104.44.56 — Enrich with FIFO-allocated payments (payment columns auto-show if any payment exists)
    sales = await _enrich_sales_with_payments_fifo(sales, gst_filter or "")
    has_payments = any((s.get('total_received') or 0) > 0 for s in sales)

    # v104.44.51 — Detect if any split entries exist (so we add Pakka/Kaccha breakdown cols)
    has_split = any(_safe_num(s.get('billed_amount')) > 0 and _safe_num(s.get('kaccha_amount')) > 0 for s in sales)
    show_pakka_col = has_split and gst_filter not in ("PKA", "KCA")
    show_kaccha_col = has_split and gst_filter not in ("PKA", "KCA")

    # Fetch oil premium data for Rice Bran
    oil_map = {}
    if product == "Rice Bran":
        op_query = {}
        if kms_year: op_query["kms_year"] = kms_year
        if season: op_query["season"] = season
        op_items = await db.oil_premium.find(op_query, {"_id": 0}).to_list(10000)
        for op in op_items:
            key = op.get("voucher_no") or op.get("rst_no") or ""
            if key: oil_map[key] = op
    has_oil = bool(oil_map) and any(oil_map.get(s.get('voucher_no') or '') or oil_map.get(s.get('rst_no') or '') for s in sales)

    # Branding
    branding = await db.branding.find_one({}, {"_id": 0}) or {}
    company = branding.get("company_name", "Rice Mill")
    address = branding.get("address", "")
    phone = branding.get("phone", "")

    wb = Workbook(); ws = wb.active
    ws.title = f"{product or 'By-Product'} Sales"

    thin = Side(style='thin', color='B0C4DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="F0F6FC", end_color="F0F6FC", fill_type="solid")
    total_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    # v104.44.51 — Color-coded fills/fonts for Pakka/Kaccha/Tax
    pakka_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # light green
    kaccha_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # light pink
    tax_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")  # light amber
    pakka_font = Font(size=9, color="2E7D32", bold=True)
    kaccha_font = Font(size=9, color="C62828", bold=True)
    tax_font = Font(size=9, color="EF6C00", bold=True)
    total_amt_font = Font(size=9, color="1B5E20", bold=True)

    # Title with mode badge
    mode_label = f" [{gst_filter}]" if gst_filter in ("PKA", "KCA") else " [ALL]"
    title = f"{product or 'By-Product'} Sale Register{mode_label}"
    if kms_year: title += f" - FY {kms_year}"
    if season: title += f" ({season})"

    # Detect which optional columns have any data across ALL sales
    has_bill_no = any(s.get('bill_number') for s in sales)
    has_billing_date = any(s.get('billing_date') for s in sales)
    has_rst = any(s.get('rst_no') for s in sales)
    has_vehicle = any(s.get('vehicle_no') for s in sales)
    has_bill_from = any(s.get('bill_from') for s in sales)
    has_dest = any(s.get('destination') for s in sales)
    has_bags = any(s.get('bags', 0) for s in sales)
    has_tax = any(s.get('tax_amount', 0) for s in sales)
    has_cash = any(s.get('cash_paid', 0) for s in sales)
    has_diesel = any(s.get('diesel_paid', 0) for s in sales)
    has_adv = any(s.get('advance', 0) for s in sales)
    has_remark = any(s.get('remark') for s in sales)

    # Build dynamic headers and column config
    cols = [('V.No', 8, 'voucher_no')]
    cols.append(('Date', 10, 'date'))
    if has_bill_no: cols.append(('Bill No', 10, 'bill_number'))
    if has_billing_date: cols.append(('Billing Date', 10, 'billing_date'))
    if has_rst: cols.append(('RST No', 8, 'rst_no'))
    if has_vehicle: cols.append(('Vehicle No', 12, 'vehicle_no'))
    if has_bill_from: cols.append(('Bill From', 14, 'bill_from'))
    cols.append(('Party Name', 16, 'party_name'))
    if has_dest: cols.append(('Destination', 14, 'destination'))
    cols.append(('N/W (Qtl)', 9, 'net_weight_qtl'))
    if has_bags: cols.append(('Bags', 7, 'bags'))
    cols.append(('Rate/Qtl', 9, 'rate_per_qtl'))
    if show_pakka_col: cols.append(('PKA Amt', 12, 'billed_amount'))
    if show_kaccha_col: cols.append(('KCA Amt', 12, 'kaccha_amount'))
    if not show_pakka_col and not show_kaccha_col:
        cols.append(('Amount', 12, 'amount'))
    if has_tax: cols.append(('Tax', 9, 'tax_amount'))
    cols.append(('Total', 12, 'total'))
    if has_cash: cols.append(('Cash', 10, 'cash_paid'))
    if has_diesel: cols.append(('Diesel', 10, 'diesel_paid'))
    if has_adv: cols.append(('Advance', 10, 'advance'))
    # v104.44.52 — PKA mode me Balance + Oil columns hide (not relevant for billed-only view)
    # v104.44.53 — Balance ko Premium ke baad (last) move kiya, premium-adjusted
    if gst_filter != "PKA":
        if has_oil:
            cols.append(('Oil%', 8, 'oil_pct'))
            cols.append(('Diff%', 8, 'oil_diff'))
            cols.append(('Premium', 12, 'oil_premium'))
        cols.append(('Balance', 12, 'balance_final'))
        # v104.44.56 — Payment summary columns (only when at least one sale has received payments)
        if has_payments:
            cols.append(('Last Pmt', 10, 'last_payment_date'))
            cols.append(('Received', 12, 'total_received'))
            cols.append(('Pending', 12, 'pending_balance'))
    if has_remark: cols.append(('Remark', 16, 'remark'))

    headers = [c[0] for c in cols]
    widths = [c[1] for c in cols]
    keys = [c[2] for c in cols]
    ncols = len(headers)

    # v104.44.51 — Build filter summary subtitle
    flt_parts = []
    if date_from or date_to:
        flt_parts.append(f"Period: {date_from or '...'} to {date_to or '...'}")
    if party_name: flt_parts.append(f"Party: {party_name}")
    if vehicle_no: flt_parts.append(f"Vehicle: {vehicle_no}")
    if bill_from: flt_parts.append(f"Bill From: {bill_from}")
    if destination: flt_parts.append(f"Destination: {destination}")
    if rst_no: flt_parts.append(f"RST: {rst_no}")
    filter_summary = "  •  ".join(flt_parts) if flt_parts else ""

    row = 6 if filter_summary else 5
    # Merge header rows to match ncols
    last_col_letter = get_column_letter(ncols)
    ws.merge_cells(f'A1:{last_col_letter}1')
    c1 = ws.cell(row=1, column=1, value=company.upper())
    c1.font = Font(bold=True, size=14, color="1F4E79"); c1.alignment = Alignment(horizontal='center')
    if address:
        ws.merge_cells(f'A2:{last_col_letter}2')
        c2 = ws.cell(row=2, column=1, value=f"{address}  |  {phone}")
        c2.font = Font(size=9, color="666666"); c2.alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A3:{last_col_letter}3')
    c3 = ws.cell(row=3, column=1, value=title)
    c3.font = Font(bold=True, size=12, color="FFFFFF")
    # Mode color: PKA=emerald, KCA=rose, ALL=blue
    if gst_filter == "PKA":
        c3.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    elif gst_filter == "KCA":
        c3.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    else:
        c3.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    c3.alignment = Alignment(horizontal='center')

    # Filter summary subtitle row
    if filter_summary:
        ws.merge_cells(f'A4:{last_col_letter}4')
        c4 = ws.cell(row=4, column=1, value=filter_summary)
        c4.font = Font(italic=True, size=9, color="555555")
        c4.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        c4.alignment = Alignment(horizontal='center')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    # Data rows
    t_nw = t_bags = t_amount = t_billed = t_kaccha = t_tax = t_total = t_cash = t_diesel = t_adv = t_bal = t_bal_final = t_oil_premium = t_received = t_pending = 0
    for idx, s in enumerate(sales):
        r = row + 1 + idx
        fill = alt_fill if idx % 2 == 0 else None
        op = oil_map.get(s.get('voucher_no') or '') or oil_map.get(s.get('rst_no') or '')
        prem = (op.get('premium_amount', 0) or 0) if op else 0
        # v104.44.53 — balance_final = balance + premium (premium adjusted balance)
        bal_final = round((s.get('balance', 0) or 0) + prem, 2)
        t_nw += s.get('net_weight_kg', 0); t_bags += s.get('bags', 0)
        t_amount += s.get('amount', 0); t_billed += s.get('billed_amount', 0); t_kaccha += s.get('kaccha_amount', 0)
        t_tax += s.get('tax_amount', 0); t_total += s.get('total', 0)
        t_cash += s.get('cash_paid', 0); t_diesel += s.get('diesel_paid', 0)
        t_adv += s.get('advance', 0); t_bal += s.get('balance', 0)
        t_bal_final += bal_final
        if op: t_oil_premium += prem
        # v104.44.56 — accumulate payment totals
        recv = s.get('total_received', 0) or 0
        pend = s.get('pending_balance', 0) or 0
        t_received += recv; t_pending += pend
        for col_idx, key in enumerate(keys, 1):
            if key == 'voucher_no': val = s.get('voucher_no', '') or ''
            elif key == 'date': val = fmt_date(s.get('date', ''))
            elif key == 'billing_date': val = fmt_date(s.get('billing_date', ''))
            elif key == 'net_weight_qtl': val = round((s.get('net_weight_qtl', 0) or s.get('net_weight_kg', 0)/100), 2)
            elif key == 'oil_pct': val = op.get('actual_oil_pct', '') if op else ''
            elif key == 'oil_diff': val = round(op.get('difference_pct', 0), 2) if op else ''
            elif key == 'oil_premium': val = round(prem, 2) if op else ''
            elif key == 'balance_final': val = bal_final
            elif key == 'last_payment_date': val = fmt_date(s.get('last_payment_date', '')) if s.get('last_payment_date') else ''
            elif key == 'total_received': val = recv if recv > 0 else ''
            elif key == 'pending_balance': val = pend
            else: val = s.get(key, 0) if key in ('net_weight_kg','bags','rate_per_qtl','amount','billed_amount','kaccha_amount','tax_amount','total','cash_paid','diesel_paid','advance','balance') else s.get(key, '')
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.border = border
            if fill: cell.fill = fill
            # v104.44.51 — Color-coded for Pakka/Kaccha/Tax/Total cells
            if key == 'billed_amount':
                cell.fill = pakka_fill; cell.font = pakka_font
            elif key == 'kaccha_amount':
                cell.fill = kaccha_fill; cell.font = kaccha_font
            elif key == 'tax_amount' and val:
                cell.fill = tax_fill; cell.font = tax_font
            elif key == 'total':
                cell.font = total_amt_font
            elif key == 'balance_final':
                # red if positive (party owes), green if zero/negative
                cell.font = Font(size=9, bold=True, color=("C62828" if val > 0 else "1B5E20"))
            elif key == 'total_received' and val:
                cell.fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
                cell.font = Font(size=9, bold=True, color="00838F")
            elif key == 'pending_balance':
                cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                cell.font = Font(size=9, bold=True, color=("E65100" if val > 0 else "1B5E20"))
            if key in ('net_weight_kg','net_weight_qtl','bags','rate_per_qtl','amount','billed_amount','kaccha_amount','tax_amount','total','cash_paid','diesel_paid','advance','balance','balance_final','total_received','pending_balance','oil_pct','oil_diff','oil_premium'):
                cell.alignment = Alignment(horizontal='right')
            if key in ('amount','billed_amount','kaccha_amount','tax_amount','total','cash_paid','diesel_paid','advance','balance','balance_final','total_received','pending_balance','oil_premium'):
                cell.number_format = '#,##0.00'
            if key == 'oil_premium' and op and prem < 0:
                cell.font = Font(size=9, color="FF0000")

    # Total row
    tr = row + 1 + len(sales)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=tr, column=col_idx)
        cell.border = border; cell.fill = total_fill; cell.font = Font(bold=True, size=9, color="FFFFFF")
    # Set total values in correct columns
    for col_idx, key in enumerate(keys, 1):
        if key == 'date': ws.cell(row=tr, column=col_idx, value="TOTAL")
        elif key == 'net_weight_kg': ws.cell(row=tr, column=col_idx, value=round(t_nw, 2)).alignment = Alignment(horizontal='right')
        elif key == 'net_weight_qtl': ws.cell(row=tr, column=col_idx, value=round(t_nw/100, 2)).alignment = Alignment(horizontal='right')
        elif key == 'bags': ws.cell(row=tr, column=col_idx, value=t_bags).alignment = Alignment(horizontal='right')
        elif key == 'amount': ws.cell(row=tr, column=col_idx, value=round(t_amount, 2)).alignment = Alignment(horizontal='right')
        elif key == 'billed_amount': ws.cell(row=tr, column=col_idx, value=round(t_billed, 2)).alignment = Alignment(horizontal='right')
        elif key == 'kaccha_amount': ws.cell(row=tr, column=col_idx, value=round(t_kaccha, 2)).alignment = Alignment(horizontal='right')
        elif key == 'tax_amount': ws.cell(row=tr, column=col_idx, value=round(t_tax, 2)).alignment = Alignment(horizontal='right')
        elif key == 'total': ws.cell(row=tr, column=col_idx, value=round(t_total, 2)).alignment = Alignment(horizontal='right')
        elif key == 'cash_paid': ws.cell(row=tr, column=col_idx, value=round(t_cash, 2)).alignment = Alignment(horizontal='right')
        elif key == 'diesel_paid': ws.cell(row=tr, column=col_idx, value=round(t_diesel, 2)).alignment = Alignment(horizontal='right')
        elif key == 'advance': ws.cell(row=tr, column=col_idx, value=round(t_adv, 2)).alignment = Alignment(horizontal='right')
        elif key == 'balance': ws.cell(row=tr, column=col_idx, value=round(t_bal, 2)).alignment = Alignment(horizontal='right')
        elif key == 'balance_final': ws.cell(row=tr, column=col_idx, value=round(t_bal_final, 2)).alignment = Alignment(horizontal='right')
        elif key == 'total_received': ws.cell(row=tr, column=col_idx, value=round(t_received, 2)).alignment = Alignment(horizontal='right')
        elif key == 'pending_balance': ws.cell(row=tr, column=col_idx, value=round(t_pending, 2)).alignment = Alignment(horizontal='right')
        elif key == 'oil_premium':
            c = ws.cell(row=tr, column=col_idx, value=round(t_oil_premium, 2))
            c.alignment = Alignment(horizontal='right')
            c.number_format = '#,##0.00'

    # Column widths
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 🎯 v104.44.9 — Apply consolidated polish (BP sale register)
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    fn = f"{(product or 'byproduct').lower().replace(' ','_')}_sale_register_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/bp-sale-register/export/pdf")
async def export_bp_sales_pdf(product: str = "", kms_year: str = "", season: str = "",
    date_from: str = "", date_to: str = "", billing_date_from: str = "", billing_date_to: str = "",
    rst_no: str = "", vehicle_no: str = "", bill_from: str = "", party_name: str = "", destination: str = "",
    gst_filter: Optional[str] = None):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from fastapi.responses import Response

    query = {}
    if product: query["product"] = product
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if rst_no: query["rst_no"] = {"$regex": rst_no, "$options": "i"}
    if vehicle_no: query["vehicle_no"] = {"$regex": vehicle_no, "$options": "i"}
    if bill_from: query["bill_from"] = {"$regex": bill_from, "$options": "i"}
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if destination: query["destination"] = {"$regex": destination, "$options": "i"}
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    if billing_date_from or billing_date_to:
        query["billing_date"] = {}
        if billing_date_from: query["billing_date"]["$gte"] = billing_date_from
        if billing_date_to: query["billing_date"]["$lte"] = billing_date_to
    sales = await db.bp_sale_register.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    # v104.44.44 — Row-level PKA/KCA
    if gst_filter == "PKA":
        sales = [_project_pakka_view(s) for s in sales if (_safe_num(s.get("billed_amount")) > 0 or _safe_num(s.get("gst_percent")) > 0)]
    elif gst_filter == "KCA":
        sales = [_project_kaccha_view(s) for s in sales if (_safe_num(s.get("kaccha_amount")) > 0 or (_safe_num(s.get("billed_amount")) == 0 and _safe_num(s.get("gst_percent")) == 0))]

    # v104.44.56 — Enrich with FIFO-allocated payments
    sales = await _enrich_sales_with_payments_fifo(sales, gst_filter or "")
    has_payments = any((s.get('total_received') or 0) > 0 for s in sales)

    # Fetch oil premium data for Rice Bran
    oil_map_pdf = {}
    if product == "Rice Bran":
        op_q = {}
        if kms_year: op_q["kms_year"] = kms_year
        if season: op_q["season"] = season
        op_list = await db.oil_premium.find(op_q, {"_id": 0}).to_list(10000)
        for op in op_list:
            key = op.get("voucher_no") or op.get("rst_no") or ""
            if key: oil_map_pdf[key] = op
    has_oil_pdf = bool(oil_map_pdf) and any(oil_map_pdf.get(s.get('voucher_no') or '') or oil_map_pdf.get(s.get('rst_no') or '') for s in sales)

    branding = await db.branding.find_one({}, {"_id": 0}) or {}

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=12, rightMargin=12, topMargin=15, bottomMargin=15)
    elements = []
    styles = getSampleStyleSheet()

    # Use shared branded header (company + address + phone + custom_fields like proprietor, GST, etc.)
    from utils.export_helpers import get_pdf_company_header
    elements.extend(get_pdf_company_header(branding))

    # v104.44.51 — Title with mode badge (PKA emerald / KCA rose / ALL blue)
    mode_label = f" [{gst_filter}]" if gst_filter in ("PKA", "KCA") else " [ALL]"
    title = f"{product or 'By-Product'} Sale Register{mode_label}"
    if kms_year: title += f" - FY {kms_year}"
    if season: title += f" ({season})"
    if gst_filter == "PKA":
        title_bg = colors.HexColor('#2E7D32')
    elif gst_filter == "KCA":
        title_bg = colors.HexColor('#C62828')
    else:
        title_bg = colors.HexColor('#2E75B6')
    title_style = ParagraphStyle('RegTitle', parent=styles['Heading2'], fontSize=10,
        textColor=colors.white, backColor=title_bg, spaceAfter=4, alignment=1,
        borderPadding=(3, 3, 3, 3))
    elements.append(Paragraph(title, title_style))

    # v104.44.51 — Filter summary subtitle
    flt_parts = []
    if date_from or date_to:
        flt_parts.append(f"Period: {date_from or '...'} to {date_to or '...'}")
    if party_name: flt_parts.append(f"Party: {party_name}")
    if vehicle_no: flt_parts.append(f"Vehicle: {vehicle_no}")
    if bill_from: flt_parts.append(f"Bill From: {bill_from}")
    if destination: flt_parts.append(f"Destination: {destination}")
    if rst_no: flt_parts.append(f"RST: {rst_no}")
    filter_summary = "  •  ".join(flt_parts)
    if filter_summary:
        sub_style = ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=7,
            textColor=colors.HexColor('#555555'), backColor=colors.HexColor('#F5F5F5'),
            alignment=1, borderPadding=(2, 2, 2, 2), spaceAfter=3)
        elements.append(Paragraph(f"<i>{filter_summary}</i>", sub_style))
    elements.append(Spacer(1, 3))

    # Detect which optional columns have data
    has_bill_no = any(s.get('bill_number') for s in sales)
    has_rst = any(s.get('rst_no') for s in sales)
    has_vehicle = any(s.get('vehicle_no') for s in sales)
    has_bill_from = any(s.get('bill_from') for s in sales)
    has_dest = any(s.get('destination') for s in sales)
    has_bags = any(s.get('bags', 0) for s in sales)
    has_tax = any(s.get('tax_amount', 0) for s in sales)
    has_cash = any(s.get('cash_paid', 0) for s in sales)
    has_diesel = any(s.get('diesel_paid', 0) for s in sales)
    has_adv = any(s.get('advance', 0) for s in sales)

    # v104.44.51 — Detect split entries (Pakka+Kaccha breakdown columns in ALL view)
    has_split = any(_safe_num(s.get('billed_amount')) > 0 and _safe_num(s.get('kaccha_amount')) > 0 for s in sales)
    show_pakka_col = has_split and gst_filter not in ("PKA", "KCA")
    show_kaccha_col = has_split and gst_filter not in ("PKA", "KCA")

    # Build dynamic columns: (header, width, key)
    pdf_cols = [('V.No', 28, 'voucher_no'), ('Date', 42, 'date')]
    if has_bill_no: pdf_cols.append(('Bill No', 40, 'bill_number'))
    if has_rst: pdf_cols.append(('RST', 28, 'rst_no'))
    if has_vehicle: pdf_cols.append(('Vehicle', 48, 'vehicle_no'))
    if has_bill_from: pdf_cols.append(('Bill From', 55, 'bill_from'))
    pdf_cols.append(('Party', 65, 'party_name'))
    if has_dest: pdf_cols.append(('Destination', 50, 'destination'))
    pdf_cols.append(('N/W(Qtl)', 40, 'net_weight_qtl'))
    if has_bags: pdf_cols.append(('Bags', 28, 'bags'))
    pdf_cols.append(('Rate/Q', 38, 'rate_per_qtl'))
    if show_pakka_col: pdf_cols.append(('PKA', 50, 'billed_amount'))
    if show_kaccha_col: pdf_cols.append(('KCA', 50, 'kaccha_amount'))
    if not show_pakka_col and not show_kaccha_col:
        pdf_cols.append(('Amount', 50, 'amount'))
    if has_tax: pdf_cols.append(('Tax', 35, 'tax_amount'))
    pdf_cols.append(('Total', 50, 'total'))
    if has_cash: pdf_cols.append(('Cash', 38, 'cash_paid'))
    if has_diesel: pdf_cols.append(('Diesel', 38, 'diesel_paid'))
    if has_adv: pdf_cols.append(('Adv', 32, 'advance'))
    # v104.44.52 — PKA mode me Balance + Oil columns hide
    # v104.44.53 — Balance ko Premium ke baad (last) move kiya, premium-adjusted
    if gst_filter != "PKA":
        if has_oil_pdf:
            pdf_cols.append(('Oil%', 30, 'oil_pct'))
            pdf_cols.append(('Diff%', 30, 'oil_diff'))
            pdf_cols.append(('Premium', 45, 'oil_premium'))
        pdf_cols.append(('Balance', 50, 'balance_final'))
        # v104.44.56 — payment columns
        if has_payments:
            pdf_cols.append(('Last Pmt', 42, 'last_payment_date'))
            pdf_cols.append(('Recvd', 45, 'total_received'))
            pdf_cols.append(('Pending', 50, 'pending_balance'))

    headers = [c[0] for c in pdf_cols]
    col_widths = [c[1] for c in pdf_cols]
    col_keys = [c[2] for c in pdf_cols]

    # Auto-fit: scale columns to fit A4 landscape (842pt - 24pt margins = 818pt)
    usable_width = 818
    total_w = sum(col_widths)
    if total_w > usable_width:
        scale = usable_width / total_w
        col_widths = [round(w * scale) for w in col_widths]

    data = [headers]
    t_nw = t_bags = t_amt = t_billed = t_kaccha = t_tax = t_total = t_cash = t_diesel = t_adv = t_bal = t_bal_final = t_oil_prem_pdf = t_received_pdf = t_pending_pdf = 0
    for idx, s in enumerate(sales):
        op = oil_map_pdf.get(s.get('voucher_no') or '') or oil_map_pdf.get(s.get('rst_no') or '')
        prem = (op.get('premium_amount', 0) or 0) if op else 0
        bal_final = round((s.get('balance', 0) or 0) + prem, 2)
        recv = s.get('total_received', 0) or 0
        pend = s.get('pending_balance', 0) or 0
        t_received_pdf += recv; t_pending_pdf += pend
        t_nw += s.get('net_weight_kg', 0); t_bags += s.get('bags', 0)
        t_amt += s.get('amount', 0); t_billed += s.get('billed_amount', 0); t_kaccha += s.get('kaccha_amount', 0)
        t_tax += s.get('tax_amount', 0); t_total += s.get('total', 0)
        t_cash += s.get('cash_paid', 0); t_diesel += s.get('diesel_paid', 0)
        t_adv += s.get('advance', 0); t_bal += s.get('balance', 0)
        t_bal_final += bal_final
        if op: t_oil_prem_pdf += prem
        row_data = []
        for key in col_keys:
            if key == 'voucher_no': row_data.append(s.get('voucher_no', '') or '')
            elif key == 'date': row_data.append(fmt_date(s.get('date', '')))
            elif key == 'party_name': row_data.append((s.get('party_name', '') or '')[:16])
            elif key == 'bill_from': row_data.append((s.get('bill_from', '') or '')[:14])
            elif key == 'destination': row_data.append((s.get('destination', '') or '')[:12])
            elif key in ('amount', 'billed_amount', 'kaccha_amount', 'tax_amount', 'total', 'balance'):
                v = s.get(key, 0) or 0
                row_data.append(f"{v:,.0f}" if v else '')
            elif key == 'balance_final':
                row_data.append(f"{bal_final:,.0f}" if bal_final else '0')
            elif key == 'last_payment_date':
                row_data.append(fmt_date(s.get('last_payment_date', '')) if s.get('last_payment_date') else '')
            elif key == 'total_received':
                row_data.append(f"{recv:,.0f}" if recv else '')
            elif key == 'pending_balance':
                row_data.append(f"{pend:,.0f}" if pend else '0')
            elif key == 'oil_pct': row_data.append(f"{op.get('actual_oil_pct', '')}%" if op else '')
            elif key == 'oil_diff':
                if op:
                    d = op.get('difference_pct', 0)
                    row_data.append(f"{'+' if d > 0 else ''}{d:.2f}%")
                else: row_data.append('')
            elif key == 'oil_premium': row_data.append(f"{prem:,.0f}" if op else '')
            elif key == 'net_weight_qtl':
                qtl = s.get('net_weight_qtl', 0) or (s.get('net_weight_kg', 0) or 0) / 100
                row_data.append(f"{qtl:,.2f}")
            else: row_data.append(s.get(key, 0) if key in ('net_weight_kg','bags','rate_per_qtl','cash_paid','diesel_paid','advance') else s.get(key, ''))
        data.append(row_data)

    # Total row
    total_row = []
    for key in col_keys:
        if key == 'date': total_row.append('TOTAL')
        elif key == 'net_weight_kg': total_row.append(round(t_nw, 0))
        elif key == 'net_weight_qtl': total_row.append(round(t_nw / 100, 2))
        elif key == 'bags': total_row.append(t_bags)
        elif key == 'amount': total_row.append(f"{t_amt:,.0f}")
        elif key == 'billed_amount': total_row.append(f"{t_billed:,.0f}" if t_billed else '')
        elif key == 'kaccha_amount': total_row.append(f"{t_kaccha:,.0f}" if t_kaccha else '')
        elif key == 'tax_amount': total_row.append(f"{t_tax:,.0f}")
        elif key == 'total': total_row.append(f"{t_total:,.0f}")
        elif key == 'cash_paid': total_row.append(round(t_cash, 0))
        elif key == 'diesel_paid': total_row.append(round(t_diesel, 0))
        elif key == 'advance': total_row.append(round(t_adv, 0))
        elif key == 'balance': total_row.append(f"{t_bal:,.0f}")
        elif key == 'balance_final': total_row.append(f"{t_bal_final:,.0f}")
        elif key == 'oil_premium': total_row.append(f"{t_oil_prem_pdf:,.0f}")
        elif key == 'last_payment_date': total_row.append('')
        elif key == 'total_received': total_row.append(f"{t_received_pdf:,.0f}" if t_received_pdf else '')
        elif key == 'pending_balance': total_row.append(f"{t_pending_pdf:,.0f}")
        else: total_row.append('')
    data.append(total_row)

    table = RLTable(data, colWidths=col_widths, repeatRows=1)

    # Find first numeric column index for right-align
    first_num = next((i for i, k in enumerate(col_keys) if k in ('net_weight_kg','net_weight_qtl','bags','rate_per_qtl','amount','billed_amount','kaccha_amount','tax_amount','total','cash_paid','diesel_paid','advance','balance','balance_final','total_received','pending_balance')), len(col_keys))

    nrows = len(data)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (first_num, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EBF1F8')]),
        ('BACKGROUND', (0, nrows - 1), (-1, nrows - 1), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, nrows - 1), (-1, nrows - 1), colors.white),
        ('FONTNAME', (0, nrows - 1), (-1, nrows - 1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]
    # v104.44.51 — Color-coded PKA/KCA/Tax/Total cells (data rows only)
    try:
        pakka_idx = col_keys.index('billed_amount')
        style_cmds.append(('BACKGROUND', (pakka_idx, 1), (pakka_idx, -2), colors.HexColor('#D0EBD2')))
        style_cmds.append(('TEXTCOLOR', (pakka_idx, 1), (pakka_idx, -2), colors.HexColor('#1B5E20')))
        style_cmds.append(('FONTNAME', (pakka_idx, 1), (pakka_idx, -2), 'Helvetica-Bold'))
    except ValueError: pass
    try:
        kac_idx = col_keys.index('kaccha_amount')
        style_cmds.append(('BACKGROUND', (kac_idx, 1), (kac_idx, -2), colors.HexColor('#FFD6D6')))
        style_cmds.append(('TEXTCOLOR', (kac_idx, 1), (kac_idx, -2), colors.HexColor('#B71C1C')))
        style_cmds.append(('FONTNAME', (kac_idx, 1), (kac_idx, -2), 'Helvetica-Bold'))
    except ValueError: pass
    try:
        tax_idx = col_keys.index('tax_amount')
        style_cmds.append(('BACKGROUND', (tax_idx, 1), (tax_idx, -2), colors.HexColor('#FFE8B0')))
        style_cmds.append(('TEXTCOLOR', (tax_idx, 1), (tax_idx, -2), colors.HexColor('#E65100')))
        style_cmds.append(('FONTNAME', (tax_idx, 1), (tax_idx, -2), 'Helvetica-Bold'))
    except ValueError: pass
    try:
        tot_idx = col_keys.index('total')
        style_cmds.append(('TEXTCOLOR', (tot_idx, 1), (tot_idx, -2), colors.HexColor('#0D47A1')))
        style_cmds.append(('FONTNAME', (tot_idx, 1), (tot_idx, -2), 'Helvetica-Bold'))
    except ValueError: pass
    # v104.44.53 — Balance (after premium) styling: red bold for positive, green for zero/negative
    try:
        bf_idx = col_keys.index('balance_final')
        style_cmds.append(('BACKGROUND', (bf_idx, 1), (bf_idx, -2), colors.HexColor('#FFF3E0')))
        style_cmds.append(('TEXTCOLOR', (bf_idx, 1), (bf_idx, -2), colors.HexColor('#C62828')))
        style_cmds.append(('FONTNAME', (bf_idx, 1), (bf_idx, -2), 'Helvetica-Bold'))
    except ValueError: pass
    # v104.44.56 — Payment columns styling
    try:
        rcv_idx = col_keys.index('total_received')
        style_cmds.append(('BACKGROUND', (rcv_idx, 1), (rcv_idx, -2), colors.HexColor('#E0F7FA')))
        style_cmds.append(('TEXTCOLOR', (rcv_idx, 1), (rcv_idx, -2), colors.HexColor('#00838F')))
        style_cmds.append(('FONTNAME', (rcv_idx, 1), (rcv_idx, -2), 'Helvetica-Bold'))
    except ValueError: pass
    try:
        pnd_idx = col_keys.index('pending_balance')
        style_cmds.append(('BACKGROUND', (pnd_idx, 1), (pnd_idx, -2), colors.HexColor('#FFE0B2')))
        style_cmds.append(('TEXTCOLOR', (pnd_idx, 1), (pnd_idx, -2), colors.HexColor('#E65100')))
        style_cmds.append(('FONTNAME', (pnd_idx, 1), (pnd_idx, -2), 'Helvetica-Bold'))
    except ValueError: pass
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    # v104.44.54 — Payment Summary footer removed (user feedback: "ganda lag raha hai")

    # Generated date
    elements.append(Spacer(1, 4))
    gen_style = ParagraphStyle('Gen', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#999999'))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", gen_style))

    doc.build(elements)
    buffer.seek(0)
    fn = f"{(product or 'byproduct').lower().replace(' ','_')}_sale_register_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


# v104.44.56 — Option C: Party Statement Excel Export (chronological ledger A4)
@router.get("/bp-sale-register/export/statement-excel")
async def export_party_statement_excel(party: str, kms_year: str = "", season: str = "",
                                        gst_filter: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from fastapi.responses import Response
    stmt = await get_bp_party_statement(party=party, kms_year=kms_year, season=season, gst_filter=gst_filter)
    branding = await db.branding.find_one({}, {"_id": 0}) or {}
    company = (branding.get('company_name') or 'Rice Mill').upper()
    address = branding.get('address', '') or ''
    phone = branding.get('phone', '') or ''

    wb = Workbook(); ws = wb.active; ws.title = f"{party[:25]} Statement"
    border = Border(left=Side(style='thin', color='B0C4DE'), right=Side(style='thin', color='B0C4DE'),
                    top=Side(style='thin', color='B0C4DE'), bottom=Side(style='thin', color='B0C4DE'))
    cols = [('Date', 14), ('Sub-Ledger', 18), ('Type', 8), ('Description', 38), ('Debit (Dr)', 14), ('Credit (Cr)', 14), ('Balance', 14)]
    ncols = len(cols)
    for i, (h, w) in enumerate(cols, 1): ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # Row 1: company
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, company); c.font = Font(bold=True, size=14, color='1F4E79'); c.alignment = Alignment(horizontal='center')
    # Row 2: address
    if address or phone:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(2, 1, f"{address}  |  {phone}"); c.font = Font(size=9, color='666666'); c.alignment = Alignment(horizontal='center')
    # Row 3: title
    mode = gst_filter or 'ALL'
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    title = f"PARTY STATEMENT — {party}  [{mode}]"
    if kms_year: title += f"   FY {kms_year}"
    if season: title += f"   ({season})"
    c = ws.cell(3, 1, title); c.font = Font(bold=True, size=12, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='2E75B6'); c.alignment = Alignment(horizontal='center')

    # Row 4: summary
    s = stmt['summary']
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
    summary_txt = f"Entries: {s['entry_count']}   |   Total Debit: ₹{s['total_debit']:,.2f}   |   Total Credit: ₹{s['total_credit']:,.2f}   |   Closing Balance: ₹{s['closing_balance']:,.2f}"
    c = ws.cell(4, 1, summary_txt); c.font = Font(italic=True, size=9, color='555555')
    c.fill = PatternFill('solid', fgColor='F5F5F5'); c.alignment = Alignment(horizontal='center')

    # Row 5: header
    header_row = 5
    for i, (h, _) in enumerate(cols, 1):
        c = ws.cell(header_row, i, h)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F4E79')
        c.alignment = Alignment(horizontal='center')
        c.border = border

    # Data
    for idx, e in enumerate(stmt['entries']):
        r = header_row + 1 + idx
        alt = PatternFill('solid', fgColor='F0F6FC') if idx % 2 == 0 else None
        debit = e['amount'] if e['flow'] == 'Dr' else 0
        credit = e['amount'] if e['flow'] == 'Cr' else 0
        vals = [fmt_date(e['date']), e.get('party_name', ''), e['flow'], e.get('description', ''),
                debit, credit, e['running_balance']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v); c.font = Font(size=9); c.border = border
            if alt: c.fill = alt
            if ci >= 5: c.alignment = Alignment(horizontal='right'); c.number_format = '#,##0.00'
            if ci == 5 and v: c.font = Font(size=9, bold=True, color='1B5E20')  # Debit green
            elif ci == 6 and v: c.font = Font(size=9, bold=True, color='C62828')  # Credit red
            elif ci == 7: c.font = Font(size=9, bold=True, color='0D47A1')  # Balance dark blue

    # Total row
    tr = header_row + 1 + len(stmt['entries'])
    for i in range(1, ncols + 1):
        c = ws.cell(tr, i); c.fill = PatternFill('solid', fgColor='2E75B6'); c.font = Font(bold=True, size=9, color='FFFFFF'); c.border = border
    ws.cell(tr, 1, 'CLOSING')
    ws.cell(tr, 5, s['total_debit']).number_format = '#,##0.00'
    ws.cell(tr, 6, s['total_credit']).number_format = '#,##0.00'
    ws.cell(tr, 7, s['closing_balance']).number_format = '#,##0.00'
    for i in range(5, 8): ws.cell(tr, i).alignment = Alignment(horizontal='right')

    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.auto_filter.ref = f"A{header_row}:{ws.cell(header_row, ncols).column_letter}{tr - 1}"
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"{party.lower().replace(' ', '_')}_statement_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buf.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fn}'})


# v104.44.56 — Option C: Party Statement PDF Export (A4 portrait, professional)
@router.get("/bp-sale-register/export/statement-pdf")
async def export_party_statement_pdf(party: str, kms_year: str = "", season: str = "",
                                      gst_filter: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle
    from io import BytesIO
    from fastapi.responses import Response
    stmt = await get_bp_party_statement(party=party, kms_year=kms_year, season=season, gst_filter=gst_filter)
    branding = await db.branding.find_one({}, {"_id": 0}) or {}

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    elements = []; styles = getSampleStyleSheet()

    from utils.export_helpers import get_pdf_company_header
    elements.extend(get_pdf_company_header(branding))

    mode = gst_filter or 'ALL'
    title_bg = colors.HexColor('#2E7D32') if mode == 'PKA' else (colors.HexColor('#C62828') if mode == 'KCA' else colors.HexColor('#2E75B6'))
    title_style = ParagraphStyle('StmtTitle', parent=styles['Heading2'], fontSize=11,
        textColor=colors.white, backColor=title_bg, spaceAfter=4, alignment=1, borderPadding=(4, 4, 4, 4))
    title = f"PARTY STATEMENT — {party}  [{mode}]"
    if kms_year: title += f"   FY {kms_year}"
    if season: title += f"   ({season})"
    elements.append(Paragraph(title, title_style))

    s = stmt['summary']
    sub_style = ParagraphStyle('SubT', parent=styles['Normal'], fontSize=8, alignment=1,
        textColor=colors.HexColor('#555555'), backColor=colors.HexColor('#F5F5F5'), borderPadding=(3, 3, 3, 3))
    summary_txt = f"<b>Entries:</b> {s['entry_count']}  |  <b>Total Debit:</b> <font color='#1B5E20'>₹{s['total_debit']:,.2f}</font>  |  <b>Total Credit:</b> <font color='#C62828'>₹{s['total_credit']:,.2f}</font>  |  <b>Closing Balance:</b> <font color='red'>₹{s['closing_balance']:,.2f}</font>"
    elements.append(Paragraph(summary_txt, sub_style))
    elements.append(Spacer(1, 5))

    # Table
    headers = ['Date', 'Sub-Ledger', 'Type', 'Description', 'Debit (Dr)', 'Credit (Cr)', 'Balance']
    col_widths = [55, 70, 28, 230, 65, 65, 65]
    data = [headers]
    for e in stmt['entries']:
        debit = f"{e['amount']:,.0f}" if e['flow'] == 'Dr' else ''
        credit = f"{e['amount']:,.0f}" if e['flow'] == 'Cr' else ''
        data.append([fmt_date(e['date']), e.get('party_name', '')[:18], e['flow'], e.get('description', '')[:55], debit, credit, f"{e['running_balance']:,.0f}"])
    data.append(['CLOSING', '', '', '', f"{s['total_debit']:,.0f}", f"{s['total_credit']:,.0f}", f"{s['closing_balance']:,.0f}"])

    table = RLTable(data, colWidths=col_widths, repeatRows=1)
    nrows = len(data)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EBF1F8')]),
        ('BACKGROUND', (0, nrows - 1), (-1, nrows - 1), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, nrows - 1), (-1, nrows - 1), colors.white),
        ('FONTNAME', (0, nrows - 1), (-1, nrows - 1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (4, 1), (4, -2), colors.HexColor('#1B5E20')),  # Dr green
        ('TEXTCOLOR', (5, 1), (5, -2), colors.HexColor('#C62828')),  # Cr red
        ('TEXTCOLOR', (6, 1), (6, -2), colors.HexColor('#0D47A1')),  # Balance blue
        ('FONTNAME', (4, 1), (6, -2), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    elements.append(Spacer(1, 6))
    gen_style = ParagraphStyle('Gen', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#999999'))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", gen_style))

    doc.build(elements); buffer.seek(0)
    fn = f"{party.lower().replace(' ', '_')}_statement_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={fn}'})
