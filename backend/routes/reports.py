from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import *
import uuid, io, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.report_helper import get_columns, get_entry_row, get_total_row, get_excel_headers, get_pdf_headers, get_excel_widths, get_pdf_widths_mm, col_count

router = APIRouter()

# ============ PHASE 4: REPORTING ============

@router.get("/reports/cmr-vs-dc")
async def report_cmr_vs_dc(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Compare milling output (CMR) vs DC allotment and deliveries"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    # Milling data
    milling = await db.milling_entries.find(query, {"_id": 0}).to_list(5000)
    total_paddy_milled = round(sum(e.get("paddy_input_qntl", 0) for e in milling), 2)
    total_rice_produced = round(sum(e.get("rice_qntl", 0) for e in milling), 2)
    total_frk_used = round(sum(e.get("frk_used_qntl", 0) for e in milling), 2)
    total_cmr = round(sum(e.get("cmr_delivery_qntl", 0) for e in milling), 2)
    avg_outturn = round(total_cmr / total_paddy_milled * 100, 2) if total_paddy_milled > 0 else 0
    # DC data
    dcs = await db.dc_entries.find(query, {"_id": 0}).to_list(1000)
    deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    total_dc_allotted = round(sum(d.get("quantity_qntl", 0) for d in dcs), 2)
    total_dc_delivered = round(sum(d.get("quantity_qntl", 0) for d in deliveries), 2)
    total_dc_pending = round(total_dc_allotted - total_dc_delivered, 2)
    # Comparison
    cmr_surplus = round(total_cmr - total_dc_allotted, 2)
    delivery_gap = round(total_cmr - total_dc_delivered, 2)  # CMR ready but not delivered
    # By-product revenue
    bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(5000)
    bp_revenue = round(sum(s.get("total_amount", 0) for s in bp_sales), 2)
    return {
        "milling": {"total_paddy_milled": total_paddy_milled, "total_rice_produced": total_rice_produced, "total_frk_used": total_frk_used, "total_cmr_ready": total_cmr, "avg_outturn_pct": avg_outturn, "milling_count": len(milling)},
        "dc": {"total_allotted": total_dc_allotted, "total_delivered": total_dc_delivered, "total_pending": total_dc_pending, "dc_count": len(dcs), "delivery_count": len(deliveries)},
        "comparison": {"cmr_vs_dc_allotted": cmr_surplus, "cmr_vs_dc_delivered": delivery_gap},
        "byproduct_revenue": bp_revenue
    }


@router.get("/reports/season-pnl")
async def report_season_pnl(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Season-wise Profit & Loss summary"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    # Income
    msp_payments = await db.msp_payments.find(query, {"_id": 0}).to_list(5000)
    msp_income = round(sum(p.get("amount", 0) for p in msp_payments), 2)
    bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(5000)
    bp_income = round(sum(s.get("total_amount", 0) for s in bp_sales), 2)
    # Expenses
    frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(5000)
    frk_cost = round(sum(p.get("total_amount", 0) for p in frk_purchases), 2)
    gunny_bags = await db.gunny_bags.find(query, {"_id": 0}).to_list(5000)
    gunny_cost = round(sum(g.get("amount", 0) for g in gunny_bags if g.get("txn_type") == "in"), 2)
    cash_txns = await db.cash_transactions.find(query, {"_id": 0}).to_list(10000)
    cash_expenses = round(sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "nikasi"), 2)
    cash_income_other = round(sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "jama"), 2)
    # Truck/Agent payments from mill entries
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(10000)
    truck_payments = round(sum(e.get("tp_paid", 0) for e in entries), 2)
    agent_payments = round(sum(e.get("agent_paid", 0) for e in entries), 2)
    total_income = round(msp_income + bp_income + cash_income_other, 2)
    total_expenses = round(frk_cost + gunny_cost + cash_expenses + truck_payments + agent_payments, 2)
    net_pnl = round(total_income - total_expenses, 2)
    return {
        "income": {"msp_payments": msp_income, "byproduct_sales": bp_income, "cash_book_jama": cash_income_other, "total": total_income},
        "expenses": {"frk_purchases": frk_cost, "gunny_bags": gunny_cost, "cash_book_nikasi": cash_expenses, "truck_payments": truck_payments, "agent_payments": agent_payments, "total": total_expenses},
        "net_pnl": net_pnl, "profit": net_pnl >= 0
    }


@router.get("/reports/cmr-vs-dc/excel")
async def export_cmr_vs_dc_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, COLORS, BORDER_THIN)
    data = await report_cmr_vs_dc(kms_year=kms_year, season=season)
    wb = Workbook(); ws = wb.active; ws.title = "CMR vs DC"
    ncols = 4
    title = "CMR vs DC Report / सीएमआर vs डीसी"
    if kms_year: title += f" - FY {kms_year}"
    style_excel_title(ws, title, ncols)
    
    ws.cell(row=4, column=1, value="MILLING OUTPUT").font = Font(bold=True, size=11, color=COLORS['title_text'])
    items = [("Paddy Milled (Q)", data["milling"]["total_paddy_milled"]), ("Rice Produced (Q)", data["milling"]["total_rice_produced"]),
             ("FRK Used (Q)", data["milling"]["total_frk_used"]), ("CMR Ready (Q)", data["milling"]["total_cmr_ready"]),
             ("Avg Outturn %", data["milling"]["avg_outturn_pct"]), ("Milling Count", data["milling"]["milling_count"])]
    for i, (label, val) in enumerate(items, 5):
        ws.cell(row=i, column=1, value=label).border = BORDER_THIN
        c = ws.cell(row=i, column=2, value=val); c.border = BORDER_THIN; c.alignment = Alignment(horizontal='right')
    
    row = 12
    ws.cell(row=row, column=1, value="DC ALLOTMENT & DELIVERY").font = Font(bold=True, size=11, color=COLORS['subtitle_text'])
    items2 = [("DC Allotted (Q)", data["dc"]["total_allotted"]), ("DC Delivered (Q)", data["dc"]["total_delivered"]),
              ("DC Pending (Q)", data["dc"]["total_pending"]), ("Total DCs", data["dc"]["dc_count"]), ("Total Deliveries", data["dc"]["delivery_count"])]
    for i, (label, val) in enumerate(items2, row+1):
        ws.cell(row=i, column=1, value=label).border = BORDER_THIN
        c = ws.cell(row=i, column=2, value=val); c.border = BORDER_THIN; c.alignment = Alignment(horizontal='right')
    
    row = 19
    ws.cell(row=row, column=1, value="COMPARISON").font = Font(bold=True, size=11, color=COLORS['date_text'])
    ws.cell(row=row+1, column=1, value="CMR vs DC Allotted").border = BORDER_THIN
    ws.cell(row=row+1, column=2, value=data["comparison"]["cmr_vs_dc_allotted"]).border = BORDER_THIN
    ws.cell(row=row+2, column=1, value="CMR vs DC Delivered").border = BORDER_THIN
    ws.cell(row=row+2, column=2, value=data["comparison"]["cmr_vs_dc_delivered"]).border = BORDER_THIN
    ws.cell(row=row+3, column=1, value="By-Product Revenue (Rs.)").border = BORDER_THIN
    ws.cell(row=row+3, column=2, value=data["byproduct_revenue"]).border = BORDER_THIN
    for letter in ['A','B','C','D']: ws.column_dimensions[letter].width = 22
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cmr_vs_dc_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/reports/cmr-vs-dc/pdf")
async def export_cmr_vs_dc_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib import colors
    from io import BytesIO
    from utils.export_helpers import get_pdf_table_style
    from utils.branding_helper import get_pdf_company_header_from_db
    data = await report_cmr_vs_dc(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=30, bottomMargin=30)
    elements = []; styles = get_pdf_styles()
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph("CMR vs DC Report / सीएमआर vs डीसी", styles['Title'])); elements.append(Spacer(1, 12))
    rows = [['Metric', 'Value'],
        ['--- MILLING ---', ''], ['Paddy Milled (Q)', data['milling']['total_paddy_milled']], ['Rice Produced (Q)', data['milling']['total_rice_produced']],
        ['FRK Used (Q)', data['milling']['total_frk_used']], ['CMR Ready (Q)', data['milling']['total_cmr_ready']], ['Outturn %', data['milling']['avg_outturn_pct']],
        ['--- DC ---', ''], ['DC Allotted (Q)', data['dc']['total_allotted']], ['DC Delivered (Q)', data['dc']['total_delivered']], ['DC Pending (Q)', data['dc']['total_pending']],
        ['--- COMPARISON ---', ''], ['CMR vs DC Allotted', data['comparison']['cmr_vs_dc_allotted']], ['CMR vs DC Delivered', data['comparison']['cmr_vs_dc_delivered']],
        ['By-Product Revenue', f"Rs.{data['byproduct_revenue']}"]]
    table = RLTable(rows, colWidths=[200, 150])
    style_cmds = get_pdf_table_style(len(rows))
    style_cmds.append(('ALIGN',(1,0),(1,-1),'RIGHT'))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cmr_vs_dc_{datetime.now().strftime('%Y%m%d')}.pdf"})


@router.get("/reports/season-pnl/excel")
async def export_season_pnl_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, COLORS, BORDER_THIN)
    data = await report_season_pnl(kms_year=kms_year, season=season)
    wb = Workbook(); ws = wb.active; ws.title = "Season P&L"
    ncols = 3
    title = "Season P&L Report / मौसम लाभ-हानि"
    if kms_year: title += f" - FY {kms_year}"
    style_excel_title(ws, title, ncols)
    
    row = 4
    ws.cell(row=row, column=1, value="INCOME").font = Font(bold=True, size=11, color=COLORS['jama_text'])
    for label, val in [("MSP Payments", data["income"]["msp_payments"]), ("By-Product Sales", data["income"]["byproduct_sales"]),
                        ("Cash Book Jama", data["income"]["cash_book_jama"]), ("TOTAL INCOME", data["income"]["total"])]:
        row += 1
        ws.cell(row=row, column=1, value=label).border = BORDER_THIN
        c = ws.cell(row=row, column=2, value=val); c.border = BORDER_THIN; c.number_format = '#,##0.00'
        if label.startswith("TOTAL"): ws.cell(row=row, column=1).font = Font(bold=True); c.font = Font(bold=True)
    row += 2
    ws.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, size=11, color=COLORS['nikasi_text'])
    for label, val in [("FRK Purchases", data["expenses"]["frk_purchases"]), ("Gunny Bags", data["expenses"]["gunny_bags"]),
                        ("Cash Book Nikasi", data["expenses"]["cash_book_nikasi"]), ("Truck Payments", data["expenses"]["truck_payments"]),
                        ("Agent Payments", data["expenses"]["agent_payments"]), ("TOTAL EXPENSES", data["expenses"]["total"])]:
        row += 1
        ws.cell(row=row, column=1, value=label).border = BORDER_THIN
        c = ws.cell(row=row, column=2, value=val); c.border = BORDER_THIN; c.number_format = '#,##0.00'
        if label.startswith("TOTAL"): ws.cell(row=row, column=1).font = Font(bold=True); c.font = Font(bold=True)
    row += 2
    pnl_label = "NET PROFIT" if data["profit"] else "NET LOSS"
    ws.cell(row=row, column=1, value=pnl_label).font = Font(bold=True, size=12, color=COLORS['jama_text'] if data["profit"] else COLORS['nikasi_text'])
    ws.cell(row=row, column=2, value=data["net_pnl"]).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    for letter in ['A','B','C']: ws.column_dimensions[letter].width = 22
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=season_pnl_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/reports/season-pnl/pdf")
async def export_season_pnl_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib import colors
    from io import BytesIO
    from utils.export_helpers import get_pdf_table_style
    from utils.branding_helper import get_pdf_company_header_from_db
    data = await report_season_pnl(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=30, bottomMargin=30)
    elements = []; styles = get_pdf_styles()
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph("Season P&L Report / मौसम लाभ-हानि", styles['Title'])); elements.append(Spacer(1, 12))
    elements.append(Paragraph("INCOME", styles['Heading2'])); elements.append(Spacer(1, 4))
    idata = [['Source', 'Amount (Rs.)'], ['MSP Payments', data['income']['msp_payments']], ['By-Product Sales', data['income']['byproduct_sales']],
             ['Cash Book Jama', data['income']['cash_book_jama']], ['TOTAL', data['income']['total']]]
    it = RLTable(idata, colWidths=[200, 120])
    it.setStyle(TableStyle(get_pdf_table_style(len(idata))))
    elements.append(it); elements.append(Spacer(1, 12))
    elements.append(Paragraph("EXPENSES", styles['Heading2'])); elements.append(Spacer(1, 4))
    edata = [['Category', 'Amount (Rs.)'], ['FRK Purchases', data['expenses']['frk_purchases']], ['Gunny Bags', data['expenses']['gunny_bags']],
             ['Cash Book Nikasi', data['expenses']['cash_book_nikasi']], ['Truck Payments', data['expenses']['truck_payments']],
             ['Agent Payments', data['expenses']['agent_payments']], ['TOTAL', data['expenses']['total']]]
    et = RLTable(edata, colWidths=[200, 120])
    et.setStyle(TableStyle(get_pdf_table_style(len(edata))))
    elements.append(et); elements.append(Spacer(1, 15))
    pnl_color = colors.HexColor('#166534') if data['profit'] else colors.HexColor('#991b1b')
    pdata = [['NET ' + ('PROFIT' if data['profit'] else 'LOSS'), f"Rs.{data['net_pnl']}"]]
    pt = RLTable(pdata, colWidths=[200, 120])
    pt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),pnl_color),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,0),14),('FONTNAME',(0,0),(-1,0),'FreeSansBold'),('ALIGN',(1,0),(1,0),'RIGHT')]))
    elements.append(pt); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=season_pnl_{datetime.now().strftime('%Y%m%d')}.pdf"})




# ============ AGENT & MANDI WISE REPORT ============

@router.get("/reports/agent-mandi-wise")
async def report_agent_mandi_wise(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Agent & Mandi wise report with individual trip entries"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        query["date"] = date_q

    entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", -1).to_list(5000)

    if search:
        s = search.lower()
        entries = [e for e in entries if s in (e.get("mandi_name","")).lower() or s in (e.get("agent_name","")).lower()]

    # Group by mandi_name
    mandi_map = {}
    for e in entries:
        mn = e.get("mandi_name", "Unknown")
        if mn not in mandi_map:
            mandi_map[mn] = {"mandi_name": mn, "agent_name": e.get("agent_name", ""), "entries": [], "totals": {
                "total_qntl": 0, "total_bag": 0, "total_g_deposite": 0, "total_gbw_cut": 0,
                "total_plastic_bag": 0, "total_p_pkt_cut": 0, "total_mill_w": 0,
                "total_moisture_cut": 0, "total_final_w": 0, "total_tp_weight": 0,
                "total_g_issued": 0, "total_cash_paid": 0, "total_diesel_paid": 0,
                "total_disc_dust_poll": 0, "entry_count": 0
            }}
        t = mandi_map[mn]["totals"]
        t["total_qntl"] += e.get("qntl", 0)
        t["total_bag"] += e.get("bag", 0)
        t["total_g_deposite"] += e.get("g_deposite", 0)
        t["total_gbw_cut"] += e.get("gbw_cut", 0) or 0
        t["total_plastic_bag"] += e.get("plastic_bag", 0) or 0
        t["total_p_pkt_cut"] += e.get("p_pkt_cut", 0) or 0
        t["total_mill_w"] += e.get("mill_w", 0)
        t["total_moisture_cut"] += e.get("moisture_cut", 0) or 0
        t["total_final_w"] += e.get("final_w", 0)
        t["total_tp_weight"] += float(e.get("tp_weight", 0) or 0)
        t["total_g_issued"] += e.get("g_issued", 0)
        t["total_cash_paid"] += e.get("cash_paid", 0) or 0
        t["total_diesel_paid"] += e.get("diesel_paid", 0) or 0
        t["total_disc_dust_poll"] += e.get("disc_dust_poll", 0) or 0
        t["entry_count"] += 1
        mandi_map[mn]["entries"].append({
            "date": e.get("date", ""),
            "truck_no": e.get("truck_no", ""),
            "qntl": round(e.get("qntl", 0), 2),
            "bag": e.get("bag", 0),
            "g_deposite": e.get("g_deposite", 0),
            "gbw_cut": round(e.get("gbw_cut", 0) or 0, 2),
            "plastic_bag": e.get("plastic_bag", 0) or 0,
            "p_pkt_cut": round(e.get("p_pkt_cut", 0) or 0, 2),
            "mill_w": round(e.get("mill_w", 0), 2),
            "moisture_cut_percent": round(e.get("moisture_cut_percent", 0) or 0, 2),
            "moisture_cut": round(e.get("moisture_cut", 0) or 0, 2),
            "cutting_percent": round(e.get("cutting_percent", 0) or 0, 2),
            "disc_dust_poll": round(e.get("disc_dust_poll", 0) or 0, 2),
            "final_w": round(e.get("final_w", 0), 2),
            "tp_weight": float(e.get("tp_weight", 0) or 0),
            "g_issued": e.get("g_issued", 0),
            "cash_paid": e.get("cash_paid", 0) or 0,
            "diesel_paid": e.get("diesel_paid", 0) or 0,
        })

    # Round totals
    result = []
    for mn, data in mandi_map.items():
        for k in data["totals"]:
            data["totals"][k] = round(data["totals"][k], 2)
        result.append(data)

    result.sort(key=lambda x: x["mandi_name"])

    # Fetch mandi targets for extra QNTL calculation
    target_query = {}
    if kms_year: target_query["kms_year"] = kms_year
    if season: target_query["season"] = season
    targets = await db.mandi_targets.find(target_query, {"_id": 0}).to_list(500)
    target_map = {t["mandi_name"]: t for t in targets}

    # Check existing pvt entries from this feature
    pvt_entries = await db.private_paddy.find({"source": "agent_extra"}, {"_id": 0, "mandi_name": 1}).to_list(500)
    pvt_mandi_set = set(p.get("mandi_name", "") for p in pvt_entries)

    for m in result:
        mn = m["mandi_name"]
        target = target_map.get(mn, {})
        target_qntl = round(target.get("target_qntl", 0), 2)
        cutting_pct = round(target.get("cutting_percent", 0), 2)
        expected_total = round(target.get("expected_total", 0) or (target_qntl + target_qntl * cutting_pct / 100), 2)
        actual_final_w_qntl = round(m["totals"]["total_final_w"] / 100, 2)  # final_w is in kg, convert to QNTL
        extra_qntl = round(max(0, actual_final_w_qntl - expected_total), 2) if expected_total > 0 else 0
        m["target_qntl"] = target_qntl
        m["cutting_percent"] = cutting_pct
        m["expected_total"] = expected_total
        m["actual_final_qntl"] = actual_final_w_qntl
        m["extra_qntl"] = extra_qntl
        m["pvt_moved"] = mn in pvt_mandi_set
        # Include last entry details for pvt move
        if m["entries"]:
            last_entry = m["entries"][-1]  # last entry (oldest date, last truck)
            m["last_truck"] = {
                "truck_no": last_entry.get("truck_no", ""),
                "date": last_entry.get("date", ""),
                "qntl": last_entry.get("qntl", 0),
                "bag": last_entry.get("bag", 0),
                "agent_name": last_entry.get("agent_name", m.get("agent_name", "")),
                "mandi_name": last_entry.get("mandi_name", mn),
            }

    # Grand totals
    grand = {"total_qntl": 0, "total_bag": 0, "total_g_deposite": 0, "total_gbw_cut": 0,
             "total_plastic_bag": 0, "total_p_pkt_cut": 0, "total_mill_w": 0,
             "total_moisture_cut": 0, "total_final_w": 0, "total_tp_weight": 0, "total_g_issued": 0,
             "total_cash_paid": 0, "total_diesel_paid": 0, "total_disc_dust_poll": 0, "entry_count": 0}
    for m in result:
        for k in grand:
            grand[k] += m["totals"][k]
    for k in grand:
        grand[k] = round(grand[k], 2)
    grand["total_extra_qntl"] = round(sum(m.get("extra_qntl", 0) for m in result), 2)

    return {"mandis": result, "grand_totals": grand}


@router.post("/reports/agent-mandi-wise/move-to-pvt")
async def move_extra_to_pvt(request: Request):
    """Move extra QNTL (above target) to Private Paddy Purchase"""
    body = await request.json()
    mandi_name = body.get("mandi_name")
    agent_name = body.get("agent_name", "")
    extra_qntl = body.get("extra_qntl", 0)
    rate = body.get("rate", 0)
    kms_year = body.get("kms_year", "")
    season = body.get("season", "") or "Kharif"
    username = body.get("username", "admin")
    last_truck = body.get("last_truck", {})

    if not mandi_name or extra_qntl <= 0 or rate <= 0:
        return {"success": False, "detail": "Mandi name, extra QNTL aur rate required hai"}

    total_amount = round(extra_qntl * rate, 2)

    # Check if already moved
    existing = await db.private_paddy.find_one({"mandi_name": mandi_name, "source": "agent_extra", "kms_year": kms_year, "season": season}, {"_id": 0})
    if existing:
        return {"success": False, "detail": f"{mandi_name} ka extra QNTL pehle se Pvt Purchase mein move ho chuka hai"}

    pvt_entry = {
        "id": str(uuid.uuid4()),
        "date": last_truck.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        "party_name": f"{agent_name} ({mandi_name})",
        "mandi_name": mandi_name,
        "agent_name": agent_name,
        "truck_no": last_truck.get("truck_no", ""),
        "kg": round(extra_qntl * 100, 2),
        "qntl": round(extra_qntl, 2),
        "final_qntl": round(extra_qntl, 2),
        "quantity_qntl": round(extra_qntl, 2),
        "rate_per_qntl": round(rate, 2),
        "total_amount": total_amount,
        "balance": total_amount,
        "bag": last_truck.get("bag", 0),
        "paid_amount": 0,
        "status": "pending",
        "source": "agent_extra",
        "note": f"Agent extra - Target se {extra_qntl}Q zyada ({last_truck.get('truck_no', '')})",
        "kms_year": kms_year,
        "season": season,
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.private_paddy.insert_one(pvt_entry)

    # Auto-create Party Ledger Jama entry - what we owe this party (NOT cash - no cash movement yet)
    party_name = pvt_entry["party_name"]
    jama_entry = {
        "id": str(uuid.uuid4()),
        "date": pvt_entry["date"],
        "account": "ledger",
        "txn_type": "jama",
        "category": party_name,
        "party_type": "Pvt Paddy Purchase",
        "description": f"Paddy Purchase: {party_name} - {extra_qntl}Q @ Rs.{rate}/Q = Rs.{total_amount}",
        "amount": round_amount(total_amount),
        "reference": f"pvt_party_jama:{pvt_entry['id'][:8]}",
        "kms_year": kms_year,
        "season": season,
        "created_by": username,
        "linked_entry_id": pvt_entry["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(jama_entry)

    return {"success": True, "message": f"{extra_qntl}Q @ Rs.{rate}/Q = Rs.{total_amount} Pvt Purchase mein move ho gaya ({agent_name} - {mandi_name})"}



@router.get("/reports/agent-mandi-wise/excel")
async def export_agent_mandi_wise_excel(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, mandis: Optional[str] = None):
    from io import BytesIO
    data = await report_agent_mandi_wise(kms_year=kms_year, season=season, search=search, date_from=date_from, date_to=date_to)
    # Filter to only expanded mandis if specified
    if mandis:
        mandi_names = [m.strip() for m in mandis.split(',') if m.strip()]
        if mandi_names:
            data["mandis"] = [m for m in data["mandis"] if m["mandi_name"] in mandi_names]
            # Recalculate grand totals for filtered mandis
            gt = {}
            for key in data["grand_totals"]:
                gt[key] = 0
            for m in data["mandis"]:
                for key in gt:
                    gt[key] += m["totals"].get(key, 0)
            gt["entry_count"] = sum(m["totals"]["entry_count"] for m in data["mandis"])
            gt["total_extra_qntl"] = round(sum(m.get("extra_qntl", 0) for m in data["mandis"]), 2)
            data["grand_totals"] = gt
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    wb = Workbook(); ws = wb.active; ws.title = "Agent Mandi Report"

    cols = get_columns("agent_mandi_report")
    ncols = col_count(cols)
    headers = get_excel_headers(cols)
    widths = get_excel_widths(cols)

    title = "Agent & Mandi Wise Report / एजेंट और मंडी"
    if kms_year: title += f" | FY: {kms_year}"
    if season: title += f" | {season}"
    style_excel_title(ws, title, ncols)

    row = 4
    for mandi_data in data["mandis"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        lbl = f"{mandi_data['mandi_name']} - Agent: {mandi_data['agent_name']} ({mandi_data['totals']['entry_count']} entries)"
        if mandi_data.get("target_qntl"):
            lbl += f" | Target: {mandi_data['target_qntl']}Q | Final W: {mandi_data.get('actual_final_qntl',0)}Q | Extra: {mandi_data.get('extra_qntl',0)}Q"
        from openpyxl.styles import PatternFill
        cell = ws.cell(row=row, column=1, value=lbl)
        cell.fill = PatternFill(start_color=COLORS['header_bg'], fill_type='solid')
        cell.font = Font(bold=True, color=COLORS['header_text'], size=10)
        row += 1

        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row, column=col_idx, value=h)
        style_excel_header_row(ws, row, ncols)
        row += 1

        data_start = row
        for entry in mandi_data["entries"]:
            vals = get_entry_row(entry, cols)
            for col_idx, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col_idx, value=v)
                if cols[col_idx-1]["align"] == "right": c.alignment = Alignment(horizontal='right')
            row += 1
        if mandi_data["entries"]:
            style_excel_data_rows(ws, data_start, row - 1, ncols, headers)

        t = mandi_data["totals"]
        total_vals = get_total_row(t, cols)
        ws.cell(row=row, column=1, value="TOTAL / कुल")
        for col_idx, val in enumerate(total_vals, 1):
            if val is not None and val != "":
                c = ws.cell(row=row, column=col_idx, value=val)
                c.alignment = Alignment(horizontal='right')
        style_excel_total_row(ws, row, ncols)
        row += 2

    g = data["grand_totals"]
    grand_vals = get_total_row(g, cols)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value=f"GRAND TOTAL ({g['entry_count']} entries)")
    for col_idx, val in enumerate(grand_vals, 1):
        if val is not None and val != "" and col_idx > 2:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.alignment = Alignment(horizontal='right')
    style_excel_total_row(ws, row, ncols)

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=agent_mandi_report_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/reports/agent-mandi-wise/pdf")
async def export_agent_mandi_wise_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, mandis: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO

    report_data = await report_agent_mandi_wise(kms_year=kms_year, season=season, search=search, date_from=date_from, date_to=date_to)
    # Filter to only expanded mandis if specified
    if mandis:
        mandi_names = [m.strip() for m in mandis.split(',') if m.strip()]
        if mandi_names:
            report_data["mandis"] = [m for m in report_data["mandis"] if m["mandi_name"] in mandi_names]
            gt = {}
            for key in report_data["grand_totals"]:
                gt[key] = 0
            for m in report_data["mandis"]:
                for key in gt:
                    gt[key] += m["totals"].get(key, 0)
            gt["entry_count"] = sum(m["totals"]["entry_count"] for m in report_data["mandis"])
            gt["total_extra_qntl"] = round(sum(m.get("extra_qntl", 0) for m in report_data["mandis"]), 2)
            report_data["grand_totals"] = gt
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=8*mm, rightMargin=8*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []; styles = get_pdf_styles()

    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    title = f"Agent & Mandi Wise Report"
    if kms_year: title += f" | FY: {kms_year}"
    if season: title += f" | {season}"
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#D97706'), alignment=TA_CENTER)
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 8))

    cols = get_columns("agent_mandi_report")
    ncols = col_count(cols)
    headers = get_pdf_headers(cols)
    from utils.export_helpers import get_pdf_table_style
    col_widths = [w*mm for w in get_pdf_widths_mm(cols)]

    for mandi_data in report_data["mandis"]:
        mandi_label = f"{mandi_data['mandi_name']} - Agent: {mandi_data['agent_name']} ({mandi_data['totals']['entry_count']} entries)"
        if mandi_data.get("target_qntl"):
            mandi_label += f" | Target: {mandi_data['target_qntl']}Q | Final W: {mandi_data.get('actual_final_qntl',0)}Q | Extra: {mandi_data.get('extra_qntl',0)}Q"
        mandi_row = [[Paragraph(f"<b>{mandi_label}</b>", styles['Normal'])] + [''] * (ncols - 1)]
        mt = RLTable(mandi_row, colWidths=col_widths)
        mt.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('SPAN', (0, 0), (-1, 0)),
            ('TOPPADDING', (0, 0), (-1, 0), 4), ('BOTTOMPADDING', (0, 0), (-1, 0), 4)]))
        elements.append(mt)

        table_data = [headers]
        for entry in mandi_data["entries"]:
            table_data.append([str(v) for v in get_entry_row(entry, cols)])

        t = mandi_data["totals"]
        total_vals = get_total_row(t, cols)
        total_row = []
        for i, val in enumerate(total_vals):
            if i == 0: total_row.append("TOTAL / कुल")
            elif val is not None: total_row.append(str(val))
            else: total_row.append("")
        table_data.append(total_row)

        tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)
        first_right = next((i for i, c in enumerate(cols) if c["align"] == "right"), 2)
        cols_info = [{'header': h} for h in headers]
        style_cmds = get_pdf_table_style(len(table_data), cols_info)
        style_cmds.append(('ALIGN', (first_right, 1), (-1, -1), 'RIGHT'))
        tbl.setStyle(TableStyle(style_cmds))
        elements.append(tbl)
        elements.append(Spacer(1, 8))

    g = report_data["grand_totals"]
    grand_vals = get_total_row(g, cols)
    grand_row = []
    for i, val in enumerate(grand_vals):
        if i == 0: grand_row.append(f"GRAND TOTAL ({g['entry_count']})")
        elif i == 1: grand_row.append("")
        elif val is not None: grand_row.append(str(val))
        else: grand_row.append("")
    grand_data = [grand_row]
    gt = RLTable(grand_data, colWidths=col_widths)
    gt.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8), ('ALIGN', (first_right, 0), (-1, 0), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, 0), 4), ('BOTTOMPADDING', (0, 0), (-1, 0), 4)]))
    elements.append(gt)

    doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=agent_mandi_report_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ WEIGHT DISCREPANCY REPORT ============

@router.get("/reports/weight-discrepancy")
async def weight_discrepancy_report(
    kms_year: Optional[str] = None, season: Optional[str] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    agent: Optional[str] = None, mandi: Optional[str] = None
):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if agent: query["agent_name"] = agent
    if mandi: query["mandi_name"] = mandi
    if date_from or date_to:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to: dq["$lte"] = date_to
        if dq: query["date"] = dq

    entries = await db.entries.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    discrepancies = []
    total_diff_qntl = 0
    for e in entries:
        tp_wt = float(e.get("tp_weight", 0) or 0)
        qntl = float(e.get("qntl", 0) or 0)
        if tp_wt > 0 and qntl > 0:
            diff = round(qntl - tp_wt, 2)
            if abs(diff) > 0:
                discrepancies.append({
                    "date": e.get("date", ""),
                    "truck_no": e.get("truck_no", ""),
                    "rst_no": e.get("rst_no", ""),
                    "tp_no": e.get("tp_no", ""),
                    "agent_name": e.get("agent_name", ""),
                    "mandi_name": e.get("mandi_name", ""),
                    "tp_weight": tp_wt,
                    "qntl": qntl,
                    "diff_qntl": diff,
                    "diff_kg": round(diff * 100, 0),
                })
                total_diff_qntl += diff

    return {
        "discrepancies": discrepancies,
        "total_count": len(discrepancies),
        "total_entries_with_tp": sum(1 for e in entries if float(e.get("tp_weight", 0) or 0) > 0),
        "total_diff_qntl": round(total_diff_qntl, 2),
        "total_diff_kg": round(total_diff_qntl * 100, 0),
    }


@router.get("/reports/weight-discrepancy/excel")
async def weight_discrepancy_excel(
    kms_year: Optional[str] = None, season: Optional[str] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    agent: Optional[str] = None, mandi: Optional[str] = None
):
    data = await weight_discrepancy_report(kms_year, season, date_from, date_to, agent, mandi)
    wb = Workbook()
    ws = wb.active
    ws.title = "Weight Discrepancy"
    headers = ["Date", "Truck No", "RST", "TP No", "Agent", "Mandi", "TP Wt (Q)", "Entry QNTL", "Diff (Q)", "Diff (KG)"]
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Weight Discrepancy Report / वजन फर्क रिपोर्ट")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    # Summary
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=f"Total Discrepancies: {data['total_count']} | Total Diff: {data['total_diff_qntl']} Q ({data['total_diff_kg']} KG)")
    # Headers
    hdr_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    # Data
    red_font = Font(color="FF0000", bold=True)
    for idx, d in enumerate(data["discrepancies"], 4):
        vals = [d["date"], d["truck_no"], d["rst_no"], d["tp_no"], d["agent_name"], d["mandi_name"], d["tp_weight"], d["qntl"], d["diff_qntl"], d["diff_kg"]]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=idx, column=i, value=v)
            if i >= 9:
                cell.font = red_font
            cell.border = Border(bottom=Side(style="thin", color="DDDDDD"))
    # Totals
    tot_row = 4 + len(data["discrepancies"])
    ws.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=tot_row, column=9, value=data["total_diff_qntl"]).font = Font(bold=True, color="FF0000", size=11)
    ws.cell(row=tot_row, column=10, value=data["total_diff_kg"]).font = Font(bold=True, color="FF0000", size=11)
    # Widths
    widths = [12, 14, 8, 8, 16, 22, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fn = f"weight_discrepancy_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/reports/weight-discrepancy/pdf")
async def weight_discrepancy_pdf(
    kms_year: Optional[str] = None, season: Optional[str] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    agent: Optional[str] = None, mandi: Optional[str] = None
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from utils.branding_helper import get_pdf_header_elements_from_db

    data = await weight_discrepancy_report(kms_year, season, date_from, date_to, agent, mandi)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    sub = f"Discrepancies: {data['total_count']} | Total Diff: {data['total_diff_qntl']} Q ({data['total_diff_kg']} KG)"
    elements.extend(await get_pdf_header_elements_from_db("Weight Discrepancy Report / वजन फर्क", sub))
    elements.append(Spacer(1, 4*mm))

    headers = ["Date", "Truck", "RST", "TP", "Agent", "Mandi", "TP Wt(Q)", "QNTL", "Diff(Q)", "Diff(KG)"]
    col_widths = [28*mm, 30*mm, 18*mm, 18*mm, 30*mm, 45*mm, 22*mm, 22*mm, 22*mm, 22*mm]
    table_data = [headers]
    for d in data["discrepancies"]:
        table_data.append([d["date"], d["truck_no"], d["rst_no"], d["tp_no"], d["agent_name"], d["mandi_name"],
            f"{d['tp_weight']:.2f}", f"{d['qntl']:.2f}", f"{d['diff_qntl']:.2f}", f"{d['diff_kg']:.0f}"])
    # Totals row
    table_data.append(["TOTAL", "", "", "", "", f"{data['total_count']} entries", "", "",
        f"{data['total_diff_qntl']:.2f}", f"{data['total_diff_kg']:.0f}"])

    t = RTable(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f3460')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('ALIGN', (6, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f8f8')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTSIZE', (0, -1), (-1, -1), 7),
    ]
    # Highlight diff columns in red
    for i in range(1, len(table_data) - 1):
        style_cmds.append(('TEXTCOLOR', (8, i), (9, i), colors.HexColor('#dc2626')))
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    doc.build(elements); buffer.seek(0)
    fn = f"weight_discrepancy_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


# ============ MANDI WISE CUSTODY REGISTER ============

@router.get("/reports/mandi-custody-register")
async def mandi_custody_register(kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Date-wise mandi procurement register: each row = date, columns = mandis, total, prog.total"""
    q = {}
    if kms_year:
        q["kms_year"] = kms_year
    if season:
        q["season"] = season
    entries = await db.mill_entries.find(q, {"_id": 0, "date": 1, "mandi_name": 1, "final_w": 1, "qntl": 1}).to_list(100000)
    if date_from:
        entries = [e for e in entries if (e.get("date") or "") >= date_from]
    if date_to:
        entries = [e for e in entries if (e.get("date") or "") <= date_to]

    # Collect unique mandis
    all_mandis = sorted(set(e.get("mandi_name", "").strip() for e in entries if e.get("mandi_name", "").strip()))

    # Group by date
    from collections import defaultdict
    date_map = defaultdict(lambda: defaultdict(float))
    for e in entries:
        d = (e.get("date") or "")[:10]
        m = (e.get("mandi_name") or "").strip()
        qntl = float(e.get("final_w") or e.get("qntl") or 0)
        if d and m:
            date_map[d][m] += round(qntl, 2)

    dates = sorted(date_map.keys())
    rows = []
    prog_total = 0.0
    for d in dates:
        mandi_vals = {}
        day_total = 0.0
        for m in all_mandis:
            val = round(date_map[d].get(m, 0), 2)
            mandi_vals[m] = val
            day_total += val
        day_total = round(day_total, 2)
        prog_total = round(prog_total + day_total, 2)
        rows.append({
            "date": d,
            "mandis": mandi_vals,
            "total": day_total,
            "prog_total": prog_total,
        })

    return {"mandis": all_mandis, "rows": rows, "grand_total": prog_total}


@router.get("/reports/mandi-custody-register/pdf")
async def mandi_custody_register_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    # Get branding for company name
    branding_doc = await db.branding.find_one({}, {"_id": 0})
    company_name = (branding_doc or {}).get("company_name", "Mill Entry System")
    tagline = (branding_doc or {}).get("tagline", "")

    data = await mandi_custody_register(kms_year, season, date_from, date_to)
    mandis = data["mandis"]
    rows = data["rows"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    elements = []

    # Company Header
    company_style = ParagraphStyle("company", fontSize=14, fontName="Helvetica-Bold", alignment=1, spaceAfter=1, textColor=colors.HexColor('#1a365d'))
    elements.append(Paragraph(company_name.upper(), company_style))
    if tagline:
        elements.append(Paragraph(tagline, ParagraphStyle("tagline", fontSize=7, alignment=1, spaceAfter=2, textColor=colors.HexColor('#64748b'))))

    # Title
    elements.append(Paragraph("MANDI WISE CUSTODY REGISTER", ParagraphStyle("title", fontSize=11, fontName="Helvetica-Bold", alignment=1, spaceAfter=2, textColor=colors.HexColor('#334155'))))
    sub = f"FY: {kms_year or 'All'} | Season: {season or 'All'}"
    if date_from or date_to:
        sub += f" | Period: {date_from or 'Start'} to {date_to or 'End'}"
    elements.append(Paragraph(sub, ParagraphStyle("sub", fontSize=7, alignment=1, spaceAfter=6, textColor=colors.HexColor('#64748b'))))

    # Table
    headers = ["Date"] + mandis + ["TOTAL", "PROG. TOTAL"]
    hdr_style = ParagraphStyle("hdr", fontSize=7, leading=9, alignment=1, fontName="Helvetica-Bold", textColor=colors.white)
    cell_style = ParagraphStyle("cell", fontSize=7, leading=9, alignment=1)
    cell_left = ParagraphStyle("cellL", fontSize=7, leading=9)
    cell_bold = ParagraphStyle("cellB", fontSize=7, leading=9, alignment=1, fontName="Helvetica-Bold")

    table_data = [[Paragraph(h, hdr_style) for h in headers]]

    for r in rows:
        # Format date as DD/MM/YYYY
        d = r["date"]
        parts = d.split("-")
        date_str = f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else d
        row = [Paragraph(date_str, cell_left)]
        for m in mandis:
            v = r["mandis"].get(m, 0)
            row.append(Paragraph(f"{v:,.2f}" if v else "-", cell_style))
        row.append(Paragraph(f"{r['total']:,.2f}", cell_bold))
        row.append(Paragraph(f"{r['prog_total']:,.2f}", cell_bold))
        table_data.append(row)

    # Grand Total row
    grand_row = [Paragraph("GRAND TOTAL", ParagraphStyle("gt", fontSize=7, leading=9, fontName="Helvetica-Bold", textColor=colors.white))]
    for m in mandis:
        m_total = sum(r["mandis"].get(m, 0) for r in rows)
        grand_row.append(Paragraph(f"{m_total:,.2f}" if m_total else "-", ParagraphStyle("gtc", fontSize=7, leading=9, alignment=1, fontName="Helvetica-Bold", textColor=colors.white)))
    grand_row.append(Paragraph(f"{data['grand_total']:,.2f}", ParagraphStyle("gtt", fontSize=7, leading=9, alignment=1, fontName="Helvetica-Bold", textColor=colors.HexColor('#fbbf24'))))
    grand_row.append(Paragraph(f"{data['grand_total']:,.2f}", ParagraphStyle("gtp", fontSize=7, leading=9, alignment=1, fontName="Helvetica-Bold", textColor=colors.HexColor('#60a5fa'))))
    table_data.append(grand_row)

    # Column widths
    avail = landscape(A4)[0] - 24*mm
    date_w = 24*mm
    total_w = 22*mm
    prog_w = 24*mm
    mandi_w = max((avail - date_w - total_w - prog_w) / max(len(mandis), 1), 20*mm)
    col_widths = [date_w] + [mandi_w]*len(mandis) + [total_w, prog_w]

    last_row = len(table_data) - 1
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        # Grand total row
        ('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor('#334155')),
        ('TEXTCOLOR', (0, last_row), (-1, last_row), colors.white),
        # General
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, last_row - 1), [colors.white, colors.HexColor('#f8fafc')]),
        # TOTAL column highlight
        ('BACKGROUND', (-2, 1), (-2, last_row - 1), colors.HexColor('#fef9c3')),
        # PROG.TOTAL column highlight
        ('BACKGROUND', (-1, 1), (-1, last_row - 1), colors.HexColor('#dbeafe')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        # Header border
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1e3a5f')),
        # Grand total top border
        ('LINEABOVE', (0, last_row), (-1, last_row), 1.5, colors.HexColor('#334155')),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    fn = f"mandi_custody_register_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@router.get("/reports/mandi-custody-register/excel")
async def mandi_custody_register_excel(kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    data = await mandi_custody_register(kms_year, season, date_from, date_to)
    mandis = data["mandis"]
    rows = data["rows"]

    branding_doc = await db.branding.find_one({}, {"_id": 0})
    company_name = (branding_doc or {}).get("company_name", "Mill Entry System")
    tagline = (branding_doc or {}).get("tagline", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Mandi Custody Register"

    headers = ["Date"] + mandis + ["TOTAL", "PROG. TOTAL"]
    n_cols = len(headers)

    # Row 1: Company Name
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c1 = ws.cell(row=1, column=1, value=company_name.upper())
    c1.font = Font(bold=True, size=14, color="1a365d")
    c1.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Tagline
    if tagline:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        c2 = ws.cell(row=2, column=1, value=tagline)
        c2.font = Font(size=9, color="64748b")
        c2.alignment = Alignment(horizontal="center")

    # Row 3: Title
    title_row = 3
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=n_cols)
    c3 = ws.cell(row=title_row, column=1, value="MANDI WISE CUSTODY REGISTER")
    c3.font = Font(bold=True, size=12, color="334155")
    c3.alignment = Alignment(horizontal="center")

    # Row 4: FY / Season
    info_row = 4
    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=n_cols)
    sub = f"FY: {kms_year or 'All'} | Season: {season or 'All'}"
    if date_from or date_to:
        sub += f" | Period: {date_from or 'Start'} to {date_to or 'End'}"
    c4 = ws.cell(row=info_row, column=1, value=sub)
    c4.font = Font(size=9, color="64748b")
    c4.alignment = Alignment(horizontal="center")

    # Row 5: Empty spacer
    data_start = 6

    # Column headers
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    hdr_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    total_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    prog_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
    grand_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='cbd5e1'), right=Side(style='thin', color='cbd5e1'),
        top=Side(style='thin', color='cbd5e1'), bottom=Side(style='thin', color='cbd5e1'))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=data_start, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    for ri, r in enumerate(rows, data_start + 1):
        # Date as DD/MM/YYYY
        d = r["date"]
        parts = d.split("-")
        date_str = f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else d
        dc = ws.cell(row=ri, column=1, value=date_str)
        dc.border = thin_border
        for mi, m in enumerate(mandis, 2):
            v = r["mandis"].get(m, 0)
            c = ws.cell(row=ri, column=mi, value=round(v, 2) if v else None)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
            c.number_format = '#,##0.00'
        tc = ws.cell(row=ri, column=len(mandis) + 2, value=r["total"])
        tc.fill = total_fill
        tc.border = thin_border
        tc.alignment = Alignment(horizontal="center")
        tc.number_format = '#,##0.00'
        tc.font = Font(bold=True)
        pc = ws.cell(row=ri, column=len(mandis) + 3, value=r["prog_total"])
        pc.fill = prog_fill
        pc.border = thin_border
        pc.alignment = Alignment(horizontal="center")
        pc.number_format = '#,##0.00'
        pc.font = Font(bold=True)

    # Grand Total row
    grand_row = data_start + 1 + len(rows)
    gc = ws.cell(row=grand_row, column=1, value="GRAND TOTAL")
    gc.font = Font(bold=True, color="FFFFFF", size=9)
    gc.fill = grand_fill
    gc.border = thin_border
    for mi, m in enumerate(mandis, 2):
        m_total = sum(r["mandis"].get(m, 0) for r in rows)
        c = ws.cell(row=grand_row, column=mi, value=round(m_total, 2) if m_total else None)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = grand_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        c.number_format = '#,##0.00'
    gtc = ws.cell(row=grand_row, column=len(mandis) + 2, value=data["grand_total"])
    gtc.font = Font(bold=True, color="fbbf24")
    gtc.fill = grand_fill
    gtc.border = thin_border
    gtc.alignment = Alignment(horizontal="center")
    gtc.number_format = '#,##0.00'
    gpc = ws.cell(row=grand_row, column=len(mandis) + 3, value=data["grand_total"])
    gpc.font = Font(bold=True, color="60a5fa")
    gpc.fill = grand_fill
    gpc.border = thin_border
    gpc.alignment = Alignment(horizontal="center")
    gpc.number_format = '#,##0.00'

    # Auto-width
    ws.column_dimensions[get_column_letter(1)].width = 14
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"mandi_custody_register_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})
