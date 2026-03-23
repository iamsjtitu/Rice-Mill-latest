from fastapi import APIRouter, HTTPException
from typing import Optional
from datetime import datetime, timezone
from database import db
import uuid

router = APIRouter()

# ============ STORE ROOMS ============

@router.get("/store-rooms")
async def get_store_rooms():
    items = await db.store_rooms.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return items

@router.post("/store-rooms")
async def create_store_room(data: dict):
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Store Room name is required")
    existing = await db.store_rooms.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Store Room already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.store_rooms.insert_one(doc)
    doc.pop("_id", None)
    return doc

@router.put("/store-rooms/{room_id}")
async def update_store_room(room_id: str, data: dict):
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Store Room name is required")
    result = await db.store_rooms.update_one({"id": room_id}, {"$set": {"name": name}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    # Update references in mill_parts and stock entries
    await db.mill_parts.update_many({"store_room": room_id}, {"$set": {"store_room_name": name}})
    await db.mill_parts_stock.update_many({"store_room": room_id}, {"$set": {"store_room_name": name}})
    return {"message": "Updated", "id": room_id}

@router.delete("/store-rooms/{room_id}")
async def delete_store_room(room_id: str):
    result = await db.store_rooms.delete_one({"id": room_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    # Unassign parts from this store room
    await db.mill_parts.update_many({"store_room": room_id}, {"$set": {"store_room": "", "store_room_name": ""}})
    await db.mill_parts_stock.update_many({"store_room": room_id}, {"$set": {"store_room": "", "store_room_name": ""}})
    return {"message": "Deleted", "id": room_id}

# ============ MILL PARTS MASTER ============

@router.post("/mill-parts")
async def create_mill_part(data: dict):
    doc = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", "").strip(),
        "category": data.get("category", "General"),
        "unit": data.get("unit", "Pcs"),
        "min_stock": float(data.get("min_stock", 0)),
        "store_room": data.get("store_room", ""),
        "store_room_name": data.get("store_room_name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    if not doc["name"]:
        raise HTTPException(status_code=400, detail="Part name is required")
    existing = await db.mill_parts.find_one({"name": {"$regex": f"^{doc['name']}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Part already exists")
    await db.mill_parts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@router.get("/mill-parts")
async def get_mill_parts():
    items = await db.mill_parts.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return items

@router.delete("/mill-parts/{part_id}")
async def delete_mill_part(part_id: str):
    result = await db.mill_parts.delete_one({"id": part_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted", "id": part_id}

@router.put("/mill-parts/{part_id}")
async def update_mill_part(part_id: str, data: dict):
    existing = await db.mill_parts.find_one({"id": part_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")
    update = {}
    if "name" in data: update["name"] = data["name"].strip()
    if "category" in data: update["category"] = data["category"]
    if "unit" in data: update["unit"] = data["unit"]
    if "min_stock" in data: update["min_stock"] = float(data["min_stock"])
    if "store_room" in data: update["store_room"] = data["store_room"]
    if "store_room_name" in data: update["store_room_name"] = data["store_room_name"]
    if update:
        await db.mill_parts.update_one({"id": part_id}, {"$set": update})
    doc = await db.mill_parts.find_one({"id": part_id}, {"_id": 0})
    return doc

# ============ MILL PARTS STOCK TRANSACTIONS ============

@router.post("/mill-parts-stock")
async def create_stock_entry(data: dict):
    doc = {
        "id": str(uuid.uuid4()),
        "date": data.get("date", ""),
        "part_name": data.get("part_name", ""),
        "txn_type": data.get("txn_type", "in"),  # "in" or "used"
        "quantity": float(data.get("quantity", 0)),
        "rate": float(data.get("rate", 0)),
        "total_amount": round(float(data.get("quantity", 0)) * float(data.get("rate", 0)), 2),
        "party_name": data.get("party_name", ""),
        "bill_no": data.get("bill_no", ""),
        "remark": data.get("remark", ""),
        "store_room": data.get("store_room", ""),
        "store_room_name": data.get("store_room_name", ""),
        "kms_year": data.get("kms_year", ""),
        "season": data.get("season", ""),
        "created_by": data.get("created_by", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    if not doc["part_name"] or doc["quantity"] <= 0:
        raise HTTPException(status_code=400, detail="Part name and quantity required")
    await db.mill_parts_stock.insert_one(doc)
    doc.pop("_id", None)

    # Update part's store_room in master when stock-in has a store_room selected
    if doc["store_room"] and doc["part_name"]:
        await db.mill_parts.update_many(
            {"name": doc["part_name"]},
            {"$set": {"store_room": doc["store_room"], "store_room_name": doc["store_room_name"]}}
        )

    # Auto-create local party account entry for purchases (txn_type=in) with party
    if doc["txn_type"] == "in" and doc["party_name"] and doc["total_amount"] > 0:
        lp = {
            "id": str(uuid.uuid4()),
            "date": doc["date"],
            "party_name": doc["party_name"],
            "txn_type": "debit",
            "amount": doc["total_amount"],
            "description": f"{doc['part_name']} x{doc['quantity']} @ Rs.{doc['rate']}",
            "source_type": "mill_part",
            "reference": f"mill_part:{doc['id'][:8]}",
            "kms_year": doc.get("kms_year", ""),
            "season": doc.get("season", ""),
            "created_by": doc.get("created_by", "system"),
            "linked_stock_id": doc["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.local_party_accounts.insert_one(lp)

        # Auto create Cash Book Jama entry (purchase from local party via mill part)
        cb = {
            "id": str(uuid.uuid4()), "date": doc["date"],
            "account": "ledger", "txn_type": "jama",
            "category": doc["party_name"], "party_type": "Local Party",
            "description": f"Mill Part: {doc['part_name']} x{doc['quantity']} @ Rs.{doc['rate']} - {doc['party_name']}",
            "amount": doc["total_amount"], "reference": f"lp_mill_part:{doc['id'][:8]}",
            "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
            "created_by": doc.get("created_by", "system"), "linked_local_party_id": lp["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(cb)

    return doc

@router.get("/mill-parts-stock")
async def get_stock_entries(part_name: Optional[str] = None, txn_type: Optional[str] = None,
                            kms_year: Optional[str] = None, season: Optional[str] = None,
                            party_name: Optional[str] = None,
                            date_from: Optional[str] = None, date_to: Optional[str] = None):
    query = {}
    if part_name: query["part_name"] = part_name
    if txn_type: query["txn_type"] = txn_type
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    items = await db.mill_parts_stock.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    return items

@router.delete("/mill-parts-stock/{entry_id}")
async def delete_stock_entry(entry_id: str):
    # Also remove linked local party entry
    await db.local_party_accounts.delete_many({"linked_stock_id": entry_id})
    result = await db.mill_parts_stock.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted", "id": entry_id}

@router.put("/mill-parts-stock/{entry_id}")
async def update_stock_entry(entry_id: str, data: dict):
    existing = await db.mill_parts_stock.find_one({"id": entry_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")
    update = {
        "date": data.get("date", existing.get("date", "")),
        "part_name": data.get("part_name", existing.get("part_name", "")),
        "txn_type": data.get("txn_type", existing.get("txn_type", "in")),
        "quantity": float(data.get("quantity", existing.get("quantity", 0))),
        "rate": float(data.get("rate", existing.get("rate", 0))),
        "party_name": data.get("party_name", existing.get("party_name", "")),
        "bill_no": data.get("bill_no", existing.get("bill_no", "")),
        "remark": data.get("remark", existing.get("remark", "")),
        "store_room": data.get("store_room", existing.get("store_room", "")),
        "store_room_name": data.get("store_room_name", existing.get("store_room_name", "")),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    update["total_amount"] = round(update["quantity"] * update["rate"], 2)
    await db.mill_parts_stock.update_one({"id": entry_id}, {"$set": update})
    # Update linked local party entry
    await db.local_party_accounts.delete_many({"linked_stock_id": entry_id})
    if update["txn_type"] == "in" and update["party_name"] and update["total_amount"] > 0:
        lp = {
            "id": str(uuid.uuid4()), "date": update["date"],
            "party_name": update["party_name"], "txn_type": "debit",
            "amount": update["total_amount"],
            "description": f"{update['part_name']} x{update['quantity']} @ Rs.{update['rate']}",
            "source_type": "mill_part",
            "reference": f"mill_part:{entry_id[:8]}",
            "kms_year": data.get("kms_year", existing.get("kms_year", "")),
            "season": data.get("season", existing.get("season", "")),
            "created_by": data.get("created_by", "system"), "linked_stock_id": entry_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.local_party_accounts.insert_one(lp)

        # Auto create Cash Book Jama entry
        cb = {
            "id": str(uuid.uuid4()), "date": update["date"],
            "account": "ledger", "txn_type": "jama",
            "category": update["party_name"], "party_type": "Local Party",
            "description": f"Mill Part: {update['part_name']} x{update['quantity']} @ Rs.{update['rate']} - {update['party_name']}",
            "amount": update["total_amount"], "reference": f"lp_mill_part:{entry_id[:8]}",
            "kms_year": data.get("kms_year", existing.get("kms_year", "")),
            "season": data.get("season", existing.get("season", "")),
            "created_by": data.get("created_by", "system"), "linked_local_party_id": lp["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(cb)

    updated = await db.mill_parts_stock.find_one({"id": entry_id}, {"_id": 0})
    return updated

# ============ STOCK SUMMARY ============

@router.get("/mill-parts/summary")
async def get_stock_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.mill_parts_stock.find(query, {"_id": 0}).to_list(10000)
    parts = await db.mill_parts.find({}, {"_id": 0}).to_list(1000)

    # Compute opening stock from previous FY
    opening_stock = {}
    if kms_year:
        fy_parts = kms_year.split('-')
        if len(fy_parts) == 2:
            try:
                prev_fy = f"{int(fy_parts[0])-1}-{int(fy_parts[1])-1}"
                prev_query = {"kms_year": prev_fy}
                if season: prev_query["season"] = season
                prev_txns = await db.mill_parts_stock.find(prev_query, {"_id": 0}).to_list(10000)
                for t in prev_txns:
                    pn = t.get("part_name", "")
                    if pn not in opening_stock:
                        opening_stock[pn] = 0
                    if t.get("txn_type") == "in":
                        opening_stock[pn] += t.get("quantity", 0)
                    else:
                        opening_stock[pn] -= t.get("quantity", 0)
            except (ValueError, IndexError):
                pass

    summary = {}
    for p in parts:
        ob = round(opening_stock.get(p["name"], 0), 2)
        summary[p["name"]] = {"part_name": p["name"], "category": p.get("category", ""),
            "unit": p.get("unit", "Pcs"), "min_stock": p.get("min_stock", 0),
            "store_room": p.get("store_room", ""), "store_room_name": p.get("store_room_name", ""),
            "opening_stock": ob,
            "stock_in": 0, "stock_used": 0, "current_stock": 0,
            "total_purchase_amount": 0, "parties": {}}

    for t in txns:
        pn = t.get("part_name", "")
        if pn not in summary:
            ob = round(opening_stock.get(pn, 0), 2)
            summary[pn] = {"part_name": pn, "category": "", "unit": "Pcs", "min_stock": 0,
                "opening_stock": ob,
                "stock_in": 0, "stock_used": 0, "current_stock": 0,
                "total_purchase_amount": 0, "parties": {}}
        if t.get("txn_type") == "in":
            summary[pn]["stock_in"] += t.get("quantity", 0)
            summary[pn]["total_purchase_amount"] += t.get("total_amount", 0)
            party = t.get("party_name", "")
            if party:
                if party not in summary[pn]["parties"]:
                    summary[pn]["parties"][party] = {"qty": 0, "amount": 0}
                summary[pn]["parties"][party]["qty"] += t.get("quantity", 0)
                summary[pn]["parties"][party]["amount"] += t.get("total_amount", 0)
        else:
            summary[pn]["stock_used"] += t.get("quantity", 0)

    result = []
    for pn, s in summary.items():
        s["stock_in"] = round(s["stock_in"], 2)
        s["stock_used"] = round(s["stock_used"], 2)
        s["current_stock"] = round(s["opening_stock"] + s["stock_in"] - s["stock_used"], 2)
        s["total_purchase_amount"] = round(s["total_purchase_amount"], 2)
        s["parties"] = [{"name": k, **v} for k, v in s["parties"].items()]
        result.append(s)

    result.sort(key=lambda x: x["part_name"])
    return result


# ============ STORE ROOM WISE REPORT ============

@router.get("/mill-parts/store-room-report")
async def get_store_room_report(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get inventory grouped by store room."""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.mill_parts_stock.find(query, {"_id": 0}).to_list(10000)
    parts = await db.mill_parts.find({}, {"_id": 0}).to_list(1000)

    # Build part -> store_room mapping from master
    part_store = {}
    for p in parts:
        part_store[p["name"]] = {
            "store_room": p.get("store_room", ""),
            "store_room_name": p.get("store_room_name", ""),
            "category": p.get("category", ""),
            "unit": p.get("unit", "Pcs"),
            "min_stock": p.get("min_stock", 0),
        }

    # Compute stock per part
    stock = {}
    for t in txns:
        pn = t.get("part_name", "")
        if pn not in stock:
            info = part_store.get(pn, {})
            stock[pn] = {
                "part_name": pn,
                "store_room": info.get("store_room", t.get("store_room", "")),
                "store_room_name": info.get("store_room_name", t.get("store_room_name", "")),
                "category": info.get("category", ""),
                "unit": info.get("unit", "Pcs"),
                "stock_in": 0, "stock_used": 0, "current_stock": 0,
            }
        if t.get("txn_type") == "in":
            stock[pn]["stock_in"] += t.get("quantity", 0)
        else:
            stock[pn]["stock_used"] += t.get("quantity", 0)

    # Include parts from master that have no txns
    for p in parts:
        if p["name"] not in stock:
            stock[p["name"]] = {
                "part_name": p["name"],
                "store_room": p.get("store_room", ""),
                "store_room_name": p.get("store_room_name", ""),
                "category": p.get("category", ""),
                "unit": p.get("unit", "Pcs"),
                "stock_in": 0, "stock_used": 0, "current_stock": 0,
            }

    # Finalize and group by store room
    room_groups = {}
    for pn, s in stock.items():
        s["stock_in"] = round(s["stock_in"], 2)
        s["stock_used"] = round(s["stock_used"], 2)
        s["current_stock"] = round(s["stock_in"] - s["stock_used"], 2)
        room_id = s["store_room"] or "__unassigned__"
        room_name = s["store_room_name"] or "Unassigned"
        if room_id not in room_groups:
            room_groups[room_id] = {"store_room_id": room_id, "store_room_name": room_name, "parts": []}
        room_groups[room_id]["parts"].append(s)

    # Sort parts within each room
    for g in room_groups.values():
        g["parts"].sort(key=lambda x: x["part_name"])

    result = sorted(room_groups.values(), key=lambda x: x["store_room_name"])
    return result


@router.get("/mill-parts/store-room-report/excel")
async def export_store_room_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    report = await get_store_room_report(kms_year, season)
    wb = Workbook()
    ws = wb.active
    ws.title = "Store Room Report"
    title = "Store Room-wise Inventory Report"
    if kms_year:
        title += f" - {kms_year}"
    if season:
        title += f" ({season})"
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color='1a365d')
    ws['A1'].alignment = Alignment(horizontal='center')

    hdr_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    room_fill = PatternFill(start_color='0e7490', end_color='0e7490', fill_type='solid')
    room_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin', color='cbd5e1'), right=Side(style='thin', color='cbd5e1'),
        top=Side(style='thin', color='cbd5e1'), bottom=Side(style='thin', color='cbd5e1')
    )
    alt_fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
    low_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')

    row = 3
    grand_total_in = 0
    grand_total_used = 0

    for group in report:
        # Store Room header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=f"  {group['store_room_name']} ({len(group['parts'])} parts)")
        c.fill = room_fill
        c.font = room_font
        c.alignment = Alignment(horizontal='left', vertical='center')
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).border = thin_border
        row += 1

        # Column headers
        headers = ['Part Name', 'Category', 'Unit', 'Stock In', 'Used', 'Current Stock']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
        row += 1

        room_total_in = 0
        room_total_used = 0
        for idx, p in enumerate(group['parts']):
            vals = [p['part_name'], p.get('category', ''), p.get('unit', 'Pcs'),
                    p['stock_in'], p['stock_used'], p['current_stock']]
            room_total_in += p['stock_in']
            room_total_used += p['stock_used']
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.border = thin_border
                c.font = Font(size=9)
                if p['current_stock'] <= 0:
                    c.fill = low_fill
                elif idx % 2 == 1:
                    c.fill = alt_fill
            row += 1

        # Room subtotal
        ws.cell(row=row, column=1, value="  Subtotal").font = Font(bold=True, size=9, color='0e7490')
        ws.cell(row=row, column=4, value=round(room_total_in, 2)).font = Font(bold=True, size=9)
        ws.cell(row=row, column=5, value=round(room_total_used, 2)).font = Font(bold=True, size=9)
        ws.cell(row=row, column=6, value=round(room_total_in - room_total_used, 2)).font = Font(bold=True, size=9)
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).border = thin_border
        row += 2
        grand_total_in += room_total_in
        grand_total_used += room_total_used

    # Grand total
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=11, color='1a365d')
    ws.cell(row=row, column=4, value=round(grand_total_in, 2)).font = Font(bold=True, size=11)
    ws.cell(row=row, column=5, value=round(grand_total_used, 2)).font = Font(bold=True, size=11)
    ws.cell(row=row, column=6, value=round(grand_total_in - grand_total_used, 2)).font = Font(bold=True, size=11)
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).border = thin_border

    widths = [22, 14, 8, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=store_room_report_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/mill-parts/store-room-report/pdf")
async def export_store_room_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    report = await get_store_room_report(kms_year, season)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    title_text = "Store Room-wise Inventory Report"
    if kms_year:
        title_text += f" - {kms_year}"
    if season:
        title_text += f" ({season})"
    elements = [Paragraph(title_text, styles['Title']), Spacer(1, 12)]

    col_widths = [140, 80, 50, 70, 70, 90]

    for group in report:
        # Room header
        room_style = ParagraphStyle('room', parent=styles['Heading3'], textColor=colors.HexColor('#0e7490'))
        elements.append(Paragraph(f"{group['store_room_name']} ({len(group['parts'])} parts)", room_style))
        elements.append(Spacer(1, 4))

        data = [['Part Name', 'Category', 'Unit', 'Stock In', 'Used', 'Current Stock']]
        room_in = 0
        room_used = 0
        for p in group['parts']:
            data.append([p['part_name'], p.get('category', ''), p.get('unit', 'Pcs'),
                         str(p['stock_in']), str(p['stock_used']), str(p['current_stock'])])
            room_in += p['stock_in']
            room_used += p['stock_used']

        data.append(['Subtotal', '', '', str(round(room_in, 2)), str(round(room_used, 2)),
                      str(round(room_in - room_used, 2))])

        t = RTable(data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0f2fe')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (3, 0), (5, -1), 'RIGHT'),
        ]
        # Highlight zero/negative stock rows
        for ri, p in enumerate(group['parts'], 1):
            if p['current_stock'] <= 0:
                style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#fee2e2')))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 14))

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=store_room_report_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ STOCK EXPORT ============

@router.get("/mill-parts/summary/excel")
async def export_stock_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    summary = await get_stock_summary(kms_year, season)
    wb = Workbook()
    ws = wb.active
    ws.title = "Mill Parts Stock"
    ws.merge_cells('A1:I1')
    ws['A1'] = f"Mill Parts Stock Summary{' - ' + kms_year if kms_year else ''}{' - ' + season if season else ''}"
    ws['A1'].font = Font(bold=True, size=14, color='1a365d')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Part Name', 'Category', 'Store Room', 'Unit', 'Stock In', 'Stock Used', 'Current Stock', 'Purchase Amount (Rs)', 'Parties']
    hdr_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(left=Side(style='thin', color='cbd5e1'), right=Side(style='thin', color='cbd5e1'), top=Side(style='thin', color='cbd5e1'), bottom=Side(style='thin', color='cbd5e1'))
    alt_fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    total_purchase = 0
    for idx, s in enumerate(summary):
        row = idx + 4
        vals = [s["part_name"], s["category"], s.get("store_room_name", ""), s["unit"], s["stock_in"], s["stock_used"], s["current_stock"], s["total_purchase_amount"], ', '.join(p['name'] for p in s.get('parties', []))]
        total_purchase += s["total_purchase_amount"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = thin_border
            c.font = Font(size=9)
            if idx % 2 == 1: c.fill = alt_fill

    # Totals row
    tr = len(summary) + 4
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, size=10, color='1a365d')
    ws.cell(row=tr, column=8, value=total_purchase).font = Font(bold=True, size=10, color='1a365d')
    for ci in range(1, 10):
        ws.cell(row=tr, column=ci).border = thin_border

    widths = [20, 14, 14, 8, 12, 12, 14, 18, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mill_parts_stock_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@router.get("/mill-parts/summary/pdf")
async def export_stock_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    summary = await get_stock_summary(kms_year, season)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    title_text = f"Mill Parts Stock Summary"
    if kms_year: title_text += f" - {kms_year}"
    if season: title_text += f" ({season})"
    elements = [Paragraph(title_text, styles['Title']), Spacer(1, 12)]

    data = [['Part', 'Category', 'Store Room', 'Unit', 'In', 'Used', 'Stock', 'Amount (Rs)', 'Parties']]
    total_purchase = 0
    for s in summary:
        total_purchase += s["total_purchase_amount"]
        data.append([s["part_name"], s["category"], s.get("store_room_name", ""), s["unit"], s["stock_in"], s["stock_used"], s["current_stock"],
            f'Rs.{s["total_purchase_amount"]:,.0f}', ', '.join(p['name'] for p in s.get('parties', []))])
    data.append(['TOTAL', '', '', '', '', '', '', f'Rs.{total_purchase:,.0f}', ''])

    col_widths = [85, 55, 65, 35, 40, 40, 50, 75, 100]
    t = RTable(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0f2fe')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (4, 0), (7, -1), 'RIGHT'),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=mill_parts_stock_{datetime.now().strftime('%Y%m%d')}.pdf"})

@router.get("/mill-parts-stock/export/excel")
async def export_transactions_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                                     part_name: Optional[str] = None, txn_type: Optional[str] = None,
                                     date_from: Optional[str] = None, date_to: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    query = {}
    if part_name: query["part_name"] = part_name
    if txn_type: query["txn_type"] = txn_type
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    items = await db.mill_parts_stock.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Parts Transactions"
    title = "Mill Parts Transactions"
    if part_name: title += f" - {part_name}"
    if date_from or date_to: title += f" ({date_from or '...'} to {date_to or '...'})"
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color='1a365d')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Date', 'Part Name', 'Store Room', 'Type', 'Qty', 'Rate', 'Amount (Rs)', 'Party', 'Bill No', 'Remark']
    hdr_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(left=Side(style='thin', color='cbd5e1'), right=Side(style='thin', color='cbd5e1'), top=Side(style='thin', color='cbd5e1'), bottom=Side(style='thin', color='cbd5e1'))
    alt_fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
    in_fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
    used_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.border = thin_border; c.alignment = Alignment(horizontal='center')

    total_in_amt = 0
    total_in_qty = 0
    total_used_qty = 0
    for idx, t in enumerate(items):
        row = idx + 4
        typ = 'IN' if t.get('txn_type') == 'in' else 'USED'
        amt = t.get('total_amount') or t.get('total_cost') or 0
        qty = t.get('quantity', 0)
        if t.get('txn_type') == 'in':
            total_in_amt += amt
            total_in_qty += qty
        else:
            total_used_qty += qty
        vals = [t.get('date',''), t.get('part_name',''), t.get('store_room_name','') or '', typ, qty, t.get('rate',0), amt, t.get('party_name',''), t.get('bill_no',''), t.get('remark','')]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = thin_border; c.font = Font(size=9)
            if ci == 4: c.fill = in_fill if typ == 'IN' else used_fill

    # Totals row
    tr = len(items) + 4
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, size=10, color='1a365d')
    ws.cell(row=tr, column=4, value=f"In:{total_in_qty} / Used:{total_used_qty}").font = Font(bold=True, size=9)
    ws.cell(row=tr, column=7, value=total_in_amt).font = Font(bold=True, size=10, color='1a365d')
    for ci in range(1, 11):
        ws.cell(row=tr, column=ci).border = thin_border

    widths = [12, 18, 12, 8, 8, 10, 14, 18, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mill_parts_transactions_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@router.get("/mill-parts-stock/export/pdf")
async def export_transactions_pdf(kms_year: Optional[str] = None, season: Optional[str] = None,
                                   part_name: Optional[str] = None, txn_type: Optional[str] = None,
                                   date_from: Optional[str] = None, date_to: Optional[str] = None):
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    query = {}
    if part_name: query["part_name"] = part_name
    if txn_type: query["txn_type"] = txn_type
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    items = await db.mill_parts_stock.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    title = "Mill Parts Transactions"
    if part_name: title += f" - {part_name}"
    subtitle_parts = []
    if date_from or date_to: subtitle_parts.append(f"Date: {date_from or '...'} to {date_to or '...'}")
    if kms_year: subtitle_parts.append(f"KMS: {kms_year}")
    if season: subtitle_parts.append(f"Season: {season}")
    elements = [Paragraph(title, styles['Title'])]
    if subtitle_parts:
        elements.append(Paragraph(' | '.join(subtitle_parts), styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [['Date', 'Part Name', 'Store Room', 'Type', 'Qty', 'Rate', 'Amount (Rs)', 'Party', 'Bill No', 'Remark']]
    total_amt = 0
    for t in items:
        typ = 'IN' if t.get('txn_type') == 'in' else 'USED'
        amt = t.get('total_amount') or t.get('total_cost') or 0
        if t.get('txn_type') == 'in': total_amt += amt
        data.append([t.get('date',''), t.get('part_name',''), t.get('store_room_name','') or '', typ, t.get('quantity',0), t.get('rate',0),
            f'Rs.{amt:,.0f}' if amt else '-', t.get('party_name',''), t.get('bill_no',''), t.get('remark','')])
    data.append(['TOTAL', '', '', '', '', '', f'Rs.{total_amt:,.0f}', '', '', ''])

    col_widths = [55, 75, 55, 30, 30, 40, 60, 70, 50, 70]
    tbl = RTable(data, colWidths=col_widths, repeatRows=1)

    # Style with colored IN/USED rows
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0f2fe')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (4, 0), (6, -1), 'RIGHT'),
    ]
    # Alternating row colors and IN/USED highlighting
    for i, t in enumerate(items, 1):
        bg = colors.HexColor('#f0fdf4') if t.get('txn_type') == 'in' else colors.HexColor('#fef2f2')
        if i % 2 == 0: bg = colors.HexColor('#dcfce7') if t.get('txn_type') == 'in' else colors.HexColor('#fee2e2')
        style_cmds.append(('BACKGROUND', (0, i), (-1, i), bg))

    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=mill_parts_transactions_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ SINGLE PART SUMMARY EXPORT ============

@router.get("/mill-parts/part-summary/excel")
async def export_part_summary_excel(part_name: str, kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    if not part_name:
        raise HTTPException(status_code=400, detail="part_name required")

    query = {"part_name": part_name}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.mill_parts_stock.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    part_info = await db.mill_parts.find_one({"name": part_name}, {"_id": 0})

    stock_in = sum(t.get("quantity", 0) for t in txns if t.get("txn_type") == "in")
    stock_used = sum(t.get("quantity", 0) for t in txns if t.get("txn_type") != "in")
    purchase_amt = sum(t.get("total_amount", 0) for t in txns if t.get("txn_type") == "in")
    unit = (part_info or {}).get("unit", "Pcs")
    category = (part_info or {}).get("category", "General")

    parties = {}
    for t in txns:
        if t.get("txn_type") == "in" and t.get("party_name"):
            pn = t["party_name"]
            if pn not in parties:
                parties[pn] = {"qty": 0, "amount": 0}
            parties[pn]["qty"] += t.get("quantity", 0)
            parties[pn]["amount"] += t.get("total_amount", 0)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{part_name} Summary"
    thin = Border(left=Side(style='thin', color='cbd5e1'), right=Side(style='thin', color='cbd5e1'),
                  top=Side(style='thin', color='cbd5e1'), bottom=Side(style='thin', color='cbd5e1'))
    hdr_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    alt_fill = PatternFill(start_color='f0f7ff', end_color='f0f7ff', fill_type='solid')

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{part_name} - Part Summary"
    ws['A1'].font = Font(bold=True, size=16, color='1a365d')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:F2')
    store_room_name = (part_info or {}).get("store_room_name", "")
    ws['A2'] = f"Category: {category} | Unit: {unit} | Store Room: {store_room_name or 'N/A'} | {kms_year or ''} {season or ''}"
    ws['A2'].font = Font(size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Stock Overview
    r = 4
    ws.cell(r, 1, "STOCK OVERVIEW").font = Font(bold=True, size=12, color='1a365d')
    r = 5
    for h in ["Stock In", "Stock Used", "Current Stock", "Total Purchase"]:
        c = ws.cell(r, ["Stock In", "Stock Used", "Current Stock", "Total Purchase"].index(h) + 1, h)
        c.fill = hdr_fill; c.font = hdr_font; c.border = thin; c.alignment = Alignment(horizontal='center')
    r = 6
    for i, v in enumerate([round(stock_in, 2), round(stock_used, 2), round(stock_in - stock_used, 2), f"Rs.{round(purchase_amt, 2):,.2f}"]):
        c = ws.cell(r, i + 1, v)
        c.font = Font(bold=True, size=11)
        c.border = thin
        c.alignment = Alignment(horizontal='center')

    # Party-wise Purchase
    if parties:
        r = 8
        ws.cell(r, 1, "PARTY-WISE PURCHASE").font = Font(bold=True, size=12, color='1a365d')
        r = 9
        for i, h in enumerate(["Party Name", "Quantity", "Amount (Rs.)"]):
            c = ws.cell(r, i + 1, h)
            c.fill = hdr_fill; c.font = hdr_font; c.border = thin; c.alignment = Alignment(horizontal='center')
        r = 10
        for idx, (pname, pdata) in enumerate(sorted(parties.items())):
            ws.cell(r, 1, pname).border = thin
            ws.cell(r, 2, round(pdata["qty"], 2)).border = thin
            ws.cell(r, 3, round(pdata["amount"], 2)).border = thin
            if idx % 2 == 1:
                for ci in range(1, 4):
                    ws.cell(r, ci).fill = alt_fill
            ws.cell(r, 1).font = Font(size=10, bold=True)
            ws.cell(r, 2).font = Font(size=10)
            ws.cell(r, 3).font = Font(size=10)
            r += 1
        # Party total
        ws.cell(r, 1, "TOTAL").font = Font(bold=True, size=10, color='1a365d')
        ws.cell(r, 2, round(sum(p["qty"] for p in parties.values()), 2)).font = Font(bold=True, size=10)
        ws.cell(r, 3, round(sum(p["amount"] for p in parties.values()), 2)).font = Font(bold=True, size=10)
        for ci in range(1, 4):
            ws.cell(r, ci).border = thin
        r += 1

    # Transactions
    tr = r + 1 if parties else 8
    ws.cell(tr, 1, "ALL TRANSACTIONS").font = Font(bold=True, size=12, color='1a365d')
    tr += 1
    txn_headers = ["Date", "Type", "Qty", "Rate", "Amount (Rs.)", "Party", "Bill No", "Remark"]
    for i, h in enumerate(txn_headers):
        c = ws.cell(tr, i + 1, h)
        c.fill = hdr_fill; c.font = hdr_font; c.border = thin; c.alignment = Alignment(horizontal='center')
    tr += 1
    in_fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
    used_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
    for idx, t in enumerate(txns):
        typ = "IN" if t.get("txn_type") == "in" else "USED"
        amt = t.get("total_amount") or 0
        vals = [t.get("date",""), typ, t.get("quantity",0), t.get("rate",0), amt, t.get("party_name",""), t.get("bill_no",""), t.get("remark","")]
        for ci, v in enumerate(vals):
            c = ws.cell(tr, ci + 1, v)
            c.border = thin
            c.font = Font(size=9)
            if ci == 1:
                c.fill = in_fill if typ == "IN" else used_fill
        tr += 1

    widths = [12, 8, 8, 10, 14, 18, 12, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{part_name.replace(' ', '_')}_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/mill-parts/part-summary/pdf")
async def export_part_summary_pdf(part_name: str, kms_year: Optional[str] = None, season: Optional[str] = None):
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    if not part_name:
        raise HTTPException(status_code=400, detail="part_name required")

    query = {"part_name": part_name}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.mill_parts_stock.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    part_info = await db.mill_parts.find_one({"name": part_name}, {"_id": 0})

    stock_in = sum(t.get("quantity", 0) for t in txns if t.get("txn_type") == "in")
    stock_used = sum(t.get("quantity", 0) for t in txns if t.get("txn_type") != "in")
    purchase_amt = sum(t.get("total_amount", 0) for t in txns if t.get("txn_type") == "in")
    unit = (part_info or {}).get("unit", "Pcs")
    category = (part_info or {}).get("category", "General")

    parties = {}
    for t in txns:
        if t.get("txn_type") == "in" and t.get("party_name"):
            pn = t["party_name"]
            if pn not in parties:
                parties[pn] = {"qty": 0, "amount": 0}
            parties[pn]["qty"] += t.get("quantity", 0)
            parties[pn]["amount"] += t.get("total_amount", 0)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=25, bottomMargin=25)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1a365d'), spaceAfter=4)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666666'), alignment=1, spaceAfter=10)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#1a365d'), spaceBefore=14, spaceAfter=6)

    store_room_name = (part_info or {}).get("store_room_name", "")
    elements = [
        Paragraph(f"{part_name} - Part Summary", title_style),
        Paragraph(f"Category: {category} | Unit: {unit} | Store Room: {store_room_name or 'N/A'} | {kms_year or ''} {season or ''}", subtitle_style),
    ]

    # Stock Overview
    elements.append(Paragraph("Stock Overview", section_style))
    overview_data = [
        ["Stock In", "Stock Used", "Current Stock", "Total Purchase"],
        [f"{round(stock_in, 2)} {unit}", f"{round(stock_used, 2)} {unit}",
         f"{round(stock_in - stock_used, 2)} {unit}", f"Rs.{round(purchase_amt, 2):,.2f}"],
    ]
    ot = RTable(overview_data, colWidths=[130, 130, 130, 150])
    ot.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 11),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f0f7ff')),
    ]))
    elements.append(ot)

    # Party-wise Purchase
    if parties:
        elements.append(Paragraph("Party-wise Purchase", section_style))
        party_data = [["Party Name", f"Quantity ({unit})", "Amount (Rs.)"]]
        for pname in sorted(parties):
            pdata = parties[pname]
            party_data.append([pname, round(pdata["qty"], 2), f"Rs.{round(pdata['amount'], 2):,.2f}"])
        party_data.append(["TOTAL", round(sum(p["qty"] for p in parties.values()), 2),
                           f"Rs.{round(sum(p['amount'] for p in parties.values()), 2):,.2f}"])

        pt = RTable(party_data, colWidths=[180, 100, 130])
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0f2fe')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]
        for i in range(1, len(party_data) - 1):
            if i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fafc')))
        pt.setStyle(TableStyle(style_cmds))
        elements.append(pt)

    # Transactions
    if txns:
        elements.append(Paragraph("All Transactions", section_style))
        txn_data = [["Date", "Type", "Qty", "Rate", "Amount (Rs.)", "Party", "Bill No", "Remark"]]
        for t in txns:
            typ = "IN" if t.get("txn_type") == "in" else "USED"
            amt = t.get("total_amount") or 0
            txn_data.append([t.get("date",""), typ, t.get("quantity",0), t.get("rate",0),
                             f"Rs.{amt:,.0f}" if amt else "-", t.get("party_name",""),
                             t.get("bill_no",""), (t.get("remark","") or "")[:20]])

        col_widths = [58, 35, 35, 45, 65, 100, 55, 80]
        tt = RTable(txn_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ALIGN', (2, 0), (4, -1), 'RIGHT'),
        ]
        for i, t in enumerate(txns, 1):
            bg = colors.HexColor('#f0fdf4') if t.get('txn_type') == 'in' else colors.HexColor('#fef2f2')
            if i % 2 == 0:
                bg = colors.HexColor('#dcfce7') if t.get('txn_type') == 'in' else colors.HexColor('#fee2e2')
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), bg))
        tt.setStyle(TableStyle(style_cmds))
        elements.append(tt)

    doc.build(elements)
    buf.seek(0)
    fname = f"{part_name.replace(' ', '_')}_summary_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})
