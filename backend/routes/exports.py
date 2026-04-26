from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import round_amount
from utils.date_format import fmt_date
from utils.commission import capped_tp_for_commission
import uuid, io, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()


async def get_company_name():
    branding = await db.branding.find_one({}, {"_id": 0})
    if branding:
        return branding.get("company_name", "NAVKAR AGRO"), branding.get("tagline", "")
    return "NAVKAR AGRO", "JOLKO, KESINGA"

@router.get("/dashboard/monthly-trend")
async def get_monthly_trend(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get monthly totals for trend chart"""
    match_query = {}
    if kms_year:
        match_query["kms_year"] = kms_year
    if season:
        match_query["season"] = season
    
    pipeline = []
    if match_query:
        pipeline.append({"$match": match_query})
    
    pipeline.extend([
        {
            "$addFields": {
                "month": {"$substr": ["$date", 0, 7]}  # Extract YYYY-MM
            }
        },
        {
            "$group": {
                "_id": "$month",
                "total_qntl": {"$sum": "$qntl"},
                "total_final_w_kg": {"$sum": "$final_w"},
                "total_entries": {"$sum": 1},
                "total_bag": {"$sum": "$bag"}
            }
        },
        {"$sort": {"_id": 1}}
    ])
    
    results = await db.mill_entries.aggregate(pipeline).to_list(12)
    
    monthly_data = []
    for r in results:
        if r["_id"]:
            monthly_data.append({
                "month": r["_id"],
                "total_qntl": round(r["total_qntl"], 2),
                "total_final_w": round(r["total_final_w_kg"] / 100, 2),
                "total_entries": r["total_entries"],
                "total_bag": r["total_bag"]
            })
    
    return {"monthly_data": monthly_data}


@router.get("/export/dashboard-pdf")
async def export_dashboard_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, filter: Optional[str] = None):
    """Export Dashboard PDF with optional filter (all, stock, or mandi_name)"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from utils.export_helpers import (
        get_pdf_styles, get_pdf_summary_banner, get_pdf_section_band, STAT_COLORS,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer, CondPageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    show_stock = (not filter) or filter == "all" or filter == "stock"
    show_targets = (not filter) or filter != "stock"
    target_mandi = filter if filter and filter not in ("all", "stock") else None

    buffer = io.BytesIO()
    # leftMargin=8 + frame's 6pt internal pad = 14pt effective from page edge → page-centered content (PAGE_W=180mm).
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=8*mm, rightMargin=8*mm, topMargin=12*mm, bottomMargin=12*mm)
    elements = []; styles = get_pdf_styles()
    PAGE_W = A4[0] - 16*mm  # ~580pt content width

    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    company, tagline = await get_company_name()
    filter_label = "All" if not filter or filter == "all" else ("Stock Only" if filter == "stock" else filter)

    # Sub-header (one compact line — header helper already shows tagline)
    sub = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=8.5, textColor=colors.HexColor('#475569'), alignment=TA_CENTER)
    elements.append(Paragraph(
        f"<b>DASHBOARD REPORT</b> &nbsp;|&nbsp; FY: {kms_year or 'All'} &nbsp;|&nbsp; Season: {season or 'All'} "
        f"&nbsp;|&nbsp; Filter: {filter_label} &nbsp;|&nbsp; {datetime.now().strftime('%d-%b-%Y %H:%M')}",
        sub,
    ))
    elements.append(Spacer(1, 4*mm))

    # ---- COMPUTE STOCK + TARGETS DATA ONCE (shared between KPI banner and sections) ----
    # Paddy IN
    pipe = [{"$match": query}, {"$group": {"_id": None, "total": {"$sum": "$final_w"}}}]
    mill_res = await db.mill_entries.aggregate(pipe).to_list(1)
    cmr_paddy = round((mill_res[0]["total"] / 100) if mill_res else 0, 2)
    pvt_query = dict(query); pvt_query["source"] = {"$ne": "agent_extra"}
    pvt_entries = await db.private_paddy.find(pvt_query, {"_id": 0, "qntl": 1, "bag": 1}).to_list(5000)
    pvt_paddy = round(sum(e.get("qntl", 0) - e.get("bag", 0) / 100 for e in pvt_entries), 2)
    milling_entries = await db.milling_entries.find(dict(query), {"_id": 0}).to_list(5000)
    paddy_used = round(sum(e.get("paddy_used", 0) for e in milling_entries), 2)
    rice_raw = round(sum(e.get("rice_produced", 0) for e in milling_entries if e.get("product_type") in ("raw", None)), 2)
    rice_usna = round(sum(e.get("rice_produced", 0) for e in milling_entries if e.get("product_type") == "usna"), 2)
    frk = round(sum(e.get("frk_produced", 0) or 0 for e in milling_entries), 2)
    byproduct = round(sum(e.get("byproduct_produced", 0) or 0 for e in milling_entries), 2)
    total_paddy_in = round(cmr_paddy + pvt_paddy, 2)
    paddy_avail = round(total_paddy_in - paddy_used, 2)
    gunny_entries = await db.gunny_bags.find(dict(query), {"_id": 0}).to_list(5000)
    gunny_in = sum(e.get('quantity', 0) for e in gunny_entries if e.get('txn_type') == 'in')
    gunny_out = sum(e.get('quantity', 0) for e in gunny_entries if e.get('txn_type') == 'out')

    # Targets
    tq = {}
    if target_mandi: tq["mandi_name"] = target_mandi
    if kms_year: tq["kms_year"] = kms_year
    if season: tq["season"] = season
    targets = await db.mandi_targets.find(tq, {"_id": 0}).to_list(100)
    if not targets and (kms_year or season):
        tq2 = {}
        if target_mandi: tq2["mandi_name"] = target_mandi
        targets = await db.mandi_targets.find(tq2, {"_id": 0}).to_list(100)

    # Compute target totals upfront — agent commission MUST use TP weight (achieved procurement)
    # not the abstract target_qntl. Rates default ONLY if missing/None — explicit 0 must be respected.
    # NOTE: Pending & Progress measured against govt TARGET (not Expected). Cutting% is agent's
    # extra commission, not part of the procurement target.
    def _rate(val, fallback):
        return val if val is not None else fallback
    tot = {"target": 0, "expected": 0, "achieved": 0, "pending": 0, "agent": 0, "cutting": 0}
    target_rows = []
    for tg in targets:
        eq = {"mandi_name": tg["mandi_name"]}
        if kms_year: eq["kms_year"] = kms_year
        if season: eq["season"] = season
        # Achieved (final_w / 100) — used for the Targets section's progress %
        pipe2 = [{"$match": eq}, {"$group": {"_id": None, "total": {"$sum": "$final_w"}}}]
        res = await db.mill_entries.aggregate(pipe2).to_list(1)
        achieved = round(res[0]["total"] / 100, 2) if res else 0
        # TP weight (sum of tp_weight column) — used for agent commission
        tpw_pipe = [{"$match": eq}, {"$group": {"_id": None, "total": {"$sum": "$tp_weight"}}}]
        tpw_res = await db.mill_entries.aggregate(tpw_pipe).to_list(1)
        tpw = round(tpw_res[0]["total"] if tpw_res and tpw_res[0]["total"] else 0, 2)
        expected = tg.get("expected_total", tg["target_qntl"])
        target_qntl_val = tg["target_qntl"]
        pending = round(max(0, target_qntl_val - achieved), 2)
        progress = round((achieved / target_qntl_val * 100) if target_qntl_val > 0 else 0, 1)
        cutting_pct = _rate(tg.get("cutting_percent"), 0)
        base_rate = _rate(tg.get("base_rate"), 10)
        cutting_rate = _rate(tg.get("cutting_rate"), 5)
        # Cap TP weight at (target + cutting%) — extra goes to Pvt Purchase, no agent commission on it
        capped_tp = capped_tp_for_commission(tpw, target_qntl_val, cutting_pct)
        cutting_q = round(capped_tp * cutting_pct / 100, 2)
        agent_amt = round((capped_tp * base_rate) + (cutting_q * cutting_rate), 2)
        # Total agent cutting for KPI banner = sum of (target_qntl * cutting%)
        cutting_target_q = round(target_qntl_val * cutting_pct / 100, 2)
        tot["target"] += target_qntl_val; tot["expected"] += expected
        tot["achieved"] += achieved; tot["pending"] += pending; tot["agent"] += agent_amt
        tot["cutting"] += cutting_target_q
        target_rows.append((tg, achieved, expected, pending, progress, agent_amt))
    overall_progress = round((tot["achieved"] / tot["target"] * 100) if tot["target"] > 0 else 0, 1)

    # ---- KPI HERO BANNER (top of report) ----
    progress_color = STAT_COLORS['green'] if overall_progress >= 100 else (STAT_COLORS['gold'] if overall_progress >= 50 else STAT_COLORS['red'])
    available_color = STAT_COLORS['emerald'] if paddy_avail >= 0 else STAT_COLORS['red']
    kpi_stats_top = []
    if show_stock:
        kpi_stats_top.extend([
            {'label': 'PADDY IN', 'value': f"{total_paddy_in:,.1f} Q", 'color': STAT_COLORS['blue']},
            {'label': 'PADDY USED', 'value': f"{paddy_used:,.1f} Q", 'color': STAT_COLORS['orange']},
            {'label': 'AVAILABLE', 'value': f"{paddy_avail:,.1f} Q", 'color': available_color},
            {'label': 'RICE PRODUCED', 'value': f"{(rice_raw + rice_usna):,.1f} Q", 'color': STAT_COLORS['purple']},
        ])
    if show_targets and targets:
        kpi_stats_top.extend([
            {'label': 'TARGETS', 'value': f"{tot['target']:,.0f} Q", 'color': STAT_COLORS['gold']},
            {'label': 'AGENT CUTTING', 'value': f"{tot['cutting']:,.0f} Q", 'color': STAT_COLORS['teal']},
            {'label': 'ACHIEVED', 'value': f"{tot['achieved']:,.0f} Q ({overall_progress}%)", 'color': progress_color},
            {'label': 'PENDING', 'value': f"{tot['pending']:,.0f} Q", 'color': STAT_COLORS['red']},
        ])
    if kpi_stats_top:
        banner = get_pdf_summary_banner(kpi_stats_top, total_width=PAGE_W)
        if banner:
            banner.hAlign = 'LEFT'
            elements.append(banner)
            elements.append(Spacer(1, 5*mm))

    hdr_bg = colors.HexColor('#1e3a8a')
    hdr_fg = colors.white
    tot_bg = colors.HexColor('#fef3c7')   # amber-100 highlight for total rows
    alt_bg = colors.HexColor('#f8fafc')   # slate-50 zebra rows

    # ---- STOCK SECTION ----
    if show_stock:
        elements.append(CondPageBreak(60*mm))
        elements.append(get_pdf_section_band("STOCK OVERVIEW", subtitle=f"FY {kms_year or 'All'} · {season or 'All'}", preset='orange', total_width=PAGE_W))
        elements.append(Spacer(1, 2*mm))

        data = [
            ["Item", "Source", "IN", "OUT/Used", "Available", "Unit"],
            ["Paddy", "CMR (Mill Entry)", f"{cmr_paddy:,.2f}", "—", "—", "Qntl"],
            ["Paddy", "Private Purchase", f"{pvt_paddy:,.2f}", "—", "—", "Qntl"],
            ["TOTAL PADDY", "", f"{total_paddy_in:,.2f}", f"{paddy_used:,.2f}", f"{paddy_avail:,.2f}", "Qntl"],
            ["Rice (Raw)", "Milling", f"{rice_raw:,.2f}", "—", f"{rice_raw:,.2f}", "Qntl"],
            ["Rice (Usna)", "Milling", f"{rice_usna:,.2f}", "—", f"{rice_usna:,.2f}", "Qntl"],
            ["FRK", "Milling", f"{frk:,.2f}", "—", f"{frk:,.2f}", "Qntl"],
            ["By-Products", "Milling", f"{byproduct:,.2f}", "—", f"{byproduct:,.2f}", "Qntl"],
            ["Gunny Bags", "All Sources", f"{gunny_in:,}", f"{gunny_out:,}", f"{(gunny_in - gunny_out):,}", "Bags"],
        ]
        # Distribute widths proportionally to PAGE_W
        cw = [PAGE_W * w for w in (0.18, 0.22, 0.16, 0.16, 0.18, 0.10)]
        t = RLTable(data, colWidths=cw, repeatRows=1)
        st = [
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg), ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            # TOTAL row highlight
            ('FONTNAME', (0, 3), (-1, 3), 'FreeSansBold'),
            ('BACKGROUND', (0, 3), (-1, 3), tot_bg),
            ('TEXTCOLOR', (0, 3), (-1, 3), colors.HexColor('#92400e')),
            # Color positive availability green; negative gunny red
            ('TEXTCOLOR', (4, 3), (4, 3), colors.HexColor(STAT_COLORS['emerald']) if paddy_avail >= 0 else colors.HexColor(STAT_COLORS['red'])),
        ]
        # Zebra rows for non-total data rows
        for i in (2, 5, 7, 9):
            if i < len(data):
                st.append(('BACKGROUND', (0, i), (-1, i), alt_bg))
        # Gunny negative? row index 8
        if (gunny_in - gunny_out) < 0:
            st.append(('TEXTCOLOR', (4, 8), (4, 8), colors.HexColor(STAT_COLORS['red'])))
            st.append(('FONTNAME', (4, 8), (4, 8), 'FreeSansBold'))
        t.setStyle(TableStyle(st))
        t.hAlign = 'LEFT'
        elements.append(t)
        elements.append(Spacer(1, 5*mm))

    # ---- TARGETS SECTION ----
    if show_targets:
        elements.append(CondPageBreak(60*mm))
        elements.append(get_pdf_section_band(
            f"MANDI TARGETS{' · ' + target_mandi if target_mandi else ''}",
            subtitle=f"Overall: {overall_progress}% achieved" if targets else None,
            preset='teal', total_width=PAGE_W,
        ))
        elements.append(Spacer(1, 2*mm))

        if target_rows:
            data = [["Mandi", "Govt Target (Q)", "Cut %", "Agent Cutting (Q)", "Achieved (Q)", "Pending (Q)", "Progress", "Agent Amt"]]
            tot_cutting = 0
            for tg, achieved, expected, pending, progress, agent_amt in target_rows:
                cutting_q = round(tg["target_qntl"] * tg.get("cutting_percent", 0) / 100, 2)
                tot_cutting += cutting_q
                data.append([
                    tg["mandi_name"], f"{tg['target_qntl']:,.1f}", f"{tg['cutting_percent']}%",
                    f"{cutting_q:,.1f}", f"{achieved:,.1f}", f"{pending:,.1f}",
                    f"{progress}%", f"Rs.{agent_amt:,.0f}",
                ])
            data.append(["TOTAL", f"{tot['target']:,.1f}", "—", f"{tot_cutting:,.1f}",
                f"{tot['achieved']:,.1f}", f"{tot['pending']:,.1f}", f"{overall_progress}%", f"Rs.{tot['agent']:,.0f}"])

            cw2 = [PAGE_W * w for w in (0.16, 0.10, 0.07, 0.13, 0.13, 0.13, 0.12, 0.16)]
            t = RLTable(data, colWidths=cw2, repeatRows=1)
            st = [
                ('BACKGROUND', (0, 0), (-1, 0), hdr_bg), ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
                ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                # TOTAL row highlight
                ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
                ('BACKGROUND', (0, -1), (-1, -1), tot_bg),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#92400e')),
            ]
            # Zebra
            for i in range(2, len(target_rows) + 1, 2):
                st.append(('BACKGROUND', (0, i), (-1, i), alt_bg))
            # Color progress per row
            for i, (_, _, _, _, progress, _) in enumerate(target_rows, 1):
                if progress >= 100:
                    st.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor(STAT_COLORS['emerald'])))
                    st.append(('FONTNAME', (6, i), (6, i), 'FreeSansBold'))
                elif progress < 50:
                    st.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor(STAT_COLORS['red'])))
                    st.append(('FONTNAME', (6, i), (6, i), 'FreeSansBold'))
                else:
                    st.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor(STAT_COLORS['gold'])))
            # Color overall progress in TOTAL row
            if overall_progress >= 100:
                st.append(('TEXTCOLOR', (6, -1), (6, -1), colors.HexColor(STAT_COLORS['emerald'])))
            elif overall_progress < 50:
                st.append(('TEXTCOLOR', (6, -1), (6, -1), colors.HexColor(STAT_COLORS['red'])))
            t.setStyle(TableStyle(st))
            t.hAlign = 'LEFT'
            elements.append(t)
        else:
            elements.append(Paragraph("Koi target set nahi hai", styles['Normal']))

    # Footer
    elements.append(Spacer(1, 8*mm))
    ft = ParagraphStyle('Ft', parent=styles['Normal'], fontSize=7.5, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER)
    elements.append(Paragraph(f"Generated by {company} Mill Entry System  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}", ft))

    doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=dashboard_{filter_label}_{datetime.now().strftime('%Y%m%d')}.pdf"})


@router.get("/export/summary-report-pdf")
async def export_summary_report_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Export complete summary report - Stock + Targets + Truck + Agent Payments + Grand Total"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from utils.export_helpers import (
        get_pdf_styles, get_pdf_summary_banner, get_pdf_section_band, STAT_COLORS,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, CondPageBreak, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    buffer = io.BytesIO()
    # leftMargin=8 + Frame's 6pt internal pad = 14pt effective from page edge → page-centered content (PAGE_W=180mm).
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=8*mm, rightMargin=8*mm, topMargin=10*mm, bottomMargin=10*mm)
    PAGE_W = A4[0] - 16*mm
    elements = []; styles = get_pdf_styles()

    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    company, tagline = await get_company_name()
    hdr_bg = colors.HexColor('#1e3a8a')
    hdr_fg = colors.white
    tot_bg = colors.HexColor('#fef3c7')
    alt_bg = colors.HexColor('#f8fafc')
    grand_bg = colors.HexColor('#B45309')  # amber-700 for grand total emphasis

    sub = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=8.5, textColor=colors.HexColor('#475569'), alignment=TA_CENTER)
    ft = ParagraphStyle('Ft', parent=styles['Normal'], fontSize=7.5, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER)

    # ---- SUB HEADER LINE ----
    elements.append(Paragraph(
        f"<b>COMPLETE SUMMARY REPORT</b> &nbsp;|&nbsp; FY: {kms_year or 'All'} &nbsp;|&nbsp; "
        f"Season: {season or 'All'} &nbsp;|&nbsp; {datetime.now().strftime('%d-%b-%Y %H:%M')}",
        sub,
    ))
    elements.append(Spacer(1, 4*mm))

    # ============================================================
    # Compute ALL data first (so KPI banner can show grand totals up top)
    # ============================================================
    # Section 1: Stock
    pipe = [{"$match": query}, {"$group": {"_id": None, "total": {"$sum": "$final_w"}}}]
    mill_res = await db.mill_entries.aggregate(pipe).to_list(1)
    cmr_paddy = round((mill_res[0]["total"] / 100) if mill_res else 0, 2)
    pvt_query_exp = dict(query); pvt_query_exp["source"] = {"$ne": "agent_extra"}
    pvt_entries = await db.private_paddy.find(pvt_query_exp, {"_id": 0, "qntl": 1, "bag": 1}).to_list(5000)
    pvt_paddy = round(sum(e.get("qntl", 0) - e.get("bag", 0) / 100 for e in pvt_entries), 2)
    milling_entries = await db.milling_entries.find(dict(query), {"_id": 0}).to_list(5000)
    paddy_used = round(sum(e.get("paddy_used", 0) for e in milling_entries), 2)
    rice_raw = round(sum(e.get("rice_produced", 0) for e in milling_entries if e.get("product_type") in ("raw", None)), 2)
    rice_usna = round(sum(e.get("rice_produced", 0) for e in milling_entries if e.get("product_type") == "usna"), 2)
    frk = round(sum(e.get("frk_produced", 0) or 0 for e in milling_entries), 2)
    gunny_e = await db.gunny_bags.find(dict(query), {"_id": 0}).to_list(5000)
    gunny_in = sum(e.get('quantity', 0) for e in gunny_e if e.get('txn_type') == 'in')
    gunny_out = sum(e.get('quantity', 0) for e in gunny_e if e.get('txn_type') == 'out')
    total_paddy_in = round(cmr_paddy + pvt_paddy, 2)
    paddy_avail = round(total_paddy_in - paddy_used, 2)

    # Section 2: Targets
    tq = {}
    if kms_year: tq["kms_year"] = kms_year
    if season: tq["season"] = season
    targets = await db.mandi_targets.find(tq, {"_id": 0}).to_list(100)
    if not targets and (kms_year or season):
        targets = await db.mandi_targets.find({}, {"_id": 0}).to_list(100)
    tot = {"t": 0, "e": 0, "a": 0, "p": 0, "c": 0}
    target_rows = []
    for tg in targets:
        eq = {"mandi_name": tg["mandi_name"]}
        if kms_year: eq["kms_year"] = kms_year
        if season: eq["season"] = season
        pipe2 = [{"$match": eq}, {"$group": {"_id": None, "total": {"$sum": "$final_w"}}}]
        res = await db.mill_entries.aggregate(pipe2).to_list(1)
        a = round(res[0]["total"] / 100, 2) if res else 0
        e_val = tg.get("expected_total", tg["target_qntl"])
        # Pending & Progress against govt TARGET (cutting% is agent's extra, not procurement target)
        t_val = tg["target_qntl"]
        p = round(max(0, t_val - a), 2)
        pr = round((a / t_val * 100) if t_val > 0 else 0, 1)
        # Agent cutting (Q) per mandi for KPI/TOTAL = target × cutting%
        c_q = round(t_val * tg.get("cutting_percent", 0) / 100, 2)
        tot["t"] += t_val; tot["e"] += e_val; tot["a"] += a; tot["p"] += p; tot["c"] += c_q
        target_rows.append((tg, a, e_val, p, pr))
    overall_progress = round((tot["a"] / tot["t"] * 100) if tot["t"] > 0 else 0, 1)

    # Section 3: Truck Payments
    entries = await db.mill_entries.find(query, {"_id": 0}).sort([("date", 1), ("rst_no", 1)]).to_list(1000)
    truck_total_net = truck_total_paid = truck_total_balance = 0
    truck_rows = []
    for entry in entries:
        eid = entry.get("id")
        pdoc = await db.truck_payments.find_one({"entry_id": eid}, {"_id": 0})
        rate = pdoc.get("rate_per_qntl", 0) if pdoc else 0
        paid = pdoc.get("paid_amount", 0) if pdoc else 0
        fq = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
        cash = entry.get("cash_paid", 0) or 0
        diesel = entry.get("diesel_paid", 0) or 0
        net = round(fq * rate - cash - diesel, 2)
        bal = round(max(0, net - paid), 2)
        truck_total_net += net; truck_total_paid += paid; truck_total_balance += bal
        truck_rows.append([fmt_date(entry.get("date", "")[:10]), entry.get("truck_no", "")[:12], entry.get("mandi_name", "")[:16],
            f"{fq:,.2f}", f"Rs.{net:,.0f}", f"Rs.{paid:,.0f}", f"Rs.{bal:,.0f}",
            "Paid" if bal < 0.10 else "Pending"])

    # Section 4: Agent Payments — use TP weight (actual procurement) + respect explicit 0 rates
    def _rate(val, fallback):
        return val if val is not None else fallback
    agent_total_amt = agent_total_paid = agent_total_balance = 0
    agent_rows = []
    for tg in targets:
        mandi = tg["mandi_name"]
        # TP weight from mill entries
        tpw_eq = {"mandi_name": mandi}
        if kms_year: tpw_eq["kms_year"] = kms_year
        if season: tpw_eq["season"] = season
        tpw_pipe = [{"$match": tpw_eq}, {"$group": {"_id": None, "total": {"$sum": "$tp_weight"}}}]
        tpw_res = await db.mill_entries.aggregate(tpw_pipe).to_list(1)
        tpw = round(tpw_res[0]["total"] if tpw_res and tpw_res[0]["total"] else 0, 2)
        cutting_pct = _rate(tg.get("cutting_percent"), 0)
        br = _rate(tg.get("base_rate"), 10)
        cr = _rate(tg.get("cutting_rate"), 5)
        cq = round(tpw * cutting_pct / 100, 2)
        total_amt = round((tpw * br) + (cq * cr), 2)
        pdoc = await db.agent_payments.find_one({"mandi_name": mandi, "kms_year": tg["kms_year"], "season": tg["season"]}, {"_id": 0})
        paid = pdoc.get("paid_amount", 0) if pdoc else 0
        bal = round(max(0, total_amt - paid), 2)
        agent_total_amt += total_amt; agent_total_paid += paid; agent_total_balance += bal
        agent_rows.append([mandi, f"{tpw:,.1f}", f"{cq:,.1f}", f"Rs.{br}/Rs.{cr}", f"Rs.{total_amt:,.0f}", f"Rs.{paid:,.0f}", f"Rs.{bal:,.0f}",
            "Paid" if bal <= 0 else "Pending"])

    # Grand totals
    ga = truck_total_net + agent_total_amt
    gp = truck_total_paid + agent_total_paid
    gb = truck_total_balance + agent_total_balance

    # ============================================================
    # KPI HERO BANNER (top of report)
    # ============================================================
    progress_color = STAT_COLORS['green'] if overall_progress >= 100 else (STAT_COLORS['gold'] if overall_progress >= 50 else STAT_COLORS['red'])
    paid_pct = round((gp / ga * 100) if ga > 0 else 0, 1)
    paid_pct_color = STAT_COLORS['green'] if paid_pct >= 90 else (STAT_COLORS['gold'] if paid_pct >= 50 else STAT_COLORS['red'])
    kpi_top = [
        {'label': 'PADDY IN', 'value': f"{total_paddy_in:,.0f} Q", 'color': STAT_COLORS['blue']},
        {'label': 'PADDY USED', 'value': f"{paddy_used:,.0f} Q", 'color': STAT_COLORS['orange']},
        {'label': 'TARGETS', 'value': f"{tot['t']:,.0f} Q", 'color': STAT_COLORS['gold']},
        {'label': 'AGENT CUTTING', 'value': f"{tot['c']:,.0f} Q", 'color': STAT_COLORS['teal']},
        {'label': 'ACHIEVED', 'value': f"{overall_progress}%", 'color': progress_color},
        {'label': 'GRAND TOTAL', 'value': f"Rs.{ga:,.0f}", 'color': STAT_COLORS['purple']},
        {'label': 'PAID', 'value': f"Rs.{gp:,.0f} ({paid_pct}%)", 'color': paid_pct_color},
        {'label': 'BALANCE DUE', 'value': f"Rs.{gb:,.0f}", 'color': STAT_COLORS['red']},
    ]
    banner = get_pdf_summary_banner(kpi_top, total_width=PAGE_W)
    if banner:
        banner.hAlign = 'LEFT'
        elements.append(banner)
        elements.append(Spacer(1, 5*mm))

    # ============================================================
    # SECTION 1: STOCK
    # ============================================================
    elements.append(CondPageBreak(60*mm))  # ensure band + at least 5 rows fit on page
    elements.append(get_pdf_section_band("1 · STOCK OVERVIEW", subtitle=f"Available: {paddy_avail:,.1f} Q · Rice: {(rice_raw + rice_usna):,.1f} Q", preset='orange', total_width=PAGE_W))
    elements.append(Spacer(1, 2*mm))

    stock_data = [
        ["Item", "IN", "OUT/Used", "Available", "Unit"],
        ["Paddy (CMR)", f"{cmr_paddy:,.2f}", "—", "—", "Qntl"],
        ["Paddy (Pvt)", f"{pvt_paddy:,.2f}", "—", "—", "Qntl"],
        ["TOTAL PADDY", f"{total_paddy_in:,.2f}", f"{paddy_used:,.2f}", f"{paddy_avail:,.2f}", "Qntl"],
        ["Rice (Raw)", f"{rice_raw:,.2f}", "—", f"{rice_raw:,.2f}", "Qntl"],
        ["Rice (Usna)", f"{rice_usna:,.2f}", "—", f"{rice_usna:,.2f}", "Qntl"],
        ["FRK", f"{frk:,.2f}", "—", f"{frk:,.2f}", "Qntl"],
        ["Gunny Bags", f"{gunny_in:,}", f"{gunny_out:,}", f"{(gunny_in - gunny_out):,}", "Bags"],
    ]
    sw = [PAGE_W * w for w in (0.30, 0.18, 0.18, 0.22, 0.12)]
    st_table = Table(stock_data, colWidths=sw)
    st_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), hdr_bg), ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('FONTNAME', (0, 3), (-1, 3), 'FreeSansBold'),
        ('BACKGROUND', (0, 3), (-1, 3), tot_bg),
        ('TEXTCOLOR', (0, 3), (-1, 3), colors.HexColor('#92400e')),
        ('TEXTCOLOR', (3, 3), (3, 3), colors.HexColor(STAT_COLORS['emerald']) if paddy_avail >= 0 else colors.HexColor(STAT_COLORS['red'])),
        ('BACKGROUND', (0, 2), (-1, 2), alt_bg),
        ('BACKGROUND', (0, 5), (-1, 5), alt_bg),
        ('BACKGROUND', (0, 7), (-1, 7), alt_bg),
    ]))
    st_table.hAlign = 'LEFT'
    elements.append(st_table)
    elements.append(Spacer(1, 5*mm))

    # ============================================================
    # SECTION 2: TARGETS
    # ============================================================
    elements.append(CondPageBreak(60*mm))
    elements.append(get_pdf_section_band("2 · MANDI TARGETS", subtitle=f"Overall: {overall_progress}% achieved" if targets else None, preset='teal', total_width=PAGE_W))
    elements.append(Spacer(1, 2*mm))

    if target_rows:
        tdata = [["Mandi", "Govt Target (Q)", "Cut %", "Agent Cutting (Q)", "Achieved (Q)", "Pending (Q)", "Progress"]]
        for tg, a, e_val, p, pr in target_rows:
            cutting_q = round(tg["target_qntl"] * tg.get("cutting_percent", 0) / 100, 2)
            tdata.append([tg["mandi_name"], f"{tg['target_qntl']:,.1f}", f"{tg['cutting_percent']}%", f"{cutting_q:,.1f}", f"{a:,.1f}", f"{p:,.1f}", f"{pr}%"])
        tdata.append(["TOTAL", f"{tot['t']:,.1f}", "—", f"{tot['c']:,.1f}", f"{tot['a']:,.1f}", f"{tot['p']:,.1f}", f"{overall_progress}%"])
        tw = [PAGE_W * w for w in (0.20, 0.13, 0.10, 0.15, 0.15, 0.15, 0.12)]
        tt = Table(tdata, colWidths=tw)
        tts = [
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg), ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
            ('BACKGROUND', (0, -1), (-1, -1), tot_bg),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#92400e')),
        ]
        for i in range(2, len(target_rows) + 1, 2):
            tts.append(('BACKGROUND', (0, i), (-1, i), alt_bg))
        for i, (_, _, _, _, pr) in enumerate(target_rows, 1):
            if pr >= 100:
                tts.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor(STAT_COLORS['emerald'])))
                tts.append(('FONTNAME', (6, i), (6, i), 'FreeSansBold'))
            elif pr < 50:
                tts.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor(STAT_COLORS['red'])))
                tts.append(('FONTNAME', (6, i), (6, i), 'FreeSansBold'))
            else:
                tts.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor(STAT_COLORS['gold'])))
        tt.setStyle(TableStyle(tts))
        tt.hAlign = 'LEFT'
        elements.append(tt)
    else:
        elements.append(Paragraph("No targets set", styles['Normal']))
    elements.append(Spacer(1, 5*mm))

    # ============================================================
    # SECTION 3: TRUCK PAYMENTS
    # ============================================================
    elements.append(CondPageBreak(60*mm))
    elements.append(get_pdf_section_band("3 · TRUCK PAYMENTS", subtitle=f"Balance: Rs.{truck_total_balance:,.0f}", preset='purple', total_width=PAGE_W))
    elements.append(Spacer(1, 2*mm))

    if truck_rows:
        tdata2 = [["Date", "Truck", "Mandi", "QNTL", "Net", "Paid", "Balance", "Status"]] + truck_rows
        tdata2.append(["TOTAL", "", "", "", f"Rs.{round(truck_total_net):,}", f"Rs.{round(truck_total_paid):,}", f"Rs.{round(truck_total_balance):,}", ""])
        tw2 = [PAGE_W * w for w in (0.10, 0.12, 0.14, 0.10, 0.13, 0.12, 0.13, 0.16)]
        tt2 = Table(tdata2, colWidths=tw2, repeatRows=1)
        tts2 = [
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg), ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
            ('BACKGROUND', (0, -1), (-1, -1), tot_bg),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#92400e')),
        ]
        # Color status column: Paid green / Pending red
        for i, row in enumerate(truck_rows, 1):
            status_color = STAT_COLORS['emerald'] if row[7] == "Paid" else STAT_COLORS['red']
            tts2.append(('TEXTCOLOR', (7, i), (7, i), colors.HexColor(status_color)))
            tts2.append(('FONTNAME', (7, i), (7, i), 'FreeSansBold'))
        # Zebra stripes
        for i in range(2, len(truck_rows) + 1, 2):
            tts2.append(('BACKGROUND', (0, i), (-1, i), alt_bg))
        tt2.setStyle(TableStyle(tts2))
        tt2.hAlign = 'LEFT'
        elements.append(tt2)
    else:
        elements.append(Paragraph("No truck entries", styles['Normal']))
    elements.append(Spacer(1, 5*mm))

    # ============================================================
    # SECTION 4: AGENT PAYMENTS
    # ============================================================
    elements.append(CondPageBreak(60*mm))
    elements.append(get_pdf_section_band("4 · AGENT / MANDI PAYMENTS", subtitle=f"Balance: Rs.{agent_total_balance:,.0f}", preset='rose', total_width=PAGE_W))
    elements.append(Spacer(1, 2*mm))

    if agent_rows:
        adata = [["Mandi", "TP Weight", "Cutting", "Rates", "Total", "Paid", "Balance", "Status"]] + agent_rows
        adata.append(["TOTAL", "", "", "", f"Rs.{round(agent_total_amt):,}", f"Rs.{round(agent_total_paid):,}", f"Rs.{round(agent_total_balance):,}", ""])
        aw = [PAGE_W * w for w in (0.16, 0.10, 0.10, 0.16, 0.13, 0.12, 0.13, 0.10)]
        at = Table(adata, colWidths=aw, repeatRows=1)
        ats = [
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg), ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
            ('BACKGROUND', (0, -1), (-1, -1), tot_bg),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#92400e')),
        ]
        for i, row in enumerate(agent_rows, 1):
            status_color = STAT_COLORS['emerald'] if row[7] == "Paid" else STAT_COLORS['red']
            ats.append(('TEXTCOLOR', (7, i), (7, i), colors.HexColor(status_color)))
            ats.append(('FONTNAME', (7, i), (7, i), 'FreeSansBold'))
        for i in range(2, len(agent_rows) + 1, 2):
            ats.append(('BACKGROUND', (0, i), (-1, i), alt_bg))
        at.setStyle(TableStyle(ats))
        at.hAlign = 'LEFT'
        elements.append(at)
    else:
        elements.append(Paragraph("No agent payments", styles['Normal']))
    elements.append(Spacer(1, 5*mm))

    # ============================================================
    # SECTION 5: GRAND TOTAL
    # ============================================================
    elements.append(CondPageBreak(70*mm))  # band + 4-row table all together
    elements.append(get_pdf_section_band("5 · GRAND TOTAL", subtitle=f"Outstanding: Rs.{gb:,.0f} ({100 - paid_pct:.1f}%)", preset='amber', total_width=PAGE_W))
    elements.append(Spacer(1, 2*mm))

    gdata = [
        ["Category", "Total Amount", "Paid", "Balance"],
        ["Truck Payments", f"Rs.{round(truck_total_net):,}", f"Rs.{round(truck_total_paid):,}", f"Rs.{round(truck_total_balance):,}"],
        ["Agent Payments", f"Rs.{round(agent_total_amt):,}", f"Rs.{round(agent_total_paid):,}", f"Rs.{round(agent_total_balance):,}"],
        ["GRAND TOTAL", f"Rs.{round(ga):,}", f"Rs.{round(gp):,}", f"Rs.{round(gb):,}"],
    ]
    gw = [PAGE_W * w for w in (0.30, 0.25, 0.20, 0.25)]
    gt = Table(gdata, colWidths=gw)
    gt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), hdr_bg), ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'), ('FONTSIZE', (0, 0), (-1, -1), 9.5),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        # GRAND TOTAL row gets emphatic amber bg + white text + bold
        ('BACKGROUND', (0, -1), (-1, -1), grand_bg), ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        # Zebra
        ('BACKGROUND', (0, 2), (-1, 2), alt_bg),
    ]))
    gt.hAlign = 'LEFT'
    elements.append(gt)

    # Footer
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(f"Generated by {company} Mill Entry System  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}", ft))

    doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=summary_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"})


@router.get("/export/truck-owner-excel")
async def export_truck_owner_excel(
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export truck owner consolidated payments to Excel"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort([("date", 1), ("rst_no", 1)]).to_list(1000)
    
    # Group by truck_no
    truck_data = {}
    for entry in entries:
        truck_no = entry.get("truck_no", "Unknown")
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        rate = payment_doc.get("rate_per_qntl", 0) if payment_doc else 0
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round(final_qntl * rate, 2)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions, 2)
        balance = round(max(0, net_amount - paid_amount), 2)
        
        if truck_no not in truck_data:
            truck_data[truck_no] = {
                "truck_no": truck_no,
                "trips": 0,
                "total_qntl": 0,
                "total_gross": 0,
                "total_deductions": 0,
                "total_net": 0,
                "total_paid": 0,
                "total_balance": 0
            }
        
        truck_data[truck_no]["trips"] += 1
        truck_data[truck_no]["total_qntl"] += final_qntl
        truck_data[truck_no]["total_gross"] += gross_amount
        truck_data[truck_no]["total_deductions"] += deductions
        truck_data[truck_no]["total_net"] += net_amount
        truck_data[truck_no]["total_paid"] += paid_amount
        truck_data[truck_no]["total_balance"] += balance
    
    consolidated = list(truck_data.values())
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Truck Owner Payments"
    
    # Styles
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:I1')
    company_name, tagline = await get_company_name()
    ws['A1'] = f"TRUCK OWNER CONSOLIDATED PAYMENTS - {company_name} | FY: {kms_year or 'All'} | {season or 'All'}"
    ws['A1'].font = Font(bold=True, size=14, color="0891B2")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:I2')
    ws['A2'] = "Ek truck ke saare trips ka combined payment"
    ws['A2'].font = Font(size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Truck No", "Total Trips", "Total QNTL", "Gross Amount", "Deductions", "Net Payable", "Paid", "Balance", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    grand_net = 0
    grand_paid = 0
    grand_balance = 0
    
    for row_idx, t in enumerate(consolidated, 5):
        status = "Paid" if t["total_balance"] < 0.10 else ("Partial" if t["total_paid"] > 0 else "Pending")
        
        ws.cell(row=row_idx, column=1, value=t["truck_no"]).font = Font(bold=True, size=11)
        ws.cell(row=row_idx, column=2, value=t["trips"]).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=3, value=round(t["total_qntl"], 2))
        ws.cell(row=row_idx, column=4, value=round(t["total_gross"], 2))
        ws.cell(row=row_idx, column=5, value=round(t["total_deductions"], 2))
        ws.cell(row=row_idx, column=6, value=round(t["total_net"], 2)).font = Font(bold=True)
        ws.cell(row=row_idx, column=7, value=round(t["total_paid"], 2))
        ws.cell(row=row_idx, column=8, value=round(t["total_balance"], 2)).font = Font(bold=True, color="DC2626" if t["total_balance"] > 0 else "059669")
        ws.cell(row=row_idx, column=9, value=status)
        
        grand_net += t["total_net"]
        grand_paid += t["total_paid"]
        grand_balance += t["total_balance"]
    
    # Grand Total row
    total_row = len(consolidated) + 5
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(consolidated)).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=6, value=round(grand_net, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=round(grand_paid, 2)).font = Font(bold=True, color="059669")
    ws.cell(row=total_row, column=8, value=round(grand_balance, 2)).font = Font(bold=True, color="DC2626")
    
    for col in range(1, 10):
        ws.cell(row=total_row, column=col).fill = total_fill
    
    # Column widths
    widths = [15, 12, 12, 14, 12, 14, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"truck_owner_consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/truck-owner-pdf")
async def export_truck_owner_pdf(
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    truck_no: Optional[str] = None,
):
    """Export truck owner consolidated payments to PDF.
    Pass `truck_no` to filter to a single truck (used by WhatsApp/Group share so
    each owner gets only their own truck's report, not all trucks)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if truck_no:
        query["truck_no"] = truck_no
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort([("date", 1), ("rst_no", 1)]).to_list(1000)
    
    # Group by truck_no
    truck_data = {}
    for entry in entries:
        truck_no = entry.get("truck_no", "Unknown")
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        rate = payment_doc.get("rate_per_qntl", 0) if payment_doc else 0
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round(final_qntl * rate, 2)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions, 2)
        balance = round(max(0, net_amount - paid_amount), 2)
        
        if truck_no not in truck_data:
            truck_data[truck_no] = {
                "truck_no": truck_no,
                "trips": 0,
                "total_qntl": 0,
                "total_gross": 0,
                "total_deductions": 0,
                "total_net": 0,
                "total_paid": 0,
                "total_balance": 0
            }
        
        truck_data[truck_no]["trips"] += 1
        truck_data[truck_no]["total_qntl"] += final_qntl
        truck_data[truck_no]["total_gross"] += gross_amount
        truck_data[truck_no]["total_deductions"] += deductions
        truck_data[truck_no]["total_net"] += net_amount
        truck_data[truck_no]["total_paid"] += paid_amount
        truck_data[truck_no]["total_balance"] += balance
    
    consolidated = list(truck_data.values())
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    elements = []
    styles = get_pdf_styles()
    
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#0891B2'), alignment=1)
    company_name, tagline = await get_company_name()
    elements.append(Paragraph(f"TRUCK OWNER CONSOLIDATED PAYMENTS - {company_name}", title_style))
    elements.append(Paragraph(f"FY: {kms_year or 'All'} | Season: {season or 'All'} | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", ParagraphStyle('Info', parent=styles['Normal'], fontSize=9, alignment=1, textColor=colors.gray)))
    elements.append(Spacer(1, 20))
    
    # Table
    table_data = [["Truck No", "Trips", "Total QNTL", "Gross", "Deductions", "Net Payable", "Paid", "Balance", "Status"]]
    
    grand_net = 0
    grand_paid = 0
    grand_balance = 0
    
    for t in consolidated:
        status = "PAID" if t["total_balance"] < 0.10 else ("PARTIAL" if t["total_paid"] > 0 else "PENDING")
        table_data.append([
            t["truck_no"],
            str(t["trips"]),
            f"{t['total_qntl']:.2f}",
            f"Rs.{t['total_gross']:.0f}",
            f"Rs.{t['total_deductions']:.0f}",
            f"Rs.{t['total_net']:.0f}",
            f"Rs.{t['total_paid']:.0f}",
            f"Rs.{t['total_balance']:.0f}",
            status
        ])
        grand_net += t["total_net"]
        grand_paid += t["total_paid"]
        grand_balance += t["total_balance"]
    
    # Grand Total
    table_data.append([
        "GRAND TOTAL",
        str(len(consolidated)),
        "",
        "",
        "",
        f"Rs.{grand_net:.0f}",
        f"Rs.{grand_paid:.0f}",
        f"Rs.{grand_balance:.0f}",
        ""
    ])
    
    table = Table(table_data, colWidths=[80, 50, 70, 70, 70, 80, 70, 70, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0891B2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'FreeSans'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
        ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"truck_owner_consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


