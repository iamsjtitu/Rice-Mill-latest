from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import Response
from typing import Optional
from datetime import datetime, timezone
from database import db
import uuid

router = APIRouter()

# ============ LOCAL PARTY ACCOUNTS ============

@router.get("/local-party/summary")
async def get_local_party_summary(kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Party-wise summary: total debit, total paid, balance with FY carry-forward"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        query["date"] = date_q
    txns = await db.local_party_accounts.find(query, {"_id": 0}).to_list(10000)

    # Compute opening balance from previous FY per party
    opening_balances = {}
    if kms_year and not date_from and not date_to:
        fy_parts = kms_year.split('-')
        if len(fy_parts) == 2:
            try:
                prev_fy = f"{int(fy_parts[0])-1}-{int(fy_parts[1])-1}"
                prev_query = {"kms_year": prev_fy}
                if season: prev_query["season"] = season
                prev_txns = await db.local_party_accounts.find(prev_query, {"_id": 0}).to_list(10000)
                for t in prev_txns:
                    pn = t.get("party_name", "").strip()
                    if not pn: continue
                    if pn not in opening_balances:
                        opening_balances[pn] = 0
                    if t.get("txn_type") == "debit":
                        opening_balances[pn] += t.get("amount", 0)
                    elif t.get("txn_type") == "payment":
                        opening_balances[pn] -= t.get("amount", 0)
            except (ValueError, IndexError):
                pass

    party_map = {}
    # First add parties that have opening balances from previous FY
    for pn, ob in opening_balances.items():
        if round(ob, 2) != 0:
            party_map[pn] = {"party_name": pn, "opening_balance": round(ob, 2), "total_debit": 0, "total_paid": 0, "balance": 0, "txn_count": 0}

    for t in txns:
        pn = t.get("party_name", "").strip()
        if not pn:
            continue
        if pn not in party_map:
            ob = round(opening_balances.get(pn, 0), 2)
            party_map[pn] = {"party_name": pn, "opening_balance": ob, "total_debit": 0, "total_paid": 0, "balance": 0, "txn_count": 0}
        if t.get("txn_type") == "debit":
            party_map[pn]["total_debit"] += t.get("amount", 0)
        elif t.get("txn_type") == "payment":
            party_map[pn]["total_paid"] += t.get("amount", 0)
        party_map[pn]["txn_count"] += 1

    # Use ledger as source of truth for total_paid (includes manual Cash Book payments)
    all_party_names = list(party_map.keys())
    if all_party_names:
        ledger_query_lp = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": all_party_names}}
        if kms_year: ledger_query_lp["kms_year"] = kms_year
        if season: ledger_query_lp["season"] = season
        ledger_payments = await db.cash_transactions.find(ledger_query_lp, {"_id": 0}).to_list(50000)
        ledger_paid_map = {}
        for lp in ledger_payments:
            pn = lp.get("category", "")
            ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lp.get("amount", 0)
        for pn in party_map:
            ledger_paid = round(ledger_paid_map.get(pn, 0), 2)
            if ledger_paid > party_map[pn]["total_paid"]:
                party_map[pn]["total_paid"] = ledger_paid

    parties = []
    for pn, s in party_map.items():
        s["total_debit"] = round(s["total_debit"], 2)
        s["total_paid"] = round(s["total_paid"], 2)
        s["balance"] = round(s.get("opening_balance", 0) + s["total_debit"] - s["total_paid"], 2)
        parties.append(s)

    parties.sort(key=lambda x: x["balance"], reverse=True)

    grand_ob = sum(p.get("opening_balance", 0) for p in parties)
    grand_debit = sum(p["total_debit"] for p in parties)
    grand_paid = sum(p["total_paid"] for p in parties)
    return {
        "parties": parties,
        "grand_opening_balance": round(grand_ob, 2),
        "grand_total_debit": round(grand_debit, 2),
        "grand_total_paid": round(grand_paid, 2),
        "grand_balance": round(grand_ob + grand_debit - grand_paid, 2)
    }


@router.get("/local-party/transactions")
async def get_local_party_transactions(party_name: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Get all transactions, optionally filtered by party. Includes manual cashbook payments."""
    query = {}
    if party_name: query["party_name"] = {"$regex": f"^{party_name}$", "$options": "i"}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        query["date"] = date_q
    txns = await db.local_party_accounts.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(10000)

    # Also include manual cashbook payments (ledger nikasi not from lp_settlement)
    if party_name:
        existing_refs = set(t.get("reference", "") for t in txns if t.get("reference"))
        existing_linked_ids = set(t.get("id", "") for t in txns)
        # Extract IDs from references for dedup matching
        existing_ref_ids = set()
        for t in txns:
            ref = t.get("reference", "")
            if ref and ":" in ref:
                existing_ref_ids.add(ref.split(":", 1)[1])
        cb_query = {"account": "ledger", "txn_type": "nikasi",
                    "category": {"$regex": f"^{party_name}$", "$options": "i"}}
        if kms_year: cb_query["kms_year"] = kms_year
        if season: cb_query["season"] = season
        if date_from or date_to:
            date_q = {}
            if date_from: date_q["$gte"] = date_from
            if date_to: date_q["$lte"] = date_to
            cb_query["date"] = date_q
        cb_payments = await db.cash_transactions.find(cb_query, {"_id": 0}).to_list(10000)
        for cb in cb_payments:
            linked_id = cb.get("linked_local_party_id", "")
            if linked_id and linked_id in existing_linked_ids:
                continue
            ref = cb.get("reference", "")
            if ref.startswith("local_party_ledger:") or ref.startswith("local_party:"):
                continue
            # Skip voucher payment ledger entries if the corresponding voucher_payment exists
            if ref.startswith("voucher_payment_ledger:"):
                ref_id = ref.split(":", 1)[1]
                if ref_id in existing_ref_ids:
                    continue
            if ref and ref in existing_refs:
                continue
            txns.append({
                "id": cb.get("id", ""),
                "date": cb.get("date", ""),
                "party_name": party_name,
                "txn_type": "payment",
                "amount": cb.get("amount", 0),
                "description": cb.get("description", "CashBook Payment"),
                "source_type": "cashbook",
                "reference": cb.get("reference", ""),
                "kms_year": cb.get("kms_year", ""),
                "season": cb.get("season", ""),
                "created_by": cb.get("created_by", ""),
                "created_at": cb.get("created_at", "")
            })
        txns.sort(key=lambda x: (x.get("date", ""), x.get("created_at", "")), reverse=True)

    return txns


@router.get("/local-party/report/{party_name}")
async def get_party_report(party_name: str, kms_year: Optional[str] = None, season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Detailed party-wise report with running balance for printing"""
    query = {"party_name": {"$regex": f"^{party_name}$", "$options": "i"}}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        query["date"] = date_q
    txns = await db.local_party_accounts.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    # Also include manual cashbook payments (ledger nikasi not from lp_settlement)
    existing_lp_ids = set()
    existing_refs = set()
    existing_ref_ids = set()  # Extract IDs from references like "voucher_payment:xxx"
    for t in txns:
        if t.get("txn_type") == "payment" and t.get("source_type") == "settlement":
            existing_lp_ids.add(t.get("id", ""))
        ref = t.get("reference", "")
        if ref:
            existing_refs.add(ref)
            # Extract the ID portion from reference for matching ledger entries
            if ":" in ref:
                ref_id = ref.split(":", 1)[1]
                existing_ref_ids.add(ref_id)
    cb_query = {"account": "ledger", "txn_type": "nikasi",
                "category": {"$regex": f"^{party_name}$", "$options": "i"}}
    if kms_year: cb_query["kms_year"] = kms_year
    if season: cb_query["season"] = season
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        cb_query["date"] = date_q
    cb_payments = await db.cash_transactions.find(cb_query, {"_id": 0}).to_list(10000)
    # Exclude cashbook payments that are already reflected in local_party_accounts
    existing_linked_ids = set(t.get("id", "") for t in txns)
    for cb in cb_payments:
        linked_id = cb.get("linked_local_party_id", "")
        if linked_id and linked_id in existing_linked_ids:
            continue
        ref = cb.get("reference", "")
        if ref.startswith("local_party_ledger:") or ref.startswith("local_party:"):
            continue
        # Skip voucher payment ledger entries if the corresponding voucher_payment exists
        if ref.startswith("voucher_payment_ledger:"):
            ref_id = ref.split(":", 1)[1]
            if ref_id in existing_ref_ids:
                continue
        if ref and ref in existing_refs:
            continue
        txns.append({
            "id": cb.get("id", ""),
            "date": cb.get("date", ""),
            "party_name": party_name,
            "txn_type": "payment",
            "amount": cb.get("amount", 0),
            "description": cb.get("description", "CashBook Payment"),
            "source_type": "cashbook",
            "reference": cb.get("reference", ""),
            "kms_year": cb.get("kms_year", ""),
            "season": cb.get("season", ""),
            "created_by": cb.get("created_by", ""),
            "created_at": cb.get("created_at", "")
        })
    txns.sort(key=lambda x: (x.get("date", ""), x.get("created_at", "")))

    running_balance = 0
    report_rows = []
    for t in txns:
        if t.get("txn_type") == "debit":
            running_balance += t.get("amount", 0)
        elif t.get("txn_type") == "payment":
            running_balance -= t.get("amount", 0)
        report_rows.append({
            **t,
            "running_balance": round(running_balance, 2)
        })

    total_debit = sum(t["amount"] for t in txns if t.get("txn_type") == "debit")
    total_paid = sum(t["amount"] for t in txns if t.get("txn_type") == "payment")

    return {
        "party_name": party_name,
        "transactions": report_rows,
        "total_debit": round(total_debit, 2),
        "total_paid": round(total_paid, 2),
        "balance": round(total_debit - total_paid, 2),
        "total_entries": len(txns)
    }




@router.post("/local-party/manual")
async def add_manual_purchase(request: Request):
    """Manually add a purchase (debit) entry for a local party"""
    data = await request.json()
    party_name = data.get("party_name", "").strip()
    amount = float(data.get("amount", 0))
    if not party_name or amount <= 0:
        raise HTTPException(status_code=400, detail="Party name aur amount (>0) required hai")

    doc = {
        "id": str(uuid.uuid4()),
        "date": data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        "party_name": party_name,
        "txn_type": "debit",
        "amount": round(amount, 2),
        "description": data.get("description", "Manual Purchase"),
        "source_type": "manual",
        "reference": "",
        "kms_year": data.get("kms_year", ""),
        "season": data.get("season", ""),
        "created_by": data.get("created_by", "system"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.local_party_accounts.insert_one(doc)
    doc.pop("_id", None)

    # Auto create Cash Book Jama entry (purchase from local party)
    cb = {
        "id": str(uuid.uuid4()),
        "date": doc["date"],
        "account": "ledger",
        "txn_type": "jama",
        "category": party_name,
        "party_type": "Local Party",
        "description": f"Purchase: {party_name} - {data.get('description', 'Manual Purchase')} Rs.{amount}",
        "amount": round(amount, 2),
        "reference": f"lp_purchase:{doc['id'][:8]}",
        "kms_year": data.get("kms_year", ""),
        "season": data.get("season", ""),
        "created_by": data.get("created_by", "system"),
        "linked_local_party_id": doc["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(cb)

    return doc


@router.post("/local-party/settle")
async def settle_local_party(request: Request):
    """Settle (pay) a local party - auto creates Cash Book nikasi entry"""
    data = await request.json()
    party_name = data.get("party_name", "").strip()
    amount = float(data.get("amount", 0))
    if not party_name or amount <= 0:
        raise HTTPException(status_code=400, detail="Party name aur amount (>0) required hai")

    date = data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    notes = data.get("notes", "")
    kms_year = data.get("kms_year", "")
    season = data.get("season", "")
    username = data.get("created_by", "system")

    # Create payment entry in local party accounts
    pay_txn = {
        "id": str(uuid.uuid4()),
        "date": date,
        "party_name": party_name,
        "txn_type": "payment",
        "amount": round(amount, 2),
        "description": f"Payment to {party_name}" + (f" - {notes}" if notes else ""),
        "source_type": "settlement",
        "reference": "",
        "kms_year": kms_year,
        "season": season,
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.local_party_accounts.insert_one(pay_txn)

    # Auto create Cash Book entry (Nikasi)
    cb = {
        "id": str(uuid.uuid4()),
        "date": date,
        "account": "cash",
        "txn_type": "nikasi",
        "category": party_name,
        "party_type": "Local Party",
        "description": f"Local Party Payment: {party_name} - Rs.{amount}" + (f" ({notes})" if notes else ""),
        "amount": round(amount, 2),
        "reference": f"local_party:{pay_txn['id'][:8]}",
        "kms_year": kms_year,
        "season": season,
        "created_by": username,
        "linked_local_party_id": pay_txn["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(cb)

    # Ledger Nikasi - reduce party outstanding
    ledger_cb = {
        "id": str(uuid.uuid4()),
        "date": date,
        "account": "ledger",
        "txn_type": "nikasi",
        "category": party_name,
        "party_type": "Local Party",
        "description": f"Local Party Payment: {party_name} - Rs.{amount}" + (f" ({notes})" if notes else ""),
        "amount": round(amount, 2),
        "reference": f"local_party_ledger:{pay_txn['id'][:8]}",
        "kms_year": kms_year,
        "season": season,
        "created_by": username,
        "linked_local_party_id": pay_txn["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(ledger_cb)

    return {"success": True, "message": f"Rs.{amount} payment to {party_name} recorded", "txn_id": pay_txn["id"]}


@router.delete("/local-party/{txn_id}")
async def delete_local_party_txn(txn_id: str):
    txn = await db.local_party_accounts.find_one({"id": txn_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    # If it was a settlement payment, also delete linked cash book entry
    if txn.get("txn_type") == "payment" and txn.get("source_type") == "settlement":
        await db.cash_transactions.delete_many({"linked_local_party_id": txn_id})

    await db.local_party_accounts.delete_one({"id": txn_id})
    return {"message": "Deleted", "id": txn_id}


# ============ EXPORT ============

@router.get("/local-party/excel")
async def export_local_party_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from io import BytesIO

    summary = await get_local_party_summary(kms_year=kms_year, season=season)
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.local_party_accounts.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Local Party Account"
    hf = PatternFill(start_color="065f46", end_color="065f46", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:F1')
    ws['A1'] = "Local Party Account / स्थानीय पार्टी खाता"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Summary
    ws.cell(row=3, column=1, value="Party Summary").font = Font(bold=True, size=11)
    for col, h in enumerate(['Party Name', 'Total Debit (Rs.)', 'Total Paid (Rs.)', 'Balance (Rs.)', 'Entries'], 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = hf
        c.font = hfont
        c.border = tb
    row = 5
    for p in summary.get("parties", []):
        for col, v in enumerate([p["party_name"], p["total_debit"], p["total_paid"], p["balance"], p["txn_count"]], 1):
            ws.cell(row=row, column=col, value=v).border = tb
        row += 1
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=summary.get("grand_total_debit", 0)).font = Font(bold=True)
    ws.cell(row=row, column=2).border = tb
    ws.cell(row=row, column=3, value=summary.get("grand_total_paid", 0)).font = Font(bold=True)
    ws.cell(row=row, column=3).border = tb
    ws.cell(row=row, column=4, value=summary.get("grand_balance", 0)).font = Font(bold=True, color="FF0000")
    ws.cell(row=row, column=4).border = tb
    row += 2

    # Transactions
    ws.cell(row=row, column=1, value="Transactions").font = Font(bold=True, size=11)
    row += 1
    for col, h in enumerate(['Date', 'Party', 'Type', 'Amount (Rs.)', 'Description', 'Source'], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hf
        c.font = hfont
        c.border = tb
    row += 1
    for t in txns:
        vals = [
            t.get("date", ""),
            t.get("party_name", ""),
            "Payment" if t.get("txn_type") == "payment" else "Purchase",
            t.get("amount", 0),
            t.get("description", ""),
            t.get("source_type", "")
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v).border = tb
        row += 1

    for letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[letter].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=local_party_account_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/local-party/pdf")
async def export_local_party_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO

    summary = await get_local_party_summary(kms_year=kms_year, season=season)
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.local_party_accounts.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Local Party Account / स्थानीय पार्टी खाता", styles['Title']))
    elements.append(Spacer(1, 12))

    # Summary table
    sum_data = [['Party Name', 'Total Debit', 'Total Paid', 'Balance', 'Entries']]
    for p in summary.get("parties", []):
        sum_data.append([p["party_name"], f"Rs.{p['total_debit']}", f"Rs.{p['total_paid']}", f"Rs.{p['balance']}", str(p["txn_count"])])
    sum_data.append(['GRAND TOTAL', f"Rs.{summary.get('grand_total_debit', 0)}", f"Rs.{summary.get('grand_total_paid', 0)}", f"Rs.{summary.get('grand_balance', 0)}", ''])
    st = Table(sum_data, colWidths=[180, 100, 100, 100, 60])
    st.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#065f46')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (3, -1), (3, -1), colors.red),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 20))

    # Transaction table
    elements.append(Paragraph("Transactions", styles['Heading2']))
    t_data = [['Date', 'Party', 'Type', 'Amount', 'Description', 'Source']]
    for t in txns:
        t_data.append([
            t.get("date", ""),
            t.get("party_name", "")[:20],
            "Payment" if t.get("txn_type") == "payment" else "Purchase",
            f"Rs.{t.get('amount', 0)}",
            t.get("description", "")[:35],
            t.get("source_type", "")
        ])
    tt = Table(t_data, colWidths=[70, 120, 55, 70, 200, 70])
    tt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#065f46')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecfdf5')]),
    ]))
    elements.append(tt)

    doc.build(elements)
    buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=local_party_account_{datetime.now().strftime('%Y%m%d')}.pdf"})
