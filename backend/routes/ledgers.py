from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import round_amount
import uuid
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils import export_helpers as _eh_default_font  # noqa: F401
from openpyxl.utils import get_column_letter
from utils.report_helper import get_columns, get_entry_row, get_total_row, get_excel_headers, get_pdf_headers, get_excel_widths, get_pdf_widths_mm, col_count
from utils.date_format import fmt_date

router = APIRouter()

# ============ PHASE 5: CONSOLIDATED LEDGERS ============

@router.get("/reports/outstanding")
async def report_outstanding(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Outstanding Report - all pending payments/deliveries across modules"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    # DC pending deliveries
    dcs = await db.dc_entries.find(query, {"_id": 0}).to_list(1000)
    all_dels = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    dc_outstanding = []
    for dc in dcs:
        delivered = round(sum(d.get("quantity_qntl", 0) for d in all_dels if d.get("dc_id") == dc["id"]), 2)
        pending = round(dc["quantity_qntl"] - delivered, 2)
        if pending > 0:
            dc_outstanding.append({"dc_number": dc.get("dc_number", ""), "allotted": dc["quantity_qntl"], "delivered": delivered, "pending": pending, "deadline": dc.get("deadline", ""), "rice_type": dc.get("rice_type", "")})
    dc_pending_total = round(sum(d["pending"] for d in dc_outstanding), 2)

    # MSP payment pending
    msp_payments = await db.msp_payments.find(query, {"_id": 0}).to_list(5000)
    total_delivered_qntl = round(sum(d.get("quantity_qntl", 0) for d in all_dels), 2)
    total_msp_paid_qty = round(sum(p.get("quantity_qntl", 0) for p in msp_payments), 2)
    total_msp_paid_amt = round(sum(p.get("amount", 0) for p in msp_payments), 2)
    msp_pending_qty = round(total_delivered_qntl - total_msp_paid_qty, 2)

    # Truck payment pending (entries with cash_paid < expected)
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(10000)
    truck_map = {}
    for e in entries:
        truck = e.get("truck_no", "Unknown")
        if truck not in truck_map:
            truck_map[truck] = {"truck_no": truck, "total_trips": 0, "total_qty_qntl": 0, "total_cash_paid": 0, "total_diesel_paid": 0}
        truck_map[truck]["total_trips"] += 1
        truck_map[truck]["total_qty_qntl"] = round(truck_map[truck]["total_qty_qntl"] + (e.get("final_w", 0) / 100), 2)
        truck_map[truck]["total_cash_paid"] = round(truck_map[truck]["total_cash_paid"] + (e.get("cash_paid", 0)), 2)
        truck_map[truck]["total_diesel_paid"] = round(truck_map[truck]["total_diesel_paid"] + (e.get("diesel_paid", 0)), 2)

    # Agent summary
    agent_map = {}
    for e in entries:
        agent = e.get("agent_name", "Unknown")
        if not agent: agent = "Unknown"
        if agent not in agent_map:
            agent_map[agent] = {"agent_name": agent, "total_entries": 0, "total_qty_qntl": 0}
        agent_map[agent]["total_entries"] += 1
        agent_map[agent]["total_qty_qntl"] = round(agent_map[agent]["total_qty_qntl"] + (e.get("final_w", 0) / 100), 2)

    # FRK purchase outstanding
    frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(5000)
    frk_party_map = {}
    for p in frk_purchases:
        party = p.get("party_name", "Unknown")
        if party not in frk_party_map:
            frk_party_map[party] = {"party_name": party, "total_qty": 0, "total_amount": 0}
        frk_party_map[party]["total_qty"] = round(frk_party_map[party]["total_qty"] + (p.get("quantity_qntl", 0)), 2)
        frk_party_map[party]["total_amount"] = round(frk_party_map[party]["total_amount"] + (p.get("total_amount", 0)), 2)

    return {
        "dc_outstanding": {"items": dc_outstanding, "total_pending_qntl": dc_pending_total, "count": len(dc_outstanding)},
        "msp_outstanding": {"total_delivered_qntl": total_delivered_qntl, "total_paid_qty": total_msp_paid_qty, "total_paid_amount": total_msp_paid_amt, "pending_qty": msp_pending_qty},
        "trucks": list(truck_map.values()),
        "agents": list(agent_map.values()),
        "frk_parties": list(frk_party_map.values()),
    }


@router.get("/reports/party-ledger")
async def report_party_ledger(party_name: Optional[str] = None, party_type: Optional[str] = None,
                                kms_year: Optional[str] = None, season: Optional[str] = None,
                                date_from: Optional[str] = None, date_to: Optional[str] = None):
    """Party Ledger - all transactions for a specific party or all parties"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        if date_q: query["date"] = date_q

    ledger = []

    # Paddy entries (Truck only - Agent payments come from Cash Book)
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(10000)
    # Paddy entries - Truck ledger (cash + diesel payments go here)
    if not party_type or party_type == "truck":
        for e in entries:
            truck = e.get("truck_no", "")
            if not truck: continue
            if party_name and truck.lower() != party_name.lower(): continue
            total_paid = round(e.get("cash_paid", 0) + e.get("diesel_paid", 0), 2)
            if total_paid > 0:
                ledger.append({"date": e.get("date", ""), "party_name": truck, "party_type": "Truck",
                    "description": f"Mandi: {e.get('mandi_name','')} | Cash: {e.get('cash_paid',0)} Diesel: {e.get('diesel_paid',0)}",
                    "debit": 0, "credit": total_paid, "ref": e.get("id", "")[:8]})

    # Cash Book categories → Party Ledger (auto)
    if not party_type or party_type == "cash_party":
        cash_query = dict(query)
        cash_txns = await db.cash_transactions.find(cash_query, {"_id": 0}).to_list(10000)
        for t in cash_txns:
            cat = (t.get("category") or "").strip()
            if not cat: continue
            # Skip system categories (Cash Payment, Diesel Payment etc.)
            if cat.lower() in ("cash payment", "diesel payment", "cash paid", "diesel", "cash paid (entry)", "diesel (entry)"): continue
            # Skip types that have their own dedicated sections
            if t.get("party_type") in ("Agent", "Hemali", "Sale Book", "Purchase Voucher"): continue
            if party_name and cat.lower() != party_name.lower(): continue
            # Skip auto-ledger entries (they are duplicates with reversed txn_type)
            if "_ledger:" in (t.get("reference") or ""): continue
            is_jama = t.get("txn_type") == "jama"
            ledger.append({"date": t.get("date", ""), "party_name": cat, "party_type": "Cash Party",
                "description": t.get("description", "") or f"{'Jama' if is_jama else 'Nikasi'}: ₹{t.get('amount',0)}",
                "debit": round(t.get("amount", 0), 2) if not is_jama else 0,
                "credit": round(t.get("amount", 0), 2) if is_jama else 0,
                "ref": t.get("id", "")[:8]})

    # Agent payments (from cash_transactions with party_type=Agent)
    if not party_type or party_type == "Agent":
        agent_query = dict(query)
        agent_query["party_type"] = "Agent"
        agent_txns = await db.cash_transactions.find(agent_query, {"_id": 0}).to_list(10000)
        for t in agent_txns:
            cat = (t.get("category") or "").strip()
            if not cat: continue
            if party_name and cat.lower() != party_name.lower(): continue
            # Skip auto-ledger entries (they are duplicates with reversed txn_type)
            if "_ledger:" in (t.get("reference") or ""): continue
            is_jama = t.get("txn_type") == "jama"
            ledger.append({"date": t.get("date", ""), "party_name": cat, "party_type": "Agent",
                "description": t.get("description", "") or f"{'Jama' if is_jama else 'Nikasi'}: ₹{t.get('amount',0)}",
                "debit": round(t.get("amount", 0), 2) if not is_jama else 0,
                "credit": round(t.get("amount", 0), 2) if is_jama else 0,
                "ref": t.get("id", "")[:8]})

    # FRK purchases
    if not party_type or party_type == "frk_party":
        frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(5000)
        for p in frk_purchases:
            party = p.get("party_name", "")
            if not party: continue
            if party_name and party.lower() != party_name.lower(): continue
            ledger.append({"date": p.get("date", ""), "party_name": party, "party_type": "FRK Seller",
                "description": f"FRK: {p.get('quantity_qntl',0)}Q @ ₹{p.get('rate_per_qntl',0)}/Q",
                "debit": round(p.get("total_amount", 0), 2), "credit": 0, "ref": p.get("id", "")[:8]})

    # By-product sales
    if not party_type or party_type == "buyer":
        bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(5000)
        for s in bp_sales:
            buyer = s.get("buyer_name", "")
            if not buyer: continue
            if party_name and buyer.lower() != party_name.lower(): continue
            ledger.append({"date": s.get("date", ""), "party_name": buyer, "party_type": "Buyer",
                "description": f"{(s.get('product','')).capitalize()}: {s.get('quantity_qntl',0)}Q @ ₹{s.get('rate_per_qntl',0)}/Q",
                "debit": 0, "credit": round(s.get("total_amount", 0), 2), "ref": s.get("id", "")[:8]})

    # Private Paddy Purchase (debit = purchase amount, credit = advance paid only)
    if not party_type or party_type == "pvt_paddy":
        pvt_paddy = await db.private_paddy.find(query, {"_id": 0}).to_list(5000)
        for p in pvt_paddy:
            party = p.get("party_name", "")
            mandi = p.get("mandi_name", "")
            if not party: continue
            display_name = f"{party} - {mandi}" if mandi else party
            if party_name and display_name.lower() != party_name.lower() and party.lower() != party_name.lower(): continue
            # Debit: total purchase amount (what we owe)
            ledger.append({"date": p.get("date", ""), "party_name": display_name, "party_type": "Pvt Paddy Purchase",
                "description": f"Paddy: {p.get('final_qntl',0)}Q @ Rs.{p.get('rate_per_qntl',0)}/Q = Rs.{p.get('total_amount',0)}",
                "debit": round(p.get("total_amount", 0), 2), "credit": 0, "ref": p.get("id", "")[:8]})
            # Credit: advance paid only (cash/diesel go to truck payment, not party)
            advance_paid = float(p.get("paid_amount", 0) or 0)
            if advance_paid > 0:
                ledger.append({"date": p.get("date", ""), "party_name": display_name, "party_type": "Pvt Paddy Purchase",
                    "description": f"Advance Paid: Rs.{advance_paid}",
                    "debit": 0, "credit": round(advance_paid, 2), "ref": p.get("id", "")[:8]})

    # Rice Sale (debit = 0, credit = sale amount)
    if not party_type or party_type == "rice_buyer":
        rice_sales = await db.rice_sales.find(query, {"_id": 0}).to_list(5000)
        for s in rice_sales:
            party = s.get("party_name", "")
            if not party: continue
            if party_name and party.lower() != party_name.lower(): continue
            ledger.append({"date": s.get("date", ""), "party_name": party, "party_type": "Rice Buyer",
                "description": f"Rice Sale: {s.get('quantity_qntl',0)}Q ({s.get('rice_type','')}) @ ₹{s.get('rate_per_qntl',0)}/Q = ₹{s.get('total_amount',0)}",
                "debit": 0, "credit": round(s.get("total_amount", 0), 2), "ref": s.get("id", "")[:8]})

    # Private Payments (received/paid)
    if not party_type or party_type in ("pvt_paddy", "rice_buyer", "pvt_payment"):
        pvt_payments = await db.private_payments.find(query if kms_year or season else {}, {"_id": 0}).to_list(5000)
        for pay in pvt_payments:
            pn = pay.get("party_name", "")
            if not pn: continue
            if party_name and pn.lower() != party_name.lower(): continue
            if pay.get("ref_type") == "paddy_purchase":
                if party_type and party_type not in ("pvt_paddy", "pvt_payment"): continue
                ledger.append({"date": pay.get("date", ""), "party_name": pn, "party_type": "Pvt Paddy Purchase",
                    "description": f"Payment: Rs.{pay.get('amount',0)} ({pay.get('mode','cash')})",
                    "debit": 0, "credit": round(pay.get("amount", 0), 2), "ref": pay.get("id", "")[:8]})
            elif pay.get("ref_type") == "rice_sale":
                if party_type and party_type not in ("rice_buyer", "pvt_payment"): continue
                ledger.append({"date": pay.get("date", ""), "party_name": pn, "party_type": "Rice Buyer",
                    "description": f"Payment Received: Rs.{pay.get('amount',0)} ({pay.get('mode','cash')})",
                    "debit": round(pay.get("amount", 0), 2), "credit": 0, "ref": pay.get("id", "")[:8]})

    # Hemali Sardar payments (from hemali_payments directly)
    if not party_type or party_type == "Hemali":
        hemali_query = dict(query)
        hemali_query["status"] = "paid"
        hemali_payments = await db.hemali_payments.find(hemali_query, {"_id": 0}).to_list(10000)
        for p in hemali_payments:
            sn = p.get("sardar_name", "")
            if not sn: continue
            if party_name and sn.lower() != party_name.lower(): continue
            items_desc = ", ".join(f"{i.get('item_name','')} x{i.get('quantity',0)}" for i in p.get("items", []))
            # Credit: amount paid to sardar (we paid, so credit in their account)
            ledger.append({"date": p.get("date", ""), "party_name": sn, "party_type": "Hemali",
                "description": f"Hemali Payment: {items_desc} | Total: Rs.{p.get('total',0)} | Paid: Rs.{p.get('amount_paid',0)}",
                "debit": 0, "credit": round(p.get("amount_paid", 0), 2), "ref": p.get("id", "")[:8]})
            # Debit: if advance given (sardar owes us)
            if p.get("new_advance", 0) > 0:
                ledger.append({"date": p.get("date", ""), "party_name": sn, "party_type": "Hemali",
                    "description": f"Advance Given: Rs.{p.get('new_advance',0)} (extra paid)",
                    "debit": round(p.get("new_advance", 0), 2), "credit": 0, "ref": p.get("id", "")[:8]})
            # Credit: if advance deducted (reduces what sardar owes)
            if p.get("advance_deducted", 0) > 0:
                ledger.append({"date": p.get("date", ""), "party_name": sn, "party_type": "Hemali",
                    "description": f"Advance Deducted: Rs.{p.get('advance_deducted',0)}",
                    "debit": 0, "credit": round(p.get("advance_deducted", 0), 2), "ref": p.get("id", "")[:8]})

    # Sale Book parties (from local_party_accounts)
    if not party_type or party_type == "sale_book":
        lp_query = dict(query)
        lp_query["source_type"] = {"$in": [
            "sale_voucher", "sale_voucher_payment",
            "bp_sale", "bp_sale_advance",
            "bp_sale_pka", "bp_sale_pka_payment",
            "bp_sale_ka", "bp_sale_ka_payment", "bp_sale_ka_oil_premium",
        ]}
        if party_name: lp_query["party_name"] = {"$regex": f"^{party_name}$", "$options": "i"}
        lp_txns = await db.local_party_accounts.find(lp_query, {"_id": 0}).to_list(10000)
        for t in lp_txns:
            pn = t.get("party_name", "")
            if not pn: continue
            src = t.get("source_type", "")
            amt = round(t.get("amount", 0), 2)
            tx = (t.get("txn_type") or "").lower()
            # Map to debit/credit based on txn_type stored:
            # debit  = party owes us (sale) — Debit column
            # payment= money received / credit (reduces what party owes) — Credit column
            if tx == "debit":
                ledger.append({"date": t.get("date", ""), "party_name": pn, "party_type": "Sale Book",
                    "description": t.get("description", "") or f"Sale: Rs.{amt}",
                    "debit": amt, "credit": 0, "ref": t.get("id", "")[:8]})
            else:
                ledger.append({"date": t.get("date", ""), "party_name": pn, "party_type": "Sale Book",
                    "description": t.get("description", "") or f"Payment: Rs.{amt}",
                    "debit": 0, "credit": amt, "ref": t.get("id", "")[:8]})

    # Purchase Voucher parties (from local_party_accounts)
    if not party_type or party_type == "purchase_voucher":
        pv_query = dict(query)
        pv_query["source_type"] = {"$in": ["purchase_voucher", "purchase_voucher_payment"]}
        if party_name: pv_query["party_name"] = {"$regex": f"^{party_name}$", "$options": "i"}
        pv_txns = await db.local_party_accounts.find(pv_query, {"_id": 0}).to_list(10000)
        for t in pv_txns:
            pn = t.get("party_name", "")
            if not pn: continue
            src = t.get("source_type", "")
            amt = round(t.get("amount", 0), 2)
            if src == "purchase_voucher":
                # Purchase amount: we owe party → Credit
                ledger.append({"date": t.get("date", ""), "party_name": pn, "party_type": "Purchase Voucher",
                    "description": t.get("description", "") or f"Purchase: Rs.{amt}",
                    "debit": 0, "credit": amt, "ref": t.get("id", "")[:8]})
            elif src == "purchase_voucher_payment":
                # Payment made: reduces what we owe → Debit
                ledger.append({"date": t.get("date", ""), "party_name": pn, "party_type": "Purchase Voucher",
                    "description": t.get("description", "") or f"Payment: Rs.{amt}",
                    "debit": amt, "credit": 0, "ref": t.get("id", "")[:8]})

    ledger.sort(key=lambda x: x.get("date", ""))

    # Party list for filter
    parties = set()
    for item in ledger: parties.add((item["party_name"], item["party_type"]))
    party_list = [{"name": n, "type": t} for n, t in sorted(parties)]

    return {"ledger": ledger, "party_list": party_list, "total_debit": round(sum(l["debit"] for l in ledger), 2),
            "total_credit": round(sum(l["credit"] for l in ledger), 2)}


@router.get("/reports/outstanding/excel")
async def export_outstanding_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    data = await report_outstanding(kms_year=kms_year, season=season)
    wb = Workbook(); ws = wb.active; ws.title = "Outstanding Report"
    ncols = 6
    
    title = "Outstanding Report / बकाया रिपोर्ट"
    if kms_year: title += f" | FY {kms_year}"
    style_excel_title(ws, title, ncols)
    
    # DC Outstanding
    row = 4
    ws.cell(row=row, column=1, value="DC PENDING DELIVERIES").font = Font(bold=True, size=11, color=COLORS['nikasi_text'])
    row += 1
    dc_headers = ['DC No', 'Allotted(Q)', 'Delivered(Q)', 'Pending(Q)', 'Deadline', 'Type']
    for col, h in enumerate(dc_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_excel_header_row(ws, row, ncols)
    row += 1; dc_data_start = row
    for d in data["dc_outstanding"]["items"]:
        for col, v in enumerate([d["dc_number"], d["allotted"], d["delivered"], d["pending"], d["deadline"], d["rice_type"]], 1):
            ws.cell(row=row, column=col, value=v)
        row += 1
    if data["dc_outstanding"]["items"]:
        style_excel_data_rows(ws, dc_data_start, row - 1, ncols, dc_headers)
    ws.cell(row=row, column=1, value="Total Pending")
    ws.cell(row=row, column=4, value=data["dc_outstanding"]["total_pending_qntl"])
    style_excel_total_row(ws, row, ncols)
    
    # MSP Outstanding
    row += 2
    ws.cell(row=row, column=1, value="MSP PAYMENT PENDING").font = Font(bold=True, size=11, color=COLORS['subtitle_text'])
    row += 1
    msp_headers = ['Metric', 'Value']
    for col, h in enumerate(msp_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_excel_header_row(ws, row, 2)
    row += 1; msp_start = row
    for label, val in [("Total Delivered (Q)", data["msp_outstanding"]["total_delivered_qntl"]),
                        ("Paid Qty (Q)", data["msp_outstanding"]["total_paid_qty"]),
                        ("Paid Amount (Rs.)", data["msp_outstanding"]["total_paid_amount"]),
                        ("Pending Qty (Q)", data["msp_outstanding"]["pending_qty"])]:
        ws.cell(row=row, column=1, value=label); ws.cell(row=row, column=2, value=val)
        row += 1
    style_excel_data_rows(ws, msp_start, row - 1, 2, msp_headers)
    
    # Trucks
    row += 1
    ws.cell(row=row, column=1, value="TRUCK SUMMARY").font = Font(bold=True, size=11, color=COLORS['date_text'])
    row += 1
    truck_headers = ['Truck No', 'Trips', 'Qty(Q)', 'Cash Paid', 'Diesel Paid']
    for col, h in enumerate(truck_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_excel_header_row(ws, row, 5)
    row += 1; truck_start = row
    for t in data["trucks"]:
        for col, v in enumerate([t["truck_no"], t["total_trips"], t["total_qty_qntl"], t["total_cash_paid"], t["total_diesel_paid"]], 1):
            ws.cell(row=row, column=col, value=v)
        row += 1
    if data["trucks"]:
        style_excel_data_rows(ws, truck_start, row - 1, 5, truck_headers)
    
    for i in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=outstanding_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/reports/outstanding/pdf")
async def export_outstanding_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib import colors
    from io import BytesIO
    from utils.export_helpers import get_pdf_table_style
    from utils.branding_helper import get_pdf_company_header_from_db
    data = await report_outstanding(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = get_pdf_styles()
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph("Outstanding Report / बकाया रिपोर्ट", styles['Title'])); elements.append(Spacer(1, 12))
    # DC pending
    elements.append(Paragraph("DC Pending Deliveries", styles['Heading2'])); elements.append(Spacer(1, 4))
    ddata = [['DC No', 'Allotted(Q)', 'Delivered(Q)', 'Pending(Q)', 'Deadline', 'Type']]
    for d in data["dc_outstanding"]["items"]:
        ddata.append([d["dc_number"], d["allotted"], d["delivered"], d["pending"], fmt_date(d["deadline"]), d["rice_type"]])
    ddata.append(['TOTAL', '', '', data["dc_outstanding"]["total_pending_qntl"], '', ''])
    dt = RLTable(ddata, colWidths=[60, 60, 60, 60, 60, 50])
    dt.setStyle(TableStyle(get_pdf_table_style(len(ddata))))
    elements.append(dt); elements.append(Spacer(1, 12))
    # MSP
    elements.append(Paragraph("MSP Payment Pending", styles['Heading2'])); elements.append(Spacer(1, 4))
    mdata = [['Metric', 'Value'], ['Delivered(Q)', data['msp_outstanding']['total_delivered_qntl']], ['Paid Qty(Q)', data['msp_outstanding']['total_paid_qty']],
             ['Paid Amount(Rs.)', data['msp_outstanding']['total_paid_amount']], ['Pending Qty(Q)', data['msp_outstanding']['pending_qty']]]
    mt = RLTable(mdata, colWidths=[150, 100])
    mt.setStyle(TableStyle(get_pdf_table_style(len(mdata))))
    elements.append(mt); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=outstanding_{datetime.now().strftime('%Y%m%d')}.pdf"})


@router.get("/reports/party-ledger/excel")
async def export_party_ledger_excel(party_name: Optional[str] = None, party_type: Optional[str] = None,
                                     kms_year: Optional[str] = None, season: Optional[str] = None,
                                     date_from: Optional[str] = None, date_to: Optional[str] = None):
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS)
    
    data = await report_party_ledger(party_name=party_name, party_type=party_type, kms_year=kms_year, season=season, date_from=date_from, date_to=date_to)
    
    cols = get_columns("party_ledger_report")
    ncols = col_count(cols)
    headers = get_excel_headers(cols)
    widths = get_excel_widths(cols)
    
    wb = Workbook(); ws = wb.active; ws.title = "Party Ledger"
    
    title = "Party Ledger / खाता बही"
    if party_name: title += f" - {party_name}"
    if kms_year: title += f" | FY {kms_year}"
    subtitle = ""
    if date_from or date_to:
        date_parts = []
        if date_from: date_parts.append(f"From: {date_from}")
        if date_to: date_parts.append(f"To: {date_to}")
        subtitle = " | ".join(date_parts)
    style_excel_title(ws, title, ncols, subtitle)
    
    header_row = 4
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_excel_header_row(ws, header_row, ncols)
    
    data_start = header_row + 1
    row_num = data_start
    for l in data["ledger"]:
        row_data = dict(l)
        if row_data.get("debit", 0) == 0: row_data["debit"] = ""
        if row_data.get("credit", 0) == 0: row_data["credit"] = ""
        if row_data.get("party_type") == "Pvt Paddy Purchase":
            row_data["party_name"] = f"[Pvt] {row_data.get('party_name', '')}"
        vals = get_entry_row(row_data, cols)
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col_idx, value=v)
            if cols[col_idx-1]["align"] == "right": c.alignment = Alignment(horizontal='right')
            if cols[col_idx-1]["type"] == "number" and isinstance(v, (int, float)): c.number_format = '#,##0.00'
        row_num += 1
    
    if data["ledger"]:
        style_excel_data_rows(ws, data_start, row_num - 1, ncols, headers)
    
    totals = {"total_debit": data["total_debit"], "total_credit": data["total_credit"]}
    total_vals = get_total_row(totals, cols)
    ws.cell(row=row_num, column=1, value="TOTAL / कुल")
    for col_idx, val in enumerate(total_vals, 1):
        if val is not None:
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.alignment = Alignment(horizontal='right')
    style_excel_total_row(ws, row_num, ncols)
    
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=party_ledger_{datetime.now().strftime('%Y%m%d')}.xlsx"})


async def _generate_party_ledger_pdf_bytes(party_name=None, party_type=None, kms_year=None, season=None, date_from=None, date_to=None):
    """Generate party ledger PDF bytes - shared between download endpoint and WhatsApp route."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from io import BytesIO
    from utils.export_helpers import get_pdf_table_style
    data = await report_party_ledger(party_name=party_name, party_type=party_type, kms_year=kms_year, season=season, date_from=date_from, date_to=date_to)
    
    cols = get_columns("party_ledger_report")
    headers = get_pdf_headers(cols)
    col_widths = [w*mm for w in get_pdf_widths_mm(cols)]
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=8*mm, rightMargin=8*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []; styles = get_pdf_styles()
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    title = "Party Ledger / खाता बही"
    if party_name: title += f" - {party_name}"
    if date_from or date_to:
        date_parts = []
        if date_from: date_parts.append(f"From: {date_from}")
        if date_to: date_parts.append(f"To: {date_to}")
        title += f" ({' | '.join(date_parts)})"
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#1a365d'), alignment=TA_CENTER)
    elements.append(Paragraph(title, title_style)); elements.append(Spacer(1, 8))
    
    desc_style = ParagraphStyle('desc', fontName='FreeSans', fontSize=6.5, leading=8, alignment=TA_LEFT)
    party_style = ParagraphStyle('party', fontName='FreeSans', fontSize=6.5, leading=8, alignment=TA_LEFT)
    
    table_data = [headers]
    for l in data["ledger"]:
        row_data = dict(l)
        if row_data.get("debit", 0) == 0: row_data["debit"] = "-"
        if row_data.get("credit", 0) == 0: row_data["credit"] = "-"
        if row_data.get("party_type") == "Pvt Paddy Purchase":
            row_data["party_name"] = f"[Pvt] {row_data.get('party_name', '')}"
        row_vals = get_entry_row(row_data, cols)
        out = []
        for i, v in enumerate(row_vals):
            if cols[i]["field"] == "description":
                out.append(Paragraph(str(v), desc_style))
            elif cols[i]["field"] == "party_name":
                out.append(Paragraph(str(v), party_style))
            else:
                out.append(str(v) if v != "" else "")
        table_data.append(out)
    
    totals = {"total_debit": data["total_debit"], "total_credit": data["total_credit"]}
    total_vals = get_total_row(totals, cols)
    total_row = []
    for i, val in enumerate(total_vals):
        if i == 0: total_row.append("TOTAL / कुल")
        elif val is not None: total_row.append(str(val))
        else: total_row.append("")
    table_data.append(total_row)
    
    first_right = next((i for i, c in enumerate(cols) if c["align"] == "right"), 4)
    cols_info = [{'header': h} for h in headers]
    pdf_style = get_pdf_table_style(len(table_data), cols_info)
    pdf_style.append(('ALIGN', (first_right, 0), (-1, -1), 'RIGHT'))
    pdf_style.append(('VALIGN', (0,0), (-1,-1), 'TOP'))
    table = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(pdf_style))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return buffer.getvalue()


@router.get("/reports/party-ledger/pdf")
async def export_party_ledger_pdf(party_name: Optional[str] = None, party_type: Optional[str] = None,
                                    kms_year: Optional[str] = None, season: Optional[str] = None,
                                    date_from: Optional[str] = None, date_to: Optional[str] = None):
    pdf_bytes = await _generate_party_ledger_pdf_bytes(party_name=party_name, party_type=party_type, kms_year=kms_year, season=season, date_from=date_from, date_to=date_to)
    return Response(content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=party_ledger_{datetime.now().strftime('%Y%m%d')}.pdf"})


# Include the router in the main app
