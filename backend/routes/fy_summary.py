from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
from datetime import datetime
from typing import Optional
from database import db

router = APIRouter()

def get_prev_fy(kms_year: str) -> str:
    parts = kms_year.split('-')
    if len(parts) == 2:
        try:
            return f"{int(parts[0])-1}-{int(parts[1])-1}"
        except (ValueError, IndexError):
            pass
    return ""

def get_next_fy(kms_year: str) -> str:
    parts = kms_year.split('-')
    if len(parts) == 2:
        try:
            return f"{int(parts[0])+1}-{int(parts[1])+1}"
        except (ValueError, IndexError):
            pass
    return ""


@router.get("/fy-summary")
async def get_fy_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    prev_fy = get_prev_fy(kms_year) if kms_year else ""
    prev_q = {}
    if prev_fy: prev_q["kms_year"] = prev_fy
    if season: prev_q["season"] = season

    # Load saved opening balances for this FY
    saved_ob = None
    if kms_year:
        saved_ob = await db.opening_balances.find_one({"kms_year": kms_year}, {"_id": 0})

    # ===== 1. CASH & BANK =====
    cash_txns = await db.cash_transactions.find(query, {"_id": 0}).to_list(100000)
    cash_in = sum(t.get("amount", 0) for t in cash_txns if t.get("account") == "cash" and t.get("txn_type") == "jama")
    cash_out = sum(t.get("amount", 0) for t in cash_txns if t.get("account") == "cash" and t.get("txn_type") == "nikasi")
    bank_in = sum(t.get("amount", 0) for t in cash_txns if t.get("account") == "bank" and t.get("txn_type") == "jama")
    bank_out = sum(t.get("amount", 0) for t in cash_txns if t.get("account") == "bank" and t.get("txn_type") == "nikasi")

    ob_cash, ob_bank = 0.0, 0.0
    if saved_ob:
        ob_cash = saved_ob.get("cash", 0.0)
        ob_bank = saved_ob.get("bank", 0.0)
    elif prev_fy:
        prev_cash = await db.cash_transactions.find(prev_q, {"_id": 0}).to_list(100000)
        p_ci = sum(t.get("amount", 0) for t in prev_cash if t.get("account") == "cash" and t.get("txn_type") == "jama")
        p_co = sum(t.get("amount", 0) for t in prev_cash if t.get("account") == "cash" and t.get("txn_type") == "nikasi")
        p_bi = sum(t.get("amount", 0) for t in prev_cash if t.get("account") == "bank" and t.get("txn_type") == "jama")
        p_bo = sum(t.get("amount", 0) for t in prev_cash if t.get("account") == "bank" and t.get("txn_type") == "nikasi")
        prev_saved = await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0})
        if prev_saved:
            ob_cash = round(prev_saved.get("cash", 0) + p_ci - p_co, 2)
            ob_bank = round(prev_saved.get("bank", 0) + p_bi - p_bo, 2)
        else:
            ob_cash = round(p_ci - p_co, 2)
            ob_bank = round(p_bi - p_bo, 2)

    cash_section = {
        "opening_cash": round(ob_cash, 2), "cash_in": round(cash_in, 2), "cash_out": round(cash_out, 2),
        "closing_cash": round(ob_cash + cash_in - cash_out, 2),
        "opening_bank": round(ob_bank, 2), "bank_in": round(bank_in, 2), "bank_out": round(bank_out, 2),
        "closing_bank": round(ob_bank + bank_in - bank_out, 2),
    }

    # ===== 2. PADDY STOCK =====
    mill_entries = await db.mill_entries.find(query, {"_id": 0, "qntl": 1, "bag": 1}).to_list(10000)
    paddy_in = round(sum(e.get("qntl", 0) - e.get("bag", 0) / 100 for e in mill_entries), 2)
    milling_entries = await db.milling_entries.find(query, {"_id": 0}).to_list(10000)
    paddy_used = round(sum(e.get("paddy_input_qntl", 0) for e in milling_entries), 2)

    ob_paddy = 0.0
    if saved_ob and "paddy_stock" in saved_ob:
        ob_paddy = saved_ob.get("paddy_stock", 0.0)
    elif prev_fy:
        prev_saved = await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0})
        prev_me = await db.mill_entries.find(prev_q, {"_id": 0, "qntl": 1, "bag": 1}).to_list(10000)
        prev_mi = await db.milling_entries.find(prev_q, {"_id": 0, "paddy_input_qntl": 1}).to_list(10000)
        prev_paddy_in = sum(e.get("qntl", 0) - e.get("bag", 0) / 100 for e in prev_me)
        prev_paddy_used = sum(e.get("paddy_input_qntl", 0) for e in prev_mi)
        prev_ob_paddy = prev_saved.get("paddy_stock", 0) if prev_saved else 0
        ob_paddy = round(prev_ob_paddy + prev_paddy_in - prev_paddy_used, 2)

    paddy_section = {
        "opening_stock": ob_paddy, "paddy_in": paddy_in, "paddy_used": paddy_used,
        "closing_stock": round(ob_paddy + paddy_in - paddy_used, 2)
    }

    # ===== 3. MILLING SUMMARY =====
    total_rice = round(sum(e.get("rice_qntl", 0) for e in milling_entries), 2)
    total_frk_used = round(sum(e.get("frk_used_qntl", 0) for e in milling_entries), 2)
    total_cmr = round(sum(e.get("cmr_delivery_qntl", 0) for e in milling_entries), 2)
    avg_outturn = round(total_cmr / paddy_used * 100, 2) if paddy_used > 0 else 0

    milling_section = {
        "total_paddy_milled": paddy_used, "total_rice_produced": total_rice,
        "total_frk_used": total_frk_used, "total_cmr_delivered": total_cmr,
        "avg_outturn": avg_outturn, "total_entries": len(milling_entries)
    }

    # ===== 4. FRK STOCK =====
    frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(1000)
    frk_bought = round(sum(p.get("quantity_qntl", 0) for p in frk_purchases), 2)
    frk_cost = round(sum(p.get("total_amount", 0) for p in frk_purchases), 2)

    ob_frk = 0.0
    if saved_ob and "frk_stock" in saved_ob:
        ob_frk = saved_ob.get("frk_stock", 0.0)
    elif prev_fy:
        prev_saved = await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0})
        prev_frk = await db.frk_purchases.find(prev_q, {"_id": 0}).to_list(1000)
        prev_frk_bought = sum(p.get("quantity_qntl", 0) for p in prev_frk)
        prev_mil = await db.milling_entries.find(prev_q, {"_id": 0, "frk_used_qntl": 1}).to_list(1000)
        prev_frk_used = sum(e.get("frk_used_qntl", 0) for e in prev_mil)
        prev_ob_frk = prev_saved.get("frk_stock", 0) if prev_saved else 0
        ob_frk = round(prev_ob_frk + prev_frk_bought - prev_frk_used, 2)

    frk_section = {
        "opening_stock": ob_frk, "purchased": frk_bought, "used": total_frk_used,
        "closing_stock": round(ob_frk + frk_bought - total_frk_used, 2), "total_cost": frk_cost
    }

    # ===== 5. BYPRODUCT STOCK =====
    byproduct_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(1000)
    products = ["bran", "kunda", "broken", "kanki", "husk"]
    byproduct_section = {}
    saved_bp = saved_ob.get("byproducts", {}) if saved_ob else {}

    prev_milling_bp = []
    prev_bp_sales = []
    if not saved_bp and prev_fy:
        prev_milling_bp = await db.milling_entries.find(prev_q, {"_id": 0}).to_list(1000)
        prev_bp_sales = await db.byproduct_sales.find(prev_q, {"_id": 0}).to_list(1000)

    for p in products:
        produced = round(sum(e.get(f"{p}_qntl", 0) for e in milling_entries), 2)
        sold = round(sum(s.get("quantity_qntl", 0) for s in byproduct_sales if s.get("product") == p), 2)
        revenue = round(sum(s.get("total_amount", 0) for s in byproduct_sales if s.get("product") == p), 2)
        ob = 0.0
        if saved_bp:
            ob = saved_bp.get(p, 0.0)
        elif prev_fy:
            prev_saved = await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0})
            prev_bp_ob = prev_saved.get("byproducts", {}).get(p, 0) if prev_saved else 0
            prev_prod = sum(e.get(f"{p}_qntl", 0) for e in prev_milling_bp)
            prev_sold = sum(s.get("quantity_qntl", 0) for s in prev_bp_sales if s.get("product") == p)
            ob = round(prev_bp_ob + prev_prod - prev_sold, 2)
        byproduct_section[p] = {
            "opening_stock": ob, "produced": produced, "sold": sold,
            "closing_stock": round(ob + produced - sold, 2), "revenue": revenue
        }

    # ===== 6. MILL PARTS STOCK =====
    parts = await db.mill_parts.find({}, {"_id": 0}).to_list(1000)
    parts_txns = await db.mill_parts_stock.find(query, {"_id": 0}).to_list(10000)
    saved_mp = saved_ob.get("mill_parts", {}) if saved_ob else {}

    prev_parts_txns = []
    if not saved_mp and prev_fy:
        prev_parts_txns = await db.mill_parts_stock.find(prev_q, {"_id": 0}).to_list(10000)

    parts_section = []
    for part in parts:
        pn = part["name"]
        s_in = sum(t.get("quantity", 0) for t in parts_txns if t.get("part_name") == pn and t.get("txn_type") == "in")
        s_out = sum(t.get("quantity", 0) for t in parts_txns if t.get("part_name") == pn and t.get("txn_type") != "in")
        ob = 0.0
        if saved_mp:
            ob = saved_mp.get(pn, 0.0)
        elif prev_fy:
            prev_saved = await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0})
            prev_mp_ob = prev_saved.get("mill_parts", {}).get(pn, 0) if prev_saved else 0
            p_in = sum(t.get("quantity", 0) for t in prev_parts_txns if t.get("part_name") == pn and t.get("txn_type") == "in")
            p_out = sum(t.get("quantity", 0) for t in prev_parts_txns if t.get("part_name") == pn and t.get("txn_type") != "in")
            ob = round(prev_mp_ob + p_in - p_out, 2)
        parts_section.append({
            "name": pn, "unit": part.get("unit", "Pcs"), "opening_stock": ob,
            "stock_in": round(s_in, 2), "stock_used": round(s_out, 2),
            "closing_stock": round(ob + s_in - s_out, 2)
        })

    # ===== 7. DIESEL ACCOUNTS =====
    diesel_txns = await db.diesel_accounts.find(query, {"_id": 0}).to_list(5000)
    pumps = await db.diesel_pumps.find({}, {"_id": 0}).to_list(100)
    saved_diesel = saved_ob.get("diesel", {}) if saved_ob else {}

    prev_diesel = []
    if not saved_diesel and prev_fy:
        prev_diesel = await db.diesel_accounts.find(prev_q, {"_id": 0}).to_list(5000)

    diesel_section = []
    known_pump_ids = set(pump["id"] for pump in pumps)
    for pump in pumps:
        pid = pump["id"]
        pt_txns = [t for t in diesel_txns if t.get("pump_id") == pid]
        td = sum(t.get("amount", 0) for t in pt_txns if t.get("txn_type") == "debit")
        tp = sum(t.get("amount", 0) for t in pt_txns if t.get("txn_type") == "payment")
        ob = 0.0
        if saved_diesel:
            ob = saved_diesel.get(pid, 0.0)
        elif prev_fy:
            prev_saved = await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0})
            prev_d_ob = prev_saved.get("diesel", {}).get(pid, 0) if prev_saved else 0
            pp = [t for t in prev_diesel if t.get("pump_id") == pid]
            prev_td = sum(t.get("amount", 0) for t in pp if t.get("txn_type") == "debit")
            prev_tp = sum(t.get("amount", 0) for t in pp if t.get("txn_type") == "payment")
            ob = round(prev_d_ob + prev_td - prev_tp, 2)
        diesel_section.append({
            "pump_name": pump["name"], "pump_id": pid, "opening_balance": ob,
            "total_diesel": round(td, 2), "total_paid": round(tp, 2),
            "closing_balance": round(ob + td - tp, 2)
        })
    # Include orphaned diesel accounts (entries with pump_id not in diesel_pumps)
    orphan_pumps = {}
    for dt in diesel_txns:
        pid = dt.get("pump_id") or "default"
        if pid not in known_pump_ids:
            if pid not in orphan_pumps:
                orphan_pumps[pid] = {"pump_name": dt.get("pump_name") or "Default Pump", "debit": 0, "payment": 0}
            if dt.get("txn_type") == "debit":
                orphan_pumps[pid]["debit"] += dt.get("amount", 0)
            elif dt.get("txn_type") == "payment":
                orphan_pumps[pid]["payment"] += dt.get("amount", 0)
    for pid, v in orphan_pumps.items():
        ob = saved_diesel.get(pid, 0.0)
        diesel_section.append({
            "pump_name": v["pump_name"], "pump_id": pid, "opening_balance": ob,
            "total_diesel": round(v["debit"], 2), "total_paid": round(v["payment"], 2),
            "closing_balance": round(ob + v["debit"] - v["payment"], 2)
        })

    # ===== 8. LOCAL PARTY ACCOUNTS =====
    lp_txns = await db.local_party_accounts.find(query, {"_id": 0}).to_list(10000)
    saved_lp = saved_ob.get("local_party", {}) if saved_ob else {}

    prev_lp = []
    if not saved_lp and prev_fy:
        prev_lp = await db.local_party_accounts.find(prev_q, {"_id": 0}).to_list(10000)

    lp_map = {}
    for t in lp_txns:
        pn = (t.get("party_name", "")).strip()
        if not pn: continue
        if pn not in lp_map:
            lp_map[pn] = {"debit": 0, "paid": 0}
        if t.get("txn_type") == "debit": lp_map[pn]["debit"] += t.get("amount", 0)
        elif t.get("txn_type") == "payment": lp_map[pn]["paid"] += t.get("amount", 0)

    if saved_lp:
        all_lp_parties = set(list(lp_map.keys()) + [k for k, v in saved_lp.items() if round(v, 2) != 0])
        lp_total_ob = sum(round(saved_lp.get(p, 0), 2) for p in all_lp_parties)
    else:
        prev_lp_map = {}
        for t in prev_lp:
            pn = (t.get("party_name", "")).strip()
            if not pn: continue
            if pn not in prev_lp_map: prev_lp_map[pn] = 0
            if t.get("txn_type") == "debit": prev_lp_map[pn] += t.get("amount", 0)
            elif t.get("txn_type") == "payment": prev_lp_map[pn] -= t.get("amount", 0)
        all_lp_parties = set(list(lp_map.keys()) + [k for k, v in prev_lp_map.items() if round(v, 2) != 0])
        lp_total_ob = sum(round(prev_lp_map.get(p, 0), 2) for p in all_lp_parties)

    lp_total_debit = sum(lp_map.get(p, {}).get("debit", 0) for p in all_lp_parties)
    lp_total_paid = sum(lp_map.get(p, {}).get("paid", 0) for p in all_lp_parties)
    local_party_section = {
        "party_count": len(all_lp_parties),
        "opening_balance": round(lp_total_ob, 2),
        "total_debit": round(lp_total_debit, 2), "total_paid": round(lp_total_paid, 2),
        "closing_balance": round(lp_total_ob + lp_total_debit - lp_total_paid, 2)
    }

    # ===== 9. STAFF ADVANCES =====
    staff_list = await db.staff.find({"active": True}, {"_id": 0}).to_list(100)
    all_advances = await db.staff_advance.find(query, {"_id": 0}).to_list(5000)
    all_payments = await db.staff_payments.find(query, {"_id": 0}).to_list(5000)
    saved_staff = saved_ob.get("staff", {}) if saved_ob else {}

    prev_adv = []
    prev_pay = []
    if not saved_staff and prev_fy:
        prev_adv = await db.staff_advance.find(prev_q, {"_id": 0}).to_list(5000)
        prev_pay = await db.staff_payments.find(prev_q, {"_id": 0}).to_list(5000)

    staff_section = []
    for s in staff_list:
        sid = s["id"]
        adv = sum(a.get("amount", 0) for a in all_advances if a.get("staff_id") == sid)
        ded = sum(p.get("advance_deducted", 0) for p in all_payments if p.get("staff_id") == sid)
        ob = 0.0
        if saved_staff:
            ob = saved_staff.get(sid, 0.0)
        elif prev_fy:
            prev_saved = await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0})
            prev_s_ob = prev_saved.get("staff", {}).get(sid, 0) if prev_saved else 0
            p_a = sum(a.get("amount", 0) for a in prev_adv if a.get("staff_id") == sid)
            p_d = sum(p.get("advance_deducted", 0) for p in prev_pay if p.get("staff_id") == sid)
            ob = round(prev_s_ob + p_a - p_d, 2)
        staff_section.append({
            "name": s["name"], "staff_id": sid, "opening_balance": ob,
            "total_advance": round(adv, 2), "total_deducted": round(ded, 2),
            "closing_balance": round(ob + adv - ded, 2)
        })

    # ===== 10. PRIVATE TRADING =====
    priv_paddy = await db.private_paddy.find(query, {"_id": 0}).to_list(5000)
    rice_sales = await db.rice_sales.find(query, {"_id": 0}).to_list(5000)
    pp_total = round(sum(p.get("total_amount", 0) for p in priv_paddy), 2)
    pp_paid = round(sum(p.get("paid_amount", 0) for p in priv_paddy), 2)
    rs_total = round(sum(r.get("total_amount", 0) for r in rice_sales), 2)
    rs_paid = round(sum(r.get("paid_amount", 0) for r in rice_sales), 2)

    private_section = {
        "paddy_purchase_amount": pp_total, "paddy_paid": pp_paid, "paddy_balance": round(pp_total - pp_paid, 2),
        "paddy_qty": round(sum(p.get("quantity_qntl", 0) for p in priv_paddy), 2),
        "rice_sale_amount": rs_total, "rice_received": rs_paid, "rice_balance": round(rs_total - rs_paid, 2),
        "rice_qty": round(sum(r.get("quantity_qntl", 0) for r in rice_sales), 2),
    }

    # ===== 11. LEDGER PARTIES (Cashbook Ledger) =====
    ledger_txns = [t for t in cash_txns if t.get("account") == "ledger"]
    saved_ledger = saved_ob.get("ledger_parties", {}) if saved_ob else {}

    ledger_map = {}
    for t in ledger_txns:
        cat = (t.get("category", "")).strip()
        if not cat: continue
        if cat not in ledger_map:
            ledger_map[cat] = {"party_name": cat, "party_type": t.get("party_type", ""), "jama": 0, "nikasi": 0}
        if t.get("txn_type") == "jama":
            ledger_map[cat]["jama"] += t.get("amount", 0)
        else:
            ledger_map[cat]["nikasi"] += t.get("amount", 0)
        if not ledger_map[cat]["party_type"] and t.get("party_type"):
            ledger_map[cat]["party_type"] = t["party_type"]

    # Compute opening balances for ledger parties
    if not saved_ledger and prev_fy:
        prev_ledger_txns = await db.cash_transactions.find(
            {**prev_q, "account": "ledger"}, {"_id": 0, "category": 1, "txn_type": 1, "amount": 1}
        ).to_list(100000)
        prev_saved_lp = (await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0}) or {}).get("ledger_parties", {})
        prev_ledger_map = {}
        for t in prev_ledger_txns:
            cat = (t.get("category", "")).strip()
            if not cat: continue
            if cat not in prev_ledger_map: prev_ledger_map[cat] = 0
            if t.get("txn_type") == "jama": prev_ledger_map[cat] += t.get("amount", 0)
            else: prev_ledger_map[cat] -= t.get("amount", 0)
        saved_ledger = {}
        for cat in set(list(prev_ledger_map.keys()) + list(prev_saved_lp.keys())):
            val = round(prev_saved_lp.get(cat, 0) + prev_ledger_map.get(cat, 0), 2)
            if val != 0:
                saved_ledger[cat] = val

    all_ledger_parties = set(list(ledger_map.keys()) + [k for k, v in saved_ledger.items() if round(v, 2) != 0])
    ledger_section = []
    for cat in sorted(all_ledger_parties):
        info = ledger_map.get(cat, {"party_name": cat, "party_type": "", "jama": 0, "nikasi": 0})
        ob = saved_ledger.get(cat, 0.0)
        ledger_section.append({
            "party_name": cat, "party_type": info.get("party_type", ""),
            "opening_balance": round(ob, 2),
            "total_jama": round(info["jama"], 2), "total_nikasi": round(info["nikasi"], 2),
            "closing_balance": round(ob + info["jama"] - info["nikasi"], 2)
        })

    ledger_summary = {
        "total_parties": len(ledger_section),
        "total_opening": round(sum(l["opening_balance"] for l in ledger_section), 2),
        "total_jama": round(sum(l["total_jama"] for l in ledger_section), 2),
        "total_nikasi": round(sum(l["total_nikasi"] for l in ledger_section), 2),
        "total_closing": round(sum(l["closing_balance"] for l in ledger_section), 2),
        "parties": ledger_section
    }

    return {
        "kms_year": kms_year or "", "season": season or "",
        "cash_bank": cash_section,
        "paddy_stock": paddy_section,
        "milling": milling_section,
        "frk_stock": frk_section,
        "byproducts": byproduct_section,
        "mill_parts": parts_section,
        "diesel": diesel_section,
        "local_party": local_party_section,
        "staff_advances": staff_section,
        "private_trading": private_section,
        "ledger_parties": ledger_summary
    }


@router.get("/fy-summary/balance-sheet")
async def get_balance_sheet(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Tally-style Balance Sheet with Liabilities and Assets"""
    summary = await get_fy_summary(kms_year=kms_year, season=season)
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    cb = summary["cash_bank"]
    ps = summary["paddy_stock"]
    frk = summary["frk_stock"]
    bp = summary["byproducts"]
    mp = summary["mill_parts"]
    diesel_list = summary["diesel"]
    lp = summary["local_party"]
    staff_list = summary["staff_advances"]
    pt = summary["private_trading"]
    ledger = summary["ledger_parties"]

    # ===== TRUCK ACCOUNTS (calculated from entries + truck_payments) =====
    entries = await db.mill_entries.find(query, {"_id": 0, "id": 1, "truck_no": 1, "final_w": 1, "qntl": 1, "bag": 1, "diesel_paid": 1, "cash_paid": 1, "g_deposite": 1, "rate_per_qntl": 1}).to_list(50000)
    truck_map = {}
    for e in entries:
        tn = (e.get("truck_no") or "").strip()
        if not tn: continue
        if tn not in truck_map: truck_map[tn] = {"gross": 0, "deductions": 0, "paid": 0}
        # Get rate from truck_payments or use default
        tp_doc = await db.truck_payments.find_one({"entry_id": e.get("id")}, {"_id": 0, "rate_per_qntl": 1})
        rate = (tp_doc or {}).get("rate_per_qntl") or e.get("rate_per_qntl") or 32
        qntl = round((e.get("qntl") or 0) - (e.get("bag") or 0) / 100, 2)
        truck_map[tn]["gross"] += round(qntl * rate, 2)
        truck_map[tn]["deductions"] += (e.get("diesel_paid") or 0) + (e.get("cash_paid") or 0) + (e.get("g_deposite") or 0)

    all_truck_nos = list(truck_map.keys())
    if all_truck_nos:
        # Get paid from all cash_transactions for these trucks
        all_truck_txns = await db.cash_transactions.find(
            {**query, "txn_type": "nikasi", "category": {"$in": all_truck_nos}}, {"_id": 0}
        ).to_list(50000)
        for tn in all_truck_nos:
            truck_txns = [t for t in all_truck_txns if (t.get("category") or "").strip() == tn]
            # Ledger nikasi (excluding deduction refs)
            ledger_paid = sum(t.get("amount", 0) for t in truck_txns
                if t.get("account") == "ledger"
                and not (t.get("reference") or "").startswith("truck_diesel_ded:")
                and not (t.get("reference") or "").startswith("truck_cash_ded:")
                and not (t.get("reference") or "").startswith("truck_deposit_ded:"))
            # Cash/bank owner payments
            cash_paid = sum(t.get("amount", 0) for t in truck_txns
                if t.get("account") in ("cash", "bank")
                and ((t.get("reference") or "").startswith("truck_owner:") or (t.get("reference") or "").startswith("truck_pay:")))
            truck_map[tn]["paid"] = max(ledger_paid, cash_paid)

    truck_accounts = [{"name": tn, "total": round(v["gross"], 2), "paid": round(v["deductions"] + v["paid"], 2), "balance": round(v["gross"] - v["deductions"] - v["paid"], 2)} for tn, v in sorted(truck_map.items())]
    truck_total_balance = round(sum(t["balance"] for t in truck_accounts), 2)

    # ===== AGENT/MANDI ACCOUNTS (total from targets, paid from cash_transactions) =====
    mandi_map = {}
    mandi_targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(5000)
    for t in mandi_targets:
        mn = (t.get("mandi_name") or "").strip()
        if not mn: continue
        if mn not in mandi_map: mandi_map[mn] = {"total": 0, "paid": 0}
        cutting_qntl = (t.get("target_qntl") or 0) * (t.get("cutting_percent") or 0) / 100
        cr = t.get("cutting_rate")
        mandi_map[mn]["total"] += (t.get("target_qntl") or 0) * (t.get("base_rate") or 10) + cutting_qntl * (cr if cr is not None else 5)
    # Also include mandis from entries without targets
    entries_with_mandi = await db.mill_entries.find(query, {"_id": 0, "mandi_name": 1}).to_list(50000)
    for e in entries_with_mandi:
        mn = (e.get("mandi_name") or "").strip()
        if not mn: continue
        if mn not in mandi_map: mandi_map[mn] = {"total": 0, "paid": 0}
    # Calculate paid from ALL nikasi transactions matching mandi name
    for mn in list(mandi_map.keys()):
        paid_txns = await db.cash_transactions.find(
            {**query, "txn_type": "nikasi",
             "category": {"$regex": f"^{mn}$", "$options": "i"}}, {"_id": 0, "amount": 1, "account": 1, "reference": 1}
        ).to_list(5000)
        ledger_paid = sum(t.get("amount", 0) for t in paid_txns if t.get("account") == "ledger")
        cash_paid = sum(t.get("amount", 0) for t in paid_txns
            if t.get("account") in ("cash", "bank")
            and ((t.get("reference") or "").startswith("agent_pay:") or (t.get("reference") or "").startswith("agent_markpaid:")))
        mandi_map[mn]["paid"] = max(ledger_paid, cash_paid)
    agent_accounts = [{"name": mn, "total": round(v["total"], 2), "paid": round(v["paid"], 2), "balance": round(v["total"] - v["paid"], 2)} for mn, v in sorted(mandi_map.items())]
    agent_total_balance = round(sum(a["balance"] for a in agent_accounts), 2)

    # ===== DC ACCOUNTS =====
    dc_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    dc_entries_list = await db.dc_entries.find(query, {"_id": 0}).to_list(5000)
    dc_map = {}
    for dc in dc_entries_list:
        party = dc.get("party_name") or dc.get("supplier_name") or ""
        if not party: continue
        if party not in dc_map: dc_map[party] = {"total": 0, "paid": 0}
        dc_map[party]["total"] += dc.get("total_amount", 0) or 0
        dc_map[party]["paid"] += dc.get("paid_amount", 0) or 0
    dc_accounts = [{"name": p, "total": round(v["total"], 2), "paid": round(v["paid"], 2), "balance": round(v["total"] - v["paid"], 2)} for p, v in sorted(dc_map.items())]
    dc_total_balance = round(sum(d["balance"] for d in dc_accounts), 2)

    # ===== MSP PAYMENTS (Government) =====
    msp_docs = await db.msp_payments.find(query, {"_id": 0}).to_list(1000)
    msp_total = round(sum(m.get("amount", 0) for m in msp_docs), 2)
    msp_received = round(sum(m.get("received_amount", 0) for m in msp_docs), 2)

    # ===== SEPARATE LEDGER PARTIES INTO DEBTORS (positive balance = we gave more) vs CREDITORS (negative = they gave more) =====
    sundry_debtors = []  # jama > nikasi → they owe us
    sundry_creditors = []  # nikasi > jama → we owe them
    for l in ledger["parties"]:
        bal = l["closing_balance"]
        if bal > 0:
            sundry_debtors.append(l)
        elif bal < 0:
            # Exclude Truck Lease from general creditors (handled separately)
            if l.get("party_type") != "Truck Lease":
                sundry_creditors.append({"party_name": l["party_name"], "party_type": l["party_type"], "amount": abs(bal)})

    # ===== BUILD BALANCE SHEET =====

    # --- LIABILITIES ---
    liabilities = []

    # Capital Account
    capital = cb["opening_cash"] + cb["opening_bank"]
    liabilities.append({"group": "Capital Account", "amount": round(capital, 2), "children": [
        {"name": "Opening Cash", "amount": cb["opening_cash"]},
        {"name": "Opening Bank", "amount": cb["opening_bank"]},
    ]})

    # Sundry Creditors (Ledger parties we owe + Local Party + Diesel)
    creditors_children = []
    creditors_total = 0
    if lp["closing_balance"] > 0:
        creditors_children.append({"name": "Local Party Accounts", "amount": lp["closing_balance"]})
        creditors_total += lp["closing_balance"]
    for sc in sundry_creditors:
        creditors_children.append({"name": sc["party_name"], "amount": sc["amount"]})
        creditors_total += sc["amount"]
    for d in diesel_list:
        if d["closing_balance"] > 0:
            creditors_children.append({"name": f"Diesel - {d['pump_name']}", "amount": d["closing_balance"]})
            creditors_total += d["closing_balance"]
    if pt["paddy_balance"] > 0:
        creditors_children.append({"name": "Pvt Paddy Purchase Payable", "amount": pt["paddy_balance"]})
        creditors_total += pt["paddy_balance"]
    if truck_total_balance > 0:
        for t in truck_accounts:
            if t["balance"] > 0:
                creditors_children.append({"name": f"Truck - {t['name']}", "amount": t["balance"]})
        creditors_total += sum(t["balance"] for t in truck_accounts if t["balance"] > 0)
    if agent_total_balance > 0:
        for a in agent_accounts:
            if a["balance"] > 0:
                creditors_children.append({"name": f"Agent - {a['name']}", "amount": a["balance"]})
        creditors_total += sum(a["balance"] for a in agent_accounts if a["balance"] > 0)
    if dc_total_balance > 0:
        for d in dc_accounts:
            if d["balance"] > 0:
                creditors_children.append({"name": f"DC - {d['name']}", "amount": d["balance"]})
        creditors_total += sum(d["balance"] for d in dc_accounts if d["balance"] > 0)
    # Truck Lease balances
    lease_query = {}
    if kms_year: lease_query["kms_year"] = kms_year
    if season: lease_query["season"] = season
    lease_query["status"] = "active"
    active_leases = await db.truck_leases.find(lease_query, {"_id": 0}).to_list(500)
    lease_total_balance = 0
    for lease in active_leases:
        from routes.truck_lease import get_months_between
        months = get_months_between(lease.get("start_date", ""), lease.get("end_date", ""))
        total_rent = len(months) * lease.get("monthly_rent", 0)
        lease_payments = await db.truck_lease_payments.find({"lease_id": lease["id"]}, {"_id": 0, "amount": 1}).to_list(5000)
        paid = sum(p.get("amount", 0) for p in lease_payments)
        balance = round(total_rent - paid, 2)
        if balance > 0:
            creditors_children.append({"name": f"Truck Lease - {lease['truck_no']}", "amount": balance})
            lease_total_balance += balance
    creditors_total += lease_total_balance
    liabilities.append({"group": "Sundry Creditors", "amount": round(creditors_total, 2), "children": creditors_children})

    total_liabilities = round(sum(l["amount"] for l in liabilities), 2)

    # --- ASSETS ---
    assets = []

    # Cash & Bank
    assets.append({"group": "Cash & Bank Balances", "amount": round(cb["closing_cash"] + cb["closing_bank"], 2), "children": [
        {"name": "Cash-in-Hand", "amount": cb["closing_cash"]},
        {"name": "Bank Accounts", "amount": cb["closing_bank"]},
    ]})

    # Stock-in-Hand
    stock_children = []
    stock_total = 0
    if ps["closing_stock"] > 0:
        stock_children.append({"name": "Paddy Stock", "amount": ps["closing_stock"], "unit": "Qtl"})
        stock_total += ps["closing_stock"]
    if frk["closing_stock"] > 0:
        stock_children.append({"name": "FRK Stock", "amount": frk["closing_stock"], "unit": "Qtl"})
        stock_total += frk["closing_stock"]
    for p_name, v in bp.items():
        if v["closing_stock"] > 0:
            stock_children.append({"name": f"Byproduct - {p_name.capitalize()}", "amount": v["closing_stock"], "unit": "Qtl"})
            stock_total += v["closing_stock"]
    for p in mp:
        if p["closing_stock"] > 0:
            stock_children.append({"name": f"Mill Part - {p['name']}", "amount": p["closing_stock"], "unit": p.get("unit", "Pcs")})
            stock_total += p["closing_stock"]
    assets.append({"group": "Stock-in-Hand", "amount": round(stock_total, 2), "children": stock_children})

    # Sundry Debtors (Ledger parties who owe us + Rice sale receivables + MSP)
    debtors_children = []
    debtors_total = 0
    for sd in sundry_debtors:
        debtors_children.append({"name": sd["party_name"], "amount": sd["closing_balance"]})
        debtors_total += sd["closing_balance"]
    if pt["rice_balance"] > 0:
        debtors_children.append({"name": "Rice Sale Receivable", "amount": pt["rice_balance"]})
        debtors_total += pt["rice_balance"]
    if msp_total - msp_received > 0:
        debtors_children.append({"name": "MSP Receivable", "amount": round(msp_total - msp_received, 2)})
        debtors_total += msp_total - msp_received
    if lp["closing_balance"] < 0:
        debtors_children.append({"name": "Local Party Advance", "amount": abs(lp["closing_balance"])})
        debtors_total += abs(lp["closing_balance"])
    assets.append({"group": "Sundry Debtors", "amount": round(debtors_total, 2), "children": debtors_children})

    # Staff Advances
    staff_total = round(sum(s["closing_balance"] for s in staff_list if s["closing_balance"] > 0), 2)
    staff_children = [{"name": s["name"], "amount": s["closing_balance"]} for s in staff_list if s["closing_balance"] > 0]
    if staff_total > 0:
        assets.append({"group": "Loans & Advances", "amount": staff_total, "children": staff_children})

    total_assets = round(sum(a["amount"] for a in assets), 2)

    # Profit & Loss (difference to balance)
    diff = round(total_assets - total_liabilities, 2)
    if diff > 0:
        liabilities.append({"group": "Profit & Loss A/c (Surplus)", "amount": diff, "children": []})
        total_liabilities = round(total_liabilities + diff, 2)
    elif diff < 0:
        assets.append({"group": "Profit & Loss A/c (Deficit)", "amount": abs(diff), "children": []})
        total_assets = round(total_assets + abs(diff), 2)

    return {
        "kms_year": kms_year or "", "season": season or "",
        "as_on_date": datetime.now().strftime("%d-%m-%Y"),
        "liabilities": liabilities, "total_liabilities": total_liabilities,
        "assets": assets, "total_assets": total_assets,
        "truck_accounts": truck_accounts, "agent_accounts": agent_accounts, "dc_accounts": dc_accounts
    }


@router.get("/fy-summary/balance-sheet/pdf")
async def export_balance_sheet_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO

    data = await get_balance_sheet(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()

    def fmt(n): return f"{(n or 0):,.2f}"

    elements.append(Paragraph(f"<b>Balance Sheet</b> - KMS {kms_year or 'All'}", styles['Title']))
    elements.append(Paragraph(f"As on: {data['as_on_date']}", styles['Normal']))
    elements.append(Spacer(1, 10))

    # Build side data
    def build_side_rows(groups, total_val):
        rows = []
        for g in groups:
            rows.append([g['group'], fmt(g['amount'])])
            for c in g.get('children', []):
                rows.append([f"    {c['name']}", fmt(c['amount'])])
        rows.append(['TOTAL', fmt(total_val)])
        return rows

    liab_rows = build_side_rows(data['liabilities'], data['total_liabilities'])
    asset_rows = build_side_rows(data['assets'], data['total_assets'])

    # Pad shorter side with empty rows
    max_rows = max(len(liab_rows), len(asset_rows))
    while len(liab_rows) < max_rows: liab_rows.append(['', ''])
    while len(asset_rows) < max_rows: asset_rows.append(['', ''])

    # Build side-by-side table: [Liab Particulars | Liab Amount | | Asset Particulars | Asset Amount]
    header = ['LIABILITIES', 'Amount (Rs.)', '', 'ASSETS', 'Amount (Rs.)']
    combined = [header]
    for i in range(max_rows):
        combined.append([liab_rows[i][0], liab_rows[i][1], '', asset_rows[i][0], asset_rows[i][1]])

    col_w = [210, 90, 10, 210, 90]
    t = RLTable(combined, colWidths=col_w, repeatRows=1)

    style_list = [
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#dc2626')),
        ('BACKGROUND', (3, 0), (4, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('GRID', (0, 0), (1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('GRID', (3, 0), (4, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]

    # Style group headers (bold, grey bg) and totals
    for i in range(1, len(combined)):
        row = combined[i]
        # Liability side
        if row[0] and not row[0].startswith('    '):
            style_list.append(('FONTNAME', (0, i), (1, i), 'Helvetica-Bold'))
            if row[0] == 'TOTAL':
                style_list.append(('BACKGROUND', (0, i), (1, i), colors.HexColor('#fecaca')))
            else:
                style_list.append(('BACKGROUND', (0, i), (1, i), colors.HexColor('#f1f5f9')))
        # Asset side
        if row[3] and not row[3].startswith('    '):
            style_list.append(('FONTNAME', (3, i), (4, i), 'Helvetica-Bold'))
            if row[3] == 'TOTAL':
                style_list.append(('BACKGROUND', (3, i), (4, i), colors.HexColor('#a7f3d0')))
            else:
                style_list.append(('BACKGROUND', (3, i), (4, i), colors.HexColor('#f1f5f9')))

    t.setStyle(TableStyle(style_list))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    fname = f"Balance_Sheet_{kms_year or 'all'}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/fy-summary/balance-sheet/excel")
async def export_balance_sheet_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    data = await get_balance_sheet(kms_year=kms_year, season=season)
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 18

    hdr_font = Font(name='Calibri', bold=True, size=14, color='1a365d')
    liab_hdr = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    asset_hdr = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    group_font = Font(name='Calibri', bold=True, size=10)
    child_font = Font(name='Calibri', size=9)
    total_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    liab_fill = PatternFill(start_color='dc2626', end_color='dc2626', fill_type='solid')
    asset_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    liab_total_fill = PatternFill(start_color='fecaca', end_color='fecaca', fill_type='solid')
    asset_total_fill = PatternFill(start_color='a7f3d0', end_color='a7f3d0', fill_type='solid')
    group_fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
    num_fmt = '#,##0.00'
    right_align = Alignment(horizontal='right')

    # Title
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value=f"Balance Sheet - KMS {kms_year or 'All'}").font = hdr_font
    row += 1
    ws.cell(row=row, column=1, value=f"As on: {data['as_on_date']}").font = Font(size=10, color='666666')
    row += 2

    # Headers
    ws.cell(row=row, column=1, value='LIABILITIES').font = liab_hdr
    ws.cell(row=row, column=2, value='Amount (Rs.)').font = liab_hdr
    ws.cell(row=row, column=2).alignment = right_align
    ws.cell(row=row, column=4, value='ASSETS').font = asset_hdr
    ws.cell(row=row, column=5, value='Amount (Rs.)').font = asset_hdr
    ws.cell(row=row, column=5).alignment = right_align
    for c in [1, 2]: ws.cell(row=row, column=c).fill = liab_fill
    for c in [4, 5]: ws.cell(row=row, column=c).fill = asset_fill
    row += 1

    # Build side rows
    def build_rows(groups, total_val):
        rows = []
        for g in groups:
            rows.append(('group', g['group'], g['amount']))
            for c in g.get('children', []):
                rows.append(('child', f"    {c['name']}", c['amount']))
        rows.append(('total', 'TOTAL', total_val))
        return rows

    liab_rows = build_rows(data['liabilities'], data['total_liabilities'])
    asset_rows = build_rows(data['assets'], data['total_assets'])
    max_rows = max(len(liab_rows), len(asset_rows))

    for i in range(max_rows):
        # Liabilities side
        if i < len(liab_rows):
            rtype, name, amt = liab_rows[i]
            ws.cell(row=row, column=1, value=name).font = group_font if rtype in ('group', 'total') else child_font
            ws.cell(row=row, column=2, value=amt).font = group_font if rtype in ('group', 'total') else child_font
            ws.cell(row=row, column=2).number_format = num_fmt
            ws.cell(row=row, column=2).alignment = right_align
            if rtype == 'group':
                ws.cell(row=row, column=1).fill = group_fill
                ws.cell(row=row, column=2).fill = group_fill
            elif rtype == 'total':
                ws.cell(row=row, column=1).fill = liab_total_fill
                ws.cell(row=row, column=2).fill = liab_total_fill
                ws.cell(row=row, column=1).font = Font(bold=True, size=11, color='dc2626')
                ws.cell(row=row, column=2).font = Font(bold=True, size=11, color='dc2626')

        # Assets side
        if i < len(asset_rows):
            rtype, name, amt = asset_rows[i]
            ws.cell(row=row, column=4, value=name).font = group_font if rtype in ('group', 'total') else child_font
            ws.cell(row=row, column=5, value=amt).font = group_font if rtype in ('group', 'total') else child_font
            ws.cell(row=row, column=5).number_format = num_fmt
            ws.cell(row=row, column=5).alignment = right_align
            if rtype == 'group':
                ws.cell(row=row, column=4).fill = group_fill
                ws.cell(row=row, column=5).fill = group_fill
            elif rtype == 'total':
                ws.cell(row=row, column=4).fill = asset_total_fill
                ws.cell(row=row, column=5).fill = asset_total_fill
                ws.cell(row=row, column=4).font = Font(bold=True, size=11, color='059669')
                ws.cell(row=row, column=5).font = Font(bold=True, size=11, color='059669')
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fname = f"Balance_Sheet_{kms_year or 'all'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.post("/fy-summary/carry-forward")
async def carry_forward_fy(data: dict):
    """Close current FY and carry forward all closing balances as opening balances for next FY"""
    kms_year = data.get("kms_year")
    if not kms_year:
        raise HTTPException(status_code=400, detail="kms_year is required")

    next_fy = get_next_fy(kms_year)
    if not next_fy:
        raise HTTPException(status_code=400, detail="Invalid kms_year format")

    # Get current FY summary to extract all closing balances
    summary = await get_fy_summary(kms_year=kms_year)

    cb = summary["cash_bank"]
    ps = summary["paddy_stock"]
    frk = summary["frk_stock"]
    bp = summary["byproducts"]
    mp = summary["mill_parts"]
    diesel = summary["diesel"]
    lp = summary["local_party"]
    staff = summary["staff_advances"]
    pt = summary["private_trading"]
    ledger = summary["ledger_parties"]

    # Build opening balance document for next FY
    ob_doc = {
        "kms_year": next_fy,
        "cash": cb["closing_cash"],
        "bank": cb["closing_bank"],
        "bank_details": {},
        "paddy_stock": ps["closing_stock"],
        "frk_stock": frk["closing_stock"],
        "byproducts": {p: v["closing_stock"] for p, v in bp.items()},
        "mill_parts": {p["name"]: p["closing_stock"] for p in mp},
        "diesel": {d["pump_id"]: d["closing_balance"] for d in diesel},
        "local_party": {},
        "staff": {s["staff_id"]: s["closing_balance"] for s in staff if s["closing_balance"] != 0},
        "ledger_parties": {l["party_name"]: l["closing_balance"] for l in ledger["parties"] if l["closing_balance"] != 0},
        "private_trading": {
            "paddy_balance": pt["paddy_balance"],
            "rice_balance": pt["rice_balance"]
        },
        "carried_from": kms_year,
        "updated_at": datetime.utcnow().isoformat()
    }

    # Save local party balances per party
    lp_txns = await db.local_party_accounts.find({"kms_year": kms_year}, {"_id": 0}).to_list(10000)
    lp_party_map = {}
    for t in lp_txns:
        pn = (t.get("party_name", "")).strip()
        if not pn: continue
        if pn not in lp_party_map: lp_party_map[pn] = 0
        if t.get("txn_type") == "debit": lp_party_map[pn] += t.get("amount", 0)
        elif t.get("txn_type") == "payment": lp_party_map[pn] -= t.get("amount", 0)
    # Add prev OB
    saved_ob = await db.opening_balances.find_one({"kms_year": kms_year}, {"_id": 0})
    if saved_ob and saved_ob.get("local_party"):
        for pn, val in saved_ob["local_party"].items():
            lp_party_map[pn] = lp_party_map.get(pn, 0) + val
    ob_doc["local_party"] = {k: round(v, 2) for k, v in lp_party_map.items() if round(v, 2) != 0}

    # Also get bank details from current summary
    try:
        from routes.cashbook import get_cash_book_summary
        cb_summary = await get_cash_book_summary(kms_year=kms_year)
        bank_details = cb_summary.get("bank_details", {})
        ob_doc["bank_details"] = {bn: round(bd.get("balance", 0), 2) for bn, bd in bank_details.items()}
    except Exception:
        pass

    await db.opening_balances.update_one({"kms_year": next_fy}, {"$set": ob_doc}, upsert=True)

    return {
        "message": f"Closing balances of {kms_year} carried forward to {next_fy}",
        "next_fy": next_fy,
        "opening_balances": ob_doc
    }


@router.get("/fy-summary/pdf")
async def export_fy_summary_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO

    data = await get_fy_summary(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=25, rightMargin=25, topMargin=25, bottomMargin=25)
    elements = []
    styles = getSampleStyleSheet()

    hdr_bg = colors.HexColor('#1a365d')
    hdr_text = colors.white
    total_bg = colors.HexColor('#e0f2fe')

    def fmt(n):
        return f"{(n or 0):,.2f}"

    def section_table(title, headers, rows, col_widths):
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(title, styles['Heading3']))
        elements.append(Spacer(1, 4))
        all_data = [headers] + rows
        t = RLTable(all_data, colWidths=col_widths, repeatRows=1)
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), hdr_text),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        for i in range(1, len(all_data)):
            if i % 2 == 0:
                style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fafc')))
        if len(rows) > 1:
            style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
            style.append(('BACKGROUND', (0, -1), (-1, -1), total_bg))
        t.setStyle(TableStyle(style))
        elements.append(t)

    # Title
    title_text = "FY Summary - Balance Sheet"
    if kms_year: title_text += f" | KMS {kms_year}"
    if season: title_text += f" | {season}"
    elements.append(Paragraph(title_text, styles['Title']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 6))

    # 1. Cash & Bank
    cb = data.get("cash_bank", {})
    section_table("1. Cash & Bank (Rs.)",
        ['Account', 'Opening', 'Inflow', 'Outflow', 'Closing'],
        [
            ['Cash', fmt(cb.get('opening_cash')), fmt(cb.get('cash_in')), fmt(cb.get('cash_out')), fmt(cb.get('closing_cash'))],
            ['Bank', fmt(cb.get('opening_bank')), fmt(cb.get('bank_in')), fmt(cb.get('bank_out')), fmt(cb.get('closing_bank'))],
            ['TOTAL', fmt((cb.get('opening_cash',0)+cb.get('opening_bank',0))),
             fmt((cb.get('cash_in',0)+cb.get('bank_in',0))),
             fmt((cb.get('cash_out',0)+cb.get('bank_out',0))),
             fmt((cb.get('closing_cash',0)+cb.get('closing_bank',0)))],
        ],
        [65, 80, 80, 80, 80])

    # 2. Paddy Stock
    ps = data.get("paddy_stock", {})
    section_table("2. Paddy Stock (Qtl)",
        ['Item', 'Opening', 'In', 'Used', 'Closing'],
        [['Paddy', fmt(ps.get('opening_stock')), fmt(ps.get('paddy_in')), fmt(ps.get('paddy_used')), fmt(ps.get('closing_stock'))]],
        [65, 80, 80, 80, 80])

    # 3. FRK Stock
    frk = data.get("frk_stock", {})
    section_table("3. FRK Stock (Qtl)",
        ['Item', 'Opening', 'Purchased', 'Used', 'Closing', 'Cost (Rs.)'],
        [['FRK', fmt(frk.get('opening_stock')), fmt(frk.get('purchased')), fmt(frk.get('used')), fmt(frk.get('closing_stock')), fmt(frk.get('total_cost'))]],
        [55, 70, 70, 70, 70, 80])

    # 4. Milling Summary
    ml = data.get("milling", {})
    section_table("4. Milling Summary",
        ['Entries', 'Paddy Milled', 'Rice Produced', 'FRK Used', 'CMR Delivered', 'Avg Outturn%'],
        [[str(ml.get('total_entries', 0)), fmt(ml.get('total_paddy_milled')), fmt(ml.get('total_rice_produced')),
          fmt(ml.get('total_frk_used')), fmt(ml.get('total_cmr_delivered')), fmt(ml.get('avg_outturn'))]],
        [55, 75, 75, 70, 80, 70])

    # 5. Byproduct Stock
    bp = data.get("byproducts", {})
    bp_rows = []
    for name, v in bp.items():
        bp_rows.append([name.capitalize(), fmt(v.get('opening_stock')), fmt(v.get('produced')), fmt(v.get('sold')),
                        fmt(v.get('closing_stock')), fmt(v.get('revenue'))])
    if bp_rows:
        section_table("5. Byproduct Stock (Qtl)",
            ['Product', 'Opening', 'Produced', 'Sold', 'Closing', 'Revenue (Rs.)'],
            bp_rows, [60, 65, 65, 65, 65, 80])

    # 6. Mill Parts Stock
    mp = data.get("mill_parts", [])
    if mp:
        mp_rows = [[p['name'], p.get('unit','Pcs'), fmt(p.get('opening_stock')), fmt(p.get('stock_in')),
                     fmt(p.get('stock_used')), fmt(p.get('closing_stock'))] for p in mp]
        section_table("6. Mill Parts Stock",
            ['Part', 'Unit', 'Opening', 'In', 'Used', 'Closing'],
            mp_rows, [80, 40, 65, 65, 65, 65])

    # 7. Diesel Accounts
    diesel = data.get("diesel", [])
    if diesel:
        d_rows = [[d['pump_name'], fmt(d.get('opening_balance')), fmt(d.get('total_diesel')),
                    fmt(d.get('total_paid')), fmt(d.get('closing_balance'))] for d in diesel]
        section_table("7. Diesel Accounts (Rs.)",
            ['Pump', 'Opening', 'Diesel', 'Paid', 'Balance'],
            d_rows, [90, 80, 80, 80, 80])

    # 8. Local Party Accounts
    lp = data.get("local_party", {})
    section_table("8. Local Party Accounts (Rs.)",
        ['Metric', 'Value'],
        [['Total Parties', str(lp.get('party_count', 0))],
         ['Opening Balance', fmt(lp.get('opening_balance'))],
         ['Total Debit', fmt(lp.get('total_debit'))],
         ['Total Paid', fmt(lp.get('total_paid'))],
         ['Closing Balance', fmt(lp.get('closing_balance'))]],
        [120, 120])

    # 9. Staff Advances
    staff = data.get("staff_advances", [])
    if staff:
        s_rows = [[s['name'], fmt(s.get('opening_balance')), fmt(s.get('total_advance')),
                    fmt(s.get('total_deducted')), fmt(s.get('closing_balance'))] for s in staff]
        section_table("9. Staff Advances (Rs.)",
            ['Staff', 'Opening', 'Advance', 'Deducted', 'Balance'],
            s_rows, [90, 80, 80, 80, 80])

    # 10. Ledger Parties
    ledger = data.get("ledger_parties", {})
    lp_list = ledger.get("parties", [])
    if lp_list:
        l_rows = [[l['party_name'], l.get('party_type', ''), fmt(l.get('opening_balance')), fmt(l.get('total_jama')),
                    fmt(l.get('total_nikasi')), fmt(l.get('closing_balance'))] for l in lp_list]
        l_rows.append(['TOTAL', '', fmt(ledger.get('total_opening')), fmt(ledger.get('total_jama')),
                        fmt(ledger.get('total_nikasi')), fmt(ledger.get('total_closing'))])
        section_table("10. Ledger Parties (Rs.)",
            ['Party', 'Type', 'Opening', 'Jama', 'Nikasi', 'Balance'],
            l_rows, [80, 55, 65, 65, 65, 65])

    # 11. Private Trading
    pt = data.get("private_trading", {})
    section_table("11. Private Trading (Rs.)",
        ['Category', 'Qty (Qtl)', 'Amount', 'Paid/Received', 'Balance'],
        [
            ['Paddy Purchase', fmt(pt.get('paddy_qty')), fmt(pt.get('paddy_purchase_amount')), fmt(pt.get('paddy_paid')), fmt(pt.get('paddy_balance'))],
            ['Rice Sales', fmt(pt.get('rice_qty')), fmt(pt.get('rice_sale_amount')), fmt(pt.get('rice_received')), fmt(pt.get('rice_balance'))],
        ],
        [80, 65, 80, 80, 80])

    doc.build(elements)
    buffer.seek(0)
    fname = f"FY_Summary_{kms_year or 'all'}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})
