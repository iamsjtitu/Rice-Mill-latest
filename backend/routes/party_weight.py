"""
v104.44.70 — Party Weight Register
Tracks weight recorded at party's own dharam kaata (independent of our mill scale).
Used for shortage/excess tracking per voucher.

v104.44.93 — Excel/PDF exports + filter params + Auto-adjust to BP sale on save.
"""
import os
import io
import uuid
from datetime import datetime, timezone
from typing import Optional
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

MONGO_URL = os.environ['MONGO_URL']
client = AsyncIOMotorClient(MONGO_URL)
db = client[os.environ['DB_NAME']]
router = APIRouter(tags=["party-weight"])


class PartyWeightEntry(BaseModel):
    id: Optional[str] = None
    product: str
    voucher_no: str
    date: str = ""
    party_name: str = ""
    vehicle_no: str = ""
    rst_no: str = ""
    our_net_weight_kg: float = 0
    party_net_weight_kg: float = 0
    shortage_kg: float = 0
    excess_kg: float = 0
    remark: str = ""
    kms_year: str = ""
    season: str = ""
    created_at: str = ""
    updated_at: str = ""
    created_by: str = ""


def _compute_diff(our_kg: float, party_kg: float) -> dict:
    """Positive diff → shortage (party got less), negative → excess."""
    diff = round(our_kg - party_kg, 2)
    return {
        "shortage_kg": max(0, diff),
        "excess_kg": abs(min(0, diff)),
    }


# ============================================================
# v104.44.93 — Auto-Adjust to BP Sale Bill
# ============================================================

async def _apply_auto_adjust(party_weight_id: str, product: str, voucher_no: str,
                              kms_year: str, party_name: str, date_str: str,
                              shortage_kg: float, excess_kg: float, season: str = "",
                              username: str = "") -> dict:
    """
    Apply weight shortage/excess to the linked BP sale.

    Behavior:
    - Split-billed BP sale (PKA + KCA): reduces kaccha_weight_kg by shortage_kg (or adds excess_kg).
      Recomputes kaccha_amount and total. PKA portion untouched.
    - Solo PKA BP sale: creates a cash_transactions JAMA/NIKASI entry on "{party} (KCA)" virtual
      ledger so party Statement reflects the credit/debit.

    Returns: dict { 'mode': 'split'|'solo_pka'|'skipped', 'amount': float, 'message': str }.
    Idempotent: tagged with reference="party_weight:{id}" for later reversal.
    """
    if shortage_kg <= 0 and excess_kg <= 0:
        return {"mode": "skipped", "amount": 0.0, "message": "No diff"}

    # Only BP sales are auto-adjustable for now (Pvt Rice — TODO)
    bp = await db.bp_sale_register.find_one({"voucher_no": voucher_no, "product": product, "kms_year": kms_year}, {"_id": 0})
    if not bp:
        return {"mode": "skipped", "amount": 0.0, "message": "BP sale not found"}

    # Signed delta: positive shortage means we should reduce billable weight; negative excess means we should add
    delta_kg = -shortage_kg if shortage_kg > 0 else excess_kg  # in kg, negative for shortage

    if bp.get("split_billing"):
        # MODE: split — adjust kaccha portion only
        new_kaccha_kg = max(0.0, float(bp.get("kaccha_weight_kg", 0) or 0) + delta_kg)
        kaccha_rate = float(bp.get("kaccha_rate_per_qtl", 0) or 0) or float(bp.get("rate_per_qtl", 0) or 0)
        new_kaccha_amount = round((new_kaccha_kg / 100.0) * kaccha_rate, 2)
        # Recompute total: pakka_amount + pakka_tax + kaccha_amount
        pakka_amount = float(bp.get("amount", 0) or 0)
        pakka_tax = float(bp.get("tax_amount", 0) or 0)
        new_total = round(pakka_amount + pakka_tax + new_kaccha_amount, 2)
        # Save back the running balance
        new_balance = round(new_total - float(bp.get("advance", 0) or 0), 2)
        await db.bp_sale_register.update_one(
            {"id": bp["id"]},
            {"$set": {
                "kaccha_weight_kg": new_kaccha_kg,
                "kaccha_amount": new_kaccha_amount,
                "total": new_total,
                "balance": new_balance,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "auto_adjust_party_weight_id": party_weight_id,
            }}
        )
        return {
            "mode": "split",
            "amount": round((delta_kg / 100.0) * kaccha_rate, 2),
            "message": f"KCA weight adjusted by {delta_kg:+.2f} Kg → new KCA amt ₹{new_kaccha_amount:,.2f}",
        }

    # MODE: solo PKA — create cash_transactions virtual KCA ledger entry
    pakka_rate = float(bp.get("rate_per_qtl", 0) or 0)
    abs_kg = abs(delta_kg)
    adjustment_amount = round((abs_kg / 100.0) * pakka_rate, 2)
    if adjustment_amount <= 0:
        return {"mode": "skipped", "amount": 0.0, "message": "Zero adjustment amount"}

    # Shortage: party gets credit (we owe party) → JAMA entry on {party} (KCA)
    # Excess:   party owes more (party debit) → NIKASI entry
    txn_type = "jama" if shortage_kg > 0 else "nikasi"
    sign_word = "Shortage" if shortage_kg > 0 else "Excess"
    desc = f"Weight {sign_word} adjustment ({abs_kg:.2f} Kg @ ₹{pakka_rate:,.0f}/Qtl) - Voucher #{voucher_no}"
    txn = {
        "id": str(uuid.uuid4()),
        "date": date_str or bp.get("date", ""),
        "account": "ledger",  # virtual ledger entry
        "txn_type": txn_type,
        "category": f"{party_name} (KCA)" if party_name else "",
        "party_type": "BP Sale",
        "description": desc,
        "amount": adjustment_amount,
        "kms_year": kms_year or bp.get("kms_year", ""),
        "season": season or bp.get("season", ""),
        "reference": f"party_weight:{party_weight_id}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": username or "system",
    }
    await db.cash_transactions.insert_one(txn)
    return {
        "mode": "solo_pka",
        "amount": adjustment_amount if shortage_kg > 0 else -adjustment_amount,
        "message": f"Virtual KCA ledger {txn_type.upper()} ₹{adjustment_amount:,.2f}",
    }


async def _reverse_auto_adjust(party_weight_id: str):
    """Reverse a previously applied auto-adjust (for update/delete flows).

    For solo_pka mode: deletes the cash_transactions entry tagged with reference.
    For split mode: restores kaccha_weight_kg to its original value (best-effort —
      we re-fetch shortage_kg from existing party_weight record before mutation).
    """
    # Drop any cash_transactions entry tied to this party_weight
    await db.cash_transactions.delete_many({"reference": f"party_weight:{party_weight_id}"})
    # Re-find any BP sale that was auto-adjusted by this entry; revert by adding back the delta
    pw = await db.party_weights.find_one({"id": party_weight_id}, {"_id": 0})
    if not pw or not pw.get("auto_adjusted"):
        return
    bp = await db.bp_sale_register.find_one({"auto_adjust_party_weight_id": party_weight_id}, {"_id": 0})
    if not bp:
        return
    old_short = float(pw.get("shortage_kg", 0) or 0)
    old_excess = float(pw.get("excess_kg", 0) or 0)
    revert_delta = old_short - old_excess  # add back what we removed
    new_kaccha_kg = max(0.0, float(bp.get("kaccha_weight_kg", 0) or 0) + revert_delta)
    kaccha_rate = float(bp.get("kaccha_rate_per_qtl", 0) or 0) or float(bp.get("rate_per_qtl", 0) or 0)
    new_kaccha_amount = round((new_kaccha_kg / 100.0) * kaccha_rate, 2)
    pakka_amount = float(bp.get("amount", 0) or 0)
    pakka_tax = float(bp.get("tax_amount", 0) or 0)
    new_total = round(pakka_amount + pakka_tax + new_kaccha_amount, 2)
    new_balance = round(new_total - float(bp.get("advance", 0) or 0), 2)
    await db.bp_sale_register.update_one(
        {"id": bp["id"]},
        {"$set": {"kaccha_weight_kg": new_kaccha_kg, "kaccha_amount": new_kaccha_amount,
                   "total": new_total, "balance": new_balance,
                   "updated_at": datetime.now(timezone.utc).isoformat()},
         "$unset": {"auto_adjust_party_weight_id": ""}}
    )


async def _fetch_sale_info(product: str, voucher_no: str, kms_year: str = "") -> Optional[dict]:
    """Look up sale entry in bp_sale_register or sale_vouchers (for Pvt Rice)."""
    voucher_no = (voucher_no or "").strip()
    if not voucher_no:
        return None
    # Try BP register first
    q = {"voucher_no": voucher_no}
    if product:
        q["product"] = product
    if kms_year:
        q["kms_year"] = kms_year
    s = await db.bp_sale_register.find_one(q, {"_id": 0})
    if s:
        return {
            "voucher_no": s.get("voucher_no", ""),
            "date": s.get("date", ""),
            "party_name": s.get("party_name", ""),
            "vehicle_no": s.get("vehicle_no", ""),
            "rst_no": s.get("rst_no", ""),
            "net_weight_kg": float(s.get("net_weight_kg", 0) or 0),
            "kms_year": s.get("kms_year", ""),
            "season": s.get("season", ""),
            "source": "bp_sale_register",
        }
    # Fallback to sale_vouchers (Pvt Rice / Govt Rice)
    q2 = {"voucher_no": voucher_no}
    if kms_year:
        q2["kms_year"] = kms_year
    sv = await db.sale_vouchers.find_one(q2, {"_id": 0})
    if sv:
        return {
            "voucher_no": sv.get("voucher_no", ""),
            "date": sv.get("date", ""),
            "party_name": sv.get("party_name", ""),
            "vehicle_no": sv.get("vehicle_no", ""),
            "rst_no": sv.get("rst_no", ""),
            "net_weight_kg": float(sv.get("net_weight_kg", 0) or 0),
            "kms_year": sv.get("kms_year", ""),
            "season": sv.get("season", ""),
            "source": "sale_vouchers",
        }
    return None


@router.get("/party-weight/lookup")
async def lookup_voucher(voucher_no: str, product: str = "", kms_year: str = ""):
    """Auto-fetch sale info for a voucher_no (used when user types in form)."""
    info = await _fetch_sale_info(product, voucher_no, kms_year)
    if not info:
        raise HTTPException(status_code=404, detail=f"Voucher #{voucher_no} not found")
    return info


@router.get("/party-weight")
async def list_party_weights(product: str = "", kms_year: str = "", season: str = "",
                              date_from: str = "", date_to: str = "", party_name: str = "",
                              voucher_no: str = "", vehicle_no: str = ""):
    q = {}
    if product: q["product"] = product
    if kms_year: q["kms_year"] = kms_year
    if season: q["season"] = season
    if date_from: q.setdefault("date", {}).update({"$gte": date_from})
    if date_to: q.setdefault("date", {}).update({"$lte": date_to})
    if party_name: q["party_name"] = {"$regex": party_name, "$options": "i"}
    if voucher_no: q["voucher_no"] = {"$regex": voucher_no, "$options": "i"}
    if vehicle_no: q["vehicle_no"] = {"$regex": vehicle_no, "$options": "i"}
    items = await db.party_weights.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return items


@router.post("/party-weight")
async def create_party_weight(data: dict, username: str = "", role: str = ""):
    voucher_no = (data.get("voucher_no") or "").strip()
    product = (data.get("product") or "").strip()
    if not voucher_no:
        raise HTTPException(status_code=400, detail="Voucher No. required")
    if not product:
        raise HTTPException(status_code=400, detail="Product required")

    # Duplicate check per (product + voucher_no + kms_year)
    kms = data.get("kms_year", "")
    dup = await db.party_weights.find_one({"product": product, "voucher_no": voucher_no, "kms_year": kms}, {"_id": 0, "id": 1})
    if dup:
        raise HTTPException(status_code=400, detail=f"Party Weight entry for Voucher #{voucher_no} already exists")

    # Auto-enrich from sale record
    info = await _fetch_sale_info(product, voucher_no, kms)
    our_kg = float(data.get("our_net_weight_kg", 0) or (info and info.get("net_weight_kg", 0)) or 0)
    party_kg = float(data.get("party_net_weight_kg", 0) or 0)
    diff = _compute_diff(our_kg, party_kg)

    auto_adjust = bool(data.get("auto_adjust", True))  # default ON
    pw_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": pw_id,
        "product": product,
        "voucher_no": voucher_no,
        "date": data.get("date") or (info and info.get("date", "")) or "",
        "party_name": data.get("party_name") or (info and info.get("party_name", "")) or "",
        "vehicle_no": data.get("vehicle_no") or (info and info.get("vehicle_no", "")) or "",
        "rst_no": data.get("rst_no") or (info and info.get("rst_no", "")) or "",
        "our_net_weight_kg": our_kg,
        "party_net_weight_kg": party_kg,
        "shortage_kg": diff["shortage_kg"],
        "excess_kg": diff["excess_kg"],
        "remark": data.get("remark", ""),
        "kms_year": kms,
        "season": data.get("season", "") or (info and info.get("season", "")) or "",
        "auto_adjusted": False,
        "adjust_mode": "",
        "adjust_amount": 0.0,
        "created_at": now,
        "updated_at": now,
        "created_by": username,
    }
    await db.party_weights.insert_one(doc)

    # v104.44.93 — Apply auto-adjust if requested and BP sale exists
    if auto_adjust and (diff["shortage_kg"] > 0 or diff["excess_kg"] > 0):
        adj = await _apply_auto_adjust(
            party_weight_id=pw_id, product=product, voucher_no=voucher_no,
            kms_year=kms, party_name=doc["party_name"], date_str=doc["date"],
            shortage_kg=diff["shortage_kg"], excess_kg=diff["excess_kg"],
            season=doc["season"], username=username,
        )
        if adj.get("mode") not in ("skipped", None):
            await db.party_weights.update_one(
                {"id": pw_id},
                {"$set": {
                    "auto_adjusted": True,
                    "adjust_mode": adj.get("mode", ""),
                    "adjust_amount": adj.get("amount", 0.0),
                }}
            )
            doc["auto_adjusted"] = True
            doc["adjust_mode"] = adj.get("mode", "")
            doc["adjust_amount"] = adj.get("amount", 0.0)
            doc["adjust_message"] = adj.get("message", "")

    doc.pop("_id", None)
    return doc


@router.put("/party-weight/{entry_id}")
async def update_party_weight(entry_id: str, data: dict, username: str = "", role: str = ""):
    existing = await db.party_weights.find_one({"id": entry_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")

    # v104.44.93 — Reverse any prior auto-adjust before computing new
    if existing.get("auto_adjusted"):
        await _reverse_auto_adjust(entry_id)

    our_kg = float(data.get("our_net_weight_kg", existing.get("our_net_weight_kg", 0)) or 0)
    party_kg = float(data.get("party_net_weight_kg", existing.get("party_net_weight_kg", 0)) or 0)
    diff = _compute_diff(our_kg, party_kg)
    auto_adjust = bool(data.get("auto_adjust", existing.get("auto_adjusted", True)))

    updates = {
        "our_net_weight_kg": our_kg,
        "party_net_weight_kg": party_kg,
        "shortage_kg": diff["shortage_kg"],
        "excess_kg": diff["excess_kg"],
        "remark": data.get("remark", existing.get("remark", "")),
        "party_name": data.get("party_name", existing.get("party_name", "")),
        "date": data.get("date", existing.get("date", "")),
        "auto_adjusted": False,
        "adjust_mode": "",
        "adjust_amount": 0.0,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.party_weights.update_one({"id": entry_id}, {"$set": updates})

    # Re-apply if requested
    if auto_adjust and (diff["shortage_kg"] > 0 or diff["excess_kg"] > 0):
        adj = await _apply_auto_adjust(
            party_weight_id=entry_id, product=existing.get("product", ""),
            voucher_no=existing.get("voucher_no", ""), kms_year=existing.get("kms_year", ""),
            party_name=updates["party_name"], date_str=updates["date"],
            shortage_kg=diff["shortage_kg"], excess_kg=diff["excess_kg"],
            season=existing.get("season", ""), username=username,
        )
        if adj.get("mode") not in ("skipped", None):
            await db.party_weights.update_one(
                {"id": entry_id},
                {"$set": {"auto_adjusted": True, "adjust_mode": adj.get("mode", ""),
                           "adjust_amount": adj.get("amount", 0.0)}}
            )
            updates["auto_adjusted"] = True
            updates["adjust_mode"] = adj.get("mode", "")
            updates["adjust_amount"] = adj.get("amount", 0.0)
    merged = {**existing, **updates}
    merged.pop("_id", None)
    return merged


@router.delete("/party-weight/{entry_id}")
async def delete_party_weight(entry_id: str, username: str = "", role: str = ""):
    # v104.44.93 — Reverse any auto-adjust before deletion
    existing = await db.party_weights.find_one({"id": entry_id}, {"_id": 0})
    if existing and existing.get("auto_adjusted"):
        await _reverse_auto_adjust(entry_id)
    res = await db.party_weights.delete_one({"id": entry_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"deleted": True}


# ============================================================
# v104.44.93 — Excel & PDF Export
# ============================================================

async def _query_with_filters(product: str = "", kms_year: str = "", season: str = "",
                              date_from: str = "", date_to: str = "", party_name: str = "",
                              voucher_no: str = "", vehicle_no: str = ""):
    q = {}
    if product: q["product"] = product
    if kms_year: q["kms_year"] = kms_year
    if season: q["season"] = season
    if date_from: q.setdefault("date", {}).update({"$gte": date_from})
    if date_to: q.setdefault("date", {}).update({"$lte": date_to})
    if party_name: q["party_name"] = {"$regex": party_name, "$options": "i"}
    if voucher_no: q["voucher_no"] = {"$regex": voucher_no, "$options": "i"}
    if vehicle_no: q["vehicle_no"] = {"$regex": vehicle_no, "$options": "i"}
    items = await db.party_weights.find(q, {"_id": 0}).sort("date", 1).to_list(20000)
    return items


@router.get("/party-weight/export/excel")
async def export_party_weight_excel(product: str = "", kms_year: str = "", season: str = "",
                                     date_from: str = "", date_to: str = "", party_name: str = "",
                                     voucher_no: str = "", vehicle_no: str = ""):
    items = await _query_with_filters(product, kms_year, season, date_from, date_to,
                                       party_name, voucher_no, vehicle_no)
    wb = Workbook()
    ws = wb.active
    ws.title = "Party Weight Register"

    # Theme
    navy = "1E3A8A"; soft = "F8FAFC"; total_bg = "FEF3C7"
    short_bg = "FEE2E2"; excess_bg = "DCFCE7"
    thin = Side(style="thin", color="CBD5E1")
    thick = Side(style="medium", color=navy)
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    title_txt = f"Party Weight Register — {product or 'All Products'}"
    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1, value=title_txt)
    c.font = Font(bold=True, size=16, color=navy)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle / filters
    parts = []
    if kms_year: parts.append(f"KMS: {kms_year}")
    if season: parts.append(f"Season: {season}")
    if date_from or date_to: parts.append(f"Date: {date_from or 'start'} → {date_to or 'today'}")
    if party_name: parts.append(f"Party: {party_name}")
    if voucher_no: parts.append(f"Voucher: {voucher_no}")
    if vehicle_no: parts.append(f"Vehicle: {vehicle_no}")
    if parts:
        ws.merge_cells("A2:J2")
        c = ws.cell(row=2, column=1, value="  •  ".join(parts))
        c.font = Font(size=10, italic=True, color="64748B")
        c.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Date", "Voucher", "Party", "Vehicle", "RST",
               "Our N/W (Kg)", "Party N/W (Kg)", "Shortage (Kg)", "Excess (Kg)", "Remark"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    ws.row_dimensions[4].height = 22

    # Data
    total_short = 0.0
    total_excess = 0.0
    for i, it in enumerate(items, start=5):
        short_kg = float(it.get("shortage_kg", 0) or 0)
        excess_kg = float(it.get("excess_kg", 0) or 0)
        total_short += short_kg
        total_excess += excess_kg
        vals = [
            it.get("date", ""), it.get("voucher_no", ""), it.get("party_name", ""),
            it.get("vehicle_no", ""), it.get("rst_no", ""),
            float(it.get("our_net_weight_kg", 0) or 0),
            float(it.get("party_net_weight_kg", 0) or 0),
            short_kg, excess_kg, it.get("remark", ""),
        ]
        # Background color based on shortage/excess
        if short_kg > 0:
            row_fill = PatternFill(start_color=short_bg, end_color=short_bg, fill_type="solid")
        elif excess_kg > 0:
            row_fill = PatternFill(start_color=excess_bg, end_color=excess_bg, fill_type="solid")
        elif i % 2 == 1:
            row_fill = PatternFill(start_color=soft, end_color=soft, fill_type="solid")
        else:
            row_fill = None
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=ci, value=v)
            cell.font = Font(size=9)
            cell.border = border_thin
            if row_fill:
                cell.fill = row_fill
            if ci >= 6 and ci <= 9:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[i].height = 18

    # Totals row
    tr = len(items) + 5
    total_vals = ["TOTALS", "", "", "", "", "", "", round(total_short, 2), round(total_excess, 2), ""]
    for ci, v in enumerate(total_vals, 1):
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.font = Font(bold=True, size=11, color=navy)
        cell.fill = PatternFill(start_color=total_bg, end_color=total_bg, fill_type="solid")
        cell.border = Border(left=thin, right=thin, top=thick, bottom=thick)
        if ci >= 6:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[tr].height = 22

    # Footer
    ws.cell(row=tr + 2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M')} UTC  •  Records: {len(items)}").font = Font(size=8, italic=True, color="64748B")
    ws.merge_cells(f"A{tr + 2}:J{tr + 2}")

    # Column widths
    widths = [11, 12, 24, 14, 8, 14, 14, 14, 14, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="party_weight_{product or "all"}_{kms_year or "all"}.xlsx"'})


@router.get("/party-weight/export/pdf")
async def export_party_weight_pdf(product: str = "", kms_year: str = "", season: str = "",
                                   date_from: str = "", date_to: str = "", party_name: str = "",
                                   voucher_no: str = "", vehicle_no: str = ""):
    items = await _query_with_filters(product, kms_year, season, date_from, date_to,
                                       party_name, voucher_no, vehicle_no)
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), topMargin=1.2 * cm,
                             bottomMargin=1.2 * cm, leftMargin=1.0 * cm, rightMargin=1.0 * cm)
    elems = []
    styles = getSampleStyleSheet()

    navy = colors.HexColor("#1E3A8A")
    soft = colors.HexColor("#F8FAFC")
    short_bg = colors.HexColor("#FEE2E2")
    excess_bg = colors.HexColor("#DCFCE7")
    total_bg = colors.HexColor("#FEF3C7")
    grey = colors.HexColor("#64748B")

    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=18, textColor=navy,
                                   alignment=1, spaceAfter=4)
    subtitle_style = ParagraphStyle("subtitle", parent=styles["Normal"], fontSize=10,
                                       textColor=grey, alignment=1, spaceAfter=8)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=8,
                                  textColor=navy, alignment=1)

    elems.append(Paragraph(f"<b>Party Weight Register — {product or 'All Products'}</b>", title_style))

    parts = []
    if kms_year: parts.append(f"KMS: {kms_year}")
    if season: parts.append(f"Season: {season}")
    if date_from or date_to: parts.append(f"Date: {date_from or 'start'} → {date_to or 'today'}")
    if party_name: parts.append(f"Party: {party_name}")
    if voucher_no: parts.append(f"Voucher: {voucher_no}")
    if vehicle_no: parts.append(f"Vehicle: {vehicle_no}")
    elems.append(Paragraph("  •  ".join(parts) if parts else "All records", subtitle_style))

    # Stats strip
    total_short = sum(float(it.get("shortage_kg", 0) or 0) for it in items)
    total_excess = sum(float(it.get("excess_kg", 0) or 0) for it in items)
    short_count = sum(1 for it in items if (it.get("shortage_kg") or 0) > 0)
    excess_count = sum(1 for it in items if (it.get("excess_kg") or 0) > 0)
    stats_data = [[
        Paragraph(f"<b>Records</b><br/><font size=12 color='#1E3A8A'>{len(items)}</font>", meta_style),
        Paragraph(f"<b>Shortage Cases</b><br/><font size=12 color='#DC2626'>{short_count}</font>", meta_style),
        Paragraph(f"<b>Excess Cases</b><br/><font size=12 color='#16A34A'>{excess_count}</font>", meta_style),
        Paragraph(f"<b>Total Shortage</b><br/><font size=12 color='#DC2626'>{total_short:,.2f} Kg</font>", meta_style),
        Paragraph(f"<b>Total Excess</b><br/><font size=12 color='#16A34A'>{total_excess:,.2f} Kg</font>", meta_style),
    ]]
    stats_table = Table(stats_data, colWidths=[5.4 * cm] * 5)
    stats_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), soft),
        ("BOX", (0, 0), (-1, -1), 0.8, navy),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, navy),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elems.append(stats_table)
    elems.append(Spacer(1, 0.3 * cm))

    # Main table
    headers = ["Date", "Voucher", "Party", "Vehicle", "RST",
               "Our N/W (Kg)", "Party N/W (Kg)", "Shortage", "Excess", "Remark"]
    table_data = [headers]
    for it in items:
        short_kg = float(it.get("shortage_kg", 0) or 0)
        excess_kg = float(it.get("excess_kg", 0) or 0)
        table_data.append([
            (it.get("date", "") or "")[-5:].replace("-", "/") if it.get("date") else "",
            it.get("voucher_no", "") or "-",
            (it.get("party_name", "") or "")[:24],
            (it.get("vehicle_no", "") or "")[:12],
            it.get("rst_no", "") or "",
            f"{float(it.get('our_net_weight_kg', 0) or 0):,.2f}",
            f"{float(it.get('party_net_weight_kg', 0) or 0):,.2f}",
            f"{short_kg:,.2f}" if short_kg else "—",
            f"{excess_kg:,.2f}" if excess_kg else "—",
            (it.get("remark", "") or "")[:30],
        ])
    table_data.append(["TOTALS", "", "", "", "", "", "",
                       f"{total_short:,.2f}", f"{total_excess:,.2f}", ""])

    col_widths_cm = [1.6, 1.7, 4.6, 2.0, 1.2, 2.4, 2.4, 2.2, 2.2, 5.0]
    t = Table(table_data, colWidths=[w * cm for w in col_widths_cm], repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (5, 1), (-3, -1), "RIGHT"),
        ("ALIGN", (0, 1), (4, -1), "LEFT"),
        ("FONTSIZE", (0, 1), (-1, -2), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, soft]),
        ("BOX", (0, 0), (-1, -1), 0.8, navy),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        # Totals row
        ("BACKGROUND", (0, -1), (-1, -1), total_bg),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 10),
        ("TEXTCOLOR", (0, -1), (-1, -1), navy),
        ("LINEABOVE", (0, -1), (-1, -1), 1.2, navy),
        ("LINEBELOW", (0, -1), (-1, -1), 1.2, navy),
    ])
    # Highlight shortage/excess rows
    for idx, it in enumerate(items, start=1):
        short_kg = float(it.get("shortage_kg", 0) or 0)
        excess_kg = float(it.get("excess_kg", 0) or 0)
        if short_kg > 0:
            style.add("BACKGROUND", (0, idx), (-1, idx), short_bg)
        elif excess_kg > 0:
            style.add("BACKGROUND", (0, idx), (-1, idx), excess_bg)
    t.setStyle(style)
    elems.append(t)

    # Footer
    elems.append(Spacer(1, 0.4 * cm))
    elems.append(Paragraph(f"<i>Generated: {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M')} UTC</i>",
                              ParagraphStyle("footer", parent=styles["Normal"], fontSize=7, textColor=grey, alignment=1)))

    doc.build(elems)
    out.seek(0)
    return StreamingResponse(out, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="party_weight_{product or "all"}_{kms_year or "all"}.pdf"'})
