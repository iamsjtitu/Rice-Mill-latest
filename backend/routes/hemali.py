from models import round_amount
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone
from database import db
from pydantic import BaseModel, Field, ConfigDict
from utils.date_format import fmt_date
import uuid
import io

router = APIRouter()


# ============ MODELS ============

class HemaliItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    rate: float
    unit: str = "bag"
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


# ============ HEMALI ITEMS (Rate Config) ============

@router.get("/hemali/items")
async def get_hemali_items():
    items = await db.hemali_items.find({"is_active": {"$ne": False}}, {"_id": 0}).to_list(500)
    return items


@router.post("/hemali/items")
async def add_hemali_item(request: Request):
    data = await request.json()
    name = (data.get("name") or "").strip()
    rate = data.get("rate")
    if not name or rate is None:
        raise HTTPException(status_code=400, detail="Name aur rate required")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "rate": float(rate),
        "unit": data.get("unit", "bag"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.hemali_items.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/hemali/items/{item_id}")
async def update_hemali_item(item_id: str, request: Request):
    data = await request.json()
    update = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if "name" in data:
        update["name"] = data["name"]
    if "rate" in data:
        update["rate"] = float(data["rate"])
    if "unit" in data:
        update["unit"] = data["unit"]
    result = await db.hemali_items.update_one({"id": item_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    item = await db.hemali_items.find_one({"id": item_id}, {"_id": 0})
    return item


@router.delete("/hemali/items/{item_id}")
async def delete_hemali_item(item_id: str):
    result = await db.hemali_items.update_one({"id": item_id}, {"$set": {"is_active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deactivated"}


# ============ ADVANCE BALANCE ============

async def get_advance_balance(sardar_name: str, kms_year: str = "", season: str = ""):
    query = {"sardar_name": sardar_name, "status": "paid"}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    payments = await db.hemali_payments.find(query, {"_id": 0, "new_advance": 1, "advance_deducted": 1}).to_list(10000)
    advance = 0
    for p in payments:
        advance += (p.get("new_advance") or 0) - (p.get("advance_deducted") or 0)
    return round(advance * 100) / 100


@router.get("/hemali/advance")
async def get_hemali_advance(sardar_name: str = "", kms_year: str = "", season: str = ""):
    if not sardar_name:
        return {"advance": 0}
    advance = await get_advance_balance(sardar_name, kms_year, season)
    return {"advance": advance, "sardar_name": sardar_name}


# ============ HEMALI PAYMENTS ============

@router.get("/hemali/payments")
async def get_hemali_payments(
    kms_year: str = "", season: str = "",
    from_date: str = "", to_date: str = "",
    sardar_name: str = ""
):
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if sardar_name:
        query["sardar_name"] = sardar_name
    if from_date or to_date:
        date_q = {}
        if from_date:
            date_q["$gte"] = from_date
        if to_date:
            date_q["$lte"] = to_date
        query["date"] = date_q

    payments = await db.hemali_payments.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    return payments


@router.post("/hemali/payments")
async def create_hemali_payment(request: Request):
    d = await request.json()
    sardar_name = (d.get("sardar_name") or "").strip()
    if not sardar_name:
        raise HTTPException(status_code=400, detail="Sardar name required")
    items = d.get("items") or []
    if not items:
        raise HTTPException(status_code=400, detail="Items select karein")

    total = round(sum((float(i.get("quantity") or 0)) * (float(i.get("rate") or 0)) for i in items), 2)
    kms_year = d.get("kms_year", "")
    season = d.get("season", "")
    prev_advance = await get_advance_balance(sardar_name, kms_year, season)
    advance_deducted = min(prev_advance, total)
    amount_payable = round(total - advance_deducted, 2)
    amount_paid = float(d.get("amount_paid") or amount_payable)
    new_advance = round(max(0, amount_paid - amount_payable), 2)

    now = datetime.now(timezone.utc).isoformat()
    payment_id = str(uuid.uuid4())
    created_by = d.get("created_by", "")

    # Generate receipt_no: HEM-YYYY-NNNN (sequence per calendar year)
    payment_date = d.get("date", now.split("T")[0])
    receipt_year = str(payment_date)[:4]
    receipt_prefix = f"HEM-{receipt_year}-"
    existing = await db.hemali_payments.find(
        {"receipt_no": {"$regex": f"^{receipt_prefix}"}},
        {"_id": 0, "receipt_no": 1}
    ).to_list(10000)
    max_seq = 0
    for e in existing:
        try:
            n = int(str(e.get("receipt_no", "")).split("-")[-1])
            if n > max_seq:
                max_seq = n
        except (ValueError, IndexError):
            pass
    receipt_no = f"{receipt_prefix}{max_seq + 1:04d}"

    payment = {
        "id": payment_id,
        "receipt_no": receipt_no,
        "sardar_name": sardar_name,
        "date": payment_date,
        "items": [
            {
                "item_name": i.get("item_name", ""),
                "rate": float(i.get("rate") or 0),
                "quantity": float(i.get("quantity") or 0),
                "amount": round_amount(float(i.get("quantity") or 0) * float(i.get("rate") or 0)),
            }
            for i in items
        ],
        "total": total,
        "advance_before": prev_advance,
        "advance_deducted": advance_deducted,
        "amount_payable": amount_payable,
        "amount_paid": amount_paid,
        "new_advance": new_advance,
        "status": "unpaid",
        "kms_year": kms_year,
        "season": season,
        "created_by": created_by,
        "created_at": now,
        "updated_at": now,
    }
    await db.hemali_payments.insert_one(payment)
    payment.pop("_id", None)

    # Create local_party_accounts debit entry on creation (so party is visible)
    items_desc = ", ".join(f"{i['item_name']} x{i['quantity']}" for i in payment["items"])
    await db.local_party_accounts.insert_one({
        "id": str(uuid.uuid4()),
        "date": payment["date"],
        "party_name": "Hemali Payment",
        "txn_type": "debit",
        "amount": total,
        "description": f"{sardar_name} - {items_desc} | Total: Rs.{total:.0f}",
        "reference": f"hemali_debit:{payment_id}",
        "source_type": "hemali",
        "kms_year": kms_year,
        "season": season,
        "created_by": created_by,
        "created_at": now,
    })

    # Also create the Ledger "jama" entry immediately (liability: we owe Sardar the work amount)
    # This makes the unpaid Hemali visible in Cash Book > Ledger right after creation.
    # The matching "nikasi" (payment) entry is added later during mark-paid.
    await db.cash_transactions.insert_one({
        "id": str(uuid.uuid4()),
        "date": payment["date"],
        "account": "ledger",
        "txn_type": "jama",
        "amount": total,
        "category": "Hemali Payment",
        "party_type": "Hemali",
        "description": f"{sardar_name} - {items_desc} | Total: Rs.{total:.0f}",
        "reference": f"hemali_work:{payment_id}",
        "kms_year": kms_year,
        "season": season,
        "created_by": created_by,
        "created_at": now,
        "updated_at": now,
    })

    return payment


async def _create_cash_entries(p, round_off=0):
    """Create cash book + ledger + local party entries for a paid hemali payment."""
    now = datetime.now(timezone.utc).isoformat()
    pid = p["id"]
    sardar = p["sardar_name"]
    items_desc = ", ".join(f"{i['item_name']} x{i['quantity']}" for i in p.get("items", []))
    total_settled = round(p.get("amount_paid", 0) + round_off, 2)
    base = {
        "kms_year": p.get("kms_year", ""), "season": p.get("season", ""),
        "created_by": p.get("created_by", ""), "created_at": now, "updated_at": now,
    }

    # 1. Cash Book: Nikasi (cash going out) - actual cash amount
    await db.cash_transactions.insert_one({
        "id": str(uuid.uuid4()), "date": p["date"], "account": "cash", "txn_type": "nikasi",
        "amount": p["amount_paid"], "category": "Hemali Payment", "party_type": "Hemali",
        "description": f"Hemali: {sardar} - {items_desc}",
        "reference": f"hemali_payment:{pid}", **base,
    })

    # 2. Ledger: Jama (work amount) — shows in Party Ledger tab
    #    UPSERT in case this is a legacy payment that didn't get a work entry on CREATE.
    await db.cash_transactions.update_one(
        {"reference": f"hemali_work:{pid}"},
        {"$set": {
            "date": p["date"], "account": "ledger", "txn_type": "jama",
            "amount": p.get("total", 0), "category": "Hemali Payment", "party_type": "Hemali",
            "description": f"{sardar} - {items_desc} | Total: Rs.{p.get('total',0):.0f}",
            "reference": f"hemali_work:{pid}", **base,
        }, "$setOnInsert": {"id": str(uuid.uuid4())}},
        upsert=True,
    )

    # 3. Ledger: Nikasi (payment) - includes round off for correct balance
    adv_info = ""
    if p.get("advance_deducted", 0) > 0:
        adv_info += f" | Adv Deducted: Rs.{p['advance_deducted']:.0f}"
    if p.get("new_advance", 0) > 0:
        adv_info += f" | New Advance: Rs.{p['new_advance']:.0f}"
    await db.cash_transactions.insert_one({
        "id": str(uuid.uuid4()), "date": p["date"], "account": "ledger", "txn_type": "nikasi",
        "amount": total_settled, "category": "Hemali Payment", "party_type": "Hemali",
        "description": f"{sardar} - Paid Rs.{total_settled:.0f}{adv_info}" + (f" (Cash: {p.get('amount_paid',0):.0f}, RoundOff: {round_off})" if round_off else ""),
        "reference": f"hemali_paid:{pid}", **base,
    })

    # 4. Local Party: payment entry with total (includes round off)
    await db.local_party_accounts.update_one(
        {"reference": f"hemali_debit:{pid}"},
        {"$set": {
            "amount": p.get("total", 0),
            "description": f"{sardar} - {items_desc} | Total: Rs.{p.get('total',0):.0f}",
        }}
    )
    await db.local_party_accounts.insert_one({
        "id": str(uuid.uuid4()), "date": p["date"],
        "party_name": "Hemali Payment", "txn_type": "payment",
        "amount": total_settled,
        "description": f"{sardar} - Paid Rs.{total_settled:.0f}{adv_info}",
        "reference": f"hemali_paid:{pid}", "source_type": "hemali",
        **{k: base[k] for k in ("kms_year", "season", "created_by", "created_at")},
    })


async def _remove_cash_entries(payment_id):
    """Remove cashbook + ledger PAYMENT entries (undo). Keeps the 'hemali_work' jama
    ledger entry because the work was actually done — only the payment is undone."""
    await db.cash_transactions.delete_many({
        "reference": {"$in": [
            f"hemali_payment:{payment_id}",
            f"hemali_paid:{payment_id}",
        ]}
    })
    # Only remove payment entry, keep debit (it was created on payment creation)
    await db.local_party_accounts.delete_many({
        "reference": f"hemali_paid:{payment_id}"
    })


# ============ MARK PAID ============

@router.put("/hemali/payments/{payment_id}/mark-paid")
async def mark_hemali_paid(payment_id: str, request: Request):
    p = await db.hemali_payments.find_one({"id": payment_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Payment not found")
    if p.get("status") == "paid":
        raise HTTPException(status_code=400, detail="Payment already paid")

    d = await request.json() if request.headers.get("content-type") == "application/json" else {}
    amount_paid = float(d.get("amount_paid") or p.get("amount_paid") or p.get("amount_payable", 0))
    new_advance = round(max(0, amount_paid - p.get("amount_payable", 0)), 2)
    round_off = float(d.get("round_off") or 0)

    await db.hemali_payments.update_one(
        {"id": payment_id},
        {"$set": {
            "status": "paid",
            "amount_paid": amount_paid,
            "new_advance": new_advance,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    updated = await db.hemali_payments.find_one({"id": payment_id}, {"_id": 0})
    await _create_cash_entries(updated, round_off=round_off)

    return {"message": "Payment marked as paid", "id": payment_id, "amount_paid": amount_paid, "new_advance": new_advance}


# ============ UNDO PAYMENT ============

@router.put("/hemali/payments/{payment_id}/undo")
async def undo_hemali_payment(payment_id: str):
    p = await db.hemali_payments.find_one({"id": payment_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Payment not found")
    if p.get("status") != "paid":
        raise HTTPException(status_code=400, detail="Payment already unpaid")

    await db.hemali_payments.update_one(
        {"id": payment_id},
        {"$set": {"status": "unpaid", "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    await _remove_cash_entries(payment_id)
    return {"message": "Payment undone", "id": payment_id}


@router.put("/hemali/payments/{payment_id}")
async def update_hemali_payment(payment_id: str, request: Request):
    """Edit an unpaid hemali payment (items, sardar, date etc.)"""
    p = await db.hemali_payments.find_one({"id": payment_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Payment not found")
    if p.get("status") == "paid":
        raise HTTPException(status_code=400, detail="Paid payment edit nahi ho sakti. Pehle undo karein.")

    d = await request.json()
    sardar_name = (d.get("sardar_name") or p.get("sardar_name", "")).strip()
    items = d.get("items") or p.get("items", [])
    date_val = d.get("date") or p.get("date", "")
    kms_year = d.get("kms_year") or p.get("kms_year", "")
    season = d.get("season") or p.get("season", "")

    total = round(sum((float(i.get("quantity") or 0)) * (float(i.get("rate") or 0)) for i in items), 2)
    prev_advance = await get_advance_balance(sardar_name, kms_year, season)
    advance_deducted = min(prev_advance, total)
    amount_payable = round(total - advance_deducted, 2)
    amount_paid = float(d.get("amount_paid") or amount_payable)
    new_advance = round(max(0, amount_paid - amount_payable), 2)

    update_doc = {
        "sardar_name": sardar_name,
        "date": date_val,
        "items": [
            {
                "item_name": i.get("item_name", ""),
                "rate": float(i.get("rate") or 0),
                "quantity": float(i.get("quantity") or 0),
                "amount": round_amount(float(i.get("quantity") or 0) * float(i.get("rate") or 0)),
            }
            for i in items
        ],
        "total": total,
        "advance_before": prev_advance,
        "advance_deducted": advance_deducted,
        "amount_payable": amount_payable,
        "amount_paid": amount_paid,
        "new_advance": new_advance,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.hemali_payments.update_one({"id": payment_id}, {"$set": update_doc})
    updated = await db.hemali_payments.find_one({"id": payment_id}, {"_id": 0})
    return updated


@router.delete("/hemali/payments/{payment_id}")
async def delete_hemali_payment(payment_id: str):
    p = await db.hemali_payments.find_one({"id": payment_id})
    if not p:
        raise HTTPException(status_code=404, detail="Payment not found")
    await _remove_cash_entries(payment_id)
    # Full delete: also remove the work ledger entry + debit
    await db.cash_transactions.delete_many({"reference": f"hemali_work:{payment_id}"})
    await db.local_party_accounts.delete_many({"reference": f"hemali_debit:{payment_id}"})
    await db.hemali_payments.delete_one({"id": payment_id})
    return {"message": "Deleted", "id": payment_id}


# ============ SARDAR LIST ============

@router.get("/hemali/sardars")
async def get_hemali_sardars():
    pipeline = [
        {"$match": {"sardar_name": {"$ne": None}}},
        {"$group": {"_id": "$sardar_name"}},
        {"$sort": {"_id": 1}},
    ]
    results = await db.hemali_payments.aggregate(pipeline).to_list(500)
    return [r["_id"] for r in results if r["_id"]]


# ============ MONTHLY SUMMARY ============

@router.get("/hemali/monthly-summary")
async def hemali_monthly_summary(kms_year: str = "", season: str = "", sardar_name: str = "", month: str = ""):
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if sardar_name:
        query["sardar_name"] = sardar_name
    if month:
        query["date"] = {"$regex": f"^{month}"}

    payments = await db.hemali_payments.find(query, {"_id": 0}).to_list(50000)

    # Group by sardar -> month
    sardars = {}
    for p in payments:
        sn = p.get("sardar_name", "Unknown")
        date_str = p.get("date", "")
        month_key = date_str[:7] if len(date_str) >= 7 else "Unknown"  # YYYY-MM

        if sn not in sardars:
            sardars[sn] = {"sardar_name": sn, "months": {}, "grand_total_work": 0, "grand_total_paid": 0, "grand_total_advance_given": 0, "grand_total_advance_deducted": 0}

        if month_key not in sardars[sn]["months"]:
            sardars[sn]["months"][month_key] = {
                "month": month_key,
                "total_payments": 0, "paid_payments": 0, "unpaid_payments": 0,
                "total_work": 0, "total_paid": 0,
                "advance_given": 0, "advance_deducted": 0,
                "items_breakdown": {},
            }

        m = sardars[sn]["months"][month_key]
        m["total_payments"] += 1
        is_paid = p.get("status") == "paid"

        # Work is done regardless of payment status — count it always
        m["total_work"] += p.get("total", 0)
        sardars[sn]["grand_total_work"] += p.get("total", 0)

        if is_paid:
            m["paid_payments"] += 1
            m["total_paid"] += p.get("amount_paid", 0)
            m["advance_given"] += p.get("new_advance", 0)
            m["advance_deducted"] += p.get("advance_deducted", 0)
            sardars[sn]["grand_total_paid"] += p.get("amount_paid", 0)
            sardars[sn]["grand_total_advance_given"] += p.get("new_advance", 0)
            sardars[sn]["grand_total_advance_deducted"] += p.get("advance_deducted", 0)
        else:
            m["unpaid_payments"] += 1

        # Items breakdown
        for item in p.get("items", []):
            iname = item.get("item_name", "")
            if iname not in m["items_breakdown"]:
                m["items_breakdown"][iname] = {"quantity": 0, "amount": 0}
            m["items_breakdown"][iname]["quantity"] += item.get("quantity", 0)
            m["items_breakdown"][iname]["amount"] += item.get("amount", 0)

    # Convert months dict to sorted list
    result = []
    for sn, data in sorted(sardars.items()):
        months_list = sorted(data["months"].values(), key=lambda x: x["month"], reverse=True)
        for m in months_list:
            m["total_work"] = round(m["total_work"], 2)
            m["total_paid"] = round(m["total_paid"], 2)
            m["advance_given"] = round(m["advance_given"], 2)
            m["advance_deducted"] = round(m["advance_deducted"], 2)
        # Current advance balance
        current_advance = 0
        for p in payments:
            if p.get("sardar_name") == sn and p.get("status") == "paid":
                current_advance += (p.get("new_advance") or 0) - (p.get("advance_deducted") or 0)

        result.append({
            "sardar_name": sn,
            "months": months_list,
            "grand_total_work": round(data["grand_total_work"], 2),
            "grand_total_paid": round(data["grand_total_paid"], 2),
            "grand_total_advance_given": round(data["grand_total_advance_given"], 2),
            "grand_total_advance_deducted": round(data["grand_total_advance_deducted"], 2),
            "current_advance_balance": round(current_advance, 2),
        })

    return result


@router.get("/hemali/monthly-summary/pdf")
async def hemali_monthly_summary_pdf(kms_year: str = "", season: str = "", sardar_name: str = "", month: str = ""):
    data = await hemali_monthly_summary(kms_year=kms_year, season=season, sardar_name=sardar_name, month=month)

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles, get_pdf_table_style
    from reportlab.lib.styles import ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=15, bottomMargin=15)
    styles = get_pdf_styles()
    elements = []

    # Branded header (company name + tagline + custom fields from settings)
    from utils.branding_helper import get_pdf_header_elements_from_db
    subtitle_parts = []
    if kms_year:
        subtitle_parts.append(f"KMS Year: {kms_year}")
    if season:
        subtitle_parts.append(f"Season: {season.title()}")
    if sardar_name:
        subtitle_parts.append(f"Sardar: {sardar_name}")
    subtitle = "  |  ".join(subtitle_parts) if subtitle_parts else "All Sardars"
    elements.extend(await get_pdf_header_elements_from_db("Hemali Monthly Summary", subtitle))
    elements.append(Spacer(1, 6))

    if not data:
        elements.append(Paragraph("Koi data nahi mila is filter ke liye",
                                  ParagraphStyle("nd", parent=styles["Normal"], fontSize=11,
                                                 textColor=colors.HexColor("#94a3b8"), alignment=1)))
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=hemali_monthly_summary.pdf"})

    # Page width calculation: A4 landscape = 842pt, minus 20pt margins each side = 802pt usable.
    # Use this as base width so EVERYTHING (sardar header band, data table, banner) spans full page width.
    PAGE_W = 802

    # Per-sardar tables — accumulate grand totals across all sardars
    grand_work = grand_paid = grand_adv_given = grand_adv_deducted = 0
    grand_payments_total = grand_payments_paid = 0
    for sardar in data:
        # Sardar name pill + advance balance — full page width
        sardar_hdr = RTable(
            [[
                Paragraph(f"<b>SARDAR:</b> {sardar['sardar_name']}",
                          ParagraphStyle("sh", parent=styles["Normal"], fontSize=10,
                                         textColor=colors.white, fontName="Helvetica-Bold")),
                Paragraph(f"Current Advance Balance: <b>Rs. {sardar['current_advance_balance']:,.2f}</b>",
                          ParagraphStyle("sa", parent=styles["Normal"], fontSize=9,
                                         textColor=colors.white, alignment=2)),
            ]],
            colWidths=[PAGE_W * 0.5, PAGE_W * 0.5]
        )
        sardar_hdr.hAlign = 'LEFT'
        sardar_hdr.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#d97706")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(sardar_hdr)
        elements.append(Spacer(1, 2))

        headers = ["Month", "Payments\n(Paid/Total)", "Total Work", "Total Paid", "Adv. Given", "Adv. Deducted"]
        rows = [headers]
        for m in sardar["months"]:
            rows.append([
                m["month"],
                f"{m['paid_payments']}/{m['total_payments']}",
                f"Rs. {m['total_work']:,.2f}",
                f"Rs. {m['total_paid']:,.2f}",
                f"Rs. {m['advance_given']:,.2f}",
                f"Rs. {m['advance_deducted']:,.2f}",
            ])
            grand_payments_total += m.get("total_payments", 0)
            grand_payments_paid += m.get("paid_payments", 0)
        rows.append([
            "TOTAL", "",
            f"Rs. {sardar['grand_total_work']:,.2f}",
            f"Rs. {sardar['grand_total_paid']:,.2f}",
            f"Rs. {sardar['grand_total_advance_given']:,.2f}",
            f"Rs. {sardar['grand_total_advance_deducted']:,.2f}",
        ])
        cols_info = [{"header": h} for h in headers]
        style_cmds = get_pdf_table_style(len(rows), cols_info)
        style_cmds.append(("ALIGN", (1, 0), (-1, -1), "RIGHT"))
        style_cmds.append(("ALIGN", (0, 0), (0, -1), "LEFT"))
        # Distribute 802pt across 6 columns: Month gets ~12%, Payments ~14%, then 18.5% each for 4 amount cols
        col_widths_table = [PAGE_W * 0.12, PAGE_W * 0.14, PAGE_W * 0.185, PAGE_W * 0.185, PAGE_W * 0.185, PAGE_W * 0.185]
        t = RTable(rows, colWidths=col_widths_table, repeatRows=1)
        t.hAlign = 'LEFT'
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 14))

        grand_work += sardar.get("grand_total_work", 0)
        grand_paid += sardar.get("grand_total_paid", 0)
        grand_adv_given += sardar.get("grand_total_advance_given", 0)
        grand_adv_deducted += sardar.get("grand_total_advance_deducted", 0)

    # ===== GRAND SUMMARY BANNER (across all sardars) — light theme, centered =====
    from utils.export_helpers import get_pdf_summary_banner, fmt_inr, STAT_COLORS
    summary_stats = [
        {'label': 'TOTAL SARDARS', 'value': str(len(data)), 'color': STAT_COLORS['primary']},
        {'label': 'PAYMENTS', 'value': f"{grand_payments_paid}/{grand_payments_total}", 'color': STAT_COLORS['blue']},
        {'label': 'GROSS WORK', 'value': fmt_inr(grand_work), 'color': STAT_COLORS['gold']},
        {'label': 'TOTAL PAID', 'value': fmt_inr(grand_paid), 'color': STAT_COLORS['emerald']},
        {'label': 'ADV. GIVEN', 'value': fmt_inr(grand_adv_given), 'color': STAT_COLORS['orange']},
        {'label': 'ADV. DEDUCTED', 'value': fmt_inr(grand_adv_deducted), 'color': STAT_COLORS['purple']},
        {'label': 'OUTSTANDING', 'value': fmt_inr(grand_work - grand_paid - grand_adv_deducted), 'color': STAT_COLORS['red']},
    ]
    elements.append(Spacer(1, 4))
    # Banner narrower than data tables so it's visually centered on the page.
    # (PAGE_W=802 - 80pt margin = 722pt banner, with 40pt margin on each side).
    banner = get_pdf_summary_banner(summary_stats, total_width=PAGE_W - 80)
    if banner:
        elements.append(banner)  # helper sets hAlign='CENTER' by default → page-centered

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=hemali_monthly_summary.pdf"})


@router.get("/hemali/monthly-summary/excel")
async def hemali_monthly_summary_excel(kms_year: str = "", season: str = "", sardar_name: str = "", month: str = ""):
    data = await hemali_monthly_summary(kms_year=kms_year, season=season, sardar_name=sardar_name, month=month)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS)
    from utils.branding_helper import get_excel_branding

    branding = await get_excel_branding()
    ncols = 6

    subtitle_parts = []
    if kms_year:
        subtitle_parts.append(f"KMS Year: {kms_year}")
    if season:
        subtitle_parts.append(f"Season: {season.title()}")
    if sardar_name:
        subtitle_parts.append(f"Sardar: {sardar_name}")
    subtitle = "  |  ".join(subtitle_parts) if subtitle_parts else "All Sardars"
    style_excel_title(ws, "Hemali Monthly Summary / हेमाली मासिक सारांश", ncols,
                      subtitle=subtitle, branding=branding)

    # style_excel_title returns starting row; use a safe row 5 (header rows 1-4 used by branding)
    row_n = ws.max_row + 2 if ws.max_row else 5

    grand_work = grand_paid = grand_adv_given = grand_adv_deducted = 0
    grand_payments_total = grand_payments_paid = 0
    for sardar in data:
        # Sardar header band (orange)
        cell = ws.cell(row=row_n, column=1, value=f"SARDAR: {sardar['sardar_name']}")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="D97706")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        adv_cell = ws.cell(row=row_n, column=5,
                           value=f"Current Advance: Rs. {sardar['current_advance_balance']:,.2f}")
        adv_cell.font = Font(bold=True, size=10, color="FFFFFF")
        adv_cell.fill = PatternFill("solid", fgColor="D97706")
        adv_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=row_n, column=6).fill = PatternFill("solid", fgColor="D97706")
        ws.cell(row=row_n, column=2).fill = PatternFill("solid", fgColor="D97706")
        ws.cell(row=row_n, column=3).fill = PatternFill("solid", fgColor="D97706")
        ws.cell(row=row_n, column=4).fill = PatternFill("solid", fgColor="D97706")
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=4)
        ws.merge_cells(start_row=row_n, start_column=5, end_row=row_n, end_column=6)
        ws.row_dimensions[row_n].height = 22
        row_n += 1

        headers_list = ["Month", "Payments (Paid/Total)", "Total Work", "Total Paid", "Adv. Given", "Adv. Deducted"]
        for ci, h in enumerate(headers_list, 1):
            ws.cell(row=row_n, column=ci, value=h)
        style_excel_header_row(ws, row_n, ncols)
        row_n += 1
        data_start = row_n
        for m in sardar["months"]:
            vals = [m["month"], f"{m['paid_payments']}/{m['total_payments']}", m["total_work"], m["total_paid"], m["advance_given"], m["advance_deducted"]]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=row_n, column=ci, value=v)
            row_n += 1
            grand_payments_total += m.get("total_payments", 0)
            grand_payments_paid += m.get("paid_payments", 0)
        if sardar["months"]:
            style_excel_data_rows(ws, data_start, row_n - 1, ncols, headers_list)
        ws.cell(row=row_n, column=1, value="TOTAL")
        ws.cell(row=row_n, column=3, value=sardar["grand_total_work"])
        ws.cell(row=row_n, column=4, value=sardar["grand_total_paid"])
        ws.cell(row=row_n, column=5, value=sardar["grand_total_advance_given"])
        ws.cell(row=row_n, column=6, value=sardar["grand_total_advance_deducted"])
        style_excel_total_row(ws, row_n, ncols)
        row_n += 2

        grand_work += sardar.get("grand_total_work", 0)
        grand_paid += sardar.get("grand_total_paid", 0)
        grand_adv_given += sardar.get("grand_total_advance_given", 0)
        grand_adv_deducted += sardar.get("grand_total_advance_deducted", 0)

    for w, col_letter in [(14, "A"), (22, "B"), (16, "C"), (16, "D"), (16, "E"), (16, "F")]:
        ws.column_dimensions[col_letter].width = w

    # ===== GRAND SUMMARY BANNER (across all sardars) =====
    if data:
        from utils.export_helpers import add_excel_summary_banner, fmt_inr
        sum_stats = [
            {'label': 'Total Sardars', 'value': str(len(data))},
            {'label': 'Payments', 'value': f"{grand_payments_paid}/{grand_payments_total}"},
            {'label': 'Gross Work', 'value': fmt_inr(grand_work)},
            {'label': 'Total Paid', 'value': fmt_inr(grand_paid)},
            {'label': 'Adv. Given', 'value': fmt_inr(grand_adv_given)},
            {'label': 'Adv. Deducted', 'value': fmt_inr(grand_adv_deducted)},
            {'label': 'Outstanding', 'value': fmt_inr(grand_work - grand_paid - grand_adv_deducted)},
        ]
        add_excel_summary_banner(ws, row_n + 1, ncols, sum_stats)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=hemali_monthly_summary.xlsx"})


# ============ PRINT RECEIPT ============

@router.get("/hemali/payments/{payment_id}/print")
async def print_hemali_receipt(payment_id: str):
    p = await db.hemali_payments.find_one({"id": payment_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Payment not found")

    from reportlab.lib.pagesizes import A5
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer, HRFlowable
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle

    def fmt_d(d):
        if not d: return ''
        parts = str(d).split('-')
        return f"{parts[2]}-{parts[1]}-{parts[0]}" if len(parts) == 3 else d

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A5, leftMargin=25, rightMargin=25, topMargin=20, bottomMargin=20)
    styles = get_pdf_styles()
    elements = []
    orange = colors.HexColor("#d97706")
    dark = colors.HexColor("#1a365d")
    red_c = colors.HexColor("#dc2626")
    green_c = colors.HexColor("#16a34a")
    grey_c = colors.HexColor("#6b7280")

    # ═══════════════════ HEADER SECTION ═══════════════════
    # Company branding
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())

    # Receipt title banner (dark navy background, white text)
    title_tbl = RTable(
        [[Paragraph("HEMALI PAYMENT RECEIPT",
                    ParagraphStyle("tt", parent=styles["Normal"], fontSize=12,
                                   textColor=colors.white, alignment=1,
                                   fontName="Helvetica-Bold"))]],
        colWidths=[349]
    )
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), dark),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(title_tbl)

    # Receipt No. + Status banner (2 columns — amber receipt no. left, status right)
    status_text = "PAID" if p.get("status") == "paid" else "UNPAID"
    status_bg = green_c if p.get("status") == "paid" else red_c
    rcpt_no_str = str(p.get("receipt_no") or "—")

    banner_tbl = RTable([[
        Paragraph(f"<b>Receipt No.</b><br/><font size='12' color='#d97706'><b>{rcpt_no_str}</b></font>",
                  ParagraphStyle("rn", parent=styles["Normal"], fontSize=8, textColor=grey_c, alignment=0)),
        Paragraph(status_text,
                  ParagraphStyle("sb", parent=styles["Normal"], fontSize=13, textColor=colors.white,
                                 alignment=1, fontName="Helvetica-Bold")),
    ]], colWidths=[195, 154])
    banner_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#fef3c7")),
        ("BACKGROUND", (1, 0), (1, 0), status_bg),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(banner_tbl)
    elements.append(Spacer(1, 10))

    # ═══════════════════ INFO GRID (2x2) ═══════════════════
    label_s = ParagraphStyle("lbl", parent=styles["Normal"], fontSize=7, textColor=grey_c,
                             fontName="Helvetica-Bold")
    val_s = ParagraphStyle("val", parent=styles["Normal"], fontSize=10, textColor=dark,
                           fontName="Helvetica-Bold")

    items_count = len(p.get("items", []))
    total_qty = sum(float(i.get("quantity") or 0) for i in p.get("items", []))

    info_data = [
        [Paragraph("RECEIPT DATE", label_s), Paragraph("SARDAR NAME", label_s)],
        [Paragraph(fmt_d(p.get("date", "")), val_s), Paragraph(p.get("sardar_name", "—"), val_s)],
        [Paragraph("ITEMS COUNT", label_s), Paragraph("TOTAL QUANTITY", label_s)],
        [Paragraph(str(items_count), val_s), Paragraph(f"{total_qty:,.0f}", val_s)],
    ]
    info_t = RTable(info_data, colWidths=[174.5, 174.5])
    info_t.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("LINEBELOW", (0, 1), (-1, 1), 0.5, colors.HexColor("#e2e8f0")),
        ("LINEAFTER", (0, 0), (0, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
        ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#f8fafc")),
    ]))
    elements.append(info_t)
    elements.append(Spacer(1, 10))

    # ═══════════════════ ITEMS TABLE ═══════════════════
    items_rows = [[
        Paragraph("<b>ITEM</b>", ParagraphStyle("ih", parent=styles["Normal"], fontSize=8, textColor=colors.white, fontName="Helvetica-Bold")),
        Paragraph("<b>QTY</b>", ParagraphStyle("ih2", parent=styles["Normal"], fontSize=8, textColor=colors.white, alignment=2, fontName="Helvetica-Bold")),
        Paragraph("<b>RATE</b>", ParagraphStyle("ih3", parent=styles["Normal"], fontSize=8, textColor=colors.white, alignment=2, fontName="Helvetica-Bold")),
        Paragraph("<b>AMOUNT</b>", ParagraphStyle("ih4", parent=styles["Normal"], fontSize=8, textColor=colors.white, alignment=2, fontName="Helvetica-Bold")),
    ]]
    for i in p.get("items", []):
        items_rows.append([
            i.get("item_name", ""),
            f"{float(i.get('quantity', 0)):,.0f}",
            f"Rs. {float(i.get('rate', 0)):,.2f}",
            f"Rs. {float(i.get('amount', 0)):,.0f}"
        ])

    it = RTable(items_rows, colWidths=[135, 60, 70, 84], repeatRows=1)
    it.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), dark),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(it)
    elements.append(Spacer(1, 12))

    # ═══════════════════ SUMMARY BOXES (color-coded) ═══════════════════
    def tile(label, value, bg_color, text_color, value_size=11):
        return RTable([[
            Paragraph(label, ParagraphStyle(f"tl_{label}", parent=styles["Normal"], fontSize=7,
                                             textColor=grey_c, alignment=1, fontName="Helvetica-Bold")),
        ], [
            Paragraph(value, ParagraphStyle(f"tv_{label}", parent=styles["Normal"], fontSize=value_size,
                                             textColor=text_color, alignment=1, fontName="Helvetica-Bold")),
        ]], colWidths=[None])

    gross = float(p.get("total") or 0)
    adv_ded = float(p.get("advance_deducted") or 0)
    payable = float(p.get("amount_payable") or 0)
    is_paid = p.get("status") == "paid"
    # For UNPAID payments, amount_paid is just a placeholder (= amount_payable). Display 0 until truly paid.
    paid = float(p.get("amount_paid") or 0) if is_paid else 0.0
    new_adv = float(p.get("new_advance") or 0) if is_paid else 0.0

    def money_tile(label, value, bg, fg, size=11):
        inner = [
            [Paragraph(label, ParagraphStyle(f"ml_{label}", parent=styles["Normal"], fontSize=6,
                                              textColor=grey_c, alignment=1, fontName="Helvetica-Bold"))],
            [Paragraph(value, ParagraphStyle(f"mv_{label}", parent=styles["Normal"], fontSize=size,
                                              textColor=fg, alignment=1, fontName="Helvetica-Bold"))],
        ]
        t = RTable(inner, colWidths=[110])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), bg),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    # Row 1: Gross | Adv Deducted | Net Payable
    row1 = RTable([[
        money_tile("GROSS AMOUNT", f"Rs. {gross:,.0f}", colors.HexColor("#eff6ff"), dark),
        money_tile("ADV. DEDUCTED", f"- Rs. {adv_ded:,.0f}" if adv_ded else "—",
                   colors.HexColor("#fef2f2"), red_c if adv_ded else grey_c),
        money_tile("NET PAYABLE", f"Rs. {payable:,.0f}",
                   colors.HexColor("#fef3c7"), orange, 12),
    ]], colWidths=[116.3, 116.3, 116.3])
    row1.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ]))
    elements.append(row1)
    elements.append(Spacer(1, 4))

    # Row 2: Amount Paid | New Advance | Balance (only meaningful for PAID receipts)
    bal = payable - paid
    bal_label = "SETTLED" if (is_paid and bal <= 0) else f"Rs. {bal:,.0f}"
    row2 = RTable([[
        money_tile("AMOUNT PAID", f"Rs. {paid:,.0f}",
                   colors.HexColor("#f0fdf4"), green_c if is_paid else grey_c, 12),
        money_tile("NEW ADVANCE", f"Rs. {new_adv:,.0f}" if new_adv else "—",
                   colors.HexColor("#fefce8"), orange if new_adv else grey_c),
        money_tile("BALANCE", bal_label,
                   colors.HexColor("#f8fafc"),
                   green_c if (is_paid and bal <= 0) else red_c),
    ]], colWidths=[116.3, 116.3, 116.3])
    row2.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ]))
    elements.append(row2)
    elements.append(Spacer(1, 18))

    # ═══════════════════ SIGNATURES ═══════════════════
    sig_rows = [[
        Paragraph("Sardar Signature",
                  ParagraphStyle("sig", parent=styles["Normal"], fontSize=7,
                                 textColor=grey_c, alignment=1)),
        Paragraph("Authorized Signature",
                  ParagraphStyle("sig2", parent=styles["Normal"], fontSize=7,
                                 textColor=grey_c, alignment=1)),
    ]]
    sig_t = RTable(sig_rows, colWidths=[174, 174])
    sig_t.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (0, 0), 0.5, grey_c),
        ("LINEABOVE", (1, 0), (1, 0), 0.5, grey_c),
        ("TOPPADDING", (0, 0), (-1, -1), 20),
    ]))
    elements.append(sig_t)
    elements.append(Spacer(1, 6))

    # Footer
    elements.append(Paragraph("This is a computer generated receipt",
                              ParagraphStyle("ft", parent=styles["Normal"], fontSize=6,
                                             textColor=grey_c, alignment=1)))

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=hemali_receipt_{payment_id[:8]}.pdf"})


# ============ PDF EXPORT ============

@router.get("/hemali/export/pdf")
async def export_hemali_pdf(
    kms_year: str = "", season: str = "",
    from_date: str = "", to_date: str = "",
    sardar_name: str = ""
):
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if sardar_name:
        query["sardar_name"] = sardar_name
    if from_date or to_date:
        date_q = {}
        if from_date:
            date_q["$gte"] = from_date
        if to_date:
            date_q["$lte"] = to_date
        query["date"] = date_q

    payments = await db.hemali_payments.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    def fmt_d(d):
        if not d: return ''
        parts = str(d).split('-')
        return f"{parts[2]}-{parts[1]}-{parts[0]}" if len(parts) == 3 else d

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=15, bottomMargin=15)
    styles = get_pdf_styles()
    elements = []

    from utils.branding_helper import get_pdf_header_elements_from_db
    meta_parts = []
    if kms_year:
        meta_parts.append(f"KMS Year: {kms_year}")
    if season:
        meta_parts.append(f"Season: {season.title()}")
    if from_date or to_date:
        meta_parts.append(f"Period: {from_date or '-'} to {to_date or '-'}")
    if sardar_name:
        meta_parts.append(f"Sardar: {sardar_name}")
    subtitle = "  |  ".join(meta_parts) if meta_parts else "All Records"
    elements.extend(await get_pdf_header_elements_from_db("Hemali Payment Report", subtitle))
    elements.append(Spacer(1, 6))

    if not payments:
        elements.append(Paragraph("Koi payments nahi mile is filter ke liye",
                                  ParagraphStyle("nd", parent=styles["Normal"], fontSize=11,
                                                 textColor=colors.HexColor("#94a3b8"), alignment=1)))
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=hemali_payments.pdf"})

    headers = ["#", "Receipt No.", "Date", "Sardar", "Items", "Total", "Adv Deduct", "Payable", "Paid", "New Adv", "Status"]
    rows = [headers]
    grand_total = grand_paid = grand_payable = grand_adv_ded = grand_new_adv = 0
    paid_count = unpaid_count = 0
    for idx, p in enumerate(payments, 1):
        items_str = ", ".join(f"{i.get('item_name','')} x{i.get('quantity',0)}" for i in p.get("items", []))
        is_paid = p.get("status") == "paid"
        # For UNPAID: paid is just placeholder; show "—"
        paid_disp = f"Rs. {p.get('amount_paid', 0):,.0f}" if is_paid else "—"
        new_adv_disp = f"Rs. {p.get('new_advance', 0):,.0f}" if is_paid and p.get("new_advance", 0) > 0 else "—"
        rows.append([
            str(idx), p.get("receipt_no", "-"), fmt_d(p.get("date", "")), p.get("sardar_name", ""), items_str,
            f"Rs. {p.get('total',0):,.0f}",
            f"Rs. {p.get('advance_deducted',0):,.0f}" if p.get("advance_deducted", 0) > 0 else "—",
            f"Rs. {p.get('amount_payable',0):,.0f}",
            paid_disp, new_adv_disp,
            "PAID" if is_paid else "UNPAID",
        ])
        grand_total += p.get("total", 0)
        grand_payable += p.get("amount_payable", 0)
        grand_adv_ded += p.get("advance_deducted", 0)
        if is_paid:
            grand_paid += p.get("amount_paid", 0)
            grand_new_adv += p.get("new_advance", 0)
            paid_count += 1
        else:
            unpaid_count += 1
    rows.append(["", "", "", "TOTAL", "",
                 f"Rs. {grand_total:,.0f}",
                 f"Rs. {grand_adv_ded:,.0f}" if grand_adv_ded > 0 else "—",
                 f"Rs. {grand_payable:,.0f}",
                 f"Rs. {grand_paid:,.0f}",
                 f"Rs. {grand_new_adv:,.0f}" if grand_new_adv > 0 else "—",
                 ""])

    from utils.export_helpers import get_pdf_table_style

    t = RTable(rows, colWidths=[22, 70, 55, 75, 170, 60, 60, 60, 60, 60, 50], repeatRows=1)
    cols_info = [{'header': h} for h in headers]
    style_cmds = get_pdf_table_style(len(rows), cols_info)
    # Right-align numeric columns
    style_cmds.append(("ALIGN", (5, 0), (-2, -1), "RIGHT"))
    # Center status column
    style_cmds.append(("ALIGN", (-1, 0), (-1, -1), "CENTER"))
    # Color status cells: PAID green, UNPAID red
    for i, p in enumerate(payments, 1):
        if p.get("status") == "paid":
            style_cmds.append(("TEXTCOLOR", (-1, i), (-1, i), colors.HexColor("#16a34a")))
            style_cmds.append(("FONTNAME", (-1, i), (-1, i), "Helvetica-Bold"))
        else:
            style_cmds.append(("TEXTCOLOR", (-1, i), (-1, i), colors.HexColor("#dc2626")))
            style_cmds.append(("FONTNAME", (-1, i), (-1, i), "Helvetica-Bold"))
    style_cmds.extend([
        ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    # ===== Beautiful single-line summary banner =====
    elements.append(Spacer(1, 12))
    outstanding = grand_payable - grand_paid
    from utils.export_helpers import get_pdf_summary_banner, STAT_COLORS
    summary_stats = [
        {'label': 'TOTAL ENTRIES', 'value': str(len(payments)), 'color': STAT_COLORS['primary']},
        {'label': 'PAID', 'value': str(paid_count), 'color': STAT_COLORS['emerald']},
        {'label': 'UNPAID', 'value': str(unpaid_count), 'color': STAT_COLORS['red']},
        {'label': 'GROSS WORK', 'value': f"Rs. {grand_total:,.0f}", 'color': STAT_COLORS['gold']},
        {'label': 'ADV. DEDUCTED', 'value': f"Rs. {grand_adv_ded:,.0f}", 'color': STAT_COLORS['orange']},
        {'label': 'PAYABLE', 'value': f"Rs. {grand_payable:,.0f}", 'color': STAT_COLORS['blue']},
        {'label': 'TOTAL PAID', 'value': f"Rs. {grand_paid:,.0f}", 'color': STAT_COLORS['green']},
        {'label': 'OUTSTANDING', 'value': f"Rs. {outstanding:,.0f}", 'color': STAT_COLORS['purple']},
    ]
    banner = get_pdf_summary_banner(summary_stats, total_width=sum([22, 70, 55, 75, 170, 60, 60, 60, 60, 60, 50]))
    if banner:
        elements.append(banner)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=hemali_payments.pdf"},
    )


# ============ EXCEL EXPORT ============

@router.get("/hemali/export/excel")
async def export_hemali_excel(
    kms_year: str = "", season: str = "",
    from_date: str = "", to_date: str = "",
    sardar_name: str = ""
):
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if sardar_name:
        query["sardar_name"] = sardar_name
    if from_date or to_date:
        date_q = {}
        if from_date:
            date_q["$gte"] = from_date
        if to_date:
            date_q["$lte"] = to_date
        query["date"] = date_q

    payments = await db.hemali_payments.find(query, {"_id": 0}).sort("date", 1).to_list(10000)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Hemali Payments"
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row)
    from utils.branding_helper import get_excel_branding

    branding = await get_excel_branding()
    ncols = 11

    subtitle_parts = []
    if kms_year:
        subtitle_parts.append(f"KMS Year: {kms_year}")
    if season:
        subtitle_parts.append(f"Season: {season.title()}")
    if from_date or to_date:
        subtitle_parts.append(f"Period: {from_date or '-'} to {to_date or '-'}")
    if sardar_name:
        subtitle_parts.append(f"Sardar: {sardar_name}")
    subtitle = "  |  ".join(subtitle_parts) if subtitle_parts else "All Records"
    style_excel_title(ws, "Hemali Payment Report / हेमाली भुगतान रिपोर्ट", ncols,
                      subtitle=subtitle, branding=branding)

    header_row_n = ws.max_row + 2 if ws.max_row else 5
    headers = ["#", "Receipt No.", "Date", "Sardar", "Items", "Total", "Adv Deducted", "Payable", "Paid", "New Advance", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row_n, column=i, value=h)
    style_excel_header_row(ws, header_row_n, ncols)

    data_start = header_row_n + 1
    row_n = data_start
    grand_total = grand_paid = grand_payable = grand_adv_ded = grand_new_adv = 0
    paid_count = unpaid_count = 0
    for idx, p in enumerate(payments, 1):
        items_str = ", ".join(f"{i.get('item_name','')} x{i.get('quantity',0)}" for i in p.get("items", []))
        is_paid = p.get("status") == "paid"
        status_txt = "PAID" if is_paid else "UNPAID"
        # For UNPAID payments, paid/new_advance are placeholders → show 0
        vals = [
            idx, p.get("receipt_no", "-"), fmt_date(p.get("date", "")),
            p.get("sardar_name", ""), items_str,
            p.get("total", 0), p.get("advance_deducted", 0), p.get("amount_payable", 0),
            p.get("amount_paid", 0) if is_paid else 0,
            p.get("new_advance", 0) if is_paid else 0,
            status_txt,
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row_n, column=ci, value=v)
        # Color status cell
        status_cell = ws.cell(row=row_n, column=ncols)
        if is_paid:
            status_cell.font = Font(bold=True, color="FFFFFF")
            status_cell.fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        else:
            status_cell.font = Font(bold=True, color="FFFFFF")
            status_cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        grand_total += p.get("total", 0)
        grand_payable += p.get("amount_payable", 0)
        grand_adv_ded += p.get("advance_deducted", 0)
        if is_paid:
            grand_paid += p.get("amount_paid", 0)
            grand_new_adv += p.get("new_advance", 0)
            paid_count += 1
        else:
            unpaid_count += 1
        row_n += 1

    if payments:
        style_excel_data_rows(ws, data_start, row_n - 1, ncols, headers)
    else:
        ws.cell(row=data_start, column=1, value="Koi payments nahi mile is filter ke liye").font = Font(italic=True, color="94A3B8")
        ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=ncols)

    if payments:
        # Comprehensive totals row with all numeric columns
        ws.cell(row=row_n, column=5, value="TOTAL")
        ws.cell(row=row_n, column=6, value=grand_total)
        ws.cell(row=row_n, column=7, value=grand_adv_ded)
        ws.cell(row=row_n, column=8, value=grand_payable)
        ws.cell(row=row_n, column=9, value=grand_paid)
        ws.cell(row=row_n, column=10, value=grand_new_adv)
        ws.cell(row=row_n, column=11, value=f"{paid_count} Paid / {unpaid_count} Unpaid")
        style_excel_total_row(ws, row_n, ncols)

        # ===== Beautiful single-line summary banner (below totals) =====
        from utils.export_helpers import add_excel_summary_banner
        outstanding = grand_payable - grand_paid
        sum_stats = [
            {'label': 'Total Entries', 'value': str(len(payments))},
            {'label': 'Paid', 'value': str(paid_count)},
            {'label': 'Unpaid', 'value': str(unpaid_count)},
            {'label': 'Gross Work', 'value': f"Rs.{grand_total:,.2f}"},
            {'label': 'Total Paid', 'value': f"Rs.{grand_paid:,.2f}"},
            {'label': 'Outstanding', 'value': f"Rs.{outstanding:,.2f}"},
        ]
        add_excel_summary_banner(ws, row_n + 2, ncols, sum_stats)

    for w, col_letter in [(5, "A"), (14, "B"), (12, "C"), (18, "D"), (40, "E"), (14, "F"), (16, "G"), (14, "H"), (14, "I"), (16, "J"), (10, "K")]:
        ws.column_dimensions[col_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=hemali_payments.xlsx"},
    )
