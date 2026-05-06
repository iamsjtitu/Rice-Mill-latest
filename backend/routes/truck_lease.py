from fastapi import APIRouter, HTTPException
from database import db
from datetime import datetime, timezone
from typing import Optional
from utils.date_format import fmt_date
import uuid

router = APIRouter()

def gen_id(): return str(uuid.uuid4())
def now_iso(): return datetime.now(timezone.utc).isoformat()

def get_months_between(start_date_str, end_date_str=None):
    """Generate list of YYYY-MM months from start to end (or current month)"""
    try:
        start = datetime.strptime(start_date_str[:7], "%Y-%m")
    except:
        return []
    end = datetime.now()
    if end_date_str:
        try: end = datetime.strptime(end_date_str[:7], "%Y-%m")
        except: pass
    months = []
    current = start
    while current <= end:
        months.append(current.strftime("%Y-%m"))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    return months


# v104.44.101 — Compute trips done by a leased truck across ALL collections
# (mill_entries, private_paddy, dc_entries, dc_deliveries, bp_sale_register).
# Only entries whose date falls within the lease's start_date..end_date window
# are counted as "leased period trips".
async def _compute_lease_trips(lease):
    """Return aggregated trips for a lease.

    Returns dict with: trips (list), total_qntl, trip_count, mandi_breakdown,
    first_date, last_date.
    """
    truck_no = (lease.get("truck_no") or "").strip().upper()
    if not truck_no:
        return {"trips": [], "trip_count": 0, "total_qntl": 0,
                "mandi_breakdown": [], "first_date": "", "last_date": ""}

    # v104.44.106 — Case-INSENSITIVE truck_no match (legacy entries may be lowercase)
    import re as _re
    truck_re = {"$regex": f"^{_re.escape(truck_no)}$", "$options": "i"}

    start = lease.get("start_date", "") or ""
    end = lease.get("end_date", "") or "9999-12-31"

    def in_window(date_str):
        if not date_str: return False
        d = date_str[:10]
        return start <= d <= end

    trips = []

    # 1. mill_entries (paddy from mandis)
    async for e in db.mill_entries.find({"truck_no": truck_re}, {"_id": 0}):
        if not in_window(e.get("date", "")): continue
        qntl = float(e.get("qntl", 0) or 0)
        bag = float(e.get("bag", 0) or 0)
        net_qntl = round(qntl - bag/100, 2)
        trips.append({
            "date": e.get("date", ""), "rst_no": e.get("rst_no", ""),
            "source": e.get("mandi_name", "") or e.get("mandi", ""),
            "source_type": "Mill (Mandi)",
            "party": "", "qntl": net_qntl, "bag": int(bag),
        })

    # 2. private_paddy
    async for e in db.private_paddy.find({"truck_no": truck_re}, {"_id": 0}):
        if not in_window(e.get("date", "")): continue
        qntl = float(e.get("qntl", 0) or 0)
        bag = float(e.get("bag", 0) or 0)
        net_qntl = round(qntl - bag/100, 2)
        trips.append({
            "date": e.get("date", ""), "rst_no": e.get("rst_no", ""),
            "source": e.get("party_name", ""),
            "source_type": "Private Paddy",
            "party": e.get("party_name", ""),
            "qntl": net_qntl, "bag": int(bag),
        })

    # 3. dc_entries (incoming DC stock)
    async for e in db.dc_entries.find({"$or": [
        {"truck_no": truck_re}, {"vehicle_no": truck_re}
    ]}, {"_id": 0}):
        if not in_window(e.get("date", "")): continue
        qntl = float(e.get("quantity_qntl", 0) or e.get("qntl", 0) or 0)
        trips.append({
            "date": e.get("date", ""), "rst_no": e.get("dc_no", "") or e.get("rst_no", ""),
            "source": e.get("party_name", "") or e.get("from_party", ""),
            "source_type": "DC In",
            "party": e.get("party_name", ""), "qntl": qntl, "bag": 0,
        })

    # 4. dc_deliveries (outgoing DC dispatches)
    async for e in db.dc_deliveries.find({"$or": [
        {"truck_no": truck_re}, {"vehicle_no": truck_re}
    ]}, {"_id": 0}):
        if not in_window(e.get("date", "")): continue
        qntl = float(e.get("quantity_qntl", 0) or e.get("qntl", 0) or 0)
        trips.append({
            "date": e.get("date", ""), "rst_no": e.get("dc_no", "") or e.get("rst_no", ""),
            "source": e.get("destination", "") or e.get("to_party", ""),
            "source_type": "DC Out",
            "party": e.get("party_name", ""), "qntl": qntl, "bag": 0,
        })

    # 5. bp_sale_register (BP sale dispatch — vehicle_no field)
    async for e in db.bp_sale_register.find({"vehicle_no": truck_re}, {"_id": 0}):
        if not in_window(e.get("date", "")): continue
        qntl = float(e.get("net_weight_qtl", 0) or 0)
        if not qntl:
            qntl = round(float(e.get("net_weight_kg", 0) or 0) / 100, 2)
        trips.append({
            "date": e.get("date", ""), "rst_no": e.get("voucher_no", "") or e.get("rst_no", ""),
            "source": e.get("destination", "") or e.get("party_name", ""),
            "source_type": f"BP Sale - {e.get('product', 'Bran')}",
            "party": e.get("party_name", ""), "qntl": qntl, "bag": int(e.get("bags", 0) or 0),
        })

    # Sort by date ASC
    trips.sort(key=lambda t: t.get("date", ""))

    # Mandi/source breakdown
    breakdown = {}
    for t in trips:
        key = t["source"] or "(unknown)"
        breakdown.setdefault(key, {"source": key, "trips": 0, "qntl": 0})
        breakdown[key]["trips"] += 1
        breakdown[key]["qntl"] += t["qntl"]
    mandi_breakdown = sorted(
        [{"source": v["source"], "trips": v["trips"], "qntl": round(v["qntl"], 2)}
         for v in breakdown.values()],
        key=lambda x: x["qntl"], reverse=True,
    )

    total_qntl = round(sum(t["qntl"] for t in trips), 2)
    return {
        "trips": trips, "trip_count": len(trips), "total_qntl": total_qntl,
        "mandi_breakdown": mandi_breakdown,
        "first_date": trips[0]["date"] if trips else "",
        "last_date": trips[-1]["date"] if trips else "",
    }


# ========== TRUCK LEASES CRUD ==========

@router.get("/truck-leases")
async def get_truck_leases(kms_year: Optional[str] = None, season: Optional[str] = None, status: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if status: query["status"] = status
    leases = await db.truck_leases.find(query, {"_id": 0}).to_list(500)
    return sorted(leases, key=lambda x: x.get("created_at", ""), reverse=True)


@router.post("/truck-leases")
async def create_truck_lease(data: dict):
    lease = {
        "id": gen_id(),
        "truck_no": (data.get("truck_no") or "").strip().upper(),
        "owner_name": (data.get("owner_name") or "").strip(),
        "monthly_rent": float(data.get("monthly_rent") or 0),
        "start_date": data.get("start_date", ""),
        "end_date": data.get("end_date", ""),
        "advance_deposit": float(data.get("advance_deposit") or 0),
        "status": "active",
        "kms_year": data.get("kms_year", ""),
        "season": data.get("season", ""),
        "created_by": data.get("created_by", ""),
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    if not lease["truck_no"]:
        raise HTTPException(status_code=400, detail="Truck number is required")
    if lease["monthly_rent"] <= 0:
        raise HTTPException(status_code=400, detail="Monthly rent must be > 0")
    # Check duplicate active lease for same truck
    existing = await db.truck_leases.find_one({"truck_no": lease["truck_no"], "status": "active"}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail=f"Truck {lease['truck_no']} already has an active lease")
    await db.truck_leases.insert_one(lease)
    lease.pop("_id", None)
    return lease


@router.put("/truck-leases/{lease_id}")
async def update_truck_lease(lease_id: str, data: dict):
    existing = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lease not found")
    updates = {}
    for field in ["truck_no", "owner_name", "monthly_rent", "start_date", "end_date", "advance_deposit", "status"]:
        if field in data:
            if field == "truck_no":
                updates[field] = (data[field] or "").strip().upper()
            elif field in ("monthly_rent", "advance_deposit"):
                updates[field] = float(data[field] or 0)
            else:
                updates[field] = data[field]
    updates["updated_at"] = now_iso()
    await db.truck_leases.update_one({"id": lease_id}, {"$set": updates})
    return {**existing, **updates}


@router.delete("/truck-leases/{lease_id}")
async def delete_truck_lease(lease_id: str):
    result = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Lease not found")
    await db.truck_leases.delete_one({"id": lease_id})
    # Also delete related payment records
    await db.truck_lease_payments.delete_many({"lease_id": lease_id})
    return {"message": "Lease deleted", "id": lease_id}


# v104.44.105 — End an active lease NOW (set end_date = today, status = ended)
@router.post("/truck-leases/{lease_id}/end-now")
async def end_lease_now(lease_id: str):
    existing = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lease not found")
    today = datetime.now().strftime("%Y-%m-%d")
    await db.truck_leases.update_one(
        {"id": lease_id},
        {"$set": {"end_date": today, "status": "ended", "updated_at": now_iso()}},
    )
    return {**existing, "end_date": today, "status": "ended"}


# v104.44.106 — Count + cleanup driver advance entries (cash/diesel given on
# individual trips) for a leased truck. These are NOT rent payments and
# should not appear in the truck's party ledger when truck is leased.
@router.get("/truck-leases/{lease_id}/driver-advances")
async def get_driver_advances(lease_id: str):
    """Returns count + total of OLD truck_cash_de + truck_diesel_de entries
    in cash_transactions for this lease's truck."""
    lease = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not lease:
        raise HTTPException(status_code=404, detail="Lease not found")
    truck_no = (lease.get("truck_no") or "").upper()
    if not truck_no:
        return {"count": 0, "total": 0, "entries": []}
    import re as _re
    query = {
        "category": {"$regex": _re.escape(truck_no), "$options": "i"},
        "reference": {"$regex": "^truck_(cash|diesel)_de"},
    }
    entries = await db.cash_transactions.find(query, {"_id": 0}).to_list(5000)
    total = round(sum(e.get("amount", 0) or 0 for e in entries), 2)
    return {"count": len(entries), "total": total, "entries": entries}


@router.post("/truck-leases/{lease_id}/driver-advances/cleanup")
async def cleanup_driver_advances(lease_id: str):
    """Delete OLD truck_cash_de + truck_diesel_de cash_transactions for this
    leased truck. These get auto-created as a 'Truck OD0...' party ledger from
    mill_entries cash_paid/diesel_paid — for leased trucks they're driver
    expenses (separate from owner rent) and should not clutter the truck's
    ledger view."""
    lease = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not lease:
        raise HTTPException(status_code=404, detail="Lease not found")
    truck_no = (lease.get("truck_no") or "").upper()
    if not truck_no:
        return {"deleted": 0}
    import re as _re
    result = await db.cash_transactions.delete_many({
        "category": {"$regex": _re.escape(truck_no), "$options": "i"},
        "reference": {"$regex": "^truck_(cash|diesel)_de"},
    })
    # Also clean local_party_accounts mirrors if any
    lp_result = await db.local_party_accounts.delete_many({
        "party_name": {"$regex": _re.escape(truck_no), "$options": "i"},
        "reference": {"$regex": "^truck_(cash|diesel)_de"},
    })
    return {"deleted": result.deleted_count, "lp_deleted": lp_result.deleted_count}


# v104.44.105 — Get trips for a specific month (for monthly drilldown)
@router.get("/truck-leases/{lease_id}/trips/by-month/{month}")
async def get_lease_trips_by_month(lease_id: str, month: str):
    """month format: YYYY-MM. Returns trips done in that calendar month
    that fall within the lease window."""
    lease = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not lease:
        raise HTTPException(status_code=404, detail="Lease not found")
    full = await _compute_lease_trips(lease)
    month_trips = [t for t in full["trips"] if (t.get("date") or "").startswith(month)]
    breakdown = {}
    for t in month_trips:
        key = t["source"] or "(unknown)"
        breakdown.setdefault(key, {"source": key, "trips": 0, "qntl": 0})
        breakdown[key]["trips"] += 1
        breakdown[key]["qntl"] += t["qntl"]
    return {
        "month": month,
        "trips": month_trips,
        "trip_count": len(month_trips),
        "total_qntl": round(sum(t["qntl"] for t in month_trips), 2),
        "total_bags": sum(t.get("bag", 0) or 0 for t in month_trips),
        "mandi_breakdown": sorted(
            [{"source": v["source"], "trips": v["trips"], "qntl": round(v["qntl"], 2)}
             for v in breakdown.values()],
            key=lambda x: x["qntl"], reverse=True,
        ),
    }


# ========== LEASE PAYMENT SUMMARY (monthly breakdown) ==========

# v104.44.105 — Find cash_transactions that are payments to this lease's owner/truck
# (when user pays from Cash Book directly without using "Pay Lease" button).
async def _fetch_cashbook_payments_for_lease(lease):
    """Returns cash_transactions matching the lease's truck/owner that should
    count as lease payments. Excludes auto_ledger duplicates and entries already
    linked via lease_pay reference (those are in truck_lease_payments)."""
    truck_no = (lease.get("truck_no") or "").upper()
    owner = (lease.get("owner_name") or "").strip()
    if not truck_no:
        return []
    import re as _re
    or_clauses = []
    # Exact category match
    if owner: or_clauses.append({"category": owner})
    or_clauses.append({"category": f"Truck Lease - {truck_no}"})
    # Case-insensitive truck_no in category
    or_clauses.append({"category": {"$regex": _re.escape(truck_no), "$options": "i"}})
    # Case-insensitive owner in category
    if owner: or_clauses.append({"category": {"$regex": _re.escape(owner), "$options": "i"}})
    # Truck_no mentioned in description (e.g., "Lease payment for OD04K2455")
    or_clauses.append({"description": {"$regex": _re.escape(truck_no), "$options": "i"}})
    query = {
        "$and": [
            {"$or": or_clauses},
            {"txn_type": "nikasi"},
            # Skip auto_ledger mirrors (avoids double-counting)
            {"reference": {"$not": {"$regex": "^auto_ledger:"}}},
            # Skip already-linked lease_pay entries (those are in truck_lease_payments)
            {"linked_payment_id": {"$not": {"$regex": f"^truck_lease:{lease['id']}:"}}},
            # v104.44.106 — Skip driver advance entries (cash/diesel given on trip)
            # — these are NOT lease rent payments and shouldn't count.
            {"reference": {"$not": {"$regex": "^truck_(cash|diesel)_de"}}},
        ]
    }
    return await db.cash_transactions.find(query, {"_id": 0}).to_list(5000)


@router.get("/truck-leases/{lease_id}/payments")
async def get_lease_payments(lease_id: str):
    lease = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not lease:
        raise HTTPException(status_code=404, detail="Lease not found")
    
    months = get_months_between(lease.get("start_date", ""), lease.get("end_date", ""))
    payments = await db.truck_lease_payments.find({"lease_id": lease_id}, {"_id": 0}).to_list(5000)

    # v104.44.105 — Also pull matching Cash Book payments (manual payments from cashbook)
    cb_payments = await _fetch_cashbook_payments_for_lease(lease)

    # Group payments by month
    month_paid = {}
    for p in payments:
        m = p.get("month", "")
        month_paid[m] = month_paid.get(m, 0) + p.get("amount", 0)
    # Cash book payments → infer month from date
    for cb in cb_payments:
        m = (cb.get("date") or "")[:7]
        if m: month_paid[m] = month_paid.get(m, 0) + (cb.get("amount", 0) or 0)
    
    monthly_records = []
    total_rent = 0
    total_paid = 0
    for m in months:
        rent = lease.get("monthly_rent", 0)
        paid = round(month_paid.get(m, 0), 2)
        balance = round(rent - paid, 2)
        status = "paid" if balance <= 0 else ("partial" if paid > 0 else "pending")
        monthly_records.append({"month": m, "rent": rent, "paid": paid, "balance": max(0, balance), "status": status})
        total_rent += rent
        total_paid += paid
    
    return {
        "lease": lease,
        "monthly_records": monthly_records,
        "total_rent": round(total_rent, 2),
        "total_paid": round(total_paid, 2),
        "total_balance": round(max(0, total_rent - total_paid), 2),
        "advance_deposit": lease.get("advance_deposit", 0)
    }


# ========== MAKE PAYMENT ==========

@router.post("/truck-leases/{lease_id}/pay")
async def make_lease_payment(lease_id: str, data: dict):
    lease = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not lease:
        raise HTTPException(status_code=404, detail="Lease not found")
    
    amount = float(data.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be > 0")
    
    month = data.get("month", datetime.now().strftime("%Y-%m"))
    account = data.get("account", "cash")
    bank_name = data.get("bank_name", "")
    payment_date = data.get("payment_date", datetime.now().strftime("%Y-%m-%d"))
    notes = data.get("notes", "")
    
    payment_id = gen_id()
    payment = {
        "id": payment_id,
        "lease_id": lease_id,
        "truck_no": lease["truck_no"],
        "owner_name": lease.get("owner_name", ""),
        "month": month,
        "amount": amount,
        "account": account,
        "bank_name": bank_name,
        "payment_date": payment_date,
        "notes": notes,
        "kms_year": lease.get("kms_year", ""),
        "season": lease.get("season", ""),
        "created_at": now_iso()
    }
    await db.truck_lease_payments.insert_one(payment)
    payment.pop("_id", None)
    
    # Create Cash Book nikasi entry
    txn_id = gen_id()
    cash_txn = {
        "id": txn_id,
        "date": payment_date,
        "account": account,
        "txn_type": "nikasi",
        "category": f"Truck Lease - {lease['truck_no']}",
        "party_type": "Truck Lease",
        "description": f"Lease payment {month} - {lease.get('owner_name', '')}",
        "amount": amount,
        "reference": f"lease_pay:{lease_id[:8]}",
        "linked_payment_id": f"truck_lease:{lease_id}:{month}:{payment_id}",
        "bank_name": bank_name,
        "kms_year": lease.get("kms_year", ""),
        "season": lease.get("season", ""),
        "created_by": data.get("created_by", ""),
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    await db.cash_transactions.insert_one(cash_txn)
    
    # Create auto-ledger entry (nikasi)
    ledger_entry = {
        "id": gen_id(),
        "date": payment_date,
        "account": "ledger",
        "txn_type": "nikasi",
        "category": f"Truck Lease - {lease['truck_no']}",
        "party_type": "Truck Lease",
        "description": f"Lease payment {month} - {lease.get('owner_name', '')}",
        "amount": amount,
        "reference": f"auto_ledger:{txn_id[:8]}",
        "kms_year": lease.get("kms_year", ""),
        "season": lease.get("season", ""),
        "created_at": now_iso(),
        "updated_at": now_iso()
    }
    await db.cash_transactions.insert_one(ledger_entry)
    
    return {"payment": payment, "cash_txn_id": txn_id, "message": f"Payment of Rs.{amount} recorded for {month}"}


# ========== PAYMENT HISTORY ==========

@router.get("/truck-leases/{lease_id}/history")
async def get_lease_payment_history(lease_id: str):
    payments = await db.truck_lease_payments.find({"lease_id": lease_id}, {"_id": 0}).to_list(5000)

    # v104.44.105 — Merge in matching Cash Book payments (so user sees ALL payments,
    # including ones made directly from Cash Book without "Pay Lease" button).
    lease = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    cb_extras = []
    if lease:
        cb_payments = await _fetch_cashbook_payments_for_lease(lease)
        for cb in cb_payments:
            cb_extras.append({
                "id": cb.get("id"),
                "lease_id": lease_id,
                "truck_no": lease["truck_no"],
                "owner_name": lease.get("owner_name", ""),
                "month": (cb.get("date") or "")[:7],
                "amount": cb.get("amount", 0),
                "account": cb.get("account", ""),
                "bank_name": cb.get("bank_name", ""),
                "payment_date": cb.get("date", ""),
                "notes": cb.get("description", ""),
                "source": "cashbook",  # marker so frontend can show badge
                "created_at": cb.get("created_at") or cb.get("date", ""),
            })
    all_payments = list(payments) + cb_extras
    return sorted(all_payments, key=lambda x: x.get("created_at", ""), reverse=True)


# v104.44.101 — Trips done by leased truck during lease window
@router.get("/truck-leases/{lease_id}/trips")
async def get_lease_trips(lease_id: str):
    """Returns aggregated trips (mill, private paddy, dc in/out, bp sales)
    where this leased truck was used DURING the lease period."""
    lease = await db.truck_leases.find_one({"id": lease_id}, {"_id": 0})
    if not lease:
        raise HTTPException(status_code=404, detail="Lease not found")
    return await _compute_lease_trips(lease)


# ========== CHECK IF TRUCK IS LEASED ==========

@router.get("/truck-leases/check/{truck_no}")
async def check_truck_leased(truck_no: str):
    lease = await db.truck_leases.find_one(
        {"truck_no": truck_no.upper(), "status": "active"}, {"_id": 0}
    )
    return {"is_leased": bool(lease), "lease": lease}


# ========== ALL LEASES SUMMARY (for Balance Sheet) ==========

@router.get("/truck-leases/summary")
async def get_leases_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {"status": "active"}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    leases = await db.truck_leases.find(query, {"_id": 0}).to_list(500)
    
    summary = []
    total_rent = 0
    total_paid = 0
    for lease in leases:
        months = get_months_between(lease.get("start_date", ""), lease.get("end_date", ""))
        rent = len(months) * lease.get("monthly_rent", 0)
        payments = await db.truck_lease_payments.find({"lease_id": lease["id"]}, {"_id": 0, "amount": 1}).to_list(5000)
        paid = sum(p.get("amount", 0) for p in payments)
        balance = round(rent - paid, 2)
        summary.append({
            "truck_no": lease["truck_no"],
            "owner_name": lease.get("owner_name", ""),
            "total_months": len(months),
            "monthly_rent": lease.get("monthly_rent", 0),
            "total_rent": round(rent, 2),
            "total_paid": round(paid, 2),
            "balance": max(0, balance),
            "advance_deposit": lease.get("advance_deposit", 0)
        })
        total_rent += rent
        total_paid += paid
    
    return {"leases": summary, "total_rent": round(total_rent, 2), "total_paid": round(total_paid, 2), "total_balance": round(max(0, total_rent - total_paid), 2)}



# ========== PDF EXPORT ==========

@router.get("/truck-leases/export/pdf")
async def export_leases_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from fastapi.responses import StreamingResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    import io

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    leases = await db.truck_leases.find(query, {"_id": 0}).to_list(500)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30)
    styles = get_pdf_styles()
    elements = []

    from utils.export_helpers import get_pdf_table_style
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph("Truck Lease Report", styles['Title']))
    if kms_year: elements.append(Paragraph(f"Year: {kms_year} | Season: {season or 'All'}", styles['Normal']))
    elements.append(Spacer(1, 12))

    header = ['Truck No.', 'Owner', 'Monthly Rent', 'Start', 'End', 'Advance', 'Status', 'Total Due', 'Paid', 'Balance']
    data = [header]
    grand_total = 0
    grand_paid = 0

    for lease in leases:
        months = get_months_between(lease.get("start_date", ""), lease.get("end_date", ""))
        total_rent = len(months) * lease.get("monthly_rent", 0)
        payments = await db.truck_lease_payments.find({"lease_id": lease["id"]}, {"_id": 0, "amount": 1}).to_list(5000)
        paid = sum(p.get("amount", 0) for p in payments)
        balance = round(total_rent - paid, 2)
        grand_total += total_rent
        grand_paid += paid
        data.append([
            lease.get("truck_no", ""), lease.get("owner_name", ""),
            f"Rs.{lease.get('monthly_rent', 0):,.0f}", fmt_date(lease.get("start_date", "")),
            fmt_date(lease.get("end_date", "")) or "Ongoing", f"Rs.{lease.get('advance_deposit', 0):,.0f}",
            lease.get("status", "").upper(),
            f"Rs.{total_rent:,.0f}", f"Rs.{paid:,.0f}", f"Rs.{max(0, balance):,.0f}"
        ])

    data.append(['', '', '', '', '', '', 'TOTAL', f"Rs.{grand_total:,.0f}", f"Rs.{grand_paid:,.0f}", f"Rs.{max(0, grand_total - grand_paid):,.0f}"])

    col_w = [65, 80, 70, 65, 65, 60, 45, 65, 60, 65]
    t = Table(data, colWidths=col_w)
    pdf_style = get_pdf_table_style(len(data))
    t.setStyle(TableStyle(pdf_style))
    elements.append(t)

    # ===== Beautiful single-line summary banner =====
    from utils.export_helpers import get_pdf_summary_banner, fmt_inr, STAT_COLORS
    bal = max(0, grand_total - grand_paid)
    active = sum(1 for l in leases if l.get('status', '').lower() == 'active')
    closed = len(leases) - active
    banner_stats = [
        {'label': 'TOTAL LEASES', 'value': str(len(leases)), 'color': STAT_COLORS['primary']},
        {'label': 'ACTIVE', 'value': str(active), 'color': STAT_COLORS['emerald']},
        {'label': 'CLOSED', 'value': str(closed), 'color': STAT_COLORS['orange']},
        {'label': 'TOTAL DUE', 'value': fmt_inr(grand_total), 'color': STAT_COLORS['gold']},
        {'label': 'PAID', 'value': fmt_inr(grand_paid), 'color': STAT_COLORS['green']},
        {'label': 'BALANCE', 'value': fmt_inr(bal), 'color': STAT_COLORS['red']},
    ]
    elements.append(Spacer(1, 6))
    banner = get_pdf_summary_banner(banner_stats, total_width=sum(col_w))
    if banner:
        elements.append(banner)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=truck_lease_report.pdf"})


# ========== EXCEL EXPORT ==========

@router.get("/truck-leases/export/excel")
async def export_leases_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    import io

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    leases = await db.truck_leases.find(query, {"_id": 0}).to_list(500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Truck Leases"

    from utils.export_helpers import style_excel_title, style_excel_header_row, style_excel_data_rows
    ncols = 12
    style_excel_title(ws, "Truck Lease Report", ncols)

    # Header at row 4
    headers = ['Truck No.', 'Owner', 'Monthly Rent', 'Start Date', 'End Date', 'Status',
               'Trips', 'Total Qntl', 'Total Months', 'Total Due', 'Total Paid', 'Balance']
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    ncols = len(headers)
    style_excel_header_row(ws, 4, ncols)

    row = 5
    for lease in leases:
        months = get_months_between(lease.get("start_date", ""), lease.get("end_date", ""))
        total_rent = len(months) * lease.get("monthly_rent", 0)
        payments = await db.truck_lease_payments.find({"lease_id": lease["id"]}, {"_id": 0, "amount": 1}).to_list(5000)
        paid = sum(p.get("amount", 0) for p in payments)
        balance = round(total_rent - paid, 2)
        trip_data = await _compute_lease_trips(lease)
        ws.cell(row=row, column=1, value=lease.get("truck_no", ""))
        ws.cell(row=row, column=2, value=lease.get("owner_name", ""))
        ws.cell(row=row, column=3, value=lease.get("monthly_rent", 0))
        ws.cell(row=row, column=4, value=fmt_date(lease.get("start_date", "")))
        ws.cell(row=row, column=5, value=fmt_date(lease.get("end_date", "")) or "Ongoing")
        ws.cell(row=row, column=6, value=lease.get("status", "").upper())
        ws.cell(row=row, column=7, value=trip_data["trip_count"])
        ws.cell(row=row, column=8, value=trip_data["total_qntl"])
        ws.cell(row=row, column=9, value=len(months))
        ws.cell(row=row, column=10, value=total_rent)
        ws.cell(row=row, column=11, value=paid)
        ws.cell(row=row, column=12, value=max(0, balance))
        row += 1

    for c in range(1, ncols + 1):
        ws.column_dimensions[chr(64 + c)].width = 14

    style_excel_data_rows(ws, 5, row - 1, ncols)

    # ===== Beautiful single-line summary banner =====
    if leases:
        from utils.export_helpers import add_excel_summary_banner, fmt_inr
        # Aggregate from already-computed loop above is gone; recompute
        gt = gp = 0
        for ls in leases:
            mts = get_months_between(ls.get("start_date", ""), ls.get("end_date", ""))
            tr = len(mts) * ls.get("monthly_rent", 0)
            pmts = await db.truck_lease_payments.find({"lease_id": ls["id"]}, {"_id": 0, "amount": 1}).to_list(5000)
            gt += tr
            gp += sum(p.get("amount", 0) for p in pmts)
        active = sum(1 for ls in leases if ls.get('status', '').lower() == 'active')
        sum_stats = [
            {'label': 'Total Leases', 'value': str(len(leases))},
            {'label': 'Active', 'value': str(active)},
            {'label': 'Closed', 'value': str(len(leases) - active)},
            {'label': 'Total Due', 'value': fmt_inr(gt)},
            {'label': 'Paid', 'value': fmt_inr(gp)},
            {'label': 'Balance', 'value': fmt_inr(max(0, gt - gp))},
        ]
        add_excel_summary_banner(ws, row + 1, ncols, sum_stats)

    buf = io.BytesIO()
    # 🎯 v104.44.9 — Apply consolidated multi-record polish
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": f"attachment; filename=truck_lease_report.xlsx"})
