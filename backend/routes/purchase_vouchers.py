from fastapi import APIRouter, HTTPException, Request
from database import db
from typing import Optional
from pydantic import BaseModel, Field, ConfigDict
from datetime import datetime, timezone
import uuid

router = APIRouter()

# ============ PURCHASE VOUCHERS ============

class PurchaseItemCreate(BaseModel):
    item_name: str
    quantity: float = 0
    rate: float = 0
    unit: str = "Qntl"

class PurchaseVoucherCreate(BaseModel):
    date: str
    party_name: str
    invoice_no: str = ""
    rst_no: str = ""
    items: list[PurchaseItemCreate] = []
    gst_type: str = "none"
    cgst_percent: float = 0
    sgst_percent: float = 0
    igst_percent: float = 0
    truck_no: str = ""
    cash_paid: float = 0
    diesel_paid: float = 0
    advance: float = 0
    eway_bill_no: str = ""
    remark: str = ""
    kms_year: str = ""
    season: str = ""

class PurchaseVoucher(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    voucher_no: int = 0
    invoice_no: str = ""
    rst_no: str = ""
    date: str = ""
    party_name: str = ""
    items: list = []
    subtotal: float = 0
    gst_type: str = "none"
    cgst_percent: float = 0
    sgst_percent: float = 0
    igst_percent: float = 0
    cgst_amount: float = 0
    sgst_amount: float = 0
    igst_amount: float = 0
    total: float = 0
    truck_no: str = ""
    cash_paid: float = 0
    diesel_paid: float = 0
    advance: float = 0
    eway_bill_no: str = ""
    paid_amount: float = 0
    balance: float = 0
    remark: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


async def _get_default_pump():
    pump = await db.diesel_accounts.find_one({}, {"_id": 0, "pump_name": 1}, sort=[("_id", -1)])
    return pump.get("pump_name", "Diesel Pump") if pump else "Diesel Pump"


async def _create_purchase_ledger_entries(d, doc_id, vno, items, username):
    """Create all accounting entries for a purchase voucher"""
    party = (d.get('party_name') or '').strip()
    cash = d.get('cash_paid', 0) or 0
    diesel = d.get('diesel_paid', 0) or 0
    advance = d.get('advance', 0) or 0
    total = d.get('total', 0) or 0
    truck = (d.get('truck_no') or '').strip()
    now_iso = datetime.now(timezone.utc).isoformat()
    base = {"kms_year": d.get('kms_year', ''), "season": d.get('season', ''), "created_by": username, "created_at": now_iso, "updated_at": now_iso}
    inv = d.get('invoice_no', '')
    desc_suffix = f" | Inv:{inv}" if inv else ""
    items_str = ', '.join(i['item_name'] for i in items)
    entries = []

    # 1. Party Ledger JAMA: total purchase amount (we owe the party)
    if party and total > 0:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "jama",
            "amount": total, "category": party, "party_type": "Purchase Voucher",
            "description": f"Purchase #{vno} - {items_str}{desc_suffix}",
            "reference": f"purchase_voucher:{doc_id}", **base
        })

    # 2. Advance paid to party: Party Ledger NIKASI (reduces what we owe) + Cash NIKASI (cash going out)
    if advance > 0 and party:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": advance, "category": party, "party_type": "Purchase Voucher",
            "description": f"Advance paid - Purchase #{vno}{desc_suffix}",
            "reference": f"purchase_voucher_adv:{doc_id}", **base
        })
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "cash", "txn_type": "nikasi",
            "amount": advance, "category": party, "party_type": "Purchase Voucher",
            "description": f"Advance paid - Purchase #{vno}{desc_suffix}",
            "reference": f"purchase_voucher_adv_cash:{doc_id}", **base
        })

    # 3. Cash paid → Cash NIKASI (cash going out)
    if cash > 0:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "cash", "txn_type": "nikasi",
            "amount": cash, "category": truck or party, "party_type": "Truck" if truck else "Purchase Voucher",
            "description": f"Truck cash - Purchase #{vno}{desc_suffix}",
            "reference": f"purchase_voucher_cash:{doc_id}", **base
        })

    # 4. Diesel paid → Diesel Pump Ledger JAMA + diesel_accounts entry
    if diesel > 0:
        pump_name = await _get_default_pump()
        pump_doc = await db.diesel_accounts.find_one({"pump_name": pump_name}, {"_id": 0, "pump_id": 1})
        pump_id = pump_doc.get("pump_id", "") if pump_doc else ""
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "jama",
            "amount": diesel, "category": pump_name, "party_type": "Diesel",
            "description": f"Diesel for truck - Purchase #{vno} - {party}{desc_suffix}",
            "reference": f"purchase_voucher_diesel:{doc_id}", **base
        })
        diesel_entry = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "pump_id": pump_id, "pump_name": pump_name,
            "truck_no": truck, "agent_name": party,
            "amount": diesel, "txn_type": "debit",
            "description": f"Diesel for Purchase #{vno} - {party}{desc_suffix}",
            "reference": f"purchase_voucher_diesel:{doc_id}",
            **base
        }
        await db.diesel_accounts.insert_one(diesel_entry)

    # 5. Truck cash+diesel → Truck Ledger NIKASI (deductions from future bhada)
    if cash > 0 and truck:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": cash, "category": truck, "party_type": "Truck",
            "description": f"Truck cash deduction - Purchase #{vno}{desc_suffix}",
            "reference": f"purchase_truck_cash:{doc_id}", **base
        })
    if diesel > 0 and truck:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": diesel, "category": truck, "party_type": "Truck",
            "description": f"Truck diesel deduction - Purchase #{vno}{desc_suffix}",
            "reference": f"purchase_truck_diesel:{doc_id}", **base
        })
    # Create truck_payments entry (rate will be set later by admin)
    truck_total = cash + diesel
    if truck_total > 0 and truck:
        truck_entry = {
            "entry_id": doc_id,
            "truck_no": truck, "date": d.get('date', ''),
            "cash_taken": cash, "diesel_taken": diesel,
            "gross_amount": 0, "deductions": truck_total,
            "net_amount": 0, "paid_amount": 0,
            "balance_amount": 0, "status": "pending",
            "source": "Purchase Voucher",
            "description": f"Purchase #{vno} - {party}{desc_suffix}",
            "reference": f"purchase_voucher_truck:{doc_id}",
            **base
        }
        await db.truck_payments.insert_one(truck_entry)

    for entry in entries:
        await db.cash_transactions.insert_one(entry)

    # Create local_party_accounts entry for purchase voucher (we owe = debit)
    if party and total > 0:
        lp = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "party_name": party, "txn_type": "debit",
            "amount": total, "description": f"Purchase #{vno} - {items_str}{desc_suffix}",
            "source_type": "purchase_voucher", "reference": f"purchase_voucher:{doc_id}",
            "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
            "created_by": username, "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.local_party_accounts.insert_one(lp)
    # If advance paid, add payment entry in local_party
    if advance > 0 and party:
        lp_adv = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "party_name": party, "txn_type": "payment",
            "amount": advance, "description": f"Advance paid - Purchase #{vno}{desc_suffix}",
            "source_type": "purchase_voucher_advance", "reference": f"purchase_voucher_adv:{doc_id}",
            "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
            "created_by": username, "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.local_party_accounts.insert_one(lp_adv)


@router.get("/purchase-book")
async def get_purchase_vouchers(kms_year: Optional[str] = None, season: Optional[str] = None,
                                party_name: Optional[str] = None, invoice_no: Optional[str] = None,
                                search: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if invoice_no: query["invoice_no"] = {"$regex": invoice_no, "$options": "i"}
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"invoice_no": {"$regex": search, "$options": "i"}},
            {"truck_no": {"$regex": search, "$options": "i"}},
            {"rst_no": {"$regex": search, "$options": "i"}},
        ]
    vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Get ledger-based paid amounts per party (includes Cash Book manual payments)
    party_names = list(set(v.get("party_name", "") for v in vouchers if v.get("party_name")))
    ledger_paid_map = {}
    if party_names:
        lq = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": party_names}}
        if kms_year: lq["kms_year"] = kms_year
        if season: lq["season"] = season
        ledger_txns = await db.cash_transactions.find(lq, {"_id": 0}).to_list(50000)
        for lt in ledger_txns:
            pn = lt.get("category", "")
            ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lt.get("amount", 0)
    
    for v in vouchers:
        pn = v.get("party_name", "")
        if pn:
            total_paid = round(ledger_paid_map.get(pn, 0), 2)
            v["ledger_paid"] = total_paid
            v["ledger_balance"] = round((v.get("total", 0)) - total_paid, 2)
    
    return vouchers


@router.post("/purchase-book")
async def create_purchase_voucher(input: PurchaseVoucherCreate, username: str = "", role: str = ""):
    d = input.model_dump()

    items = []
    subtotal = 0
    for item in d.get('items', []):
        amount = round(item.get('quantity', 0) * item.get('rate', 0), 2)
        items.append({**item, "amount": amount})
        subtotal += amount

    d['items'] = items
    d['subtotal'] = round(subtotal, 2)

    cgst_amt = round(subtotal * d.get('cgst_percent', 0) / 100, 2) if d.get('gst_type') == 'cgst_sgst' else 0
    sgst_amt = round(subtotal * d.get('sgst_percent', 0) / 100, 2) if d.get('gst_type') == 'cgst_sgst' else 0
    igst_amt = round(subtotal * d.get('igst_percent', 0) / 100, 2) if d.get('gst_type') == 'igst' else 0
    d['cgst_amount'] = cgst_amt
    d['sgst_amount'] = sgst_amt
    d['igst_amount'] = igst_amt
    total = round(subtotal + cgst_amt + sgst_amt + igst_amt, 2)
    d['total'] = total

    advance = d.get('advance', 0) or 0
    d['paid_amount'] = round(advance, 2)
    d['balance'] = round(total - advance, 2)
    d['created_by'] = username

    last = await db.purchase_vouchers.find_one(sort=[("voucher_no", -1)], projection={"_id": 0, "voucher_no": 1})
    d['voucher_no'] = (last.get('voucher_no', 0) if last else 0) + 1

    obj = PurchaseVoucher(**d)
    doc = obj.model_dump()
    await db.purchase_vouchers.insert_one(doc)
    doc.pop('_id', None)

    await _create_purchase_ledger_entries(d, doc.get('id', ''), d['voucher_no'], items, username)
    return doc


@router.put("/purchase-book/{voucher_id}")
async def update_purchase_voucher(voucher_id: str, input: PurchaseVoucherCreate, username: str = "", role: str = ""):
    existing = await db.purchase_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Voucher not found")

    d = input.model_dump()
    items = []
    subtotal = 0
    for item in d.get('items', []):
        amount = round(item.get('quantity', 0) * item.get('rate', 0), 2)
        items.append({**item, "amount": amount})
        subtotal += amount

    d['items'] = items
    d['subtotal'] = round(subtotal, 2)
    cgst_amt = round(subtotal * d.get('cgst_percent', 0) / 100, 2) if d.get('gst_type') == 'cgst_sgst' else 0
    sgst_amt = round(subtotal * d.get('sgst_percent', 0) / 100, 2) if d.get('gst_type') == 'cgst_sgst' else 0
    igst_amt = round(subtotal * d.get('igst_percent', 0) / 100, 2) if d.get('gst_type') == 'igst' else 0
    d['cgst_amount'] = cgst_amt
    d['sgst_amount'] = sgst_amt
    d['igst_amount'] = igst_amt
    total = round(subtotal + cgst_amt + sgst_amt + igst_amt, 2)
    d['total'] = total
    advance = d.get('advance', 0) or 0
    d['paid_amount'] = round(advance, 2)
    d['balance'] = round(total - advance, 2)
    d['updated_at'] = datetime.now(timezone.utc).isoformat()

    await db.purchase_vouchers.update_one({"id": voucher_id}, {"$set": d})

    # Delete old accounting entries and recreate
    await db.cash_transactions.delete_many({"reference": {"$regex": f"purchase_voucher.*:{voucher_id}"}})
    await db.diesel_accounts.delete_many({"reference": {"$regex": f"purchase_voucher.*:{voucher_id}"}})
    await db.truck_payments.delete_many({"reference": {"$regex": f"purchase_voucher.*:{voucher_id}"}})
    await db.local_party_accounts.delete_many({"reference": {"$regex": f"purchase_voucher.*:{voucher_id}"}})
    vno = existing.get('voucher_no', 0)
    await _create_purchase_ledger_entries(d, voucher_id, vno, items, username)

    updated = await db.purchase_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    return updated


@router.delete("/purchase-book/{voucher_id}")
async def delete_purchase_voucher(voucher_id: str, username: str = "", role: str = ""):
    existing = await db.purchase_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Voucher not found")
    await db.purchase_vouchers.delete_one({"id": voucher_id})
    await db.cash_transactions.delete_many({"reference": {"$regex": f"purchase_voucher.*:{voucher_id}"}})
    await db.diesel_accounts.delete_many({"reference": {"$regex": f"purchase_voucher.*:{voucher_id}"}})
    await db.truck_payments.delete_many({"reference": {"$regex": f"purchase_voucher.*:{voucher_id}"}})
    await db.local_party_accounts.delete_many({"reference": {"$regex": f"purchase_voucher.*:{voucher_id}"}})
    return {"message": "Purchase voucher deleted", "id": voucher_id}


@router.post("/purchase-book/delete-bulk")
async def bulk_delete_purchase_vouchers(request: Request):
    data = await request.json()
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    deleted = 0
    for vid in ids:
        existing = await db.purchase_vouchers.find_one({"id": vid}, {"_id": 0})
        if existing:
            await db.purchase_vouchers.delete_one({"id": vid})
            await db.cash_transactions.delete_many({"reference": {"$regex": f"purchase_voucher.*:{vid}"}})
            await db.diesel_accounts.delete_many({"reference": {"$regex": f"purchase_voucher.*:{vid}"}})
            await db.truck_payments.delete_many({"reference": {"$regex": f"purchase_voucher.*:{vid}"}})
            await db.local_party_accounts.delete_many({"reference": {"$regex": f"purchase_voucher.*:{vid}"}})
            deleted += 1
    return {"message": f"{deleted} purchase vouchers deleted", "deleted": deleted}

@router.get("/purchase-book/item-suggestions")
async def get_item_suggestions():
    """Get unique item names from purchase vouchers for autocomplete"""
    pipeline = [
        {"$unwind": "$items"},
        {"$group": {"_id": "$items.item_name"}},
        {"$sort": {"_id": 1}}
    ]
    results = await db.purchase_vouchers.aggregate(pipeline).to_list(1000)
    return [r["_id"] for r in results if r["_id"]]


@router.get("/purchase-book/stock-items")
async def get_purchase_stock_items(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get all stock items with available quantities for purchase form"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    from utils.stock_calculator import (
        calc_cmr_paddy_in, calc_pvt_paddy_in, calc_paddy_used,
        calc_rice_produced, calc_govt_delivered, calc_pvt_rice_sold,
        calc_sale_voucher_items, calc_purchase_voucher_items,
        calc_byproduct_produced, calc_byproduct_sold, calc_frk_in, BY_PRODUCTS
    )

    milling = await db.milling_entries.find(query, {"_id": 0}).to_list(10000)
    dc = await db.dc_entries.find(query, {"_id": 0}).to_list(10000)
    pvt_sales = await db.rice_sales.find(query, {"_id": 0}).to_list(10000)
    sale_vouchers = await db.sale_vouchers.find(query, {"_id": 0}).to_list(10000)
    bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(10000)
    purchase_vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).to_list(10000)
    mill_entries = await db.mill_entries.find(query, {"_id": 0}).to_list(10000)
    pvt_paddy = await db.private_paddy.find(query, {"_id": 0}).to_list(10000)

    cmr_paddy_in = calc_cmr_paddy_in(mill_entries)
    pvt_paddy_in = calc_pvt_paddy_in(pvt_paddy)
    paddy_used = calc_paddy_used(milling)
    usna_produced = calc_rice_produced(milling, 'usna')
    raw_produced = calc_rice_produced(milling, 'raw')
    govt_delivered = calc_govt_delivered(dc)
    pvt_sold_usna = calc_pvt_rice_sold(pvt_sales, 'usna')
    pvt_sold_raw = calc_pvt_rice_sold(pvt_sales, 'raw')
    sb_sold = calc_sale_voucher_items(sale_vouchers)
    pv_bought = calc_purchase_voucher_items(purchase_vouchers)
    bp_produced = calc_byproduct_produced(milling)
    bp_sold_map = calc_byproduct_sold(bp_sales)

    items = []

    # Paddy
    paddy_total_in = round(cmr_paddy_in + pvt_paddy_in + pv_bought.get("Paddy", 0), 2)
    items.append({"name": "Paddy", "available_qntl": round(paddy_total_in - paddy_used, 2), "unit": "Qntl"})

    # Rice
    usna_avail = round(usna_produced + pv_bought.get("Rice (Usna)", 0) - govt_delivered - pvt_sold_usna - sb_sold.get("Rice (Usna)", 0), 2)
    items.append({"name": "Rice (Usna)", "available_qntl": usna_avail, "unit": "Qntl"})
    raw_avail = round(raw_produced + pv_bought.get("Rice (Raw)", 0) - pvt_sold_raw - sb_sold.get("Rice (Raw)", 0), 2)
    items.append({"name": "Rice (Raw)", "available_qntl": raw_avail, "unit": "Qntl"})

    # By-products
    for p in products:
        produced = bp_produced.get(p, 0)
        sold_bp = round(bp_sold_map.get(p, 0), 2)
        sold_sb = sb_sold.get(p.title(), 0)
        purchased = pv_bought.get(p.title(), 0)
        avail = round(produced + purchased - sold_bp - sold_sb, 2)
        items.append({"name": p.title(), "available_qntl": avail, "unit": "Qntl"})

    # FRK
    frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(10000) if await db.frk_purchases.count_documents(query) > 0 else []
    frk_in = round(sum(e.get('quantity_qntl', 0) or e.get('quantity', 0) for e in frk_purchases), 2)
    frk_purchased_pv = pv_bought.get("FRK", 0)
    frk_sold_sb = sb_sold.get("FRK", 0)
    items.append({"name": "FRK", "available_qntl": round(frk_in + frk_purchased_pv - frk_sold_sb, 2), "unit": "Qntl"})

    # FRK Rice (common purchase item)
    frk_rice_bought = pv_bought.get("FRK Rice", 0)
    frk_rice_sold = sb_sold.get("FRK Rice", 0)
    if frk_rice_bought > 0 or frk_rice_sold > 0:
        items.append({"name": "FRK Rice", "available_qntl": round(frk_rice_bought - frk_rice_sold, 2), "unit": "Qntl"})

    # Custom items from previous purchases (not already covered)
    known_items = {"Paddy", "Rice (Usna)", "Rice (Raw)", "FRK", "FRK Rice"} | {p.title() for p in products}
    for item_name, qty in pv_bought.items():
        if item_name not in known_items and item_name:
            sold = sb_sold.get(item_name, 0)
            items.append({"name": item_name, "available_qntl": round(qty - sold, 2), "unit": "Qntl"})

    return items


# ============ STOCK SUMMARY ============

@router.get("/stock-summary")
async def get_stock_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    # ===== FETCH OPENING STOCK =====
    ob = {}
    if kms_year:
        ob_doc = await db.opening_stock.find_one({"kms_year": kms_year}, {"_id": 0})
        if ob_doc:
            ob = ob_doc.get("stocks", {})
    # Map opening stock keys to stock item names
    ob_paddy = float(ob.get("paddy", 0))
    ob_usna = float(ob.get("rice_usna", ob.get("rice", 0)))
    ob_raw = float(ob.get("rice_raw", 0))
    ob_bran = float(ob.get("bran", 0))
    ob_kunda = float(ob.get("kunda", 0))
    ob_broken = float(ob.get("broken", 0))
    ob_kanki = float(ob.get("kanki", 0))
    ob_husk = float(ob.get("husk", 0))
    ob_frk = float(ob.get("frk", 0))

    from utils.stock_calculator import (
        calc_cmr_paddy_in, calc_pvt_paddy_in, calc_paddy_used,
        calc_rice_produced, calc_govt_delivered, calc_pvt_rice_sold,
        calc_sale_voucher_items, calc_purchase_voucher_items,
        calc_byproduct_produced, calc_byproduct_sold, calc_frk_in, BY_PRODUCTS
    )

    milling = await db.milling_entries.find(query, {"_id": 0}).to_list(10000)
    dc = await db.dc_entries.find(query, {"_id": 0}).to_list(10000)
    pvt_sales = await db.rice_sales.find(query, {"_id": 0}).to_list(10000)
    sale_vouchers = await db.sale_vouchers.find(query, {"_id": 0}).to_list(10000)
    bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(10000)
    purchase_vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).to_list(10000)
    mill_entries = await db.mill_entries.find(query, {"_id": 0}).to_list(10000)
    pvt_paddy = await db.private_paddy.find({**query, "source": {"$ne": "agent_extra"}}, {"_id": 0}).to_list(10000)

    cmr_paddy_in = calc_cmr_paddy_in(mill_entries)
    pvt_paddy_in = calc_pvt_paddy_in(pvt_paddy)
    paddy_used_milling = calc_paddy_used(milling)
    usna_produced = calc_rice_produced(milling, 'usna')
    raw_produced = calc_rice_produced(milling, 'raw')
    govt_delivered = calc_govt_delivered(dc)
    pvt_sold_usna = calc_pvt_rice_sold(pvt_sales, 'usna')
    pvt_sold_raw = calc_pvt_rice_sold(pvt_sales, 'raw')
    sb_sold = calc_sale_voucher_items(sale_vouchers)
    pv_bought = calc_purchase_voucher_items(purchase_vouchers)
    bp_produced = calc_byproduct_produced(milling)
    bp_sold_map = calc_byproduct_sold(bp_sales)
    products = BY_PRODUCTS

    frk_purchases = []
    if await db.frk_purchases.count_documents(query) > 0:
        frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(10000)
    frk_in = calc_frk_in(frk_purchases)

    # Build stock items list
    stock_items = []
    ob_map = {}  # opening balance per item

    # Paddy
    pv_paddy = round(pv_bought.get("Paddy", 0), 2)
    paddy_total_in = round(cmr_paddy_in + pvt_paddy_in + pv_paddy, 2)
    stock_items.append({
        "name": "Paddy", "category": "Raw Material",
        "opening": ob_paddy,
        "in_qty": paddy_total_in, "out_qty": paddy_used_milling,
        "available": round(ob_paddy + paddy_total_in - paddy_used_milling, 2), "unit": "Qntl",
        "details": f"OB: {ob_paddy}Q + CMR: {cmr_paddy_in}Q + Pvt: {pvt_paddy_in}Q + Purchase: {pv_paddy}Q - Milling: {paddy_used_milling}Q"
    })

    # Rice Usna
    pv_usna = round(pv_bought.get("Rice (Usna)", 0), 2)
    usna_sold_total = round(govt_delivered + pvt_sold_usna + sb_sold.get("Rice (Usna)", 0), 2)
    usna_avail = round(ob_usna + usna_produced + pv_usna - usna_sold_total, 2)
    stock_items.append({
        "name": "Rice (Usna)", "category": "Finished",
        "opening": ob_usna,
        "in_qty": round(usna_produced + pv_usna, 2), "out_qty": usna_sold_total,
        "available": usna_avail, "unit": "Qntl",
        "details": f"OB: {ob_usna}Q + Milling: {usna_produced}Q + Purchase: {pv_usna}Q - DC: {govt_delivered}Q - Pvt: {pvt_sold_usna}Q - Sale: {sb_sold.get('Rice (Usna)', 0)}Q"
    })

    # Rice Raw
    pv_raw = round(pv_bought.get("Rice (Raw)", 0), 2)
    raw_sold_total = round(pvt_sold_raw + sb_sold.get("Rice (Raw)", 0), 2)
    raw_avail = round(ob_raw + raw_produced + pv_raw - raw_sold_total, 2)
    stock_items.append({
        "name": "Rice (Raw)", "category": "Finished",
        "opening": ob_raw,
        "in_qty": round(raw_produced + pv_raw, 2), "out_qty": raw_sold_total,
        "available": raw_avail, "unit": "Qntl",
        "details": f"OB: {ob_raw}Q + Milling: {raw_produced}Q + Purchase: {pv_raw}Q - Pvt: {pvt_sold_raw}Q - Sale: {sb_sold.get('Rice (Raw)', 0)}Q"
    })

    # By-products - Dynamic categories
    bp_cats = await db.byproduct_categories.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    if not bp_cats:
        bp_cats = [{"id": p, "name": p.title()} for p in ["bran", "kunda", "broken", "kanki", "husk"]]
    for cat in bp_cats:
        p = cat["id"]
        display_name = cat.get("name", p.title())
        produced = bp_produced.get(p, 0)
        sold_bp = round(bp_sold_map.get(p, 0), 2)
        sold_sb = sb_sold.get(display_name, 0) + sb_sold.get(p.title(), 0) + sb_sold.get(p, 0)
        purchased = pv_bought.get(display_name, 0) + pv_bought.get(p.title(), 0) + pv_bought.get(p, 0)
        item_ob = float(ob.get(p, 0))
        total_in = round(produced + purchased, 2)
        total_out = round(sold_bp + sold_sb, 2)
        avail = round(item_ob + total_in - total_out, 2)
        stock_items.append({
            "name": display_name, "category": "By-Product",
            "opening": item_ob,
            "in_qty": total_in, "out_qty": total_out,
            "available": avail, "unit": "Qntl",
            "details": f"OB: {item_ob}Q + Milling: {produced}Q + Purchased: {purchased}Q - Sold: {sold_bp}Q - Sale Voucher: {sold_sb}Q"
        })

    # FRK
    frk_purchased_pv = pv_bought.get("FRK", 0)
    frk_total_in = round(frk_in + frk_purchased_pv, 2)
    frk_sold_sb = sb_sold.get("FRK", 0)
    stock_items.append({
        "name": "FRK", "category": "By-Product",
        "opening": ob_frk,
        "in_qty": frk_total_in, "out_qty": frk_sold_sb,
        "available": round(ob_frk + frk_total_in - frk_sold_sb, 2), "unit": "Qntl",
        "details": f"OB: {ob_frk}Q + FRK Purchase: {frk_in}Q + Purchase Voucher: {frk_purchased_pv}Q - Sale Voucher: {frk_sold_sb}Q"
    })

    # Custom items from purchase vouchers (not already covered above)
    known_items = {"Paddy", "Rice (Usna)", "Rice (Raw)", "FRK"} | {p.title() for p in products}
    for item_name, qty in pv_bought.items():
        if item_name not in known_items:
            sold = sb_sold.get(item_name, 0)
            stock_items.append({
                "name": item_name, "category": "Custom",
                "opening": 0,
                "in_qty": round(qty, 2), "out_qty": round(sold, 2),
                "available": round(qty - sold, 2), "unit": "Qntl",
                "details": f"Purchased: {qty}Q - Sold: {sold}Q"
            })

    # Gunny Bags excluded from stock summary (tracked separately)

    return {"items": stock_items}


# ============ PURCHASE BOOK PDF EXPORT ============

@router.get("/purchase-book/export/pdf")
async def export_purchase_book_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None):
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import io

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"invoice_no": {"$regex": search, "$options": "i"}},
        ]
    vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).sort("voucher_no", 1).to_list(10000)

    # Ledger-based paid amounts
    party_names = list(set(v.get("party_name", "") for v in vouchers if v.get("party_name")))
    ledger_paid_map = {}
    if party_names:
        ledger_q = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": party_names}, "party_type": "Purchase Book"}
        if kms_year: ledger_q["kms_year"] = kms_year
        ledger_txns = await db.cash_transactions.find(ledger_q, {"_id": 0}).to_list(50000)
        for lt in ledger_txns:
            pn = lt.get("category", "")
            ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lt.get("amount", 0)

    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=6*mm, rightMargin=6*mm, topMargin=8*mm, bottomMargin=6*mm)
    elements = []
    styles = get_pdf_styles()

    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())

    meta_parts = ["Purchase Book"]
    if kms_year: meta_parts.append(f"FY: {kms_year}")
    if season: meta_parts.append(season)
    meta_parts.append(f"Date: {datetime.now().strftime('%d-%m-%Y')}")
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#555555'), alignment=TA_CENTER, spaceAfter=4)
    elements.append(Paragraph(" | ".join(meta_parts), meta_style))

    cell_s = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=6, leading=8)
    cell_r = ParagraphStyle('CellR', parent=styles['Normal'], fontSize=6, leading=8, alignment=TA_RIGHT)
    cell_rb = ParagraphStyle('CellRB', parent=styles['Normal'], fontSize=6, leading=8, alignment=TA_RIGHT, fontName='FreeSansBold')

    headers = ['#', 'Date', 'Inv', 'Party', 'Items', 'Truck', 'Total', 'Adv', 'Cash', 'Diesel', 'Ledger Paid', 'Balance', 'Status']
    table_data = [headers]
    # Optimized for A4 landscape
    col_widths = [12*mm, 18*mm, 16*mm, 28*mm, 52*mm, 20*mm, 20*mm, 16*mm, 16*mm, 16*mm, 20*mm, 20*mm, 16*mm]

    g = {"total": 0, "adv": 0, "cash": 0, "diesel": 0, "paid": 0, "bal": 0}
    for v in vouchers:
        items_str = ', '.join(f"{i['item_name']} ({i.get('quantity', 0)} Qntl)" for i in v.get('items', []))
        dp = str(v.get('date', '')).split('-')
        dt = f"{dp[2]}/{dp[1]}/{dp[0]}" if len(dp) == 3 else v.get('date', '')
        total = v.get('total', 0) or 0
        pn = v.get('party_name', '')
        ledger_paid = round(ledger_paid_map.get(pn, 0), 2)
        ledger_bal = round(total - ledger_paid, 2)
        status = "Paid" if ledger_bal <= 0 and total > 0 else "Pending"
        g["total"] += total
        g["adv"] += v.get('advance', 0) or 0
        g["cash"] += v.get('cash_paid', 0) or 0
        g["diesel"] += v.get('diesel_paid', 0) or 0
        g["paid"] += ledger_paid
        g["bal"] += ledger_bal
        table_data.append([
            Paragraph(f"{v.get('voucher_no','')}", cell_s),
            Paragraph(dt, cell_s),
            Paragraph(str(v.get('invoice_no', '')), cell_s),
            Paragraph(f"<b>{pn}</b>", cell_s),
            Paragraph(items_str, cell_s),
            Paragraph(str(v.get('truck_no', '')), cell_s),
            Paragraph(f"{total:,.0f}", cell_rb),
            Paragraph(f"{v.get('advance',0) or 0:,.0f}", cell_r),
            Paragraph(f"{v.get('cash_paid',0) or 0:,.0f}", cell_r),
            Paragraph(f"{v.get('diesel_paid',0) or 0:,.0f}", cell_r),
            Paragraph(f"{ledger_paid:,.0f}", cell_r),
            Paragraph(f"{ledger_bal:,.0f}", cell_rb),
            Paragraph(status, cell_s),
        ])

    table_data.append([
        Paragraph(f"<b>TOTAL ({len(vouchers)})</b>", cell_s), '', '', '', '', '',
        Paragraph(f"<b>{g['total']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['adv']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['cash']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['diesel']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['paid']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['bal']:,.0f}</b>", cell_rb),
        '',
    ])

    tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e7d32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#2e7d32')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
    ]
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=6, textColor=colors.HexColor('#999999'), alignment=TA_CENTER, spaceBefore=6)
    elements.append(Paragraph(f"{company} - Purchase Book | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", footer_style))

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=purchase_book.pdf"})


# ============ PURCHASE BOOK EXCEL EXPORT ============

@router.get("/purchase-book/export/excel")
async def export_purchase_book_excel(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None):
    from fastapi.responses import Response
    import io
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"invoice_no": {"$regex": search, "$options": "i"}},
        ]
    vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).sort("voucher_no", 1).to_list(10000)
    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")

    # Ledger-based paid amounts
    party_names = list(set(v.get("party_name", "") for v in vouchers if v.get("party_name")))
    ledger_paid_map = {}
    if party_names:
        ledger_q = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": party_names}, "party_type": "Purchase Book"}
        if kms_year: ledger_q["kms_year"] = kms_year
        ledger_txns = await db.cash_transactions.find(ledger_q, {"_id": 0}).to_list(50000)
        for lt in ledger_txns:
            pn = lt.get("category", "")
            ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lt.get("amount", 0)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from utils.export_helpers import (style_excel_title, style_excel_header_row,
            style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)

        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Book"

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        cols = ['#', 'Date', 'Inv No.', 'Party', 'Items (Qntl)', 'Truck', 'Total', 'Advance', 'Cash', 'Diesel', 'Ledger Paid', 'Balance', 'Status']
        widths = [5, 9, 9, 16, 30, 10, 10, 9, 8, 8, 10, 10, 7]
        ncols = len(cols)

        title = f"{company} - Purchase Book / खरीद बही"
        subtitle = f"FY: {kms_year or 'All'} | {season or 'All'} | {datetime.now().strftime('%d-%m-%Y')}"
        style_excel_title(ws, title, ncols, subtitle)

        for i, (col, w) in enumerate(zip(cols, widths), 1):
            ws.cell(row=4, column=i, value=col)
            ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w
        style_excel_header_row(ws, 4, ncols)

        g = {"total": 0, "adv": 0, "cash": 0, "diesel": 0, "paid": 0, "bal": 0}
        data_start = 5
        for ri, v in enumerate(vouchers, data_start):
            items_str = ', '.join(f"{i['item_name']} ({i.get('quantity', 0)} Qntl)" for i in v.get('items', []))
            dp = str(v.get('date', '')).split('-')
            dt = f"{dp[2]}/{dp[1]}/{dp[0]}" if len(dp) == 3 else v.get('date', '')
            total = v.get('total', 0) or 0
            pn = v.get('party_name', '')
            ledger_paid = round(ledger_paid_map.get(pn, 0), 2)
            ledger_bal = round(total - ledger_paid, 2)
            status = "Paid" if ledger_bal <= 0 and total > 0 else "Pending"
            g["total"] += total
            g["adv"] += v.get('advance', 0) or 0
            g["cash"] += v.get('cash_paid', 0) or 0
            g["diesel"] += v.get('diesel_paid', 0) or 0
            g["paid"] += ledger_paid
            g["bal"] += ledger_bal

            row_data = [v.get('voucher_no', ''), dt, v.get('invoice_no', ''), pn, items_str,
                        v.get('truck_no', ''), total, v.get('advance', 0) or 0, v.get('cash_paid', 0) or 0,
                        v.get('diesel_paid', 0) or 0, ledger_paid, ledger_bal, status]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ci >= 7 and ci <= 12:
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(val, (int, float)): cell.number_format = '#,##0'

        if vouchers:
            style_excel_data_rows(ws, data_start, data_start + len(vouchers) - 1, ncols, cols)

        tr = len(vouchers) + data_start
        ws.cell(row=tr, column=1, value=f"TOTAL ({len(vouchers)} vouchers)")
        for ci, val in enumerate([g['total'], g['adv'], g['cash'], g['diesel'], g['paid'], g['bal'], ''], 7):
            cell = ws.cell(row=tr, column=ci, value=val)
            cell.alignment = Alignment(horizontal='right')
            if isinstance(val, (int, float)): cell.number_format = '#,##0'
        style_excel_total_row(ws, tr, ncols)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(content=buf.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename=purchase_book_{datetime.now().strftime('%Y%m%d')}.xlsx"})
    except ImportError:
        return {"error": "openpyxl not installed"}


# ============ SINGLE PURCHASE VOUCHER INVOICE PDF ============
# NOTE: This route must come AFTER /purchase-book/export/pdf to avoid route conflicts

@router.get("/purchase-book/{voucher_id}/pdf")
async def export_single_purchase_voucher_pdf(voucher_id: str):
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import io

    v = await db.purchase_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    if not v: raise HTTPException(status_code=404, detail="Voucher not found")

    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "")
    address = branding.get("address", "")
    phone = branding.get("phone", "")
    gstin = branding.get("gstin", "")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm)
    elements = []
    styles = get_pdf_styles()
    green = '#2e7d32'

    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())

    label_s = ParagraphStyle('L', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#555'))
    val_s = ParagraphStyle('V', parent=styles['Normal'], fontSize=10, fontName='FreeSansBold')
    cell_s = ParagraphStyle('C', parent=styles['Normal'], fontSize=9, leading=12)
    cell_r = ParagraphStyle('CR', parent=styles['Normal'], fontSize=9, leading=12, alignment=TA_RIGHT)
    cell_rb = ParagraphStyle('CRB', parent=styles['Normal'], fontSize=10, leading=12, alignment=TA_RIGHT, fontName='FreeSansBold')
    hd_s = ParagraphStyle('H', parent=styles['Normal'], fontSize=9, leading=12, fontName='FreeSansBold', textColor=colors.white)

    elements.append(Spacer(1, 4*mm))

    inv_title = ParagraphStyle('IT', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor(green), alignment=TA_CENTER, spaceAfter=3)
    elements.append(Paragraph("PURCHASE VOUCHER", inv_title))

    dp = str(v.get('date', '')).split('-')
    date_str = f"{dp[2]}/{dp[1]}/{dp[0]}" if len(dp) == 3 else v.get('date', '')

    info_data = [
        [Paragraph('<b>Voucher No:</b>', label_s), Paragraph(f"#{v.get('voucher_no', '')}", val_s),
         Paragraph('<b>Date:</b>', label_s), Paragraph(date_str, val_s)],
        [Paragraph('<b>Party/Seller:</b>', label_s), Paragraph(v.get('party_name', ''), val_s),
         Paragraph('<b>Invoice No:</b>', label_s), Paragraph(v.get('invoice_no', ''), val_s)],
        [Paragraph('<b>Truck No:</b>', label_s), Paragraph(v.get('truck_no', ''), val_s),
         Paragraph('<b>RST No:</b>', label_s), Paragraph(v.get('rst_no', ''), val_s)],
    ]
    if v.get('eway_bill_no'):
        info_data.append([Paragraph('<b>E-Way Bill:</b>', label_s), Paragraph(v.get('eway_bill_no', ''), val_s), '', ''])

    info_tbl = RLTable(info_data, colWidths=[30*mm, 55*mm, 30*mm, 55*mm])
    info_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), 
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#ddd')),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 5*mm))

    items_header = [Paragraph('<b>S.No</b>', hd_s), Paragraph('<b>Item Name</b>', hd_s),
                    Paragraph('<b>Quantity (Qntl)</b>', hd_s), Paragraph('<b>Rate (Rs)</b>', hd_s),
                    Paragraph('<b>Amount (Rs)</b>', hd_s)]
    items_data = [items_header]
    for idx, item in enumerate(v.get('items', []), 1):
        qty = item.get('quantity', 0) or 0
        rate = item.get('rate', 0) or 0
        amt = round(qty * rate, 2)
        items_data.append([
            Paragraph(str(idx), cell_s), Paragraph(item.get('item_name', ''), cell_s),
            Paragraph(f"{qty} Qntl", cell_r), Paragraph(f"{rate:,.2f}", cell_r),
            Paragraph(f"{amt:,.2f}", cell_rb)
        ])

    items_tbl = RLTable(items_data, colWidths=[15*mm, 60*mm, 30*mm, 30*mm, 35*mm], repeatRows=1)
    items_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(green)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    for i in range(1, len(items_data)):
        if i % 2 == 0:
            items_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
    items_tbl.setStyle(TableStyle(items_style))
    elements.append(items_tbl)
    elements.append(Spacer(1, 3*mm))

    subtotal = v.get('subtotal', 0) or sum((i.get('quantity', 0) or 0) * (i.get('rate', 0) or 0) for i in v.get('items', []))
    cgst = v.get('cgst_amount', 0) or 0
    sgst = v.get('sgst_amount', 0) or 0
    igst = v.get('igst_amount', 0) or 0
    total = v.get('total', 0) or 0
    advance = v.get('advance', 0) or 0
    cash = v.get('cash_paid', 0) or 0
    diesel = v.get('diesel_paid', 0) or 0
    balance = v.get('balance', 0) or 0

    tot_s = ParagraphStyle('TS', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT)
    tot_b = ParagraphStyle('TB', parent=styles['Normal'], fontSize=11, alignment=TA_RIGHT, fontName='FreeSansBold', textColor=colors.HexColor(green))

    total_data = [[Paragraph('Subtotal:', tot_s), Paragraph(f"Rs. {subtotal:,.2f}", tot_s)]]
    if cgst > 0: total_data.append([Paragraph(f"CGST ({v.get('cgst_percent',0)}%):", tot_s), Paragraph(f"Rs. {cgst:,.2f}", tot_s)])
    if sgst > 0: total_data.append([Paragraph(f"SGST ({v.get('sgst_percent',0)}%):", tot_s), Paragraph(f"Rs. {sgst:,.2f}", tot_s)])
    if igst > 0: total_data.append([Paragraph(f"IGST ({v.get('igst_percent',0)}%):", tot_s), Paragraph(f"Rs. {igst:,.2f}", tot_s)])
    total_data.append([Paragraph('<b>Grand Total:</b>', tot_b), Paragraph(f"<b>Rs. {total:,.2f}</b>", tot_b)])
    if advance > 0: total_data.append([Paragraph('Advance Paid:', tot_s), Paragraph(f"Rs. {advance:,.2f}", tot_s)])
    total_data.append([Paragraph('<b>Balance Due:</b>', tot_b), Paragraph(f"<b>Rs. {balance:,.2f}</b>", tot_b)])

    tot_tbl = RLTable(total_data, colWidths=[120*mm, 50*mm])
    tot_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), 
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor(green)),
    ]))
    elements.append(tot_tbl)

    if v.get('remark'):
        elements.append(Spacer(1, 3*mm))
        elements.append(Paragraph(f"<b>Remark:</b> {v['remark']}", label_s))

    elements.append(Spacer(1, 15*mm))
    sig_s = ParagraphStyle('Sig', alignment=TA_CENTER, fontSize=9)
    sig_b = ParagraphStyle('SigB', alignment=TA_CENTER, fontSize=9, fontName='FreeSansBold')
    sig_data = [[Paragraph('Received By', sig_s), Paragraph(f'For {company}', sig_b)]]
    sig_tbl = RLTable(sig_data, colWidths=[85*mm, 85*mm])
    sig_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), ('LINEABOVE', (0, 0), (0, 0), 0.5, colors.black), ('LINEABOVE', (1, 0), (1, 0), 0.5, colors.black)]))
    elements.append(sig_tbl)

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=purchase_voucher_{v.get('voucher_no','')}.pdf"})


# ============ STOCK SUMMARY PDF EXPORT ============

@router.get("/stock-summary/export/pdf")
async def export_stock_summary_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import io

    data = await get_stock_summary(kms_year, season)
    items = data.get("items", [])

    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=15*mm, bottomMargin=12*mm)
    elements = []
    styles = get_pdf_styles()

    # Company header
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())

    meta_parts = ["Stock Summary Report"]
    if kms_year: meta_parts.append(f"FY: {kms_year}")
    if season: meta_parts.append(season)
    meta_parts.append(f"Date: {datetime.now().strftime('%d-%m-%Y')}")
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#555555'), alignment=TA_CENTER, spaceAfter=10)
    elements.append(Paragraph(" | ".join(meta_parts), meta_style))
    elements.append(Spacer(1, 5))

    # Category colors
    cat_colors = {
        "Raw Material": colors.HexColor('#F59E0B'),
        "Finished": colors.HexColor('#10B981'),
        "By-Product": colors.HexColor('#3B82F6'),
        "Custom": colors.HexColor('#8B5CF6'),
    }

    # Group by category
    grouped = {}
    for item in items:
        cat = item.get('category', 'Other')
        if cat not in grouped: grouped[cat] = []
        grouped[cat].append(item)

    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10)
    cell_r = ParagraphStyle('CellR', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_RIGHT)
    cell_b = ParagraphStyle('CellB', parent=styles['Normal'], fontSize=9, leading=11, alignment=TA_RIGHT)
    detail_style = ParagraphStyle('Detail', parent=styles['Normal'], fontSize=6, leading=8, textColor=colors.HexColor('#888888'))

    col_widths = [45*mm, 30*mm, 30*mm, 35*mm, 46*mm]

    for cat_name, cat_items in grouped.items():
        # Category header
        cat_color = cat_colors.get(cat_name, colors.HexColor('#666666'))
        cat_style = ParagraphStyle('CatTitle', parent=styles['Normal'], fontSize=10, textColor=cat_color, spaceAfter=3, spaceBefore=8)
        elements.append(Paragraph(f"<b>{cat_name}</b> ({len(cat_items)} items)", cat_style))

        # Table
        table_data = [["Item", "In (Qntl)", "Out (Qntl)", "Available", "Details"]]
        for item in cat_items:
            avail = item.get('available', 0)
            avail_str = f"{avail} {item.get('unit', 'Qntl')}"
            table_data.append([
                Paragraph(f"<b>{item['name']}</b>", cell_style),
                Paragraph(f"{item.get('in_qty', 0)}", cell_r),
                Paragraph(f"{item.get('out_qty', 0)}", cell_r),
                Paragraph(f"<b>{avail_str}</b>", cell_b),
                Paragraph(item.get('details', ''), detail_style),
            ])

        tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
            # Color available based on positive/negative
            avail_val = cat_items[i-1].get('available', 0)
            if avail_val < 0:
                style_cmds.append(('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#DC2626')))
            else:
                style_cmds.append(('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#059669')))

        tbl.setStyle(TableStyle(style_cmds))
        elements.append(tbl)
        elements.append(Spacer(1, 8))

    # Footer
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#999999'), alignment=TA_CENTER, spaceBefore=15)
    elements.append(Paragraph(f"{company} - Stock Summary | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", footer_style))

    doc.build(elements)
    pdf_bytes = buf.getvalue()

    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=stock_summary.pdf"})


# ============ STOCK SUMMARY EXCEL EXPORT ============

@router.get("/stock-summary/export/excel")
async def export_stock_summary_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from fastapi.responses import Response
    import io
    data = await get_stock_summary(kms_year, season)
    items = data.get("items", [])

    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, COLORS, BORDER_THIN)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Summary"
    ncols = 6

    title = f"{company} - Stock Summary / स्टॉक सारांश"
    subtitle = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    style_excel_title(ws, title, ncols, subtitle)

    headers = ["Item", "Category", "In (Qntl)", "Out (Qntl)", "Available (Qntl)", "Details"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_excel_header_row(ws, 4, ncols)

    grouped = {}
    for item in items:
        cat = item.get('category', 'Other')
        if cat not in grouped: grouped[cat] = []
        grouped[cat].append(item)

    row = 5
    for cat_name, cat_items in grouped.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=cat_name)
        cell.font = Font(bold=True, size=10, color=COLORS['header_text'])
        cell.fill = PatternFill(start_color=COLORS['header_bg'], fill_type='solid')
        row += 1
        data_start = row

        for item in cat_items:
            avail = item.get('available', 0)
            vals = [item['name'], item.get('category', ''), item.get('in_qty', 0), item.get('out_qty', 0), avail, item.get('details', '')]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.border = BORDER_THIN
                if c in (3, 4, 5):
                    cell.alignment = Alignment(horizontal='right')
            row += 1

        if cat_items:
            style_excel_data_rows(ws, data_start, row - 1, ncols, headers)
        row += 1

    widths = [22, 14, 14, 14, 18, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=stock_summary.xlsx"})
