from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import round_amount
from pydantic import BaseModel, ConfigDict, Field
from utils.optimistic_lock import optimistic_update, stamp_version
from utils.audit import log_audit
import uuid
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils import export_helpers as _eh_default_font  # noqa: F401
from openpyxl.utils import get_column_letter
from utils.report_helper import get_columns, get_entry_row, get_total_row, get_excel_headers, get_pdf_headers, get_excel_widths, get_pdf_widths_mm, col_count

router = APIRouter()

# ============ BANK ACCOUNTS MANAGEMENT ============

@router.get("/bank-accounts")
async def get_bank_accounts():
    accounts = await db.bank_accounts.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return accounts

@router.post("/bank-accounts")
async def add_bank_account(request: Request):
    data = await request.json()
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Bank name is required")
    existing = await db.bank_accounts.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Bank already exists")
    doc = {"id": str(uuid.uuid4()), "name": name, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.bank_accounts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@router.delete("/bank-accounts/{bank_id}")
async def delete_bank_account(bank_id: str):
    result = await db.bank_accounts.delete_one({"id": bank_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bank account not found")
    return {"message": "Bank account deleted", "id": bank_id}


# ============ OWNER ACCOUNTS MANAGEMENT (Pvt / Drawing accounts) ============
# Owner accounts are tracked as PARTIES in the cashbook ledger (e.g. "Titu",
# "Mahesh"). Listing them centrally lets the frontend autocomplete owner names
# and auto-classify transactions with party_type="Owner" so the Party Ledger
# tab can produce a clean per-owner statement.

@router.get("/owner-accounts")
async def get_owner_accounts():
    accounts = await db.owner_accounts.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return accounts


@router.post("/owner-accounts")
async def add_owner_account(request: Request):
    data = await request.json()
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Owner account name is required")
    existing = await db.owner_accounts.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Owner account already exists")
    doc = {"id": str(uuid.uuid4()), "name": name, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.owner_accounts.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@router.delete("/owner-accounts/{owner_id}")
async def delete_owner_account(owner_id: str):
    result = await db.owner_accounts.delete_one({"id": owner_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Owner account not found")
    return {"message": "Owner account deleted", "id": owner_id}


@router.post("/owner-accounts/convert-from-ledger")
async def convert_ledger_to_owner_account(request: Request):
    """Convert an existing party/ledger (e.g. 'Titu') into an Owner Account.

    Strategy:
    1. Create the owner account if it doesn't exist
    2. Find all REAL cashbook txns (account in cash/bank) where category matches the name
    3. Switch each: account=owner, owner_name=<name>, FLIP txn_type
       (Owner accounting is inverted from cash/bank — see compute_owner_balances)
    4. Update auto_ledger pairs to match the flipped txn_type
    5. Returns counts so user can verify

    Body: {"name": "Titu", "dry_run": false}
    """
    data = await request.json()
    name = (data.get("name") or "").strip()
    dry_run = bool(data.get("dry_run", False))
    if not name:
        raise HTTPException(status_code=400, detail="Ledger name required")

    # 1. Ensure owner account exists
    existing_owner = await db.owner_accounts.find_one(
        {"name": {"$regex": f"^{name}$", "$options": "i"}}, {"_id": 0}
    )
    owner_id = existing_owner.get("id") if existing_owner else None

    # 2. Find matching real cashbook txns (NOT auto_ledger entries)
    txns = await db.cash_transactions.find(
        {
            "category": {"$regex": f"^{name}$", "$options": "i"},
            "account": {"$in": ["cash", "bank"]}
        },
        {"_id": 0, "id": 1, "txn_type": 1, "amount": 1, "date": 1, "category": 1, "account": 1}
    ).to_list(50000)

    total_amount = sum(t.get("amount", 0) for t in txns)
    cash_count = sum(1 for t in txns if t.get("account") == "cash")
    bank_count = sum(1 for t in txns if t.get("account") == "bank")

    preview = {
        "owner_already_exists": existing_owner is not None,
        "matching_txn_count": len(txns),
        "cash_txn_count": cash_count,
        "bank_txn_count": bank_count,
        "total_amount": round(total_amount, 2),
    }

    if dry_run:
        return {"success": True, "dry_run": True, "preview": preview}

    if not owner_id:
        new_doc = {"id": str(uuid.uuid4()), "name": name,
                   "created_at": datetime.now(timezone.utc).isoformat()}
        await db.owner_accounts.insert_one(new_doc.copy())
        owner_id = new_doc["id"]

    # 3. Flip + update each real txn AND its auto_ledger pair
    converted = 0
    now_iso = datetime.now(timezone.utc).isoformat()
    for txn in txns:
        old_type = txn.get("txn_type", "")
        new_type = "jama" if old_type == "nikasi" else "nikasi"
        txn_id = txn["id"]
        await db.cash_transactions.update_one(
            {"id": txn_id},
            {"$set": {"account": "owner", "owner_name": name,
                      "txn_type": new_type, "updated_at": now_iso}}
        )
        # Mirror flip on the auto_ledger pair (same prefix matching)
        await db.cash_transactions.update_many(
            {"reference": f"auto_ledger:{txn_id[:8]}"},
            {"$set": {"txn_type": new_type, "owner_name": name, "updated_at": now_iso}}
        )
        converted += 1

    return {
        "success": True,
        "owner_id": owner_id,
        "name": name,
        "converted": converted,
        "preview": preview,
    }


# ============ CASH BOOK / DAILY CASH & BANK REGISTER ============

class CashTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    account: str  # "cash", "bank", or "ledger"
    txn_type: str  # "jama" (credit/in) or "nikasi" (debit/out)
    category: str = ""  # Party name
    party_type: str = ""  # "Truck", "Agent", "Local Party", "Diesel", "Manual"
    description: str = ""
    amount: float = 0
    reference: str = ""
    bank_name: str = ""  # Which bank account (for bank transactions)
    owner_name: str = ""  # Which owner account (for account=='owner' txns)
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


@router.post("/cash-book")
async def add_cash_transaction(txn: CashTransaction, username: str = "", role: str = "", round_off: float = 0):
    from services.cashbook_service import (
        detect_party_type, backfill_party_type,
        create_auto_ledger_entry, process_diesel_auto_entry, process_pvt_paddy_auto_payment
    )
    txn_dict = txn.model_dump()
    txn_dict['created_by'] = username
    txn_dict['amount'] = round(txn_dict['amount'], 2)
    category = txn_dict.get('category', '').strip()

    # v104.44.97 — Generate default description BEFORE insert so auto_ledger
    # mirror entry uses the same description (prevents BP Sale "with-payments"
    # dedup mismatch which causes "double amount briefly visible until refresh").
    if not txn_dict.get('description') and category:
        acct = txn_dict.get('account', 'cash')
        ttype = txn_dict.get('txn_type', '')
        if acct == 'owner':
            owner = txn_dict.get('owner_name', 'Owner')
            txn_dict['description'] = (
                f"{owner} received from {category}" if ttype == 'jama'
                else f"{owner} paid to {category}"
            )
        else:
            txn_dict['description'] = (
                f"{acct.capitalize()} received from {category}" if ttype == 'jama'
                else f"{acct.capitalize()} payment to {category}"
            )
    
    # Auto-detect party_type if not provided
    if not txn_dict.get('party_type') and category:
        txn_dict['party_type'] = await detect_party_type(category)
        if txn_dict.get('party_type'):
            await backfill_party_type(category, txn_dict['party_type'])
    
    await db.cash_transactions.insert_one(stamp_version(txn_dict))
    await log_audit("cash_transactions", txn_dict["id"], "create", txn_dict.get("created_by", ""), new_data=txn_dict)
    txn_dict.pop('_id', None)
    
    # Auto-create double-entry ledger entry
    await create_auto_ledger_entry(txn_dict, round_off)
    
    # Auto-create diesel payment entry
    await process_diesel_auto_entry(txn_dict, round_off, username)
    
    # Auto-update private paddy payment
    await process_pvt_paddy_auto_payment(txn_dict)

    return txn_dict


@router.get("/cash-book")
async def get_cash_transactions(kms_year: Optional[str] = None, season: Optional[str] = None,
                                 account: Optional[str] = None, txn_type: Optional[str] = None,
                                 category: Optional[str] = None, party_type: Optional[str] = None,
                                 date_from: Optional[str] = None, date_to: Optional[str] = None,
                                 exclude_round_off: Optional[str] = None,
                                 page: int = 1, page_size: int = 200):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if txn_type: query["txn_type"] = txn_type
    # Owner-account parties are stored with account="owner" + owner_name=<owner>;
    # category holds the OTHER party. When frontend sends party_type="Owner",
    # match owner_name OR direct category=Owner txns (e.g. mill cash from Titu).
    # Exclude auto_ledger duplicates so each real txn shows exactly once.
    is_owner_query = bool(category and party_type == "Owner")
    if is_owner_query:
        query["$and"] = [
            {"$or": [
                {"owner_name": category, "account": "owner"},
                {"category": category, "party_type": "Owner", "account": {"$in": ["cash", "bank"]}},
            ]},
            {"reference": {"$not": {"$regex": "^auto_ledger:"}}},
        ]
    else:
        # Account filter ignored for owner ledger views (handled inside $or above)
        if account:
            query["account"] = account
        if category:
            query["category"] = category
        if party_type:
            query["party_type"] = party_type
    if exclude_round_off == "true" and not party_type:
        query["party_type"] = {"$ne": "Round Off"}
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        if date_q: query["date"] = date_q
    total_count = await db.cash_transactions.count_documents(query)
    # page_size=0 means return all (for summary/category filtering)
    if page_size <= 0:
        txns = await db.cash_transactions.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(50000)
        return {"transactions": txns, "total": total_count, "page": 1, "page_size": total_count, "total_pages": 1}
    if page < 1: page = 1
    skip = (page - 1) * page_size
    txns = await db.cash_transactions.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).skip(skip).limit(page_size).to_list(page_size)
    return {"transactions": txns, "total": total_count, "page": page, "page_size": page_size, "total_pages": max(1, (total_count + page_size - 1) // page_size)}


@router.get("/cash-book/summary")
async def get_cash_book_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    from services.cashbook_service import compute_account_totals, compute_bank_details, compute_opening_balances
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.cash_transactions.find(query, {"_id": 0}).to_list(10000)
    
    totals = compute_account_totals(txns)
    cash_in, cash_out = totals["cash_in"], totals["cash_out"]
    bank_in, bank_out = totals["bank_in"], totals["bank_out"]
    
    bank_accounts = await db.bank_accounts.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    bank_names = [b["name"] for b in bank_accounts]
    bank_details = compute_bank_details(totals["real_txns"], bank_names)
    
    opening_cash, opening_bank, opening_bank_details = await compute_opening_balances(kms_year)
    
    # Add opening balances per bank to bank_details
    for bn in bank_details:
        ob_val = opening_bank_details.get(bn, 0)
        bank_details[bn]["opening"] = ob_val
        bank_details[bn]["balance"] = round(ob_val + bank_details[bn]["in"] - bank_details[bn]["out"], 2)
    for bn, ob_val in opening_bank_details.items():
        if bn not in bank_details and (ob_val or 0) > 0:
            bank_details[bn] = {"in": 0, "out": 0, "opening": ob_val, "balance": ob_val}
    
    return {
        "opening_cash": opening_cash,
        "opening_bank": opening_bank,
        "opening_bank_details": opening_bank_details,
        "cash_in": round(cash_in, 2), "cash_out": round(cash_out, 2),
        "cash_balance": round(opening_cash + cash_in - cash_out, 2),
        "bank_in": round(bank_in, 2), "bank_out": round(bank_out, 2),
        "bank_balance": round(opening_bank + bank_in - bank_out, 2),
        "bank_details": bank_details,
        "total_balance": round((opening_cash + cash_in - cash_out) + (opening_bank + bank_in - bank_out), 2),
        "total_transactions": len(txns)
    }

@router.get("/cash-book/opening-balance")
async def get_opening_balance(kms_year: str):
    saved = await db.opening_balances.find_one({"kms_year": kms_year}, {"_id": 0})
    if saved:
        return {"cash": saved.get("cash", 0), "bank": saved.get("bank", 0), "bank_details": saved.get("bank_details", {}), "source": "manual"}
    parts = kms_year.split('-')
    if len(parts) == 2:
        try:
            prev_fy = f"{int(parts[0])-1}-{int(parts[1])-1}"
            prev_txns = await db.cash_transactions.find({"kms_year": prev_fy}, {"_id": 0}).to_list(10000)
            prev_real = [t for t in prev_txns if t.get('party_type') != 'Round Off']
            prev_cash_in = sum(t['amount'] for t in prev_real if t.get('account') == 'cash' and t.get('txn_type') == 'jama')
            prev_cash_out = sum(t['amount'] for t in prev_real if t.get('account') == 'cash' and t.get('txn_type') == 'nikasi')
            prev_bank_in = sum(t['amount'] for t in prev_real if t.get('account') == 'bank' and t.get('txn_type') == 'jama')
            prev_bank_out = sum(t['amount'] for t in prev_real if t.get('account') == 'bank' and t.get('txn_type') == 'nikasi')
            prev_ob = await db.opening_balances.find_one({"kms_year": prev_fy}, {"_id": 0})
            ob_cash = prev_ob.get("cash", 0) if prev_ob else 0
            ob_bank = prev_ob.get("bank", 0) if prev_ob else 0
            return {"cash": round(ob_cash + prev_cash_in - prev_cash_out, 2), "bank": round(ob_bank + prev_bank_in - prev_bank_out, 2), "bank_details": {}, "source": "auto"}
        except (ValueError, IndexError):
            pass
    return {"cash": 0, "bank": 0, "bank_details": {}, "source": "none"}

@router.put("/cash-book/opening-balance")
async def save_opening_balance(data: dict):
    kms_year = data.get("kms_year")
    if not kms_year:
        raise HTTPException(status_code=400, detail="kms_year is required")
    bank_details = data.get("bank_details", {})
    total_bank = sum(float(v) for v in bank_details.values()) if bank_details else float(data.get("bank", 0))
    doc = {
        "kms_year": kms_year,
        "cash": float(data.get("cash", 0)),
        "bank": round(total_bank, 2),
        "bank_details": bank_details,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.opening_balances.update_one({"kms_year": kms_year}, {"$set": doc}, upsert=True)
    return doc

@router.post("/cash-book/delete-bulk")
async def delete_cash_transactions_bulk(request: Request):
    body = await request.json()
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="No ids provided")
    # Revert hemali payments for any hemali cashbook entries being deleted
    hemali_txns = await db.cash_transactions.find(
        {"id": {"$in": ids}, "reference": {"$regex": "^hemali_payment:"}}, {"_id": 0}
    ).to_list(1000)
    for txn in hemali_txns:
        hemali_pid = txn["reference"].replace("hemali_payment:", "")
        await db.hemali_payments.update_one(
            {"id": hemali_pid},
            {"$set": {"status": "unpaid", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        await db.cash_transactions.delete_many({
            "reference": {"$in": [f"hemali_work:{hemali_pid}", f"hemali_paid:{hemali_pid}"]}
        })
        await db.local_party_accounts.delete_many({
            "reference": f"hemali_paid:{hemali_pid}"
        })
    result = await db.cash_transactions.delete_many({"id": {"$in": ids}})
    return {"message": f"{result.deleted_count} transactions deleted", "deleted": result.deleted_count}


@router.delete("/cash-book/{txn_id}")
async def delete_cash_transaction(txn_id: str, username: str = "", role: str = ""):
    from services.cashbook_service import revert_pvt_paddy_payment, revert_rice_sale_payment, revert_linked_payments, revert_hemali_payment
    from services.edit_lock import check_edit_lock
    txn = await db.cash_transactions.find_one({"id": txn_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    can_edit, message = await check_edit_lock(txn, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)

    await log_audit("cash_transactions", txn_id, "delete", username, old_data=txn)
    
    party_type = txn.get("party_type", "")

    # Revert linked payment amounts based on party_type
    if party_type == "Pvt Paddy Purchase":
        await revert_pvt_paddy_payment(txn)
    elif party_type == "Rice Sale":
        await revert_rice_sale_payment(txn)

    # Revert truck/agent/lease payment amounts
    await revert_linked_payments(txn)
    
    # Delete auto-created ledger entry
    await db.cash_transactions.delete_many({"reference": f"auto_ledger:{txn_id[:8]}"})

    # Revert hemali payment
    await revert_hemali_payment(txn)

    await db.cash_transactions.delete_one({"id": txn_id})
    return {"message": "Transaction deleted", "id": txn_id}


@router.put("/cash-book/{txn_id}")
async def update_cash_transaction(txn_id: str, request: Request, username: str = "", role: str = ""):
    from services.edit_lock import check_edit_lock
    body = await request.json()
    client_v = body.pop("_v", None)
    body.pop("_id", None)
    body.pop("id", None)
    old_txn = await db.cash_transactions.find_one({"id": txn_id}, {"_id": 0})
    if not old_txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    can_edit, message = await check_edit_lock(old_txn, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    body["updated_at"] = datetime.now(timezone.utc).isoformat()
    body["updated_by"] = username or body.get("updated_by", "")
    if "amount" in body:
        body["amount"] = round(float(body["amount"]), 2)
    await optimistic_update(db.cash_transactions, txn_id, body, client_v)
    if old_txn:
        await log_audit("cash_transactions", txn_id, "update", username, old_data=old_txn, new_data=body)
    # Update auto-created ledger entry too
    ledger_body = {k: v for k, v in body.items() if k not in ('account', 'reference', '_v')}
    # Keep same txn_type for auto-ledger (no reversal - party's khata matches direction)
    await db.cash_transactions.update_many({"reference": f"auto_ledger:{txn_id[:8]}"}, {"$set": ledger_body})
    updated = await db.cash_transactions.find_one({"id": txn_id}, {"_id": 0})
    return updated


@router.get("/cash-book/categories")
async def get_cash_book_categories():
    cats = await db.cash_book_categories.find({}, {"_id": 0}).to_list(500)
    return cats


@router.get("/cash-book/agent-names")
async def get_agent_names_for_cashbook(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Return mandi names from mandi_targets and unique agent names from entries for Cash Book suggestions"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    # Get mandi names from targets
    targets = await db.mandi_targets.find(query, {"_id": 0, "mandi_name": 1}).to_list(200)
    mandi_names = list(set(t["mandi_name"] for t in targets if t.get("mandi_name")))
    # Also get unique truck_no and agent_name from entries
    trucks = await db.mill_entries.distinct("truck_no", query) if query else await db.mill_entries.distinct("truck_no")
    agents = await db.mill_entries.distinct("agent_name", query) if query else await db.mill_entries.distinct("agent_name")
    return {"mandi_names": sorted(mandi_names), "truck_numbers": sorted([t for t in trucks if t]), "agent_names": sorted([a for a in agents if a])}


@router.post("/cash-book/categories")
async def add_cash_book_category(request: Request):
    data = await request.json()
    name = (data.get("name") or "").strip()
    cat_type = data.get("type", "")  # cash_jama, cash_nikasi, bank_jama, bank_nikasi
    if not name or not cat_type:
        raise HTTPException(status_code=400, detail="Name and type required")
    existing = await db.cash_book_categories.find_one({"name": name, "type": cat_type})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    cat = {"id": str(uuid.uuid4()), "name": name, "type": cat_type, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.cash_book_categories.insert_one(cat)
    cat.pop('_id', None)
    return cat


@router.delete("/cash-book/categories/{cat_id}")
async def delete_cash_book_category(cat_id: str):
    result = await db.cash_book_categories.delete_one({"id": cat_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted", "id": cat_id}


@router.get("/cash-book/party-summary")
async def get_party_summary(kms_year: Optional[str] = None, season: Optional[str] = None, party_type: Optional[str] = None):
    """Get Tally-style party summary - total jama, nikasi, outstanding per party"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if party_type: query["party_type"] = party_type
    
    txns = await db.cash_transactions.find(query, {"_id": 0}).to_list(100000)
    
    # Group by category (party name)
    party_map = {}
    for t in txns:
        cat = t.get("category", "").strip()
        if not cat: continue
        # Skip auto-ledger entries (duplicates with reversed txn_type)
        if "_ledger:" in (t.get("reference") or ""): continue
        if cat not in party_map:
            party_map[cat] = {"party_name": cat, "party_type": t.get("party_type", ""), "total_jama": 0, "total_nikasi": 0, "balance": 0, "txn_count": 0}
        if t.get("txn_type") == "jama":
            party_map[cat]["total_jama"] += t.get("amount", 0)
        else:
            party_map[cat]["total_nikasi"] += t.get("amount", 0)
        party_map[cat]["txn_count"] += 1
        # Update party_type if empty
        if not party_map[cat]["party_type"] and t.get("party_type"):
            party_map[cat]["party_type"] = t["party_type"]
    
    # === Add Owner Accounts as virtual parties (their statement comes from
    # transactions where account=="owner" and owner_name=<owner>) ===
    # IMPORTANT: From the Owner's ledger perspective the direction is FLIPPED:
    #   account=owner + nikasi (Owner paid mill's vendor) → Owner's contribution ↑ (JAMA)
    #   account=owner + jama   (Owner received from a party, mill paid via owner) → Owner withdrew (NIKASI)
    # This way Owner's ledger reads naturally: "kitna paisa Owner ne mill mein
    # daala vs kitna nikala", with running balance = mill ka karz Owner ki taraf.
    for t in txns:
        if t.get("account") != "owner":
            continue
        owner = (t.get("owner_name") or "").strip()
        if not owner:
            continue
        # Skip auto-ledger entries
        if "_ledger:" in (t.get("reference") or ""):
            continue
        if owner not in party_map:
            party_map[owner] = {"party_name": owner, "party_type": "Owner",
                                "total_jama": 0, "total_nikasi": 0,
                                "balance": 0, "txn_count": 0}
        else:
            # Force party_type to Owner if owner_name matches
            party_map[owner]["party_type"] = "Owner"
        if t.get("txn_type") == "nikasi":
            # Owner paid out for mill → CONTRIBUTION (jama in Owner's ledger)
            party_map[owner]["total_jama"] += t.get("amount", 0)
        else:
            # Owner received money via mill → WITHDRAWAL (nikasi from Owner's ledger)
            party_map[owner]["total_nikasi"] += t.get("amount", 0)
        party_map[owner]["txn_count"] += 1

    # Auto-detect party_type for parties with empty type by checking ALL transactions
    empty_type_parties = [cat for cat, p in party_map.items() if not p["party_type"]]
    if empty_type_parties:
        all_txns_for_types = await db.cash_transactions.find(
            {"category": {"$in": empty_type_parties}, "party_type": {"$ne": "", "$exists": True}},
            {"_id": 0, "category": 1, "party_type": 1}
        ).to_list(10000)
        type_lookup = {}
        for t in all_txns_for_types:
            cat = t.get("category", "").strip()
            if cat and t.get("party_type") and cat not in type_lookup:
                type_lookup[cat] = t["party_type"]
        
        # Also check other collections for party type detection
        for cat in empty_type_parties:
            if cat in type_lookup:
                party_map[cat]["party_type"] = type_lookup[cat]
                continue
            # Check private_paddy
            paddy_check = await db.private_paddy.find_one({"party_name": cat}, {"_id": 0, "party_name": 1})
            if paddy_check:
                party_map[cat]["party_type"] = "Pvt Paddy Purchase"
                continue
            # Check rice_sales
            rice_check = await db.rice_sales.find_one({"party_name": cat}, {"_id": 0, "party_name": 1})
            if rice_check:
                party_map[cat]["party_type"] = "Rice Sale"
                continue
            # Check local_party_accounts
            local_check = await db.local_party_accounts.find_one({"party_name": cat}, {"_id": 0, "party_name": 1})
            if local_check:
                party_map[cat]["party_type"] = "Local Party"
                continue
            # Check diesel_accounts
            diesel_check = await db.diesel_accounts.find_one({"pump_name": cat}, {"_id": 0, "pump_name": 1})
            if diesel_check:
                party_map[cat]["party_type"] = "Diesel"
                continue
    
    # Calculate balance and sort
    parties = []
    for p in party_map.values():
        p["total_jama"] = round(p["total_jama"], 2)
        p["total_nikasi"] = round(p["total_nikasi"], 2)
        p["balance"] = round(p["total_jama"] - p["total_nikasi"], 2)
        parties.append(p)
    
    parties.sort(key=lambda x: abs(x["balance"]), reverse=True)
    
    # Summary totals
    total_jama = round(sum(p["total_jama"] for p in parties), 2)
    total_nikasi = round(sum(p["total_nikasi"] for p in parties), 2)
    total_outstanding = round(sum(p["balance"] for p in parties if p["balance"] != 0), 2)
    settled_count = sum(1 for p in parties if p["balance"] == 0)
    pending_count = sum(1 for p in parties if p["balance"] != 0)
    
    return {
        "parties": parties,
        "summary": {
            "total_parties": len(parties),
            "settled_count": settled_count,
            "pending_count": pending_count,
            "total_jama": total_jama,
            "total_nikasi": total_nikasi,
            "total_outstanding": total_outstanding
        }
    }


@router.get("/cash-book/party-summary/pdf")
async def export_party_summary_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, party_type: Optional[str] = None, status: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from utils.export_helpers import get_pdf_styles
    from io import BytesIO
    
    result = await get_party_summary(kms_year, season, party_type)
    parties = result["parties"]
    if status == "settled": parties = [p for p in parties if p["balance"] == 0]
    elif status == "pending": parties = [p for p in parties if p["balance"] != 0]
    summary = result["summary"]
    
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = get_pdf_styles()
    elements = []
    
    # Company Header + Title
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    elements.append(Paragraph("Party Summary / पार्टी सारांश", styles['Title']))
    filter_text = ""
    if party_type: filter_text += f"Party Type: {party_type} | "
    if kms_year: filter_text += f"Year: {kms_year} | "
    if season: filter_text += f"Season: {season}"
    if filter_text: elements.append(Paragraph(filter_text, styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Summary cards
    sum_data = [
        ['Total Parties', 'Settled (0 Balance)', 'Pending', 'Total Jama', 'Total Nikasi', 'Outstanding'],
        [str(summary['total_parties']), str(summary['settled_count']), str(summary['pending_count']),
         f"Rs.{summary['total_jama']:,.2f}", f"Rs.{summary['total_nikasi']:,.2f}", f"Rs.{summary['total_outstanding']:,.2f}"]
    ]
    from utils.export_helpers import get_pdf_table_style
    t = RLTable(sum_data, colWidths=[80, 90, 60, 90, 90, 90])
    t.setStyle(TableStyle(get_pdf_table_style(len(sum_data))))
    elements.append(t); elements.append(Spacer(1, 16))
    
    # Party table
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT
    name_style = ParagraphStyle('name', fontName='FreeSans', fontSize=7, leading=8.5, alignment=TA_LEFT)
    
    data = [['#', 'Party Name', 'Party Type', 'Jama (Cr)', 'Nikasi (Dr)', 'Balance (Rs)', 'Txns', 'Status']]
    for i, p in enumerate(parties, 1):
        status = 'Settled' if p['balance'] == 0 else 'Pending'
        data.append([str(i), Paragraph(p['party_name'], name_style), p['party_type'], f"{p['total_jama']:,.2f}", f"{p['total_nikasi']:,.2f}",
                      f"{p['balance']:,.2f}", str(p['txn_count']), status])
    
    table = RLTable(data, colWidths=[25, 120, 70, 75, 75, 75, 35, 50], repeatRows=1)
    cols_info = [{'header': h} for h in data[0]]
    style_cmds = get_pdf_table_style(len(data), cols_info)
    style_cmds.extend([
        ('ALIGN', (3,0), (5,-1), 'RIGHT'), ('ALIGN', (6,0), (6,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ])
    table.setStyle(TableStyle(style_cmds))
    # Color rows based on status
    for i, p in enumerate(parties, 1):
        if p['balance'] == 0:
            table.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), ('BACKGROUND', (0,i), (-1,i), colors.HexColor('#f0fff4'))]))
        elif p['balance'] < 0:
            table.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), ('BACKGROUND', (0,i), (-1,i), colors.HexColor('#fff5f5'))]))
    
    elements.append(table)
    doc.build(elements)
    
    from starlette.responses import Response
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=party_summary_{datetime.now().strftime('%Y%m%d')}.pdf"})


@router.get("/cash-book/party-summary/excel")
async def export_party_summary_excel(kms_year: Optional[str] = None, season: Optional[str] = None, party_type: Optional[str] = None, status: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, COLORS, BORDER_THIN)
    
    result = await get_party_summary(kms_year, season, party_type)
    parties = result["parties"]
    if status == "settled": parties = [p for p in parties if p["balance"] == 0]
    elif status == "pending": parties = [p for p in parties if p["balance"] != 0]
    summary = result["summary"]
    
    wb = Workbook(); ws = wb.active; ws.title = "Party Summary"
    ncols = 8
    
    style_excel_title(ws, "Party Summary / पार्टी सारांश", ncols)
    
    # Summary
    ws.cell(row=4, column=1, value="Total Parties").font = Font(bold=True)
    ws.cell(row=4, column=2, value=summary['total_parties'])
    ws.cell(row=4, column=3, value="Settled").font = Font(bold=True)
    ws.cell(row=4, column=4, value=summary['settled_count'])
    ws.cell(row=4, column=5, value="Pending").font = Font(bold=True)
    ws.cell(row=4, column=6, value=summary['pending_count'])
    ws.cell(row=5, column=1, value="Total Outstanding").font = Font(bold=True, color=COLORS['nikasi_text'])
    ws.cell(row=5, column=2, value=summary['total_outstanding']).number_format = '#,##0.00'
    
    # Headers
    row = 7
    headers = ['#', 'Party Name', 'Party Type', 'Jama (Cr)', 'Nikasi (Dr)', 'Balance (Rs)', 'Transactions', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_excel_header_row(ws, row, ncols)
    
    data_start = row + 1
    for i, p in enumerate(parties, 1):
        row += 1
        status_val = 'Settled' if p['balance'] == 0 else 'Pending'
        for col, v in enumerate([i, p['party_name'], p['party_type'], p['total_jama'], p['total_nikasi'], p['balance'], p['txn_count'], status_val], 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER_THIN
            if col in [4,5,6]: c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal='right')
    
    if parties:
        style_excel_data_rows(ws, data_start, row, ncols, headers)
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    # 🎯 v104.44.9 — Apply consolidated multi-record polish (auto-filter + freeze + gridlines)
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from starlette.responses import Response
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=party_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"})



@router.post("/cash-book/fix-empty-party-types")
async def fix_empty_party_types():
    """Fix all cash_transactions entries with empty party_type by running auto-detection"""
    import re
    fixed = 0
    empty_entries = await db.cash_transactions.find(
        {"$or": [{"party_type": ""}, {"party_type": None}, {"party_type": {"$exists": False}}]},
        {"_id": 0, "id": 1, "category": 1}
    ).to_list(100000)
    
    # Group by category to avoid repeated lookups
    cats = {}
    for e in empty_entries:
        cat = (e.get("category") or "").strip()
        if cat:
            cats.setdefault(cat, []).append(e["id"])
    
    for cat, ids in cats.items():
        detected = ""
        cat_rgx = {"$regex": f"^{re.escape(cat)}$", "$options": "i"}
        cat_contains = {"$regex": re.escape(cat), "$options": "i"}
        
        # Check existing transactions with party_type
        existing = await db.cash_transactions.find_one(
            {"category": cat_rgx, "party_type": {"$nin": ["", None]}},
            {"_id": 0, "party_type": 1}
        )
        if existing and existing.get("party_type"):
            detected = existing["party_type"]
        else:
            if await db.private_paddy.find_one({"party_name": cat_rgx}):
                detected = "Pvt Paddy Purchase"
            elif await db.rice_sales.find_one({"party_name": cat_rgx}):
                detected = "Rice Sale"
            elif await db.diesel_accounts.find_one({"pump_name": cat_rgx}):
                detected = "Diesel"
            elif await db.local_party_accounts.find_one({"party_name": cat_rgx}):
                detected = "Local Party"
            elif await db.truck_payments.find_one({"truck_no": cat_rgx}):
                detected = "Truck"
            elif await db.mandi_targets.find_one({"mandi_name": cat_rgx}):
                detected = "Agent"
            elif await db.private_paddy.find_one({"party_name": cat_contains}):
                detected = "Pvt Paddy Purchase"
            elif await db.rice_sales.find_one({"party_name": cat_contains}):
                detected = "Rice Sale"
            elif await db.diesel_accounts.find_one({"pump_name": cat_contains}):
                detected = "Diesel"
            elif await db.local_party_accounts.find_one({"party_name": cat_contains}):
                detected = "Local Party"
            elif await db.private_payments.find_one({"party_name": cat_rgx}):
                detected = "Pvt Paddy Purchase"
            else:
                detected = "Cash Party"
        
        if detected:
            result = await db.cash_transactions.update_many(
                {"id": {"$in": ids}},
                {"$set": {"party_type": detected}}
            )
            fixed += result.modified_count
    
    return {"success": True, "fixed_count": fixed, "categories_processed": len(cats)}



@router.post("/cash-book/fix-auto-ledger-direction")
async def fix_auto_ledger_direction():
    """Fix existing auto_ledger entries that had reversed txn_type.
    Old logic: Cash Jama → Ledger Nikasi (wrong)
    New logic: Cash Jama → Ledger Jama (correct - party's khata matches direction)
    This flips all auto_ledger entries to match the original cash/bank entry's txn_type."""
    fixed = 0
    auto_ledger_entries = await db.cash_transactions.find(
        {"reference": {"$regex": "^auto_ledger:"}},
        {"_id": 0, "id": 1, "reference": 1, "txn_type": 1}
    ).to_list(100000)
    
    for entry in auto_ledger_entries:
        # Find the original cash/bank entry by matching the reference prefix
        ref = entry.get("reference", "")
        orig_id_prefix = ref.replace("auto_ledger:", "")
        if not orig_id_prefix:
            continue
        # Find original entry
        original = await db.cash_transactions.find_one(
            {"id": {"$regex": f"^{orig_id_prefix}"}, "account": {"$in": ["cash", "bank"]}},
            {"_id": 0, "txn_type": 1}
        )
        if original and original.get("txn_type") != entry.get("txn_type"):
            # The auto_ledger has reversed txn_type - fix it to match original
            await db.cash_transactions.update_one(
                {"id": entry["id"]},
                {"$set": {"txn_type": original["txn_type"]}}
            )
            fixed += 1
    
    return {"success": True, "fixed_count": fixed, "total_auto_ledger": len(auto_ledger_entries)}

@router.post("/cash-book/cleanup-round-off-entries")
async def cleanup_round_off_entries():
    """Delete all separate round_off entries from cash_transactions.
    Round off info is preserved in the main transaction's description."""
    result = await db.cash_transactions.delete_many({
        "$or": [
            {"party_type": "Round Off"},
            {"category": "Round Off"},
            {"reference": {"$regex": "^round_off:"}},
        ]
    })
    return {"success": True, "deleted_count": result.deleted_count}

@router.post("/cash-book/auto-fix")
async def auto_fix_all():
    """Master auto-fix: runs on every app startup to fix ALL data inconsistencies.
    Idempotent - safe to run multiple times."""
    fixes = {"auto_ledger_direction": 0, "round_off_cleaned": 0, "pvt_jama_created": 0, "duplicate_removed": 0}

    # 1. Fix auto_ledger direction (jama/nikasi should match original)
    auto_ledgers = await db.cash_transactions.find(
        {"reference": {"$regex": "^auto_ledger:"}}, {"_id": 0, "id": 1, "reference": 1, "txn_type": 1}
    ).to_list(100000)
    for entry in auto_ledgers:
        prefix = entry.get("reference", "").replace("auto_ledger:", "")
        if not prefix:
            continue
        original = await db.cash_transactions.find_one(
            {"id": {"$regex": f"^{prefix}"}, "account": {"$in": ["cash", "bank"]}}, {"_id": 0, "txn_type": 1}
        )
        if original and original.get("txn_type") != entry.get("txn_type"):
            await db.cash_transactions.update_one({"id": entry["id"]}, {"$set": {"txn_type": original["txn_type"]}})
            fixes["auto_ledger_direction"] += 1

    # 2. Clean up round_off entries
    ro = await db.cash_transactions.delete_many({
        "$or": [{"party_type": "Round Off"}, {"category": "Round Off"}, {"reference": {"$regex": "^round_off:"}}]
    })
    fixes["round_off_cleaned"] = ro.deleted_count

    # 3. Create missing pvt_party_jama entries for private paddy purchases (including agent_extra)
    pvt_entries = await db.private_paddy.find(
        {}, {"_id": 0}
    ).to_list(100000)
    for pvt in pvt_entries:
        total_amt = float(pvt.get("total_amount", 0) or 0)
        if total_amt <= 0:
            continue
        entry_id = pvt.get("id", "")
        if not entry_id:
            continue
        # Fix missing qntl/final_qntl fields for agent_extra entries
        if pvt.get("source") == "agent_extra" and not pvt.get("final_qntl") and pvt.get("quantity_qntl"):
            qty = float(pvt.get("quantity_qntl", 0))
            await db.private_paddy.update_one(
                {"id": entry_id},
                {"$set": {
                    "final_qntl": round(qty, 2),
                    "qntl": round(qty, 2),
                    "kg": round(qty * 100, 2),
                    "balance": round(total_amt - float(pvt.get("paid_amount", 0) or 0), 2)
                }}
            )
            fixes["agent_extra_fields_fixed"] = fixes.get("agent_extra_fields_fixed", 0) + 1
        # Fix empty season
        if not pvt.get("season"):
            await db.private_paddy.update_one({"id": entry_id}, {"$set": {"season": "Kharif"}})
            pvt["season"] = "Kharif"
            fixes["season_fixed"] = fixes.get("season_fixed", 0) + 1
        # Check if jama entry already exists
        existing = await db.cash_transactions.find_one(
            {"reference": f"pvt_party_jama:{entry_id[:8]}"}, {"_id": 0, "id": 1}
        )
        if not existing:
            party = pvt.get("party_name", "") or "Pvt Paddy"
            qntl = pvt.get("qntl", 0) or pvt.get("kg", 0) / 100 if pvt.get("kg") else 0
            rate = pvt.get("rate_per_qntl", 0) or pvt.get("rate", 0) or 0
            desc = f"Paddy Purchase: {party} - {qntl}Q @ Rs.{rate}/Q = Rs.{total_amt}" if qntl and rate else f"Paddy Purchase: {party} - Rs.{total_amt}"
            await db.cash_transactions.insert_one({
                "id": str(uuid.uuid4()), "date": pvt.get("date", ""),
                "account": "ledger", "txn_type": "jama",
                "category": party, "party_type": "Pvt Paddy Purchase",
                "description": desc, "amount": round_amount(total_amt), "bank_name": "",
                "reference": f"pvt_party_jama:{entry_id[:8]}",
                "kms_year": pvt.get("kms_year", ""), "season": pvt.get("season", "") or "Kharif",
                "created_by": "auto-fix", "linked_entry_id": entry_id,
                "created_at": pvt.get("created_at", ""), "updated_at": pvt.get("updated_at", ""),
            })
            fixes["pvt_jama_created"] += 1

    # 3b. Fix existing pvt_party_jama entries: ensure they are 'ledger' (NOT 'cash')
    # Paddy purchase = liability, not cash movement
    result = await db.cash_transactions.update_many(
        {"reference": {"$regex": "^pvt_party_jama"}, "account": "cash"},
        {"$set": {"account": "ledger"}}
    )
    if result.modified_count > 0:
        fixes["pvt_jama_account_fixed"] = result.modified_count

    # 3c. Remove duplicate pvt_party_jama_ledger entries (leftover from old double-entry logic)
    dupe_ledger = await db.cash_transactions.find(
        {"reference": {"$regex": "^pvt_party_jama_ledger:"}}, {"_id": 0, "id": 1}
    ).to_list(100000)
    if dupe_ledger:
        await db.cash_transactions.delete_many({"reference": {"$regex": "^pvt_party_jama_ledger:"}})
        fixes["duplicate_ledger_removed"] = len(dupe_ledger)

    # 3d. Fix empty season in cash_transactions for pvt_party_jama entries
    result2 = await db.cash_transactions.update_many(
        {"reference": {"$regex": "^pvt_party_jama:"}, "season": {"$in": ["", None]}},
        {"$set": {"season": "Kharif"}}
    )
    if result2.modified_count > 0:
        fixes["cash_txn_season_fixed"] = result2.modified_count

    # 4. Remove duplicate ledger entries (same reference, same amount, same date)
    all_ledger = await db.cash_transactions.find(
        {"account": "ledger"}, {"_id": 0, "id": 1, "reference": 1, "amount": 1, "date": 1, "category": 1}
    ).to_list(100000)
    seen = set()
    for t in all_ledger:
        key = f"{t.get('reference','')}|{t.get('amount',0)}|{t.get('date','')}|{t.get('category','')}"
        if key in seen and t.get("reference", ""):
            await db.cash_transactions.delete_one({"id": t["id"]})
            fixes["duplicate_removed"] += 1
        else:
            seen.add(key)

    # 5. Clean orphaned pvt_party_jama entries (private_paddy was deleted but cash_transactions remain)
    pvt_refs = await db.cash_transactions.find(
        {"reference": {"$regex": "^pvt_party_jama"}}, {"_id": 0, "id": 1, "linked_entry_id": 1, "reference": 1}
    ).to_list(100000)
    pvt_ids = set(e.get("id", "") for e in pvt_entries)
    for ref_entry in pvt_refs:
        linked_id = ref_entry.get("linked_entry_id", "")
        if linked_id and linked_id not in pvt_ids:
            await db.cash_transactions.delete_one({"id": ref_entry["id"]})
            fixes["orphan_cleaned"] = fixes.get("orphan_cleaned", 0) + 1

    # 6. Fix duplicate party names in cash_transactions
    # e.g. "Kridha (Kesinga) - Kesinga" should be merged into "Kridha (Kesinga)"
    import re as _re
    all_categories = await db.cash_transactions.distinct("category")
    all_categories = [c for c in all_categories if c and c.strip()]
    merge_map = {}  # long_name -> short_name
    for cat in all_categories:
        if " - " not in cat:
            continue
        parts = cat.rsplit(" - ", 1)
        base_name = parts[0].strip()
        suffix = parts[1].strip()
        # If the suffix is already contained in the base name (case-insensitive), it's a duplicate
        if suffix and suffix.lower() in base_name.lower() and base_name in all_categories:
            merge_map[cat] = base_name
    for long_name, short_name in merge_map.items():
        result = await db.cash_transactions.update_many(
            {"category": long_name},
            {"$set": {"category": short_name}}
        )
        if result.modified_count > 0:
            fixes["duplicate_party_merged"] = fixes.get("duplicate_party_merged", 0) + result.modified_count

    # 7. Clean orphaned auto_ledger entries (original cash entry was deleted)
    auto_ledger_all = await db.cash_transactions.find(
        {"reference": {"$regex": "^auto_ledger:"}}, {"_id": 0, "id": 1, "reference": 1}
    ).to_list(100000)
    for al_entry in auto_ledger_all:
        prefix = al_entry.get("reference", "").replace("auto_ledger:", "")
        if not prefix:
            continue
        original = await db.cash_transactions.find_one(
            {"id": {"$regex": f"^{prefix}"}, "account": {"$in": ["cash", "bank"]}}, {"_id": 0, "id": 1}
        )
        if not original:
            await db.cash_transactions.delete_one({"id": al_entry["id"]})
            fixes["orphan_auto_ledger_cleaned"] = fixes.get("orphan_auto_ledger_cleaned", 0) + 1

    # 8. Recalculate paid_amount/balance/payment_status for all private_paddy entries
    pvt_entries_fresh = await db.private_paddy.find({}, {"_id": 0}).to_list(100000)
    for pvt in pvt_entries_fresh:
        entry_id = pvt.get("id", "")
        if not entry_id:
            continue
        total_amt = float(pvt.get("total_amount", 0) or 0)
        if total_amt <= 0:
            continue
        # Sum all payment sources
        pay_sum = 0
        # a. private_payments
        payments = await db.private_payments.find(
            {"ref_id": entry_id, "ref_type": "paddy_purchase"}, {"_id": 0, "amount": 1, "round_off": 1}
        ).to_list(1000)
        for p in payments:
            pay_sum += float(p.get("amount", 0) or 0) + float(p.get("round_off", 0) or 0)
        # b. advance entries (pvt_paddy_adv)
        adv = await db.cash_transactions.find(
            {"linked_entry_id": entry_id, "reference": {"$regex": "^pvt_paddy_adv:"}, "account": "cash"},
            {"_id": 0, "amount": 1}
        ).to_list(100)
        for a in adv:
            pay_sum += float(a.get("amount", 0) or 0)
        # c. mark-paid entries
        mark = await db.cash_transactions.find(
            {"reference": {"$regex": f"^mark_paid:{entry_id[:8]}"}, "account": "cash"},
            {"_id": 0, "amount": 1}
        ).to_list(100)
        for m in mark:
            pay_sum += float(m.get("amount", 0) or 0)
        # d. manual cashbook entries
        manual = await db.cash_transactions.find(
            {"cashbook_pvt_linked": entry_id, "account": {"$in": ["cash", "bank"]}},
            {"_id": 0, "amount": 1}
        ).to_list(100)
        for mc in manual:
            pay_sum += float(mc.get("amount", 0) or 0)
        pay_sum = round(pay_sum, 2)
        stored_paid = round(float(pvt.get("paid_amount", 0) or 0), 2)
        if abs(pay_sum - stored_paid) > 0.5:
            new_balance = round(total_amt - pay_sum, 2)
            new_status = "paid" if pay_sum >= total_amt else "pending"
            await db.private_paddy.update_one(
                {"id": entry_id},
                {"$set": {"paid_amount": pay_sum, "balance": new_balance, "payment_status": new_status}}
            )
            fixes["paid_amount_recalculated"] = fixes.get("paid_amount_recalculated", 0) + 1

    # 9. Clean orphaned private_payments (ref_id points to non-existent entry)
    all_pvt_ids = set(p.get("id", "") for p in pvt_entries_fresh)
    all_payments = await db.private_payments.find({}, {"_id": 0, "id": 1, "ref_id": 1}).to_list(100000)
    for pay in all_payments:
        if pay.get("ref_id") and pay["ref_id"] not in all_pvt_ids:
            # Check rice_sales too
            rice = await db.rice_sales.find_one({"id": pay["ref_id"]}, {"_id": 0, "id": 1})
            if not rice:
                await db.private_payments.delete_one({"id": pay["id"]})
                fixes["orphan_payments_cleaned"] = fixes.get("orphan_payments_cleaned", 0) + 1

    total = sum(fixes.values())
    return {"success": True, "total_fixes": total, "details": fixes}

@router.post("/cash-book/migrate-ledger-entries")
async def migrate_ledger_entries():
    """Migrate old local_party, truck, agent, diesel entries to cash_transactions"""
    migrated = {"local_party_debit": 0, "local_party_payment": 0, "diesel_payment": 0, "total": 0}
    
    # 1. Migrate local_party debit entries (purchase side - Jama)
    lp_debits = await db.local_party_accounts.find({"txn_type": "debit"}, {"_id": 0}).to_list(50000)
    for t in lp_debits:
        existing = await db.cash_transactions.find_one({"linked_local_party_id": t["id"], "txn_type": "jama"})
        if not existing:
            cb = {
                "id": str(uuid.uuid4()), "date": t.get("date", ""),
                "account": "ledger", "txn_type": "jama",
                "category": t.get("party_name", ""), "party_type": "Local Party",
                "description": f"Purchase: {t.get('party_name','')} - {t.get('description','')} Rs.{t.get('amount',0)}",
                "amount": round_amount(t.get("amount", 0)),
                "reference": f"lp_migrate:{t['id'][:8]}",
                "kms_year": t.get("kms_year", ""), "season": t.get("season", ""),
                "created_by": "migration", "linked_local_party_id": t["id"],
                "created_at": t.get("created_at", datetime.now(timezone.utc).isoformat()),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(cb)
            migrated["local_party_debit"] += 1
    
    # 2. Migrate local_party payment entries that don't have linked cash_transactions
    lp_payments = await db.local_party_accounts.find({"txn_type": "payment"}, {"_id": 0}).to_list(50000)
    for t in lp_payments:
        existing = await db.cash_transactions.find_one({"linked_local_party_id": t["id"], "txn_type": "nikasi"})
        if not existing:
            cb = {
                "id": str(uuid.uuid4()), "date": t.get("date", ""),
                "account": "cash", "txn_type": "nikasi",
                "category": t.get("party_name", ""), "party_type": "Local Party",
                "description": f"Local Party Payment: {t.get('party_name','')} - Rs.{t.get('amount',0)}",
                "amount": round_amount(t.get("amount", 0)),
                "reference": f"lp_pay_migrate:{t['id'][:8]}",
                "kms_year": t.get("kms_year", ""), "season": t.get("season", ""),
                "created_by": "migration", "linked_local_party_id": t["id"],
                "created_at": t.get("created_at", datetime.now(timezone.utc).isoformat()),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(cb)
            migrated["local_party_payment"] += 1
    
    # 3. Migrate diesel payment entries that don't have linked cash_transactions
    diesel_payments = await db.diesel_accounts.find({"txn_type": "payment"}, {"_id": 0}).to_list(50000)
    for t in diesel_payments:
        existing = await db.cash_transactions.find_one({"linked_diesel_payment_id": t["id"]})
        if not existing:
            pump = await db.diesel_pumps.find_one({"id": t.get("pump_id")}, {"_id": 0})
            pump_name = pump["name"] if pump else t.get("pump_id", "")
            cb = {
                "id": str(uuid.uuid4()), "date": t.get("date", ""),
                "account": "cash", "txn_type": "nikasi",
                "category": pump_name, "party_type": "Diesel",
                "description": f"Diesel Payment: {pump_name} - Rs.{t.get('amount',0)}",
                "amount": round_amount(t.get("amount", 0)),
                "reference": f"diesel_migrate:{t['id'][:8]}",
                "kms_year": t.get("kms_year", ""), "season": t.get("season", ""),
                "created_by": "migration", "linked_diesel_payment_id": t["id"],
                "created_at": t.get("created_at", datetime.now(timezone.utc).isoformat()),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(cb)
            migrated["diesel_payment"] += 1
    
    # 4. Migrate truck entries (Jama - paddy delivered by truck, we owe them)
    migrated["truck_jama"] = 0
    truck_entries = await db.mill_entries.find({}, {"_id": 0}).to_list(50000)
    truck_payments_col = await db.truck_payments.find({}, {"_id": 0}).to_list(50000)
    tp_map = {p.get("entry_id"): p for p in truck_payments_col}
    for entry in truck_entries:
        entry_id = entry.get("id", "")
        truck_no = entry.get("truck_no", "")
        if not truck_no: continue
        # Check if Jama entry already exists
        existing = await db.cash_transactions.find_one({"reference": {"$regex": f"truck_entry:{entry_id[:8]}"}})
        
        payment_doc = tp_map.get(entry_id, {})
        rate = payment_doc.get("rate_per_qntl", 0) if payment_doc else 0
        final_qntl = round(entry.get("qntl", 0), 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        # Calculate gross amount - use rate*qntl if rate exists, else use paid + remaining + advances
        if rate > 0:
            gross_amount = round(final_qntl * rate, 2)
        elif payment_doc:
            # Total owed = paid_amount + remaining + cash_taken + diesel_taken
            gross_amount = round(
                payment_doc.get("paid_amount", 0) + payment_doc.get("remaining", 0) + cash_taken + diesel_taken, 2)
        else:
            gross_amount = 0
        
        if not existing and gross_amount > 0:
            cb = {
                "id": str(uuid.uuid4()), "date": entry.get("date", ""),
                "account": "ledger", "txn_type": "jama",
                "category": truck_no, "party_type": "Truck",
                "description": f"Truck Entry: {truck_no} - {final_qntl}Q @ Rs.{rate}/Q = Rs.{gross_amount}",
                "amount": round_amount(gross_amount), "reference": f"truck_entry:{entry_id[:8]}",
                "kms_year": entry.get("kms_year", ""), "season": entry.get("season", ""),
                "created_by": "migration",
                "created_at": entry.get("created_at", datetime.now(timezone.utc).isoformat()),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(cb)
            migrated["truck_jama"] += 1
        
        # Also add deductions (cash_taken, diesel_taken) as Nikasi
        if cash_taken > 0:
            existing_ded = await db.cash_transactions.find_one({"reference": f"truck_cash_ded:{entry_id[:8]}"})
            if not existing_ded:
                cb_ded = {
                    "id": str(uuid.uuid4()), "date": entry.get("date", ""),
                    "account": "cash", "txn_type": "nikasi",
                    "category": truck_no, "party_type": "Truck",
                    "description": f"Truck Cash Taken: {truck_no} - Rs.{cash_taken}",
                    "amount": round_amount(cash_taken), "reference": f"truck_cash_ded:{entry_id[:8]}",
                    "kms_year": entry.get("kms_year", ""), "season": entry.get("season", ""),
                    "created_by": "migration",
                    "created_at": entry.get("created_at", datetime.now(timezone.utc).isoformat()),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                await db.cash_transactions.insert_one(cb_ded)
        
        # Diesel deduction entry
        if diesel_taken > 0:
            existing_dd = await db.cash_transactions.find_one({"reference": f"truck_diesel_ded:{entry_id[:8]}"})
            if not existing_dd:
                cb_dd = {
                    "id": str(uuid.uuid4()), "date": entry.get("date", ""),
                    "account": "cash", "txn_type": "nikasi",
                    "category": truck_no, "party_type": "Truck",
                    "description": f"Truck Diesel Advance: {truck_no} - Rs.{diesel_taken}",
                    "amount": round_amount(diesel_taken), "reference": f"truck_diesel_ded:{entry_id[:8]}",
                    "kms_year": entry.get("kms_year", ""), "season": entry.get("season", ""),
                    "created_by": "migration",
                    "created_at": entry.get("created_at", datetime.now(timezone.utc).isoformat()),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                await db.cash_transactions.insert_one(cb_dd)
    
    # 5. Migrate agent entries (Jama - agent commission, we owe them)
    migrated["agent_jama"] = 0
    mandi_targets = await db.mandi_targets.find({}, {"_id": 0}).to_list(50000)
    mandi_map = {m.get("mandi_name"): m for m in mandi_targets}
    # Calculate agent commission from mill_entries
    agent_entries = {}
    for entry in truck_entries:
        mandi = entry.get("mandi_name", "")
        if not mandi: continue
        if mandi not in agent_entries:
            agent_entries[mandi] = {"total_qntl": 0, "entries": []}
        agent_entries[mandi]["total_qntl"] += entry.get("qntl", 0)
        agent_entries[mandi]["entries"].append(entry)
    
    for mandi, data in agent_entries.items():
        existing = await db.cash_transactions.find_one({"reference": {"$regex": f"agent_comm:{mandi[:12]}"}})
        if existing: continue
        target = mandi_map.get(mandi, {})
        base_rate = target.get("base_rate", 0)
        comm_amount = round(data["total_qntl"] * base_rate, 2)
        if comm_amount <= 0: continue
        
        last_entry = data["entries"][-1]
        cb = {
            "id": str(uuid.uuid4()), "date": last_entry.get("date", ""),
            "account": "ledger", "txn_type": "jama",
            "category": mandi, "party_type": "Agent",
            "description": f"Agent Commission: {mandi} - {data['total_qntl']}Q @ Rs.{base_rate}/Q",
            "amount": comm_amount, "reference": f"agent_comm:{mandi[:12]}",
            "kms_year": last_entry.get("kms_year", ""), "season": last_entry.get("season", ""),
            "created_by": "migration",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(cb)
        migrated["agent_jama"] += 1
    
    # 6. Update existing cash_transactions that have old-style categories
    # Extract party name from description and update category
    import re
    old_style = await db.cash_transactions.find(
        {"category": {"$in": ["Local Party Payment", "Truck Payment", "Agent Payment", "Diesel Payment"]}},
        {"_id": 0}
    ).to_list(50000)
    for t in old_style:
        party_name = t.get("category", "")
        pt = ""
        desc = t.get("description", "")
        if t["category"] == "Local Party Payment":
            pt = "Local Party"
            m = re.search(r"Local Party Payment: (.+?) -", desc)
            if m: party_name = m.group(1).strip()
        elif t["category"] == "Truck Payment":
            pt = "Truck"
            m = re.search(r"Truck Payment: (.+?)(?:\s*-|\s*\()", desc)
            if m: party_name = m.group(1).strip()
        elif t["category"] == "Agent Payment":
            pt = "Agent"
            m = re.search(r"Agent Payment: (.+?)(?:\s*-|\s*\()", desc)
            if m: party_name = m.group(1).strip()
        elif t["category"] == "Diesel Payment":
            pt = "Diesel"
            m = re.search(r"Diesel Payment: (.+?) -", desc)
            if m: party_name = m.group(1).strip()
        await db.cash_transactions.update_one(
            {"id": t["id"]},
            {"$set": {"category": party_name, "party_type": pt}}
        )
    
    # Also set party_type for entries that have it empty
    await db.cash_transactions.update_many(
        {"party_type": {"$exists": False}},
        {"$set": {"party_type": ""}}
    )
    
    migrated["old_categories_fixed"] = len(old_style)
    migrated["total"] = migrated["local_party_debit"] + migrated["local_party_payment"] + migrated["diesel_payment"] + migrated["truck_jama"] + migrated["agent_jama"] + len(old_style)
    return {"success": True, "migrated": migrated}


@router.get("/cash-book/excel")
async def export_cash_book_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                                  account: Optional[str] = None, txn_type: Optional[str] = None,
                                  category: Optional[str] = None, party_type: Optional[str] = None,
                                  date_from: Optional[str] = None, date_to: Optional[str] = None):
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, style_excel_header_row, 
        style_excel_data_rows, style_excel_total_row, style_excel_summary_header, COLORS)
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if txn_type: query["txn_type"] = txn_type
    # Owner-account parties: combine account=owner + cash/bank txns with category=<owner>.
    is_owner_view = bool(category and party_type == "Owner")
    if is_owner_view:
        query["$and"] = [
            {"$or": [
                {"owner_name": category, "account": "owner"},
                {"category": category, "party_type": "Owner", "account": {"$in": ["cash", "bank"]}},
            ]},
            {"reference": {"$not": {"$regex": "^auto_ledger:"}}},
        ]
    else:
        if account:
            query["account"] = account
        if category:
            query["category"] = category
        if party_type:
            query["party_type"] = party_type
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        query["date"] = date_q
    txns = await db.cash_transactions.find(query, {"_id": 0}).sort([("date", 1), ("created_at", 1)]).to_list(10000)
    # Owner ledger view → flip txn_type for display ONLY for account=owner entries
    # (cash/bank entries with category=<owner> are already in Owner perspective).
    if is_owner_view:
        for t in txns:
            if t.get("account") == "owner":
                t["txn_type"] = "jama" if t.get("txn_type") == "nikasi" else "nikasi"
    summary = await get_cash_book_summary(kms_year=kms_year, season=season)
    
    run_bal = 0
    rows = []
    for t in txns:
        jama = t['amount'] if t['txn_type'] == 'jama' else 0
        nikasi = t['amount'] if t['txn_type'] == 'nikasi' else 0
        run_bal += jama - nikasi
        rows.append({
            "date": t.get("date", ""),
            "account_label": "Ledger" if t.get("account") == "ledger" else ("Cash" if t.get("account") == "cash" else ("Owner" if t.get("account") == "owner" else "Bank")),
            "type_label": "Jama" if t.get("txn_type") == "jama" else "Nikasi",
            "category": t.get("category", ""),
            "party_type": t.get("party_type", ""),
            "description": t.get("description", ""),
            "jama": jama if jama > 0 else "",
            "nikasi": nikasi if nikasi > 0 else "",
            "balance": round(run_bal, 2),
            "reference": t.get("reference", ""),
        })
    
    cols = get_columns("cashbook_report")
    ncols = col_count(cols)
    headers = get_excel_headers(cols)
    widths = get_excel_widths(cols)
    
    wb = Workbook(); ws = wb.active; ws.title = "Cash Book"
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title section
    title = "Daily Cash Book / रोज़नामचा"
    if category: title += f" - {category}"
    if kms_year: title += f" | FY {kms_year}"
    subtitle = ""
    if date_from or date_to:
        date_parts = []
        if date_from: date_parts.append(f"From: {date_from}")
        if date_to: date_parts.append(f"To: {date_to}")
        subtitle = " | ".join(date_parts)
    style_excel_title(ws, title, ncols, subtitle)

    # Transactions section (Summary table removed - not useful for party ledger views)
    row_num = 4
    ws.cell(row=row_num, column=1, value="Transactions / लेनदेन").font = Font(bold=True, size=11, color=COLORS['title_text'])
    ws.cell(row=row_num, column=ncols, value=f"{len(rows)} entries").font = Font(size=9, italic=True, color='888888')
    ws.cell(row=row_num, column=ncols).alignment = Alignment(horizontal='right')
    row_num += 1
    
    # Header row
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row_num, column=col_idx, value=h)
    style_excel_header_row(ws, row_num, ncols)
    data_start = row_num + 1
    row_num += 1
    
    for r in rows:
        vals = get_entry_row(r, cols)
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col_idx, value=v)
            if cols[col_idx-1]["align"] == "right": c.alignment = Alignment(horizontal='right')
            if cols[col_idx-1]["type"] == "number" and isinstance(v, (int, float)): c.number_format = '#,##0.00'
        row_num += 1
    
    # Style data rows with color coding
    style_excel_data_rows(ws, data_start, row_num - 1, ncols, headers)
    
    # Total row
    totals = {
        "total_jama": round(sum(t['amount'] for t in txns if t['txn_type'] == 'jama'), 2),
        "total_nikasi": round(sum(t['amount'] for t in txns if t['txn_type'] == 'nikasi'), 2),
        "closing_balance": round(run_bal, 2),
    }
    total_vals = get_total_row(totals, cols)
    ws.cell(row=row_num, column=1, value="TOTAL / कुल")
    for col_idx, val in enumerate(total_vals, 1):
        if val is not None:
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.alignment = Alignment(horizontal='right')
    style_excel_total_row(ws, row_num, ncols)
    
    # ===== Beautiful single-line teal summary banner =====
    from utils.export_helpers import add_excel_summary_banner, fmt_inr
    net = totals["total_jama"] - totals["total_nikasi"]
    sum_stats = [
        {'label': 'Total Entries', 'value': str(len(rows))},
        {'label': 'Total Jama', 'value': fmt_inr(totals["total_jama"])},
        {'label': 'Total Nikasi', 'value': fmt_inr(totals["total_nikasi"])},
        {'label': 'Net Movement', 'value': fmt_inr(net)},
        {'label': 'Closing Balance', 'value': fmt_inr(totals["closing_balance"])},
    ]
    add_excel_summary_banner(ws, row_num + 2, ncols, sum_stats)

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # 🎯 v104.44.9 — Apply consolidated multi-record polish (auto-filter + freeze + gridlines)
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cash_book_{datetime.now().strftime('%Y%m%d')}.xlsx"})


async def _generate_cash_book_pdf_bytes(kms_year=None, season=None, account=None, txn_type=None,
                                        category=None, party_type=None, date_from=None, date_to=None):
    """Generate cash book PDF bytes - shared between download endpoint and WhatsApp route."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from io import BytesIO
    from utils.export_helpers import get_pdf_table_style
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if txn_type: query["txn_type"] = txn_type
    # Owner-account parties: combine account=owner + cash/bank txns with category=<owner>.
    is_owner_view = bool(category and party_type == "Owner")
    if is_owner_view:
        query["$and"] = [
            {"$or": [
                {"owner_name": category, "account": "owner"},
                {"category": category, "party_type": "Owner", "account": {"$in": ["cash", "bank"]}},
            ]},
            {"reference": {"$not": {"$regex": "^auto_ledger:"}}},
        ]
    else:
        if account:
            query["account"] = account
        if category:
            query["category"] = category
        if party_type:
            query["party_type"] = party_type
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        query["date"] = date_q
    txns = await db.cash_transactions.find(query, {"_id": 0}).sort([("date", 1), ("created_at", 1)]).to_list(10000)
    # Owner ledger view → flip txn_type for display ONLY for account=owner entries
    if is_owner_view:
        for t in txns:
            if t.get("account") == "owner":
                t["txn_type"] = "jama" if t.get("txn_type") == "nikasi" else "nikasi"
    summary = await get_cash_book_summary(kms_year=kms_year, season=season)
    
    run_bal = 0
    rows = []
    for t in txns:
        jama = t['amount'] if t['txn_type'] == 'jama' else 0
        nikasi = t['amount'] if t['txn_type'] == 'nikasi' else 0
        run_bal += jama - nikasi
        rows.append({
            "date": t.get("date", ""),
            "account_label": "Ledger" if t.get("account") == "ledger" else ("Cash" if t.get("account") == "cash" else ("Owner" if t.get("account") == "owner" else "Bank")),
            "type_label": "Jama" if t.get("txn_type") == "jama" else "Nikasi",
            "category": t.get("category", ""),
            "party_type": t.get("party_type", ""),
            "description": t.get("description", ""),
            "jama": jama if jama > 0 else "",
            "nikasi": nikasi if nikasi > 0 else "",
            "balance": round(run_bal, 2),
            "reference": t.get("reference", ""),
        })
    
    cols = get_columns("cashbook_report")
    headers = get_pdf_headers(cols)
    col_widths = [w*mm for w in get_pdf_widths_mm(cols)]
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=8*mm, rightMargin=8*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []; styles = get_pdf_styles()
    
    # Company Header + Title
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    title = "Daily Cash Book / रोज़नामचा"
    if category: title += f" - {category}"
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=11, textColor=colors.white,
        alignment=TA_CENTER, backColor=colors.HexColor('#0891b2'), spaceAfter=4, spaceBefore=2)
    
    subtitle_parts = [title + (f" | FY {kms_year}" if kms_year else "")]
    if date_from or date_to:
        dp = []
        if date_from: dp.append(f"From: {date_from}")
        if date_to: dp.append(f"To: {date_to}")
        subtitle_parts.append(" | ".join(dp))
    elements.append(Paragraph(" | ".join(subtitle_parts), title_style))
    elements.append(Spacer(1, 6))

    # Summary table removed (not useful for party ledger views)

    # Transactions table
    desc_style = ParagraphStyle('desc', fontName='FreeSans', fontSize=6, leading=7.5, alignment=TA_LEFT)
    party_style = ParagraphStyle('party', fontName='FreeSans', fontSize=6, leading=7.5, alignment=TA_LEFT)
    
    table_data = [headers]
    for r in rows:
        row_vals = get_entry_row(r, cols)
        out = []
        for i, v in enumerate(row_vals):
            if cols[i]["field"] == "description":
                out.append(Paragraph(str(v), desc_style))
            elif cols[i]["field"] == "category":
                out.append(Paragraph(str(v), party_style))
            else:
                out.append(str(v) if v != "" else "")
        table_data.append(out)
    
    tj = round(sum(t['amount'] for t in txns if t['txn_type'] == 'jama'), 2)
    tn = round(sum(t['amount'] for t in txns if t['txn_type'] == 'nikasi'), 2)
    totals = {"total_jama": tj, "total_nikasi": tn, "closing_balance": round(run_bal, 2)}
    total_vals = get_total_row(totals, cols)
    total_row = []
    for i, val in enumerate(total_vals):
        if i == 0: total_row.append("TOTAL / कुल")
        elif val is not None: total_row.append(str(val))
        else: total_row.append("")
    table_data.append(total_row)
    
    first_right = next((i for i, c in enumerate(cols) if c["align"] == "right"), 3)
    num_rows = len(table_data)
    style_cmds = get_pdf_table_style(num_rows)
    style_cmds.append(('ALIGN', (first_right, 1), (-1, -1), 'RIGHT'))
    
    tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    # ===== Beautiful single-line summary banner =====
    from utils.export_helpers import get_pdf_summary_banner, fmt_inr, STAT_COLORS
    net = tj - tn
    page_inner_w = sum(col_widths)
    summary_stats = [
        {'label': 'TOTAL ENTRIES', 'value': str(len(rows)), 'color': STAT_COLORS['primary']},
        {'label': 'TOTAL JAMA', 'value': fmt_inr(tj), 'color': STAT_COLORS['emerald']},
        {'label': 'TOTAL NIKASI', 'value': fmt_inr(tn), 'color': STAT_COLORS['red']},
        {'label': 'NET MOVEMENT', 'value': fmt_inr(net), 'color': STAT_COLORS['gold'] if net >= 0 else STAT_COLORS['orange']},
        {'label': 'CLOSING BALANCE', 'value': fmt_inr(run_bal), 'color': STAT_COLORS['blue'] if run_bal >= 0 else STAT_COLORS['red']},
    ]
    elements.append(Spacer(1, 8))
    banner = get_pdf_summary_banner(summary_stats, total_width=page_inner_w)
    if banner:
        elements.append(banner)

    doc.build(elements); buffer.seek(0)
    return buffer.getvalue()


@router.get("/cash-book/pdf")
async def export_cash_book_pdf(kms_year: Optional[str] = None, season: Optional[str] = None,
                                account: Optional[str] = None, txn_type: Optional[str] = None,
                                category: Optional[str] = None, party_type: Optional[str] = None,
                                date_from: Optional[str] = None, date_to: Optional[str] = None):
    pdf_bytes = await _generate_cash_book_pdf_bytes(
        kms_year=kms_year, season=season, account=account, txn_type=txn_type,
        category=category, party_type=party_type, date_from=date_from, date_to=date_to)
    return Response(content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cash_book_{datetime.now().strftime('%Y%m%d')}.pdf"})


