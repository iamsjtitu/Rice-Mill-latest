from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from database import db, USERS, print_pages
from models import *
from utils.optimistic_lock import optimistic_update, stamp_version
from utils.audit import log_audit
from utils.date_format import fmt_date
import uuid, io, csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()


async def get_company_name():
    branding = await db.branding.find_one({}, {"_id": 0})
    if branding:
        return branding.get("company_name", "NAVKAR AGRO"), branding.get("tagline", "")
    return "NAVKAR AGRO", "JOLKO, KESINGA"

# ============ MILL ENTRIES CRUD ============

@router.get("/")
async def root():
    return {"message": "Mill Entry API - Navkar Agro"}


@router.post("/entries")
async def create_entry(input: MillEntryCreate, username: str = "", role: str = ""):
    entry_dict = input.model_dump()
    
    # Duplicate RST check - same rst_no + kms_year not allowed
    rst = str(entry_dict.get("rst_no", "")).strip()
    kms = entry_dict.get("kms_year", "")
    if rst:
        existing = await db.mill_entries.find_one({"rst_no": rst, "kms_year": kms}, {"_id": 0, "id": 1, "rst_no": 1})
        if existing:
            raise HTTPException(status_code=400, detail=f"RST #{rst} pehle se entry hai")
    
    # Duplicate TP check - same tp_no + kms_year not allowed
    tp = str(entry_dict.get("tp_no", "")).strip()
    if tp:
        existing_tp = await db.mill_entries.find_one({"tp_no": tp, "kms_year": kms}, {"_id": 0, "id": 1, "rst_no": 1})
        if existing_tp:
            raise HTTPException(status_code=400, detail=f"TP No. {tp} pehle se entry hai")
    
    entry_dict = calculate_auto_fields(entry_dict)
    entry_dict['created_by'] = username
    
    entry_obj = MillEntry(**entry_dict)
    doc = stamp_version(entry_obj.model_dump())
    
    await db.mill_entries.insert_one(doc)
    await log_audit("mill_entries", doc["id"], "create", entry_dict.get("created_by", ""), new_data=doc)
    
    truck_no = doc.get("truck_no", "")
    
    # Auto Jama (Ledger) entry for truck purchase - what we owe the truck
    final_qntl = round(doc.get("qntl", 0) - doc.get("bag", 0) / 100, 2)
    if final_qntl > 0 and truck_no:
        # Look up rate from existing truck_payments for same truck_no + mandi
        existing_rate_doc = await db.truck_payments.find_one(
            {"entry_id": {"$in": [e["id"] async for e in db.mill_entries.find({"truck_no": truck_no, "mandi_name": doc.get("mandi_name", "")}, {"_id": 0, "id": 1})]}},
            {"_id": 0, "rate_per_qntl": 1}
        ) if truck_no else None
        rate = existing_rate_doc.get("rate_per_qntl", 32) if existing_rate_doc else 32
        gross_amount = round_amount(final_qntl * rate)
        
        cash_taken = float(doc.get("cash_paid", 0) or 0)
        diesel_taken = float(doc.get("diesel_paid", 0) or 0)
        deductions = cash_taken + diesel_taken
        
        jama_entry = {
            "id": str(uuid.uuid4()), "date": doc.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "account": "ledger", "txn_type": "jama", "category": truck_no,
            "party_type": "Truck",
            "description": f"Truck Entry: {truck_no} - {final_qntl}Q @ Rs.{rate}" + (f" (Ded: Rs.{deductions})" if deductions > 0 else ""),
            "amount": round_amount(gross_amount), "reference": f"truck_entry:{doc['id'][:8]}",
            "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
            "created_by": username or "system", "linked_entry_id": doc["id"],
            "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(jama_entry)
        jama_entry.pop("_id", None)
        await log_audit("cash_transactions", jama_entry["id"], "create", username, new_data=jama_entry)
        if diesel_taken > 0:
            diesel_ded = {
                "id": str(uuid.uuid4()), "date": doc.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
                "account": "ledger", "txn_type": "nikasi", "category": truck_no,
                "party_type": "Truck",
                "description": f"Truck Diesel Advance: {truck_no} - Rs.{diesel_taken}",
                "amount": round_amount(diesel_taken), "reference": f"truck_diesel_ded:{doc['id'][:8]}",
                "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
                "created_by": username or "system", "linked_entry_id": doc["id"],
                "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(diesel_ded)
            diesel_ded.pop("_id", None)
            await log_audit("cash_transactions", diesel_ded["id"], "create", username, new_data=diesel_ded)
    
    # Auto Cash Book entry for cash_paid
    cash_paid = float(doc.get("cash_paid", 0) or 0)
    if cash_paid > 0:
        cb = {
            "id": str(uuid.uuid4()), "date": doc.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "account": "cash", "txn_type": "nikasi", "category": doc.get("truck_no", "Cash Paid (Entry)"),
            "party_type": "Truck",
            "description": f"Cash Paid: Truck {doc.get('truck_no','')} - Mandi {doc.get('mandi_name','')} - Rs.{cash_paid}",
            "amount": round_amount(cash_paid), "reference": f"entry_cash:{doc['id'][:8]}",
            "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
            "created_by": username or "system", "linked_entry_id": doc["id"],
            "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(cb)
        cb.pop("_id", None)
        await log_audit("cash_transactions", cb["id"], "create", username, new_data=cb)
        # Also create Ledger Nikasi entry for cash deduction (counted against truck balance)
        if truck_no:
            cash_ded = {
                "id": str(uuid.uuid4()), "date": doc.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
                "account": "ledger", "txn_type": "nikasi", "category": truck_no,
                "party_type": "Truck",
                "description": f"Truck Cash Advance: {truck_no} - Rs.{cash_paid}",
                "amount": round_amount(cash_paid), "reference": f"truck_cash_ded:{doc['id'][:8]}",
                "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
                "created_by": username or "system", "linked_entry_id": doc["id"],
                "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(cash_ded)
            cash_ded.pop("_id", None)
            await log_audit("cash_transactions", cash_ded["id"], "create", username, new_data=cash_ded)
    
    # Auto Diesel Account entry for diesel_paid
    diesel_paid = float(doc.get("diesel_paid", 0) or 0)
    if diesel_paid > 0:
        # Get default pump
        default_pump = await db.diesel_pumps.find_one({"is_default": True}, {"_id": 0})
        pump_name = default_pump["name"] if default_pump else "Default Pump"
        pump_id = default_pump["id"] if default_pump else "default"
        diesel_txn = {
            "id": str(uuid.uuid4()), "date": doc.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "pump_id": pump_id, "pump_name": pump_name,
            "truck_no": doc.get("truck_no", ""), "agent_name": doc.get("agent_name", ""),
            "mandi_name": doc.get("mandi_name", ""),
            "amount": round_amount(diesel_paid), "txn_type": "debit",
            "description": f"Diesel: Truck {doc.get('truck_no','')} - Mandi {doc.get('mandi_name','')}",
            "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
            "created_by": username or "system", "linked_entry_id": doc["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.diesel_accounts.insert_one(diesel_txn)
        
        # Also create JAMA (Ledger) entry in cash_transactions for diesel pump
        diesel_jama = {
            "id": str(uuid.uuid4()), "date": doc.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "account": "ledger", "txn_type": "jama", "category": pump_name,
            "party_type": "Diesel",
            "description": f"Diesel Fill: Truck {doc.get('truck_no','')} - {pump_name} - Rs.{diesel_paid}",
            "amount": round_amount(diesel_paid), "reference": f"diesel_fill:{doc['id'][:8]}",
            "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
            "created_by": username or "system", "linked_entry_id": doc["id"],
            "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(diesel_jama)
        diesel_jama.pop("_id", None)
        await log_audit("cash_transactions", diesel_jama["id"], "create", username, new_data=diesel_jama)
    await _create_gunny_entries_for_mill(doc, username)
    
    # Return doc (which has _v) instead of entry_obj
    doc.pop("_id", None)
    return doc


async def _create_gunny_entries_for_mill(doc, username=""):
    """Auto-create gunny bag entries for bag (IN) and g_issued (OUT) in a mill entry."""
    entry_id = doc["id"]
    agent = doc.get("agent_name", "")
    mandi = doc.get("mandi_name", "")
    source = f"{agent} - {mandi}" if agent and mandi else (agent or mandi or "")
    truck = doc.get("truck_no", "")
    base = {
        "bag_type": "old", "rate": 0, "amount": 0, "notes": "Auto from Mill Entry",
        "kms_year": doc.get("kms_year", ""), "season": doc.get("season", ""),
        "created_by": username or "system", "linked_entry_id": entry_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    bag_in = int(float(doc.get("bag", 0) or 0))
    if bag_in > 0:
        entry = {**base, "id": str(uuid.uuid4()), "date": doc.get("date", ""),
                 "txn_type": "in", "quantity": bag_in, "source": source, "reference": truck}
        await db.gunny_bags.insert_one(entry)

    g_issued = int(float(doc.get("g_issued", 0) or 0))
    if g_issued > 0:
        entry = {**base, "id": str(uuid.uuid4()), "date": doc.get("date", ""),
                 "txn_type": "out", "quantity": g_issued, "source": source, "reference": truck}
        await db.gunny_bags.insert_one(entry)



@router.post("/entries/fix-cash-ledger")
async def fix_cash_ledger_entries():
    """Backfill missing ledger nikasi entries for cash_paid on existing mill entries"""
    entries = await db.mill_entries.find({}, {"_id": 0}).to_list(50000)
    fixed = 0
    for entry in entries:
        cash_paid = float(entry.get("cash_paid", 0) or 0)
        truck_no = entry.get("truck_no", "")
        entry_id = entry.get("id", "")
        if cash_paid > 0 and truck_no and entry_id:
            existing = await db.cash_transactions.find_one({"reference": f"truck_cash_ded:{entry_id[:8]}"})
            if not existing:
                cash_ded = {
                    "id": str(uuid.uuid4()), "date": entry.get("date", ""),
                    "account": "ledger", "txn_type": "nikasi", "category": truck_no,
                    "party_type": "Truck",
                    "description": f"Truck Cash Advance: {truck_no} - Rs.{cash_paid}",
                    "amount": round_amount(cash_paid), "reference": f"truck_cash_ded:{entry_id[:8]}",
                    "kms_year": entry.get("kms_year", ""), "season": entry.get("season", ""),
                    "created_by": "system", "linked_entry_id": entry_id,
                    "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
                }
                await db.cash_transactions.insert_one(cash_ded)
                fixed += 1
    return {"success": True, "fixed_count": fixed}


@router.post("/entries/import-excel")
async def import_entries_from_excel(
    file: UploadFile = File(...),
    kms_year: str = Form(""),
    season: str = Form(""),
    username: str = Form("admin"),
    preview_only: str = Form("false")
):
    """Import mill entries from Excel file. Also auto-creates cash book & diesel entries."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sirf Excel file (.xlsx) upload karein")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Find header row (look for DATE in first 5 rows)
    header_row = None
    for r in range(1, min(6, ws.max_row + 1)):
        for c in range(1, min(5, ws.max_column + 1)):
            val = str(ws.cell(r, c).value or "").strip().upper()
            if val == "DATE":
                header_row = r
                break
        if header_row:
            break

    if not header_row:
        raise HTTPException(status_code=400, detail="Header row nahi mila. 'DATE' column hona chahiye.")

    # Build column map from header
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(header_row, c).value or "").strip().upper()
        if "DATE" in val:
            col_map["date"] = c
        elif "TRUCK" in val:
            col_map["truck_no"] = c
        elif "AGENT" in val:
            col_map["agent_name"] = c
        elif "MANDI" in val:
            col_map["mandi_name"] = c
        elif "NETT" in val or val == "KG":
            col_map["kg"] = c
        elif val == "BAG":
            col_map["bag"] = c
        elif "DEPOSITE" in val or "G.DEP" in val:
            col_map["g_deposite"] = c
        elif "GBW" in val:
            col_map["gbw_cut"] = c
        elif "CUTTING" in val and "QNTL" not in val:
            col_map["cutting_percent"] = c
        elif "G.ISSUED" in val or "ISSUED" in val:
            col_map["g_issued"] = c
        elif "MOISTURE" in val:
            col_map["moisture"] = c
        elif "DISC" in val or "DUST" in val:
            col_map["disc_dust_poll"] = c
        elif "CASH" in val:
            col_map["cash_paid"] = c
        elif "DIESEL" in val:
            col_map["diesel_paid"] = c
        elif "REMARK" in val:
            col_map["remark"] = c

    def parse_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        if not s or s == "-":
            return None
        for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"]:
            try:
                return datetime.strptime(s.split(" ")[0], fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    def safe_float(val):
        if val is None or str(val).strip() in ("", "-", "None"):
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    def safe_int(val):
        if val is None or str(val).strip() in ("", "-", "None"):
            return 0
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    entries = []
    skipped = 0
    for r in range(header_row + 1, ws.max_row + 1):
        date_val = ws.cell(r, col_map.get("date", 1)).value
        date_str = parse_date(date_val)
        truck_col = col_map.get("truck_no", 2)
        truck_val = str(ws.cell(r, truck_col).value or "").strip()

        if not date_str or not truck_val:
            skipped += 1
            continue

        cutting_raw = safe_float(ws.cell(r, col_map.get("cutting_percent", 10)).value if "cutting_percent" in col_map else 0)
        # If cutting is in decimal form (e.g., 0.0526), convert to percent (5.26)
        cutting_pct = cutting_raw * 100 if 0 < cutting_raw < 1 else cutting_raw

        entry_data = {
            "date": date_str,
            "kms_year": kms_year,
            "season": season,
            "truck_no": truck_val,
            "agent_name": str(ws.cell(r, col_map.get("agent_name", 3)).value or "").strip(),
            "mandi_name": str(ws.cell(r, col_map.get("mandi_name", 4)).value or "").strip(),
            "kg": safe_float(ws.cell(r, col_map.get("kg", 5)).value if "kg" in col_map else 0),
            "bag": safe_int(ws.cell(r, col_map.get("bag", 6)).value if "bag" in col_map else 0),
            "g_deposite": safe_float(ws.cell(r, col_map.get("g_deposite", 7)).value if "g_deposite" in col_map else 0),
            "gbw_cut": safe_float(ws.cell(r, col_map.get("gbw_cut", 8)).value if "gbw_cut" in col_map else 0),
            "cutting_percent": cutting_pct,
            "g_issued": safe_float(ws.cell(r, col_map.get("g_issued", 12)).value if "g_issued" in col_map else 0),
            "moisture": safe_float(ws.cell(r, col_map.get("moisture", 13)).value if "moisture" in col_map else 0),
            "disc_dust_poll": safe_float(ws.cell(r, col_map.get("disc_dust_poll", 14)).value if "disc_dust_poll" in col_map else 0),
            "cash_paid": safe_float(ws.cell(r, col_map.get("cash_paid", 16)).value if "cash_paid" in col_map else 0),
            "diesel_paid": safe_float(ws.cell(r, col_map.get("diesel_paid", 17)).value if "diesel_paid" in col_map else 0),
            "remark": str(ws.cell(r, col_map.get("remark", 18)).value or "").strip() if "remark" in col_map else "",
        }
        entries.append(entry_data)

    if preview_only == "true":
        return {"preview": True, "count": len(entries), "skipped": skipped,
                "sample": entries[:10], "columns_detected": list(col_map.keys())}

    # Import all entries with auto cash book and diesel entries
    imported = 0
    cash_count = 0
    diesel_count = 0
    default_pump = await db.diesel_pumps.find_one({"is_default": True}, {"_id": 0})
    pump_name = default_pump["name"] if default_pump else "Default Pump"
    pump_id = default_pump["id"] if default_pump else "default"

    for entry_data in entries:
        entry_data = calculate_auto_fields(entry_data)
        entry_data["created_by"] = username
        entry_obj = MillEntry(**entry_data)
        doc = entry_obj.model_dump()
        await db.mill_entries.insert_one(doc)

        cash_paid = float(doc.get("cash_paid", 0) or 0)
        if cash_paid > 0:
            cb = {
                "id": str(uuid.uuid4()), "date": doc["date"],
                "account": "cash", "txn_type": "nikasi", "category": doc.get("truck_no", "Cash Paid (Entry)"),
                "party_type": "Truck",
                "description": f"Cash Paid: Truck {doc.get('truck_no','')} - Mandi {doc.get('mandi_name','')} - Rs.{cash_paid}",
                "amount": round_amount(cash_paid), "reference": f"entry_cash:{doc['id'][:8]}",
                "kms_year": kms_year, "season": season,
                "created_by": username, "linked_entry_id": doc["id"],
                "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(cb)
            cash_count += 1

        diesel_paid = float(doc.get("diesel_paid", 0) or 0)
        if diesel_paid > 0:
            diesel_txn = {
                "id": str(uuid.uuid4()), "date": doc["date"],
                "pump_id": pump_id, "pump_name": pump_name,
                "truck_no": doc.get("truck_no", ""), "agent_name": doc.get("agent_name", ""),
            "mandi_name": doc.get("mandi_name", ""),
                "amount": round_amount(diesel_paid), "txn_type": "debit",
                "description": f"Diesel: Truck {doc.get('truck_no','')} - Mandi {doc.get('mandi_name','')}",
                "kms_year": kms_year, "season": season,
                "created_by": username, "linked_entry_id": doc["id"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.diesel_accounts.insert_one(diesel_txn)
            # Diesel JAMA ledger entry
            diesel_jama = {
                "id": str(uuid.uuid4()), "date": doc["date"],
                "account": "ledger", "txn_type": "jama", "category": pump_name,
                "party_type": "Diesel",
                "description": f"Diesel Fill: Truck {doc.get('truck_no','')} - {pump_name} - Rs.{diesel_paid}",
                "amount": round_amount(diesel_paid), "reference": f"diesel_fill:{doc['id'][:8]}",
                "kms_year": kms_year, "season": season,
                "created_by": username, "linked_entry_id": doc["id"],
                "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(diesel_jama)
            diesel_count += 1

        # Auto Gunny Bag entries for g_issued and g_deposite
        await _create_gunny_entries_for_mill(doc, username)

        imported += 1

    return {
        "success": True,
        "imported": imported,
        "skipped": skipped,
        "cash_book_entries": cash_count,
        "diesel_entries": diesel_count,
        "message": f"{imported} entries import ho gaye! Cash Book: {cash_count}, Diesel: {diesel_count}"
    }



@router.get("/entries")
async def get_entries(
    truck_no: Optional[str] = None,
    rst_no: Optional[str] = None,
    tp_no: Optional[str] = None,
    agent_name: Optional[str] = None,
    mandi_name: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    page_size: int = 200
):
    query = {}
    
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if rst_no:
        query["rst_no"] = {"$regex": rst_no, "$options": "i"}
    if tp_no:
        query["tp_no"] = {"$regex": tp_no, "$options": "i"}
    if agent_name:
        query["agent_name"] = {"$regex": agent_name, "$options": "i"}
    if mandi_name:
        query["mandi_name"] = {"$regex": mandi_name, "$options": "i"}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    # Date range filter
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        if date_query:
            query["date"] = date_query
    
    total_count = await db.mill_entries.count_documents(query)
    if page_size <= 0:
        entries = await db.mill_entries.find(query, {"_id": 0}).to_list(50000)
        entries.sort(key=lambda e: (e.get("date", "")[:10], int(e.get("rst_no") or 0)), reverse=True)
        return {"entries": entries, "total": total_count, "page": 1, "page_size": total_count, "total_pages": 1}
    if page < 1: page = 1
    skip = (page - 1) * page_size

    all_entries = await db.mill_entries.find(query, {"_id": 0}).to_list(50000)
    all_entries.sort(key=lambda e: (e.get("date", ""), int(e.get("rst_no") or 0)), reverse=True)
    total = len(all_entries)
    entries = all_entries[skip:skip + page_size]
    # Ensure every entry has a persistent 'id' field
    for e in entries:
        if not e.get("id"):
            new_id = str(uuid.uuid4())
            e["id"] = new_id
            await db.mill_entries.update_one(
                {"_id": (await db.mill_entries.find_one({"truck_no": e.get("truck_no"), "date": e.get("date"), "created_at": e.get("created_at")}, {"_id": 1}) or {}).get("_id")},
                {"$set": {"id": new_id}}
            )
    return {"entries": entries, "total": total, "page": page, "page_size": page_size, "total_pages": max(1, (total + page_size - 1) // page_size)}


@router.get("/entries/check-duplicate")
async def check_duplicate_rst_tp(rst_no: str = "", tp_no: str = "", kms_year: str = "", exclude_id: str = ""):
    """Real-time check if RST or TP already exists."""
    result = {"rst_exists": False, "tp_exists": False, "rst_entry": None, "tp_entry": None, "tp_rst_no": None}
    if rst_no.strip():
        q = {"rst_no": rst_no.strip(), "kms_year": kms_year}
        if exclude_id:
            q["id"] = {"$ne": exclude_id}
        found = await db.mill_entries.find_one(q, {"_id": 0, "id": 1, "rst_no": 1, "truck_no": 1})
        if found:
            result["rst_exists"] = True
            result["rst_entry"] = f"RST #{rst_no} - {found.get('truck_no', '')}"
    if tp_no.strip():
        q = {"tp_no": tp_no.strip(), "kms_year": kms_year}
        if exclude_id:
            q["id"] = {"$ne": exclude_id}
        found = await db.mill_entries.find_one(q, {"_id": 0, "id": 1, "rst_no": 1, "truck_no": 1})
        if found:
            result["tp_exists"] = True
            result["tp_rst_no"] = found.get('rst_no', '?')
            result["tp_entry"] = f"RST #{found.get('rst_no', '?')} - {found.get('truck_no', '')}"
    return result


@router.get("/entries/{entry_id}")
async def get_entry(entry_id: str):
    entry = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    return entry


@router.put("/entries/{entry_id}")
async def update_entry(entry_id: str, request: Request, username: str = "", role: str = ""):
    raw_body = await request.json()
    client_v = raw_body.pop("_v", None)

    existing = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Check permission
    can_edit, message = can_edit_entry(existing, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    
    update_data = {k: v for k, v in raw_body.items() if v is not None}
    
    # Duplicate RST/TP check (exclude self)
    kms = update_data.get("kms_year", existing.get("kms_year", ""))
    rst = str(update_data.get("rst_no", existing.get("rst_no", ""))).strip()
    if rst:
        dup_rst = await db.mill_entries.find_one({"rst_no": rst, "kms_year": kms, "id": {"$ne": entry_id}}, {"_id": 0, "id": 1})
        if dup_rst:
            raise HTTPException(status_code=400, detail=f"RST #{rst} pehle se entry hai")
    tp = str(update_data.get("tp_no", existing.get("tp_no", ""))).strip()
    if tp:
        dup_tp = await db.mill_entries.find_one({"tp_no": tp, "kms_year": kms, "id": {"$ne": entry_id}}, {"_id": 0, "id": 1, "rst_no": 1})
        if dup_tp:
            raise HTTPException(status_code=400, detail=f"TP No. {tp} pehle se entry hai")
    
    # Merge existing data with updates
    merged_data = {**existing, **update_data}
    merged_data = calculate_auto_fields(merged_data)
    merged_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    merged_data.pop("_id", None)
    
    await optimistic_update(db.mill_entries, entry_id, merged_data, client_v)
    await log_audit("mill_entries", entry_id, "update", raw_body.get("username", username), old_data=existing, new_data=merged_data)
    
    # Update auto cash book entries for this entry (delete all and recreate)
    await db.cash_transactions.delete_many({"linked_entry_id": entry_id})
    
    truck_no = merged_data.get("truck_no", "")
    
    # Recreate Jama (Ledger) entry for truck purchase
    final_qntl = round(merged_data.get("qntl", 0) - merged_data.get("bag", 0) / 100, 2)
    if final_qntl > 0 and truck_no:
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        gross_amount = round_amount(final_qntl * rate)
        cash_taken = float(merged_data.get("cash_paid", 0) or 0)
        diesel_taken = float(merged_data.get("diesel_paid", 0) or 0)
        deductions = cash_taken + diesel_taken
        
        jama_entry = {
            "id": str(uuid.uuid4()), "date": merged_data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "account": "ledger", "txn_type": "jama", "category": truck_no,
            "party_type": "Truck",
            "description": f"Truck Entry: {truck_no} - {final_qntl}Q @ Rs.{rate}" + (f" (Ded: Rs.{deductions})" if deductions > 0 else ""),
            "amount": round_amount(gross_amount), "reference": f"truck_entry:{entry_id[:8]}",
            "kms_year": merged_data.get("kms_year", ""), "season": merged_data.get("season", ""),
            "created_by": username or "system", "linked_entry_id": entry_id,
            "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(jama_entry)
        jama_entry.pop("_id", None)
        await log_audit("cash_transactions", jama_entry["id"], "create", username, new_data=jama_entry)
        
        if diesel_taken > 0:
            diesel_ded = {
                "id": str(uuid.uuid4()), "date": merged_data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
                "account": "ledger", "txn_type": "nikasi", "category": truck_no,
                "party_type": "Truck",
                "description": f"Truck Diesel Advance: {truck_no} - Rs.{diesel_taken}",
                "amount": round_amount(diesel_taken), "reference": f"truck_diesel_ded:{entry_id[:8]}",
                "kms_year": merged_data.get("kms_year", ""), "season": merged_data.get("season", ""),
                "created_by": username or "system", "linked_entry_id": entry_id,
                "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(diesel_ded)
            diesel_ded.pop("_id", None)
            await log_audit("cash_transactions", diesel_ded["id"], "create", username, new_data=diesel_ded)
    
    # Recreate Cash Book Nikasi entry for cash_paid
    cash_paid = float(merged_data.get("cash_paid", 0) or 0)
    if cash_paid > 0:
        cb = {
            "id": str(uuid.uuid4()), "date": merged_data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "account": "cash", "txn_type": "nikasi", "category": merged_data.get("truck_no", "Cash Paid (Entry)"),
            "party_type": "Truck",
            "description": f"Cash Paid: Truck {merged_data.get('truck_no','')} - Mandi {merged_data.get('mandi_name','')} - Rs.{cash_paid}",
            "amount": round_amount(cash_paid), "reference": f"entry_cash:{entry_id[:8]}",
            "kms_year": merged_data.get("kms_year", ""), "season": merged_data.get("season", ""),
            "created_by": username or "system", "linked_entry_id": entry_id,
            "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(cb)
        cb.pop("_id", None)
        await log_audit("cash_transactions", cb["id"], "create", username, new_data=cb)
        # Also create Ledger Nikasi entry for cash deduction (counted against truck balance)
        if truck_no:
            cash_ded = {
                "id": str(uuid.uuid4()), "date": merged_data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
                "account": "ledger", "txn_type": "nikasi", "category": truck_no,
                "party_type": "Truck",
                "description": f"Truck Cash Advance: {truck_no} - Rs.{cash_paid}",
                "amount": round_amount(cash_paid), "reference": f"truck_cash_ded:{entry_id[:8]}",
                "kms_year": merged_data.get("kms_year", ""), "season": merged_data.get("season", ""),
                "created_by": username or "system", "linked_entry_id": entry_id,
                "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.cash_transactions.insert_one(cash_ded)
            cash_ded.pop("_id", None)
            await log_audit("cash_transactions", cash_ded["id"], "create", username, new_data=cash_ded)
    
    # Update auto diesel account entry for diesel_paid
    await db.diesel_accounts.delete_many({"linked_entry_id": entry_id})
    diesel_paid = float(merged_data.get("diesel_paid", 0) or 0)
    if diesel_paid > 0:
        default_pump = await db.diesel_pumps.find_one({"is_default": True}, {"_id": 0})
        pump_name = default_pump["name"] if default_pump else "Default Pump"
        pump_id = default_pump["id"] if default_pump else "default"
        diesel_txn = {
            "id": str(uuid.uuid4()), "date": merged_data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "pump_id": pump_id, "pump_name": pump_name,
            "truck_no": merged_data.get("truck_no", ""), "agent_name": merged_data.get("agent_name", ""),
            "mandi_name": merged_data.get("mandi_name", ""),
            "amount": round_amount(diesel_paid), "txn_type": "debit",
            "description": f"Diesel: Truck {merged_data.get('truck_no','')} - Mandi {merged_data.get('mandi_name','')}",
            "kms_year": merged_data.get("kms_year", ""), "season": merged_data.get("season", ""),
            "created_by": username or "system", "linked_entry_id": entry_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.diesel_accounts.insert_one(diesel_txn)
        
        # Also create JAMA (Ledger) entry in cash_transactions for diesel pump
        diesel_jama = {
            "id": str(uuid.uuid4()), "date": merged_data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            "account": "ledger", "txn_type": "jama", "category": pump_name,
            "party_type": "Diesel",
            "description": f"Diesel Fill: Truck {merged_data.get('truck_no','')} - {pump_name} - Rs.{diesel_paid}",
            "amount": round_amount(diesel_paid), "reference": f"diesel_fill:{entry_id[:8]}",
            "kms_year": merged_data.get("kms_year", ""), "season": merged_data.get("season", ""),
            "created_by": username or "system", "linked_entry_id": entry_id,
            "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cash_transactions.insert_one(diesel_jama)
        diesel_jama.pop("_id", None)
        await log_audit("cash_transactions", diesel_jama["id"], "create", username, new_data=diesel_jama)
    
    # Update auto gunny bag entries (delete old + recreate)
    await db.gunny_bags.delete_many({"linked_entry_id": entry_id})
    await _create_gunny_entries_for_mill(merged_data | {"id": entry_id}, username)
    
    # Sync cash/diesel to linked vehicle_weight entry (same RST + kms_year)
    entry_rst = str(merged_data.get("rst_no", "")).strip()
    entry_kms = merged_data.get("kms_year", "")
    if entry_rst:
        try:
            rst_int = int(entry_rst)
            vw_query = {"rst_no": rst_int, "kms_year": entry_kms}
            vw_existing = await db["vehicle_weights"].find_one(vw_query, {"_id": 0, "id": 1})
            if vw_existing:
                vw_update_fields = {
                    "cash_paid": float(merged_data.get("cash_paid", 0) or 0),
                    "diesel_paid": float(merged_data.get("diesel_paid", 0) or 0),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                await db["vehicle_weights"].update_one({"id": vw_existing["id"]}, {"$set": vw_update_fields})
                await log_audit("vehicle_weights", vw_existing["id"], "update", username,
                    old_data={"cash_paid": vw_existing.get("cash_paid"), "diesel_paid": vw_existing.get("diesel_paid")},
                    new_data=vw_update_fields)
        except (ValueError, TypeError):
            pass
    
    updated = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    return updated


@router.delete("/entries/{entry_id}")
async def delete_entry(entry_id: str, username: str = "", role: str = ""):
    existing = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Check permission
    can_edit, message = can_edit_entry(existing, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    
    result = await db.mill_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    await log_audit("mill_entries", entry_id, "delete", username, old_data=existing)
    
    # Clean up linked auto entries
    await db.cash_transactions.delete_many({"linked_entry_id": entry_id})
    await db.diesel_accounts.delete_many({"linked_entry_id": entry_id})
    await db.gunny_bags.delete_many({"linked_entry_id": entry_id})
    
    return {"message": "Entry deleted successfully"}


@router.get("/totals", response_model=TotalsResponse)
async def get_totals(
    truck_no: Optional[str] = None,
    agent_name: Optional[str] = None,
    mandi_name: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    match_query = {}
    
    if truck_no:
        match_query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if agent_name:
        match_query["agent_name"] = {"$regex": agent_name, "$options": "i"}
    if mandi_name:
        match_query["mandi_name"] = {"$regex": mandi_name, "$options": "i"}
    if kms_year:
        match_query["kms_year"] = kms_year
    if season:
        match_query["season"] = season
    
    # Date range filter
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        if date_query:
            match_query["date"] = date_query
    
    pipeline = []
    if match_query:
        pipeline.append({"$match": match_query})
    
    pipeline.append({
        "$group": {
            "_id": None,
            "total_kg": {"$sum": "$kg"},
            "total_qntl": {"$sum": "$qntl"},
            "total_bag": {"$sum": "$bag"},
            "total_g_deposite": {"$sum": "$g_deposite"},
            "total_gbw_cut": {"$sum": "$gbw_cut"},
            "total_mill_w": {"$sum": "$mill_w"},
            "total_p_pkt_cut": {"$sum": "$p_pkt_cut"},
            "total_cutting": {"$sum": "$cutting"},
            "total_disc_dust_poll": {"$sum": "$disc_dust_poll"},
            "total_final_w": {"$sum": "$final_w"},
            "total_g_issued": {"$sum": "$g_issued"},
            "total_cash_paid": {"$sum": "$cash_paid"},
            "total_diesel_paid": {"$sum": "$diesel_paid"}
        }
    })
    
    result = await db.mill_entries.aggregate(pipeline).to_list(1)
    
    if result:
        totals = result[0]
        del totals['_id']
        return TotalsResponse(**totals)
    
    return TotalsResponse()


# ============ AUTO-SUGGEST ENDPOINTS ============

@router.get("/suggestions/trucks")
async def get_truck_suggestions(q: str = ""):
    if len(q) < 1:
        trucks_mill = await db.mill_entries.distinct("truck_no")
        trucks_vw = await db.vehicle_weights.distinct("vehicle_no")
    else:
        trucks_mill = await db.mill_entries.distinct("truck_no", {"truck_no": {"$regex": q, "$options": "i"}})
        trucks_vw = await db.vehicle_weights.distinct("vehicle_no", {"vehicle_no": {"$regex": q, "$options": "i"}})
    combined = list(set([t for t in (trucks_mill + trucks_vw) if t]))
    combined.sort()
    return {"suggestions": combined}


@router.get("/suggestions/agents")
async def get_agent_suggestions(q: str = ""):
    if len(q) < 1:
        agents = await db.mill_entries.distinct("agent_name")
        vw_parties = await db.vehicle_weights.distinct("party_name")
    else:
        agents = await db.mill_entries.distinct("agent_name", {"agent_name": {"$regex": q, "$options": "i"}})
        vw_parties = await db.vehicle_weights.distinct("party_name", {"party_name": {"$regex": q, "$options": "i"}})
    combined = list(set([a for a in (agents + vw_parties) if a]))
    combined.sort()
    return {"suggestions": combined}


@router.get("/suggestions/mandis")
async def get_mandi_suggestions(q: str = "", agent_name: str = ""):
    query = {}
    if q:
        query["mandi_name"] = {"$regex": q, "$options": "i"}
    if agent_name:
        query["agent_name"] = agent_name
    
    mandis = await db.mill_entries.distinct("mandi_name", query if query else None)
    # Also get farmer_name from vehicle_weights as source suggestions
    vw_query = {}
    if q:
        vw_query["farmer_name"] = {"$regex": q, "$options": "i"}
    if agent_name:
        vw_query["party_name"] = agent_name
    vw_sources = await db.vehicle_weights.distinct("farmer_name", vw_query if vw_query else None)
    combined = list(set([m for m in (mandis + vw_sources) if m]))
    combined.sort()
    return {"suggestions": combined}


@router.get("/suggestions/kms_years")
async def get_kms_year_suggestions():
    years = await db.mill_entries.distinct("kms_year")
    return {"suggestions": [y for y in years if y]}


# ============ EXPORT ENDPOINTS ============

@router.get("/export/excel")
async def export_excel(
    truck_no: Optional[str] = None,
    rst_no: Optional[str] = None,
    tp_no: Optional[str] = None,
    agent_name: Optional[str] = None,
    mandi_name: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    report_title: Optional[str] = None
):
    """Export entries to styled Excel file"""
    query = {}
    
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if rst_no:
        query["rst_no"] = {"$regex": rst_no, "$options": "i"}
    if tp_no:
        query["tp_no"] = {"$regex": tp_no, "$options": "i"}
    if agent_name:
        query["agent_name"] = {"$regex": agent_name, "$options": "i"}
    if mandi_name:
        query["mandi_name"] = {"$regex": mandi_name, "$options": "i"}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if date_from or date_to:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to: dq["$lte"] = date_to
        if dq: query["date"] = dq
    
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(1000)
    entries.sort(key=lambda e: (e.get("date", "")[:10], int(e.get("rst_no") or 0)))
    
    # Create workbook
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Mill Entries"
    
    ncols = 19
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title
    company_name, tagline = await get_company_name()
    default_title = "Mill Entries / मिल एंट्री"
    display_title = report_title or default_title
    title = f"{company_name} - {display_title}"
    subtitle = f"FY: {kms_year or 'All'} | {season or 'All Seasons'} | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    style_excel_title(ws, title, ncols, subtitle)
    
    # Headers
    headers = [
        "Date", "Truck No", "RST No", "TP No", "Agent", "Mandi", "QNTL", "BAG", "G.Dep",
        "GBW Cut", "P.Pkt", "P.Pkt Cut", "Mill W", "Moist%", "M.Cut", "Cut%", 
        "D/D/P", "Final W", "G.Issued"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_excel_header_row(ws, 4, ncols)
    
    # Data rows
    data_start = 5
    row_num = data_start
    
    for idx, entry in enumerate(entries):
        row_data = [
            fmt_date(entry.get('date', '')),
            entry.get('truck_no', ''),
            entry.get('rst_no', ''),
            entry.get('tp_no', ''),
            entry.get('agent_name', ''),
            entry.get('mandi_name', ''),
            round(entry.get('qntl', 0), 2),
            entry.get('bag', 0),
            entry.get('g_deposite', 0),
            round(entry.get('gbw_cut', 0), 2),
            entry.get('plastic_bag', 0),
            round(entry.get('p_pkt_cut', 0) / 100, 2),
            round(entry.get('mill_w', 0) / 100, 2),
            entry.get('moisture', 0),
            round(entry.get('moisture_cut', 0) / 100, 2) if entry.get('moisture_cut') else 0,
            entry.get('cutting_percent', 0),
            entry.get('disc_dust_poll', 0),
            round(entry.get('final_w', 0) / 100, 2),
            entry.get('g_issued', 0)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = BORDER_THIN
            if col >= 6: cell.alignment = right_align
        
        row_num += 1
    
    if entries:
        style_excel_data_rows(ws, data_start, row_num - 1, ncols, headers)
    
    # Totals row
    totals = await get_totals(truck_no, agent_name, mandi_name, kms_year, season, date_from, date_to)
    totals_data = [
        "TOTAL", "", "", "", "", "",
        round(totals.total_qntl, 2),
        totals.total_bag,
        totals.total_g_deposite,
        round(totals.total_gbw_cut, 2),
        "-",
        round(totals.total_p_pkt_cut / 100, 2) if hasattr(totals, 'total_p_pkt_cut') else "-",
        round(totals.total_mill_w / 100, 2),
        "-",
        "-",
        "-",
        totals.total_disc_dust_poll,
        round(totals.total_final_w / 100, 2),
        totals.total_g_issued
    ]
    
    for col, value in enumerate(totals_data, 1):
        ws.cell(row=row_num, column=col, value=value)
        if col >= 7:
            ws.cell(row=row_num, column=col).alignment = right_align
    style_excel_total_row(ws, row_num, ncols)
    
    # Column widths - A4 optimized (19 cols)
    col_widths = [10, 12, 9, 9, 10, 16, 9, 6, 6, 7, 6, 7, 8, 6, 7, 6, 6, 9, 7]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"mill_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf")
async def export_pdf(
    truck_no: Optional[str] = None,
    rst_no: Optional[str] = None,
    tp_no: Optional[str] = None,
    agent_name: Optional[str] = None,
    mandi_name: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    report_title: Optional[str] = None
):
    """Export entries to styled PDF file (A4 Landscape)"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    
    query = {}
    
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if rst_no:
        query["rst_no"] = {"$regex": rst_no, "$options": "i"}
    if tp_no:
        query["tp_no"] = {"$regex": tp_no, "$options": "i"}
    if agent_name:
        query["agent_name"] = {"$regex": agent_name, "$options": "i"}
    if mandi_name:
        query["mandi_name"] = {"$regex": mandi_name, "$options": "i"}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if date_from or date_to:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to: dq["$lte"] = date_to
        if dq: query["date"] = dq
    
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(1000)
    entries.sort(key=lambda e: (e.get("date", "")[:10], int(e.get("rst_no") or 0)))
    totals = await get_totals(truck_no, agent_name, mandi_name, kms_year, season, date_from, date_to)
    
    # Create PDF buffer
    buffer = io.BytesIO()
    
    # A4 Landscape
    page_width, page_height = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=8*mm,
        rightMargin=8*mm,
        topMargin=8*mm,
        bottomMargin=8*mm
    )
    
    elements = []
    styles = get_pdf_styles()
    
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.white,
        alignment=TA_CENTER,
        spaceAfter=2*mm
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#475569'),
        alignment=TA_CENTER,
        spaceAfter=3*mm
    )
    
    from utils.export_helpers import get_pdf_table_style
    
    # Title table with themed background
    company_name, tagline = await get_company_name()
    default_title = "Mill Entries / मिल एंट्री"
    display_title = report_title or default_title
    title_text = f"{company_name} - {display_title} | FY: {kms_year or 'All'} | {season or 'All Seasons'}"
    title_data = [[Paragraph(f"<b>{title_text}</b>", title_style)]]
    title_table = Table(title_data, colWidths=[page_width - 16*mm])
    title_table.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), 
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1B4F72')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(title_table)
    
    # Date
    date_text = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    elements.append(Paragraph(date_text, subtitle_style))
    
    # Table headers
    headers = [
        "Date", "Truck", "RST", "TP", "Agent", "Mandi", "QNTL", "BAG", "G.Dep",
        "GBW", "P.Pkt", "P.Cut", "Mill W", "M%", "M.Cut", "C%", 
        "D/D/P", "Final W", "G.Iss"
    ]
    
    # Build data rows
    table_data = [headers]
    
    for entry in entries:
        row = [
            fmt_date(entry.get('date', '')[:10]) if entry.get('date') else '',
            entry.get('truck_no', '')[:14] if entry.get('truck_no') else '',
            entry.get('rst_no', '')[:8] if entry.get('rst_no') else '',
            entry.get('tp_no', '')[:8] if entry.get('tp_no') else '',
            entry.get('agent_name', '')[:10] if entry.get('agent_name') else '',
            entry.get('mandi_name', '')[:16] if entry.get('mandi_name') else '',
            f"{entry.get('qntl', 0):.2f}",
            str(entry.get('bag', 0)),
            str(entry.get('g_deposite', 0)),
            f"{entry.get('gbw_cut', 0):.1f}",
            str(entry.get('plastic_bag', 0)),
            f"{entry.get('p_pkt_cut', 0) / 100:.2f}",
            f"{entry.get('mill_w', 0) / 100:.2f}",
            f"{entry.get('moisture', 0):.0f}",
            f"{(entry.get('moisture_cut', 0) or 0) / 100:.2f}",
            f"{entry.get('cutting_percent', 0):.1f}",
            str(entry.get('disc_dust_poll', 0)),
            f"{entry.get('final_w', 0) / 100:.2f}",
            str(entry.get('g_issued', 0))
        ]
        table_data.append(row)
    
    # Totals row
    totals_row = [
        "TOTAL", "", "", "", "", "",
        f"{totals.total_qntl:.2f}",
        str(totals.total_bag),
        str(int(totals.total_g_deposite)),
        f"{totals.total_gbw_cut:.1f}",
        "-",
        f"{totals.total_p_pkt_cut / 100:.2f}" if hasattr(totals, 'total_p_pkt_cut') else "-",
        f"{totals.total_mill_w / 100:.2f}",
        "-",
        "-",
        "-",
        str(int(totals.total_disc_dust_poll)),
        f"{totals.total_final_w / 100:.2f}",
        str(int(totals.total_g_issued))
    ]
    table_data.append(totals_row)
    
    # Column widths (19 columns for A4 landscape with margins)
    col_widths = [15*mm, 18*mm, 11*mm, 11*mm, 15*mm, 27*mm, 13*mm, 9*mm, 9*mm, 11*mm, 
                  9*mm, 11*mm, 13*mm, 9*mm, 11*mm, 9*mm, 9*mm, 13*mm, 11*mm]
    
    # Create table
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Use centralized styling
    cols_info = [{'header': h} for h in headers]
    style_commands = get_pdf_table_style(len(table_data), cols_info)
    style_commands.extend([
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('ALIGN', (6, 1), (-1, -1), 'RIGHT'),
    ])
    
    main_table.setStyle(TableStyle(style_commands))
    elements.append(main_table)
    
    # Build PDF
    doc.build(elements)
    
    buffer.seek(0)
    filename = f"mill_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/truck-payments-excel")
async def export_truck_payments_excel(
    truck_no: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export truck payments to styled Excel file"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort([("date", 1), ("rst_no", 1), ("created_at", 1)]).to_list(1000)
    
    # Build payment data
    payments_data = []
    total_net = 0
    total_paid = 0
    total_balance = 0
    
    for entry in entries:
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round_amount(final_qntl * rate)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions)
        balance = round(max(0, net_amount - paid_amount), 2)
        status = "Paid" if balance < 0.10 else ("Partial" if paid_amount > 0 else "Pending")
        
        total_net += net_amount
        total_paid += paid_amount
        total_balance += balance
        
        payments_data.append({
            "date": fmt_date(entry.get("date", "")),
            "truck_no": entry.get("truck_no", ""),
            "mandi_name": entry.get("mandi_name", ""),
            "final_qntl": final_qntl,
            "rate": rate,
            "gross": gross_amount,
            "cash": cash_taken,
            "diesel": diesel_taken,
            "deductions": deductions,
            "net": net_amount,
            "paid": paid_amount,
            "balance": balance,
            "status": status
        })
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Truck Payments"
    
    # Styles
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    paid_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    pending_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:M1')
    company_name, tagline = await get_company_name()
    ws['A1'] = f"TRUCK PAYMENTS - {company_name} | FY: {kms_year or 'All'} | {season or 'All'}"
    ws['A1'].font = Font(bold=True, size=14, color="D97706")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Date", "Truck No", "Mandi", "Final QNTL", "Rate", "Gross", "Cash", "Diesel", "Deductions", "Net Amount", "Paid", "Balance", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, p in enumerate(payments_data, 4):
        ws.cell(row=row_idx, column=1, value=p["date"])
        ws.cell(row=row_idx, column=2, value=p["truck_no"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value=p["mandi_name"])
        ws.cell(row=row_idx, column=4, value=p["final_qntl"])
        ws.cell(row=row_idx, column=5, value=f"₹{p['rate']}")
        ws.cell(row=row_idx, column=6, value=p["gross"])
        ws.cell(row=row_idx, column=7, value=p["cash"])
        ws.cell(row=row_idx, column=8, value=p["diesel"])
        ws.cell(row=row_idx, column=9, value=p["deductions"])
        ws.cell(row=row_idx, column=10, value=p["net"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=11, value=p["paid"])
        ws.cell(row=row_idx, column=12, value=p["balance"]).font = Font(bold=True, color="DC2626" if p["balance"] > 0 else "059669")
        status_cell = ws.cell(row=row_idx, column=13, value=p["status"])
        if p["status"] == "Paid":
            status_cell.fill = paid_fill
        elif p["status"] == "Pending":
            status_cell.fill = pending_fill
    
    # Totals row
    total_row = len(payments_data) + 4
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=10, value=round(total_net, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=11, value=round(total_paid, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=12, value=round(total_balance, 2)).font = Font(bold=True, color="DC2626")
    for col in range(1, 14):
        ws.cell(row=total_row, column=col).fill = total_fill
    
    # Column widths
    col_widths = [12, 14, 14, 12, 8, 10, 8, 8, 10, 12, 10, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"truck_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/truck-payments-pdf")
async def export_truck_payments_pdf(
    truck_no: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export truck payments to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.enums import TA_CENTER
    
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort([("date", 1), ("rst_no", 1), ("created_at", 1)]).to_list(1000)
    
    # Build payment data
    payments_data = []
    total_net = 0
    total_paid = 0
    total_balance = 0
    
    for entry in entries:
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("qntl", 0) - entry.get("bag", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round_amount(final_qntl * rate)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions)
        balance = round(max(0, net_amount - paid_amount), 2)
        status = "Paid" if balance < 0.10 else ("Partial" if paid_amount > 0 else "Pending")
        
        total_net += net_amount
        total_paid += paid_amount
        total_balance += balance
        
        payments_data.append([
            fmt_date(entry.get("date", "")[:10]),
            entry.get("truck_no", "")[:12],
            entry.get("mandi_name", "")[:12],
            f"{final_qntl}",
            f"Rs.{rate}",
            f"Rs.{gross_amount}",
            f"-Rs.{deductions}",
            f"Rs.{net_amount}",
            f"Rs.{paid_amount}",
            f"Rs.{balance}",
            status
        ])
    
    # Create PDF
    buffer = io.BytesIO()
    page_width, page_height = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    elements = []
    styles = get_pdf_styles()
    
    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.white, alignment=TA_CENTER)
    company_name, tagline = await get_company_name()
    title_data = [[Paragraph(f"<b>TRUCK PAYMENTS - {company_name} | FY: {kms_year or 'All'} | {season or 'All'}</b>", title_style)]]
    title_table = Table(title_data, colWidths=[page_width - 20*mm])
    title_table.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), 
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D97706')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(title_table)
    elements.append(Table([[""]], colWidths=[page_width], rowHeights=[5*mm]))
    
    # Headers
    headers = ["Date", "Truck No", "Mandi", "QNTL", "Rate", "Gross", "Deduct", "Net", "Paid", "Balance", "Status"]
    table_data = [headers] + payments_data
    
    # Totals
    table_data.append(["TOTAL", "", "", "", "", "", "", f"Rs.{round(total_net, 2)}", f"Rs.{round(total_paid, 2)}", f"Rs.{round(total_balance, 2)}", ""])
    
    col_widths = [18*mm, 22*mm, 22*mm, 14*mm, 12*mm, 16*mm, 16*mm, 18*mm, 16*mm, 18*mm, 14*mm]
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
        ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
    ]
    
    # Alternating rows and status colors
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
        if payments_data[i-1][-1] == "Paid":
            style_commands.append(('BACKGROUND', (-1, i), (-1, i), colors.HexColor('#D1FAE5')))
        elif payments_data[i-1][-1] == "Pending":
            style_commands.append(('BACKGROUND', (-1, i), (-1, i), colors.HexColor('#FEE2E2')))
    
    main_table.setStyle(TableStyle(style_commands))
    elements.append(main_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"truck_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/entries/bulk-delete")
async def bulk_delete_entries(entry_ids: List[str], username: str = "", role: str = ""):
    """Bulk delete entries"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can bulk delete")
    
    result = await db.mill_entries.delete_many({"id": {"$in": entry_ids}})
    return {"message": f"{result.deleted_count} entries deleted successfully", "deleted_count": result.deleted_count}


# ============ MANDI TARGET ENDPOINTS ============

@router.get("/mandi-targets", response_model=List[MandiTarget])
async def get_mandi_targets(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get all mandi targets"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).sort("mandi_name", 1).to_list(100)
    return targets


@router.post("/mandi-targets", response_model=MandiTarget)
async def create_mandi_target(input: MandiTargetCreate, username: str = "", role: str = ""):
    """Create a new mandi target (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin target set kar sakta hai")
    
    # Check if target already exists for this mandi + kms_year + season
    existing = await db.mandi_targets.find_one({
        "mandi_name": input.mandi_name,
        "kms_year": input.kms_year,
        "season": input.season
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(status_code=400, detail=f"{input.mandi_name} ka target already set hai is FY Year aur Season ke liye")
    
    # Calculate expected total
    expected_total = round(input.target_qntl + (input.target_qntl * input.cutting_percent / 100), 2)
    
    target_obj = MandiTarget(
        mandi_name=input.mandi_name,
        target_qntl=input.target_qntl,
        cutting_percent=input.cutting_percent,
        expected_total=expected_total,
        base_rate=input.base_rate,
        cutting_rate=input.cutting_rate,
        kms_year=input.kms_year,
        season=input.season,
        created_by=username
    )
    
    doc = target_obj.model_dump()
    await db.mandi_targets.insert_one(doc)
    return target_obj


@router.put("/mandi-targets/{target_id}", response_model=MandiTarget)
async def update_mandi_target(target_id: str, input: MandiTargetUpdate, username: str = "", role: str = ""):
    """Update a mandi target (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin target update kar sakta hai")
    
    existing = await db.mandi_targets.find_one({"id": target_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Target not found")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    merged = {**existing, **update_data}
    
    # Recalculate expected total
    merged["expected_total"] = round(merged["target_qntl"] + (merged["target_qntl"] * merged["cutting_percent"] / 100), 2)
    
    await db.mandi_targets.update_one({"id": target_id}, {"$set": merged})
    updated = await db.mandi_targets.find_one({"id": target_id}, {"_id": 0})
    return updated


@router.delete("/mandi-targets/{target_id}")
async def delete_mandi_target(target_id: str, username: str = "", role: str = ""):
    """Delete a mandi target (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin target delete kar sakta hai")
    
    result = await db.mandi_targets.delete_one({"id": target_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Target not found")
    return {"message": "Target deleted successfully"}


@router.get("/mandi-targets/summary", response_model=List[MandiTargetSummary])
async def get_mandi_target_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get mandi target vs achieved summary for dashboard"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    
    summaries = []
    for target in targets:
        # Get achieved sum for this mandi
        entry_query = {
            "mandi_name": target["mandi_name"],
            "kms_year": target["kms_year"],
            "season": target["season"]
        }
        
        pipeline = [
            {"$match": entry_query},
            {"$group": {"_id": None, "total_final_w": {"$sum": "$final_w"}}}
        ]
        
        result = await db.mill_entries.aggregate(pipeline).to_list(1)
        achieved_kg = result[0]["total_final_w"] if result else 0
        achieved_qntl = round(achieved_kg / 100, 2)
        
        expected_total = target["expected_total"]
        pending_qntl = round(max(0, expected_total - achieved_qntl), 2)
        progress_percent = round((achieved_qntl / expected_total * 100) if expected_total > 0 else 0, 1)
        
        # Calculate agent payment amounts
        target_qntl = target["target_qntl"]
        cutting_qntl = round(target_qntl * target["cutting_percent"] / 100, 2)
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        target_amount = round(target_qntl * base_rate, 2)
        cutting_amount = round(cutting_qntl * cutting_rate, 2)
        total_agent_amount = round(target_amount + cutting_amount, 2)
        
        summaries.append(MandiTargetSummary(
            id=target["id"],
            mandi_name=target["mandi_name"],
            target_qntl=target_qntl,
            cutting_percent=target["cutting_percent"],
            expected_total=expected_total,
            achieved_qntl=achieved_qntl,
            pending_qntl=pending_qntl,
            progress_percent=progress_percent,
            base_rate=base_rate,
            cutting_rate=cutting_rate,
            target_amount=target_amount,
            cutting_qntl=cutting_qntl,
            cutting_amount=cutting_amount,
            total_agent_amount=total_agent_amount,
            kms_year=target["kms_year"],
            season=target["season"]
        ))
    
    return summaries


# ============ DASHBOARD ENDPOINTS ============

@router.get("/dashboard/agent-totals")
async def get_agent_totals(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get agent-wise totals for bar chart"""
    match_query = {}
    if kms_year:
        match_query["kms_year"] = kms_year
    if season:
        match_query["season"] = season
    
    pipeline = []
    if match_query:
        pipeline.append({"$match": match_query})
    
    pipeline.extend([
        {
            "$group": {
                "_id": "$agent_name",
                "total_qntl": {"$sum": "$qntl"},
                "total_final_w_kg": {"$sum": "$final_w"},
                "total_entries": {"$sum": 1},
                "total_bag": {"$sum": "$bag"}
            }
        },
        {"$sort": {"total_final_w_kg": -1}}
    ])
    
    results = await db.mill_entries.aggregate(pipeline).to_list(50)
    
    agent_totals = []
    for r in results:
        if r["_id"]:  # Skip empty agent names
            agent_totals.append({
                "agent_name": r["_id"],
                "total_qntl": round(r["total_qntl"], 2),
                "total_final_w": round(r["total_final_w_kg"] / 100, 2),
                "total_entries": r["total_entries"],
                "total_bag": r["total_bag"]
            })
    
    return {"agent_totals": agent_totals}


@router.get("/dashboard/date-range-totals")
async def get_date_range_totals(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Get totals for a date range (for date filter reporting)"""
    match_query = {}
    
    if start_date and end_date:
        match_query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        match_query["date"] = {"$gte": start_date}
    elif end_date:
        match_query["date"] = {"$lte": end_date}
    
    if kms_year:
        match_query["kms_year"] = kms_year
    if season:
        match_query["season"] = season
    
    pipeline = []
    if match_query:
        pipeline.append({"$match": match_query})
    
    pipeline.append({
        "$group": {
            "_id": None,
            "total_kg": {"$sum": "$kg"},
            "total_qntl": {"$sum": "$qntl"},
            "total_bag": {"$sum": "$bag"},
            "total_final_w": {"$sum": "$final_w"},
            "total_entries": {"$sum": 1}
        }
    })
    
    result = await db.mill_entries.aggregate(pipeline).to_list(1)
    
    if result:
        data = result[0]
        return {
            "total_kg": round(data["total_kg"], 2),
            "total_qntl": round(data["total_qntl"], 2),
            "total_bag": data["total_bag"],
            "total_final_w": round(data["total_final_w"] / 100, 2),
            "total_entries": data["total_entries"],
            "start_date": start_date,
            "end_date": end_date
        }
    
    return {
        "total_kg": 0,
        "total_qntl": 0,
        "total_bag": 0,
        "total_final_w": 0,
        "total_entries": 0,
        "start_date": start_date,
        "end_date": end_date
    }


