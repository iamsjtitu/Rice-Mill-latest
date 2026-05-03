from fastapi import APIRouter, HTTPException, Request
from database import db
from typing import Optional
from pydantic import BaseModel, Field, ConfigDict
from datetime import datetime, timezone
import uuid

router = APIRouter()


def _vlbl(v: dict) -> str:
    """Return display label for a sale voucher: explicit label or fallback S-NNN."""
    lbl = (v.get('voucher_no_label') or '').strip()
    if lbl:
        return lbl
    n = v.get('voucher_no', 0) or 0
    return f"S-{n:03d}"

# ============ GST SETTINGS ============

class GSTSettings(BaseModel):
    cgst_percent: float = 0
    sgst_percent: float = 0
    igst_percent: float = 0

@router.get("/gst-settings")
async def get_gst_settings():
    settings = await db.settings.find_one({"key": "gst"}, {"_id": 0})
    if not settings:
        return {"cgst_percent": 0, "sgst_percent": 0, "igst_percent": 0}
    return {"cgst_percent": settings.get("cgst_percent", 0), "sgst_percent": settings.get("sgst_percent", 0), "igst_percent": settings.get("igst_percent", 0)}

@router.put("/gst-settings")
async def update_gst_settings(data: GSTSettings):
    await db.settings.update_one(
        {"key": "gst"},
        {"$set": {"key": "gst", "cgst_percent": data.cgst_percent, "sgst_percent": data.sgst_percent, "igst_percent": data.igst_percent}},
        upsert=True
    )
    return {"success": True, "message": "GST settings updated"}

# ============ SALE BOOK ============

class SaleItemCreate(BaseModel):
    item_name: str
    quantity: float = 0
    rate: float = 0
    unit: str = "KG"
    hsn_code: str = ""
    gst_percent: float = 0
    oil_percent: float = 0

class SaleVoucherCreate(BaseModel):
    date: str
    party_name: str
    voucher_no_label: str = ""  # Editable display label e.g. S-001 (auto-generated if empty)
    invoice_no: str = ""
    bill_book: str = ""
    destination: str = ""
    buyer_gstin: str = ""
    buyer_address: str = ""
    items: list[SaleItemCreate] = []
    gst_type: str = "none"
    cgst_percent: float = 0
    sgst_percent: float = 0
    igst_percent: float = 0
    truck_no: str = ""
    rst_no: str = ""
    remark: str = ""
    cash_paid: float = 0
    diesel_paid: float = 0
    advance: float = 0
    eway_bill_no: str = ""
    kms_year: str = ""
    season: str = ""

class SaleVoucher(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    voucher_no: int = 0
    # Display label like `S-001`. Editable by user; auto-generated if empty.
    voucher_no_label: str = ""
    invoice_no: str = ""
    bill_book: str = ""
    destination: str = ""
    date: str = ""
    party_name: str = ""
    buyer_gstin: str = ""
    buyer_address: str = ""
    items: list = []
    subtotal: float = 0
    gst_type: str = "none"
    cgst_percent: float = 0
    sgst_percent: float = 0
    igst_percent: float = 0
    cgst_amount: float = 0
    sgst_amount: float = 0
    igst_amount: float = 0
    total: float = 0
    truck_no: str = ""
    rst_no: str = ""
    remark: str = ""
    cash_paid: float = 0
    diesel_paid: float = 0
    advance: float = 0
    eway_bill_no: str = ""
    paid_amount: float = 0
    balance: float = 0
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


@router.get("/sale-book/stock-items")
async def get_stock_items(kms_year: Optional[str] = None, season: Optional[str] = None):
    from utils.stock_calculator import (
        calc_rice_produced, calc_govt_delivered, calc_pvt_rice_sold,
        calc_sale_voucher_items, calc_purchase_voucher_items,
        calc_byproduct_produced, calc_byproduct_sold, calc_frk_in, BY_PRODUCTS
    )
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    # ===== FETCH OPENING STOCK =====
    ob = {}
    if kms_year:
        ob_doc = await db.opening_stock.find_one({"kms_year": kms_year}, {"_id": 0})
        if ob_doc:
            ob = ob_doc.get("stocks", {})
    ob_usna = float(ob.get("rice_usna", ob.get("rice", 0)))
    ob_raw = float(ob.get("rice_raw", 0))
    ob_bran = float(ob.get("bran", 0))
    ob_kunda = float(ob.get("kunda", 0))
    ob_broken = float(ob.get("broken", 0))
    ob_kanki = float(ob.get("kanki", 0))
    ob_husk = float(ob.get("husk", 0))
    ob_frk = float(ob.get("frk", 0))

    milling = await db.milling_entries.find(query, {"_id": 0}).to_list(10000)
    dc = await db.dc_entries.find(query, {"_id": 0}).to_list(10000)
    pvt_sales = await db.rice_sales.find(query, {"_id": 0}).to_list(10000)
    sale_vouchers = await db.sale_vouchers.find(query, {"_id": 0}).to_list(10000)
    purchase_vouchers = await db.purchase_vouchers.find(query, {"_id": 0}).to_list(10000)
    bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(10000)

    usna_produced = calc_rice_produced(milling, 'usna')
    raw_produced = calc_rice_produced(milling, 'raw')
    govt_delivered = calc_govt_delivered(dc)
    pvt_sold_usna = calc_pvt_rice_sold(pvt_sales, 'usna')
    pvt_sold_raw = calc_pvt_rice_sold(pvt_sales, 'raw')
    sb_sold = calc_sale_voucher_items(sale_vouchers)
    pv_bought = calc_purchase_voucher_items(purchase_vouchers)
    # Dynamic by-product categories (fetch early for calc_byproduct_produced)
    cats = await db.byproduct_categories.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    if not cats:
        cats = [{"id": "bran", "name": "Bran"}, {"id": "kunda", "name": "Kunda"}, {"id": "broken", "name": "Broken"}, {"id": "kanki", "name": "Kanki"}, {"id": "husk", "name": "Husk"}]
    bp_produced = calc_byproduct_produced(milling, categories=cats)
    bp_sold_map = calc_byproduct_sold(bp_sales)

    items = []
    usna_avail = round(ob_usna + usna_produced + pv_bought.get("Rice (Usna)", 0) - govt_delivered - pvt_sold_usna - sb_sold.get("Rice (Usna)", 0), 2)
    raw_avail = round(ob_raw + raw_produced + pv_bought.get("Rice (Raw)", 0) - pvt_sold_raw - sb_sold.get("Rice (Raw)", 0), 2)
    items.append({"name": "Rice (Usna)", "available_qntl": usna_avail, "unit": "Qntl"})
    items.append({"name": "Rice (Raw)", "available_qntl": raw_avail, "unit": "Qntl"})

    for cat in cats:
        p = cat["id"]
        display_name = cat.get("name", p.title())
        produced = bp_produced.get(p, 0)
        # Check multiple name variants for sale/purchase matching
        purchased = pv_bought.get(display_name, 0) + pv_bought.get(p.title(), 0) + pv_bought.get(p, 0)
        sold_bp = round(bp_sold_map.get(p, 0), 2)
        sold_sb = sb_sold.get(display_name, 0) + sb_sold.get(p.title(), 0) + sb_sold.get(p, 0)
        item_ob = float(ob.get(p, 0))
        avail = round(item_ob + produced + purchased - sold_bp - sold_sb, 2)
        items.append({"name": display_name, "available_qntl": avail, "unit": "Qntl"})

    frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(10000) if await db.frk_purchases.count_documents(query) > 0 else []
    frk_in = calc_frk_in(frk_purchases)
    frk_pv = pv_bought.get("FRK", 0)
    frk_sold_sb = sb_sold.get("FRK", 0)
    items.append({"name": "FRK", "available_qntl": round(ob_frk + frk_in + frk_pv - frk_sold_sb, 2), "unit": "Qntl"})

    # Custom items from Purchase Vouchers not already covered
    known_items = {"Rice (Usna)", "Rice (Raw)", "FRK"} | {cat.get("name", cat["id"].title()) for cat in cats} | {cat["id"] for cat in cats} | {cat["id"].title() for cat in cats}
    for item_name, qty in pv_bought.items():
        if item_name not in known_items and item_name:
            sold = sb_sold.get(item_name, 0)
            items.append({"name": item_name, "available_qntl": round(qty - sold, 2), "unit": "Qntl"})

    return items


@router.get("/sale-book")
async def get_sale_vouchers(kms_year: Optional[str] = None, season: Optional[str] = None,
                            party_name: Optional[str] = None, invoice_no: Optional[str] = None,
                            rst_no: Optional[str] = None, search: Optional[str] = None,
                            item_category: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if invoice_no: query["invoice_no"] = {"$regex": invoice_no, "$options": "i"}
    if rst_no: query["rst_no"] = {"$regex": rst_no, "$options": "i"}
    if item_category:
        query["items.item_name"] = {"$regex": f"^{item_category}$", "$options": "i"}
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"invoice_no": {"$regex": search, "$options": "i"}},
            {"rst_no": {"$regex": search, "$options": "i"}},
            {"truck_no": {"$regex": search, "$options": "i"}},
        ]
    vouchers = await db.sale_vouchers.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Get ledger-based paid amounts per party (includes Cash Book manual payments)
    party_names = list(set(v.get("party_name", "") for v in vouchers if v.get("party_name")))
    ledger_paid_map = {}
    if party_names:
        ledger_query = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": party_names}, "party_type": "Sale Book"}
        if kms_year: ledger_query["kms_year"] = kms_year
        if season: ledger_query["season"] = season
        ledger_txns = await db.cash_transactions.find(ledger_query, {"_id": 0}).to_list(50000)
        for lt in ledger_txns:
            pn = lt.get("category", "")
            ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lt.get("amount", 0)
    
    for v in vouchers:
        pn = v.get("party_name", "")
        if pn:
            total_paid = round(ledger_paid_map.get(pn, 0), 2)
            v["ledger_paid"] = total_paid
            v["ledger_balance"] = round((v.get("total", 0)) - total_paid, 2)
    
    return vouchers


async def _get_default_pump():
    """Get default diesel pump name"""
    pump = await db.diesel_accounts.find_one({}, {"_id": 0, "pump_name": 1}, sort=[("_id", -1)])
    return pump.get("pump_name", "Diesel Pump") if pump else "Diesel Pump"


async def _create_sale_ledger_entries(d, doc_id, vno, items, username):
    """Create all accounting entries for a sale voucher"""
    party = (d.get('party_name') or '').strip()
    cash = d.get('cash_paid', 0) or 0
    diesel = d.get('diesel_paid', 0) or 0
    advance = d.get('advance', 0) or 0
    total = d.get('total', 0) or 0
    truck = (d.get('truck_no') or '').strip()
    now_iso = datetime.now(timezone.utc).isoformat()
    base = {"kms_year": d.get('kms_year', ''), "season": d.get('season', ''), "created_by": username, "created_at": now_iso, "updated_at": now_iso}
    inv = d.get('invoice_no', '')
    desc_suffix = f" | Inv:{inv}" if inv else ""
    items_str = ', '.join(i['item_name'] for i in items)
    entries = []

    # 1. Party Ledger JAMA: total sale amount (party owes us)
    if party and total > 0:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "jama",
            "amount": total, "category": party, "party_type": "Sale Book",
            "description": f"Sale #{vno} - {items_str}{desc_suffix}",
            "reference": f"sale_voucher:{doc_id}", **base
        })

    # 2. Advance from party: Party Ledger NIKASI (reduces party debt)
    if advance > 0 and party:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": advance, "category": party, "party_type": "Sale Book",
            "description": f"Advance received - Sale #{vno}{desc_suffix}",
            "reference": f"sale_voucher_adv:{doc_id}", **base
        })
        # Cash JAMA: advance cash received (money comes into cash box)
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "cash", "txn_type": "jama",
            "amount": advance, "category": party, "party_type": "Sale Book",
            "description": f"Advance received - Sale #{vno}{desc_suffix}",
            "reference": f"sale_voucher_adv_cash:{doc_id}", **base
        })

    # 3. Cash paid to truck → Cash NIKASI (cash going out)
    if cash > 0:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "cash", "txn_type": "nikasi",
            "amount": cash, "category": truck or party, "party_type": "Truck" if truck else "Sale Book",
            "description": f"Truck cash - Sale #{vno}{desc_suffix}",
            "reference": f"sale_voucher_cash:{doc_id}", **base
        })

    # 4. Diesel paid → Diesel Pump Ledger JAMA (we owe pump) + diesel_accounts entry
    if diesel > 0:
        pump_name = await _get_default_pump()
        # Get pump_id
        pump_doc = await db.diesel_accounts.find_one({"pump_name": pump_name}, {"_id": 0, "pump_id": 1})
        pump_id = pump_doc.get("pump_id", "") if pump_doc else ""
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "jama",
            "amount": diesel, "category": pump_name, "party_type": "Diesel",
            "description": f"Diesel for truck - Sale #{vno} - {party}{desc_suffix}",
            "reference": f"sale_voucher_diesel:{doc_id}", **base
        })
        # Also create entry in diesel_accounts collection (txn_type=debit so it counts in summary)
        diesel_entry = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "pump_id": pump_id, "pump_name": pump_name,
            "truck_no": truck, "agent_name": party,
            "amount": diesel, "txn_type": "debit",
            "description": f"Diesel for Sale #{vno} - {party}{desc_suffix}",
            "reference": f"sale_voucher_diesel:{doc_id}",
            **base
        }
        await db.diesel_accounts.insert_one(diesel_entry)

    # 5. Truck cash+diesel → Truck Ledger NIKASI (deductions from future bhada)
    if cash > 0 and truck:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": cash, "category": truck, "party_type": "Truck",
            "description": f"Truck cash deduction - Sale #{vno}{desc_suffix}",
            "reference": f"sale_truck_cash:{doc_id}", **base
        })
    if diesel > 0 and truck:
        entries.append({
            "id": str(uuid.uuid4()), "date": d.get('date', ''), "account": "ledger", "txn_type": "nikasi",
            "amount": diesel, "category": truck, "party_type": "Truck",
            "description": f"Truck diesel deduction - Sale #{vno}{desc_suffix}",
            "reference": f"sale_truck_diesel:{doc_id}", **base
        })
    # Create truck_payments entry (rate will be set later by admin)
    truck_total = cash + diesel
    if truck_total > 0 and truck:
        truck_entry = {
            "entry_id": doc_id,
            "truck_no": truck, "date": d.get('date', ''),
            "cash_taken": cash, "diesel_taken": diesel,
            "gross_amount": 0, "deductions": truck_total,
            "net_amount": 0, "paid_amount": 0,
            "balance_amount": 0, "status": "pending",
            "source": "Sale Book",
            "description": f"Sale #{vno} - {party}{desc_suffix}",
            "reference": f"sale_voucher_truck:{doc_id}",
            **base
        }
        await db.truck_payments.insert_one(truck_entry)

    for entry in entries:
        await db.cash_transactions.insert_one(entry)

    # Create local_party_accounts entry for sale voucher (party owes us = debit)
    if party and total > 0:
        lp = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "party_name": party, "txn_type": "debit",
            "amount": total, "description": f"Sale #{vno} - {items_str}{desc_suffix}",
            "source_type": "sale_voucher", "reference": f"sale_voucher:{doc_id}",
            "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
            "created_by": username, "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.local_party_accounts.insert_one(lp)
    # If advance received, add payment entry in local_party
    if advance > 0 and party:
        lp_adv = {
            "id": str(uuid.uuid4()), "date": d.get('date', ''),
            "party_name": party, "txn_type": "payment",
            "amount": advance, "description": f"Advance received - Sale #{vno}{desc_suffix}",
            "source_type": "sale_voucher_advance", "reference": f"sale_voucher_adv:{doc_id}",
            "kms_year": d.get('kms_year', ''), "season": d.get('season', ''),
            "created_by": username, "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.local_party_accounts.insert_one(lp_adv)


def _compute_sale_gst(d):
    """Compute per-item GST and voucher totals. Modifies d in-place."""
    raw_items = d.get('items', [])
    gst_type = d.get('gst_type', 'none')
    items = []
    subtotal = 0
    total_gst = 0
    for item in raw_items:
        qty = item.get('quantity', 0) or 0
        rate = item.get('rate', 0) or 0
        amount = round(qty * rate, 2)
        gst_pct = item.get('gst_percent', 0) or 0
        item_gst = round(amount * gst_pct / 100, 2) if gst_type != 'none' else 0
        items.append({**item, "amount": amount, "gst_amount": item_gst})
        subtotal += amount
        total_gst += item_gst
    d['items'] = items
    d['subtotal'] = round(subtotal, 2)
    if gst_type == 'cgst_sgst':
        d['cgst_amount'] = round(total_gst / 2, 2)
        d['sgst_amount'] = round(total_gst / 2, 2)
        d['igst_amount'] = 0
    elif gst_type == 'igst':
        d['cgst_amount'] = 0
        d['sgst_amount'] = 0
        d['igst_amount'] = round(total_gst, 2)
    else:
        d['cgst_amount'] = 0
        d['sgst_amount'] = 0
        d['igst_amount'] = 0
    total = round(subtotal + d['cgst_amount'] + d['sgst_amount'] + d['igst_amount'], 2)
    d['total'] = total
    advance = d.get('advance', 0) or 0
    d['paid_amount'] = round(advance, 2)
    d['balance'] = round(total - advance, 2)
    return items


@router.get("/sale-book/next-voucher-label")
async def next_sale_voucher_label():
    """Preview the next sale voucher label `S-NNN` (zero-padded to 3 digits).
    Sequence is based on internal integer voucher_no. If user creates with custom
    voucher_no_label, the int sequence still continues independently."""
    last = await db.sale_vouchers.find_one(sort=[("voucher_no", -1)], projection={"_id": 0, "voucher_no": 1})
    n = (last.get('voucher_no', 0) if last else 0) + 1
    return {"voucher_no_label": f"S-{n:03d}", "voucher_no": n}


@router.post("/sale-book")
async def create_sale_voucher(input: SaleVoucherCreate, username: str = "", role: str = ""):
    d = input.model_dump()
    items = _compute_sale_gst(d)
    d['created_by'] = username
    
    last = await db.sale_vouchers.find_one(sort=[("voucher_no", -1)], projection={"_id": 0, "voucher_no": 1})
    d['voucher_no'] = (last.get('voucher_no', 0) if last else 0) + 1
    # Auto-generate display label if user didn't provide a custom one.
    if not (d.get('voucher_no_label') or '').strip():
        d['voucher_no_label'] = f"S-{d['voucher_no']:03d}"
    
    obj = SaleVoucher(**d)
    doc = obj.model_dump()
    await db.sale_vouchers.insert_one(doc)
    doc.pop('_id', None)
    
    await _create_sale_ledger_entries(d, doc.get('id', ''), d['voucher_no'], items, username)
    
    return doc


@router.delete("/sale-book/{voucher_id}")
async def delete_sale_voucher(voucher_id: str, username: str = "", role: str = ""):
    from services.edit_lock import check_edit_lock
    existing = await db.sale_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Voucher not found")
    can_edit, message = await check_edit_lock(existing, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    await db.sale_vouchers.delete_one({"id": voucher_id})
    await db.cash_transactions.delete_many({"reference": {"$regex": f"sale_voucher.*:{voucher_id}"}})
    await db.diesel_accounts.delete_many({"reference": {"$regex": f"sale_voucher.*:{voucher_id}"}})
    await db.truck_payments.delete_many({"reference": {"$regex": f"sale_voucher.*:{voucher_id}"}})
    await db.local_party_accounts.delete_many({"reference": {"$regex": f"sale_voucher.*:{voucher_id}"}})
    return {"message": "Sale voucher deleted", "id": voucher_id}


@router.post("/sale-book/delete-bulk")
async def bulk_delete_sale_vouchers(request: Request):
    data = await request.json()
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    deleted = 0
    for vid in ids:
        existing = await db.sale_vouchers.find_one({"id": vid}, {"_id": 0})
        if existing:
            await db.sale_vouchers.delete_one({"id": vid})
            await db.cash_transactions.delete_many({"reference": {"$regex": f"sale_voucher.*:{vid}"}})
            await db.diesel_accounts.delete_many({"reference": {"$regex": f"sale_voucher.*:{vid}"}})
            await db.truck_payments.delete_many({"reference": {"$regex": f"sale_voucher.*:{vid}"}})
            await db.local_party_accounts.delete_many({"reference": {"$regex": f"sale_voucher.*:{vid}"}})
            deleted += 1
    return {"message": f"{deleted} sale vouchers deleted", "deleted": deleted}


@router.put("/sale-book/{voucher_id}")
async def update_sale_voucher(voucher_id: str, input: SaleVoucherCreate, username: str = "", role: str = ""):
    from services.edit_lock import check_edit_lock
    existing = await db.sale_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Voucher not found")
    can_edit, message = await check_edit_lock(existing, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)

    d = input.model_dump()
    items = _compute_sale_gst(d)
    d['updated_at'] = datetime.now(timezone.utc).isoformat()
    # If user cleared the label field, fall back to auto S-NNN based on existing voucher_no.
    if not (d.get('voucher_no_label') or '').strip():
        d['voucher_no_label'] = f"S-{existing.get('voucher_no', 0):03d}"

    await db.sale_vouchers.update_one({"id": voucher_id}, {"$set": d})
    
    # Delete old entries and recreate
    await db.cash_transactions.delete_many({"reference": {"$regex": f"sale_voucher.*:{voucher_id}"}})
    await db.diesel_accounts.delete_many({"reference": {"$regex": f"sale_voucher.*:{voucher_id}"}})
    await db.truck_payments.delete_many({"reference": {"$regex": f"sale_voucher.*:{voucher_id}"}})
    await db.local_party_accounts.delete_many({"reference": {"$regex": f"sale_voucher.*:{voucher_id}"}})
    vno = existing.get('voucher_no', 0)
    await _create_sale_ledger_entries(d, voucher_id, vno, items, username)
    
    updated = await db.sale_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    return updated


# ============ SALE BOOK PDF EXPORT (A4 Fit, Professional) ============

@router.get("/sale-book/export/pdf")
async def export_sale_book_pdf(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None,
                                date_from: Optional[str] = None, date_to: Optional[str] = None,
                                item_category: Optional[str] = None, party_name: Optional[str] = None):
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles; from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import io

    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    # v104.44.40 — Date range + party + category filter support for export
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    if item_category:
        # v104.44.40 — escape regex special chars (parens like "Rice (Raw)")
        import re as _re
        query["items.item_name"] = {"$regex": f"^{_re.escape(item_category)}$", "$options": "i"}
    if party_name:
        query["party_name"] = {"$regex": party_name, "$options": "i"}
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"invoice_no": {"$regex": search, "$options": "i"}},
            {"rst_no": {"$regex": search, "$options": "i"}},
        ]
    vouchers = await db.sale_vouchers.find(query, {"_id": 0}).sort("voucher_no", 1).to_list(10000)

    # Get ledger-based paid amounts
    party_names = list(set(v.get("party_name", "") for v in vouchers if v.get("party_name")))
    ledger_paid_map = {}
    if party_names:
        ledger_q = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": party_names}, "party_type": "Sale Book"}
        if kms_year: ledger_q["kms_year"] = kms_year
        ledger_txns = await db.cash_transactions.find(ledger_q, {"_id": 0}).to_list(50000)
        for lt in ledger_txns:
            pn = lt.get("category", "")
            ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lt.get("amount", 0)

    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=6*mm, rightMargin=6*mm, topMargin=8*mm, bottomMargin=6*mm)
    elements = []
    styles = get_pdf_styles()

    from utils.branding_helper import get_pdf_company_header_from_db
    elements.extend(await get_pdf_company_header_from_db())

    meta_parts = ["Sale Book"]
    if kms_year: meta_parts.append(f"FY: {kms_year}")
    if season: meta_parts.append(season)
    meta_parts.append(f"Date: {datetime.now().strftime('%d-%m-%Y')}")
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#555555'), alignment=TA_CENTER, spaceAfter=4)
    elements.append(Paragraph(" | ".join(meta_parts), meta_style))

    cell_s = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=6, leading=8)
    cell_r = ParagraphStyle('CellR', parent=styles['Normal'], fontSize=6, leading=8, alignment=TA_RIGHT)
    cell_b = ParagraphStyle('CellB', parent=styles['Normal'], fontSize=6, leading=8, fontName='FreeSansBold')
    cell_rb = ParagraphStyle('CellRB', parent=styles['Normal'], fontSize=6, leading=8, alignment=TA_RIGHT, fontName='FreeSansBold')

    headers = ['#', 'Date', 'Inv', 'Party', 'Items', 'Truck/RST', 'Total', 'Adv', 'Cash', 'Diesel', 'Paid', 'Balance', 'Status']
    table_data = [headers]
    # Optimized for A4 landscape
    col_widths = [12*mm, 18*mm, 16*mm, 28*mm, 52*mm, 22*mm, 20*mm, 16*mm, 16*mm, 16*mm, 20*mm, 20*mm, 16*mm]

    g = {"total": 0, "adv": 0, "cash": 0, "diesel": 0, "paid": 0, "bal": 0}
    for v in vouchers:
        items_str = ', '.join(f"{i['item_name']} ({i['quantity']} Qntl)" for i in v.get('items', []))
        dp = str(v.get('date', '')).split('-')
        fd = f"{dp[2]}/{dp[1]}/{dp[0]}" if len(dp) == 3 else v.get('date', '')
        truck_rst = v.get('truck_no', '')
        if v.get('rst_no'): truck_rst += f"/{v['rst_no']}"
        total = v.get('total', 0) or 0
        pn = v.get('party_name', '')
        ledger_paid = round(ledger_paid_map.get(pn, 0), 2)
        ledger_bal = round(total - ledger_paid, 2)
        status = "Paid" if ledger_bal <= 0 and total > 0 else "Pending"
        g["total"] += total
        g["adv"] += v.get('advance', 0) or 0
        g["cash"] += v.get('cash_paid', 0) or 0
        g["diesel"] += v.get('diesel_paid', 0) or 0
        g["paid"] += ledger_paid
        g["bal"] += ledger_bal
        table_data.append([
            Paragraph(_vlbl(v), cell_s),
            Paragraph(fd, cell_s),
            Paragraph(str(v.get('invoice_no', '')), cell_s),
            Paragraph(f"<b>{pn}</b>", cell_s),
            Paragraph(items_str, cell_s),
            Paragraph(truck_rst, cell_s),
            Paragraph(f"{total:,.0f}", cell_rb),
            Paragraph(f"{v.get('advance',0) or 0:,.0f}", cell_r),
            Paragraph(f"{v.get('cash_paid',0) or 0:,.0f}", cell_r),
            Paragraph(f"{v.get('diesel_paid',0) or 0:,.0f}", cell_r),
            Paragraph(f"{ledger_paid:,.0f}", cell_r),
            Paragraph(f"{ledger_bal:,.0f}", cell_rb),
            Paragraph(status, cell_s),
        ])

    # Total row
    table_data.append([
        Paragraph(f"<b>TOTAL ({len(vouchers)})</b>", cell_b), '', '', '', '', '',
        Paragraph(f"<b>{g['total']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['adv']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['cash']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['diesel']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['paid']:,.0f}</b>", cell_rb),
        Paragraph(f"<b>{g['bal']:,.0f}</b>", cell_rb),
        '',
    ])

    tbl = RLTable(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
    ]
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    # ===== Beautiful single-line summary banner =====
    from utils.export_helpers import get_pdf_summary_banner, fmt_inr, STAT_COLORS
    page_inner_w = sum(col_widths)
    banner_stats = [
        {'label': 'TOTAL ENTRIES', 'value': str(len(vouchers)), 'color': STAT_COLORS['primary']},
        {'label': 'GROSS SALE', 'value': fmt_inr(g['total']), 'color': STAT_COLORS['gold']},
        {'label': 'ADVANCE', 'value': fmt_inr(g['adv']), 'color': STAT_COLORS['orange']},
        {'label': 'CASH PAID', 'value': fmt_inr(g['cash']), 'color': STAT_COLORS['green']},
        {'label': 'DIESEL', 'value': fmt_inr(g['diesel']), 'color': STAT_COLORS['purple']},
        {'label': 'TOTAL PAID', 'value': fmt_inr(g['paid']), 'color': STAT_COLORS['emerald']},
        {'label': 'OUTSTANDING', 'value': fmt_inr(g['bal']), 'color': STAT_COLORS['red']},
    ]
    elements.append(Spacer(1, 4))
    banner = get_pdf_summary_banner(banner_stats, total_width=page_inner_w)
    if banner:
        elements.append(banner)

    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=6, textColor=colors.HexColor('#999999'), alignment=TA_CENTER, spaceBefore=6)
    elements.append(Paragraph(f"{company} - Sale Book | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", footer_style))

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=sale_book.pdf"})


# ============ SALE BOOK EXCEL EXPORT ============

@router.get("/sale-book/export/excel")
async def export_sale_book_excel(kms_year: Optional[str] = None, season: Optional[str] = None, search: Optional[str] = None,
                                  date_from: Optional[str] = None, date_to: Optional[str] = None,
                                  item_category: Optional[str] = None, party_name: Optional[str] = None):
    from fastapi.responses import Response
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    # v104.44.40 — Date range + party + category filter support for export
    if date_from or date_to:
        query["date"] = {}
        if date_from: query["date"]["$gte"] = date_from
        if date_to: query["date"]["$lte"] = date_to
    if item_category:
        # v104.44.40 — escape regex special chars (parens like "Rice (Raw)")
        import re as _re
        query["items.item_name"] = {"$regex": f"^{_re.escape(item_category)}$", "$options": "i"}
    if party_name:
        query["party_name"] = {"$regex": party_name, "$options": "i"}
    if search:
        query["$or"] = [
            {"party_name": {"$regex": search, "$options": "i"}},
            {"invoice_no": {"$regex": search, "$options": "i"}},
            {"rst_no": {"$regex": search, "$options": "i"}},
        ]
    vouchers = await db.sale_vouchers.find(query, {"_id": 0}).sort("voucher_no", 1).to_list(10000)
    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")

    # Ledger-based paid amounts
    party_names = list(set(v.get("party_name", "") for v in vouchers if v.get("party_name")))
    ledger_paid_map = {}
    if party_names:
        ledger_q = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": party_names}, "party_type": "Sale Book"}
        if kms_year: ledger_q["kms_year"] = kms_year
        ledger_txns = await db.cash_transactions.find(ledger_q, {"_id": 0}).to_list(50000)
        for lt in ledger_txns:
            pn = lt.get("category", "")
            ledger_paid_map[pn] = ledger_paid_map.get(pn, 0) + lt.get("amount", 0)
    
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sale Book"
    
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    
    # Header
    cols = ['#', 'Date', 'Inv No.', 'Party', 'Items (Qntl)', 'Truck/RST', 'Total', 'Advance', 'Cash', 'Diesel', 'Ledger Paid', 'Balance', 'Status']
    widths = [5, 9, 9, 16, 30, 12, 10, 9, 8, 8, 10, 10, 7]
    ncols = len(cols)
    
    title = f"{company} - Sale Book / बिक्री बही"
    subtitle = f"FY: {kms_year or 'All'} | {season or 'All'} | {datetime.now().strftime('%d-%m-%Y')}"
    style_excel_title(ws, title, ncols, subtitle)
    
    for i, (col, w) in enumerate(zip(cols, widths), 1):
        ws.cell(row=4, column=i, value=col)
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w
    style_excel_header_row(ws, 4, ncols)
    
    g = {"total": 0, "adv": 0, "cash": 0, "diesel": 0, "paid": 0, "bal": 0}
    data_start = 5
    for ri, v in enumerate(vouchers, data_start):
        items_str = ', '.join(f"{i['item_name']} ({i['quantity']} Qntl)" for i in v.get('items', []))
        dp = str(v.get('date', '')).split('-')
        fd = f"{dp[2]}/{dp[1]}/{dp[0]}" if len(dp) == 3 else v.get('date', '')
        truck_rst = v.get('truck_no', '')
        if v.get('rst_no'): truck_rst += f"/{v['rst_no']}"
        total = v.get('total', 0) or 0
        pn = v.get('party_name', '')
        ledger_paid = round(ledger_paid_map.get(pn, 0), 2)
        ledger_bal = round(total - ledger_paid, 2)
        status = "Paid" if ledger_bal <= 0 and total > 0 else "Pending"
        g["total"] += total
        g["adv"] += v.get('advance', 0) or 0
        g["cash"] += v.get('cash_paid', 0) or 0
        g["diesel"] += v.get('diesel_paid', 0) or 0
        g["paid"] += ledger_paid
        g["bal"] += ledger_bal
        
        row_data = [_vlbl(v), fd, v.get('invoice_no',''), pn, items_str, truck_rst,
                    total, v.get('advance', 0) or 0, v.get('cash_paid', 0) or 0, v.get('diesel_paid', 0) or 0,
                    ledger_paid, ledger_bal, status]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci >= 7 and ci <= 12:
                cell.alignment = Alignment(horizontal='right')
                if isinstance(val, (int, float)): cell.number_format = '#,##0'
    
    if vouchers:
        style_excel_data_rows(ws, data_start, data_start + len(vouchers) - 1, ncols, cols)
    
    tr = len(vouchers) + data_start
    ws.cell(row=tr, column=1, value=f"TOTAL ({len(vouchers)} vouchers)")
    for ci, val in enumerate([g['total'], g['adv'], g['cash'], g['diesel'], g['paid'], g['bal'], ''], 7):
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.alignment = Alignment(horizontal='right')
        if isinstance(val, (int, float)): cell.number_format = '#,##0'
    style_excel_total_row(ws, tr, ncols)
    
    # ===== Beautiful single-line summary banner =====
    if vouchers:
        from utils.export_helpers import add_excel_summary_banner, fmt_inr
        sum_stats = [
            {'label': 'Total Entries', 'value': str(len(vouchers))},
            {'label': 'Gross Sale', 'value': fmt_inr(g['total'])},
            {'label': 'Advance', 'value': fmt_inr(g['adv'])},
            {'label': 'Cash Paid', 'value': fmt_inr(g['cash'])},
            {'label': 'Diesel', 'value': fmt_inr(g['diesel'])},
            {'label': 'Total Paid', 'value': fmt_inr(g['paid'])},
            {'label': 'Outstanding', 'value': fmt_inr(g['bal'])},
        ]
        add_excel_summary_banner(ws, tr + 2, ncols, sum_stats)

    # 🎯 v104.44.9 — Apply consolidated multi-record polish (sale book is multi-row)
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sale_book_{datetime.now().strftime('%Y%m%d')}.xlsx"})


# ============ OPENING BALANCE ============

class OpeningBalanceCreate(BaseModel):
    party_name: str
    party_type: str = ""
    amount: float = 0
    balance_type: str = "jama"
    kms_year: str = ""
    season: str = ""
    note: str = ""


# ============ SINGLE SALE VOUCHER INVOICE PDF ============

def _num_to_words_inr(n):
    """Convert number to Indian currency words."""
    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
            'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
            'Seventeen', 'Eighteen', 'Nineteen']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    def _chunk(num):
        if num == 0: return ''
        if num < 20: return ones[num]
        if num < 100: return tens[num // 10] + (' ' + ones[num % 10] if num % 10 else '')
        return ones[num // 100] + ' Hundred' + (' and ' + _chunk(num % 100) if num % 100 else '')
    n = int(round(n))
    if n == 0: return 'Rupees Zero only'
    parts = []
    if n >= 10000000:
        parts.append(_chunk(n // 10000000) + ' Crore'); n %= 10000000
    if n >= 100000:
        parts.append(_chunk(n // 100000) + ' Lakh'); n %= 100000
    if n >= 1000:
        parts.append(_chunk(n // 1000) + ' Thousand'); n %= 1000
    if n > 0:
        parts.append(_chunk(n))
    return 'Rupees ' + ' '.join(parts) + ' only'


@router.get("/sale-book/{voucher_id}/pdf")
async def export_single_sale_voucher_pdf(voucher_id: str):
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    import io

    v = await db.sale_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    if not v: raise HTTPException(status_code=404, detail="Voucher not found")

    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "")

    gst_co = await db.settings.find_one({"key": "gst_company"}, {"_id": 0}) or {}
    co_gstin = gst_co.get("gstin", branding.get("gstin", ""))
    co_address = gst_co.get("address", branding.get("address", ""))
    co_phone = gst_co.get("phone", branding.get("phone", ""))
    co_state = gst_co.get("state_name", "")
    co_state_code = gst_co.get("state_code", "")
    co_bank_name = gst_co.get("bank_name", "")
    co_bank_acc = gst_co.get("bank_account", "")
    co_bank_ifsc = gst_co.get("bank_ifsc", "")

    gst_type = v.get('gst_type', 'none')
    has_gst = gst_type != 'none'

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm)
    elements = []
    styles = get_pdf_styles()

    BK = colors.HexColor('#222222')
    GR = colors.HexColor('#888888')
    LN = colors.HexColor('#CCCCCC')
    LN2 = colors.HexColor('#999999')
    W = 180 * mm

    s_co = ParagraphStyle('co', fontName='FreeSansBold', fontSize=14, textColor=BK, alignment=TA_CENTER, leading=18)
    s_sub = ParagraphStyle('su', fontName='FreeSans', fontSize=8, textColor=GR, alignment=TA_CENTER, leading=10)
    s_title = ParagraphStyle('ti', fontName='FreeSansBold', fontSize=11, textColor=BK, alignment=TA_CENTER, leading=14, spaceBefore=2, spaceAfter=2)
    s_lbl = ParagraphStyle('lb', fontName='FreeSans', fontSize=8, textColor=GR, leading=10)
    s_val = ParagraphStyle('vl', fontName='FreeSansBold', fontSize=9, textColor=BK, leading=12)
    s_hd = ParagraphStyle('hd', fontName='FreeSansBold', fontSize=8.5, textColor=BK, leading=11)
    s_hd_r = ParagraphStyle('hdr', fontName='FreeSansBold', fontSize=8.5, textColor=BK, leading=11, alignment=TA_RIGHT)
    s_c = ParagraphStyle('cl', fontName='FreeSans', fontSize=8.5, textColor=BK, leading=11)
    s_cr = ParagraphStyle('cr', fontName='FreeSans', fontSize=8.5, textColor=BK, leading=11, alignment=TA_RIGHT)
    s_cb = ParagraphStyle('cb', fontName='FreeSansBold', fontSize=9, textColor=BK, leading=12, alignment=TA_RIGHT)
    s_tot_l = ParagraphStyle('tl', fontName='FreeSansBold', fontSize=9, textColor=BK, alignment=TA_RIGHT, leading=12)
    s_tot_v = ParagraphStyle('tv', fontName='FreeSansBold', fontSize=10, textColor=BK, alignment=TA_RIGHT, leading=13)
    s_foot = ParagraphStyle('ft', fontName='FreeSans', fontSize=7, textColor=GR, alignment=TA_CENTER, leading=9)

    # ========== COMPANY HEADER ==========
    elements.append(Paragraph(company, s_co))
    if tagline:
        elements.append(Paragraph(tagline, s_sub))
    if co_address:
        elements.append(Paragraph(co_address, s_sub))
    parts = []
    if co_gstin: parts.append(f"GSTIN: {co_gstin}")
    if co_phone: parts.append(f"Ph: {co_phone}")
    if co_state: parts.append(f"State: {co_state} ({co_state_code})")
    if parts:
        elements.append(Paragraph(" | ".join(parts), s_sub))
    elements.append(Spacer(1, 4*mm))

    # ========== TITLE ==========
    title_text = "Tax Invoice" if has_gst else "Sale Invoice"
    elements.append(Paragraph(f"<b>{title_text}</b>", s_title))
    elements.append(Spacer(1, 3*mm))

    # ========== DETAILS (2-col grid) ==========
    dp = str(v.get('date', '')).split('-')
    date_str = f"{dp[2]}/{dp[1]}/{dp[0]}" if len(dp) == 3 else v.get('date', '')

    def _row(l1, v1, l2, v2):
        return [Paragraph(f'<b>{l1}:</b>', s_lbl), Paragraph(str(v1), s_val),
                Paragraph(f'<b>{l2}:</b>', s_lbl), Paragraph(str(v2), s_val)]

    det = [
        _row("Invoice No", v.get('invoice_no', '') or _vlbl(v), "Date", date_str),
        _row("Party Name", v.get('party_name', ''), "Voucher No", _vlbl(v)),
    ]
    if v.get('buyer_gstin') or v.get('buyer_address'):
        det.append(_row("Buyer GSTIN", v.get('buyer_gstin', '-'), "Buyer Address", v.get('buyer_address', '-')))
    det.append(_row("Truck No", v.get('truck_no', '') or '-', "RST No", v.get('rst_no', '') or '-'))
    if v.get('eway_bill_no'):
        det.append(_row("E-Way Bill", v['eway_bill_no'], "", ""))

    det_tbl = RLTable(det, colWidths=[28*mm, 60*mm, 28*mm, 60*mm])
    det_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 3), ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, LN)]))
    elements.append(det_tbl)
    elements.append(Spacer(1, 5*mm))

    # ========== INVOICE ITEMS TABLE ==========
    elements.append(Paragraph('<b>Invoice Items</b>', ParagraphStyle('it', fontName='FreeSansBold', fontSize=11, textColor=BK, spaceAfter=3)))

    if has_gst:
        hdrs = ['Description', 'Qty', 'Unit', 'Rate', 'HSN Code', 'Total']
        h_row = [Paragraph(f'<b>{h}</b>', s_hd if i < 4 else (s_hd_r if i >= 4 else s_hd)) for i, h in enumerate(hdrs)]
        items_data = [h_row]
        cw = [52*mm, 18*mm, 14*mm, 22*mm, 24*mm, 32*mm]
    else:
        hdrs = ['Description', 'Quantity', 'Unit', 'Rate', 'Total']
        h_row = [Paragraph(f'<b>{h}</b>', s_hd if i < 3 else s_hd_r) for i, h in enumerate(hdrs)]
        items_data = [h_row]
        cw = [62*mm, 22*mm, 16*mm, 28*mm, 34*mm]

    for item in v.get('items', []):
        qty = item.get('quantity', 0) or 0
        rate = item.get('rate', 0) or 0
        amt = round(qty * rate, 2)
        if has_gst:
            items_data.append([
                Paragraph(item.get('item_name', ''), s_c),
                Paragraph(f"{qty:g}", s_c), Paragraph(item.get('unit', 'Qntl'), s_c),
                Paragraph(f"{rate:,.2f}", s_cr),
                Paragraph(item.get('hsn_code', ''), s_c),
                Paragraph(f"\u20B9{amt:,.2f} INR", s_cb)
            ])
        else:
            items_data.append([
                Paragraph(item.get('item_name', ''), s_c),
                Paragraph(f"{qty:g}", s_c), Paragraph(item.get('unit', 'Qntl'), s_c),
                Paragraph(f"{rate:,.2f}", s_cr),
                Paragraph(f"\u20B9{amt:,.2f} INR", s_cb)
            ])

    # Sub Total row
    subtotal = v.get('subtotal', 0) or 0
    ncols = len(hdrs)
    sub_row = [''] * ncols
    sub_row[1] = Paragraph('<b>Sub Total</b>', s_tot_l)
    sub_row[-1] = Paragraph(f'<b>\u20B9{subtotal:,.2f} INR</b>', s_tot_v)
    items_data.append(sub_row)

    items_tbl = RLTable(items_data, colWidths=cw)
    tbl_s = [
        ('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
        ('LINEBELOW', (0,0), (-1,0), 1.2, BK),
        ('LINEBELOW', (0,1), (-1,-2), 0.3, LN),
        ('LINEABOVE', (0,-1), (-1,-1), 1.2, BK),
        ('LINEBELOW', (0,-1), (-1,-1), 0.3, LN),
        ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('SPAN', (0,-1), (-2 if ncols > 2 else 0, -1)),
    ]
    # Merge sub total label across columns
    if ncols == 6:
        tbl_s.append(('SPAN', (0,-1), (3,-1)))
    elif ncols == 5:
        tbl_s.append(('SPAN', (0,-1), (2,-1)))
    items_tbl.setStyle(TableStyle(tbl_s))
    elements.append(items_tbl)
    elements.append(Spacer(1, 2*mm))

    # ========== GST SECTION (bordered box like reference) ==========
    if has_gst:
        cgst_val = v.get('cgst_amount', 0) or 0
        sgst_val = v.get('sgst_amount', 0) or 0
        igst_val = v.get('igst_amount', 0) or 0
        total_gst = cgst_val + sgst_val + igst_val

        # Build GST mini-table with CGST | SGST | IGST columns
        gst_hdr_cells = []
        gst_val_cells = []
        if gst_type == 'cgst_sgst':
            # Compute average rate per item for display
            items_list = v.get('items', [])
            avg_pct = sum(i.get('gst_percent', 0) for i in items_list) / max(len(items_list), 1) if items_list else 0
            half_pct = avg_pct / 2
            gst_hdr_cells = [
                Paragraph(f'<b>CGST ({half_pct:g}%)</b>', ParagraphStyle('gh', fontName='FreeSansBold', fontSize=8, alignment=TA_CENTER, leading=10)),
                Paragraph(f'<b>SGST ({half_pct:g}%)</b>', ParagraphStyle('gh2', fontName='FreeSansBold', fontSize=8, alignment=TA_CENTER, leading=10)),
                Paragraph('<b>IGST</b>', ParagraphStyle('gh3', fontName='FreeSansBold', fontSize=8, alignment=TA_CENTER, leading=10)),
            ]
            gst_val_cells = [
                Paragraph(f'\u20B9{cgst_val:,.2f} INR', ParagraphStyle('gv', fontName='FreeSans', fontSize=8.5, alignment=TA_CENTER, leading=11)),
                Paragraph(f'\u20B9{sgst_val:,.2f} INR', ParagraphStyle('gv2', fontName='FreeSans', fontSize=8.5, alignment=TA_CENTER, leading=11)),
                Paragraph('-', ParagraphStyle('gv3', fontName='FreeSans', fontSize=8.5, alignment=TA_CENTER, leading=11)),
            ]
        else:
            items_list = v.get('items', [])
            avg_pct = sum(i.get('gst_percent', 0) for i in items_list) / max(len(items_list), 1) if items_list else 0
            gst_hdr_cells = [
                Paragraph('<b>CGST</b>', ParagraphStyle('gh', fontName='FreeSansBold', fontSize=8, alignment=TA_CENTER, leading=10)),
                Paragraph('<b>SGST</b>', ParagraphStyle('gh2', fontName='FreeSansBold', fontSize=8, alignment=TA_CENTER, leading=10)),
                Paragraph(f'<b>IGST ({avg_pct:g}%)</b>', ParagraphStyle('gh3', fontName='FreeSansBold', fontSize=8, alignment=TA_CENTER, leading=10)),
            ]
            gst_val_cells = [
                Paragraph('-', ParagraphStyle('gv', fontName='FreeSans', fontSize=8.5, alignment=TA_CENTER, leading=11)),
                Paragraph('-', ParagraphStyle('gv2', fontName='FreeSans', fontSize=8.5, alignment=TA_CENTER, leading=11)),
                Paragraph(f'\u20B9{igst_val:,.2f} INR', ParagraphStyle('gv3', fontName='FreeSans', fontSize=8.5, alignment=TA_CENTER, leading=11)),
            ]

        gst_inner = RLTable([gst_hdr_cells, gst_val_cells], colWidths=[32*mm, 32*mm, 32*mm])
        gst_inner.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
            ('GRID', (0,0), (-1,-1), 0.5, LN2),
            ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))

        # Outer GST row: [blank] [GST label] [inner table] [total gst]
        gst_outer = RLTable([
            ['', Paragraph('<b>GST</b>', ParagraphStyle('gl', fontName='FreeSansBold', fontSize=9, alignment=TA_RIGHT, leading=12)),
             gst_inner,
             Paragraph(f'<b>\u20B9{total_gst:,.2f} INR</b>', s_tot_v)]
        ], colWidths=[18*mm, 28*mm, 100*mm, 34*mm])
        gst_outer.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
            ('LINEABOVE', (0,0), (-1,0), 0.3, LN),
            ('LINEBELOW', (0,0), (-1,0), 0.3, LN),
            ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
        elements.append(gst_outer)
        elements.append(Spacer(1, 2*mm))

    # ========== TOTAL AMOUNT INCL. GST ==========
    total = v.get('total', 0) or 0
    total_row = RLTable([
        [Paragraph('<b>Total Amount Incl. GST :</b>', s_tot_l), Paragraph(f'<b>\u20B9{total:,.2f} INR</b>', s_tot_v)]
    ], colWidths=[130*mm, 50*mm])
    total_row.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
        ('LINEABOVE', (0,0), (-1,0), 1.2, BK), ('LINEBELOW', (0,0), (-1,0), 0.3, LN),
        ('TOPPADDING', (0,0), (-1,-1), 8), ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elements.append(total_row)

    # ========== AMOUNT IN WORDS ==========
    words_text = _num_to_words_inr(total)
    words_row = RLTable([
        [Paragraph('<b>Total Amount Incl. GST (in words) :</b>', s_tot_l),
         Paragraph(f'<b>{words_text}</b>', ParagraphStyle('aw', fontName='FreeSansBold', fontSize=9, alignment=TA_RIGHT, textColor=BK, leading=12))]
    ], colWidths=[90*mm, 90*mm])
    words_row.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
        ('LINEBELOW', (0,0), (-1,0), 0.3, LN),
        ('TOPPADDING', (0,0), (-1,-1), 8), ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elements.append(words_row)

    # ========== DEDUCTIONS ==========
    advance = v.get('advance', 0) or 0
    cash = v.get('cash_paid', 0) or 0
    diesel = v.get('diesel_paid', 0) or 0
    balance = v.get('balance', 0) or 0

    deduction_rows = []
    if advance > 0:
        deduction_rows.append([Paragraph('<b>Advance (Party se) :</b>', s_tot_l), Paragraph(f'\u20B9{advance:,.2f} INR', s_cb)])
    if cash > 0:
        deduction_rows.append([Paragraph('<b>Cash (Truck ko) :</b>', s_tot_l), Paragraph(f'\u20B9{cash:,.2f} INR', s_cb)])
    if diesel > 0:
        deduction_rows.append([Paragraph('<b>Diesel (Pump se) :</b>', s_tot_l), Paragraph(f'\u20B9{diesel:,.2f} INR', s_cb)])
    deduction_rows.append([Paragraph('<b>Balance Due :</b>', s_tot_l), Paragraph(f'<b>\u20B9{balance:,.2f} INR</b>', s_tot_v)])

    ded_tbl = RLTable(deduction_rows, colWidths=[130*mm, 50*mm])
    ded_style = [('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
        ('LINEBELOW', (0,0), (-1,-1), 0.3, LN),
        ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEABOVE', (0,-1), (-1,-1), 1.2, BK)]
    ded_tbl.setStyle(TableStyle(ded_style))
    elements.append(ded_tbl)

    # ========== BANK DETAILS ==========
    if has_gst and co_bank_name:
        elements.append(Spacer(1, 5*mm))
        bk_s = ParagraphStyle('bk', fontName='FreeSans', fontSize=8, textColor=GR, leading=11)
        elements.append(Paragraph(f'<b>Bank Details:</b> {co_bank_name} | A/c: {co_bank_acc} | IFSC: {co_bank_ifsc}', bk_s))

    if v.get('remark'):
        elements.append(Spacer(1, 3*mm))
        elements.append(Paragraph(f'<b>Remark:</b> {v["remark"]}', s_lbl))

    # ========== SIGNATURES ==========
    elements.append(Spacer(1, 18*mm))
    sig_s = ParagraphStyle('sg', fontName='FreeSans', fontSize=8.5, alignment=TA_CENTER, textColor=GR)
    sig_b = ParagraphStyle('sgb', fontName='FreeSansBold', fontSize=8.5, alignment=TA_CENTER, textColor=BK)
    sig_tbl = RLTable([
        [Paragraph("Receiver's Signature", sig_s), '', Paragraph(f'For {company}', sig_b)],
        ['', '', ''],
        [Paragraph('_____________________', sig_s), '', Paragraph('_____________________', sig_s)],
        ['', '', Paragraph('Authorized Signatory', sig_s)],
    ], colWidths=[55*mm, 60*mm, 55*mm])
    sig_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ('TOPPADDING', (0,0), (-1,-1), 1), ('BOTTOMPADDING', (0,0), (-1,-1), 1)]))
    elements.append(sig_tbl)

    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph('This is a computer generated invoice.', s_foot))

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=sale_invoice_{v.get('voucher_no','')}.pdf"})


@router.get("/opening-balances")
async def get_opening_balances(kms_year: Optional[str] = None):
    query = {"is_opening_balance": True}
    if kms_year: query["kms_year"] = kms_year
    return await db.cash_transactions.find(query, {"_id": 0}).sort("category", 1).to_list(10000)

@router.post("/opening-balances")
async def create_opening_balance(data: OpeningBalanceCreate, username: str = "", role: str = ""):
    entry = {
        "id": str(uuid.uuid4()),
        "date": "",
        "account": "ledger",
        "txn_type": data.balance_type,
        "amount": abs(data.amount),
        "category": data.party_name.strip(),
        "party_type": data.party_type or "Cash Party",
        "description": f"Opening Balance - {data.note}" if data.note else "Opening Balance",
        "reference": "opening_balance",
        "is_opening_balance": True,
        "kms_year": data.kms_year,
        "season": data.season,
        "created_by": username,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.cash_transactions.insert_one(entry)
    entry.pop('_id', None)
    return entry

@router.delete("/opening-balances/{entry_id}")
async def delete_opening_balance(entry_id: str, username: str = "", role: str = ""):
    existing = await db.cash_transactions.find_one({"id": entry_id, "is_opening_balance": True}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Opening balance not found")
    await db.cash_transactions.delete_one({"id": entry_id})
    return {"message": "Opening balance deleted", "id": entry_id}


# ============ WHATSAPP SALE VOUCHER SEND ============

@router.post("/sale-book/{voucher_id}/whatsapp-send")
async def whatsapp_send_sale_voucher(voucher_id: str, request: Request):
    """Generate PDF, upload to tmpfiles.org, and send via WhatsApp."""
    import httpx, io
    body = await request.json()
    phone = body.get("phone", "").strip()

    v = await db.sale_vouchers.find_one({"id": voucher_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Voucher not found")

    branding = await db.settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    company = branding.get("company_name", "Mill Entry System")

    # Build WhatsApp text message
    gst_type = v.get('gst_type', 'none')
    has_gst = gst_type != 'none'
    date_str = v.get('date', '')
    dp = date_str.split('-')
    if len(dp) == 3:
        date_str = f"{dp[2]}/{dp[1]}/{dp[0]}"

    text = (
        f"*{company}*\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"*{'TAX INVOICE' if has_gst else 'SALE INVOICE'}*\n"
        f"Voucher: {_vlbl(v)}\n"
        f"Date: {date_str}\n"
    )
    if v.get('invoice_no'):
        text += f"Invoice No: {v['invoice_no']}\n"
    text += f"━━━━━━━━━━━━━━━━\n"
    text += f"Party: *{v.get('party_name', '')}*\n"
    if v.get('buyer_gstin'):
        text += f"GSTIN: {v['buyer_gstin']}\n"
    text += f"━━━━━━━━━━━━━━━━\n"

    for item in v.get('items', []):
        qty = item.get('quantity', 0)
        rate = item.get('rate', 0)
        amt = round(qty * rate, 2)
        text += f"{item.get('item_name', '')} | {qty}Q x Rs.{rate:,.0f} = Rs.{amt:,.0f}"
        if has_gst and item.get('gst_percent'):
            text += f" (+{item.get('gst_percent')}% GST)"
        text += "\n"

    text += f"━━━━━━━━━━━━━━━━\n"
    text += f"Taxable: Rs.{v.get('subtotal', 0):,.2f}\n"
    cgst = v.get('cgst_amount', 0)
    sgst = v.get('sgst_amount', 0)
    igst = v.get('igst_amount', 0)
    if cgst > 0:
        text += f"CGST: Rs.{cgst:,.2f}\n"
    if sgst > 0:
        text += f"SGST: Rs.{sgst:,.2f}\n"
    if igst > 0:
        text += f"IGST: Rs.{igst:,.2f}\n"
    text += f"*Grand Total: Rs.{v.get('total', 0):,.2f}*\n"
    if v.get('advance', 0) > 0:
        text += f"Advance: Rs.{v.get('advance', 0):,.2f}\n"
    text += f"*Balance: Rs.{v.get('balance', 0):,.2f}*\n"
    text += f"\nThank you\n{company}"

    # Generate PDF
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []
    styles = get_pdf_styles()
    blue = '#1a5276'

    from utils.branding_helper import get_pdf_company_header_from_db as _gph
    elements.extend(await _gph())

    gst_co = await db.settings.find_one({"key": "gst_company"}, {"_id": 0}) or {}
    co_gstin = gst_co.get("gstin", "")
    co_state = gst_co.get("state_name", "")
    co_state_code = gst_co.get("state_code", "")

    if co_gstin:
        gstin_s = ParagraphStyle('GS', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#333'))
        elements.append(Paragraph(f"GSTIN: {co_gstin} | State: {co_state} ({co_state_code})", gstin_s))
    elements.append(Spacer(1, 3*mm))

    inv_title = ParagraphStyle('IT', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor(blue), alignment=TA_CENTER, spaceBefore=0, spaceAfter=2)
    elements.append(Paragraph("TAX INVOICE" if has_gst else "SALE INVOICE", inv_title))

    label_s = ParagraphStyle('L', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#555'))
    val_s = ParagraphStyle('V', parent=styles['Normal'], fontSize=9, fontName='FreeSansBold')
    cell_s = ParagraphStyle('C', parent=styles['Normal'], fontSize=8, leading=11)
    cell_r = ParagraphStyle('CR', parent=styles['Normal'], fontSize=8, leading=11, alignment=TA_RIGHT)
    cell_rb = ParagraphStyle('CRB', parent=styles['Normal'], fontSize=9, leading=11, alignment=TA_RIGHT, fontName='FreeSansBold')
    hd_s = ParagraphStyle('H', parent=styles['Normal'], fontSize=8, leading=11, fontName='FreeSansBold', textColor=colors.white)

    info_data = [
        [Paragraph('<b>Voucher:</b>', label_s), Paragraph(_vlbl(v), val_s),
         Paragraph('<b>Date:</b>', label_s), Paragraph(date_str, val_s)],
        [Paragraph('<b>Party:</b>', label_s), Paragraph(v.get('party_name', ''), val_s),
         Paragraph('<b>Invoice:</b>', label_s), Paragraph(v.get('invoice_no', ''), val_s)],
    ]
    if v.get('buyer_gstin') or v.get('buyer_address'):
        info_data.append([Paragraph('<b>Buyer GSTIN:</b>', label_s), Paragraph(v.get('buyer_gstin', ''), val_s),
                         Paragraph('<b>Address:</b>', label_s), Paragraph(v.get('buyer_address', ''), val_s)])

    info_tbl = RLTable(info_data, colWidths=[28*mm, 55*mm, 28*mm, 55*mm])
    info_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('LINEBELOW', (0,-1), (-1,-1), 1, colors.HexColor('#ddd'))]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 4*mm))

    if has_gst:
        headers = [Paragraph('<b>S.No</b>', hd_s), Paragraph('<b>Item</b>', hd_s), Paragraph('<b>HSN</b>', hd_s),
                   Paragraph('<b>Qty</b>', hd_s), Paragraph('<b>Rate</b>', hd_s), Paragraph('<b>Taxable</b>', hd_s),
                   Paragraph('<b>GST%</b>', hd_s), Paragraph('<b>GST</b>', hd_s), Paragraph('<b>Total</b>', hd_s)]
        col_w = [10*mm, 35*mm, 20*mm, 18*mm, 18*mm, 22*mm, 14*mm, 20*mm, 25*mm]
    else:
        headers = [Paragraph('<b>S.No</b>', hd_s), Paragraph('<b>Item</b>', hd_s),
                   Paragraph('<b>Qty</b>', hd_s), Paragraph('<b>Rate</b>', hd_s), Paragraph('<b>Amount</b>', hd_s)]
        col_w = [15*mm, 60*mm, 30*mm, 30*mm, 35*mm]

    items_data = [headers]
    for idx, item in enumerate(v.get('items', []), 1):
        qty = item.get('quantity', 0) or 0
        rate = item.get('rate', 0) or 0
        amt = round(qty * rate, 2)
        if has_gst:
            gst_pct = item.get('gst_percent', 0)
            gst_amt = item.get('gst_amount', 0) or round(amt * gst_pct / 100, 2)
            items_data.append([Paragraph(str(idx), cell_s), Paragraph(item.get('item_name', ''), cell_s),
                Paragraph(item.get('hsn_code', ''), cell_s), Paragraph(f"{qty}", cell_r), Paragraph(f"{rate:,.2f}", cell_r),
                Paragraph(f"{amt:,.2f}", cell_r), Paragraph(f"{gst_pct}%", cell_r),
                Paragraph(f"{gst_amt:,.2f}", cell_r), Paragraph(f"{round(amt+gst_amt,2):,.2f}", cell_rb)])
        else:
            items_data.append([Paragraph(str(idx), cell_s), Paragraph(item.get('item_name', ''), cell_s),
                Paragraph(f"{qty} Q", cell_r), Paragraph(f"{rate:,.2f}", cell_r), Paragraph(f"{amt:,.2f}", cell_rb)])

    items_tbl = RLTable(items_data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor(blue)), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0,0), (-1,-1), 3), ('BOTTOMPADDING', (0,0), (-1,-1), 3), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elements.append(items_tbl)
    elements.append(Spacer(1, 3*mm))

    tot_s = ParagraphStyle('TS', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT)
    tot_b = ParagraphStyle('TB', parent=styles['Normal'], fontSize=11, alignment=TA_RIGHT, fontName='FreeSansBold', textColor=colors.HexColor(blue))

    total_data = [[Paragraph('Taxable:', tot_s), Paragraph(f"Rs. {v.get('subtotal', 0):,.2f}", tot_s)]]
    if cgst > 0: total_data.append([Paragraph('CGST:', tot_s), Paragraph(f"Rs. {cgst:,.2f}", tot_s)])
    if sgst > 0: total_data.append([Paragraph('SGST:', tot_s), Paragraph(f"Rs. {sgst:,.2f}", tot_s)])
    if igst > 0: total_data.append([Paragraph('IGST:', tot_s), Paragraph(f"Rs. {igst:,.2f}", tot_s)])
    total_data.append([Paragraph('<b>Grand Total:</b>', tot_b), Paragraph(f"<b>Rs. {v.get('total', 0):,.2f}</b>", tot_b)])
    total_data.append([Paragraph('<b>Balance:</b>', tot_b), Paragraph(f"<b>Rs. {v.get('balance', 0):,.2f}</b>", tot_b)])

    tot_tbl = RLTable(total_data, colWidths=[120*mm, 50*mm])
    tot_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('LINEABOVE', (0,-1), (-1,-1), 1, colors.HexColor(blue))]))
    elements.append(tot_tbl)

    elements.append(Spacer(1, 15*mm))
    sig_data = [[Paragraph('Received By', ParagraphStyle('Sig', alignment=TA_CENTER, fontSize=9)),
                 Paragraph(f'For {company}', ParagraphStyle('Sig', alignment=TA_CENTER, fontSize=9, fontName='FreeSansBold'))]]
    sig_tbl = RLTable(sig_data, colWidths=[85*mm, 85*mm])
    sig_tbl.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), 'FreeSans'), ('LINEABOVE', (0,0), (0,0), 0.5, colors.black), ('LINEABOVE', (1,0), (1,0), 0.5, colors.black)]))
    elements.append(sig_tbl)

    doc.build(elements)
    pdf_bytes = buf.getvalue()

    # Upload PDF to tmpfiles.org
    pdf_url = ""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            files = {"file": (f"sale_invoice_{v.get('voucher_no','')}.pdf", pdf_bytes, "application/pdf")}
            resp = await client.post("https://tmpfiles.org/api/v1/upload", files=files)
            if resp.status_code == 200:
                data = resp.json()
                url = data.get("data", {}).get("url", "")
                if url:
                    pdf_url = url.replace("tmpfiles.org/", "tmpfiles.org/dl/")
    except Exception as e:
        import logging
        logging.getLogger("salebook").error(f"tmpfiles upload error: {e}")

    # Send via WhatsApp
    from routes.whatsapp import _send_wa_message, _get_wa_settings
    wa_settings = await _get_wa_settings()
    if not wa_settings.get("api_key"):
        return {"success": False, "error": "WhatsApp API key set nahi hai. Settings mein jaake set karein."}

    default_numbers = wa_settings.get("default_numbers", [])
    if isinstance(default_numbers, str):
        default_numbers = [n.strip() for n in default_numbers.split(",") if n.strip()]

    targets = [phone] if phone else default_numbers
    results = []
    for num in targets:
        if num and num.strip():
            r = await _send_wa_message(num.strip(), text, pdf_url)
            results.append({"target": num, "success": r.get("success", False)})

    if not results:
        return {"success": False, "error": "Koi phone number nahi hai. Settings > WhatsApp mein set karein."}

    sc = sum(1 for r in results if r["success"])
    return {"success": sc > 0, "message": f"Sale Invoice {sc}/{len(results)} pe bhej diya!", "details": results, "pdf_url": pdf_url}
