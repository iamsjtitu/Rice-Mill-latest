"""
Total Sales Register — unified view across BP + Pvt Rice + Govt Rice sales.
v104.44.87 — Govt Rice included, Received from cash_transactions (party-level FIFO)
"""
from fastapi import APIRouter, Query
from typing import Optional
from datetime import datetime, timezone
import os, io
from collections import defaultdict
from motor.motor_asyncio import AsyncIOMotorClient

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME")
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

router = APIRouter()


def _safe_num(v):
    try: return float(v or 0)
    except Exception: return 0.0


async def _fetch_party_received(party: str, kms_year: str = "", season: str = "") -> float:
    """Total cash payments received for a party from cash_transactions (txn_type='jama').
    Premium adjustments are NOT counted here — they're folded into row 'total' instead
    (matching BP Sale Register logic where Balance = Amount + Premium).
    """
    if not party: return 0.0
    q = {"category": party, "txn_type": "jama"}
    if kms_year: q["kms_year"] = kms_year
    if season: q["season"] = season
    raws = await db.cash_transactions.find(q, {"_id": 0}).to_list(10000)
    skip_keywords = ('lab test premium', 'lab test bonus', 'oil premium', 'sale bhada')
    raws = [r for r in raws if not any(k in (r.get('description', '') or '').lower() for k in skip_keywords)]
    # Dedupe by (date, description) — prefer auto_ledger mirror
    grouped = {}
    for r in raws:
        key = (r.get('date', '') or '', (r.get('description', '') or '').strip().lower())
        ref = (r.get('reference', '') or '')
        existing = grouped.get(key)
        if existing is None:
            grouped[key] = r
        elif ref.startswith('auto_ledger:') and not (existing.get('reference', '') or '').startswith('auto_ledger:'):
            grouped[key] = r
    return round(sum(_safe_num(r.get('amount', 0)) for r in grouped.values()), 2)


async def _build_premium_map(bp_items: list) -> dict:
    """v104.44.90 — Fetch oil_premium for all vouchers/RSTs in given BP sales.
    Returns map: voucher_no_or_rst -> premium_amount (signed).
    Premium is applied only to Rice Bran sales' KCA portion (or non-split total).
    """
    if not bp_items:
        return {}
    voucher_set = set()
    rst_set = set()
    for s in bp_items:
        if (s.get('product') or '').strip() != 'Rice Bran':
            continue
        v = (s.get('voucher_no') or '').strip()
        r = str(s.get('rst_no') or '').strip()
        if v: voucher_set.add(v)
        if r: rst_set.add(r)
    if not voucher_set and not rst_set:
        return {}
    q = {"$or": [
        {"voucher_no": {"$in": list(voucher_set)}},
        {"rst_no": {"$in": list(rst_set)}},
    ]}
    items = await db.oil_premium.find(q, {"_id": 0}).to_list(10000)
    pmap = {}
    for op in items:
        v = (op.get('voucher_no') or '').strip()
        r = str(op.get('rst_no') or '').strip()
        amt = _safe_num(op.get('premium_amount'))
        if v: pmap[f"v:{v}"] = pmap.get(f"v:{v}", 0) + amt
        if r: pmap[f"r:{r}"] = pmap.get(f"r:{r}", 0) + amt
    return pmap


def _premium_for_sale(s: dict, pmap: dict) -> float:
    """Lookup oil premium for a BP sale, preferring voucher_no over rst_no."""
    if (s.get('product') or '').strip() != 'Rice Bran':
        return 0.0
    v = (s.get('voucher_no') or '').strip()
    r = str(s.get('rst_no') or '').strip()
    if v and f"v:{v}" in pmap:
        return pmap[f"v:{v}"]
    if r and f"r:{r}" in pmap:
        return pmap[f"r:{r}"]
    return 0.0


def _bp_to_rows(s: dict, premium: float = 0.0) -> list:
    """Normalize a bp_sale_register document into unified row(s).
    v104.44.90 — `premium` (oil/lab test) is folded into KCA total (split) or row total (non-split),
    matching BP Sale Register's Balance = Amount + Premium logic.
    If split_billing: returns 2 rows (PKA + KCA). Else 1 row."""
    common = {
        "source": "bp_sale",
        "id": s.get("id"),
        "date": s.get("date", ""),
        "voucher_no": s.get("voucher_no", ""),
        "bill_number": s.get("bill_number", ""),
        "billing_date": s.get("billing_date", ""),
        "rst_no": str(s.get("rst_no", "") or ""),
        "vehicle_no": s.get("vehicle_no", ""),
        "bill_from": s.get("bill_from", ""),
        "product": s.get("product", ""),
        "party_name": s.get("party_name", ""),
        "destination": s.get("destination", ""),
        "bags": int(s.get("bags", 0) or 0),
        "kms_year": s.get("kms_year", ""),
        "season": s.get("season", ""),
        "gst_type": s.get("gst_type", ""),
    }
    advance_total = float(s.get("advance", 0) or 0)
    if s.get("split_billing"):
        # Pakka
        pakka_kg = float(s.get("billed_weight_kg", 0) or 0)
        pakka_qtl = round(pakka_kg / 100, 2)
        pakka_rate = float(s.get("rate_per_qtl", 0) or 0)
        pakka_amt = round(pakka_qtl * pakka_rate, 2)
        pakka_tax = round(float(s.get("tax_amount", 0) or 0), 2)
        pakka_total = round(pakka_amt + pakka_tax, 2)
        # Kaccha
        kaccha_kg = float(s.get("kaccha_weight_kg", 0) or 0)
        kaccha_qtl = round(kaccha_kg / 100, 2)
        kaccha_rate = float(s.get("kaccha_rate_per_qtl", 0) or pakka_rate)
        kaccha_amt = round(kaccha_qtl * kaccha_rate, 2)
        kaccha_tax = 0.0
        # v104.44.90 — Premium folds into KCA total (matches BP Sale Register: Balance = Amount + Premium)
        kaccha_total = round(kaccha_amt + premium, 2)
        # Advance split proportionally by total
        combined = pakka_total + kaccha_total
        pakka_adv = round(advance_total * (pakka_total / combined), 2) if combined > 0 else 0
        kaccha_adv = round(advance_total - pakka_adv, 2)
        # v104.44.91 — All bags go to KCA row when split (PKA shows 0).
        # User rule: PKA + KCA both present → bags only in KCA. Solo PKA → PKA. Solo KCA → KCA.
        pakka_bags = 0
        kaccha_bags = common["bags"]
        return [
            {**common, "split_type": "PKA", "bags": pakka_bags,
             "net_weight_qtl": pakka_qtl, "rate_per_qtl": pakka_rate,
             "amount": pakka_amt, "tax": pakka_tax, "total": pakka_total,
             "balance": round(pakka_total - pakka_adv, 2), "advance": pakka_adv,
             "split_billing": True},
            {**common, "split_type": "KCA", "bags": kaccha_bags,
             "net_weight_qtl": kaccha_qtl, "rate_per_qtl": kaccha_rate,
             "amount": kaccha_amt, "tax": kaccha_tax, "total": kaccha_total,
             "balance": round(kaccha_total - kaccha_adv, 2), "advance": kaccha_adv,
             "split_billing": True},
        ]
    # Non-split
    amt = float(s.get("amount", 0) or 0)
    tax = float(s.get("tax_amount", 0) or 0)
    base_total = float(s.get("total", 0) or 0) or (amt + tax)
    # v104.44.90 — Premium folds into total for non-split BP rows too
    total = round(base_total + premium, 2)
    balance = round(total - advance_total, 2)
    return [{**common, "split_type": "", "bags": common["bags"],
             "net_weight_qtl": round(float(s.get("net_weight_kg", 0) or 0) / 100, 2),
             "rate_per_qtl": float(s.get("rate_per_qtl", 0) or 0),
             "amount": round(amt, 2), "tax": round(tax, 2), "total": round(total, 2),
             "balance": round(balance, 2), "advance": round(advance_total, 2),
             "split_billing": False}]


def _rs_to_row(s: dict) -> dict:
    """Normalize rice_sales document (Pvt Rice) into unified row format."""
    total = float(s.get("total_amount", 0) or 0)
    paid = float(s.get("paid_amount", 0) or 0)
    return {
        "source": "rice_sale",
        "id": s.get("id"),
        "date": s.get("date", ""),
        "voucher_no": "",
        "bill_number": "",
        "billing_date": s.get("date", ""),
        "rst_no": str(s.get("rst_no", "") or ""),
        "vehicle_no": s.get("truck_no", ""),
        "bill_from": "",
        "product": s.get("rice_type", "Pvt Rice") or "Pvt Rice",
        "party_name": s.get("party_name", ""),
        "destination": "",
        "net_weight_qtl": float(s.get("quantity_qntl", 0) or 0),
        "bags": int(s.get("bags", 0) or 0),
        "rate_per_qtl": float(s.get("rate_per_qntl", 0) or 0),
        "amount": round(total, 2),
        "tax": 0,
        "total": round(total, 2),
        "balance": round(total - paid, 2),
        "advance": round(paid, 2),
        "kms_year": s.get("kms_year", ""),
        "season": s.get("season", ""),
        "split_billing": False,
        "split_type": "",
        "gst_type": "",
    }


def _sv_to_row(s: dict) -> dict:
    """Normalize sale_vouchers document (Govt Rice / SaleBook) into unified row."""
    items = s.get("items", []) or []
    # Aggregate from items (quantity treated as weight in Qtl already per SaleBook convention,
    # but some older rows may have it in KG — normalize by checking magnitude)
    total_qtl = 0.0
    total_bags = 0
    rates = []
    product_names = []
    for it in items:
        q = float(it.get("quantity", 0) or it.get("weight_qntl", 0) or 0)
        # If unit is KG, convert to Qtl
        unit = (it.get("unit", "") or "").upper()
        if unit == "KG":
            q = q / 100
        total_qtl += q
        total_bags += int(it.get("bags", 0) or 0)
        r = float(it.get("rate", 0) or 0)
        if r > 0:
            rates.append(r)
        nm = it.get("item_name", "") or ""
        if nm and nm not in product_names:
            product_names.append(nm)
    avg_rate = round(sum(rates) / len(rates), 0) if rates else 0
    subtotal = float(s.get("subtotal", 0) or 0)
    # Tax: CGST + SGST + IGST
    tax = float(s.get("cgst_amount", 0) or 0) + float(s.get("sgst_amount", 0) or 0) + float(s.get("igst_amount", 0) or 0)
    total = float(s.get("total", 0) or 0) or round(subtotal + tax, 2)
    paid = float(s.get("paid_amount", 0) or 0)
    balance = float(s.get("balance", 0) or 0) or round(total - paid, 2)
    gst_type = s.get("gst_type", "none")
    is_kca = gst_type == "none" or gst_type == ""
    product_label = "Govt Rice" + (f" · {' / '.join(product_names[:2])}" if product_names else "")
    voucher = s.get("voucher_no_label") or (f"S-{s.get('voucher_no', 0):03d}" if s.get("voucher_no") else "")
    return {
        "source": "sale_voucher",
        "id": s.get("id"),
        "date": s.get("date", ""),
        "voucher_no": voucher,
        "bill_number": s.get("invoice_no", ""),
        "billing_date": s.get("date", ""),
        "rst_no": str(s.get("rst_no", "") or ""),
        "vehicle_no": s.get("truck_no", ""),
        "bill_from": s.get("bill_book", ""),
        "product": product_label,
        "party_name": s.get("party_name", ""),
        "destination": s.get("destination", ""),
        "net_weight_qtl": round(total_qtl, 2),
        "bags": total_bags,
        "rate_per_qtl": avg_rate,
        "amount": round(subtotal, 2),
        "tax": round(tax, 2),
        "total": round(total, 2),
        "balance": round(balance, 2),
        "advance": round(paid, 2),
        "kms_year": s.get("kms_year", ""),
        "season": s.get("season", ""),
        "split_billing": False,
        "split_type": "KCA" if is_kca else "PKA",
        "gst_type": gst_type,
    }


@router.get("/total-sales-register")
async def get_total_sales(
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    party_name: Optional[str] = None,
    product: Optional[str] = None,
    source: Optional[str] = None,  # "bp_sale" | "rice_sale" | None (both)
    search: Optional[str] = None,
):
    """Combined Total Sales Register across BP Sale Register + Pvt Rice Sales."""
    q_common = {}
    if kms_year:
        q_common["kms_year"] = kms_year
    if season:
        q_common["season"] = season
    if date_from or date_to:
        q_common["date"] = {}
        if date_from:
            q_common["date"]["$gte"] = date_from
        if date_to:
            q_common["date"]["$lte"] = date_to

    rows = []
    # BP Sale Register
    if source in (None, "bp_sale"):
        q_bp = dict(q_common)
        if product:
            q_bp["product"] = {"$regex": product, "$options": "i"}
        if party_name:
            q_bp["party_name"] = {"$regex": party_name, "$options": "i"}
        bp_items = await db.bp_sale_register.find(q_bp, {"_id": 0}).to_list(20000)
        # v104.44.90 — Build oil_premium map keyed by voucher_no/rst_no, fold into row total
        premium_map = await _build_premium_map(bp_items)
        for s in bp_items:
            prem = _premium_for_sale(s, premium_map)
            rows.extend(_bp_to_rows(s, prem))

    # Pvt Rice
    if source in (None, "rice_sale") and (not product or "rice" in product.lower()):
        q_rs = dict(q_common)
        if party_name:
            q_rs["party_name"] = {"$regex": party_name, "$options": "i"}
        rs_items = await db.rice_sales.find(q_rs, {"_id": 0}).to_list(20000)
        for s in rs_items:
            rows.append(_rs_to_row(s))

    # Govt Rice / SaleBook
    if source in (None, "sale_voucher"):
        q_sv = dict(q_common)
        if party_name:
            q_sv["party_name"] = {"$regex": party_name, "$options": "i"}
        sv_items = await db.sale_vouchers.find(q_sv, {"_id": 0}).to_list(20000)
        for s in sv_items:
            row = _sv_to_row(s)
            # product filter: if explicit product, include only if matches any item name
            if product:
                if product.lower() not in (row.get("product") or "").lower():
                    continue
            rows.append(row)

    # v104.44.87 — Allocate payments from cash_transactions per party (FIFO by date).
    # BP split rows use suffixed party_name: "PartyName (PKA)" / "PartyName (KCA)".
    # Non-split BP + Pvt Rice + Govt use plain party_name.
    def _party_key(r):
        if r.get("split_type") in ("PKA", "KCA") and r.get("source") == "bp_sale":
            return f"{(r.get('party_name') or '').strip()} ({r['split_type']})"
        return (r.get("party_name") or "").strip()

    party_rows = defaultdict(list)
    for r in rows:
        pk = _party_key(r)
        if pk:
            party_rows[(pk, r.get("kms_year", ""), r.get("season", ""))].append(r)
    for (pk, kms, ssn), group in party_rows.items():
        received_total = await _fetch_party_received(pk, kms, ssn)
        group.sort(key=lambda r: (r.get("date") or "", r.get("id") or ""))
        remaining = received_total
        for r in group:
            if remaining <= 0:
                r["advance"] = 0.0
                r["balance"] = round(float(r.get("total", 0) or 0), 2)
                continue
            alloc = min(remaining, float(r.get("total", 0) or 0))
            r["advance"] = round(alloc, 2)
            r["balance"] = round(float(r.get("total", 0) or 0) - alloc, 2)
            remaining = round(remaining - alloc, 2)

    # Free-text search across party, vehicle, rst, voucher, product
    if search:
        q = search.lower()
        rows = [
            r for r in rows
            if q in (r.get("party_name", "") or "").lower()
            or q in (r.get("vehicle_no", "") or "").lower()
            or q in (r.get("rst_no", "") or "").lower()
            or q in (r.get("voucher_no", "") or "").lower()
            or q in (r.get("product", "") or "").lower()
            or q in (r.get("bill_number", "") or "").lower()
        ]

    # Sort: date desc, then created
    rows.sort(key=lambda r: (r.get("date") or "", r.get("id") or ""), reverse=True)

    # Totals summary (sum over filtered rows)
    totals = {
        "rows_count": len(rows),
        "net_weight_qtl": round(sum(r["net_weight_qtl"] for r in rows), 2),
        "bags": sum(r["bags"] for r in rows),
        "amount": round(sum(r["amount"] for r in rows), 2),
        "tax": round(sum(r["tax"] for r in rows), 2),
        "total": round(sum(r["total"] for r in rows), 2),
        "balance": round(sum(r["balance"] for r in rows), 2),
        "received": round(sum(r["advance"] for r in rows), 2),
    }

    # Party grouping
    parties: dict = {}
    for r in rows:
        key = (r.get("party_name") or "").strip() or "(Unknown)"
        p = parties.setdefault(key, {"party_name": key, "rows": 0, "net_weight_qtl": 0.0, "bags": 0, "total": 0.0, "balance": 0.0, "received": 0.0, "products": set()})
        p["rows"] += 1
        p["net_weight_qtl"] = round(p["net_weight_qtl"] + r["net_weight_qtl"], 2)
        p["bags"] += r["bags"]
        p["total"] = round(p["total"] + r["total"], 2)
        p["balance"] = round(p["balance"] + r["balance"], 2)
        p["received"] = round(p["received"] + r["advance"], 2)
        p["products"].add(r.get("product", ""))
    parties_list = [{**v, "products": sorted([x for x in v["products"] if x])} for v in parties.values()]
    parties_list.sort(key=lambda p: p["total"], reverse=True)

    return {"rows": rows, "totals": totals, "parties": parties_list}


@router.get("/total-sales-register/export/excel")
async def export_total_sales_excel(
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    party_name: Optional[str] = None,
    product: Optional[str] = None,
    source: Optional[str] = None,
    search: Optional[str] = None,
):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = await get_total_sales(kms_year, season, date_from, date_to, party_name, product, source, search)
    rows = data["rows"]
    totals = data["totals"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Total Sales"

    # Styling palette (professional navy/gold/white scheme)
    navy = "1E3A8A"
    gold = "D97706"
    soft_bg = "F8FAFC"
    header_bg = "1E293B"
    kca_bg = "FEF3C7"
    pka_bg = "D1FAE5"
    total_bg = "FFFBEB"
    thin = Side(border_style="thin", color="CBD5E1")
    thick = Side(border_style="medium", color=navy)
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Date", "Voucher", "Bill No", "RST", "Vehicle", "Bill From", "Party", "Destination",
               "N/W (Qtl)", "Bags", "Rate/Q", "Amount", "Tax", "Total",
               "Received(T)", "Balance(T)"]

    # Title area (rows 1-3)
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "TOTAL SALES REGISTER"
    ws["A1"].font = Font(bold=True, size=18, color=navy, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    meta = f"KMS {kms_year or 'ALL'}  •  {season or 'All Seasons'}"
    if date_from or date_to:
        meta += f"  •  {date_from or 'start'} → {date_to or 'today'}"
    if party_name:
        meta += f"  •  Party: {party_name}"
    if product:
        meta += f"  •  Product: {product}"
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = meta
    ws["A2"].font = Font(size=10, italic=True, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Header row (row 4)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = border_thin
    ws.row_dimensions[4].height = 22

    # Data rows
    for i, r in enumerate(rows, start=5):
        voucher_display = r.get("voucher_no", "") or "-"
        if r.get("split_type"):
            voucher_display = f"{voucher_display} · {r['split_type']}"
        vals = [
            r.get("date", ""), voucher_display, r.get("bill_number", ""),
            r.get("rst_no", ""), r.get("vehicle_no", ""), r.get("bill_from", ""),
            r.get("party_name", ""), r.get("destination", ""),
            r.get("net_weight_qtl", 0), r.get("bags", 0), r.get("rate_per_qtl", 0),
            r.get("amount", 0), r.get("tax", 0), r.get("total", 0),
            r.get("advance", 0), r.get("balance", 0),
        ]
        # Row color based on split_type
        row_fill = None
        if r.get("split_type") == "PKA":
            row_fill = PatternFill(start_color=pka_bg, end_color=pka_bg, fill_type="solid")
        elif r.get("split_type") == "KCA":
            row_fill = PatternFill(start_color=kca_bg, end_color=kca_bg, fill_type="solid")
        elif i % 2 == 1:
            row_fill = PatternFill(start_color=soft_bg, end_color=soft_bg, fill_type="solid")

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = border_thin
            cell.font = Font(size=9)
            if row_fill:
                cell.fill = row_fill
            # right-align numeric columns (9..17)
            if c >= 9:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if c == 10:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[i].height = 18

    # Grand totals row
    # Grand totals row — Date, Voucher, BillNo, RST, Vehicle, BillFrom, Party, Destination, NW, Bags, Rate, Amount, Tax, Total, Received(T), Balance(T)
    tr = len(rows) + 5
    total_vals = ["TOTALS", "", "", "", "", "", "", "",
                  totals["net_weight_qtl"], totals["bags"], "",
                  totals["amount"], totals["tax"], totals["total"],
                  totals["received"], totals["balance"]]
    for c, v in enumerate(total_vals, 1):
        cell = ws.cell(row=tr, column=c, value=v)
        cell.font = Font(bold=True, size=11, color=navy)
        cell.fill = PatternFill(start_color=total_bg, end_color=total_bg, fill_type="solid")
        cell.border = Border(left=thin, right=thin, top=thick, bottom=thick)
        if c >= 9:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if c == 10:
                cell.number_format = '#,##0'
            else:
                cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[tr].height = 22

    # Footer
    ws.cell(row=tr + 2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M')} UTC  •  Rows: {len(rows)}").font = Font(size=8, italic=True, color="64748B")
    ws.merge_cells(f"A{tr + 2}:{last_col}{tr + 2}")

    # Column widths — Date, Voucher, BillNo, RST, Vehicle, BillFrom, Party, Destination, NW, Bags, Rate, Amount, Tax, Total, Received(T), Balance(T)
    widths = [11, 14, 11, 7, 13, 13, 22, 14, 11, 7, 10, 13, 10, 13, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes below header
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="total_sales_{kms_year or "all"}.xlsx"'})


@router.get("/total-sales-register/export/pdf")
async def export_total_sales_pdf(
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    party_name: Optional[str] = None,
    product: Optional[str] = None,
    source: Optional[str] = None,
    search: Optional[str] = None,
):
    from fastapi.responses import StreamingResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    data = await get_total_sales(kms_year, season, date_from, date_to, party_name, product, source, search)
    rows = data["rows"]
    totals = data["totals"]

    # Palette — professional navy + gold
    navy = colors.HexColor("#1E3A8A")
    navy_dark = colors.HexColor("#0F172A")
    gold = colors.HexColor("#D97706")
    soft = colors.HexColor("#F8FAFC")
    pka_green = colors.HexColor("#D1FAE5")
    kca_amber = colors.HexColor("#FEF3C7")
    total_gold = colors.HexColor("#FEF3C7")
    border_gray = colors.HexColor("#CBD5E1")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3),
        leftMargin=1.0 * cm, rightMargin=1.0 * cm,
        topMargin=1.0 * cm, bottomMargin=1.0 * cm,
        title="Total Sales Register",
    )
    elems = []

    # Header section: title + meta banner
    title_style = ParagraphStyle("title", fontName="Helvetica-Bold", alignment=TA_CENTER, fontSize=20, textColor=navy, spaceAfter=4, leading=24)
    subtitle_style = ParagraphStyle("sub", fontName="Helvetica-Oblique", alignment=TA_CENTER, fontSize=10, textColor=colors.HexColor("#475569"), spaceAfter=8)
    meta_style = ParagraphStyle("meta", fontName="Helvetica", alignment=TA_CENTER, fontSize=9, textColor=colors.HexColor("#64748B"))

    elems.append(Paragraph("TOTAL SALES REGISTER", title_style))
    meta = f"KMS {kms_year or 'ALL'}  •  {season or 'All Seasons'}"
    if date_from or date_to:
        meta += f"  •  {date_from or 'start'} → {date_to or 'today'}"
    if party_name:
        meta += f"  •  Party: {party_name}"
    if product:
        meta += f"  •  Product: {product}"
    elems.append(Paragraph(meta, subtitle_style))

    # Quick stats strip — Pending == Balance, use Balance only
    stats_data = [[
        Paragraph(f"<b>Entries</b><br/><font size=13 color='#1E3A8A'>{totals['rows_count']}</font>", meta_style),
        Paragraph(f"<b>N/W (Qtl)</b><br/><font size=13 color='#1E3A8A'>{totals['net_weight_qtl']:,.2f}</font>", meta_style),
        Paragraph(f"<b>Bags</b><br/><font size=13 color='#1E3A8A'>{totals['bags']:,}</font>", meta_style),
        Paragraph(f"<b>Total Bill</b><br/><font size=13 color='#059669'>₹ {totals['total']:,.2f}</font>", meta_style),
        Paragraph(f"<b>Received(T)</b><br/><font size=13 color='#0891B2'>₹ {totals['received']:,.2f}</font>", meta_style),
        Paragraph(f"<b>Balance(T)</b><br/><font size=13 color='{'#D97706' if (totals['balance'] or 0) > 0 else '#059669'}'>₹ {totals['balance']:,.2f}</font>", meta_style),
    ]]
    stats_table = Table(stats_data, colWidths=[6.3 * cm] * 6)
    stats_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), soft),
        ("BOX", (0, 0), (-1, -1), 0.8, navy),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, border_gray),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elems.append(stats_table)
    elems.append(Spacer(1, 0.35 * cm))

    # Data table
    headers = ["Date", "Voucher", "Bill No", "RST", "Vehicle", "Bill From", "Party", "Destination",
               "N/W (Qtl)", "Bags", "Rate/Q", "Amount", "Tax", "Total",
               "Received(T)", "Balance(T)"]
    table_data = [headers]
    split_bg_rows = []  # (row_index, is_pka)
    for r in rows:
        voucher_disp = r.get("voucher_no", "") or "-"
        if r.get("split_type"):
            voucher_disp = f"{voucher_disp} · {r['split_type']}"
        table_data.append([
            (r.get("date", "") or "")[-5:].replace("-", "/") if r.get("date") else "",
            voucher_disp,
            (r.get("bill_number", "") or "")[:12],
            r.get("rst_no", "") or "",
            r.get("vehicle_no", "") or "",
            (r.get("bill_from", "") or "")[:12],
            (r.get("party_name", "") or "")[:22],
            (r.get("destination", "") or "")[:13],
            f"{r.get('net_weight_qtl', 0):,.2f}",
            f"{r.get('bags', 0):,}",
            f"{r.get('rate_per_qtl', 0):,.0f}",
            f"{r.get('amount', 0):,.2f}",
            f"{r.get('tax', 0):,.2f}",
            f"{r.get('total', 0):,.2f}",
            f"{r.get('advance', 0):,.2f}",
            f"{r.get('balance', 0):,.2f}",
        ])
        if r.get("split_type") == "PKA":
            split_bg_rows.append((len(table_data) - 1, True))
        elif r.get("split_type") == "KCA":
            split_bg_rows.append((len(table_data) - 1, False))

    # Totals row
    table_data.append([
        "TOTALS", "", "", "", "", "", "", "",
        f"{totals['net_weight_qtl']:,.2f}", f"{totals['bags']:,}", "",
        f"{totals['amount']:,.2f}", f"{totals['tax']:,.2f}", f"{totals['total']:,.2f}",
        f"{totals['received']:,.2f}", f"{totals['balance']:,.2f}",
    ])

    # Column widths (cm) — Date,Voucher,BillNo,RST,Vehicle,BillFrom,Party,Destination,NW,Bags,Rate,Amount,Tax,Total,Recv(T),Bal(T)
    col_widths_cm = [1.5, 1.9, 1.7, 1.0, 1.8, 1.8, 4.0, 2.0, 1.7, 1.0, 1.4, 2.1, 1.4, 2.1, 2.2, 2.2]
    col_widths = [w * cm for w in col_widths_cm]

    style_cmds = [
        # Header row
        ("BACKGROUND", (0, 0), (-1, 0), navy_dark),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        # Body
        ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -2), 8),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ("ALIGN", (8, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 1), (7, -1), "LEFT"),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        # Alternating rows
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, soft]),
        # Grid
        ("INNERGRID", (0, 0), (-1, -1), 0.3, border_gray),
        ("BOX", (0, 0), (-1, -1), 0.8, navy),
        # Totals row
        ("BACKGROUND", (0, -1), (-1, -1), total_gold),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 10),
        ("TEXTCOLOR", (0, -1), (-1, -1), navy),
        ("LINEABOVE", (0, -1), (-1, -1), 1.5, navy),
        ("TOPPADDING", (0, -1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 8),
    ]
    # Split row backgrounds
    for idx, is_pka in split_bg_rows:
        style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), pka_green if is_pka else kca_amber))

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(style_cmds))
    elems.append(t)
    elems.append(Spacer(1, 0.4 * cm))

    # Legend
    legend = Paragraph(
        f"<font color='#059669'>■</font> PKA (GST Bill Portion)   "
        f"<font color='#D97706'>■</font> KCA (Kaccha Slip — No GST)   "
        f"<font color='#64748B'>Generated: {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M')} UTC  •  Rows: {len(rows)}</font>",
        meta_style,
    )
    elems.append(legend)

    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="total_sales_{kms_year or "all"}.pdf"'})
