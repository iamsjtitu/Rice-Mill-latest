from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import Response
from typing import Optional
from datetime import datetime, timezone
from database import db
from models import *
import uuid

router = APIRouter()

# ============ DIESEL PUMPS MANAGEMENT ============

@router.get("/diesel-pumps")
async def get_diesel_pumps():
    pumps = await db.diesel_pumps.find({}, {"_id": 0}).to_list(100)
    return pumps

@router.post("/diesel-pumps")
async def add_diesel_pump(request: Request):
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Pump name required")
    existing = await db.diesel_pumps.find_one({"name": name})
    if existing:
        raise HTTPException(status_code=400, detail="Pump already exists")
    pump = {
        "id": str(uuid.uuid4()),
        "name": name,
        "is_default": data.get("is_default", False),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    # If setting as default, unset others
    if pump["is_default"]:
        await db.diesel_pumps.update_many({}, {"$set": {"is_default": False}})
    await db.diesel_pumps.insert_one(pump)
    pump.pop("_id", None)
    return pump

@router.put("/diesel-pumps/{pump_id}/set-default")
async def set_default_pump(pump_id: str):
    pump = await db.diesel_pumps.find_one({"id": pump_id})
    if not pump:
        raise HTTPException(status_code=404, detail="Pump not found")
    await db.diesel_pumps.update_many({}, {"$set": {"is_default": False}})
    await db.diesel_pumps.update_one({"id": pump_id}, {"$set": {"is_default": True}})
    return {"message": "Default pump set", "pump_id": pump_id}

@router.delete("/diesel-pumps/{pump_id}")
async def delete_diesel_pump(pump_id: str):
    result = await db.diesel_pumps.delete_one({"id": pump_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pump not found")
    return {"message": "Pump deleted"}

# ============ DIESEL ACCOUNT TRANSACTIONS ============

@router.get("/diesel-accounts")
async def get_diesel_accounts(pump_id: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None,
                               date_from: Optional[str] = None, date_to: Optional[str] = None,
                               txn_type: Optional[str] = None, truck_no: Optional[str] = None):
    query = {}
    if pump_id: query["pump_id"] = pump_id
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if txn_type: query["txn_type"] = txn_type
    if truck_no: query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        if date_q: query["date"] = date_q
    txns = await db.diesel_accounts.find(query, {"_id": 0}).sort([("date", -1), ("created_at", -1)]).to_list(5000)
    return txns

@router.get("/diesel-accounts/summary")
async def get_diesel_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.diesel_accounts.find(query, {"_id": 0}).to_list(5000)
    pumps = await db.diesel_pumps.find({}, {"_id": 0}).to_list(100)
    
    # Compute opening balance from previous FY per pump
    opening_balances = {}
    if kms_year:
        fy_parts = kms_year.split('-')
        if len(fy_parts) == 2:
            try:
                prev_fy = f"{int(fy_parts[0])-1}-{int(fy_parts[1])-1}"
                prev_query = {"kms_year": prev_fy}
                if season: prev_query["season"] = season
                prev_txns = await db.diesel_accounts.find(prev_query, {"_id": 0}).to_list(5000)
                for t in prev_txns:
                    pid = t.get("pump_id", "")
                    if pid not in opening_balances:
                        opening_balances[pid] = 0
                    if t.get("txn_type") == "debit":
                        opening_balances[pid] += t.get("amount", 0)
                    elif t.get("txn_type") == "payment":
                        opening_balances[pid] -= t.get("amount", 0)
            except (ValueError, IndexError):
                pass
    
    # Get all ledger nikasi entries for diesel pumps (source of truth for payments)
    pump_names = [p["name"] for p in pumps]
    ledger_query = {"account": "ledger", "txn_type": "nikasi", "category": {"$in": pump_names}}
    if kms_year: ledger_query["kms_year"] = kms_year
    if season: ledger_query["season"] = season
    ledger_payments = await db.cash_transactions.find(ledger_query, {"_id": 0}).to_list(50000)
    
    # Group ledger payments by pump name
    ledger_paid_by_pump = {}
    for lp in ledger_payments:
        pn = lp.get("category", "")
        ledger_paid_by_pump[pn] = ledger_paid_by_pump.get(pn, 0) + lp.get("amount", 0)
    
    pump_summaries = []
    for pump in pumps:
        pid = pump["id"]
        pump_txns = [t for t in txns if t.get("pump_id") == pid]
        total_diesel = sum(t["amount"] for t in pump_txns if t.get("txn_type") == "debit")
        # Use ledger as source of truth for total_paid (includes manual Cash Book payments)
        total_paid = round(ledger_paid_by_pump.get(pump["name"], 0), 2)
        ob = round(opening_balances.get(pid, 0), 2)
        balance = round(ob + total_diesel - total_paid, 2)
        pump_summaries.append({
            "pump_id": pid, "pump_name": pump["name"], "is_default": pump.get("is_default", False),
            "opening_balance": ob,
            "total_diesel": round(total_diesel, 2), "total_paid": total_paid, "balance": balance,
            "txn_count": len([t for t in pump_txns if t.get("txn_type") == "debit"])
        })
    
    grand_ob = sum(p["opening_balance"] for p in pump_summaries)
    grand_diesel = sum(p["total_diesel"] for p in pump_summaries)
    grand_paid = sum(p["total_paid"] for p in pump_summaries)
    return {
        "pumps": pump_summaries,
        "grand_opening_balance": round(grand_ob, 2),
        "grand_total_diesel": round(grand_diesel, 2),
        "grand_total_paid": round(grand_paid, 2),
        "grand_balance": round(grand_ob + grand_diesel - grand_paid, 2)
    }

# ============ DIESEL PAYMENT / SETTLEMENT ============

@router.post("/diesel-accounts/pay")
async def make_diesel_payment(request: Request, username: str = "", role: str = ""):
    data = await request.json()
    pump_id = data.get("pump_id")
    amount = float(data.get("amount", 0))
    if not pump_id or amount <= 0:
        raise HTTPException(status_code=400, detail="pump_id and amount > 0 required")
    
    pump = await db.diesel_pumps.find_one({"id": pump_id}, {"_id": 0})
    if not pump:
        raise HTTPException(status_code=404, detail="Pump not found")
    
    kms_year = data.get("kms_year", "")
    season = data.get("season", "")
    date = data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    notes = data.get("notes", "")
    round_off = float(data.get("round_off", 0))
    total_settled = round(amount + round_off, 2)
    
    # Create payment transaction in diesel account (total including round off)
    pay_txn = {
        "id": str(uuid.uuid4()), "date": date,
        "pump_id": pump_id, "pump_name": pump["name"],
        "truck_no": "", "agent_name": "",
        "amount": total_settled, "txn_type": "payment",
        "description": f"Payment to {pump['name']} - Rs.{amount}" + (f" (Round Off: {'+' if round_off > 0 else ''}{round_off})" if round_off else "") + (f" - {notes}" if notes else ""),
        "kms_year": kms_year, "season": season,
        "created_by": username or "system",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.diesel_accounts.insert_one(pay_txn)
    
    # Auto create Cash Book entry (Nikasi) - only actual cash
    cb = {
        "id": str(uuid.uuid4()), "date": date,
        "account": "cash", "txn_type": "nikasi", "category": pump['name'],
        "party_type": "Diesel",
        "description": f"Diesel Payment: {pump['name']} - Rs.{amount}" + (f" ({notes})" if notes else ""),
        "amount": round(amount, 2), "reference": f"diesel_pay:{pay_txn['id'][:8]}",
        "kms_year": kms_year, "season": season,
        "created_by": username or "system",
        "linked_diesel_payment_id": pay_txn["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(cb)

    # Ledger Nikasi - reduce diesel outstanding (includes round off)
    ledger_cb = {
        "id": str(uuid.uuid4()), "date": date,
        "account": "ledger", "txn_type": "nikasi", "category": pump['name'],
        "party_type": "Diesel",
        "description": f"Diesel Payment: {pump['name']} - Rs.{total_settled}" + (f" (Cash: {amount}, Round Off: {round_off})" if round_off else "") + (f" ({notes})" if notes else ""),
        "amount": total_settled, "reference": f"diesel_pay_ledger:{pay_txn['id'][:8]}",
        "kms_year": kms_year, "season": season,
        "created_by": username or "system",
        "linked_diesel_payment_id": pay_txn["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(ledger_cb)
    
    return {"success": True, "message": f"Rs.{amount} payment to {pump['name']} recorded", "txn_id": pay_txn["id"]}

@router.delete("/diesel-accounts/{txn_id}")
async def delete_diesel_transaction(txn_id: str):
    txn = await db.diesel_accounts.find_one({"id": txn_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # If it was a payment, also delete linked cash book entry
    if txn.get("txn_type") == "payment":
        await db.cash_transactions.delete_many({"linked_diesel_payment_id": txn_id})
    
    await db.diesel_accounts.delete_one({"id": txn_id})
    return {"message": "Deleted", "id": txn_id}


@router.post("/diesel-accounts/delete-bulk")
async def delete_diesel_transactions_bulk(request: Request):
    body = await request.json()
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="No ids provided")
    # Also delete linked cash book entries for payment transactions
    payment_txns = await db.diesel_accounts.find({"id": {"$in": ids}, "txn_type": "payment"}, {"_id": 0}).to_list(1000)
    for pt in payment_txns:
        await db.cash_transactions.delete_many({"linked_diesel_payment_id": pt["id"]})
    result = await db.diesel_accounts.delete_many({"id": {"$in": ids}})
    return {"message": f"{result.deleted_count} transactions deleted", "deleted": result.deleted_count}


@router.get("/diesel-accounts/excel")
async def export_diesel_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from utils.export_helpers import (style_excel_title, style_excel_header_row,
        style_excel_data_rows, style_excel_total_row, COLORS, BORDER_THIN)
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.diesel_accounts.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    summary = await get_diesel_summary(kms_year=kms_year, season=season)

    wb = Workbook(); ws = wb.active; ws.title = "Diesel Account"
    ncols = 7
    
    # Title
    title = "Diesel Account / डीजल खाता"
    if kms_year: title += f" | KMS {kms_year}"
    style_excel_title(ws, title, ncols)

    # Summary section
    ws.cell(row=4, column=1, value="Pump Summary").font = Font(bold=True, size=11, color=COLORS['title_text'])
    sum_headers = ['Pump Name', 'Total Diesel (Rs.)', 'Total Paid (Rs.)', 'Balance (Rs.)', 'Entries']
    for col, h in enumerate(sum_headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_excel_header_row(ws, 5, 5)
    
    row = 6
    sum_start = row
    for p in summary.get("pumps", []):
        for col, v in enumerate([p["pump_name"] + (" (Default)" if p.get("is_default") else ""), p["total_diesel"], p["total_paid"], p["balance"], p["txn_count"]], 1):
            c = ws.cell(row=row, column=col, value=v); c.border = BORDER_THIN
        row += 1
    style_excel_data_rows(ws, sum_start, row - 1, 5, sum_headers)
    
    ws.cell(row=row, column=1, value="GRAND TOTAL")
    ws.cell(row=row, column=2, value=summary.get("grand_total_diesel", 0))
    ws.cell(row=row, column=3, value=summary.get("grand_total_paid", 0))
    ws.cell(row=row, column=4, value=summary.get("grand_balance", 0))
    style_excel_total_row(ws, row, 5)
    row += 2

    # Transactions section
    ws.cell(row=row, column=1, value="Transactions").font = Font(bold=True, size=11, color=COLORS['title_text'])
    row += 1
    txn_headers = ['Date', 'Pump', 'Type', 'Truck No', 'Agent', 'Amount (Rs.)', 'Description']
    for col, h in enumerate(txn_headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_excel_header_row(ws, row, ncols)
    row += 1
    data_start = row
    for t in txns:
        vals = [t.get("date",""), t.get("pump_name",""), "Payment" if t.get("txn_type")=="payment" else "Diesel",
                t.get("truck_no",""), t.get("agent_name",""), t.get("amount",0), t.get("description","")]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
        row += 1
    if txns:
        style_excel_data_rows(ws, data_start, row - 1, ncols, txn_headers)

    for i in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=diesel_account_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@router.get("/diesel-accounts/pdf")
async def export_diesel_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from utils.export_helpers import get_pdf_styles
    from reportlab.lib import colors
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.diesel_accounts.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    summary = await get_diesel_summary(kms_year=kms_year, season=season)

    from utils.export_helpers import get_pdf_table_style, get_pdf_company_header
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = get_pdf_styles()
    elements = []

    elements.extend(get_pdf_company_header())
    elements.append(Paragraph("Diesel Account / डीजल खाता", styles['Title']))
    elements.append(Spacer(1, 12))

    # Summary table
    sum_data = [['Pump Name', 'Total Diesel', 'Total Paid', 'Balance', 'Entries']]
    for p in summary.get("pumps", []):
        sum_data.append([p["pump_name"] + (" *" if p.get("is_default") else ""),
                         f"Rs.{p['total_diesel']}", f"Rs.{p['total_paid']}", f"Rs.{p['balance']}", str(p["txn_count"])])
    sum_data.append(['GRAND TOTAL', f"Rs.{summary.get('grand_total_diesel',0)}", f"Rs.{summary.get('grand_total_paid',0)}", f"Rs.{summary.get('grand_balance',0)}", ''])
    st = Table(sum_data, colWidths=[180, 100, 100, 100, 60])
    st.setStyle(TableStyle(get_pdf_table_style(len(sum_data))))
    elements.append(st)
    elements.append(Spacer(1, 20))

    # Transaction table
    elements.append(Paragraph("Transactions", styles['Heading2']))
    t_data = [['Date', 'Pump', 'Type', 'Truck', 'Agent', 'Amount', 'Description']]
    for t in txns:
        t_data.append([t.get("date",""), t.get("pump_name","")[:15],
                       "Payment" if t.get("txn_type")=="payment" else "Diesel",
                       t.get("truck_no",""), t.get("agent_name","")[:12],
                       f"Rs.{t.get('amount',0)}", t.get("description","")[:30]])
    tt = Table(t_data, colWidths=[70, 100, 55, 80, 80, 70, 180])
    tt.setStyle(TableStyle(get_pdf_table_style(len(t_data))))
    elements.append(tt)

    doc.build(elements)
    buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=diesel_account_{datetime.now().strftime('%Y%m%d')}.pdf"})
