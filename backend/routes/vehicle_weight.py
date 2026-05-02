"""Vehicle Weight Entry - Weighbridge management for Rice Mill"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from database import db
from datetime import datetime, timezone
import uuid
import logging
import io
import os
import base64 as b64mod
import httpx
from utils.date_format import fmt_date

router = APIRouter()
logger = logging.getLogger(__name__)

IMG_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "vw_images")
os.makedirs(IMG_DIR, exist_ok=True)


def _is_sale(trans_type: str) -> bool:
    """VW Sale dispatch detection — handles current and legacy labels."""
    t = (trans_type or "").lower()
    return "sale" in t or "dispatch" in t or "sell" in t


async def _sync_sale_bag_out(vw_entry_id: str, vw_entry: dict, username: str = "system"):
    """Sync linked gunny_bags 'out' entry for a VW sale dispatch.

    Creates / updates / deletes the bag-stock-out entry based on:
        trans_type is sale + tot_pkts > 0 + bag_type set
    Reference key: `vw_sale_bag:{rst_no}`
    """
    rst_no = vw_entry.get("rst_no")
    if rst_no is None:
        return
    ref = f"vw_sale_bag:{rst_no}"
    is_sale = _is_sale(vw_entry.get("trans_type", ""))
    bag_type = (vw_entry.get("bag_type") or "").strip()
    qty = int(vw_entry.get("tot_pkts", 0) or 0)

    # Conditions to KEEP a bag entry: must be sale, must have type, must have qty
    if not (is_sale and bag_type and qty > 0):
        # Remove any existing linked entry (sale switched off, qty cleared, etc.)
        await db.gunny_bags.delete_many({"reference": ref})
        return

    existing = await db.gunny_bags.find_one({"reference": ref}, {"_id": 0})
    now = datetime.now(timezone.utc).isoformat()
    fields = {
        "date": vw_entry.get("date", ""),
        "bag_type": bag_type,
        "txn_type": "out",
        "quantity": qty,
        "source": (vw_entry.get("party_name") or "VW Sale").strip(),
        "rate": 0,
        "amount": 0,
        "notes": f"Auto from VW Sale (RST #{rst_no})",
        "kms_year": vw_entry.get("kms_year", ""),
        "season": vw_entry.get("season", "Kharif"),
        "created_by": username or "system",
        "linked_entry_id": vw_entry_id,
        "reference": ref,
        "rst_no": str(rst_no),
        "truck_no": vw_entry.get("vehicle_no", ""),
        "updated_at": now,
    }
    if existing:
        await db.gunny_bags.update_one({"reference": ref}, {"$set": fields})
    else:
        fields["id"] = str(uuid.uuid4())
        fields["created_at"] = now
        await db.gunny_bags.insert_one(fields)


async def _sync_sale_bhada_ledger(vw_entry_id: str, vw_entry: dict, username: str = "system"):
    """Sync truck-owner ledger entry for sale-trip OR purchase-trip bhada (lump-sum truck rent).

    For Sale dispatches: mill ne truck owner ko bhada dena hota hai (e.g. ₹4000 fixed).
    For Purchase trips:  mill ne truck owner ko inbound paddy ka bhada dena hota hai.
    Both create JAMA (CR — mill owes) entries on the truck owner ledger.

    References (idempotent, distinct so a single RST never collides between sale & purchase):
        Sale     → `vw_sale_bhada:{rst_no}`
        Purchase → `vw_purchase_bhada:{rst_no}`

    Conditions to keep the entry: vehicle_no set, bhada > 0, AND trans_type=sale OR purchase.
    """
    rst_no = vw_entry.get("rst_no")
    if rst_no is None:
        return
    is_sale = _is_sale(vw_entry.get("trans_type", ""))
    trans_type_lower = (vw_entry.get("trans_type", "") or "").lower()
    is_purchase = ("purchase" in trans_type_lower) or ("receive" in trans_type_lower)
    vehicle_no = (vw_entry.get("vehicle_no") or "").strip()
    bhada = float(vw_entry.get("bhada", 0) or 0)

    # Distinct references for sale vs purchase to avoid collision on same RST.
    sale_ref = f"vw_sale_bhada:{rst_no}"
    purchase_ref = f"vw_purchase_bhada:{rst_no}"
    active_ref = sale_ref if is_sale else (purchase_ref if is_purchase else None)
    inactive_ref = purchase_ref if is_sale else (sale_ref if is_purchase else None)

    # Always clear the inactive (other-direction) reference if any
    if inactive_ref:
        await db.cash_transactions.delete_many({"reference": inactive_ref})

    if not (active_ref and vehicle_no and bhada > 0):
        if active_ref:
            await db.cash_transactions.delete_many({"reference": active_ref})
        return

    existing = await db.cash_transactions.find_one({"reference": active_ref}, {"_id": 0})
    now = datetime.now(timezone.utc).isoformat()
    label = "Sale" if is_sale else "Purchase"
    fields = {
        "date": vw_entry.get("date", ""),
        "account": "ledger",
        "txn_type": "jama",
        "category": vehicle_no,
        "party_type": "Truck",
        "amount": bhada,
        "description": f"{label} Bhada (RST #{rst_no}) → {vw_entry.get('farmer_name') or vw_entry.get('party_name','')}",
        "kms_year": vw_entry.get("kms_year", ""),
        "season": vw_entry.get("season", "Kharif"),
        "created_by": username or "system",
        "linked_entry_id": vw_entry_id,
        "reference": active_ref,
        "updated_at": now,
    }
    if existing:
        await db.cash_transactions.update_one({"reference": active_ref}, {"$set": fields})
    else:
        fields["id"] = str(uuid.uuid4())
        fields["created_at"] = now
        await db.cash_transactions.insert_one(fields)


def _save_image(entry_id: str, tag: str, b64data) -> str:
    try:
        if not b64data or not isinstance(b64data, str):
            return ""
        raw = b64data
        # Strip data URL prefix if present (data:image/jpeg;base64,...)
        if raw.startswith("data:"):
            comma_idx = raw.find(",")
            if comma_idx > 0:
                raw = raw[comma_idx + 1:]
        if not raw or len(raw) < 100:
            return ""
        filename = f"{entry_id}_{tag}.jpg"
        with open(os.path.join(IMG_DIR, filename), "wb") as f:
            f.write(b64mod.b64decode(raw))
        return filename
    except Exception as e:
        print(f"[VW] save_image error: {e}")
        return ""


def _load_image_b64(filename: str) -> str:
    if not filename:
        return ""
    fp = os.path.join(IMG_DIR, filename)
    if os.path.exists(fp):
        with open(fp, "rb") as f:
            return b64mod.b64encode(f.read()).decode()
    return ""


@router.get("/vehicle-weight/image/{filename}")
async def serve_image(filename: str):
    """Serve a saved vehicle weight image file (for WhatsApp media_url)."""
    fp = os.path.join(IMG_DIR, filename)
    if not os.path.exists(fp):
        raise HTTPException(status_code=404, detail="Image not found")
    return FileResponse(fp, media_type="image/jpeg")


@router.get("/vehicle-weight/{entry_id}/photos")
async def get_entry_photos(entry_id: str):
    """Get all photos for a vehicle weight entry as base64."""
    entry = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {
        "entry_id": entry_id,
        "rst_no": entry.get("rst_no"),
        "date": entry.get("date", ""),
        "vehicle_no": entry.get("vehicle_no", ""),
        "party_name": entry.get("party_name", ""),
        "farmer_name": entry.get("farmer_name", ""),
        "product": entry.get("product", ""),
        "trans_type": entry.get("trans_type", ""),
        "tot_pkts": entry.get("tot_pkts", 0),
        "first_wt": entry.get("first_wt", 0),
        "first_wt_time": entry.get("first_wt_time", ""),
        "second_wt": entry.get("second_wt", 0),
        "second_wt_time": entry.get("second_wt_time", ""),
        "net_wt": entry.get("net_wt", 0),
        "gross_wt": entry.get("gross_wt", 0),
        "tare_wt": entry.get("tare_wt", 0),
        "remark": entry.get("remark", ""),
        "cash_paid": entry.get("cash_paid", 0),
        "diesel_paid": entry.get("diesel_paid", 0),
        "g_issued": entry.get("g_issued", 0),
        "tp_no": entry.get("tp_no", ""),
        "tp_weight": entry.get("tp_weight", 0),
        "first_wt_front_img": _load_image_b64(entry.get("first_wt_front_img", "")),
        "first_wt_side_img": _load_image_b64(entry.get("first_wt_side_img", "")),
        "second_wt_front_img": _load_image_b64(entry.get("second_wt_front_img", "")),
        "second_wt_side_img": _load_image_b64(entry.get("second_wt_side_img", "")),
    }


async def _next_rst(kms_year: str = ""):
    """Auto-increment RST number — max+1 across ALL collections that use RST
    (vehicle_weights, sale_vouchers, purchase_vouchers, private_paddy, entries,
    by_product_sale_vouchers) to prevent duplicates from manual entry in any form.
    v104.44.29: Cross-collection scan."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    used_rsts = set()
    for coll_name in ["vehicle_weights", "sale_vouchers", "purchase_vouchers",
                      "private_paddy", "entries", "by_product_sale_vouchers"]:
        try:
            docs = await db[coll_name].find(query, {"_id": 0, "rst_no": 1}).to_list(length=50000)
            for d in docs:
                raw = d.get("rst_no", "")
                try:
                    used_rsts.add(int(str(raw).strip() or 0))
                except (ValueError, TypeError):
                    pass
        except Exception:
            pass
    max_rst = max(used_rsts) if used_rsts else 0
    return max_rst + 1


@router.get("/vehicle-weight")
async def list_weights(kms_year: str = "", status: str = "", page: int = 1, page_size: int = 200,
                       date_from: str = "", date_to: str = "", vehicle_no: str = "",
                       party_name: str = "", farmer_name: str = "", rst_no: str = "",
                       trans_type: str = ""):
    """List weight entries with pagination and filters.
    
    `trans_type` filter accepts:
      - "purchase" → matches `Receive(Purchase)` / `Receive(Pur)` / contains "purchase"/"receive"
      - "sale"     → matches `Dispatch(Sale)` / contains "sale"/"dispatch"
      - exact value (e.g. "Receive(Purchase)") → exact match
      - empty      → no filter
    """
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if status:
        query["status"] = status
    has_search = any([vehicle_no, party_name, farmer_name, rst_no])
    if not has_search and (date_from or date_to):
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        query["date"] = date_q
    if vehicle_no:
        query["vehicle_no"] = {"$regex": vehicle_no, "$options": "i"}
    if party_name:
        query["party_name"] = {"$regex": party_name, "$options": "i"}
    if farmer_name:
        query["farmer_name"] = {"$regex": farmer_name, "$options": "i"}
    if rst_no:
        try: query["rst_no"] = int(rst_no)
        except: pass
    if trans_type:
        tt = trans_type.lower().strip()
        if tt == "sale":
            query["trans_type"] = {"$regex": "sale|dispatch", "$options": "i"}
        elif tt == "purchase":
            query["trans_type"] = {"$regex": "purchase|receive", "$options": "i"}
        else:
            query["trans_type"] = trans_type
    total_count = await db["vehicle_weights"].count_documents(query)
    if page_size < 1: page_size = 200
    if page < 1: page = 1
    skip = (page - 1) * page_size
    items = await db["vehicle_weights"].find(query, {"_id": 0}).sort([("date", -1), ("rst_no", -1)]).skip(skip).limit(page_size).to_list(page_size)
    return {"entries": items, "count": len(items), "total": total_count, "page": page, "page_size": page_size, "total_pages": max(1, (total_count + page_size - 1) // page_size)}


@router.get("/vehicle-weight/pending")
async def pending_vehicles(kms_year: str = ""):
    """Get vehicles with only first weight (waiting for second weight)."""
    query = {"status": "pending"}
    if kms_year:
        query["kms_year"] = kms_year
    cursor = db["vehicle_weights"].find(query, {"_id": 0}).sort([("date", -1), ("rst_no", -1)])
    items = await cursor.to_list(length=100)
    return {"pending": items, "count": len(items)}


@router.get("/vehicle-weight/next-rst")
async def get_next_rst(kms_year: str = ""):
    """Get next RST number."""
    rst = await _next_rst(kms_year)
    return {"rst_no": rst}


@router.get("/settings/vw-date-lock")
async def get_vw_date_lock():
    doc = await db["settings"].find_one({"key": "vw_date_lock"}, {"_id": 0})
    return {"locked": doc.get("locked", False) if doc else False}


@router.put("/settings/vw-date-lock")
async def update_vw_date_lock(data: dict):
    locked = data.get("locked", False)
    await db["settings"].update_one(
        {"key": "vw_date_lock"},
        {"$set": {"key": "vw_date_lock", "locked": locked}},
        upsert=True
    )
    return {"success": True, "locked": locked}


@router.get("/vehicle-weight/rst-edit-setting")
async def get_rst_edit_setting():
    """Get manual RST edit toggle setting."""
    doc = await db["settings"].find_one({"key": "manual_rst_edit"}, {"_id": 0})
    return {"enabled": doc.get("enabled", False) if doc else False}


@router.put("/vehicle-weight/rst-edit-setting")
async def update_rst_edit_setting(data: dict):
    """Toggle manual RST edit on/off."""
    enabled = data.get("enabled", False)
    await db["settings"].update_one(
        {"key": "manual_rst_edit"},
        {"$set": {"key": "manual_rst_edit", "enabled": enabled}},
        upsert=True
    )
    return {"success": True, "enabled": enabled}


@router.get("/vehicle-weight/auto-notify-setting")
async def get_auto_notify_setting():
    """Get auto VW messaging setting with group config."""
    doc = await db["settings"].find_one({"key": "auto_vw_messaging"}, {"_id": 0})
    if doc:
        return {
            "enabled": doc.get("enabled", False),
            "wa_group_id": doc.get("wa_group_id", ""),
            "wa_group_name": doc.get("wa_group_name", ""),
            "tg_chat_ids": doc.get("tg_chat_ids", []),
        }
    return {"enabled": False, "wa_group_id": "", "wa_group_name": "", "tg_chat_ids": []}


@router.put("/vehicle-weight/auto-notify-setting")
async def update_auto_notify_setting(data: dict):
    """Update auto VW messaging setting with group config."""
    update_fields = {"key": "auto_vw_messaging"}
    if "enabled" in data:
        update_fields["enabled"] = data["enabled"]
    if "wa_group_id" in data:
        update_fields["wa_group_id"] = data["wa_group_id"]
    if "wa_group_name" in data:
        update_fields["wa_group_name"] = data["wa_group_name"]
    if "tg_chat_ids" in data:
        update_fields["tg_chat_ids"] = data["tg_chat_ids"]
    await db["settings"].update_one(
        {"key": "auto_vw_messaging"},
        {"$set": update_fields},
        upsert=True
    )
    return {"success": True, **{k: v for k, v in update_fields.items() if k != "key"}}


@router.post("/vehicle-weight/auto-notify")
async def auto_notify_weight(data: dict):
    """Auto-send weight report PDF to WhatsApp & Telegram."""
    entry_id = data.get("entry_id", "")
    weight_type = data.get("weight_type", "1st")

    entry = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    rst = entry.get("rst_no", "?")

    # Generate weight report PDF
    pdf_buf = None
    try:
        pdf_response = await weight_report_pdf(entry_id)
        pdf_buf = io.BytesIO()
        async for chunk in pdf_response.body_iterator:
            pdf_buf.write(chunk)
        pdf_buf.seek(0)
    except Exception as e:
        logger.error(f"PDF generation error: {e}")

    caption = f"*{weight_type} Weight Report - RST #{rst}*"

    results = {"whatsapp": [], "telegram": []}
    vw_config = await db["settings"].find_one({"key": "auto_vw_messaging"}, {"_id": 0}) or {}
    vw_wa_group_id = vw_config.get("wa_group_id", "")
    vw_tg_chat_ids = vw_config.get("tg_chat_ids", [])

    # Upload PDF to tmpfiles.org for public URL
    pdf_url = ""
    if pdf_buf:
        pdf_buf.seek(0)
        pdf_bytes = pdf_buf.read()
        try:
            import aiohttp
            form = aiohttp.FormData()
            form.add_field('file', pdf_bytes, filename=f'WeightReport_RST{rst}.pdf', content_type='application/pdf')
            async with aiohttp.ClientSession() as session:
                async with session.post('https://tmpfiles.org/api/v1/upload', data=form, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                    result = await resp.json()
                    logger.info(f"tmpfiles upload: {result}")
                    if result.get('status') == 'success' and result.get('data', {}).get('url'):
                        pdf_url = result['data']['url'].replace('://tmpfiles.org/', '://tmpfiles.org/dl/').replace('http://', 'https://')
        except Exception as e:
            logger.error(f"tmpfiles upload error: {e}")
            # Fallback to direct URL
            base_url = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
            pdf_url = f"{base_url}/api/vehicle-weight/{entry_id}/weight-report-pdf?t={int(datetime.now().timestamp())}" if base_url else ""
    
    logger.info(f"PDF URL for WA: {pdf_url}")

    # Send via WhatsApp - caption + PDF URL
    try:
        from routes.whatsapp import _get_wa_settings
        wa_settings = await _get_wa_settings()
        if wa_settings.get("enabled") and wa_settings.get("api_key"):
            api_key = wa_settings["api_key"]

            from routes.whatsapp import _wa_base_url
            wa_base = _wa_base_url(wa_settings)
            async with httpx.AsyncClient(timeout=30) as client:
                if vw_wa_group_id:
                    wa_data = {"groupId": vw_wa_group_id, "text": caption}
                    if pdf_url:
                        wa_data["url"] = pdf_url
                    resp = await client.post(f"{wa_base}/sendGroup",
                        data=wa_data, headers={"Authorization": f"Bearer {api_key}"})
                    logger.info(f"WA group resp: {resp.status_code} {resp.text[:200]}")
                    results["whatsapp"].append({"success": resp.status_code == 201 or resp.json().get("success")})
                else:
                    numbers = wa_settings.get("default_numbers", [])
                    for num in numbers:
                        if num:
                            wa_data = {"phonenumber": num.strip(), "text": caption}
                            if pdf_url:
                                wa_data["url"] = pdf_url
                            resp = await client.post(f"{wa_base}/sendMessage",
                                data=wa_data, headers={"Authorization": f"Bearer {api_key}"})
                            results["whatsapp"].append({"success": resp.status_code == 201 or resp.json().get("success")})
    except Exception as e:
        logger.error(f"WA auto-notify error: {e}")

    try:
        from routes.telegram import get_telegram_config
        tg_config = await get_telegram_config()
        if tg_config and tg_config.get("bot_token"):
            bot_token = tg_config["bot_token"]
            chat_ids = vw_tg_chat_ids if vw_tg_chat_ids and len(vw_tg_chat_ids) > 0 else (tg_config.get("chat_ids") or [])
            if chat_ids and pdf_buf:
                pdf_buf.seek(0)
                pdf_bytes = pdf_buf.read()
                async with httpx.AsyncClient() as client:
                    for item in chat_ids:
                        cid = str(item.get("chat_id", "")).strip()
                        if cid:
                            try:
                                files = {"document": (f"WeightReport_RST{rst}.pdf", pdf_bytes, "application/pdf")}
                                r = await client.post(f"https://api.telegram.org/bot{bot_token}/sendDocument",
                                    data={"chat_id": cid, "caption": caption}, files=files, timeout=30)
                                results["telegram"].append(r.json())
                            except Exception as te:
                                logger.error(f"TG send doc error: {te}")
    except Exception as e:
        logger.error(f"Telegram auto-notify error: {e}")

    wa_sent = sum(1 for r in results["whatsapp"] if r.get("success"))
    tg_sent = sum(1 for r in results["telegram"] if r.get("ok"))
    return {"success": True, "message": f"WA: {wa_sent} sent, TG: {tg_sent} sent", "results": results}


@router.get("/vehicle-weight/by-rst/{rst_no}")
async def get_by_rst(rst_no: int, kms_year: str = ""):
    """Lookup vehicle weight entry by RST number - used by Entries form auto-fill."""
    query = {"rst_no": rst_no}
    if kms_year:
        query["kms_year"] = kms_year
    entry = await db["vehicle_weights"].find_one(query, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="RST not found in Vehicle Weight")
    return {"success": True, "entry": entry}


@router.get("/vehicle-weight/linked-rst")
async def get_linked_rst(kms_year: str = ""):
    """Get RST numbers from Mill Entries that are linked to Vehicle Weight entries."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    # Get all RST numbers from mill_entries
    entries = await db["mill_entries"].find(query, {"_id": 0, "rst_no": 1}).to_list(50000)
    linked = set()
    for e in entries:
        r = e.get("rst_no", "")
        if r and r.strip():
            try: linked.add(int(r))
            except: pass
    return {"linked_rst": list(linked)}


@router.get("/vehicle-weight/linked-rst-sale")
async def get_linked_rst_sale(kms_year: str = ""):
    """Get RST numbers from DC Deliveries (sale dispatches) that are linked to Vehicle Weight Sale entries.
    Handles slash-joined multi-truck RSTs like '123 / 124'."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    dels = await db["dc_deliveries"].find(query, {"_id": 0, "rst_no": 1}).to_list(50000)
    linked = set()
    for d in dels:
        raw = (d.get("rst_no") or "").strip()
        if not raw:
            continue
        # Split slash-joined RSTs (multi-truck deliveries)
        for part in raw.split("/"):
            p = part.strip()
            if p:
                try: linked.add(int(p))
                except: pass
    return {"linked_rst": list(linked)}


@router.get("/vehicle-weight/linked-rst-bp-sale")
async def get_linked_rst_bp_sale(kms_year: str = ""):
    """Get RST numbers from BP Sale Register (rice bran / by-product sales) linked to Vehicle Weight Sale entries.
    BP sales reference an RST of an Auto Weight sale entry by its `rst_no` field."""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    sales = await db["bp_sale_register"].find(query, {"_id": 0, "rst_no": 1}).to_list(50000)
    linked = set()
    for s in sales:
        raw = (s.get("rst_no") or "").strip() if isinstance(s.get("rst_no"), str) else str(s.get("rst_no") or "")
        if not raw:
            continue
        for part in raw.split("/"):
            p = part.strip()
            if p:
                try: linked.add(int(p))
                except: pass
    return {"linked_rst": list(linked)}


@router.get("/vehicle-weight/pending-count")
async def get_pending_vw_count(kms_year: str = ""):
    """Count VW entries that don't have a corresponding Mill Entry."""
    vw_query = {"status": "completed"}
    me_query = {}
    if kms_year:
        vw_query["kms_year"] = kms_year
        me_query["kms_year"] = kms_year
    # Get all VW RST numbers
    vw_entries = await db["vehicle_weights"].find(vw_query, {"_id": 0, "rst_no": 1}).to_list(50000)
    vw_rsts = set()
    for e in vw_entries:
        r = e.get("rst_no")
        if r is not None:
            try: vw_rsts.add(int(r))
            except: pass
    # Get all Mill Entry RST numbers
    me_entries = await db["mill_entries"].find(me_query, {"_id": 0, "rst_no": 1}).to_list(50000)
    linked = set()
    for e in me_entries:
        r = e.get("rst_no", "")
        if r and r.strip():
            try: linked.add(int(r))
            except: pass
    pending = vw_rsts - linked
    return {"pending_count": len(pending), "total_vw": len(vw_rsts), "linked": len(linked & vw_rsts)}



@router.post("/vehicle-weight/send-manual")
async def send_manual_weight_msg(data: dict):
    """Manual send weight report PDF to WhatsApp/Telegram."""
    entry_id = data.get("entry_id", "")
    send_to_numbers = data.get("send_to_numbers", False)
    send_to_group = data.get("send_to_group", False)

    entry = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    rst = entry.get("rst_no", "?")
    first_wt = float(entry.get("first_wt", 0) or 0)
    second_wt = float(entry.get("second_wt", 0) or 0)
    net_wt = float(entry.get("net_wt", 0) or 0)
    avg_wt = round((first_wt + second_wt) / 2, 2) if (first_wt and second_wt) else 0

    caption = f"*Weight Report - RST #{rst}*\n"
    caption += f"Vehicle: {entry.get('vehicle_no','')}\n"
    caption += f"Party: {entry.get('party_name','')}\n"
    if first_wt: caption += f"1st Wt: {first_wt:,.0f} KG\n"
    if second_wt: caption += f"2nd Wt: {second_wt:,.0f} KG\n"
    if net_wt: caption += f"*Net Wt: {net_wt:,.0f} KG*\n"
    if avg_wt: caption += f"Avg Wt: {avg_wt:,.2f} KG"

    # Generate PDF
    pdf_buf = None
    try:
        pdf_response = await weight_report_pdf(entry_id)
        pdf_buf = io.BytesIO()
        async for chunk in pdf_response.body_iterator:
            pdf_buf.write(chunk)
        pdf_buf.seek(0)
    except Exception as e:
        logger.error(f"PDF generation error: {e}")

    base_url = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
    pdf_url = f"{base_url}/api/vehicle-weight/{entry_id}/weight-report-pdf" if base_url else ""

    results = {"whatsapp": [], "telegram": []}

    # WhatsApp
    try:
        from routes.whatsapp import _get_wa_settings, _send_wa_message, _send_wa_to_group
        wa_settings = await _get_wa_settings()
        if wa_settings.get("enabled") and wa_settings.get("api_key"):
            if send_to_numbers:
                numbers = wa_settings.get("default_numbers", [])
                for num in numbers:
                    if num:
                        r = await _send_wa_message(num.strip(), caption, pdf_url if pdf_url else None)
                        results["whatsapp"].append({"to": num, "success": r.get("success", False)})
            if send_to_group:
                group_id = wa_settings.get("default_group_id", "")
                if group_id:
                    r = await _send_wa_to_group(group_id, caption, pdf_url if pdf_url else None)
                    results["whatsapp"].append({"to": "group", "success": r.get("success", False)})
    except Exception as e:
        logger.error(f"WA manual send error: {e}")

    # Telegram - send PDF as document
    try:
        from routes.telegram import get_telegram_config
        tg_config = await get_telegram_config()
        if tg_config and tg_config.get("bot_token") and tg_config.get("chat_ids") and pdf_buf:
            bot_token = tg_config["bot_token"]
            chat_ids = tg_config["chat_ids"]
            pdf_buf.seek(0)
            pdf_bytes = pdf_buf.read()
            async with httpx.AsyncClient() as client:
                for item in chat_ids:
                    cid = str(item.get("chat_id", "")).strip()
                    if cid:
                        try:
                            files = {"document": (f"WeightReport_RST{rst}.pdf", pdf_bytes, "application/pdf")}
                            form_data = {"chat_id": cid, "caption": caption.replace("*", "")}
                            r = await client.post(
                                f"https://api.telegram.org/bot{bot_token}/sendDocument",
                                data=form_data, files=files, timeout=30
                            )
                            results["telegram"].append(r.json())
                        except Exception as te:
                            logger.error(f"TG send doc error: {te}")
    except Exception as e:
        logger.error(f"TG manual send error: {e}")

    wa_sent = sum(1 for r in results["whatsapp"] if r.get("success"))
    tg_sent = sum(1 for r in results["telegram"] if r.get("ok"))
    return {"success": True, "message": f"WA: {wa_sent}, TG: {tg_sent}", "results": results}


@router.post("/vehicle-weight")
async def create_weight_entry(data: dict):
    """Create new weight entry with first weight."""
    kms_year = data.get("kms_year", "")
    # Allow custom RST number, fallback to auto-increment
    custom_rst = data.get("rst_no")
    if custom_rst and int(custom_rst) > 0:
        rst_no = int(custom_rst)
        # Check for duplicate RST in same kms_year
        dup_query = {"rst_no": rst_no}
        if kms_year:
            dup_query["kms_year"] = kms_year
        existing = await db["vehicle_weights"].find_one(dup_query)
        if existing:
            raise HTTPException(status_code=400, detail=f"RST #{rst_no} already exists! Duplicate RST number.")
    else:
        rst_no = await _next_rst(kms_year)
        # Double-check no duplicate exists (race condition guard)
        dup_check = {"rst_no": rst_no}
        if kms_year:
            dup_check["kms_year"] = kms_year
        if await db["vehicle_weights"].find_one(dup_check):
            # Recalculate to be safe
            rst_no = await _next_rst(kms_year)

    # TP No. duplicate check
    tp_no_raw = (data.get("tp_no", "") or "").strip()
    if tp_no_raw:
        tp_dup_query = {"tp_no": tp_no_raw}
        if kms_year:
            tp_dup_query["kms_year"] = kms_year
        tp_existing = await db["vehicle_weights"].find_one(tp_dup_query, {"_id": 0, "rst_no": 1})
        if tp_existing:
            raise HTTPException(status_code=400, detail=f"TP No. {tp_no_raw} already RST #{tp_existing['rst_no']} mein hai! Duplicate TP allowed nahi hai.")

    entry = {
        "id": str(uuid.uuid4()),
        "rst_no": rst_no,
        "date": data.get("date", datetime.now().strftime("%Y-%m-%d")),
        "kms_year": kms_year,
        "vehicle_no": (data.get("vehicle_no", "") or "").strip().upper(),
        "party_name": (data.get("party_name", "") or "").strip(),
        "tp_no": tp_no_raw,
        "tp_weight": float(data.get("tp_weight", 0) or 0),
        "g_issued": float(data.get("g_issued", 0) or 0),
        "farmer_name": (data.get("farmer_name", "") or "").strip(),
        "product": data.get("product", "PADDY"),
        "trans_type": data.get("trans_type", "Receive(Pur)"),
        "bag_type": (data.get("bag_type", "") or "").strip(),
        "j_pkts": int(data.get("j_pkts", 0) or 0),
        "p_pkts": int(data.get("p_pkts", 0) or 0),
        "tot_pkts": int(data.get("tot_pkts", 0) or 0),
        "first_wt": float(data.get("first_wt", 0) or 0),
        "first_wt_time": datetime.now(timezone.utc).isoformat(),
        "second_wt": 0,
        "second_wt_time": "",
        "net_wt": 0,
        "remark": data.get("remark", ""),
        "cash_paid": float(data.get("cash_paid", 0) or 0),
        "diesel_paid": float(data.get("diesel_paid", 0) or 0),
        "bhada": float(data.get("bhada", 0) or 0),
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    eid = entry["id"]
    entry["first_wt_front_img"] = _save_image(eid, "1st_front", data.get("first_wt_front_img", ""))
    entry["first_wt_side_img"] = _save_image(eid, "1st_side", data.get("first_wt_side_img", ""))

    await db["vehicle_weights"].insert_one(entry)
    entry.pop("_id", None)
    # Sync gunny bag stock-out for sale dispatches
    await _sync_sale_bag_out(entry["id"], entry, data.get("username", "system"))
    # Sync truck-owner bhada ledger for sale dispatches
    await _sync_sale_bhada_ledger(entry["id"], entry, data.get("username", "system"))
    return {"success": True, "entry": entry, "message": f"RST #{rst_no} - First weight saved!"}


@router.put("/vehicle-weight/{entry_id}/second-weight")
async def update_second_weight(entry_id: str, data: dict):
    """Update second weight, cash/diesel and calculate net weight."""
    entry = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    second_wt = float(data.get("second_wt", 0) or 0)
    first_wt = entry["first_wt"]

    # Validation: 2nd weight should not be greater than 1st weight (negative net not allowed)
    if second_wt > first_wt:
        raise HTTPException(status_code=400, detail=f"2nd Weight ({int(second_wt)} KG) pehle weight ({int(first_wt)} KG) se zyada hai! Negative weight entry allowed nahi hai.")

    net_wt = abs(first_wt - second_wt)
    gross_wt = max(first_wt, second_wt)
    tare_wt = min(first_wt, second_wt)

    update_fields = {
        "second_wt": second_wt,
        "second_wt_time": datetime.now(timezone.utc).isoformat(),
        "net_wt": net_wt,
        "gross_wt": gross_wt,
        "tare_wt": tare_wt,
        "status": "completed"
    }
    # Save second weight camera photos
    front_img = _save_image(entry_id, "2nd_front", data.get("second_wt_front_img", ""))
    side_img = _save_image(entry_id, "2nd_side", data.get("second_wt_side_img", ""))
    if front_img:
        update_fields["second_wt_front_img"] = front_img
    if side_img:
        update_fields["second_wt_side_img"] = side_img
    # Update cash/diesel if provided during second weight capture
    if "cash_paid" in data:
        update_fields["cash_paid"] = float(data.get("cash_paid", 0) or 0)
    if "diesel_paid" in data:
        update_fields["diesel_paid"] = float(data.get("diesel_paid", 0) or 0)
    if "bhada" in data:
        update_fields["bhada"] = float(data.get("bhada", 0) or 0)
    if "g_issued" in data:
        update_fields["g_issued"] = float(data.get("g_issued", 0) or 0)
    if "tp_no" in data:
        new_tp = str(data.get("tp_no", "")).strip()
        if new_tp:
            tp_dup_query = {"tp_no": new_tp, "id": {"$ne": entry_id}}
            entry_kms = entry.get("kms_year", "")
            if entry_kms:
                tp_dup_query["kms_year"] = entry_kms
            tp_existing = await db["vehicle_weights"].find_one(tp_dup_query, {"_id": 0, "rst_no": 1})
            if tp_existing:
                raise HTTPException(status_code=400, detail=f"TP No. {new_tp} already RST #{tp_existing['rst_no']} mein hai! Duplicate TP allowed nahi hai.")
        update_fields["tp_no"] = new_tp
    if "tp_weight" in data:
        update_fields["tp_weight"] = float(data.get("tp_weight", 0) or 0)
    if "tot_pkts" in data:
        update_fields["tot_pkts"] = int(data.get("tot_pkts", 0) or 0)

    await db["vehicle_weights"].update_one(
        {"id": entry_id},
        {"$set": update_fields}
    )

    updated = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    # Sync truck-owner bhada ledger after second weight capture (if Sale + bhada provided)
    if updated:
        await _sync_sale_bhada_ledger(entry_id, updated, "system")
    return {"success": True, "entry": updated, "message": f"RST #{entry['rst_no']} - Net Wt: {net_wt} KG"}


@router.delete("/vehicle-weight/{entry_id}")
async def delete_weight_entry(entry_id: str, username: str = "", role: str = ""):
    """Delete a weight entry + cascade delete linked mill entry & transactions."""
    from services.edit_lock import check_edit_lock
    vw = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not vw:
        raise HTTPException(status_code=404, detail="Entry not found")
    can_edit, message = await check_edit_lock(vw, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    rst_no = vw.get("rst_no")
    kms_year = vw.get("kms_year", "")

    # Cascade: find and delete linked mill entry
    cascade_deleted = []
    if rst_no is not None:
        # rst_no can be int in VW but str in mill_entries - check both types
        mill_q = {"$or": [{"rst_no": rst_no}, {"rst_no": str(rst_no)}, {"rst_no": int(rst_no) if str(rst_no).isdigit() else rst_no}]}
        if kms_year:
            mill_q = {"$and": [mill_q, {"kms_year": kms_year}]}
        linked_entries = await db.mill_entries.find(mill_q, {"_id": 0}).to_list(10)
        for linked_entry in linked_entries:
            eid = linked_entry["id"]
            await db.mill_entries.delete_one({"id": eid})
            await db.cash_transactions.delete_many({"linked_entry_id": eid})
            await db.diesel_accounts.delete_many({"linked_entry_id": eid})
            await db.gunny_bags.delete_many({"linked_entry_id": eid})
            cascade_deleted.append(f"Mill Entry RST #{rst_no}")

    await db["vehicle_weights"].delete_one({"id": entry_id})
    # Cascade: also remove linked sale-bag-out + bhada-ledger entries if any
    if rst_no is not None:
        await db.gunny_bags.delete_many({"reference": f"vw_sale_bag:{rst_no}"})
        await db.cash_transactions.delete_many({"reference": f"vw_sale_bhada:{rst_no}"})
        await db.cash_transactions.delete_many({"reference": f"vw_purchase_bhada:{rst_no}"})
    msg = "Entry deleted"
    if cascade_deleted:
        msg += f" + {', '.join(cascade_deleted)} bhi delete kiya"
    return {"success": True, "message": msg}


@router.put("/vehicle-weight/{entry_id}/edit")
async def edit_weight_entry(entry_id: str, data: dict, username: str = "", role: str = ""):
    """Edit completed weight entry (Vehicle, Party, Product, Pkts, Cash, Diesel)."""
    from services.edit_lock import check_edit_lock
    entry = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    can_edit, message = await check_edit_lock(entry, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)

    update_fields = {}
    editable = ["vehicle_no", "party_name", "farmer_name", "product", "tot_pkts", "bag_type", "cash_paid", "diesel_paid", "bhada", "g_issued", "tp_no", "tp_weight", "remark"]
    for f in editable:
        if f in data:
            if f in ("cash_paid", "diesel_paid", "bhada", "tp_weight"):
                update_fields[f] = float(data[f] or 0)
            elif f == "tot_pkts":
                update_fields[f] = data[f]
            elif f == "tp_no":
                new_tp = str(data[f] or "").strip()
                if new_tp:
                    tp_dup_query = {"tp_no": new_tp, "id": {"$ne": entry_id}}
                    entry_kms = entry.get("kms_year", "")
                    if entry_kms:
                        tp_dup_query["kms_year"] = entry_kms
                    tp_existing = await db["vehicle_weights"].find_one(tp_dup_query, {"_id": 0, "rst_no": 1})
                    if tp_existing:
                        raise HTTPException(status_code=400, detail=f"TP No. {new_tp} already RST #{tp_existing['rst_no']} mein hai! Duplicate TP allowed nahi hai.")
                update_fields[f] = new_tp
            else:
                update_fields[f] = data[f]

    if update_fields:
        await db["vehicle_weights"].update_one({"id": entry_id}, {"$set": update_fields})

    # Cascade edit to linked Mill Entry
    rst_no = entry.get("rst_no")
    kms_year = entry.get("kms_year", "")
    if rst_no is not None:
        mill_update = {}
        field_map = {
            "vehicle_no": "truck_no",
            "party_name": "party_name",
            "farmer_name": "mandi_name",
            "tp_no": "tp_no",
            "tp_weight": "tp_weight",
            "tot_pkts": "bag",
        }
        for vw_f, mill_f in field_map.items():
            if vw_f in update_fields:
                mill_update[mill_f] = update_fields[vw_f]
        if mill_update:
            mill_q = {"$or": [{"rst_no": rst_no}, {"rst_no": str(rst_no)}, {"rst_no": int(rst_no) if str(rst_no).isdigit() else rst_no}]}
            if kms_year:
                mill_q = {"$and": [mill_q, {"kms_year": kms_year}]}
            await db.mill_entries.update_many(mill_q, {"$set": mill_update})

    updated = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    # Re-sync bag stock-out (in case bag_type / tot_pkts edited)
    await _sync_sale_bag_out(entry_id, updated, username or "system")
    # Re-sync truck-owner bhada ledger (in case bhada / vehicle_no edited)
    await _sync_sale_bhada_ledger(entry_id, updated, username or "system")
    return {"success": True, "entry": updated}


@router.get("/vehicle-weight/{entry_id}/slip-pdf")
async def weight_slip_pdf(entry_id: str, party_only: int = 0):
    """Generate weight slip PDF - A5 portrait. party_only=1 for single Party Copy (download), 0 for 2 copies (print)."""
    from reportlab.lib.pagesizes import A5
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from fastapi.responses import StreamingResponse

    entry = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    from utils.branding_helper import get_branding_data
    from utils.export_helpers import register_hindi_fonts
    register_hindi_fonts()
    branding = await get_branding_data()
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "JOLKO, KESINGA")
    custom_fields = branding.get("custom_fields", [])
    above_parts, below_parts = [], []
    for f in custom_fields:
        val = f.get("value", "").strip()
        if not val: continue
        lbl = f.get("label", "").strip()
        txt = f"{lbl}: {val}" if lbl else val
        if f.get("placement", "below") == "above":
            above_parts.append(txt)
        else:
            below_parts.append(txt)
    above_text = "  |  ".join(above_parts) if above_parts else ""
    below_text = "  |  ".join(below_parts) if below_parts else ""

    # A5 portrait = 148mm x 210mm
    W, H = A5  # (419.53, 595.28) points
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A5)

    first_wt = entry.get("first_wt", 0)
    second_wt = entry.get("second_wt", 0)
    net_wt = entry.get("net_wt", 0)
    gross_wt = entry.get("gross_wt", max(first_wt, second_wt))
    tare_wt = entry.get("tare_wt", min(first_wt, second_wt))
    cash = entry.get("cash_paid", 0) or 0
    diesel = entry.get("diesel_paid", 0) or 0
    rst = entry.get("rst_no", "")

    LM = 5*mm   # left margin
    RM = 5*mm   # right margin
    PW = W - LM - RM  # printable width ~138mm

    def draw_copy(c, top_y, copy_label, show_sig):
        """Draw one copy block starting from top_y (in points from bottom)."""
        x = LM
        y = top_y
        bh = 95*mm  # block height for each copy

        # Border box (outer)
        c.setStrokeColor(colors.HexColor("#1a1a2e"))
        c.setLineWidth(2)
        c.rect(x, y - bh, PW, bh)

        # Copy label (top right on border)
        c.setFont("Helvetica-Bold", 7)
        lw = c.stringWidth(copy_label, "Helvetica-Bold", 7)
        c.setFillColor(colors.white)
        c.rect(x + PW - lw - 14*mm, y - 1, lw + 4*mm, 6, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#888"))
        c.drawString(x + PW - lw - 12*mm, y + 0.5, copy_label)

        cy = y - 5*mm

        # Custom fields ABOVE company name
        if above_text:
            c.setFont("FreeSans", 8)
            c.setFillColor(colors.HexColor("#8B0000"))
            c.drawCentredString(W/2, cy, above_text)
            cy -= 4.5*mm

        # Company name - large bold
        c.setFont("Helvetica-Bold", 18)
        c.setFillColor(colors.HexColor("#000"))
        c.drawCentredString(W/2, cy, company)
        cy -= 5.5*mm

        # Tagline
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.gray)
        c.drawCentredString(W/2, cy, tagline)
        cy -= 3.5*mm

        # Custom fields BELOW tagline
        if below_text:
            c.setFont("FreeSans", 7)
            c.setFillColor(colors.HexColor("#374151"))
            c.drawCentredString(W/2, cy, below_text)
            cy -= 3*mm

        # Slip title
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(colors.HexColor("#333"))
        c.drawCentredString(W/2, cy, "WEIGHT SLIP / \u0924\u094c\u0932 \u092a\u0930\u094d\u091a\u0940")
        cy -= 2*mm

        # ── Bordered Info Table (4 rows x 4 cols) ──
        # Header separator line
        c.setStrokeColor(colors.HexColor("#1a1a2e"))
        c.setLineWidth(1.5)
        c.line(x, cy, x + PW, cy)

        _is_sale = ("sale" in str(entry.get("trans_type", "")).lower() or "dispatch" in str(entry.get("trans_type", "")).lower())
        rows = [
            ("RST No.", f"#{rst}", "Date / \u0926\u093f\u0928\u093e\u0902\u0915", fmt_date(entry.get("date", ""))),
            ("Vehicle / \u0917\u093e\u0921\u093c\u0940", entry.get("vehicle_no", ""), "Trans Type", entry.get("trans_type", "")),
            ("Party / \u092a\u093e\u0930\u094d\u091f\u0940", entry.get("party_name", ""), "Destination" if _is_sale else "Source/Mandi", entry.get("farmer_name", "")),
            ("Product / \u092e\u093e\u0932", entry.get("product", ""), "Bags / \u092c\u094b\u0930\u0947", str(entry.get("tot_pkts", 0))),
        ]
        g_issued = float(entry.get("g_issued", 0) or 0)
        tp_no = entry.get("tp_no", "") or ""
        tp_weight = float(entry.get("tp_weight", 0) or 0)
        remark_text = entry.get("remark", "") or ""
        if g_issued > 0:
            rows.append(("G.Issued", f"{g_issued:,.0f}", "TP No.", tp_no or "-"))
        elif tp_no:
            rows.append(("TP No.", tp_no, "", ""))
        if tp_weight > 0:
            rows.append(("TP Weight", f"{tp_weight} Q", "", ""))
        if remark_text:
            rows.append(("Remark", remark_text, "", ""))
        rh = 6*mm  # row height - taller for proper table cells
        table_x = x
        c1w = PW * 0.18   # label col 1
        c2w = PW * 0.32   # value col 1
        c3w = PW * 0.18   # label col 2
        c4w = PW * 0.32   # value col 2

        for i, (l1, v1, l2, v2) in enumerate(rows):
            row_top = cy - i * rh
            row_bottom = row_top - rh

            # Draw cell borders
            c.setStrokeColor(colors.HexColor("#999"))
            c.setLineWidth(0.5)
            # Horizontal line at bottom of row
            c.line(table_x, row_bottom, table_x + PW, row_bottom)
            # Vertical lines for column separators
            c.line(table_x + c1w, row_top, table_x + c1w, row_bottom)
            c.line(table_x + c1w + c2w, row_top, table_x + c1w + c2w, row_bottom)
            c.line(table_x + c1w + c2w + c3w, row_top, table_x + c1w + c2w + c3w, row_bottom)

            # Text vertical center in cell
            text_y = row_top - rh * 0.65

            # Label 1
            c.setFont("FreeSans", 8)
            c.setFillColor(colors.HexColor("#333"))
            c.drawString(table_x + 2*mm, text_y, l1)

            # Value 1 - bold
            c.setFont("FreeSansBold" if i == 0 else "FreeSans", 10 if i == 0 else 9)
            c.setFillColor(colors.HexColor("#000"))
            c.drawString(table_x + c1w + 2*mm, text_y, str(v1)[:22])

            # Label 2
            c.setFont("FreeSans", 8)
            c.setFillColor(colors.HexColor("#333"))
            c.drawString(table_x + c1w + c2w + 2*mm, text_y, l2)

            # Value 2 - bold
            c.setFont("FreeSansBold" if i == 0 else "FreeSans", 10 if i == 0 else 9)
            c.setFillColor(colors.HexColor("#000"))
            c.drawString(table_x + c1w + c2w + c3w + 2*mm, text_y, str(v2)[:22])

        cy -= len(rows) * rh

        # Thick line separating table from weight boxes
        c.setStrokeColor(colors.HexColor("#1a1a2e"))
        c.setLineWidth(1.5)
        c.line(x, cy, x + PW, cy)
        cy -= 1*mm

        # ── Weight boxes (Gross | Tare | Net + optional Cash/Diesel) ──
        wt_items = [
            ("GROSS / \u0915\u0941\u0932", f"{gross_wt:,.0f} KG", "#f0f0f0", "#000", "#999"),
            ("TARE / \u0916\u093e\u0932\u0940", f"{tare_wt:,.0f} KG", "#f0f0f0", "#000", "#999"),
            ("NET / \u0936\u0941\u0926\u094d\u0927", f"{net_wt:,.0f} KG", "#dcf5dc", "#1b5e20", "#2e7d32"),
        ]
        if cash > 0:
            wt_items.append(("CASH / \u0928\u0915\u0926", f"\u20b9{cash:,.0f}", "#fff8e1", "#e65100", "#f9a825"))
        if diesel > 0:
            wt_items.append(("DIESEL / \u0921\u0940\u091c\u0932", f"\u20b9{diesel:,.0f}", "#fff8e1", "#e65100", "#f9a825"))

        num_cols = len(wt_items)
        col_w = PW / num_cols
        box_h = 13*mm

        for i, (label, val, bg, fg, bc) in enumerate(wt_items):
            bx = x + i * col_w
            # Background fill
            c.setFillColor(colors.HexColor(bg))
            c.rect(bx, cy - box_h, col_w, box_h, fill=1, stroke=0)
            # Cell border
            c.setStrokeColor(colors.HexColor(bc))
            c.setLineWidth(1.2 if "NET" in label else 0.5)
            c.rect(bx, cy - box_h, col_w, box_h)
            # Label (uppercase, small)
            c.setFont("Helvetica-Bold", 7)
            c.setFillColor(colors.HexColor("#555"))
            c.drawCentredString(bx + col_w/2, cy - 4*mm, label)
            # Value (large, bold)
            fz = 14 if "NET" in label else 11 if "CASH" in label or "DIESEL" in label else 12
            c.setFont("Helvetica-Bold", fz)
            c.setFillColor(colors.HexColor(fg))
            c.drawCentredString(bx + col_w/2, cy - 10*mm, val)

        cy -= box_h + 2*mm

        # ── Signature section (only Customer copy) ──
        if show_sig:
            sig_w = 38*mm
            c.setStrokeColor(colors.HexColor("#333"))
            c.setLineWidth(0.6)
            c.line(x + 8*mm, cy - 10*mm, x + 8*mm + sig_w, cy - 10*mm)
            c.setFont("Helvetica", 6)
            c.setFillColor(colors.HexColor("#555"))
            c.drawCentredString(x + 8*mm + sig_w/2, cy - 13*mm, "Driver / \u0921\u094d\u0930\u093e\u0907\u0935\u0930")
            c.line(x + PW - 8*mm - sig_w, cy - 10*mm, x + PW - 8*mm, cy - 10*mm)
            c.drawCentredString(x + PW - 8*mm - sig_w/2, cy - 13*mm, "Authorized / \u0905\u0927\u093f\u0915\u0943\u0924")

        # Footer
        c.setFont("Helvetica", 6)
        c.setFillColor(colors.HexColor("#999"))
        c.drawCentredString(W/2, y - bh + 2*mm, f"{company} | Computer Generated")

    # Draw copies based on mode
    top_margin = 5*mm
    copy1_top = H - top_margin

    if party_only:
        # Single Party Copy - centered vertically on A5
        draw_copy(c, copy1_top, "PARTY COPY", False)
    else:
        # 2 copies: Party + Customer
        draw_copy(c, copy1_top, "PARTY COPY", False)

        # Cut line
        cut_y = copy1_top - 95*mm - 3*mm
        c.setStrokeColor(colors.HexColor("#aaa"))
        c.setDash(3, 3)
        c.setLineWidth(0.8)
        c.line(LM, cut_y, W - RM, cut_y)
        c.setDash()
        c.setFont("Helvetica", 5)
        c.setFillColor(colors.HexColor("#aaa"))
        c.drawCentredString(W/2, cut_y + 1, "- - - CUT HERE / \u0915\u093e\u091f\u0947\u0902 - - -")

        copy2_top = cut_y - 2*mm
        draw_copy(c, copy2_top, "CUSTOMER COPY", True)

    c.save()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=WeightSlip_RST{rst}.pdf"})



@router.head("/vehicle-weight/{entry_id}/weight-report-pdf")
async def weight_report_pdf_head(entry_id: str):
    """HEAD support for WhatsApp media URL validation."""
    from fastapi.responses import Response
    entry = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Not found")
    return Response(status_code=200, headers={"Content-Type": "application/pdf"})

@router.get("/vehicle-weight/{entry_id}/weight-report-pdf")
async def weight_report_pdf(entry_id: str):
    """Generate compact single-page professional weight report PDF using shared branding."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image as RLImage
    from fastapi.responses import StreamingResponse

    entry = await db["vehicle_weights"].find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    from utils.branding_helper import get_pdf_header_elements_from_db
    from utils.export_helpers import register_hindi_fonts
    from utils.date_format import fmt_date
    register_hindi_fonts()

    rst = entry.get("rst_no", "?")
    first_wt = float(entry.get("first_wt", 0) or 0)
    second_wt = float(entry.get("second_wt", 0) or 0)
    gross_wt = float(entry.get("gross_wt", max(first_wt, second_wt)) or 0)
    tare_wt = float(entry.get("tare_wt", min(first_wt, second_wt)) or 0)
    net_wt = float(entry.get("net_wt", 0) or 0)
    bags = int(entry.get("tot_pkts", 0) or 0)
    avg_wt = round(net_wt / bags, 2) if (net_wt and bags > 0) else 0
    cash = float(entry.get("cash_paid", 0) or 0)
    diesel = float(entry.get("diesel_paid", 0) or 0)
    g_issued = float(entry.get("g_issued", 0) or 0)

    def fmtTime(ts):
        if not ts: return ""
        try:
            from datetime import datetime as dt2, timedelta, timezone
            t = dt2.fromisoformat(ts.replace("Z", "+00:00"))
            ist = timezone(timedelta(hours=5, minutes=30))
            return t.astimezone(ist).strftime("%I:%M %p")
        except: return ts[:8] if len(ts) > 8 else ts

    # Load images
    img_objs = {}
    for key in ["first_wt_front_img", "first_wt_side_img", "second_wt_front_img", "second_wt_side_img"]:
        b64 = _load_image_b64(entry.get(key, ""))
        if b64:
            try:
                img_objs[key] = io.BytesIO(b64mod.b64decode(b64))
            except: pass

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=10*mm, bottomMargin=10*mm)

    elements = []

    # ── Branding Header (from settings) ──
    header_els = await get_pdf_header_elements_from_db(f"WEIGHT REPORT — RST #{rst}")
    elements.extend(header_els)

    # Styles — use NotoDeva (Hindi-capable, also renders Latin) for labels that mix English + Devanagari
    lbl_style = ParagraphStyle('lbl', fontName='NotoDeva', fontSize=8, textColor=colors.HexColor('#555'))
    val_style = ParagraphStyle('val', fontName='NotoDevaBold', fontSize=9, textColor=colors.HexColor('#000'))
    small_style = ParagraphStyle('sm', fontName='NotoDeva', fontSize=7.5, textColor=colors.HexColor('#666'))

    # ── Info Table ──
    _is_sale_slip = ("sale" in str(entry.get("trans_type", "")).lower() or "dispatch" in str(entry.get("trans_type", "")).lower())
    info_rows = [
        [Paragraph("RST No.", lbl_style), Paragraph(f"#{rst}", val_style),
         Paragraph("Date / दिनांक", lbl_style), Paragraph(fmt_date(entry.get("date", "")), val_style)],
        [Paragraph("Vehicle / गाड़ी", lbl_style), Paragraph(entry.get("vehicle_no", "-"), val_style),
         Paragraph("Trans Type", lbl_style), Paragraph(entry.get("trans_type", "-"), val_style)],
        [Paragraph("Party / पार्टी", lbl_style), Paragraph(entry.get("party_name", "-"), val_style),
         Paragraph("Destination" if _is_sale_slip else "Source / Mandi", lbl_style), Paragraph(entry.get("farmer_name", "") or "-", val_style)],
        [Paragraph("Product / माल", lbl_style), Paragraph(entry.get("product", "-"), val_style),
         Paragraph("Bags / बोरे", lbl_style), Paragraph(str(bags) if bags else "-", val_style)],
    ]
    # Conditional rows — only if value exists
    cond_row = []
    if g_issued: cond_row.extend([Paragraph("G.Issued", lbl_style), Paragraph(str(int(g_issued)), val_style)])
    else: cond_row.extend([Paragraph("", lbl_style), Paragraph("", val_style)])
    tp_no = entry.get("tp_no", "") or ""
    if tp_no: cond_row.extend([Paragraph("TP No.", lbl_style), Paragraph(tp_no, val_style)])
    else: cond_row.extend([Paragraph("", lbl_style), Paragraph("", val_style)])
    if any(c.text for c in cond_row): info_rows.append(cond_row)

    cond_row2 = []
    tp_wt = entry.get("tp_weight", "") or entry.get("tp_wt", "") or ""
    remark = entry.get("remark", "") or ""
    if tp_wt: cond_row2.extend([Paragraph("TP Weight", lbl_style), Paragraph(f"{tp_wt} Q", val_style)])
    else: cond_row2.extend([Paragraph("", lbl_style), Paragraph("", val_style)])
    if remark: cond_row2.extend([Paragraph("Remark", lbl_style), Paragraph(remark, val_style)])
    else: cond_row2.extend([Paragraph("", lbl_style), Paragraph("", val_style)])
    if any(c.text for c in cond_row2): info_rows.append(cond_row2)

    cw = (doc.width) / 4
    info_table = Table(info_rows, colWidths=[cw*0.8, cw*1.2, cw*0.8, cw*1.2])
    info_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#999')),
        ('FONTNAME', (0, 0), (-1, -1), 'FreeSans'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#f0f8ff'), colors.white]),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 4*mm))

    # ── Weight Sections ──
    wt_header_style = ParagraphStyle('wth', fontName='NotoDevaBold', fontSize=9, textColor=colors.white)

    def add_weight_section(label, wt_val, time_str, front_key, side_key, bg_color):
        # Header bar as table
        wt_bar = Table([[Paragraph(label, wt_header_style),
                         Paragraph(f"{wt_val:,.0f} KG", ParagraphStyle('wtv', fontName='FreeSansBold', fontSize=11, textColor=colors.white, alignment=TA_CENTER)),
                         Paragraph(f"Time: {fmtTime(time_str)}", ParagraphStyle('wtt', fontName='FreeSans', fontSize=8, textColor=colors.HexColor('#ddd'), alignment=TA_RIGHT))]],
                       colWidths=[doc.width*0.35, doc.width*0.30, doc.width*0.35])
        wt_bar.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), bg_color),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(wt_bar)

        # Photos
        has_front = front_key in img_objs
        has_side = side_key in img_objs
        if has_front or has_side:
            photo_w = doc.width / 2 - 2*mm
            photo_h = 32*mm
            row = []
            if has_front:
                img_objs[front_key].seek(0)
                row.append(RLImage(img_objs[front_key], width=photo_w, height=photo_h, kind='proportional'))
            else:
                row.append("")
            if has_side:
                img_objs[side_key].seek(0)
                row.append(RLImage(img_objs[side_key], width=photo_w, height=photo_h, kind='proportional'))
            else:
                row.append("")
            photo_table = Table([row], colWidths=[doc.width/2, doc.width/2])
            photo_table.setStyle(TableStyle([('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2), ('ALIGN', (0,0), (-1,-1), 'CENTER')]))
            elements.append(photo_table)
        elements.append(Spacer(1, 2*mm))

    add_weight_section("1st Weight / पहला वजन", first_wt, entry.get("first_wt_time", ""),
                       "first_wt_front_img", "first_wt_side_img", colors.HexColor("#1a5276"))

    if second_wt > 0:
        add_weight_section("2nd Weight / दूसरा वजन", second_wt, entry.get("second_wt_time", ""),
                           "second_wt_front_img", "second_wt_side_img", colors.HexColor("#34495e"))

    elements.append(Spacer(1, 3*mm))

    # ── Summary Boxes ──
    summary_items = [
        ("GROSS\nकुल", f"{gross_wt:,.0f} KG", "#dce6f0", "#1a5276"),
        ("TARE\nखाली", f"{tare_wt:,.0f} KG", "#f0e8dc", "#6d4c1d"),
        ("NET\nशुद्ध", f"{net_wt:,.0f} KG", "#d5f5d5", "#1b7a30"),
        ("AVG/BAG\nप्रति बोरा", f"{avg_wt:,.2f} KG", "#e3f2fd", "#1565c0"),
    ]
    # Add Cash/Diesel only if present
    if cash > 0:
        summary_items.append(("CASH\nनकद", f"₹{cash:,.0f}", "#fff8e1", "#e65100"))
    if diesel > 0:
        summary_items.append(("DIESEL\nडीजल", f"₹{diesel:,.0f}", "#fce4d6", "#bf360c"))

    box_data_labels = []
    box_data_values = []
    box_styles = []
    for i, (lbl, val, bg, fg) in enumerate(summary_items):
        box_data_labels.append(Paragraph(lbl, ParagraphStyle(f'bl{i}', fontName='NotoDeva', fontSize=6, textColor=colors.HexColor(fg), alignment=TA_CENTER)))
        box_data_values.append(Paragraph(val, ParagraphStyle(f'bv{i}', fontName='NotoDevaBold', fontSize=10, textColor=colors.HexColor(fg), alignment=TA_CENTER)))
        box_styles.append(('BACKGROUND', (i, 0), (i, 1), colors.HexColor(bg)))

    ncols = len(summary_items)
    bw = doc.width / ncols
    summary_table = Table([box_data_labels, box_data_values], colWidths=[bw]*ncols)
    summary_table.setStyle(TableStyle([
        *box_styles,
        ('BOX', (0, 0), (-1, -1), 0.3, colors.HexColor('#ccc')),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#ccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(summary_table)

    # Build
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=WeightReport_RST{rst}.pdf"})




# ── Bulk Export: Excel & PDF ──

async def _build_vw_query(kms_year="", status="", date_from="", date_to="",
                           vehicle_no="", party_name="", farmer_name="", rst_no=""):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if status: query["status"] = status
    if date_from or date_to:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to: dq["$lte"] = date_to
        query["date"] = dq
    if vehicle_no: query["vehicle_no"] = {"$regex": vehicle_no, "$options": "i"}
    if party_name: query["party_name"] = {"$regex": party_name, "$options": "i"}
    if farmer_name: query["farmer_name"] = {"$regex": farmer_name, "$options": "i"}
    if rst_no:
        try: query["rst_no"] = int(rst_no)
        except: pass
    return query


@router.get("/vehicle-weight/export/excel")
async def export_vw_excel(kms_year: str = "", status: str = "completed",
                          date_from: str = "", date_to: str = "",
                          vehicle_no: str = "", party_name: str = "",
                          farmer_name: str = "", rst_no: str = "",
                          trans_type: str = ""):
    """Export vehicle weight entries to Excel.
    trans_type='sale' → sale-specific columns (RST, Date, Vehicle, Party, Destination,
                       Product, Bags, Bag Type, Net Wt, Cash, Diesel, Remark).
    trans_type='purchase' → purchase columns (1st/2nd Wt, G.Issued, TP No., TP Wt etc.).
    Empty → mixed/all columns (legacy behavior).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse
    from utils.branding_helper import get_branding_data
    query = await _build_vw_query(kms_year, status, date_from, date_to, vehicle_no, party_name, farmer_name, rst_no)
    # Apply trans_type filter same as list endpoint.
    if trans_type:
        tt = trans_type.lower().strip()
        if tt == "sale":
            query["trans_type"] = {"$regex": "sale|dispatch", "$options": "i"}
        elif tt == "purchase":
            query["trans_type"] = {"$regex": "purchase|receive", "$options": "i"}
    items = await db["vehicle_weights"].find(query, {"_id": 0}).to_list(10000)
    items.sort(key=lambda e: (e.get("date", ""), int(e.get("rst_no") or 0)))

    branding = await get_branding_data()
    company_name = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "")
    custom_fields = branding.get("custom_fields", [])

    is_sale = (trans_type or "").lower().strip() == "sale"

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Weight - Sale" if is_sale else "Vehicle Weight"
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cur_row = 1
    # Above fields (placement=above)
    above_parts = []
    for f in custom_fields:
        if f.get("placement", "below") == "above":
            lbl, val = f.get("label", ""), f.get("value", "")
            if val:
                above_parts.append(f"{lbl}: {val}" if lbl else val)
    title_label = "Vehicle Weight - Sale / बिक्री" if is_sale else "Vehicle Weight / तौल पर्ची"
    n_cols = 11 if is_sale else 15
    if above_parts:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        cell = ws.cell(row=cur_row, column=1, value="  |  ".join(above_parts))
        cell.font = Font(bold=True, size=10, color="8B0000")
        cell.alignment = Alignment(horizontal='center')
        cur_row += 1

    # Company Title
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    title_cell = ws.cell(row=cur_row, column=1, value=f"{company_name} - {title_label}")
    title_cell.font = Font(bold=True, size=14, color="1a1a2e")
    title_cell.alignment = Alignment(horizontal='center')
    cur_row += 1

    # Tagline + Below fields
    below_parts = []
    for f in custom_fields:
        if f.get("placement", "below") != "above":
            lbl, val = f.get("label", ""), f.get("value", "")
            if val:
                below_parts.append(f"{lbl}: {val}" if lbl else val)
    sub_text = tagline
    if below_parts:
        sub_text += "  |  " + "  |  ".join(below_parts) if sub_text else "  |  ".join(below_parts)
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    ws.cell(row=cur_row, column=1, value=sub_text).font = Font(size=9, color="666666")
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal='center')
    cur_row += 1

    # Date/count row
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    sub = f"Date: {date_from or 'All'} to {date_to or 'All'} | Total: {len(items)} entries"
    ws.cell(row=cur_row, column=1, value=sub).font = Font(size=9, color="666666")
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal='center')
    cur_row += 1

    # Header row
    hdr_row = cur_row + 1

    if is_sale:
        headers = ["RST", "Date", "Vehicle", "Party", "Destination", "Product",
                   "Bags", "Bag Type", "Net Wt (KG)", "Bhada", "Remark"]
        n_cols = 11
    else:
        headers = ["RST", "Date", "Vehicle", "Party", "Source/Mandi", "Product", "Trans Type", "Bags",
                   "1st Wt (KG)", "2nd Wt (KG)", "Net Wt (KG)", "TP Wt (Q)", "G.Issued", "Cash", "Diesel"]
    hdr_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, e in enumerate(items, hdr_row + 1):
        if is_sale:
            vals = [e.get("rst_no",""), fmt_date(e.get("date","")), e.get("vehicle_no",""), e.get("party_name",""),
                    e.get("farmer_name",""), e.get("product",""), e.get("tot_pkts",""), e.get("bag_type",""),
                    e.get("net_wt",0), e.get("bhada",0), e.get("remark","")]
            num_start_col = 9
        else:
            vals = [e.get("rst_no",""), fmt_date(e.get("date","")), e.get("vehicle_no",""), e.get("party_name",""),
                    e.get("farmer_name",""), e.get("product",""), e.get("trans_type",""), e.get("tot_pkts",""),
                    e.get("first_wt",0), e.get("second_wt",0), e.get("net_wt",0),
                    float(e.get("tp_weight",0) or 0),
                    e.get("g_issued",0), e.get("cash_paid",0), e.get("diesel_paid",0)]
            num_start_col = 9
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = border
            # Right-align numeric columns. Sale: cols 7 (Bags), 9-10 (Net Wt, Bhada).
            if is_sale:
                if c in (7, 9, 10): cell.alignment = Alignment(horizontal='right')
            else:
                if c >= num_start_col: cell.alignment = Alignment(horizontal='right')

    # Totals row
    tot_row = hdr_row + 1 + len(items)
    tot_bags = sum(int(e.get("tot_pkts", 0) or 0) for e in items)
    tot_net = sum(float(e.get("net_wt", 0) or 0) for e in items)
    tot_cash = sum(float(e.get("cash_paid", 0) or 0) for e in items)
    tot_diesel = sum(float(e.get("diesel_paid", 0) or 0) for e in items)
    tot_bhada = sum(float(e.get("bhada", 0) or 0) for e in items)
    tot_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    tot_font = Font(bold=True, color="FFFFFF", size=10)
    if is_sale:
        tot_vals = ["", "", "", "", "", "TOTAL:", tot_bags, "", tot_net, tot_bhada, ""]
        right_from = 7
    else:
        tot_1st = sum(float(e.get("first_wt", 0) or 0) for e in items)
        tot_2nd = sum(float(e.get("second_wt", 0) or 0) for e in items)
        tot_tp = sum(float(e.get("tp_weight", 0) or 0) for e in items)
        tot_giss = sum(float(e.get("g_issued", 0) or 0) for e in items)
        tot_vals = ["", "", "", "", "", "", "TOTAL:", tot_bags, tot_1st, tot_2nd, tot_net, tot_tp, tot_giss, tot_cash, tot_diesel]
        right_from = 7
    for c, v in enumerate(tot_vals, 1):
        cell = ws.cell(row=tot_row, column=c, value=v)
        cell.font = tot_font
        cell.fill = tot_fill
        cell.border = border
        if c >= right_from: cell.alignment = Alignment(horizontal='right')

    # ===== Beautiful single-line teal summary banner =====
    if items:
        from utils.export_helpers import add_excel_summary_banner, fmt_inr
        if is_sale:
            sum_stats = [
                {'label': 'Total Entries', 'value': str(len(items))},
                {'label': 'Total Bags', 'value': f"{tot_bags:,}"},
                {'label': 'Net Wt', 'value': f"{tot_net:,.0f} KG"},
                {'label': 'Total Bhada', 'value': fmt_inr(tot_bhada)},
            ]
        else:
            sum_stats = [
                {'label': 'Total Entries', 'value': str(len(items))},
                {'label': 'Total Bags', 'value': f"{tot_bags:,}"},
                {'label': '1st Wt', 'value': f"{tot_1st:,.0f} KG"},
                {'label': '2nd Wt', 'value': f"{tot_2nd:,.0f} KG"},
                {'label': 'Net Wt', 'value': f"{tot_net:,.0f} KG"},
                {'label': 'Cash Paid', 'value': fmt_inr(tot_cash)},
                {'label': 'Diesel', 'value': fmt_inr(tot_diesel)},
            ]
        add_excel_summary_banner(ws, tot_row + 2, n_cols, sum_stats)

    # Auto width
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        cells = [c for c in col if not isinstance(c, MergedCell)]
        if not cells: continue
        max_len = max((len(str(c.value or "")) for c in cells), default=8)
        ws.column_dimensions[cells[0].column_letter].width = min(max_len + 3, 25)

    # 🎯 v104.44.9 — Apply consolidated polish (vehicle weight register)
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = "sales" if is_sale else "vehicle_weight"
    fname = f"{suffix}_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/vehicle-weight/export/pdf")
async def export_vw_pdf(kms_year: str = "", status: str = "completed",
                        date_from: str = "", date_to: str = "",
                        vehicle_no: str = "", party_name: str = "",
                        farmer_name: str = "", rst_no: str = "",
                        trans_type: str = ""):
    """Export vehicle weight entries to PDF (A4 Landscape).
    trans_type='sale' → sale-specific layout; 'purchase' → purchase layout; '' → mixed.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from fastapi.responses import StreamingResponse
    from utils.branding_helper import get_pdf_header_elements_from_db

    query = await _build_vw_query(kms_year, status, date_from, date_to, vehicle_no, party_name, farmer_name, rst_no)
    if trans_type:
        tt = trans_type.lower().strip()
        if tt == "sale":
            query["trans_type"] = {"$regex": "sale|dispatch", "$options": "i"}
        elif tt == "purchase":
            query["trans_type"] = {"$regex": "purchase|receive", "$options": "i"}
    items = await db["vehicle_weights"].find(query, {"_id": 0}).to_list(10000)
    items.sort(key=lambda e: (e.get("date", "")[:10], int(e.get("rst_no") or 0)))

    is_sale = (trans_type or "").lower().strip() == "sale"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []

    # Branding header with custom fields
    sub = f"Date: {date_from or 'All'} to {date_to or 'All'} | Total: {len(items)} entries"
    title = "Vehicle Weight - Sale / बिक्री" if is_sale else "Vehicle Weight / तौल पर्ची"
    elements.extend(await get_pdf_header_elements_from_db(title, sub))
    elements.append(Spacer(1, 3*mm))

    # Table
    if is_sale:
        headers = ["RST", "Date", "Vehicle", "Party", "Destination", "Product", "Bags", "Bag Type", "Net Wt", "Bhada", "Remark"]
    else:
        headers = ["RST", "Date", "Vehicle", "Party", "Source/Mandi", "Product", "Trans Type", "Bags", "1st Wt", "2nd Wt", "Net Wt", "TP Wt", "G.Iss", "Cash", "Diesel"]
    data = [headers]
    for e in items:
        if is_sale:
            data.append([
                e.get("rst_no",""), fmt_date(e.get("date","")), e.get("vehicle_no",""),
                e.get("party_name",""), e.get("farmer_name",""), e.get("product",""),
                e.get("tot_pkts",""), e.get("bag_type","") or "-",
                f"{e.get('net_wt',0):,.0f}",
                f"{e.get('bhada',0):,.0f}" if e.get('bhada') else "-",
                (e.get('remark','') or "")[:30],
            ])
        else:
            data.append([
                e.get("rst_no",""), fmt_date(e.get("date","")), e.get("vehicle_no",""),
                e.get("party_name",""), e.get("farmer_name",""), e.get("product",""),
                e.get("trans_type",""), e.get("tot_pkts",""),
                f"{e.get('first_wt',0):,.0f}", f"{e.get('second_wt',0):,.0f}", f"{e.get('net_wt',0):,.0f}",
                f"{float(e.get('tp_weight',0) or 0)}" if float(e.get('tp_weight',0) or 0) > 0 else "-",
                f"{e.get('g_issued',0):,.0f}" if e.get('g_issued') else "-",
                f"{e.get('cash_paid',0):,.0f}" if e.get('cash_paid') else "-",
                f"{e.get('diesel_paid',0):,.0f}" if e.get('diesel_paid') else "-"
            ])

    # Add totals row
    tot_bags = sum(int(e.get("tot_pkts", 0) or 0) for e in items)
    tot_net = sum(float(e.get("net_wt", 0) or 0) for e in items)
    tot_cash = sum(float(e.get("cash_paid", 0) or 0) for e in items)
    tot_diesel = sum(float(e.get("diesel_paid", 0) or 0) for e in items)
    tot_bhada = sum(float(e.get("bhada", 0) or 0) for e in items)
    if is_sale:
        data.append(["", "", "", "", "", "TOTAL:", str(tot_bags), "",
                     f"{tot_net:,.0f}",
                     f"{tot_bhada:,.0f}" if tot_bhada else "-", ""])
        col_widths = [35, 58, 65, 75, 70, 60, 38, 50, 55, 60, 110]
    else:
        tot_1st = sum(float(e.get("first_wt", 0) or 0) for e in items)
        tot_2nd = sum(float(e.get("second_wt", 0) or 0) for e in items)
        tot_tp = sum(float(e.get("tp_weight", 0) or 0) for e in items)
        tot_giss = sum(float(e.get("g_issued", 0) or 0) for e in items)
        data.append(["", "", "", "", "", "", "TOTAL:", str(tot_bags),
                     f"{tot_1st:,.0f}", f"{tot_2nd:,.0f}", f"{tot_net:,.0f}",
                     f"{tot_tp}" if tot_tp > 0 else "-", f"{tot_giss:,.0f}",
                     f"{tot_cash:,.0f}" if tot_cash else "-", f"{tot_diesel:,.0f}" if tot_diesel else "-"])
        col_widths = [35, 58, 65, 70, 65, 55, 52, 30, 50, 50, 50, 38, 38, 42, 42]

    t = Table(data, colWidths=col_widths, repeatRows=1)

    # Color-coded column header: Navy(info), Teal(weights), Orange(money)
    navy = colors.HexColor('#1a237e')
    teal = colors.HexColor('#004d40')
    amber = colors.HexColor('#e65100')

    if is_sale:
        # Sale layout: cols 0-7 navy(info+bags+bag_type), col 8 teal(net wt), col 9 amber(bhada), col 10 navy(remark)
        style_cmds = [
            ('BACKGROUND', (0, 0), (7, 0), navy),
            ('BACKGROUND', (8, 0), (8, 0), teal),
            ('BACKGROUND', (9, 0), (9, 0), amber),
            ('BACKGROUND', (10, 0), (10, 0), navy),
        ]
        right_align_from = 6
        net_wt_col = 8
        cash_col = 9   # bhada in sale view (re-using variable name for color logic)
        diesel_col = 9
    else:
        style_cmds = [
            ('BACKGROUND', (0, 0), (7, 0), navy),
            ('BACKGROUND', (8, 0), (11, 0), teal),
            ('BACKGROUND', (12, 0), (14, 0), amber),
        ]
        right_align_from = 7
        net_wt_col = 10
        cash_col = 13
        diesel_col = 14

    style_cmds.extend([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (right_align_from, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, 0), 0.5, colors.HexColor('#ffffff')),
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#f9a825')),
        ('INNERGRID', (0, 1), (-1, -1), 0.3, colors.HexColor('#d0d5dd')),
        ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#90a4ae')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7ff')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ])

    # Color-code specific data cells
    for row_idx in range(1, len(data)):
        # RST bold navy
        style_cmds.append(('TEXTCOLOR', (0, row_idx), (0, row_idx), navy))
        style_cmds.append(('FONTNAME', (0, row_idx), (0, row_idx), 'Helvetica-Bold'))
        # Net Wt green bold
        style_cmds.append(('TEXTCOLOR', (net_wt_col, row_idx), (net_wt_col, row_idx), colors.HexColor('#1b5e20')))
        style_cmds.append(('FONTNAME', (net_wt_col, row_idx), (net_wt_col, row_idx), 'Helvetica-Bold'))
        # Cash green bold
        if data[row_idx][cash_col] != "-":
            style_cmds.append(('TEXTCOLOR', (cash_col, row_idx), (cash_col, row_idx), colors.HexColor('#2e7d32')))
            style_cmds.append(('FONTNAME', (cash_col, row_idx), (cash_col, row_idx), 'Helvetica-Bold'))
        # Diesel orange bold
        if data[row_idx][diesel_col] != "-":
            style_cmds.append(('TEXTCOLOR', (diesel_col, row_idx), (diesel_col, row_idx), colors.HexColor('#e65100')))
            style_cmds.append(('FONTNAME', (diesel_col, row_idx), (diesel_col, row_idx), 'Helvetica-Bold'))
        if not is_sale:
            # 1st Wt blue, 2nd Wt purple (only purchase view)
            style_cmds.append(('TEXTCOLOR', (8, row_idx), (8, row_idx), colors.HexColor('#0277bd')))
            style_cmds.append(('TEXTCOLOR', (9, row_idx), (9, row_idx), colors.HexColor('#7b1fa2')))

    # Totals row styling
    last_row = len(data) - 1
    style_cmds.append(('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor('#1a1a2e')))
    style_cmds.append(('TEXTCOLOR', (0, last_row), (-1, last_row), colors.white))
    style_cmds.append(('FONTNAME', (0, last_row), (-1, last_row), 'Helvetica-Bold'))
    style_cmds.append(('FONTSIZE', (0, last_row), (-1, last_row), 8))

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    # ===== Beautiful single-line summary banner =====
    from utils.export_helpers import get_pdf_summary_banner, fmt_inr, STAT_COLORS
    page_inner_w = sum(col_widths)
    if is_sale:
        summary_stats = [
            {'label': 'TOTAL ENTRIES', 'value': str(len(items)), 'color': STAT_COLORS['primary']},
            {'label': 'TOTAL BAGS', 'value': f"{tot_bags:,}", 'color': STAT_COLORS['blue']},
            {'label': 'NET WT', 'value': f"{tot_net:,.0f}", 'color': STAT_COLORS['emerald']},
            {'label': 'TOTAL BHADA', 'value': fmt_inr(tot_bhada), 'color': STAT_COLORS['orange']},
        ]
    else:
        summary_stats = [
            {'label': 'TOTAL ENTRIES', 'value': str(len(items)), 'color': STAT_COLORS['primary']},
            {'label': 'TOTAL BAGS', 'value': f"{tot_bags:,}", 'color': STAT_COLORS['blue']},
            {'label': '1ST WT', 'value': f"{tot_1st:,.0f}", 'color': STAT_COLORS['teal']},
            {'label': '2ND WT', 'value': f"{tot_2nd:,.0f}", 'color': STAT_COLORS['purple']},
            {'label': 'NET WT', 'value': f"{tot_net:,.0f}", 'color': STAT_COLORS['emerald']},
            {'label': 'CASH PAID', 'value': fmt_inr(tot_cash), 'color': STAT_COLORS['green']},
            {'label': 'DIESEL', 'value': fmt_inr(tot_diesel), 'color': STAT_COLORS['orange']},
        ]
    elements.append(Spacer(1, 4*mm))
    banner = get_pdf_summary_banner(summary_stats, total_width=page_inner_w)
    if banner:
        elements.append(banner)

    doc.build(elements)
    buf.seek(0)
    fname = f"vehicle_weight_{date_from or 'all'}_{date_to or 'all'}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})

# ── Image Cleanup Settings & Scheduler ──

@router.get("/settings/mandi-cutting-map")
async def get_mandi_cutting_map():
    """Get mandi cutting map from database."""
    doc = await db["settings"].find_one({"key": "mandi_cutting_map"}, {"_id": 0})
    return doc.get("value", {}) if doc else {}


@router.put("/settings/mandi-cutting-map")
async def save_mandi_cutting_map(data: dict):
    """Save/update a mandi cutting entry."""
    key = data.get("key", "")
    value = data.get("value", 0)
    if not key:
        raise HTTPException(status_code=400, detail="Key missing")
    doc = await db["settings"].find_one({"key": "mandi_cutting_map"})
    current = doc.get("value", {}) if doc else {}
    current[key] = value
    await db["settings"].update_one(
        {"key": "mandi_cutting_map"},
        {"$set": {"key": "mandi_cutting_map", "value": current}},
        upsert=True
    )
    return {"success": True}


@router.get("/settings/camera-config")
async def get_camera_config():
    """Get camera config from database."""
    doc = await db["settings"].find_one({"key": "camera_config"}, {"_id": 0})
    return doc.get("value", {}) if doc else {}


@router.put("/settings/camera-config")
async def save_camera_config(data: dict):
    """Save camera config to database."""
    await db["settings"].update_one(
        {"key": "camera_config"},
        {"$set": {"key": "camera_config", "value": data}},
        upsert=True
    )
    return {"success": True}


@router.get("/settings/image-cleanup")
async def get_image_cleanup_setting():
    """Get image auto-cleanup days setting."""
    doc = await db["settings"].find_one({"key": "image_cleanup"}, {"_id": 0})
    if doc:
        return {"days": doc.get("days", 0), "enabled": doc.get("days", 0) > 0}
    return {"days": 0, "enabled": False}


@router.put("/settings/image-cleanup")
async def update_image_cleanup_setting(data: dict):
    """Set image auto-cleanup days (0 = disabled)."""
    days = int(data.get("days", 0) or 0)
    if days < 0:
        days = 0
    await db["settings"].update_one(
        {"key": "image_cleanup"},
        {"$set": {"key": "image_cleanup", "days": days}},
        upsert=True
    )
    return {"success": True, "days": days, "enabled": days > 0}


@router.post("/settings/image-cleanup/run")
async def run_image_cleanup_now():
    """Manually trigger image cleanup based on configured days."""
    doc = await db["settings"].find_one({"key": "image_cleanup"}, {"_id": 0})
    days = doc.get("days", 0) if doc else 0
    if days <= 0:
        return {"success": False, "message": "Cleanup disabled (days = 0)", "deleted": 0}
    deleted = _cleanup_old_images(days)
    return {"success": True, "message": f"{deleted} purani images delete hui", "deleted": deleted}


def _cleanup_old_images(days: int) -> int:
    """Delete image files older than N days from IMG_DIR."""
    import time
    if days <= 0:
        return 0
    cutoff = time.time() - (days * 86400)
    deleted = 0
    try:
        for fname in os.listdir(IMG_DIR):
            fpath = os.path.join(IMG_DIR, fname)
            if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
                os.remove(fpath)
                deleted += 1
        if deleted > 0:
            logger.info(f"Image cleanup: {deleted} files older than {days} days deleted")
    except Exception as e:
        logger.error(f"Image cleanup error: {e}")
    return deleted


async def image_cleanup_scheduler():
    """Background task: run cleanup once every 24 hours."""
    import asyncio
    while True:
        await asyncio.sleep(86400)  # 24 hours
        try:
            doc = await db["settings"].find_one({"key": "image_cleanup"}, {"_id": 0})
            days = doc.get("days", 0) if doc else 0
            if days > 0:
                _cleanup_old_images(days)
        except Exception as e:
            logger.error(f"Image cleanup scheduler error: {e}")


# Storage Engine (Web version always uses MongoDB - this is for frontend compatibility)
@router.get("/settings/storage-engine")
async def get_storage_engine():
    return {"engine": "mongodb"}


@router.get("/settings/weighbridge-host")
async def get_weighbridge_host():
    setting = await db["app_settings"].find_one({"setting_id": "weighbridge_host"}, {"_id": 0})
    return {"url": setting.get("value", "") if setting else ""}

@router.put("/settings/weighbridge-host")
async def set_weighbridge_host(data: dict):
    url = (data.get("url") or "").strip()
    await db["app_settings"].update_one(
        {"setting_id": "weighbridge_host"},
        {"$set": {"setting_id": "weighbridge_host", "value": url, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"success": True, "url": url}

@router.get("/weighbridge/live-weight")
async def get_live_weight():
    """Web version has no serial port - return disconnected status"""
    return {"connected": False, "weight": 0, "stable": False, "timestamp": 0}


# ════════════════════════════════════════════════════════════════════════════
# 🛻 TRUCK OWNER — Per-Trip Breakdown (Bhada-based, FIFO settlement)
# ════════════════════════════════════════════════════════════════════════════
@router.get("/truck-owner/per-trip-trucks")
async def list_trucks_with_bhada(kms_year: str = "", season: str = ""):
    """List all distinct truck numbers that have at least 1 VW entry with bhada > 0."""
    q = {"bhada": {"$gt": 0}, "vehicle_no": {"$ne": ""}}
    if kms_year:
        q["kms_year"] = kms_year
    if season:
        q["season"] = season
    cursor = db.vehicle_weights.find(q, {"_id": 0, "vehicle_no": 1, "bhada": 1, "farmer_name": 1})
    agg = {}
    async for vw in cursor:
        v = (vw.get("vehicle_no") or "").strip()
        if not v:
            continue
        a = agg.setdefault(v, {"vehicle_no": v, "trips_count": 0, "total_bhada": 0.0})
        a["trips_count"] += 1
        a["total_bhada"] += float(vw.get("bhada", 0) or 0)
    out = sorted(agg.values(), key=lambda x: x["vehicle_no"])
    return {"trucks": out}


@router.get("/truck-owner/per-trip-all")
async def per_trip_all(kms_year: str = "", season: str = "", date_from: str = "", date_to: str = "", filter_status: str = ""):
    """Aggregate per-trip view across ALL trucks. Used as the default view in Payments tab.

    Returns: { trips: [...with vehicle_no...], summary: aggregated KPIs }
    """
    q = {"bhada": {"$gt": 0}, "vehicle_no": {"$ne": ""}}
    if kms_year:
        q["kms_year"] = kms_year
    if season:
        q["season"] = season
    distinct_vnos = await db.vehicle_weights.distinct("vehicle_no", q)

    all_trips = []
    agg = {
        "total_trips": 0, "sale_count": 0, "purchase_count": 0,
        "total_bhada": 0.0, "total_paid": 0.0, "total_pending": 0.0,
        "settled_count": 0, "partial_count": 0, "pending_count": 0,
        "extra_paid_unallocated": 0.0,
    }
    for vno in distinct_vnos:
        try:
            data = await truck_owner_per_trip(vno, kms_year, season, date_from, date_to)
            for t in data["trips"]:
                t_with_v = {**t, "vehicle_no": vno}
                all_trips.append(t_with_v)
            sm = data["summary"]
            agg["total_trips"]    += sm.get("total_trips", 0)
            agg["sale_count"]     += sm.get("sale_count", 0)
            agg["purchase_count"] += sm.get("purchase_count", 0)
            agg["total_bhada"]    += sm.get("total_bhada", 0)
            agg["total_paid"]     += sm.get("total_paid", 0)
            agg["total_pending"]  += sm.get("total_pending", 0)
            agg["settled_count"]  += sm.get("settled_count", 0)
            agg["partial_count"]  += sm.get("partial_count", 0)
            agg["pending_count"]  += sm.get("pending_count", 0)
            agg["extra_paid_unallocated"] += sm.get("extra_paid_unallocated", 0)
        except Exception:
            continue

    if filter_status and filter_status != "all":
        all_trips = [t for t in all_trips if t.get("status") == filter_status]
    # Newest first
    all_trips.sort(key=lambda x: (x.get("date") or "", x.get("rst_no") or 0), reverse=True)

    for k in ("total_bhada", "total_paid", "total_pending", "extra_paid_unallocated"):
        agg[k] = round(agg[k], 2)
    return {"trips": all_trips, "summary": agg, "total_trucks": len(distinct_vnos)}


@router.get("/truck-owner/per-trip-pending-count")
async def per_trip_pending_count(kms_year: str = "", season: str = ""):
    """Total count of trips with pending/partial bhada (across all trucks).
    Used by Payments tab badge — auto-decrements on Pay action."""
    q = {"bhada": {"$gt": 0}, "vehicle_no": {"$ne": ""}}
    if kms_year:
        q["kms_year"] = kms_year
    if season:
        q["season"] = season
    # Group VWs by truck → call per-trip endpoint logic for each → count non-settled
    trucks_with_bhada = await db.vehicle_weights.distinct("vehicle_no", q)
    pending = 0
    for vno in trucks_with_bhada:
        try:
            data = await truck_owner_per_trip(vno, kms_year, season, "", "")
            pending += sum(1 for t in data["trips"] if t.get("status") != "settled")
        except Exception:
            continue
    return {"pending_count": pending}


@router.get("/truck-owner/{vehicle_no}/per-trip")
async def truck_owner_per_trip(vehicle_no: str, kms_year: str = "", season: str = "", date_from: str = "", date_to: str = ""):
    """Per-trip Bhada breakdown for a single truck owner with FIFO settlement.

    Algorithm:
      1. Fetch all VW entries for `vehicle_no` with `bhada > 0` (chronological).
      2. Fetch all `cash_transactions` for this truck owner where:
           account=ledger AND party_type=Truck AND category=vehicle_no AND txn_type=nikasi
         (Both manual payments and any auto-generated NIKASI count as "paid").
      3. FIFO-apply nikasi totals on jama trips → derive per-trip status:
            - settled  : fully covered
            - partial  : partially covered (paidAmt > 0, < bhada)
            - pending  : zero coverage

    Returns: { vehicle_no, trips:[...], summary: {...} }
    """
    vno = (vehicle_no or "").strip()
    if not vno:
        raise HTTPException(status_code=400, detail="vehicle_no required")

    vw_q = {"vehicle_no": vno, "bhada": {"$gt": 0}}
    if kms_year:
        vw_q["kms_year"] = kms_year
    if season:
        vw_q["season"] = season
    if date_from:
        vw_q.setdefault("date", {})["$gte"] = date_from
    if date_to:
        vw_q.setdefault("date", {})["$lte"] = date_to

    vws = await db.vehicle_weights.find(vw_q, {"_id": 0}).to_list(length=2000)
    # Chronological asc for FIFO
    vws.sort(key=lambda x: (x.get("date", ""), x.get("created_at", ""), x.get("rst_no", 0)))

    nikasi_q = {
        "account": "ledger",
        "party_type": "Truck",
        "category": vno,
        "txn_type": "nikasi",
    }
    if kms_year:
        nikasi_q["kms_year"] = kms_year
    nikasis = await db.cash_transactions.find(nikasi_q, {"_id": 0}).to_list(length=5000)
    nikasis.sort(key=lambda x: (x.get("date", ""), x.get("created_at", "")))

    # Step 1: Apply trip-targeted direct settlements first (reference = truck_settle_ledger:{vno}:{rst})
    direct_paid = {}  # rst_no → cumulative paid
    pool_nikasis = []
    for n in nikasis:
        ref = n.get("reference", "") or ""
        if ref.startswith("truck_settle_ledger:"):
            parts = ref.split(":")
            if len(parts) >= 3:
                try:
                    target_rst = int(parts[-1])
                    direct_paid[target_rst] = direct_paid.get(target_rst, 0) + float(n.get("amount", 0) or 0)
                    continue
                except Exception:
                    pass
        # Otherwise add to FIFO pool
        pool_nikasis.append(n)
    pool = sum(float(n.get("amount", 0) or 0) for n in pool_nikasis)
    total_paid_pool = sum(float(n.get("amount", 0) or 0) for n in nikasis)  # Both direct + pool

    trips = []
    for vw in vws:
        bhada = float(vw.get("bhada", 0) or 0)
        rst = vw.get("rst_no")
        trans_type_lower = (vw.get("trans_type", "") or "").lower()
        is_sale = ("sale" in trans_type_lower) or ("dispatch" in trans_type_lower)
        is_purchase = ("purchase" in trans_type_lower) or ("receive" in trans_type_lower)
        ttype = "sale" if is_sale else ("purchase" if is_purchase else "other")

        # Direct trip-targeted payment first
        paid = min(direct_paid.get(rst, 0), bhada)
        # Then apply pool FIFO to remaining
        remaining = bhada - paid
        if remaining > 0 and pool > 0:
            take = min(pool, remaining)
            paid += take
            pool -= take

        if paid >= bhada and bhada > 0:
            status = "settled"
        elif paid > 0:
            status = "partial"
        else:
            status = "pending"

        trips.append({
            "rst_no": rst,
            "date": vw.get("date", ""),
            "trans_type": ttype,
            "trans_type_raw": vw.get("trans_type", ""),
            "party_name": vw.get("party_name", ""),
            "farmer_name": vw.get("farmer_name", ""),
            "product": vw.get("product", ""),
            "tot_pkts": vw.get("tot_pkts", 0),
            "net_wt": vw.get("net_wt", 0),
            "bhada": bhada,
            "paid_amount": round(paid, 2),
            "pending_amount": round(bhada - paid, 2),
            "status": status,
            "vw_id": vw.get("id"),
        })
    # Newest trips first for display
    trips.sort(key=lambda x: (x.get("date") or "", x.get("rst_no") or 0), reverse=True)

    total_bhada = sum(t["bhada"] for t in trips)
    total_paid = sum(t["paid_amount"] for t in trips)
    total_pending = round(total_bhada - total_paid, 2)
    settled_count = sum(1 for t in trips if t["status"] == "settled")
    partial_count = sum(1 for t in trips if t["status"] == "partial")
    pending_count = sum(1 for t in trips if t["status"] == "pending")
    sale_count = sum(1 for t in trips if t["trans_type"] == "sale")
    purchase_count = sum(1 for t in trips if t["trans_type"] == "purchase")

    # Driver name from any nikasi (description) or first VW entry
    driver_name = ""
    return {
        "vehicle_no": vno,
        "driver_name": driver_name,
        "trips": trips,
        "summary": {
            "total_trips": len(trips),
            "sale_count": sale_count,
            "purchase_count": purchase_count,
            "total_bhada": round(total_bhada, 2),
            "total_paid": round(total_paid, 2),
            "total_pending": total_pending,
            "settled_count": settled_count,
            "partial_count": partial_count,
            "pending_count": pending_count,
            "extra_paid_unallocated": round(max(0, total_paid_pool - total_bhada), 2),
        },
    }


@router.post("/truck-owner/{vehicle_no}/settle/{rst_no}")
async def truck_owner_settle_trip(vehicle_no: str, rst_no: int, data: dict):
    """One-click settle a specific trip's bhada — creates dual cash_transactions entries
    matching the existing /truck-owner/{truck_no}/pay pattern:
      • Cash/Bank/Owner NIKASI → deducts from selected payment mode
      • Ledger NIKASI         → marks bhada settled (FIFO algo picks this up)

    Body: {
      amount?: number (defaults to pending bhada),
      account?: "cash"|"bank"|"owner",
      bank_name?: string,
      owner_name?: string,
      round_off?: number,
      note?: string,
      date?: string,
      username?: string
    }
    """
    vno = (vehicle_no or "").strip()
    if not vno or rst_no is None:
        raise HTTPException(status_code=400, detail="vehicle_no and rst_no required")

    vw = await db.vehicle_weights.find_one({"vehicle_no": vno, "rst_no": rst_no}, {"_id": 0})
    if not vw:
        raise HTTPException(status_code=404, detail="VW entry not found")
    bhada = float(vw.get("bhada", 0) or 0)
    if bhada <= 0:
        raise HTTPException(status_code=400, detail="No bhada on this trip")

    amount = float(data.get("amount") or bhada)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount > 0 hona chahiye")

    note = data.get("note", "") or ""
    round_off = float(data.get("round_off") or 0)
    pay_account = (data.get("account") or data.get("payment_mode") or "cash").lower()
    if pay_account not in ("cash", "bank", "owner"):
        pay_account = "cash"
    if pay_account == "bank" and not data.get("bank_name"):
        raise HTTPException(status_code=400, detail="Bank name select karein")
    if pay_account == "owner" and not data.get("owner_name"):
        raise HTTPException(status_code=400, detail="Owner account select karein")

    pay_bank_name = data.get("bank_name", "") if pay_account == "bank" else ""
    pay_owner_name = data.get("owner_name", "") if pay_account == "owner" else ""
    date_str = data.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    username = data.get("username") or "system"
    party_label = vw.get('farmer_name') or vw.get('party_name', '')
    desc_base = f"Bhada Settle (RST #{rst_no} → {party_label})"
    if note:
        desc_base = f"{desc_base} - {note}"

    now_iso = datetime.now(timezone.utc).isoformat()
    txn_suffix = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{vno}_{rst_no}"

    # 1. Cash/Bank/Owner NIKASI — deducts payment mode balance
    cash_txn = {
        "id": f"txn_{txn_suffix}",
        "date": date_str,
        "account": pay_account,
        "bank_name": pay_bank_name,
        "owner_name": pay_owner_name,
        "txn_type": "nikasi",
        "category": vno,
        "party_type": "Truck",
        "description": desc_base,
        "amount": amount,
        "reference": f"truck_settle:{vno}:{rst_no}",
        "linked_entry_id": vw.get("id"),
        "kms_year": vw.get("kms_year", ""),
        "season": vw.get("season", "Kharif"),
        "created_by": username,
        "created_at": now_iso,
    }
    await db.cash_transactions.insert_one(cash_txn)

    # 2. Ledger NIKASI — picked up by FIFO settlement algorithm
    owner_total = round(amount + round_off, 2)
    desc_ledger = desc_base + (f" (Pay: {amount}, Round Off: {round_off})" if round_off else "")
    ledger_txn = {
        "id": f"txn_ledger_{txn_suffix}_{uuid.uuid4().hex[:6]}",
        "date": date_str,
        "account": "ledger",
        "txn_type": "nikasi",
        "category": vno,
        "party_type": "Truck",
        "description": desc_ledger,
        "amount": owner_total,
        "reference": f"truck_settle_ledger:{vno}:{rst_no}",
        "linked_entry_id": vw.get("id"),
        "kms_year": vw.get("kms_year", ""),
        "season": vw.get("season", "Kharif"),
        "created_by": username,
        "created_at": now_iso,
    }
    await db.cash_transactions.insert_one(ledger_txn)

    return {"success": True, "settled_amount": amount, "round_off": round_off, "rst_no": rst_no, "payment_mode": pay_account}


@router.get("/truck-owner/{vehicle_no}/trip-history/{rst_no}")
async def truck_owner_trip_history(vehicle_no: str, rst_no: int, kms_year: str = ""):
    """Return all payment history for a specific RST trip (settle entries + truck-level NIKASI)."""
    vno = (vehicle_no or "").strip()
    # Direct trip-specific settle entries
    direct = await db.cash_transactions.find({
        "category": vno,
        "party_type": "Truck",
        "txn_type": "nikasi",
        "$or": [
            {"reference": f"truck_settle:{vno}:{rst_no}"},
            {"reference": f"truck_settle_ledger:{vno}:{rst_no}"},
        ],
    }, {"_id": 0}).sort("date", 1).to_list(length=200)
    return {"vehicle_no": vno, "rst_no": rst_no, "payments": direct}



# ════════════════════════════════════════════════════════════════════════════
# 🛻 TRUCK OWNER — Pending PDF / Excel / WhatsApp Export
# ════════════════════════════════════════════════════════════════════════════
async def _build_pertrip_payload(vehicle_no: str, kms_year: str = "", season: str = "", date_from: str = "", date_to: str = "", filter_status: str = ""):
    """Reuse the per-trip endpoint logic to build a payload for export."""
    full = await truck_owner_per_trip(vehicle_no, kms_year, season, date_from, date_to)
    if filter_status and filter_status != "all":
        full["trips"] = [t for t in full["trips"] if t.get("status") == filter_status]
    return full


@router.get("/truck-owner/{vehicle_no}/per-trip-pdf")
async def truck_owner_per_trip_pdf(vehicle_no: str, kms_year: str = "", season: str = "", date_from: str = "", date_to: str = "", filter_status: str = ""):
    """Generate per-trip Bhada PDF — defaults to all trips, filter_status=pending for unpaid only."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.utils import ImageReader  # noqa: F401
    from fastapi.responses import StreamingResponse
    from utils.branding_helper import get_pdf_header_elements_from_db
    from utils.export_helpers import register_hindi_fonts
    from utils.date_format import fmt_date as fmt_d
    register_hindi_fonts()

    payload = await _build_pertrip_payload(vehicle_no, kms_year, season, date_from, date_to, filter_status)
    trips = payload["trips"]
    sm = payload["summary"]

    buf = io.BytesIO()
    title_kind = "Pending Bhada" if filter_status == "pending" else "Per-Trip Bhada"
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []
    header_elems = await get_pdf_header_elements_from_db(title=f"🛻 {title_kind} — {vehicle_no}", subtitle=f"KMS: {kms_year or 'All'} · Season: {season or 'All'}")
    elements.extend(header_elems)

    # Summary banner
    summary_data = [[
        f"Trips: {sm['total_trips']}",
        f"Total Bhada: ₹{sm['total_bhada']:,.0f}",
        f"Settled: ₹{sm['total_paid']:,.0f}",
        f"Pending: ₹{sm['total_pending']:,.0f}",
    ]]
    summary_tbl = Table(summary_data, colWidths=[45*mm, 45*mm, 45*mm, 45*mm])
    summary_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('LEADING', (0,0), (-1,-1), 14),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#1a237e')),
    ]))
    elements.append(summary_tbl)
    elements.append(Spacer(1, 4*mm))

    # Trips table
    headers = ["RST", "Date", "Type", "Party", "Net Wt", "Bhada", "Paid", "Pending", "Status"]
    rows = [headers]
    for t in trips:
        rows.append([
            f"#{t['rst_no']}", fmt_d(t['date']),
            "Sale" if t['trans_type'] == 'sale' else ("Purchase" if t['trans_type'] == 'purchase' else (t.get('trans_type_raw') or '-')),
            (t.get('party_name') or t.get('farmer_name') or '-')[:28],
            f"{int(t.get('net_wt') or 0):,}" if t.get('net_wt') else '-',
            f"₹{t['bhada']:,.0f}",
            f"₹{t['paid_amount']:,.0f}" if t['paid_amount'] else '-',
            f"₹{t['pending_amount']:,.0f}" if t['pending_amount'] else '-',
            t['status'].title(),
        ])
    if len(rows) == 1:
        rows.append(["—", "Koi trip nahi", "", "", "", "", "", "", ""])
    trips_tbl = Table(rows, colWidths=[14*mm, 22*mm, 18*mm, 50*mm, 18*mm, 22*mm, 22*mm, 22*mm, 22*mm], repeatRows=1)
    trips_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#004d40')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),    # RST
        ('ALIGN', (4,1), (-2,-1), 'RIGHT'),    # numbers
        ('ALIGN', (-1,0), (-1,-1), 'CENTER'),  # Status
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('LEADING', (0,0), (-1,-1), 11),
    ]))
    # Color status cells based on value
    for i, t in enumerate(trips, start=1):
        st = t.get('status')
        if st == 'settled':
            trips_tbl.setStyle(TableStyle([('BACKGROUND', (-1,i), (-1,i), colors.HexColor('#c8e6c9')), ('TEXTCOLOR', (-1,i), (-1,i), colors.HexColor('#1b5e20'))]))
        elif st == 'partial':
            trips_tbl.setStyle(TableStyle([('BACKGROUND', (-1,i), (-1,i), colors.HexColor('#ffe0b2')), ('TEXTCOLOR', (-1,i), (-1,i), colors.HexColor('#e65100'))]))
        else:
            trips_tbl.setStyle(TableStyle([('BACKGROUND', (-1,i), (-1,i), colors.HexColor('#ffcdd2')), ('TEXTCOLOR', (-1,i), (-1,i), colors.HexColor('#b71c1c'))]))
    elements.append(trips_tbl)

    # Footer
    elements.append(Spacer(1, 4*mm))
    style_small = ParagraphStyle('small', fontSize=8, alignment=TA_CENTER, textColor=colors.grey)
    elements.append(Paragraph(f"Generated: {fmt_d(datetime.now(timezone.utc).strftime('%Y-%m-%d'))} · {len(trips)} trip(s)", style_small))

    doc.build(elements)
    buf.seek(0)
    safe_vno = "".join(c for c in vehicle_no if c.isalnum() or c in "-_").strip("-_") or "truck"
    fname = f"{safe_vno}_{title_kind.replace(' ', '_')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/truck-owner/{vehicle_no}/per-trip-excel")
async def truck_owner_per_trip_excel(vehicle_no: str, kms_year: str = "", season: str = "", date_from: str = "", date_to: str = "", filter_status: str = ""):
    """Excel export of per-trip Bhada."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    payload = await _build_pertrip_payload(vehicle_no, kms_year, season, date_from, date_to, filter_status)
    trips = payload["trips"]
    sm = payload["summary"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Per-Trip Bhada"

    title_kind = "Pending Bhada" if filter_status == "pending" else "Per-Trip Bhada"
    title_font = Font(name="Inter", bold=True, size=14, color="1a1a2e")
    sub_font = Font(name="Inter", size=10, color="555555")
    hdr_font = Font(name="Inter", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill(start_color="004D40", end_color="004D40", fill_type="solid")
    sub_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    thin = Side(style="thin", color="cccccc")
    bd = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value=f"🛻 {title_kind} — {vehicle_no}").font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws.cell(row=2, column=1, value=f"KMS: {kms_year or 'All'} · Season: {season or 'All'} · Trips: {sm['total_trips']} · Total Bhada: ₹{sm['total_bhada']:,.0f} · Settled: ₹{sm['total_paid']:,.0f} · Pending: ₹{sm['total_pending']:,.0f}").font = sub_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["RST", "Date", "Type", "Party", "Net Wt (KG)", "Bhada", "Paid", "Pending", "Status"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = bd
        c.alignment = Alignment(horizontal="center")

    for ri, t in enumerate(trips, start=5):
        vals = [
            f"#{t['rst_no']}", t['date'],
            "Sale" if t['trans_type'] == 'sale' else ("Purchase" if t['trans_type'] == 'purchase' else (t.get('trans_type_raw') or '-')),
            t.get('party_name') or t.get('farmer_name') or '-',
            int(t.get('net_wt') or 0),
            float(t['bhada']),
            float(t['paid_amount']),
            float(t['pending_amount']),
            t['status'].title(),
        ]
        for col_idx, v in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=col_idx, value=v)
            c.border = bd
            if col_idx in (5, 6, 7, 8):
                c.alignment = Alignment(horizontal="right")
                c.number_format = '#,##0'
        # Status cell color
        st_cell = ws.cell(row=ri, column=9)
        if t['status'] == 'settled':
            st_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif t['status'] == 'partial':
            st_cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        else:
            st_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    # Auto-width
    widths = {1: 10, 2: 14, 3: 12, 4: 32, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14}
    for k, w in widths.items():
        ws.column_dimensions[chr(64 + k)].width = w

    # 🎯 v104.44.9 — Apply consolidated polish (single-truck per-trip)
    from utils.export_helpers import apply_consolidated_excel_polish
    apply_consolidated_excel_polish(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_vno = "".join(c for c in vehicle_no if c.isalnum() or c in "-_").strip("-_") or "truck"
    fname = f"{safe_vno}_{title_kind.replace(' ', '_')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/truck-owner/{vehicle_no}/whatsapp-text")
async def truck_owner_whatsapp_text(vehicle_no: str, kms_year: str = "", season: str = "", filter_status: str = "pending"):
    """Generate a plain-text WhatsApp message summarizing trips for the truck owner."""
    payload = await _build_pertrip_payload(vehicle_no, kms_year, season, "", "", filter_status)
    sm = payload["summary"]
    trips = payload["trips"][:10]  # limit to 10 trips

    lines = []
    lines.append(f"🛻 *{vehicle_no}* — {'Pending' if filter_status == 'pending' else 'Per-Trip'} Bhada")
    lines.append("")
    lines.append(f"📊 Trips: {sm['total_trips']}  ·  Bhada: ₹{sm['total_bhada']:,.0f}")
    lines.append(f"✅ Paid: ₹{sm['total_paid']:,.0f}  ·  ⚠️ Pending: ₹{sm['total_pending']:,.0f}")
    lines.append("")
    if trips:
        lines.append("*Trip details:*")
        for t in trips:
            tag = "🟢" if t['trans_type'] == 'sale' else "🔵"
            status_emoji = "✅" if t['status'] == 'settled' else ("🟡" if t['status'] == 'partial' else "⚠️")
            party = (t.get('party_name') or t.get('farmer_name') or '')[:22]
            lines.append(f"{tag} RST #{t['rst_no']} · {t['date']} · {party} · ₹{t['bhada']:,.0f} {status_emoji}")
    if len(payload["trips"]) > 10:
        lines.append(f"... +{len(payload['trips']) - 10} more")
    lines.append("")
    lines.append(f"_Total pending: ₹{sm['total_pending']:,.0f}_")

    return {"text": "\n".join(lines), "vehicle_no": vehicle_no, "summary": sm}



# ════════════════════════════════════════════════════════════════════════════
# 🛻 TRUCK OWNER — All Trucks Per-Trip PDF / Excel Export (combined view)
# ════════════════════════════════════════════════════════════════════════════
async def _build_pertrip_all_payload(
    kms_year: str = "", season: str = "",
    date_from: str = "", date_to: str = "",
    filter_status: str = "", trans_type: str = "", search: str = ""
):
    """Build all-trucks per-trip payload with additional trans_type + search filtering."""
    full = await per_trip_all(kms_year, season, date_from, date_to, filter_status)
    trips = full.get("trips", [])
    tt = (trans_type or "all").lower().strip()
    if tt and tt != "all":
        trips = [t for t in trips if (t.get("trans_type") or "") == tt]
    s = (search or "").strip().lower()
    if s:
        def _match(t):
            hay = " ".join([
                str(t.get("vehicle_no") or ""),
                str(t.get("party_name") or ""),
                str(t.get("farmer_name") or ""),
                str(t.get("rst_no") or ""),
            ]).lower()
            return s in hay
        trips = [t for t in trips if _match(t)]
    # Recompute summary on the *filtered* trips so PDF/Excel banner stays in sync.
    sm = {
        "total_trips": len(trips),
        "sale_count": sum(1 for t in trips if t.get("trans_type") == "sale"),
        "purchase_count": sum(1 for t in trips if t.get("trans_type") == "purchase"),
        "total_bhada": round(sum(float(t.get("bhada") or 0) for t in trips), 2),
        "total_paid": round(sum(float(t.get("paid_amount") or 0) for t in trips), 2),
        "total_pending": round(sum(float(t.get("pending_amount") or 0) for t in trips), 2),
        "settled_count": sum(1 for t in trips if t.get("status") == "settled"),
        "partial_count": sum(1 for t in trips if t.get("status") == "partial"),
        "pending_count": sum(1 for t in trips if t.get("status") == "pending"),
    }
    return {"trips": trips, "summary": sm, "total_trucks": full.get("total_trucks", 0)}


@router.get("/truck-owner/per-trip-all/pdf")
async def truck_owner_per_trip_all_pdf(
    kms_year: str = "", season: str = "",
    date_from: str = "", date_to: str = "",
    filter_status: str = "", trans_type: str = "", search: str = ""
):
    """Generate combined per-trip Bhada PDF for ALL trucks with active filters.
    Layout: Header → Subtitle → Trips Table → KPI Summary Banner (bottom) → Footer.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from fastapi.responses import StreamingResponse
    from utils.branding_helper import get_pdf_header_elements_from_db
    from utils.export_helpers import register_hindi_fonts
    from utils.date_format import fmt_date as fmt_d
    register_hindi_fonts()

    payload = await _build_pertrip_all_payload(kms_year, season, date_from, date_to, filter_status, trans_type, search)
    trips = payload["trips"]
    sm = payload["summary"]

    # Build subtitle reflecting active filters
    flt_bits = []
    if filter_status and filter_status != "all": flt_bits.append(f"Status: {filter_status.title()}")
    if trans_type and trans_type != "all": flt_bits.append(f"Type: {trans_type.title()}")
    if search: flt_bits.append(f"Search: {search}")
    flt_label = " · ".join(flt_bits) if flt_bits else "All"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    elements = []
    header_elems = await get_pdf_header_elements_from_db(
        title="Per-Trip Bhada — All Trucks",
        subtitle=f"KMS: {kms_year or 'All'} · Season: {season or 'All'} · Filter: {flt_label} · Trips: {sm['total_trips']} · Trucks: {payload.get('total_trucks', 0)}"
    )
    elements.extend(header_elems)

    # Trips table — landscape gives us room for Truck No column
    headers = ["RST", "Date", "Truck No", "Type", "Party", "Destination", "Net Wt", "Bhada", "Paid", "Pending", "Status"]
    rows = [headers]
    for t in trips:
        rows.append([
            f"#{t['rst_no']}", fmt_d(t['date']),
            t.get('vehicle_no') or '-',
            "Sale" if t['trans_type'] == 'sale' else ("Purchase" if t['trans_type'] == 'purchase' else (t.get('trans_type_raw') or '-')),
            (t.get('party_name') or '-')[:24],
            (t.get('farmer_name') or '-')[:24],
            f"{int(t.get('net_wt') or 0):,}" if t.get('net_wt') else '-',
            f"Rs.{t['bhada']:,.0f}",
            f"Rs.{t['paid_amount']:,.0f}" if t['paid_amount'] else '-',
            f"Rs.{t['pending_amount']:,.0f}" if t['pending_amount'] else '-',
            t['status'].title(),
        ])
    if len(rows) == 1:
        rows.append(["—", "Koi trip nahi", "", "", "", "", "", "", "", "", ""])

    col_widths = [16*mm, 22*mm, 28*mm, 18*mm, 42*mm, 42*mm, 18*mm, 22*mm, 22*mm, 22*mm, 22*mm]
    trips_tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    trips_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d1b2a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (6,1), (-2,-1), 'RIGHT'),
        ('ALIGN', (-1,0), (-1,-1), 'CENTER'),
        ('ALIGN', (3,1), (3,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f7f9fc')]),
        ('LEADING', (0,0), (-1,-1), 11),
        # Bhada column emphasis
        ('TEXTCOLOR', (7,1), (7,-1), colors.HexColor('#e65100')),
        ('FONTNAME', (7,1), (7,-1), 'Helvetica-Bold'),
        # Pending column emphasis
        ('TEXTCOLOR', (9,1), (9,-1), colors.HexColor('#b71c1c')),
        # RST + Truck No bold accents
        ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0,1), (0,-1), colors.HexColor('#1a237e')),
        ('FONTNAME', (2,1), (2,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (2,1), (2,-1), colors.HexColor('#0277bd')),
    ]))
    # Color status cells based on value
    for i, t in enumerate(trips, start=1):
        st = t.get('status')
        if st == 'settled':
            trips_tbl.setStyle(TableStyle([('BACKGROUND', (-1,i), (-1,i), colors.HexColor('#c8e6c9')), ('TEXTCOLOR', (-1,i), (-1,i), colors.HexColor('#1b5e20')), ('FONTNAME', (-1,i), (-1,i), 'Helvetica-Bold')]))
        elif st == 'partial':
            trips_tbl.setStyle(TableStyle([('BACKGROUND', (-1,i), (-1,i), colors.HexColor('#ffe0b2')), ('TEXTCOLOR', (-1,i), (-1,i), colors.HexColor('#e65100')), ('FONTNAME', (-1,i), (-1,i), 'Helvetica-Bold')]))
        else:
            trips_tbl.setStyle(TableStyle([('BACKGROUND', (-1,i), (-1,i), colors.HexColor('#ffcdd2')), ('TEXTCOLOR', (-1,i), (-1,i), colors.HexColor('#b71c1c')), ('FONTNAME', (-1,i), (-1,i), 'Helvetica-Bold')]))
    elements.append(trips_tbl)

    # ── KPI SUMMARY BANNER — placed BELOW the table for visibility after data ──
    elements.append(Spacer(1, 6*mm))
    banner_data = [
        # Row 1: labels
        ['TOTAL TRIPS', 'TOTAL BHADA', 'SETTLED', 'PARTIAL', 'PENDING'],
        # Row 2: values
        [
            f"{sm['total_trips']}\nSale {sm['sale_count']} · Purchase {sm['purchase_count']}",
            f"Rs.{sm['total_bhada']:,.0f}",
            f"Rs.{sm['total_paid']:,.0f}\n{sm['settled_count']} trips",
            f"{sm['partial_count']} trips",
            f"Rs.{sm['total_pending']:,.0f}\n{sm['pending_count']} trips",
        ],
    ]
    # Each tile a different color
    tile_colors = [
        colors.HexColor('#1a237e'),  # navy — total trips
        colors.HexColor('#e65100'),  # orange — total bhada
        colors.HexColor('#1b5e20'),  # green — settled
        colors.HexColor('#f9a825'),  # amber — partial
        colors.HexColor('#b71c1c'),  # red — pending
    ]
    banner_tbl = Table(banner_data, colWidths=[55*mm]*5, rowHeights=[8*mm, 16*mm])
    banner_style = [
        # Labels row (top)
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,0), 4),
        ('BOTTOMPADDING', (0,0), (-1,0), 4),
        # Values row (bottom)
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,1), (-1,1), 12),
        ('TEXTCOLOR', (0,1), (-1,1), colors.white),
        ('ALIGN', (0,1), (-1,1), 'CENTER'),
        ('VALIGN', (0,1), (-1,1), 'MIDDLE'),
        ('TOPPADDING', (0,1), (-1,1), 6),
        ('BOTTOMPADDING', (0,1), (-1,1), 6),
        ('LEADING', (0,1), (-1,1), 12),
        ('LINEBEFORE', (0,0), (-1,-1), 2, colors.white),
        ('LINEAFTER', (0,0), (-1,-1), 2, colors.white),
        ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor('#ffffff66')),
    ]
    # Apply per-tile background
    for col_idx, tcol in enumerate(tile_colors):
        banner_style.append(('BACKGROUND', (col_idx, 0), (col_idx, -1), tcol))
    banner_tbl.setStyle(TableStyle(banner_style))
    elements.append(KeepTogether(banner_tbl))

    # Footer
    elements.append(Spacer(1, 4*mm))
    style_small = ParagraphStyle('small', fontSize=7.5, alignment=TA_CENTER, textColor=colors.HexColor('#9e9e9e'), fontName='Helvetica-Oblique')
    elements.append(Paragraph(
        f"Generated on {fmt_d(datetime.now(timezone.utc).strftime('%Y-%m-%d'))} · {len(trips)} trip(s) across {payload.get('total_trucks', 0)} truck(s) · Filter: {flt_label}",
        style_small,
    ))

    doc.build(elements)
    buf.seek(0)
    fname = f"per_trip_bhada_all_trucks{('_' + filter_status) if filter_status and filter_status != 'all' else ''}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/truck-owner/per-trip-all/excel")
async def truck_owner_per_trip_all_excel(
    kms_year: str = "", season: str = "",
    date_from: str = "", date_to: str = "",
    filter_status: str = "", trans_type: str = "", search: str = ""
):
    """Professional Excel export — all trucks per-trip Bhada with active filters.
    Layout: Branding header → Filter strip → Trips Table (with auto-filter, frozen header) → KPI Banner (BELOW) → Footer.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    payload = await _build_pertrip_all_payload(kms_year, season, date_from, date_to, filter_status, trans_type, search)
    trips = payload["trips"]
    sm = payload["summary"]

    flt_bits = []
    if filter_status and filter_status != "all": flt_bits.append(f"Status: {filter_status.title()}")
    if trans_type and trans_type != "all": flt_bits.append(f"Type: {trans_type.title()}")
    if search: flt_bits.append(f"Search: {search}")
    flt_label = " · ".join(flt_bits) if flt_bits else "All"

    # Branding from DB (best-effort)
    company = "Rice Mill"
    tagline = ""
    try:
        br = await db.branding.find_one({}, {"_id": 0})
        if br:
            company = br.get("company_name") or company
            tagline = br.get("tagline") or ""
    except Exception:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Per-Trip Bhada (All)"
    ws.sheet_view.showGridLines = False

    # Style palette
    NAVY = "0D1B2A"
    DARK_GREEN = "1B5E20"
    LIGHT_GREY = "F7F9FC"
    BORDER_GREY = "D5DBE5"
    thin = Side(style="thin", color=BORDER_GREY)
    bd = Border(top=thin, bottom=thin, left=thin, right=thin)

    title_font = Font(name="Inter", bold=True, size=18, color="FFFFFF")
    sub_font_white = Font(name="Inter", size=10, color="E0E0E0")
    filter_font = Font(name="Inter", size=10, color="455A64", italic=True)
    hdr_font = Font(name="Inter", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    company_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

    # Row 1 — Company Brand banner
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:K1")
    c = ws.cell(row=1, column=1, value=company)
    c.font = title_font; c.fill = company_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 12):
        ws.cell(row=1, column=col).fill = company_fill

    # Row 2 — Subtitle (English-only avoids Devanagari encoding issues)
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:K2")
    c2 = ws.cell(row=2, column=1, value="Per-Trip Bhada Report — All Trucks")
    c2.font = sub_font_white; c2.fill = company_fill
    c2.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 12):
        ws.cell(row=2, column=col).fill = company_fill

    # Row 3 — Filter info strip (light blue)
    ws.row_dimensions[3].height = 22
    ws.merge_cells("A3:K3")
    filter_strip = (
        f"KMS: {kms_year or 'All'}  |  Season: {season or 'All'}  |  Filter: {flt_label}  |  "
        f"Trips: {sm['total_trips']}  |  Trucks: {payload.get('total_trucks', 0)}"
    )
    c3 = ws.cell(row=3, column=1, value=filter_strip)
    c3.font = filter_font
    c3.fill = PatternFill(start_color="EAF2FA", end_color="EAF2FA", fill_type="solid")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 12):
        ws.cell(row=3, column=col).fill = PatternFill(start_color="EAF2FA", end_color="EAF2FA", fill_type="solid")

    # Row 5 — Table header (skip row 4 for spacing)
    headers = ["RST", "Date", "Truck No", "Type", "Party", "Destination", "Net Wt (KG)", "Bhada", "Paid", "Pending", "Status"]
    HEADER_ROW = 5
    ws.row_dimensions[HEADER_ROW].height = 26
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = bd
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    body_font = Font(name="Inter", size=10, color="212121")
    rst_font = Font(name="Inter", bold=True, size=10, color="1A237E")
    truck_font = Font(name="Inter", bold=True, size=10, color="0277BD")
    money_font = Font(name="Inter", bold=True, size=10, color="E65100")
    pending_font = Font(name="Inter", bold=True, size=10, color="B71C1C")
    paid_font = Font(name="Inter", size=10, color="2E7D32")
    status_settled_font = Font(name="Inter", bold=True, size=10, color="1B5E20")
    status_partial_font = Font(name="Inter", bold=True, size=10, color="E65100")
    status_pending_font = Font(name="Inter", bold=True, size=10, color="B71C1C")

    DATA_START = HEADER_ROW + 1
    for ri, t in enumerate(trips, start=DATA_START):
        is_alt = (ri - DATA_START) % 2 == 1
        row_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid") if is_alt else None
        vals = [
            f"#{t['rst_no']}", t.get('date') or '',
            t.get('vehicle_no') or '',
            "Sale" if t['trans_type'] == 'sale' else ("Purchase" if t['trans_type'] == 'purchase' else (t.get('trans_type_raw') or '-')),
            t.get('party_name') or '-',
            t.get('farmer_name') or '-',
            int(t.get('net_wt') or 0),
            float(t['bhada']),
            float(t['paid_amount']),
            float(t['pending_amount']),
            t['status'].title(),
        ]
        for col_idx, v in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=col_idx, value=v)
            c.border = bd
            if row_fill:
                c.fill = row_fill
            # Per-column formatting
            if col_idx == 1:  # RST
                c.font = rst_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:  # Date
                c.font = body_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:  # Truck No
                c.font = truck_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:  # Type
                c.font = body_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (5, 6):  # Party / Destination
                c.font = body_font
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            elif col_idx == 7:  # Net Wt
                c.font = body_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0'
            elif col_idx == 8:  # Bhada
                c.font = money_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '"₹"#,##0'
            elif col_idx == 9:  # Paid
                c.font = paid_font if t['paid_amount'] else body_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '"₹"#,##0;[Color9]"-"'
            elif col_idx == 10:  # Pending
                c.font = pending_font if t['pending_amount'] else body_font
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '"₹"#,##0;[Color9]"-"'
            elif col_idx == 11:  # Status
                if t['status'] == 'settled':
                    c.font = status_settled_font
                    c.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif t['status'] == 'partial':
                    c.font = status_partial_font
                    c.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                else:
                    c.font = status_pending_font
                    c.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center")

    DATA_END = DATA_START + max(0, len(trips)) - 1 if trips else DATA_START - 1

    # Empty-state placeholder
    if not trips:
        ws.merge_cells(f"A{DATA_START}:K{DATA_START}")
        c = ws.cell(row=DATA_START, column=1, value="No trips found for the selected filters.")
        c.font = Font(name="Inter", italic=True, size=10, color="9E9E9E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bd
        DATA_END = DATA_START

    # Auto-filter + freeze header
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{max(DATA_END, HEADER_ROW)}"
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ── KPI SUMMARY BANNER (BELOW the table) ──
    BANNER_ROW1 = DATA_END + 2  # Spacing row between table and banner
    BANNER_ROW2 = BANNER_ROW1 + 1
    ws.row_dimensions[BANNER_ROW1].height = 18
    ws.row_dimensions[BANNER_ROW2].height = 26

    # KPI tiles: each spans 2-3 columns. We have 11 cols total.
    # Layout (5 KPI tiles): A-B = Trips · C-D = Bhada · E-F = Settled · G-H = Partial · I-K = Pending
    tiles = [
        ("TOTAL TRIPS",  f"{sm['total_trips']}", "1A237E", "A", "B"),
        ("TOTAL BHADA",  f"₹{sm['total_bhada']:,.0f}", "E65100", "C", "D"),
        ("SETTLED",      f"₹{sm['total_paid']:,.0f}", "1B5E20", "E", "F"),
        ("PARTIAL",      f"{sm['partial_count']} trip(s)", "F57F17", "G", "H"),
        ("PENDING",      f"₹{sm['total_pending']:,.0f}", "B71C1C", "I", "K"),
    ]
    for label, value, color, c1, c2 in tiles:
        ws.merge_cells(f"{c1}{BANNER_ROW1}:{c2}{BANNER_ROW1}")
        lbl = ws[f"{c1}{BANNER_ROW1}"]
        lbl.value = label
        lbl.font = Font(name="Inter", bold=True, size=9, color="FFFFFF")
        lbl.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        # Also fill all merged cells with same color
        for col in range(ord(c1) - 64, ord(c2) - 64 + 1):
            ws.cell(row=BANNER_ROW1, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.merge_cells(f"{c1}{BANNER_ROW2}:{c2}{BANNER_ROW2}")
        val = ws[f"{c1}{BANNER_ROW2}"]
        val.value = value
        val.font = Font(name="Inter", bold=True, size=14, color="FFFFFF")
        val.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        val.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(ord(c1) - 64, ord(c2) - 64 + 1):
            ws.cell(row=BANNER_ROW2, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Trips composition strip (Sale vs Purchase) below the banner
    COMP_ROW = BANNER_ROW2 + 1
    ws.row_dimensions[COMP_ROW].height = 18
    ws.merge_cells(f"A{COMP_ROW}:K{COMP_ROW}")
    comp = ws.cell(row=COMP_ROW, column=1,
                   value=f"Composition: {sm['sale_count']} Sale  ·  {sm['purchase_count']} Purchase  ·  {sm['settled_count']} Settled  ·  {sm['partial_count']} Partial  ·  {sm['pending_count']} Pending")
    comp.font = Font(name="Inter", size=10, color="455A64", italic=True)
    comp.alignment = Alignment(horizontal="center", vertical="center")
    comp.fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    for col in range(1, 12):
        ws.cell(row=COMP_ROW, column=col).fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    # Footer
    FOOTER_ROW = COMP_ROW + 2
    ws.merge_cells(f"A{FOOTER_ROW}:K{FOOTER_ROW}")
    fcell = ws.cell(row=FOOTER_ROW, column=1,
                    value=f"Generated on {datetime.now(timezone.utc).strftime('%d-%b-%Y')}  ·  {sm['total_trips']} trip(s) across {payload.get('total_trucks', 0)} truck(s)  ·  Filter: {flt_label}")
    fcell.font = Font(name="Inter", size=9, italic=True, color="9E9E9E")
    fcell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths — properly tuned for content
    widths = [10, 12, 18, 11, 26, 26, 12, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"per_trip_bhada_all_trucks{('_' + filter_status) if filter_status and filter_status != 'all' else ''}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )
