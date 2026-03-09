"""
Test Suite: Staff Attendance Export Features (Iteration 31)
Tests the Staff Attendance PDF and Excel exports to verify:
1. PDF export returns valid PDF with Monthly Summary (Page 2)
   - Breakdown (P/A/H/CH) section
   - Month-wise Estimated Salary section
2. Excel export returns valid XLSX with Monthly Summary sheet
   - Breakdown (P/A/H/CH) section
   - Month-wise Estimated Salary section
3. Staff Payments export works
"""

import pytest
import requests
import os
from io import BytesIO

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')


class TestStaffAttendanceExports:
    """Staff Attendance Export API Tests"""

    def test_pdf_export_returns_200(self):
        """Test PDF export endpoint returns 200 OK"""
        response = requests.get(
            f"{BASE_URL}/api/staff/export/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31", "fmt": "pdf"}
        )
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("PDF export returns 200 OK")

    def test_pdf_export_content_type(self):
        """Test PDF export returns correct content type"""
        response = requests.get(
            f"{BASE_URL}/api/staff/export/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31", "fmt": "pdf"}
        )
        assert response.status_code == 200
        assert "application/pdf" in response.headers.get("Content-Type", ""), \
            f"Expected application/pdf, got {response.headers.get('Content-Type')}"
        print("PDF content type is application/pdf")

    def test_pdf_export_valid_pdf_header(self):
        """Test PDF export returns valid PDF file (magic header check)"""
        response = requests.get(
            f"{BASE_URL}/api/staff/export/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31", "fmt": "pdf"}
        )
        assert response.status_code == 200
        # Check PDF magic header
        assert response.content[:4] == b'%PDF', "Response is not a valid PDF file"
        print("PDF has valid %PDF header")

    def test_pdf_export_contains_monthly_summary(self):
        """Test PDF export contains Monthly Summary sections"""
        response = requests.get(
            f"{BASE_URL}/api/staff/export/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31", "fmt": "pdf"}
        )
        assert response.status_code == 200
        
        # Import PyPDF2 to extract text
        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(BytesIO(response.content))
            
            # Check for 2 pages (attendance + monthly summary)
            assert len(reader.pages) >= 2, f"Expected at least 2 pages, got {len(reader.pages)}"
            
            # Extract text from page 2 (monthly summary)
            page2_text = reader.pages[1].extract_text() if len(reader.pages) > 1 else ""
            
            # Verify Monthly Summary section
            assert "Monthly Summary" in page2_text or "Masik Saransh" in page2_text, \
                "Monthly Summary section not found in PDF page 2"
            
            # Verify Breakdown section
            assert "Breakdown" in page2_text, "Breakdown (P/A/H/CH) section not found in PDF"
            
            # Verify Month-wise Estimated Salary section
            assert "Month-wise" in page2_text or "Anumanit" in page2_text, \
                "Month-wise Estimated Salary section not found in PDF"
            
            print("PDF contains all Monthly Summary sections")
        except ImportError:
            pytest.skip("PyPDF2 not installed - skipping PDF content verification")

    def test_excel_export_returns_200(self):
        """Test Excel export endpoint returns 200 OK"""
        response = requests.get(
            f"{BASE_URL}/api/staff/export/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31", "fmt": "excel"}
        )
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("Excel export returns 200 OK")

    def test_excel_export_content_type(self):
        """Test Excel export returns correct content type"""
        response = requests.get(
            f"{BASE_URL}/api/staff/export/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31", "fmt": "excel"}
        )
        assert response.status_code == 200
        expected_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert expected_type in response.headers.get("Content-Type", ""), \
            f"Expected {expected_type}, got {response.headers.get('Content-Type')}"
        print("Excel content type is correct")

    def test_excel_export_valid_xlsx_header(self):
        """Test Excel export returns valid XLSX file (ZIP magic header check)"""
        response = requests.get(
            f"{BASE_URL}/api/staff/export/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31", "fmt": "excel"}
        )
        assert response.status_code == 200
        # Check ZIP magic header (XLSX is a ZIP file)
        assert response.content[:4] == b'PK\x03\x04', "Response is not a valid XLSX file"
        print("Excel has valid ZIP/XLSX header")

    def test_excel_export_has_monthly_summary_sheet(self):
        """Test Excel export contains Monthly Summary sheet with required sections"""
        response = requests.get(
            f"{BASE_URL}/api/staff/export/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31", "fmt": "excel"}
        )
        assert response.status_code == 200
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(response.content))
            
            # Check sheet names
            assert "Attendance" in wb.sheetnames, "Attendance sheet not found"
            assert "Monthly Summary" in wb.sheetnames, "Monthly Summary sheet not found"
            
            ws2 = wb["Monthly Summary"]
            
            # Find Breakdown and Month-wise Salary sections
            breakdown_found = False
            monthwise_salary_found = False
            
            for row in range(1, min(100, ws2.max_row + 1)):
                cell_val = ws2.cell(row=row, column=1).value
                if cell_val:
                    cell_str = str(cell_val)
                    if "Breakdown" in cell_str:
                        breakdown_found = True
                    if "Month-wise" in cell_str:
                        monthwise_salary_found = True
            
            assert breakdown_found, "Breakdown (P/A/H/CH) section not found in Excel"
            assert monthwise_salary_found, "Month-wise Estimated Salary section not found in Excel"
            
            print("Excel contains all Monthly Summary sections")
        except ImportError:
            pytest.skip("openpyxl not installed - skipping Excel content verification")

    def test_staff_payments_export_returns_200(self):
        """Test Staff Payments export endpoint returns 200 OK"""
        response = requests.get(f"{BASE_URL}/api/staff/export/payments")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("Staff Payments export returns 200 OK")

    def test_staff_payments_export_content_type(self):
        """Test Staff Payments export returns correct content type"""
        response = requests.get(f"{BASE_URL}/api/staff/export/payments")
        assert response.status_code == 200
        expected_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert expected_type in response.headers.get("Content-Type", ""), \
            f"Expected {expected_type}, got {response.headers.get('Content-Type')}"
        print("Staff Payments export content type is correct")


class TestStaffAPIBasics:
    """Basic Staff API Tests"""

    def test_get_staff_list(self):
        """Test GET /api/staff returns staff list"""
        response = requests.get(f"{BASE_URL}/api/staff")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list), "Expected list response"
        print(f"Staff list API works - {len(data)} staff members")

    def test_get_staff_attendance(self):
        """Test GET /api/staff/attendance returns attendance records"""
        response = requests.get(
            f"{BASE_URL}/api/staff/attendance",
            params={"date_from": "2025-01-01", "date_to": "2025-01-31"}
        )
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list), "Expected list response"
        print(f"Staff attendance API works - {len(data)} records")

    def test_get_staff_payments(self):
        """Test GET /api/staff/payments returns payment records"""
        response = requests.get(f"{BASE_URL}/api/staff/payments")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list), "Expected list response"
        print(f"Staff payments API works - {len(data)} records")
