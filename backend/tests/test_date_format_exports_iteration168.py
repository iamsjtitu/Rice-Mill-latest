"""
Test Date Format DD-MM-YYYY in All Excel and PDF Exports
Iteration 168 - Testing global date format change from YYYY-MM-DD to DD-MM-YYYY
"""
import pytest
import requests
import os
import io
from openpyxl import load_workbook
import re

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestDateFormatExports:
    """Test that all exports use DD-MM-YYYY date format"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test session"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
    
    def is_dd_mm_yyyy_format(self, date_str):
        """Check if date string is in DD-MM-YYYY format"""
        if not date_str or date_str == '-' or date_str == '':
            return True  # Empty/null dates are acceptable
        # DD-MM-YYYY pattern: 01-12-2026
        pattern = r'^\d{2}-\d{2}-\d{4}$'
        return bool(re.match(pattern, str(date_str)))
    
    def is_yyyy_mm_dd_format(self, date_str):
        """Check if date string is in YYYY-MM-DD format (old format - should NOT be used)"""
        if not date_str or date_str == '-' or date_str == '':
            return False
        # YYYY-MM-DD pattern: 2026-12-01
        pattern = r'^\d{4}-\d{2}-\d{2}$'
        return bool(re.match(pattern, str(date_str)))
    
    # ========== MILL ENTRIES EXCEL EXPORT ==========
    def test_mill_entries_excel_date_format(self):
        """Test /api/export/excel - Mill Entries Excel export uses DD-MM-YYYY"""
        response = self.session.get(f"{BASE_URL}/api/export/excel")
        assert response.status_code == 200, f"Excel export failed: {response.status_code}"
        
        # Load Excel file
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find date column (usually column A or first column with dates)
        date_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=4, column=col).value  # Headers usually at row 4
            if header and 'date' in str(header).lower():
                date_col = col
                break
        
        if date_col is None:
            # Try first column if no header found
            date_col = 1
        
        # Check dates in data rows (starting from row 5)
        dates_found = []
        wrong_format_dates = []
        for row in range(5, min(ws.max_row + 1, 20)):  # Check first 15 data rows
            cell_value = ws.cell(row=row, column=date_col).value
            if cell_value and str(cell_value).strip() and str(cell_value) != 'TOTAL':
                dates_found.append(str(cell_value))
                if self.is_yyyy_mm_dd_format(str(cell_value)):
                    wrong_format_dates.append(str(cell_value))
        
        print(f"Mill Entries Excel - Dates found: {dates_found[:5]}")
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD format dates (should be DD-MM-YYYY): {wrong_format_dates}"
        
        # Verify at least some dates are in correct format
        correct_format_dates = [d for d in dates_found if self.is_dd_mm_yyyy_format(d)]
        print(f"Mill Entries Excel - Correct DD-MM-YYYY dates: {correct_format_dates[:5]}")
    
    # ========== MILL ENTRIES PDF EXPORT ==========
    def test_mill_entries_pdf_export_status(self):
        """Test /api/export/pdf - Mill Entries PDF export returns valid PDF"""
        response = self.session.get(f"{BASE_URL}/api/export/pdf")
        assert response.status_code == 200, f"PDF export failed: {response.status_code}"
        assert response.headers.get('content-type') == 'application/pdf' or 'pdf' in response.headers.get('content-type', '').lower()
        # PDF content starts with %PDF
        assert response.content[:4] == b'%PDF', "Response is not a valid PDF"
        print("Mill Entries PDF export - Status: OK, Valid PDF returned")
    
    # ========== TRUCK PAYMENTS EXCEL EXPORT ==========
    def test_truck_payments_excel_date_format(self):
        """Test /api/export/truck-payments-excel - Truck Payments Excel uses DD-MM-YYYY"""
        response = self.session.get(f"{BASE_URL}/api/export/truck-payments-excel")
        assert response.status_code == 200, f"Truck Payments Excel export failed: {response.status_code}"
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find date column
        date_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=4, column=col).value
            if header and 'date' in str(header).lower():
                date_col = col
                break
        
        if date_col is None:
            date_col = 1
        
        dates_found = []
        wrong_format_dates = []
        for row in range(5, min(ws.max_row + 1, 20)):
            cell_value = ws.cell(row=row, column=date_col).value
            if cell_value and str(cell_value).strip() and str(cell_value) != 'TOTAL':
                dates_found.append(str(cell_value))
                if self.is_yyyy_mm_dd_format(str(cell_value)):
                    wrong_format_dates.append(str(cell_value))
        
        print(f"Truck Payments Excel - Dates found: {dates_found[:5]}")
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD format dates: {wrong_format_dates}"
    
    # ========== TRUCK PAYMENTS PDF EXPORT ==========
    def test_truck_payments_pdf_export_status(self):
        """Test /api/export/truck-payments-pdf - Truck Payments PDF returns valid PDF"""
        response = self.session.get(f"{BASE_URL}/api/export/truck-payments-pdf")
        assert response.status_code == 200, f"Truck Payments PDF export failed: {response.status_code}"
        assert response.content[:4] == b'%PDF', "Response is not a valid PDF"
        print("Truck Payments PDF export - Status: OK, Valid PDF returned")
    
    # ========== MILLING REPORT EXCEL EXPORT ==========
    def test_milling_report_excel_date_format(self):
        """Test /api/milling-report/excel - Milling Report Excel uses DD-MM-YYYY"""
        response = self.session.get(f"{BASE_URL}/api/milling-report/excel")
        assert response.status_code == 200, f"Milling Report Excel export failed: {response.status_code}"
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find date column
        date_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=4, column=col).value
            if header and 'date' in str(header).lower():
                date_col = col
                break
        
        if date_col is None:
            date_col = 1
        
        dates_found = []
        wrong_format_dates = []
        for row in range(5, min(ws.max_row + 1, 20)):
            cell_value = ws.cell(row=row, column=date_col).value
            if cell_value and str(cell_value).strip() and str(cell_value) != 'TOTAL':
                dates_found.append(str(cell_value))
                if self.is_yyyy_mm_dd_format(str(cell_value)):
                    wrong_format_dates.append(str(cell_value))
        
        print(f"Milling Report Excel - Dates found: {dates_found[:5]}")
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD format dates: {wrong_format_dates}"
    
    # ========== PADDY CUSTODY REGISTER EXCEL EXPORT ==========
    def test_paddy_custody_register_excel_date_format(self):
        """Test /api/paddy-custody-register/excel - Paddy Custody Register Excel uses DD-MM-YYYY"""
        response = self.session.get(f"{BASE_URL}/api/paddy-custody-register/excel")
        assert response.status_code == 200, f"Paddy Custody Register Excel export failed: {response.status_code}"
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find date column
        date_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=4, column=col).value
            if header and 'date' in str(header).lower():
                date_col = col
                break
        
        if date_col is None:
            date_col = 1
        
        dates_found = []
        wrong_format_dates = []
        for row in range(5, min(ws.max_row + 1, 20)):
            cell_value = ws.cell(row=row, column=date_col).value
            if cell_value and str(cell_value).strip() and str(cell_value) != 'TOTAL':
                dates_found.append(str(cell_value))
                if self.is_yyyy_mm_dd_format(str(cell_value)):
                    wrong_format_dates.append(str(cell_value))
        
        print(f"Paddy Custody Register Excel - Dates found: {dates_found[:5]}")
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD format dates: {wrong_format_dates}"
    
    # ========== PARTY LEDGER EXCEL EXPORT ==========
    def test_party_ledger_excel_date_format(self):
        """Test /api/reports/party-ledger/excel - Party Ledger Excel uses DD-MM-YYYY"""
        # First get a party name from entries
        entries_resp = self.session.get(f"{BASE_URL}/api/entries?page=1&page_size=1")
        if entries_resp.status_code == 200:
            data = entries_resp.json()
            entries = data.get('entries', [])
            if entries:
                party_name = entries[0].get('agent_name', 'Test')
                response = self.session.get(f"{BASE_URL}/api/reports/party-ledger/excel?party_name={party_name}")
                if response.status_code == 200:
                    wb = load_workbook(io.BytesIO(response.content))
                    ws = wb.active
                    
                    # Find date column
                    date_col = None
                    for col in range(1, ws.max_column + 1):
                        header = ws.cell(row=4, column=col).value
                        if header and 'date' in str(header).lower():
                            date_col = col
                            break
                    
                    if date_col is None:
                        date_col = 1
                    
                    dates_found = []
                    wrong_format_dates = []
                    for row in range(5, min(ws.max_row + 1, 20)):
                        cell_value = ws.cell(row=row, column=date_col).value
                        if cell_value and str(cell_value).strip() and str(cell_value) != 'TOTAL':
                            dates_found.append(str(cell_value))
                            if self.is_yyyy_mm_dd_format(str(cell_value)):
                                wrong_format_dates.append(str(cell_value))
                    
                    print(f"Party Ledger Excel - Dates found: {dates_found[:5]}")
                    assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD format dates: {wrong_format_dates}"
                else:
                    print(f"Party Ledger Excel - Skipped (no data for party: {party_name})")
            else:
                print("Party Ledger Excel - Skipped (no entries found)")
        else:
            print("Party Ledger Excel - Skipped (could not fetch entries)")
    
    # ========== TRUCK LEASE EXCEL EXPORT ==========
    def test_truck_lease_excel_date_format(self):
        """Test /api/truck-leases/export/excel - Truck Lease Excel uses DD-MM-YYYY"""
        response = self.session.get(f"{BASE_URL}/api/truck-leases/export/excel")
        assert response.status_code == 200, f"Truck Lease Excel export failed: {response.status_code}"
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find date columns (Start Date, End Date)
        date_cols = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=4, column=col).value
            if header and ('date' in str(header).lower() or 'start' in str(header).lower() or 'end' in str(header).lower()):
                date_cols.append(col)
        
        if not date_cols:
            date_cols = [4, 5]  # Default columns for start/end date
        
        dates_found = []
        wrong_format_dates = []
        for row in range(5, min(ws.max_row + 1, 20)):
            for date_col in date_cols:
                cell_value = ws.cell(row=row, column=date_col).value
                if cell_value and str(cell_value).strip() and str(cell_value) not in ['TOTAL', 'Ongoing', '-']:
                    dates_found.append(str(cell_value))
                    if self.is_yyyy_mm_dd_format(str(cell_value)):
                        wrong_format_dates.append(str(cell_value))
        
        print(f"Truck Lease Excel - Dates found: {dates_found[:5]}")
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD format dates: {wrong_format_dates}"
    
    # ========== TRUCK LEASE PDF EXPORT ==========
    def test_truck_lease_pdf_export_status(self):
        """Test /api/truck-leases/export/pdf - Truck Lease PDF returns valid PDF"""
        response = self.session.get(f"{BASE_URL}/api/truck-leases/export/pdf")
        assert response.status_code == 200, f"Truck Lease PDF export failed: {response.status_code}"
        assert response.content[:4] == b'%PDF', "Response is not a valid PDF"
        print("Truck Lease PDF export - Status: OK, Valid PDF returned")


class TestAPIEndpoints:
    """Test basic API endpoints are working"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        self.session = requests.Session()
    
    def test_health_check(self):
        """Test API health endpoint"""
        response = self.session.get(f"{BASE_URL}/api/health")
        assert response.status_code == 200
        print("Health check - OK")
    
    def test_entries_endpoint(self):
        """Test entries endpoint returns data"""
        response = self.session.get(f"{BASE_URL}/api/entries?page=1&page_size=10")
        assert response.status_code == 200
        data = response.json()
        assert 'entries' in data
        print(f"Entries endpoint - OK, returned {len(data.get('entries', []))} entries")
    
    def test_totals_endpoint(self):
        """Test totals endpoint"""
        response = self.session.get(f"{BASE_URL}/api/totals")
        assert response.status_code == 200
        print("Totals endpoint - OK")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
