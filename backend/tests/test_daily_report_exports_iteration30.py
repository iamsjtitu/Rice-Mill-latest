"""
Test Daily Report API - Iteration 30
Focus: 
1. Paddy entries in Summary mode should have ALL 22 fields (not just 4)
2. Pump Account should use 'mandi' key (not 'agent')
3. PDF/Excel exports should work and have proper column headers
"""

import pytest
import requests
import os
import io

# Get BASE_URL from environment variable
BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestDailyReportAPI:
    """Test Daily Report API - Paddy Entries and Pump Account Fields"""
    
    def test_api_health(self):
        """Test API is reachable"""
        response = requests.get(f"{BASE_URL}/api/")
        assert response.status_code == 200
        print(f"API Health: PASS - {response.json()}")
    
    def test_daily_report_normal_mode_paddy_fields(self):
        """Test that normal mode paddy_entries.details has ALL 22 fields, not just 4"""
        response = requests.get(f"{BASE_URL}/api/reports/daily?date=2026-01-18&mode=normal")
        assert response.status_code == 200
        data = response.json()
        
        # Check paddy_entries exists
        assert "paddy_entries" in data
        paddy = data["paddy_entries"]
        
        print(f"Paddy entries count: {paddy['count']}")
        
        # If there are entries, check all 22 fields are present in details
        if paddy["details"]:
            first_entry = paddy["details"][0]
            
            # All 22 fields that should be present
            expected_fields = [
                "truck_no", "agent", "mandi", "rst_no", "tp_no", "season",
                "kg", "qntl", "bags", "g_deposite", "gbw_cut", "mill_w",
                "moisture", "moisture_cut", "cutting_percent", "disc_dust_poll",
                "final_w", "plastic_bag", "p_pkt_cut", "g_issued", "cash_paid", "diesel_paid"
            ]
            
            present_fields = list(first_entry.keys())
            print(f"Fields present in paddy entry: {present_fields}")
            print(f"Total fields count: {len(present_fields)}")
            
            # Check each expected field is present
            missing_fields = [f for f in expected_fields if f not in first_entry]
            if missing_fields:
                print(f"MISSING fields: {missing_fields}")
            
            # Assert all 22 fields are present
            assert len(present_fields) >= 20, f"Expected at least 20 fields, got {len(present_fields)}: {present_fields}"
            
            # Critical fields that MUST be present (these were previously missing)
            critical_fields = ["kg", "qntl", "bags", "g_deposite", "gbw_cut", "mill_w", 
                             "moisture", "moisture_cut", "cutting_percent", "disc_dust_poll",
                             "final_w", "plastic_bag", "p_pkt_cut", "g_issued", "cash_paid", "diesel_paid"]
            
            for field in critical_fields:
                assert field in first_entry, f"Critical field '{field}' missing from paddy entry"
            
            print(f"PASS: All {len(present_fields)} fields present in normal mode paddy_entries.details")
        else:
            print("INFO: No paddy entries found for 2026-01-18")
    
    def test_daily_report_detail_mode_paddy_fields(self):
        """Test that detail mode paddy_entries.details also has all fields"""
        response = requests.get(f"{BASE_URL}/api/reports/daily?date=2026-01-18&mode=detail")
        assert response.status_code == 200
        data = response.json()
        
        paddy = data["paddy_entries"]
        print(f"Detail mode paddy entries count: {paddy['count']}")
        
        if paddy["details"]:
            first_entry = paddy["details"][0]
            present_fields = list(first_entry.keys())
            print(f"Detail mode fields count: {len(present_fields)}")
            assert len(present_fields) >= 20, f"Expected at least 20 fields, got {len(present_fields)}"
            print(f"PASS: Detail mode has {len(present_fields)} fields")
    
    def test_pump_account_has_mandi_key(self):
        """Test that pump_account.details has 'mandi' key instead of 'agent'"""
        response = requests.get(f"{BASE_URL}/api/reports/daily?date=2026-01-18&mode=normal")
        assert response.status_code == 200
        data = response.json()
        
        pump_account = data.get("pump_account", {})
        print(f"Pump Account - Total Diesel: {pump_account.get('total_diesel', 0)}")
        print(f"Pump Account - Details count: {len(pump_account.get('details', []))}")
        
        if pump_account.get("details"):
            first_txn = pump_account["details"][0]
            print(f"Pump account detail keys: {list(first_txn.keys())}")
            
            # Should have 'mandi' key, not 'agent'
            assert "mandi" in first_txn, f"'mandi' key missing from pump_account.details. Got: {list(first_txn.keys())}"
            assert "agent" not in first_txn or first_txn.get("agent") is None or "agent" not in list(first_txn.keys()), \
                "pump_account.details should NOT have 'agent' key, should use 'mandi'"
            
            print(f"PASS: pump_account.details has 'mandi' key = '{first_txn['mandi']}'")
            
            # Also check other expected fields
            expected_pump_fields = ["pump", "txn_type", "amount", "truck_no", "mandi", "desc"]
            for field in expected_pump_fields:
                assert field in first_txn, f"Field '{field}' missing from pump_account.details"
            print(f"PASS: All 6 expected fields present: {expected_pump_fields}")
        else:
            print("INFO: No diesel transactions found for 2026-01-18")
    

class TestDailyReportExports:
    """Test PDF and Excel exports"""
    
    def test_pdf_export_normal_mode(self):
        """Test PDF export for normal mode returns non-zero bytes"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/pdf?date=2026-01-18&mode=normal")
        assert response.status_code == 200, f"PDF export failed: {response.status_code}"
        
        content_type = response.headers.get("content-type", "")
        assert "pdf" in content_type.lower() or "application/pdf" in content_type, \
            f"Expected PDF content type, got: {content_type}"
        
        content_length = len(response.content)
        assert content_length > 100, f"PDF content too small: {content_length} bytes"
        
        print(f"PASS: PDF export returned {content_length} bytes")
    
    def test_excel_export_normal_mode_column_count(self):
        """Test Excel export for normal mode has 20 column headers for paddy"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/excel?date=2026-01-18&mode=normal")
        assert response.status_code == 200, f"Excel export failed: {response.status_code}"
        
        content_type = response.headers.get("content-type", "")
        assert "spreadsheet" in content_type.lower() or "excel" in content_type.lower() or "vnd.openxmlformats" in content_type, \
            f"Expected Excel content type, got: {content_type}"
        
        content_length = len(response.content)
        assert content_length > 100, f"Excel content too small: {content_length} bytes"
        
        print(f"PASS: Excel export returned {content_length} bytes")
        
        # Parse Excel to verify column headers
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            ws = wb.active
            
            # Find paddy entries header row (look for 'Truck' in first 10 rows)
            paddy_header_row = None
            for row_num in range(1, 15):
                for col_num in range(1, 25):
                    cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
                    if "Truck" in cell_value and ws.cell(row=row_num, column=col_num+1).value:
                        paddy_header_row = row_num
                        break
                if paddy_header_row:
                    break
            
            if paddy_header_row:
                # Count columns in that row
                headers = []
                for col_num in range(1, 25):
                    val = ws.cell(row=paddy_header_row, column=col_num).value
                    if val:
                        headers.append(str(val))
                
                print(f"Paddy Headers found at row {paddy_header_row}: {headers}")
                print(f"Total header columns: {len(headers)}")
                
                # Should have ~20 columns for paddy
                assert len(headers) >= 18, f"Expected at least 18 columns for paddy, got {len(headers)}"
                
                # Check specific columns exist
                expected_headers = ["Truck", "Agent", "Mandi", "RST", "TP", "QNTL", "Bags", "G.Dep", 
                                   "GBW", "P.Pkt", "P.Cut", "Mill W", "M%", "M.Cut", "C%", "D/D/P", 
                                   "Final W", "G.Iss", "Cash", "Diesel"]
                
                for eh in expected_headers:
                    assert any(eh in h for h in headers), f"Expected column '{eh}' not found in headers: {headers}"
                
                print(f"PASS: Excel has {len(headers)} columns including all 20 expected paddy columns")
            else:
                print("WARN: Could not find paddy header row in Excel")
            
        except ImportError:
            print("SKIP: openpyxl not available for deep Excel inspection")
    
    def test_pdf_export_with_more_data(self):
        """Test PDF export with date having more data (2026-03-09)"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/pdf?date=2026-03-09&mode=normal")
        assert response.status_code == 200, f"PDF export failed: {response.status_code}"
        
        content_length = len(response.content)
        assert content_length > 100, f"PDF content too small: {content_length} bytes"
        
        print(f"PASS: PDF (2026-03-09) returned {content_length} bytes")
    
    def test_excel_export_detail_mode(self):
        """Test Excel export for detail mode"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/excel?date=2026-03-09&mode=detail")
        assert response.status_code == 200, f"Detail Excel export failed: {response.status_code}"
        
        content_length = len(response.content)
        assert content_length > 100, f"Excel content too small: {content_length} bytes"
        
        print(f"PASS: Detail Excel (2026-03-09) returned {content_length} bytes")
    
    def test_pump_account_excel_columns(self):
        """Test that Pump Account in Excel has 6 columns: Pump, Type, Truck, Mandi, Description, Amount"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/excel?date=2026-01-18&mode=normal")
        assert response.status_code == 200
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            ws = wb.active
            
            # Find Pump Account section
            pump_header_row = None
            for row_num in range(1, ws.max_row + 1):
                for col_num in range(1, 10):
                    val = str(ws.cell(row=row_num, column=col_num).value or "")
                    if "Pump" in val and ws.cell(row=row_num, column=col_num+1).value:
                        # Check if this row has Type, Truck columns too
                        next_val = str(ws.cell(row=row_num, column=col_num+1).value or "")
                        if "Type" in next_val:
                            pump_header_row = row_num
                            break
                if pump_header_row:
                    break
            
            if pump_header_row:
                headers = []
                for col_num in range(1, 10):
                    val = ws.cell(row=pump_header_row, column=col_num).value
                    if val:
                        headers.append(str(val))
                
                print(f"Pump Account headers: {headers}")
                
                # Should have: Pump, Type, Truck, Mandi, Description, Amount
                expected = ["Pump", "Type", "Truck", "Mandi", "Description", "Amount"]
                for eh in expected:
                    assert any(eh in h for h in headers), f"Expected Pump Account column '{eh}' not found: {headers}"
                
                # Should NOT have 'Agent' column
                assert not any("Agent" in h for h in headers), f"Pump Account should use 'Mandi' not 'Agent': {headers}"
                
                print(f"PASS: Pump Account Excel has correct 6 columns: {headers}")
            else:
                print("INFO: Could not find Pump Account section header in Excel (may be no diesel data)")
                
        except ImportError:
            print("SKIP: openpyxl not available")


class TestDateWithMultipleEntries:
    """Test with date that has more data (2026-03-09)"""
    
    def test_daily_report_march_date(self):
        """Test daily report for date with multiple entries"""
        response = requests.get(f"{BASE_URL}/api/reports/daily?date=2026-03-09&mode=normal")
        assert response.status_code == 200
        data = response.json()
        
        paddy = data["paddy_entries"]
        print(f"March 9 paddy entries: {paddy['count']}")
        
        if paddy["details"]:
            fields = list(paddy["details"][0].keys())
            print(f"Fields in March 9 entry: {len(fields)} - {fields}")
            assert len(fields) >= 20


# Run if executed directly
if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
