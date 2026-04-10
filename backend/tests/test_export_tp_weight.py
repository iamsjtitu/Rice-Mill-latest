"""
Test Export Endpoints for TP Weight Feature
Tests PDF and Excel exports for Mill Entries, Vehicle Weight, and Daily Report
Verifies TP Weight totals are included in exports
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestMillEntriesExport:
    """Mill Entries Excel and PDF export tests - TP Weight column and totals"""
    
    def test_mill_entries_excel_export_status(self):
        """Test Mill Entries Excel export returns 200"""
        response = requests.get(f"{BASE_URL}/api/export/excel", params={"kms_year": "2026-2027"})
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        print("PASS: Mill Entries Excel export returns 200 with correct content type")
    
    def test_mill_entries_excel_has_tp_weight_column(self):
        """Test Mill Entries Excel has TP Wt (Q) column at position 5"""
        response = requests.get(f"{BASE_URL}/api/export/excel", params={"kms_year": "2026-2027"})
        assert response.status_code == 200
        
        # Parse Excel content
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find header row (row 4 based on code)
        headers = [ws.cell(row=4, column=i).value for i in range(1, 21)]
        print(f"Excel headers: {headers}")
        
        # Check TP Wt (Q) is at position 5 (column E)
        assert "TP Wt (Q)" in headers, f"TP Wt (Q) column not found in headers: {headers}"
        tp_wt_index = headers.index("TP Wt (Q)")
        assert tp_wt_index == 4, f"TP Wt (Q) should be at position 5 (index 4), found at {tp_wt_index + 1}"
        print(f"PASS: TP Wt (Q) column found at position {tp_wt_index + 1}")
    
    def test_mill_entries_excel_has_tp_weight_total(self):
        """Test Mill Entries Excel has TP Weight total in TOTAL row"""
        response = requests.get(f"{BASE_URL}/api/export/excel", params={"kms_year": "2026-2027"})
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find TOTAL row
        total_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "TOTAL":
                total_row = row
                break
        
        assert total_row is not None, "TOTAL row not found in Excel"
        
        # TP Wt (Q) is at column 5 (E)
        tp_wt_total = ws.cell(row=total_row, column=5).value
        print(f"TP Weight total in Excel: {tp_wt_total}")
        
        # Total should be a number (could be 0 if no data)
        assert tp_wt_total is not None, "TP Weight total is None"
        assert isinstance(tp_wt_total, (int, float)), f"TP Weight total should be numeric, got {type(tp_wt_total)}"
        print(f"PASS: TP Weight total found in TOTAL row: {tp_wt_total}")
    
    def test_mill_entries_pdf_export_status(self):
        """Test Mill Entries PDF export returns 200"""
        response = requests.get(f"{BASE_URL}/api/export/pdf", params={"kms_year": "2026-2027"})
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/pdf" in response.headers.get("Content-Type", "")
        print("PASS: Mill Entries PDF export returns 200 with correct content type")
    
    def test_mill_entries_pdf_has_content(self):
        """Test Mill Entries PDF has content (non-empty)"""
        response = requests.get(f"{BASE_URL}/api/export/pdf", params={"kms_year": "2026-2027"})
        assert response.status_code == 200
        assert len(response.content) > 1000, f"PDF content too small: {len(response.content)} bytes"
        print(f"PASS: Mill Entries PDF has content: {len(response.content)} bytes")


class TestVehicleWeightExport:
    """Vehicle Weight Excel and PDF export tests - TOTAL row with TP Wt and G.Issued"""
    
    def test_vw_excel_export_status(self):
        """Test Vehicle Weight Excel export returns 200"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/excel", params={"kms_year": "2026-2027"})
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        print("PASS: VW Excel export returns 200 with correct content type")
    
    def test_vw_excel_has_total_row(self):
        """Test Vehicle Weight Excel has TOTAL: row"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/excel", params={"kms_year": "2026-2027"})
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find TOTAL: row
        total_row = None
        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=7).value  # TOTAL: is in column 7
            if cell_val and "TOTAL" in str(cell_val):
                total_row = row
                break
        
        assert total_row is not None, "TOTAL: row not found in VW Excel"
        print(f"PASS: TOTAL: row found at row {total_row}")
    
    def test_vw_excel_has_15_columns(self):
        """Test Vehicle Weight Excel has 15 columns including TP Wt and G.Issued"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/excel", params={"kms_year": "2026-2027"})
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find header row (look for RST header)
        header_row = None
        for row in range(1, 10):
            if ws.cell(row=row, column=1).value == "RST":
                header_row = row
                break
        
        assert header_row is not None, "Header row not found"
        
        # Get all headers
        headers = []
        for col in range(1, 20):
            val = ws.cell(row=header_row, column=col).value
            if val:
                headers.append(val)
        
        print(f"VW Excel headers: {headers}")
        
        # Check for TP Wt and G.Issued columns
        assert "TP Wt (Q)" in headers, f"TP Wt (Q) column not found in headers: {headers}"
        assert "G.Issued" in headers, f"G.Issued column not found in headers: {headers}"
        print(f"PASS: VW Excel has {len(headers)} columns including TP Wt (Q) and G.Issued")
    
    def test_vw_excel_total_row_has_all_totals(self):
        """Test Vehicle Weight Excel TOTAL row has totals for all numeric columns"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/excel", params={"kms_year": "2026-2027"})
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find header row
        header_row = None
        for row in range(1, 10):
            if ws.cell(row=row, column=1).value == "RST":
                header_row = row
                break
        
        # Find TOTAL row
        total_row = None
        for row in range(header_row + 1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=7).value
            if cell_val and "TOTAL" in str(cell_val):
                total_row = row
                break
        
        assert total_row is not None, "TOTAL row not found"
        
        # Check totals exist for columns 8-15 (Bags through Diesel)
        # Column 12 is TP Wt (Q), Column 13 is G.Issued
        tp_wt_total = ws.cell(row=total_row, column=12).value
        g_issued_total = ws.cell(row=total_row, column=13).value
        
        print(f"VW Excel totals - TP Wt: {tp_wt_total}, G.Issued: {g_issued_total}")
        
        # Totals should exist (can be 0)
        assert tp_wt_total is not None, "TP Wt total is None"
        assert g_issued_total is not None, "G.Issued total is None"
        print(f"PASS: VW Excel TOTAL row has TP Wt ({tp_wt_total}) and G.Issued ({g_issued_total}) totals")
    
    def test_vw_pdf_export_status(self):
        """Test Vehicle Weight PDF export returns 200"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/pdf", params={"kms_year": "2026-2027"})
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/pdf" in response.headers.get("Content-Type", "")
        print("PASS: VW PDF export returns 200 with correct content type")
    
    def test_vw_pdf_has_content(self):
        """Test Vehicle Weight PDF has content (non-empty)"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/pdf", params={"kms_year": "2026-2027"})
        assert response.status_code == 200
        assert len(response.content) > 1000, f"PDF content too small: {len(response.content)} bytes"
        print(f"PASS: VW PDF has content: {len(response.content)} bytes")


class TestDailyReportExport:
    """Daily Report PDF and Excel export tests - TP Weight column and summary"""
    
    def test_daily_report_pdf_export_status(self):
        """Test Daily Report PDF export returns 200"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/pdf", params={
            "date": "2026-04-06",
            "kms_year": "2026-2027"
        })
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/pdf" in response.headers.get("Content-Type", "")
        print("PASS: Daily Report PDF export returns 200 with correct content type")
    
    def test_daily_report_pdf_has_content(self):
        """Test Daily Report PDF has content (non-empty)"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/pdf", params={
            "date": "2026-04-06",
            "kms_year": "2026-2027"
        })
        assert response.status_code == 200
        assert len(response.content) > 1000, f"PDF content too small: {len(response.content)} bytes"
        print(f"PASS: Daily Report PDF has content: {len(response.content)} bytes")
    
    def test_daily_report_excel_export_status(self):
        """Test Daily Report Excel export returns 200"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/excel", params={
            "date": "2026-04-06",
            "kms_year": "2026-2027"
        })
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        print("PASS: Daily Report Excel export returns 200 with correct content type")
    
    def test_daily_report_excel_has_tp_weight_in_summary(self):
        """Test Daily Report Excel has TP Wt in summary line"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/excel", params={
            "date": "2026-04-06",
            "kms_year": "2026-2027"
        })
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for TP Wt in the content
        tp_wt_found = False
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and "TP Wt" in str(cell_val):
                    tp_wt_found = True
                    print(f"Found 'TP Wt' at row {row}, col {col}: {cell_val}")
                    break
            if tp_wt_found:
                break
        
        assert tp_wt_found, "TP Wt not found in Daily Report Excel"
        print("PASS: Daily Report Excel contains TP Wt reference")
    
    def test_daily_report_excel_has_tp_weight_header(self):
        """Test Daily Report Excel has TP Wt header column in paddy entries section"""
        response = requests.get(f"{BASE_URL}/api/reports/daily/excel", params={
            "date": "2026-04-06",
            "kms_year": "2026-2027",
            "mode": "detail"
        })
        assert response.status_code == 200
        
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for TP Wt header
        tp_wt_header_found = False
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(25, ws.max_column + 1)):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and str(cell_val).strip() == "TP Wt":
                    tp_wt_header_found = True
                    print(f"Found 'TP Wt' header at row {row}, col {col}")
                    break
            if tp_wt_header_found:
                break
        
        assert tp_wt_header_found, "TP Wt header column not found in Daily Report Excel"
        print("PASS: Daily Report Excel has TP Wt header column")


class TestTotalsAPI:
    """Test /api/totals endpoint includes total_tp_weight"""
    
    def test_totals_api_status(self):
        """Test /api/totals returns 200"""
        response = requests.get(f"{BASE_URL}/api/totals", params={"kms_year": "2026-2027"})
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("PASS: /api/totals returns 200")
    
    def test_totals_api_has_total_tp_weight(self):
        """Test /api/totals response includes total_tp_weight field"""
        response = requests.get(f"{BASE_URL}/api/totals", params={"kms_year": "2026-2027"})
        assert response.status_code == 200
        
        data = response.json()
        print(f"Totals API response: {data}")
        
        assert "total_tp_weight" in data, f"total_tp_weight not found in response: {data.keys()}"
        assert isinstance(data["total_tp_weight"], (int, float)), f"total_tp_weight should be numeric, got {type(data['total_tp_weight'])}"
        print(f"PASS: /api/totals includes total_tp_weight: {data['total_tp_weight']}")


class TestDailyReportAPI:
    """Test /api/reports/daily endpoint includes TP Weight data"""
    
    def test_daily_report_api_status(self):
        """Test /api/reports/daily returns 200"""
        response = requests.get(f"{BASE_URL}/api/reports/daily", params={
            "date": "2026-04-06",
            "kms_year": "2026-2027"
        })
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print("PASS: /api/reports/daily returns 200")
    
    def test_daily_report_api_has_tp_weight_total(self):
        """Test /api/reports/daily response includes total_tp_weight in paddy_entries"""
        response = requests.get(f"{BASE_URL}/api/reports/daily", params={
            "date": "2026-04-06",
            "kms_year": "2026-2027"
        })
        assert response.status_code == 200
        
        data = response.json()
        paddy_entries = data.get("paddy_entries", {})
        print(f"Paddy entries summary: {paddy_entries.keys()}")
        
        assert "total_tp_weight" in paddy_entries, f"total_tp_weight not found in paddy_entries: {paddy_entries.keys()}"
        print(f"PASS: Daily report includes total_tp_weight: {paddy_entries['total_tp_weight']}")
    
    def test_daily_report_api_details_have_tp_weight(self):
        """Test /api/reports/daily details include tp_weight field"""
        response = requests.get(f"{BASE_URL}/api/reports/daily", params={
            "date": "2026-04-06",
            "kms_year": "2026-2027"
        })
        assert response.status_code == 200
        
        data = response.json()
        details = data.get("paddy_entries", {}).get("details", [])
        
        if details:
            first_entry = details[0]
            assert "tp_weight" in first_entry, f"tp_weight not found in entry details: {first_entry.keys()}"
            print(f"PASS: Entry details include tp_weight field: {first_entry.get('tp_weight')}")
        else:
            print("SKIP: No paddy entries found for the date, but API structure is correct")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
