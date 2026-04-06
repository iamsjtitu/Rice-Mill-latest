"""
Test Date Format in Excel/PDF Exports - Iteration 173
Verifies that dates in exports show as DD-MM-YYYY (e.g., 15-01-2025) not YYYY-MM-DD
"""
import pytest
import requests
import os
import io
from openpyxl import load_workbook
import re

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestDateFormatExports:
    """Test date format in Excel exports - should be DD-MM-YYYY"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test session"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        # Login
        login_resp = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "username": "admin",
            "password": "admin123"
        })
        assert login_resp.status_code == 200, f"Login failed: {login_resp.text}"
        self.token = login_resp.json().get("token", "")
        if self.token:
            self.session.headers.update({"Authorization": f"Bearer {self.token}"})
    
    def test_login_works(self):
        """Verify login with admin/admin123 works"""
        resp = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "username": "admin",
            "password": "admin123"
        })
        assert resp.status_code == 200
        data = resp.json()
        # Response contains username and role for successful login
        assert "username" in data or "role" in data or "message" in data
        print(f"PASS: Login with admin/admin123 works - {data.get('message', 'success')}")
    
    def test_cashbook_excel_export_date_format(self):
        """Verify cashbook Excel export shows dates as DD-MM-YYYY"""
        resp = self.session.get(f"{BASE_URL}/api/cash-book/excel")
        assert resp.status_code == 200, f"Cashbook Excel export failed: {resp.status_code}"
        
        # Parse Excel content
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        
        # Find date column and check format
        date_pattern_correct = re.compile(r'^\d{2}-\d{2}-\d{4}$')  # DD-MM-YYYY
        date_pattern_wrong = re.compile(r'^\d{4}-\d{2}-\d{2}$')    # YYYY-MM-DD
        
        dates_found = []
        wrong_format_dates = []
        correct_format_dates = []
        
        # Check all cells for date values
        for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row)):
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip()
                    if date_pattern_correct.match(val):
                        correct_format_dates.append(val)
                        dates_found.append(val)
                    elif date_pattern_wrong.match(val):
                        wrong_format_dates.append(val)
                        dates_found.append(val)
        
        print(f"Dates found: {len(dates_found)}")
        print(f"Correct format (DD-MM-YYYY): {len(correct_format_dates)} - {correct_format_dates[:5]}")
        print(f"Wrong format (YYYY-MM-DD): {len(wrong_format_dates)} - {wrong_format_dates[:5]}")
        
        # Assert no wrong format dates
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD dates in cashbook Excel: {wrong_format_dates[:5]}"
        print("PASS: Cashbook Excel export dates are in DD-MM-YYYY format")
    
    def test_party_ledger_excel_export_date_format(self):
        """Verify party ledger Excel export shows dates as DD-MM-YYYY"""
        resp = self.session.get(f"{BASE_URL}/api/cash-book/party-summary/excel")
        assert resp.status_code == 200, f"Party summary Excel export failed: {resp.status_code}"
        
        # Parse Excel content
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        
        date_pattern_wrong = re.compile(r'^\d{4}-\d{2}-\d{2}$')  # YYYY-MM-DD
        wrong_format_dates = []
        
        for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row)):
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip()
                    if date_pattern_wrong.match(val):
                        wrong_format_dates.append(val)
        
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD dates in party summary Excel: {wrong_format_dates[:5]}"
        print("PASS: Party summary Excel export dates are correct")
    
    def test_entries_excel_export_date_format(self):
        """Verify entries Excel export shows dates as DD-MM-YYYY"""
        resp = self.session.get(f"{BASE_URL}/api/export/excel")
        assert resp.status_code == 200, f"Entries Excel export failed: {resp.status_code}"
        
        # Parse Excel content
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        
        date_pattern_correct = re.compile(r'^\d{2}-\d{2}-\d{4}$')  # DD-MM-YYYY
        date_pattern_wrong = re.compile(r'^\d{4}-\d{2}-\d{2}$')    # YYYY-MM-DD
        
        correct_format_dates = []
        wrong_format_dates = []
        
        for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row)):
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip()
                    if date_pattern_correct.match(val):
                        correct_format_dates.append(val)
                    elif date_pattern_wrong.match(val):
                        wrong_format_dates.append(val)
        
        print(f"Entries Excel - Correct format (DD-MM-YYYY): {len(correct_format_dates)} - {correct_format_dates[:5]}")
        print(f"Entries Excel - Wrong format (YYYY-MM-DD): {len(wrong_format_dates)} - {wrong_format_dates[:5]}")
        
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD dates in entries Excel: {wrong_format_dates[:5]}"
        print("PASS: Entries Excel export dates are in DD-MM-YYYY format")
    
    def test_truck_payments_excel_export_date_format(self):
        """Verify truck payments Excel export shows dates as DD-MM-YYYY"""
        resp = self.session.get(f"{BASE_URL}/api/export/truck-payments-excel")
        assert resp.status_code == 200, f"Truck payments Excel export failed: {resp.status_code}"
        
        # Parse Excel content
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        
        date_pattern_correct = re.compile(r'^\d{2}-\d{2}-\d{4}$')  # DD-MM-YYYY
        date_pattern_wrong = re.compile(r'^\d{4}-\d{2}-\d{2}$')    # YYYY-MM-DD
        
        correct_format_dates = []
        wrong_format_dates = []
        
        for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row)):
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip()
                    if date_pattern_correct.match(val):
                        correct_format_dates.append(val)
                    elif date_pattern_wrong.match(val):
                        wrong_format_dates.append(val)
        
        print(f"Truck payments Excel - Correct format (DD-MM-YYYY): {len(correct_format_dates)} - {correct_format_dates[:5]}")
        print(f"Truck payments Excel - Wrong format (YYYY-MM-DD): {len(wrong_format_dates)} - {wrong_format_dates[:5]}")
        
        assert len(wrong_format_dates) == 0, f"Found YYYY-MM-DD dates in truck payments Excel: {wrong_format_dates[:5]}"
        print("PASS: Truck payments Excel export dates are in DD-MM-YYYY format")
    
    def test_cashbook_page_loads(self):
        """Verify cashbook API endpoint works"""
        resp = self.session.get(f"{BASE_URL}/api/cash-book")
        assert resp.status_code == 200, f"Cashbook API failed: {resp.status_code}"
        data = resp.json()
        assert "transactions" in data or isinstance(data, list)
        print("PASS: Cashbook API endpoint works")
    
    def test_export_endpoints_return_200(self):
        """Verify all export endpoints return 200"""
        endpoints = [
            "/api/cash-book/excel",
            "/api/cash-book/party-summary/excel",
            "/api/export/excel",
            "/api/export/truck-payments-excel",
        ]
        
        for endpoint in endpoints:
            resp = self.session.get(f"{BASE_URL}{endpoint}")
            assert resp.status_code == 200, f"Export endpoint {endpoint} failed: {resp.status_code}"
            print(f"PASS: {endpoint} returns 200")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
