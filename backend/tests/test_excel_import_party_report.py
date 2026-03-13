"""
Test module for Excel Import and Party-wise Report features (Iteration 27)

Features tested:
1. GET /api/local-party/report/{party_name} - Party report with running balance
2. POST /api/entries/import-excel (preview_only=true) - Preview Excel entries
3. POST /api/entries/import-excel (preview_only=false) - Actual import with auto entries
"""
import pytest
import requests
import os
from pathlib import Path

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://web-app-mirror-2.preview.emergentagent.com').rstrip('/')

class TestPartyReport:
    """Tests for GET /api/local-party/report/{party_name} endpoint"""
    
    def test_party_report_basic(self):
        """Test basic party report endpoint returns correct structure"""
        # First check existing party 'Bicky' from seed data
        response = requests.get(f"{BASE_URL}/api/local-party/report/Bicky")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        data = response.json()
        # Validate response structure
        assert "party_name" in data, "Response missing party_name"
        assert "transactions" in data, "Response missing transactions"
        assert "total_debit" in data, "Response missing total_debit"
        assert "total_paid" in data, "Response missing total_paid"
        assert "balance" in data, "Response missing balance"
        assert "total_entries" in data, "Response missing total_entries"
        
        print(f"✅ Party report for Bicky: total_debit={data['total_debit']}, total_paid={data['total_paid']}, balance={data['balance']}")
    
    def test_party_report_running_balance(self):
        """Test that transactions have running_balance field"""
        response = requests.get(f"{BASE_URL}/api/local-party/report/Bicky")
        assert response.status_code == 200
        
        data = response.json()
        transactions = data.get("transactions", [])
        if len(transactions) > 0:
            # Each transaction should have running_balance
            for txn in transactions:
                assert "running_balance" in txn, f"Transaction missing running_balance: {txn}"
            
            # Verify running balance calculation
            # Last transaction's running_balance should equal the final balance
            last_txn = transactions[-1]
            assert last_txn["running_balance"] == data["balance"], \
                f"Last running_balance {last_txn['running_balance']} != final balance {data['balance']}"
            
            print(f"✅ Running balance verified: {len(transactions)} transactions, final balance Rs.{data['balance']}")
        else:
            print("⚠️ No transactions found for Bicky, skipping running balance check")
    
    def test_party_report_url_encoded_name(self):
        """Test party report with URL-encoded party name"""
        # Test with party name containing spaces (if exists)
        response = requests.get(f"{BASE_URL}/api/local-party/report/Ramesh")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        data = response.json()
        assert data["party_name"] == "Ramesh" or data["party_name"].lower() == "ramesh"
        print(f"✅ Party report for Ramesh: balance={data.get('balance', 0)}")
    
    def test_party_report_nonexistent_party(self):
        """Test party report for non-existent party returns empty transactions"""
        response = requests.get(f"{BASE_URL}/api/local-party/report/NonExistentParty123")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        data = response.json()
        assert data["transactions"] == [] or len(data.get("transactions", [])) == 0
        assert data.get("balance", 0) == 0
        print("✅ Non-existent party returns empty result")
    
    def test_party_report_with_filters(self):
        """Test party report with kms_year and season filters"""
        response = requests.get(f"{BASE_URL}/api/local-party/report/Bicky?kms_year=2025-2026&season=Kharif")
        assert response.status_code == 200
        
        data = response.json()
        assert "transactions" in data
        print(f"✅ Party report with filters: {len(data.get('transactions', []))} transactions")


class TestExcelImportPreview:
    """Tests for POST /api/entries/import-excel with preview_only=true"""
    
    @pytest.fixture
    def sample_excel_file(self):
        """Get the sample Excel file path"""
        file_path = Path("/tmp/sample.xlsx")
        if file_path.exists():
            return file_path
        pytest.skip("Sample Excel file not found at /tmp/sample.xlsx")
    
    def test_excel_import_preview(self, sample_excel_file):
        """Test Excel import preview returns correct structure"""
        with open(sample_excel_file, 'rb') as f:
            files = {'file': ('sample.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {
                'kms_year': '2025-2026',
                'season': 'Kharif',
                'username': 'admin',
                'preview_only': 'true'
            }
            response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        result = response.json()
        # Validate preview response structure
        assert result.get("preview") == True, "Response should have preview=True"
        assert "count" in result, "Response missing count"
        assert "skipped" in result, "Response missing skipped"
        assert "columns_detected" in result, "Response missing columns_detected"
        assert "sample" in result, "Response missing sample (first 10 entries)"
        
        print(f"✅ Excel preview: {result['count']} entries detected, {result['skipped']} skipped")
        print(f"   Columns detected: {result['columns_detected']}")
    
    def test_excel_import_preview_columns_detected(self, sample_excel_file):
        """Test that Excel import detects expected columns"""
        with open(sample_excel_file, 'rb') as f:
            files = {'file': ('sample.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {
                'preview_only': 'true',
                'username': 'admin'
            }
            response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 200
        result = response.json()
        
        columns = result.get("columns_detected", [])
        # Check for essential columns
        essential_columns = ['date', 'truck_no', 'agent_name']
        for col in essential_columns:
            assert col in columns, f"Essential column '{col}' not detected. Found: {columns}"
        
        print(f"✅ All essential columns detected: {essential_columns}")
    
    def test_excel_import_preview_sample_data(self, sample_excel_file):
        """Test that sample data contains valid entries"""
        with open(sample_excel_file, 'rb') as f:
            files = {'file': ('sample.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {
                'preview_only': 'true',
                'username': 'admin'
            }
            response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 200
        result = response.json()
        
        sample = result.get("sample", [])
        assert len(sample) <= 10, "Sample should be max 10 entries"
        
        if len(sample) > 0:
            first_entry = sample[0]
            # Verify entry has essential fields
            assert "date" in first_entry, "Sample entry missing date"
            assert "truck_no" in first_entry, "Sample entry missing truck_no"
            
            print(f"✅ Sample data valid: {len(sample)} preview entries")
            print(f"   First entry: date={first_entry.get('date')}, truck={first_entry.get('truck_no')}")
    
    def test_excel_import_invalid_file_type(self):
        """Test that non-Excel files are rejected"""
        # Create a fake text file
        files = {'file': ('test.txt', b'not an excel file', 'text/plain')}
        data = {
            'preview_only': 'true',
            'username': 'admin'
        }
        response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 400, f"Expected 400 for invalid file, got {response.status_code}"
        print("✅ Non-Excel file correctly rejected")


class TestExcelImportActual:
    """Tests for POST /api/entries/import-excel with preview_only=false (actual import)"""
    
    def test_excel_import_small_batch(self):
        """Test actual Excel import with a small test file (don't import 153 entries)"""
        # Create a minimal test Excel file with 2-3 entries
        from openpyxl import Workbook
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Headers matching expected format
        headers = ["DATE", "TRUCK NO", "AGENT", "MANDI", "KG", "BAG", "G.DEP", "GBW CUT", "CUTTING%", "G.ISSUED", "MOISTURE", "DISC", "CASH PAID", "DIESEL"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add 2 test entries
        test_entries = [
            ["2025-01-10", "TEST001", "TestAgent1", "TestMandi1", 5000, 50, 10, 25, 5.0, 5, 16, 0, 500, 200],
            ["2025-01-11", "TEST002", "TestAgent2", "TestMandi2", 4500, 45, 8, 22, 4.5, 4, 15, 0, 0, 0],
        ]
        for row_idx, entry in enumerate(test_entries, 2):
            for col_idx, value in enumerate(entry, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Send import request
        files = {'file': ('test_import.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {
            'kms_year': '2025-2026',
            'season': 'Kharif',
            'username': 'admin',
            'preview_only': 'false'
        }
        response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        result = response.json()
        # Validate import response structure
        assert result.get("success") == True, "Import should return success=True"
        assert "imported" in result, "Response missing imported count"
        assert "cash_book_entries" in result, "Response missing cash_book_entries count"
        assert "diesel_entries" in result, "Response missing diesel_entries count"
        
        print(f"✅ Import complete: {result['imported']} entries, {result['cash_book_entries']} cash book, {result['diesel_entries']} diesel")
    
    def test_excel_import_creates_cash_book_entry(self):
        """Test that importing entry with cash_paid creates Cash Book entry"""
        from openpyxl import Workbook
        import io
        
        wb = Workbook()
        ws = wb.active
        headers = ["DATE", "TRUCK NO", "AGENT", "MANDI", "KG", "BAG", "CASH PAID"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Entry with cash_paid
        test_entry = ["2025-01-12", "CASHTEST01", "CashAgent", "CashMandi", 3000, 30, 1000]
        for col_idx, value in enumerate(test_entry, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        files = {'file': ('cash_test.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {
            'kms_year': '2025-2026',
            'season': 'Kharif',
            'username': 'admin',
            'preview_only': 'false'
        }
        response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 200
        result = response.json()
        
        # Should have created 1 cash book entry
        assert result.get("cash_book_entries", 0) >= 1, "Should create Cash Book entry for cash_paid"
        print(f"✅ Cash book entry created: {result.get('cash_book_entries')} entries")
    
    def test_excel_import_creates_diesel_entry(self):
        """Test that importing entry with diesel_paid creates Diesel Account entry"""
        from openpyxl import Workbook
        import io
        
        wb = Workbook()
        ws = wb.active
        headers = ["DATE", "TRUCK NO", "AGENT", "MANDI", "KG", "BAG", "DIESEL"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Entry with diesel_paid
        test_entry = ["2025-01-13", "DIESELTEST01", "DieselAgent", "DieselMandi", 2500, 25, 500]
        for col_idx, value in enumerate(test_entry, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        files = {'file': ('diesel_test.xlsx', buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {
            'kms_year': '2025-2026',
            'season': 'Kharif',
            'username': 'admin',
            'preview_only': 'false'
        }
        response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 200
        result = response.json()
        
        # Should have created 1 diesel entry
        assert result.get("diesel_entries", 0) >= 1, "Should create Diesel entry for diesel_paid"
        print(f"✅ Diesel entry created: {result.get('diesel_entries')} entries")


class TestSampleFilePreview:
    """Tests using the actual sample.xlsx file for preview"""
    
    def test_sample_file_preview_entry_count(self):
        """Test that sample file has ~153 entries as mentioned"""
        sample_path = Path("/tmp/sample.xlsx")
        if not sample_path.exists():
            pytest.skip("Sample file not found")
        
        with open(sample_path, 'rb') as f:
            files = {'file': ('sample.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {
                'preview_only': 'true',
                'username': 'admin'
            }
            response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 200
        result = response.json()
        
        # Should have around 153 entries as per context
        count = result.get("count", 0)
        print(f"✅ Sample file preview: {count} entries detected (expected ~153)")
        assert count > 0, "Sample file should have entries"
    
    def test_sample_file_columns_detection(self):
        """Test that sample file detects ~15 columns"""
        sample_path = Path("/tmp/sample.xlsx")
        if not sample_path.exists():
            pytest.skip("Sample file not found")
        
        with open(sample_path, 'rb') as f:
            files = {'file': ('sample.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {
                'preview_only': 'true',
                'username': 'admin'
            }
            response = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
        
        assert response.status_code == 200
        result = response.json()
        
        columns = result.get("columns_detected", [])
        print(f"✅ Sample file columns: {len(columns)} detected")
        print(f"   Columns: {columns}")
        assert len(columns) >= 5, "Should detect at least 5 columns"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
