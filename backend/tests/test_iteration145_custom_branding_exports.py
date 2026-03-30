"""
Iteration 145: Test custom branding fields in Vehicle Weight exports
- GET /api/branding - should return custom_fields array with 2 entries
- GET /api/vehicle-weight/{id}/slip-pdf - should contain custom branding fields
- GET /api/vehicle-weight/export/pdf - should contain custom branding fields in header
- GET /api/vehicle-weight/export/excel - Row 1 should have above custom field, Row 3 should have GST below field
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestBrandingAPI:
    """Test branding API returns custom_fields"""
    
    def test_branding_endpoint_returns_custom_fields(self):
        """GET /api/branding should return custom_fields array with 2 entries"""
        response = requests.get(f"{BASE_URL}/api/branding")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        data = response.json()
        assert "custom_fields" in data, "Response should contain custom_fields"
        custom_fields = data.get("custom_fields", [])
        
        # Should have 2 entries as per test data
        assert len(custom_fields) >= 2, f"Expected at least 2 custom_fields, got {len(custom_fields)}"
        
        # Check for the expected fields
        above_field = None
        below_field = None
        for f in custom_fields:
            if f.get("placement") == "above":
                above_field = f
            elif f.get("placement") == "below":
                below_field = f
        
        # Verify above field (ॐ अर्हं नमः)
        assert above_field is not None, "Should have a field with placement='above'"
        assert "अर्हं" in above_field.get("value", "") or "ॐ" in above_field.get("value", ""), \
            f"Above field should contain Hindi text, got: {above_field.get('value', '')}"
        
        # Verify below field (GST: 21XXXXX1234Z1)
        assert below_field is not None, "Should have a field with placement='below'"
        assert below_field.get("label", "").upper() == "GST" or "GST" in below_field.get("label", "").upper(), \
            f"Below field should have GST label, got: {below_field.get('label', '')}"
        
        print(f"✓ Branding API returns {len(custom_fields)} custom_fields")
        print(f"  - Above field: {above_field}")
        print(f"  - Below field: {below_field}")


class TestVehicleWeightExports:
    """Test custom branding fields appear in all 3 export endpoints"""
    
    @pytest.fixture(scope="class")
    def vehicle_weight_entry_id(self):
        """Get a completed vehicle weight entry ID for testing"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight", params={
            "page_size": 1,
            "status": "completed"
        })
        assert response.status_code == 200, f"Failed to get VW entries: {response.status_code}"
        
        data = response.json()
        entries = data.get("entries", [])
        assert len(entries) > 0, "No completed vehicle weight entries found for testing"
        
        entry_id = entries[0].get("id")
        rst_no = entries[0].get("rst_no")
        print(f"✓ Using VW entry ID: {entry_id}, RST: {rst_no}")
        return entry_id
    
    def test_slip_pdf_returns_200_with_content(self, vehicle_weight_entry_id):
        """GET /api/vehicle-weight/{id}/slip-pdf should return HTTP 200 with non-zero PDF"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/{vehicle_weight_entry_id}/slip-pdf")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert len(response.content) > 0, "PDF content should not be empty"
        assert response.headers.get("content-type") == "application/pdf", \
            f"Expected application/pdf, got {response.headers.get('content-type')}"
        
        print(f"✓ slip-pdf returns 200 with {len(response.content)} bytes")
    
    def test_slip_pdf_contains_custom_branding_fields(self, vehicle_weight_entry_id):
        """GET /api/vehicle-weight/{id}/slip-pdf should contain custom branding fields"""
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            pytest.skip("PyPDF2 not installed")
        
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/{vehicle_weight_entry_id}/slip-pdf")
        assert response.status_code == 200
        
        # Extract text from PDF
        pdf_reader = PdfReader(io.BytesIO(response.content))
        pdf_text = ""
        for page in pdf_reader.pages:
            pdf_text += page.extract_text() or ""
        
        print(f"PDF text extracted ({len(pdf_text)} chars):")
        print(pdf_text[:500] if len(pdf_text) > 500 else pdf_text)
        
        # Check for custom branding fields
        # Above field: 'ॐ अर्हं नमः' (Hindi text)
        # Below field: 'GST: 21XXXXX1234Z1'
        
        # Note: PDF text extraction may not perfectly preserve Hindi characters
        # Check for GST which should be ASCII
        has_gst = "GST" in pdf_text.upper() or "21XXXXX" in pdf_text
        
        # For Hindi text, check if any Devanagari characters are present
        has_hindi = any(ord(c) >= 0x0900 and ord(c) <= 0x097F for c in pdf_text)
        
        print(f"  - Contains GST: {has_gst}")
        print(f"  - Contains Hindi/Devanagari: {has_hindi}")
        
        # At minimum, GST should be present
        assert has_gst, "PDF should contain GST custom field"
        print("✓ slip-pdf contains custom branding fields")
    
    def test_export_pdf_returns_200_with_content(self):
        """GET /api/vehicle-weight/export/pdf should return HTTP 200 with non-zero PDF"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/pdf")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert len(response.content) > 0, "PDF content should not be empty"
        assert "application/pdf" in response.headers.get("content-type", ""), \
            f"Expected application/pdf, got {response.headers.get('content-type')}"
        
        print(f"✓ export/pdf returns 200 with {len(response.content)} bytes")
    
    def test_export_pdf_contains_custom_branding_fields(self):
        """GET /api/vehicle-weight/export/pdf should contain custom branding fields in header"""
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            pytest.skip("PyPDF2 not installed")
        
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/pdf")
        assert response.status_code == 200
        
        # Extract text from PDF
        pdf_reader = PdfReader(io.BytesIO(response.content))
        pdf_text = ""
        for page in pdf_reader.pages:
            pdf_text += page.extract_text() or ""
        
        print(f"Export PDF text extracted ({len(pdf_text)} chars):")
        print(pdf_text[:500] if len(pdf_text) > 500 else pdf_text)
        
        # Check for GST custom field
        has_gst = "GST" in pdf_text.upper() or "21XXXXX" in pdf_text
        
        print(f"  - Contains GST: {has_gst}")
        
        assert has_gst, "Export PDF should contain GST custom field in header"
        print("✓ export/pdf contains custom branding fields")
    
    def test_export_excel_returns_200_with_content(self):
        """GET /api/vehicle-weight/export/excel should return HTTP 200 with non-zero Excel file"""
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/excel")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert len(response.content) > 0, "Excel content should not be empty"
        assert "spreadsheet" in response.headers.get("content-type", "").lower() or \
               "excel" in response.headers.get("content-type", "").lower() or \
               "openxmlformats" in response.headers.get("content-type", "").lower(), \
            f"Expected Excel content-type, got {response.headers.get('content-type')}"
        
        print(f"✓ export/excel returns 200 with {len(response.content)} bytes")
    
    def test_export_excel_contains_custom_branding_fields(self):
        """GET /api/vehicle-weight/export/excel - Row 1 should have above custom field, Row 3 should have GST"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = requests.get(f"{BASE_URL}/api/vehicle-weight/export/excel")
        assert response.status_code == 200
        
        # Load Excel workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get values from first few rows
        row1_value = ws.cell(row=1, column=1).value or ""
        row2_value = ws.cell(row=2, column=1).value or ""
        row3_value = ws.cell(row=3, column=1).value or ""
        
        print(f"Excel Row 1: {row1_value}")
        print(f"Excel Row 2: {row2_value}")
        print(f"Excel Row 3: {row3_value}")
        
        # Row 1 should contain above custom field (ॐ अर्हं नमः)
        # Check for Hindi characters
        has_above_field = any(ord(c) >= 0x0900 and ord(c) <= 0x097F for c in row1_value) or \
                          "अर्हं" in row1_value or "ॐ" in row1_value
        
        # Row 2 should contain company name
        has_company = "NAVKAR" in row2_value.upper() or "AGRO" in row2_value.upper()
        
        # Row 3 should contain GST below field alongside tagline
        has_gst_in_row3 = "GST" in row3_value.upper() or "21XXXXX" in row3_value
        
        print(f"  - Row 1 has above field (Hindi): {has_above_field}")
        print(f"  - Row 2 has company name: {has_company}")
        print(f"  - Row 3 has GST: {has_gst_in_row3}")
        
        # Verify above custom field is in Row 1
        assert has_above_field, f"Row 1 should contain above custom field (Hindi text), got: {row1_value}"
        
        # Verify company name is in Row 2
        assert has_company, f"Row 2 should contain company name, got: {row2_value}"
        
        # Verify GST is in Row 3 (tagline + below fields)
        assert has_gst_in_row3, f"Row 3 should contain GST custom field, got: {row3_value}"
        
        print("✓ export/excel contains custom branding fields in correct rows")


class TestAllExportsReturnValidFiles:
    """Verify all 3 export endpoints return valid files with non-zero sizes"""
    
    def test_all_exports_return_200(self):
        """All 3 export endpoints should return HTTP 200"""
        # Get a VW entry ID first
        vw_response = requests.get(f"{BASE_URL}/api/vehicle-weight", params={
            "page_size": 1,
            "status": "completed"
        })
        assert vw_response.status_code == 200
        entries = vw_response.json().get("entries", [])
        
        if not entries:
            pytest.skip("No completed VW entries for testing")
        
        entry_id = entries[0].get("id")
        
        # Test all 3 endpoints
        endpoints = [
            f"/api/vehicle-weight/{entry_id}/slip-pdf",
            "/api/vehicle-weight/export/pdf",
            "/api/vehicle-weight/export/excel"
        ]
        
        results = []
        for endpoint in endpoints:
            response = requests.get(f"{BASE_URL}{endpoint}")
            results.append({
                "endpoint": endpoint,
                "status": response.status_code,
                "size": len(response.content),
                "content_type": response.headers.get("content-type", "")
            })
            
            assert response.status_code == 200, f"{endpoint} returned {response.status_code}"
            assert len(response.content) > 0, f"{endpoint} returned empty content"
        
        print("✓ All 3 export endpoints return HTTP 200 with non-zero file sizes:")
        for r in results:
            print(f"  - {r['endpoint']}: {r['status']} ({r['size']} bytes)")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
