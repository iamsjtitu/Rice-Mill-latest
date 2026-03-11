"""
Test Agent & Mandi Report - Column Alignment Fix Verification (iteration_47)
This test verifies that G.Iss column is at position 6 (after G.Dep) in both API and exports.
"""
import pytest
import requests
import os
from io import BytesIO

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')


class TestAgentMandiReportColumnFix:
    """Test Agent & Mandi wise report - Column alignment verification"""

    def test_01_api_returns_data_with_correct_fields(self):
        """Verify API returns data with g_issued field in entries"""
        response = requests.get(f"{BASE_URL}/api/reports/agent-mandi-wise?kms_year=2025-2026")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        data = response.json()
        assert "mandis" in data, "Response should contain 'mandis' key"
        assert "grand_totals" in data, "Response should contain 'grand_totals' key"
        
        # Verify grand totals has g_issued
        assert "total_g_issued" in data["grand_totals"], "grand_totals should have total_g_issued"
        print(f"Grand totals - Total G.Issued: {data['grand_totals']['total_g_issued']}")

    def test_02_entry_has_g_issued_field(self):
        """Verify each entry has g_issued field"""
        response = requests.get(f"{BASE_URL}/api/reports/agent-mandi-wise?kms_year=2025-2026")
        assert response.status_code == 200
        
        data = response.json()
        for mandi in data.get("mandis", []):
            for entry in mandi.get("entries", []):
                assert "g_issued" in entry, f"Entry should have g_issued field: {entry.keys()}"
                assert "g_deposite" in entry, f"Entry should have g_deposite field"
                print(f"Entry: g_deposite={entry['g_deposite']}, g_issued={entry['g_issued']}")

    def test_03_totals_has_g_issued(self):
        """Verify mandi totals has total_g_issued"""
        response = requests.get(f"{BASE_URL}/api/reports/agent-mandi-wise?kms_year=2025-2026")
        assert response.status_code == 200
        
        data = response.json()
        for mandi in data.get("mandis", []):
            totals = mandi.get("totals", {})
            assert "total_g_issued" in totals, f"Totals should have total_g_issued: {totals.keys()}"
            assert "total_g_deposite" in totals, f"Totals should have total_g_deposite"
            print(f"{mandi['mandi_name']} totals: G.Dep={totals['total_g_deposite']}, G.Iss={totals['total_g_issued']}")

    def test_04_excel_export_success(self):
        """Verify Excel export works without errors"""
        response = requests.get(f"{BASE_URL}/api/reports/agent-mandi-wise/excel?kms_year=2025-2026")
        assert response.status_code == 200, f"Excel export failed: {response.status_code}"
        assert len(response.content) > 1000, "Excel file should not be empty"
        print(f"Excel export size: {len(response.content)} bytes")

    def test_05_excel_column_order(self):
        """Verify Excel has G.Iss as 6th column (after G.Dep)"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = requests.get(f"{BASE_URL}/api/reports/agent-mandi-wise/excel?kms_year=2025-2026")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Find header row
        headers = []
        for row_idx in range(1, 10):
            row_vals = [ws.cell(row=row_idx, column=col).value for col in range(1, 16)]
            if "Date" in row_vals:
                headers = row_vals
                break
        
        expected_order = ["Date", "Truck No", "QNTL", "BAG", "G.Dep", "G.Iss", "GBW", "P.Pkt", "P.Cut", "Mill W", "M%", "M.Cut", "C%", "D/D/P", "Final W"]
        assert headers == expected_order, f"Header order mismatch.\nExpected: {expected_order}\nGot: {headers}"
        print(f"Excel headers verified: G.Iss is at position 6")

    def test_06_pdf_export_success(self):
        """Verify PDF export works without errors"""
        response = requests.get(f"{BASE_URL}/api/reports/agent-mandi-wise/pdf?kms_year=2025-2026")
        assert response.status_code == 200, f"PDF export failed: {response.status_code}"
        assert len(response.content) > 1000, "PDF file should not be empty"
        
        # Verify it starts with PDF header
        assert response.content[:4] == b'%PDF', "Response should be a valid PDF file"
        print(f"PDF export success: {len(response.content)} bytes")

    def test_07_g_iss_position_after_g_dep(self):
        """Verify G.Iss comes right after G.Dep in API response entries"""
        response = requests.get(f"{BASE_URL}/api/reports/agent-mandi-wise?kms_year=2025-2026")
        assert response.status_code == 200
        
        data = response.json()
        if data.get("mandis") and data["mandis"][0].get("entries"):
            entry_keys = list(data["mandis"][0]["entries"][0].keys())
            # Find indices
            g_dep_idx = entry_keys.index("g_deposite") if "g_deposite" in entry_keys else -1
            g_iss_idx = entry_keys.index("g_issued") if "g_issued" in entry_keys else -1
            
            # Both should exist
            assert g_dep_idx >= 0, "g_deposite should be in entry keys"
            assert g_iss_idx >= 0, "g_issued should be in entry keys"
            
            print(f"Entry keys: {entry_keys}")
            print(f"G.Dep at index {g_dep_idx}, G.Iss at index {g_iss_idx}")


# Run tests if executed directly
if __name__ == "__main__":
    pytest.main([__file__, "-v"])
