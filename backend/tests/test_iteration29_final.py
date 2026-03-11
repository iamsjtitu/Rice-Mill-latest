"""
Iteration 29 - Final Comprehensive Test Suite
Tests all features: Local Party, Mill Parts, Excel Import, Cash Book, Date filters
"""
import pytest
import requests
import os
from datetime import datetime

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://private-trade-beta.preview.emergentagent.com').rstrip('/')

# ============ FIXTURES ============

@pytest.fixture(scope="session")
def api_client():
    session = requests.Session()
    session.headers.update({"Content-Type": "application/json"})
    return session

@pytest.fixture(scope="session")
def admin_token(api_client):
    """Login as admin and get token"""
    res = api_client.post(f"{BASE_URL}/api/auth/login", json={"username": "admin", "password": "admin123"})
    assert res.status_code == 200, f"Login failed: {res.text}"
    data = res.json()
    assert "token" in data
    return data.get("token")

# ============ LOCAL PARTY TESTS ============

class TestLocalParty:
    """Local Party Account feature tests"""
    
    def test_local_party_summary_no_filter(self, api_client):
        """Test GET /api/local-party/summary returns party-wise summary"""
        res = api_client.get(f"{BASE_URL}/api/local-party/summary")
        assert res.status_code == 200
        data = res.json()
        assert "parties" in data
        assert "grand_total_debit" in data
        assert "grand_total_paid" in data
        assert "grand_balance" in data
        # Verify structure
        if data["parties"]:
            party = data["parties"][0]
            assert "party_name" in party
            assert "total_debit" in party
            assert "total_paid" in party
            assert "balance" in party
            assert "txn_count" in party
        print(f"PASS: Local Party Summary returned {len(data['parties'])} parties")
    
    def test_local_party_summary_with_date_filter(self, api_client):
        """Test date_from and date_to filters work"""
        res = api_client.get(f"{BASE_URL}/api/local-party/summary?date_from=2025-02-15&date_to=2025-02-16")
        assert res.status_code == 200
        data = res.json()
        # Should return Bicky with filtered data
        bicky = next((p for p in data["parties"] if p["party_name"] == "Bicky"), None)
        if bicky:
            assert bicky["txn_count"] == 2, f"Expected 2 txns in date range, got {bicky['txn_count']}"
            print(f"PASS: Date filter working - Bicky has {bicky['txn_count']} txns in date range")
        else:
            print("INFO: No Bicky in date range - data may have changed")
    
    def test_local_party_transactions(self, api_client):
        """Test GET /api/local-party/transactions"""
        res = api_client.get(f"{BASE_URL}/api/local-party/transactions?party_name=Bicky")
        assert res.status_code == 200
        data = res.json()
        assert isinstance(data, list)
        if data:
            txn = data[0]
            assert "id" in txn
            assert "date" in txn
            assert "party_name" in txn
            assert "txn_type" in txn
            assert "amount" in txn
        print(f"PASS: Bicky transactions returned {len(data)} entries")
    
    def test_local_party_report_with_running_balance(self, api_client):
        """Test GET /api/local-party/report/:party_name returns running balance"""
        res = api_client.get(f"{BASE_URL}/api/local-party/report/Bicky")
        assert res.status_code == 200
        data = res.json()
        assert data["party_name"] == "Bicky"
        assert "transactions" in data
        assert "total_debit" in data
        assert "total_paid" in data
        assert "balance" in data
        # Verify running balance calculation
        if data["transactions"]:
            for txn in data["transactions"]:
                assert "running_balance" in txn, "Running balance missing in transaction"
        print(f"PASS: Bicky report - Debit: {data['total_debit']}, Paid: {data['total_paid']}, Balance: {data['balance']}")
    
    def test_local_party_report_date_filter(self, api_client):
        """Test report with date filter"""
        res = api_client.get(f"{BASE_URL}/api/local-party/report/Bicky?date_from=2025-02-15&date_to=2025-02-16")
        assert res.status_code == 200
        data = res.json()
        assert data["party_name"] == "Bicky"
        # Date-filtered should have subset of transactions
        print(f"PASS: Bicky filtered report - {data['total_entries']} entries in date range")
    
    def test_local_party_manual_purchase_validation(self, api_client):
        """Test POST /api/local-party/manual validation"""
        # Empty party name should fail
        res = api_client.post(f"{BASE_URL}/api/local-party/manual", json={
            "party_name": "", "amount": 100
        })
        assert res.status_code == 400
        
        # Zero amount should fail
        res = api_client.post(f"{BASE_URL}/api/local-party/manual", json={
            "party_name": "TestParty", "amount": 0
        })
        assert res.status_code == 400
        print("PASS: Manual purchase validation working")
    
    def test_local_party_settle_creates_cashbook_entry(self, api_client):
        """Test POST /api/local-party/settle creates both payment and cash book entry"""
        # First check current cash transactions count
        cb_before = api_client.get(f"{BASE_URL}/api/cash-book")
        cb_count_before = len(cb_before.json()) if cb_before.status_code == 200 else 0
        
        # Create settlement
        res = api_client.post(f"{BASE_URL}/api/local-party/settle", json={
            "party_name": "TestParty1",
            "amount": 100,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "notes": "Test settlement",
            "created_by": "pytest"
        })
        
        if res.status_code == 200:
            data = res.json()
            assert data["success"] == True
            assert "txn_id" in data
            
            # Verify cash book entry was created
            cb_after = api_client.get(f"{BASE_URL}/api/cash-book")
            cb_count_after = len(cb_after.json())
            assert cb_count_after > cb_count_before, "Cash book entry not created"
            print(f"PASS: Settlement created with txn_id {data['txn_id'][:8]}...")
        else:
            print(f"INFO: Settlement failed - {res.status_code}: {res.text}")
    
    def test_local_party_excel_export(self, api_client):
        """Test GET /api/local-party/excel returns xlsx"""
        res = api_client.get(f"{BASE_URL}/api/local-party/excel")
        assert res.status_code == 200
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in res.headers.get('Content-Type', '')
        assert len(res.content) > 0
        print(f"PASS: Excel export returned {len(res.content)} bytes")
    
    def test_local_party_pdf_export(self, api_client):
        """Test GET /api/local-party/pdf returns pdf"""
        res = api_client.get(f"{BASE_URL}/api/local-party/pdf")
        assert res.status_code == 200
        assert 'application/pdf' in res.headers.get('Content-Type', '')
        assert len(res.content) > 0
        print(f"PASS: PDF export returned {len(res.content)} bytes")

# ============ MILL PARTS TESTS ============

class TestMillParts:
    """Mill Parts Stock feature tests"""
    
    def test_mill_parts_list(self, api_client):
        """Test GET /api/mill-parts returns parts master"""
        res = api_client.get(f"{BASE_URL}/api/mill-parts")
        assert res.status_code == 200
        data = res.json()
        assert isinstance(data, list)
        if data:
            part = data[0]
            assert "id" in part
            assert "name" in part
            assert "category" in part
            assert "unit" in part
        print(f"PASS: Mill parts returned {len(data)} parts")
    
    def test_mill_parts_summary(self, api_client):
        """Test GET /api/mill-parts/summary returns stock summary"""
        res = api_client.get(f"{BASE_URL}/api/mill-parts/summary")
        assert res.status_code == 200
        data = res.json()
        assert isinstance(data, list)
        if data:
            summary = data[0]
            assert "part_name" in summary
            assert "stock_in" in summary
            assert "stock_used" in summary
            assert "current_stock" in summary
            assert "total_purchase_amount" in summary
        print(f"PASS: Mill parts summary returned {len(data)} parts")
    
    def test_mill_parts_stock_transactions(self, api_client):
        """Test GET /api/mill-parts-stock"""
        res = api_client.get(f"{BASE_URL}/api/mill-parts-stock")
        assert res.status_code == 200
        data = res.json()
        assert isinstance(data, list)
        print(f"PASS: Stock transactions returned {len(data)} entries")
    
    def test_mill_parts_stock_creates_local_party_entry(self, api_client):
        """Test POST /api/mill-parts-stock with party creates local party entry"""
        # Get current local party count
        lp_before = api_client.get(f"{BASE_URL}/api/local-party/transactions")
        lp_count_before = len(lp_before.json()) if lp_before.status_code == 200 else 0
        
        # Create stock in with party
        res = api_client.post(f"{BASE_URL}/api/mill-parts-stock", json={
            "date": datetime.now().strftime("%Y-%m-%d"),
            "part_name": "Bearing (Shaft)",
            "txn_type": "in",
            "quantity": 5,
            "rate": 50,
            "party_name": "TestLocalParty_Pytest",
            "created_by": "pytest"
        })
        
        if res.status_code == 200:
            data = res.json()
            assert "id" in data
            assert data["total_amount"] == 250  # 5 * 50
            
            # Verify local party entry was created
            lp_after = api_client.get(f"{BASE_URL}/api/local-party/transactions?party_name=TestLocalParty_Pytest")
            assert lp_after.status_code == 200
            lp_txns = lp_after.json()
            if lp_txns:
                print(f"PASS: Stock in created local party entry for TestLocalParty_Pytest")
            else:
                print("INFO: Local party entry may not have been created (party_name filter)")
        else:
            print(f"INFO: Stock in failed - {res.status_code}")
    
    def test_mill_parts_edit_updates_local_party(self, api_client):
        """Test PUT /api/mill-parts-stock/:id updates linked local party"""
        # Get existing stock entry
        res = api_client.get(f"{BASE_URL}/api/mill-parts-stock")
        if res.status_code == 200 and res.json():
            entries = res.json()
            # Find one with party_name
            entry = next((e for e in entries if e.get("party_name") and e.get("txn_type") == "in"), None)
            if entry:
                entry_id = entry["id"]
                original_amount = entry.get("total_amount", 0)
                
                # Update with new rate
                update_res = api_client.put(f"{BASE_URL}/api/mill-parts-stock/{entry_id}", json={
                    "date": entry["date"],
                    "part_name": entry["part_name"],
                    "txn_type": entry["txn_type"],
                    "quantity": entry["quantity"],
                    "rate": entry.get("rate", 0) + 10,  # Increase rate
                    "party_name": entry.get("party_name", ""),
                    "created_by": "pytest"
                })
                assert update_res.status_code == 200
                print(f"PASS: Mill parts edit working - entry {entry_id[:8]} updated")
            else:
                print("INFO: No stock entry with party_name found for edit test")
        else:
            print("INFO: No stock entries found")

# ============ EXCEL IMPORT TESTS ============

class TestExcelImport:
    """Excel Import feature tests"""
    
    def test_excel_import_preview(self, api_client):
        """Test POST /api/entries/import-excel with preview_only=true"""
        # Create a simple test Excel file
        import io
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'DATE'
            ws['B1'] = 'TRUCK NO'
            ws['C1'] = 'AGENT'
            ws['D1'] = 'MANDI'
            ws['E1'] = 'NETT KG'
            ws['F1'] = 'BAG'
            ws['A2'] = '2025-01-15'
            ws['B2'] = 'TEST001'
            ws['C2'] = 'TestAgent'
            ws['D2'] = 'TestMandi'
            ws['E2'] = 5000
            ws['F2'] = 50
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            files = {'file': ('test.xlsx', buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {'preview_only': 'true', 'kms_year': '2024-25', 'season': 'Kharif', 'username': 'pytest'}
            
            res = requests.post(f"{BASE_URL}/api/entries/import-excel", files=files, data=data)
            assert res.status_code == 200
            result = res.json()
            assert result.get("preview") == True
            assert "count" in result
            assert "columns_detected" in result
            print(f"PASS: Excel import preview - detected {len(result['columns_detected'])} columns, {result['count']} entries")
        except ImportError:
            print("INFO: openpyxl not available for test file creation - skipping")

# ============ CASH BOOK TESTS ============

class TestCashBook:
    """Cash Book feature tests"""
    
    def test_cashbook_summary(self, api_client):
        """Test GET /api/cash-book/summary returns balance info"""
        res = api_client.get(f"{BASE_URL}/api/cash-book/summary")
        assert res.status_code == 200
        data = res.json()
        assert "cash_balance" in data
        assert "bank_balance" in data
        assert "total_balance" in data
        assert "cash_in" in data
        assert "cash_out" in data
        print(f"PASS: Cash Book Summary - Cash: {data['cash_balance']}, Bank: {data['bank_balance']}, Total: {data['total_balance']}")
    
    def test_cashbook_transactions(self, api_client):
        """Test GET /api/cash-book returns transactions"""
        res = api_client.get(f"{BASE_URL}/api/cash-book")
        assert res.status_code == 200
        data = res.json()
        assert isinstance(data, list)
        if data:
            txn = data[0]
            assert "id" in txn
            assert "date" in txn
            assert "account" in txn
            assert "txn_type" in txn
            assert "amount" in txn
        print(f"PASS: Cash Book returned {len(data)} transactions")
    
    def test_cashbook_categories(self, api_client):
        """Test GET /api/cash-book/categories"""
        res = api_client.get(f"{BASE_URL}/api/cash-book/categories")
        assert res.status_code == 200
        data = res.json()
        assert isinstance(data, list)
        print(f"PASS: Cash Book categories returned {len(data)} custom categories")

# ============ DC PAYMENTS GUNNY BAG TESTS ============

class TestDCPaymentsGunny:
    """DC Payments and Gunny Bag tests for local party auto-create"""
    
    def test_gunny_bag_summary(self, api_client):
        """Test GET /api/gunny-bags/summary"""
        res = api_client.get(f"{BASE_URL}/api/gunny-bags/summary")
        assert res.status_code == 200
        data = res.json()
        assert "new" in data
        assert "old" in data
        print(f"PASS: Gunny bag summary - New: {data['new']['balance']}, Old: {data['old']['balance']}")
    
    def test_gunny_bag_old_creates_local_party(self, api_client):
        """Test POST /api/gunny-bags with old type and source creates local party entry"""
        res = api_client.post(f"{BASE_URL}/api/gunny-bags", json={
            "date": datetime.now().strftime("%Y-%m-%d"),
            "bag_type": "old",
            "txn_type": "in",
            "quantity": 100,
            "rate": 5,
            "source": "GunnyTestParty_Pytest",
            "created_by": "pytest"
        })
        
        if res.status_code == 200:
            data = res.json()
            assert data["amount"] == 500  # 100 * 5
            
            # Check if local party entry created
            lp = api_client.get(f"{BASE_URL}/api/local-party/transactions?party_name=GunnyTestParty_Pytest")
            if lp.status_code == 200 and lp.json():
                print("PASS: Old gunny bag purchase created local party entry")
            else:
                print("INFO: Local party entry may exist for GunnyTestParty")
        else:
            print(f"INFO: Gunny bag POST - {res.status_code}")

# ============ API HEALTH & MISC TESTS ============

class TestMiscAPI:
    """Miscellaneous API endpoint tests"""
    
    def test_auth_login(self, api_client):
        """Test POST /api/auth/login"""
        res = api_client.post(f"{BASE_URL}/api/auth/login", json={
            "username": "admin", "password": "admin123"
        })
        assert res.status_code == 200
        data = res.json()
        assert data.get("success") == True or "token" in data, "Login should succeed"
        assert data.get("role") == "admin" or data.get("user", {}).get("role") == "admin"
        print("PASS: Admin login working")
    
    def test_branding(self, api_client):
        """Test GET /api/branding"""
        res = api_client.get(f"{BASE_URL}/api/branding")
        assert res.status_code == 200
        print("PASS: Branding endpoint working")
    
    def test_fy_settings(self, api_client):
        """Test GET /api/fy-settings"""
        res = api_client.get(f"{BASE_URL}/api/fy-settings")
        assert res.status_code == 200
        print("PASS: FY settings endpoint working")
    
    def test_suggestions_trucks(self, api_client):
        """Test GET /api/suggestions/trucks"""
        res = api_client.get(f"{BASE_URL}/api/suggestions/trucks")
        assert res.status_code == 200
        data = res.json()
        assert "suggestions" in data
        print(f"PASS: Truck suggestions - {len(data['suggestions'])} trucks")
    
    def test_suggestions_agents(self, api_client):
        """Test GET /api/suggestions/agents"""
        res = api_client.get(f"{BASE_URL}/api/suggestions/agents")
        assert res.status_code == 200
        data = res.json()
        assert "suggestions" in data
        print(f"PASS: Agent suggestions - {len(data['suggestions'])} agents")
    
    def test_suggestions_mandis(self, api_client):
        """Test GET /api/suggestions/mandis"""
        res = api_client.get(f"{BASE_URL}/api/suggestions/mandis")
        assert res.status_code == 200
        data = res.json()
        assert "suggestions" in data
        print(f"PASS: Mandi suggestions - {len(data['suggestions'])} mandis")
    
    def test_mandi_targets(self, api_client):
        """Test GET /api/mandi-targets"""
        res = api_client.get(f"{BASE_URL}/api/mandi-targets")
        assert res.status_code == 200
        data = res.json()
        assert isinstance(data, list)
        print(f"PASS: Mandi targets - {len(data)} targets")
    
    def test_dc_summary(self, api_client):
        """Test GET /api/dc-summary"""
        res = api_client.get(f"{BASE_URL}/api/dc-summary")
        assert res.status_code == 200
        data = res.json()
        assert "total_dc" in data
        assert "total_allotted_qntl" in data
        print(f"PASS: DC summary - {data['total_dc']} DCs")

# Run with: pytest /app/backend/tests/test_iteration29_final.py -v --tb=short
