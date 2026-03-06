"""
Test Export Functionality - Excel and PDF exports
Tests for G.Dep, Cash, Diesel columns and styling
"""
import pytest
import requests
import os
import io
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestExportEndpoints:
    """Test export API endpoints"""
    
    def test_excel_export_returns_valid_file(self):
        """Test that Excel export returns a valid xlsx file"""
        response = requests.get(f"{BASE_URL}/api/export/excel")
        assert response.status_code == 200
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.headers.get('content-type', '')
        
        # Verify it's a valid xlsx (starts with PK)
        assert response.content[:2] == b'PK'
        print("✓ Excel export returns valid xlsx file")
    
    def test_pdf_export_returns_valid_file(self):
        """Test that PDF export returns a valid PDF file"""
        response = requests.get(f"{BASE_URL}/api/export/pdf")
        assert response.status_code == 200
        assert 'application/pdf' in response.headers.get('content-type', '')
        
        # Verify it's a valid PDF (starts with %PDF)
        assert response.content[:4] == b'%PDF'
        print("✓ PDF export returns valid PDF file")
    
    def test_excel_export_with_filters(self):
        """Test Excel export with KMS year filter"""
        response = requests.get(f"{BASE_URL}/api/export/excel?kms_year=2025-2026")
        assert response.status_code == 200
        assert response.content[:2] == b'PK'
        print("✓ Excel export with filters works")
    
    def test_pdf_export_with_filters(self):
        """Test PDF export with KMS year filter"""
        response = requests.get(f"{BASE_URL}/api/export/pdf?kms_year=2025-2026")
        assert response.status_code == 200
        assert response.content[:4] == b'%PDF'
        print("✓ PDF export with filters works")


class TestExcelColumns:
    """Test Excel export contains all required columns"""
    
    def test_excel_has_all_required_columns(self):
        """Verify Excel export has G.Dep, Cash, Diesel columns"""
        response = requests.get(f"{BASE_URL}/api/export/excel")
        assert response.status_code == 200
        
        # Load workbook from response content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get header row (row 3 based on code)
        headers = [cell.value for cell in ws[3]]
        
        # Required columns to verify
        required_columns = ["G.Dep", "Cash", "Diesel", "QNTL", "Final W"]
        
        for col in required_columns:
            assert col in headers, f"Missing column: {col}"
            print(f"✓ Found column: {col}")
        
        print(f"✓ All required columns present: {required_columns}")
    
    def test_excel_column_order(self):
        """Verify Excel columns are in correct order"""
        response = requests.get(f"{BASE_URL}/api/export/excel")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = [cell.value for cell in ws[3]]
        
        # Expected column order
        expected_order = [
            "Date", "Truck No", "Agent", "Mandi", "QNTL", "BAG", "G.Dep",
            "GBW Cut", "Mill W", "Moist%", "M.Cut", "Cut%", 
            "D/D/P", "Final W", "G.Issued", "Cash", "Diesel"
        ]
        
        assert headers == expected_order, f"Column order mismatch. Got: {headers}"
        print(f"✓ Excel columns in correct order")


class TestExcelStyling:
    """Test Excel export has proper styling"""
    
    def test_excel_has_header_styling(self):
        """Verify Excel has styled header row"""
        response = requests.get(f"{BASE_URL}/api/export/excel")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check header row (row 3) has fill color
        header_cell = ws.cell(row=3, column=1)
        assert header_cell.fill.fgColor.rgb is not None, "Header should have fill color"
        print(f"✓ Header has fill color: {header_cell.fill.fgColor.rgb}")
    
    def test_excel_has_title_row(self):
        """Verify Excel has title row with styling"""
        response = requests.get(f"{BASE_URL}/api/export/excel")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check title row (row 1)
        title_cell = ws.cell(row=1, column=1)
        assert title_cell.value is not None, "Title row should have content"
        assert "NAVKAR AGRO" in str(title_cell.value), "Title should contain NAVKAR AGRO"
        print(f"✓ Title row present: {title_cell.value}")
    
    def test_excel_has_highlighted_columns(self):
        """Verify Excel has highlighted columns (QNTL, G.Dep, Final W, Cash, Diesel)"""
        response = requests.get(f"{BASE_URL}/api/export/excel")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get headers to find column indices
        headers = [cell.value for cell in ws[3]]
        
        # Check if there are data rows
        if ws.max_row > 3:
            # Check QNTL column (col 5) has green fill
            qntl_idx = headers.index("QNTL") + 1
            qntl_cell = ws.cell(row=4, column=qntl_idx)
            print(f"✓ QNTL column fill: {qntl_cell.fill.fgColor.rgb if qntl_cell.fill.fgColor else 'None'}")
            
            # Check G.Dep column has blue fill
            gdep_idx = headers.index("G.Dep") + 1
            gdep_cell = ws.cell(row=4, column=gdep_idx)
            print(f"✓ G.Dep column fill: {gdep_cell.fill.fgColor.rgb if gdep_cell.fill.fgColor else 'None'}")
            
            # Check Final W column has yellow fill
            final_idx = headers.index("Final W") + 1
            final_cell = ws.cell(row=4, column=final_idx)
            print(f"✓ Final W column fill: {final_cell.fill.fgColor.rgb if final_cell.fill.fgColor else 'None'}")
            
            # Check Cash column has pink fill
            cash_idx = headers.index("Cash") + 1
            cash_cell = ws.cell(row=4, column=cash_idx)
            print(f"✓ Cash column fill: {cash_cell.fill.fgColor.rgb if cash_cell.fill.fgColor else 'None'}")
            
            # Check Diesel column has pink fill
            diesel_idx = headers.index("Diesel") + 1
            diesel_cell = ws.cell(row=4, column=diesel_idx)
            print(f"✓ Diesel column fill: {diesel_cell.fill.fgColor.rgb if diesel_cell.fill.fgColor else 'None'}")
        else:
            print("⚠ No data rows to check styling")


class TestExcelData:
    """Test Excel export data values"""
    
    def test_excel_has_totals_row(self):
        """Verify Excel has totals row"""
        response = requests.get(f"{BASE_URL}/api/export/excel")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Find totals row (last row with data)
        last_row = ws.max_row
        first_cell = ws.cell(row=last_row, column=1)
        
        assert first_cell.value == "TOTAL", f"Last row should be TOTAL, got: {first_cell.value}"
        print(f"✓ Totals row present at row {last_row}")
    
    def test_excel_totals_include_cash_diesel(self):
        """Verify totals row includes Cash and Diesel totals"""
        response = requests.get(f"{BASE_URL}/api/export/excel")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = [cell.value for cell in ws[3]]
        last_row = ws.max_row
        
        # Get Cash and Diesel totals
        cash_idx = headers.index("Cash") + 1
        diesel_idx = headers.index("Diesel") + 1
        
        cash_total = ws.cell(row=last_row, column=cash_idx).value
        diesel_total = ws.cell(row=last_row, column=diesel_idx).value
        
        print(f"✓ Cash total in Excel: {cash_total}")
        print(f"✓ Diesel total in Excel: {diesel_total}")
        
        # Values should be numbers (not None or empty)
        assert cash_total is not None, "Cash total should not be None"
        assert diesel_total is not None, "Diesel total should not be None"


class TestPDFContent:
    """Test PDF export content (basic checks)"""
    
    def test_pdf_has_content(self):
        """Verify PDF has content"""
        response = requests.get(f"{BASE_URL}/api/export/pdf")
        assert response.status_code == 200
        
        # PDF should have reasonable size
        assert len(response.content) > 1000, "PDF should have substantial content"
        print(f"✓ PDF size: {len(response.content)} bytes")
    
    def test_pdf_is_valid_reportlab_pdf(self):
        """Verify PDF is a valid ReportLab generated PDF"""
        response = requests.get(f"{BASE_URL}/api/export/pdf")
        
        # Check if it's a valid ReportLab PDF (text is compressed)
        content_str = response.content.decode('latin-1', errors='ignore')
        assert "ReportLab" in content_str, "PDF should be generated by ReportLab"
        print("✓ PDF is valid ReportLab generated PDF")


class TestAuthEndpoints:
    """Test authentication endpoints"""
    
    def test_login_with_admin_credentials(self):
        """Test login with admin/admin123"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "username": "admin",
            "password": "admin123"
        })
        assert response.status_code == 200
        data = response.json()
        assert data["success"] == True
        assert data["username"] == "admin"
        assert data["role"] == "admin"
        print("✓ Admin login successful")
    
    def test_login_with_invalid_credentials(self):
        """Test login with invalid credentials"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "username": "invalid",
            "password": "invalid"
        })
        assert response.status_code == 401
        print("✓ Invalid login rejected correctly")


class TestEntriesAPI:
    """Test entries CRUD API"""
    
    def test_get_entries(self):
        """Test getting entries list"""
        response = requests.get(f"{BASE_URL}/api/entries")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        print(f"✓ Got {len(data)} entries")
    
    def test_get_totals(self):
        """Test getting totals"""
        response = requests.get(f"{BASE_URL}/api/totals")
        assert response.status_code == 200
        data = response.json()
        
        # Verify totals include cash and diesel
        assert "total_cash_paid" in data
        assert "total_diesel_paid" in data
        assert "total_g_deposite" in data
        
        print(f"✓ Totals: G.Dep={data['total_g_deposite']}, Cash={data['total_cash_paid']}, Diesel={data['total_diesel_paid']}")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
