"""
Iteration 169 - Feature Tests
Tests for:
1. Python round_amount function (>0.50 rounds up, <=0.50 rounds down)
2. Excel export date format (DD-MM-YYYY)
3. API endpoints for entries
"""

import pytest
import requests
import os
import sys

# Add backend to path for importing models
sys.path.insert(0, '/app/backend')
from models import round_amount

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://paddy-ledger-1.preview.emergentagent.com').rstrip('/')


class TestRoundAmountFunction:
    """Test Python round_amount function - >0.50 rounds up, <=0.50 rounds down"""
    
    def test_round_amount_4000_51_rounds_up(self):
        """4000.51 should round up to 4001"""
        result = round_amount(4000.51)
        assert result == 4001, f"Expected 4001, got {result}"
    
    def test_round_amount_4000_50_rounds_down(self):
        """4000.50 should round down to 4000"""
        result = round_amount(4000.50)
        assert result == 4000, f"Expected 4000, got {result}"
    
    def test_round_amount_4000_49_rounds_down(self):
        """4000.49 should round down to 4000"""
        result = round_amount(4000.49)
        assert result == 4000, f"Expected 4000, got {result}"
    
    def test_round_amount_100_51_rounds_up(self):
        """100.51 should round up to 101"""
        result = round_amount(100.51)
        assert result == 101, f"Expected 101, got {result}"
    
    def test_round_amount_100_50_rounds_down(self):
        """100.50 should round down to 100"""
        result = round_amount(100.50)
        assert result == 100, f"Expected 100, got {result}"
    
    def test_round_amount_zero(self):
        """0 should return 0"""
        result = round_amount(0)
        assert result == 0, f"Expected 0, got {result}"
    
    def test_round_amount_none(self):
        """None should return 0"""
        result = round_amount(None)
        assert result == 0, f"Expected 0, got {result}"
    
    def test_round_amount_negative_51(self):
        """Negative -100.51 should round to -101"""
        result = round_amount(-100.51)
        assert result == -101, f"Expected -101, got {result}"
    
    def test_round_amount_negative_50(self):
        """Negative -100.50 should round to -100"""
        result = round_amount(-100.50)
        assert result == -100, f"Expected -100, got {result}"


class TestExcelExportDateFormat:
    """Test Excel export date format is DD-MM-YYYY"""
    
    def test_excel_export_returns_file(self):
        """Excel export endpoint should return a file"""
        response = requests.get(
            f"{BASE_URL}/api/export/excel",
            params={"kms_year": "2025-2026", "date_from": "2025-01-01", "date_to": "2026-12-31"}
        )
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.headers.get('content-type', '')
    
    def test_excel_export_date_format_dd_mm_yyyy(self):
        """Excel export should have dates in DD-MM-YYYY format"""
        import openpyxl
        from io import BytesIO
        
        response = requests.get(
            f"{BASE_URL}/api/export/excel",
            params={"kms_year": "2025-2026", "date_from": "2025-01-01", "date_to": "2026-12-31"}
        )
        assert response.status_code == 200
        
        # Load workbook from response content
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Find date column (usually first data column after headers)
        # Skip header rows (usually 4 rows: company name, tagline, report title, column headers)
        date_found = False
        for row_num in range(5, min(15, ws.max_row + 1)):
            date_cell = ws.cell(row=row_num, column=1).value
            if date_cell and isinstance(date_cell, str) and '-' in date_cell:
                # Check if date is in DD-MM-YYYY format
                parts = date_cell.split('-')
                if len(parts) == 3:
                    day, month, year = parts
                    # DD-MM-YYYY: day should be 2 digits, month 2 digits, year 4 digits
                    if len(day) == 2 and len(month) == 2 and len(year) == 4:
                        date_found = True
                        print(f"Found date in DD-MM-YYYY format: {date_cell}")
                        break
        
        assert date_found, "No dates found in DD-MM-YYYY format in Excel export"


class TestEntriesAPI:
    """Test entries API endpoints"""
    
    def test_entries_list(self):
        """GET /api/entries should return entries list"""
        response = requests.get(
            f"{BASE_URL}/api/entries",
            params={"kms_year": "2025-2026", "date_from": "2025-01-01", "date_to": "2026-12-31"}
        )
        assert response.status_code == 200
        data = response.json()
        assert "entries" in data or isinstance(data, list)
    
    def test_totals_endpoint(self):
        """GET /api/totals should return totals"""
        response = requests.get(
            f"{BASE_URL}/api/totals",
            params={"kms_year": "2025-2026"}
        )
        assert response.status_code == 200
        data = response.json()
        # Should have total fields
        assert "total_qntl" in data or "total_kg" in data


class TestLoginFlow:
    """Test login API"""
    
    def test_login_success(self):
        """Login with valid credentials should succeed"""
        response = requests.post(
            f"{BASE_URL}/api/auth/login",
            json={"username": "admin", "password": "admin123"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data.get("success") == True
        assert data.get("username") == "admin"
        assert data.get("role") == "admin"
    
    def test_login_invalid_password(self):
        """Login with invalid password should fail"""
        response = requests.post(
            f"{BASE_URL}/api/auth/login",
            json={"username": "admin", "password": "wrongpassword"}
        )
        assert response.status_code in [401, 400]


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
