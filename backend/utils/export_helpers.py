"""
Shared Excel & PDF styling helpers for professional colorful exports.
Used by all backend export endpoints.

Typography (v104.28.43+): Inter (UI text) + JetBrains Mono (numbers) — matches
the on-screen typography for a consistent brand feel across screen and print.
"""
import os
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# === FONT PATHS (resolves Inter/JetBrains Mono from the backend's bundled fonts dir) ===
_HERE = os.path.dirname(os.path.abspath(__file__))
_FONTS_DIR = os.path.normpath(os.path.join(_HERE, '..', 'fonts'))

def _font_path(filename):
    """Return abs path of a font in /app/backend/fonts/, or empty string if missing."""
    p = os.path.join(_FONTS_DIR, filename)
    return p if os.path.exists(p) else ''


# === REGISTER FONTS FOR PDF ===
_fonts_registered = False

def register_hindi_fonts():
    """Register Inter (UI) + JetBrains Mono (numbers) + FreeSans (Devanagari fallback)
    font families for use in ReportLab PDFs.
    """
    global _fonts_registered
    if _fonts_registered:
        return
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase.pdfmetrics import registerFontFamily

        # Inter (primary UI font) — registered as 'FreeSans' alias so existing
        # styles (which reference FreeSans/FreeSansBold) automatically pick up Inter.
        # We also register them under their proper names for new code.
        inter_reg = _font_path('Inter-Regular.ttf')
        inter_med = _font_path('Inter-Medium.ttf')
        inter_sb = _font_path('Inter-SemiBold.ttf')
        inter_bold = _font_path('Inter-Bold.ttf')
        inter_italic = _font_path('Inter-Italic.ttf')
        if inter_reg and inter_bold:
            pdfmetrics.registerFont(TTFont('Inter', inter_reg))
            pdfmetrics.registerFont(TTFont('InterMedium', inter_med or inter_reg))
            pdfmetrics.registerFont(TTFont('InterSemiBold', inter_sb or inter_bold))
            pdfmetrics.registerFont(TTFont('InterBold', inter_bold))
            if inter_italic:
                pdfmetrics.registerFont(TTFont('InterItalic', inter_italic))
            registerFontFamily('Inter', normal='Inter', bold='InterBold',
                              italic='InterItalic' if inter_italic else 'Inter',
                              boldItalic='InterBold')
            # Alias FreeSans -> Inter so legacy code paths inherit the new typography.
            pdfmetrics.registerFont(TTFont('FreeSans', inter_reg))
            pdfmetrics.registerFont(TTFont('FreeSansBold', inter_bold))
            pdfmetrics.registerFont(TTFont('FreeSansOblique', inter_italic or inter_reg))
            pdfmetrics.registerFont(TTFont('FreeSansBoldOblique', inter_bold))
            registerFontFamily('FreeSans', normal='FreeSans', bold='FreeSansBold',
                              italic='FreeSansOblique', boldItalic='FreeSansBoldOblique')
        else:
            # Fallback to system FreeSans for Devanagari support if Inter not bundled
            pdfmetrics.registerFont(TTFont('FreeSans', '/usr/share/fonts/truetype/freefont/FreeSans.ttf'))
            pdfmetrics.registerFont(TTFont('FreeSansBold', '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf'))
            pdfmetrics.registerFont(TTFont('FreeSansOblique', '/usr/share/fonts/truetype/freefont/FreeSansOblique.ttf'))
            pdfmetrics.registerFont(TTFont('FreeSansBoldOblique', '/usr/share/fonts/truetype/freefont/FreeSansBoldOblique.ttf'))
            registerFontFamily('FreeSans', normal='FreeSans', bold='FreeSansBold',
                              italic='FreeSansOblique', boldItalic='FreeSansBoldOblique')

        # JetBrains Mono (numbers)
        jbm_reg = _font_path('JetBrainsMono-Regular.ttf')
        jbm_med = _font_path('JetBrainsMono-Medium.ttf')
        jbm_bold = _font_path('JetBrainsMono-Bold.ttf')
        if jbm_reg and jbm_bold:
            pdfmetrics.registerFont(TTFont('JetBrainsMono', jbm_reg))
            pdfmetrics.registerFont(TTFont('JetBrainsMonoMedium', jbm_med or jbm_reg))
            pdfmetrics.registerFont(TTFont('JetBrainsMonoBold', jbm_bold))
            registerFontFamily('JetBrainsMono', normal='JetBrainsMono', bold='JetBrainsMonoBold',
                              italic='JetBrainsMono', boldItalic='JetBrainsMonoBold')

        _fonts_registered = True
    except Exception:
        pass


def get_pdf_styles():
    """Return getSampleStyleSheet() with FreeSans (Hindi-capable) as default font."""
    register_hindi_fonts()
    from reportlab.lib.styles import getSampleStyleSheet
    styles = getSampleStyleSheet()
    for name in styles.byName:
        style = styles[name]
        if hasattr(style, 'fontName'):
            if style.fontName == 'Helvetica':
                style.fontName = 'FreeSans'
            elif style.fontName == 'Helvetica-Bold':
                style.fontName = 'FreeSansBold'
            elif style.fontName == 'Helvetica-Oblique':
                style.fontName = 'FreeSansOblique'
            elif style.fontName == 'Helvetica-BoldOblique':
                style.fontName = 'FreeSansBoldOblique'
    return styles

# === COLOR PALETTE ===
COLORS = {
    'header_bg': '1B4F72',
    'header_text': 'FFFFFF',
    'title_bg': 'FEF3C7',
    'title_text': '1B4F72',
    'subtitle_bg': 'FFF7ED',
    'subtitle_text': 'D97706',
    'alt_row1': 'F0F7FF',
    'alt_row2': 'FFFFFF',
    'border': 'D0D5DD',
    'header_border': '0D3B66',
    # Amount
    'jama_text': '16A34A', 'jama_bg': 'DCFCE7',
    'nikasi_text': 'DC2626', 'nikasi_bg': 'FEE2E2',
    'balance_bg': 'FEFCE8', 'balance_text': '92400E',
    'date_bg': 'EFF6FF', 'date_text': '1E40AF',
    # Status
    'paid_text': '16A34A', 'paid_bg': 'DCFCE7',
    'pending_text': 'DC2626', 'pending_bg': 'FEE2E2',
    'partial_text': 'D97706', 'partial_bg': 'FEF3C7',
    # Total
    'total_bg': 'FEF3C7', 'total_border': 'F59E0B',
}

_thin = Side(style='thin', color=COLORS['border'])
_hair = Side(style='hair', color=COLORS['border'])
_med = Side(style='medium', color=COLORS['header_border'])
BORDER_THIN = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_HAIR = Border(left=_hair, right=_hair, top=_hair, bottom=_hair)
BORDER_HEADER = Border(left=_thin, right=_thin, top=_med, bottom=_med)


def style_excel_title(ws, title, ncols, subtitle="", branding=None):
    """Add a colorful title section: Company Name, Tagline + Custom Fields, Report Title.
    Always produces exactly 3 rows for backward compatibility."""
    branding = branding or {}
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "JOLKO, KESINGA")
    custom_fields = branding.get("custom_fields", [])

    # Build above-line text (fields with placement=above)
    above_parts = []
    for f in custom_fields:
        if f.get("placement", "below") == "above":
            lbl = f.get("label", "").strip()
            val = f.get("value", "").strip()
            if val:
                above_parts.append(f"{lbl}: {val}" if lbl else val)
    above_text = "  |  ".join(above_parts) if above_parts else ""

    # Build tagline + below custom fields combined text
    tagline_parts = [tagline] if tagline else []
    for f in custom_fields:
        if f.get("placement", "below") != "above":
            lbl = f.get("label", "").strip()
            val = f.get("value", "").strip()
            if val:
                tagline_parts.append(f"{lbl}: {val}" if lbl else val)
    combined_tagline = "  |  ".join(tagline_parts) if tagline_parts else ""

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)

    # Row 1: Above fields (if any) + Company Name
    row1_text = f"{above_text}\n{company}" if above_text else company
    c1 = ws.cell(row=1, column=1, value=row1_text)
    c1.font = Font(bold=True, size=18, color=COLORS['title_text'])
    c1.fill = PatternFill(start_color=COLORS['title_bg'], fill_type='solid')
    c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 48 if above_text else 36

    # Row 2: Tagline + custom fields
    c2 = ws.cell(row=2, column=1, value=combined_tagline)
    c2.font = Font(size=9, italic=True, color='555555')
    c2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22 if custom_fields else 20

    # Row 3: Report Title + date
    from datetime import datetime
    date_str = datetime.now().strftime('%d/%m/%Y')
    title_text = f"{title} | {date_str}" if subtitle == "" else f"{title} | {subtitle} | {date_str}"
    c3 = ws.cell(row=3, column=1, value=title_text)
    c3.font = Font(bold=True, size=12, color=COLORS['subtitle_text'])
    c3.fill = PatternFill(start_color=COLORS['subtitle_bg'], fill_type='solid')
    c3.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 26


def style_excel_header_row(ws, row_num, ncols):
    """Style a header row with dark background and white text."""
    ws.row_dimensions[row_num].height = 30
    hf = PatternFill(start_color=COLORS['header_bg'], fill_type='solid')
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill = hf
        c.font = Font(bold=True, size=10, color=COLORS['header_text'])
        c.border = BORDER_HEADER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_excel_data_rows(ws, start_row, end_row, ncols, headers=None):
    """Style data rows with column-type-aware coloring."""
    # Detect column types from headers
    header_lower = []
    if headers:
        header_lower = [str(h).lower() for h in headers]
    else:
        for col in range(1, ncols + 1):
            header_lower.append(str(ws.cell(row=start_row - 1, column=col).value or '').lower())
    
    is_date = [any(k in h for k in ('date', 'tarikh')) for h in header_lower]
    is_jama = [any(k in h for k in ('jama', 'credit', 'received', ' in')) for h in header_lower]
    is_nikasi = [any(k in h for k in ('nikasi', 'debit', 'paid', ' out')) for h in header_lower]
    is_balance = [any(k in h for k in ('balance', 'bal', 'bakaya')) for h in header_lower]
    is_amount = [any(k in h for k in ('amount', 'total', 'gross', 'net', 'rent', 'rate')) for h in header_lower]
    is_status = [any(k in h for k in ('status',)) for h in header_lower]
    
    for row in range(start_row, end_row + 1):
        is_even = (row - start_row) % 2 == 0
        base_bg = COLORS['alt_row1'] if is_even else COLORS['alt_row2']
        ws.row_dimensions[row].height = 22
        
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            ci = col - 1
            val = c.value
            
            c.fill = PatternFill(start_color=base_bg, fill_type='solid')
            c.border = BORDER_HAIR
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            
            if ci < len(is_date) and is_date[ci]:
                c.fill = PatternFill(start_color=COLORS['date_bg'], fill_type='solid')
                c.font = Font(size=10, color=COLORS['date_text'])
            
            if ci < len(is_jama) and is_jama[ci] and isinstance(val, (int, float)) and val > 0:
                c.fill = PatternFill(start_color=COLORS['jama_bg'], fill_type='solid')
                c.font = Font(bold=True, size=10, color=COLORS['jama_text'])
            
            if ci < len(is_nikasi) and is_nikasi[ci] and isinstance(val, (int, float)) and val > 0:
                c.fill = PatternFill(start_color=COLORS['nikasi_bg'], fill_type='solid')
                c.font = Font(bold=True, size=10, color=COLORS['nikasi_text'])
            
            if ci < len(is_balance) and is_balance[ci]:
                c.fill = PatternFill(start_color=COLORS['balance_bg'], fill_type='solid')
                c.font = Font(bold=True, size=10, color=COLORS['balance_text'])
                if isinstance(val, (int, float)) and val < 0:
                    c.font = Font(bold=True, size=10, color=COLORS['nikasi_text'])
            
            if ci < len(is_amount) and is_amount[ci] and isinstance(val, (int, float)):
                c.font = Font(bold=True, size=10)
                c.number_format = '#,##0.00'
            
            str_val = str(val or '')
            if ci < len(is_status) and is_status[ci]:
                if str_val.lower() in ('paid',):
                    c.fill = PatternFill(start_color=COLORS['paid_bg'], fill_type='solid')
                    c.font = Font(bold=True, size=10, color=COLORS['paid_text'])
                elif str_val.lower() in ('pending',):
                    c.fill = PatternFill(start_color=COLORS['pending_bg'], fill_type='solid')
                    c.font = Font(bold=True, size=10, color=COLORS['pending_text'])
                elif str_val.lower() in ('partial',):
                    c.fill = PatternFill(start_color=COLORS['partial_bg'], fill_type='solid')
                    c.font = Font(bold=True, size=10, color=COLORS['partial_text'])


def style_excel_total_row(ws, row_num, ncols):
    """Style the total row with amber background."""
    ws.row_dimensions[row_num].height = 26
    tf = PatternFill(start_color=COLORS['total_bg'], fill_type='solid')
    tb = Border(left=_thin, right=_thin, top=Side(style='medium', color=COLORS['total_border']),
                bottom=Side(style='medium', color=COLORS['total_border']))
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill = tf
        c.font = Font(bold=True, size=11)
        c.border = tb


def style_excel_summary_header(ws, row_num, ncols):
    """Style a summary sub-header row."""
    hf = PatternFill(start_color=COLORS['header_bg'], fill_type='solid')
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill = hf
        c.font = Font(bold=True, size=9, color=COLORS['header_text'])
        c.border = BORDER_THIN
        c.alignment = Alignment(horizontal='center')


def get_pdf_table_style(num_rows, cols_info=None):
    """Generate a colorful ReportLab TableStyle for PDF tables."""
    from reportlab.lib import colors as rl_colors
    
    register_hindi_fonts()
    style = [
        # Base font for all cells (Hindi support)
        ('FONTNAME', (0, 0), (-1, -1), 'FreeSans'),
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1B4F72')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'FreeSansBold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.4, rl_colors.HexColor('#D0D5DD')),
        # Total row (last row)
        ('BACKGROUND', (0, -1), (-1, -1), rl_colors.HexColor('#FEF3C7')),
        ('FONTNAME', (0, -1), (-1, -1), 'FreeSansBold'),
        ('LINEABOVE', (0, -1), (-1, -1), 1.5, rl_colors.HexColor('#F59E0B')),
    ]
    
    # Alternating row colors
    for i in range(1, num_rows):
        if i % 2 == 0:
            style.append(('BACKGROUND', (0, i), (-1, i), rl_colors.HexColor('#F0F7FF')))
    
    # Column-specific colors if cols_info provided
    if cols_info:
        for ci, col in enumerate(cols_info):
            h = col.get('header', '').lower() if isinstance(col, dict) else str(col).lower()
            if any(k in h for k in ('jama', 'credit', 'in')):
                for ri in range(1, num_rows):
                    style.append(('TEXTCOLOR', (ci, ri), (ci, ri), rl_colors.HexColor('#16A34A')))
            elif any(k in h for k in ('nikasi', 'debit', 'out')):
                for ri in range(1, num_rows):
                    style.append(('TEXTCOLOR', (ci, ri), (ci, ri), rl_colors.HexColor('#DC2626')))
    
    return style


def _build_custom_fields_row(branding, placement_filter="below"):
    """Build a ReportLab Table row for custom_fields with Left/Center/Right alignment.
    placement_filter: 'above' or 'below' to filter fields by placement."""
    register_hindi_fonts()
    from reportlab.platypus import Paragraph, Table as RLTable, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib import colors as rl_colors

    fields = branding.get("custom_fields", [])
    fields = [f for f in fields if f.get("placement", "below") == placement_filter]
    if not fields:
        return []

    left_parts, center_parts, right_parts = [], [], []
    for f in fields:
        lbl = f.get('label', '').strip()
        val = f.get('value', '').strip()
        if not val:
            continue
        txt = f"<b>{lbl}:</b> {val}" if lbl else f"{val}"
        pos = f.get("position", "center").lower()
        if pos == "left":
            left_parts.append(txt)
        elif pos == "right":
            right_parts.append(txt)
        else:
            center_parts.append(txt)

    style_l = ParagraphStyle('CFL', fontSize=8, fontName='FreeSans', textColor=rl_colors.HexColor('#374151'), alignment=TA_LEFT)
    style_c = ParagraphStyle('CFC', fontSize=8, fontName='FreeSans', textColor=rl_colors.HexColor('#374151'), alignment=TA_CENTER)
    style_r = ParagraphStyle('CFR', fontSize=8, fontName='FreeSans', textColor=rl_colors.HexColor('#374151'), alignment=TA_RIGHT)

    left_p = Paragraph("<br/>".join(left_parts) if left_parts else "", style_l)
    center_p = Paragraph("<br/>".join(center_parts) if center_parts else "", style_c)
    right_p = Paragraph("<br/>".join(right_parts) if right_parts else "", style_r)

    tbl = RLTable([[left_p, center_p, right_p]], colWidths=['33%', '34%', '33%'])
    tbl.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'FreeSans'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#D0D5DD')),
    ]))
    return [tbl]


def get_pdf_header_elements(title, branding=None, subtitle=""):
    """Return ReportLab Paragraph elements for company name, tagline, custom fields, and report title."""
    register_hindi_fonts()
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib import colors as rl_colors

    branding = branding or {}
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "JOLKO, KESINGA")

    company_style = ParagraphStyle(
        'CompanyHeader', fontSize=18, fontName='FreeSansBold',
        textColor=rl_colors.HexColor('#1B4F72'), alignment=TA_CENTER,
        spaceAfter=2, backColor=rl_colors.HexColor('#FFFBEB'),
        borderPadding=(6, 4, 6, 4),
    )
    tagline_style = ParagraphStyle(
        'TaglineHeader', fontSize=9, fontName='FreeSansOblique',
        textColor=rl_colors.HexColor('#6B7280'), alignment=TA_CENTER,
        spaceAfter=4,
    )
    title_style = ParagraphStyle(
        'ReportTitle', fontSize=12, fontName='FreeSansBold',
        textColor=rl_colors.white, alignment=TA_CENTER,
        backColor=rl_colors.HexColor('#0891B2'),
        borderPadding=(4, 3, 4, 3), spaceAfter=6,
    )

    elements = []
    # Custom fields ABOVE company name
    elements.extend(_build_custom_fields_row(branding, "above"))
    elements.append(Paragraph(company, company_style))
    elements.append(Spacer(1, 2))
    if tagline:
        elements.append(Paragraph(tagline, tagline_style))
    # Custom fields BELOW company name (default)
    elements.extend(_build_custom_fields_row(branding, "below"))
    elements.append(Paragraph(title if not subtitle else f"{title} | {subtitle}", title_style))
    elements.append(Spacer(1, 6))
    return elements


def get_pdf_company_header(branding=None):
    """Return just the company name + tagline + custom fields paragraphs for PDF (no title)."""
    register_hindi_fonts()
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib import colors as rl_colors

    branding = branding or {}
    company = branding.get("company_name", "NAVKAR AGRO")
    tagline = branding.get("tagline", "JOLKO, KESINGA")

    company_style = ParagraphStyle(
        'CompanyHdr', fontSize=18, fontName='FreeSansBold',
        textColor=rl_colors.HexColor('#1B4F72'), alignment=TA_CENTER,
        spaceAfter=2, backColor=rl_colors.HexColor('#FFFBEB'),
        borderPadding=(6, 4, 6, 4),
    )
    tagline_style = ParagraphStyle(
        'TaglineHdr', fontSize=9, fontName='FreeSansOblique',
        textColor=rl_colors.HexColor('#6B7280'), alignment=TA_CENTER,
        spaceAfter=6,
    )

    elements = []
    # Custom fields ABOVE company name
    elements.extend(_build_custom_fields_row(branding, "above"))
    elements.append(Paragraph(company, company_style))
    elements.append(Spacer(1, 2))
    if tagline:
        elements.append(Paragraph(tagline, tagline_style))
    # Custom fields BELOW company name (default)
    elements.extend(_build_custom_fields_row(branding, "below"))
    return elements


# ============================================================================
# BEAUTIFUL SINGLE-LINE SUMMARY BANNER (PDF + Excel) — LIGHT THEME
# ============================================================================
# Used at the bottom of export reports to show key statistics in a single
# professional horizontal strip with colored stat columns.
# Stat dict format: { 'label': 'TOTAL ENTRIES', 'value': '42', 'color': '#b45309' }
# 
# Light theme palette (high contrast on cream/amber-50 bg):
#   primary='#1e293b'  (dark slate)
#   green='#15803d'    (emerald-700)
#   red='#b91c1c'      (red-700)
#   gold='#b45309'     (amber-700)
#   orange='#c2410c'   (orange-700)
#   blue='#1d4ed8'     (blue-700)
#   emerald='#047857'  (emerald-700)
#   purple='#7e22ce'   (purple-700)

SUMMARY_BANNER_BG = '#FFFBEB'      # amber-50 cream
SUMMARY_LABEL_COLOR = '#64748b'    # slate-500 muted
SUMMARY_DIVIDER = '#E5E7EB'        # gray-200 hairline
SUMMARY_TOP_STRIPE = '#F59E0B'     # amber-500 gold
SUMMARY_BOTTOM_STRIPE = '#FCD34D'  # amber-300 lighter gold


def get_pdf_summary_banner(stats, total_width=None):
    """Returns a ReportLab Table representing the LIGHT-themed summary banner.
    
    Typography (v104.28.43+): label part in Inter, value part in JetBrainsMono
    with tabular alignment for a premium Stripe-grade aesthetic.
    """
    register_hindi_fonts()
    from reportlab.platypus import Table as RTable, TableStyle, Paragraph
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    if not stats:
        return None
    # Build a Paragraph per cell mixing Inter (label) + JetBrainsMono (value)
    cell_style = ParagraphStyle(
        'BannerCell', fontSize=9, fontName='FreeSansBold',
        textColor=rl_colors.HexColor(SUMMARY_LABEL_COLOR),
        alignment=TA_CENTER, leading=12,
    )

    def _esc(t):
        return str(t).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    cells = []
    try:
        # check if JetBrainsMono is registered
        from reportlab.pdfbase import pdfmetrics
        has_jbm = 'JetBrainsMonoBold' in pdfmetrics.getRegisteredFontNames()
    except Exception:
        has_jbm = False

    for s in stats:
        color = s.get('color', '#1e293b')
        label_html = f'<font color="{SUMMARY_LABEL_COLOR}">{_esc(s["label"])}</font>'
        if has_jbm:
            value_html = f'<font color="{color}" face="JetBrainsMonoBold" size="11">{_esc(s["value"])}</font>'
        else:
            value_html = f'<font color="{color}" size="11">{_esc(s["value"])}</font>'
        cells.append(Paragraph(f"{label_html}<br/>{value_html}", cell_style))

    n = len(stats)
    col_widths = [total_width / n] * n if total_width else [100] * n
    t = RTable([cells], colWidths=col_widths)
    t.hAlign = 'CENTER'

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, -1), rl_colors.HexColor(SUMMARY_BANNER_BG)),
        ('LINEABOVE', (0, 0), (-1, 0), 2, rl_colors.HexColor(SUMMARY_TOP_STRIPE)),
        ('LINEBELOW', (0, 0), (-1, 0), 1, rl_colors.HexColor(SUMMARY_BOTTOM_STRIPE)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]
    for i in range(1, n):
        style_cmds.append(('LINEBEFORE', (i, 0), (i, 0), 0.5, rl_colors.HexColor(SUMMARY_DIVIDER)))
    t.setStyle(TableStyle(style_cmds))
    return t


def add_excel_summary_banner(ws, row_num, ncols, stats):
    """Adds a LIGHT-themed single-line summary banner to an Excel worksheet at row_num.
    
    Args:
        ws: openpyxl worksheet
        row_num: row number to place the banner
        ncols: number of columns to merge across
        stats: list of {'label': str, 'value': str} dicts
    """
    if not stats:
        return
    parts = []
    for s in stats:
        parts.append(f"{s['label']}: {s['value']}")
    text = "  •  ".join(parts)
    text = f"📊  {text}"

    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = Font(bold=True, size=11, color="1E293B")  # dark slate text
    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amber-50
    cell.alignment = Alignment(horizontal='center', vertical='center')
    # Apply gold border on top + bottom for the visual frame
    cell.border = Border(
        top=Side(style='medium', color='F59E0B'),
        bottom=Side(style='thin', color='FCD34D'),
        left=Side(style='thin', color='FDE68A'),
        right=Side(style='thin', color='FDE68A'),
    )
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=ncols)
    ws.row_dimensions[row_num].height = 28


def fmt_inr(amount, with_currency=True):
    """Format a number as Indian Rupees with comma separators. e.g. 13892.0 -> 'Rs. 13,892'"""
    try:
        n = float(amount or 0)
    except Exception:
        n = 0.0
    if abs(n - int(n)) < 0.005:
        s = f"{int(n):,}"
    else:
        s = f"{n:,.2f}"
    return f"Rs. {s}" if with_currency else s


# Light-theme stat color palette (use these for value text colors)
STAT_COLORS = {
    'primary':   '#1E293B',  # slate-900 (default for white-on-dark replacement)
    'green':     '#15803D',  # emerald-700
    'red':       '#B91C1C',  # red-700
    'gold':      '#B45309',  # amber-700
    'orange':    '#C2410C',  # orange-700
    'blue':      '#1D4ED8',  # blue-700
    'emerald':   '#047857',  # emerald-700
    'purple':    '#7E22CE',  # purple-700
    'teal':      '#0F766E',  # teal-700
    'pink':      '#BE185D',  # pink-700
}



# ============================================================================
# SECTION BAND — full-width color-coded title bar used between report sections
# ============================================================================
# Each "section" inside a report (e.g. STOCK OVERVIEW, MANDI TARGETS) gets a
# prominent colored band so the eye can scan the document quickly. The band is
# a single-row Table that renders as a full-width title + optional sub-text.

# Pre-defined visual presets keyed by intent. Each preset = (bg, text, accent).
SECTION_BAND_PRESETS = {
    'navy':    ('#1E3A8A', '#FFFFFF', '#FBBF24'),  # blue-900 + amber accent
    'teal':    ('#0F766E', '#FFFFFF', '#A7F3D0'),  # teal-700
    'orange':  ('#C2410C', '#FFFFFF', '#FFEDD5'),  # orange-700
    'emerald': ('#047857', '#FFFFFF', '#A7F3D0'),  # emerald-700
    'rose':    ('#BE123C', '#FFFFFF', '#FECDD3'),  # rose-700
    'purple':  ('#6D28D9', '#FFFFFF', '#DDD6FE'),  # violet-700
    'amber':   ('#B45309', '#FFFFFF', '#FDE68A'),  # amber-700
    'slate':   ('#334155', '#FFFFFF', '#CBD5E1'),  # slate-700
}


def get_pdf_section_band(title, subtitle=None, preset='navy', total_width=None):
    """Returns a single full-width Table flowable rendering a colored section band.

    Args:
        title: Section title text (e.g. "STOCK OVERVIEW")
        subtitle: Optional small right-aligned subtitle (e.g. "Updated: 04-2026")
        preset: One of SECTION_BAND_PRESETS keys
        total_width: Total width in points (defaults to a value the caller will set)
    """
    register_hindi_fonts()
    from reportlab.platypus import Table as RTable, TableStyle
    from reportlab.lib import colors as rl_colors

    bg, fg, accent = SECTION_BAND_PRESETS.get(preset, SECTION_BAND_PRESETS['navy'])
    width = total_width or 540  # safe default for A4 portrait with 15mm margins

    if subtitle:
        # Two-cell band: title on left, subtitle on right
        row = [title.upper(), subtitle]
        col_widths = [width * 0.62, width * 0.38]
    else:
        row = [title.upper()]
        col_widths = [width]

    t = RTable([row], colWidths=col_widths)
    t.hAlign = 'LEFT'
    style = [
        ('BACKGROUND', (0, 0), (-1, -1), rl_colors.HexColor(bg)),
        ('TEXTCOLOR', (0, 0), (-1, -1), rl_colors.HexColor(fg)),
        ('FONTNAME', (0, 0), (-1, -1), 'FreeSansBold'),
        ('FONTSIZE', (0, 0), (0, 0), 11),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        # Accent stripe at left edge (visual marker)
        ('LINEBEFORE', (0, 0), (0, 0), 4, rl_colors.HexColor(accent)),
    ]
    if subtitle:
        style.append(('FONTSIZE', (1, 0), (1, 0), 8.5))
        style.append(('TEXTCOLOR', (1, 0), (1, 0), rl_colors.HexColor(accent)))
        style.append(('ALIGN', (1, 0), (1, 0), 'RIGHT'))
        style.append(('FONTNAME', (1, 0), (1, 0), 'FreeSans'))
    t.setStyle(TableStyle(style))
    return t
