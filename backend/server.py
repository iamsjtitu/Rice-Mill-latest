from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import StreamingResponse, Response, HTMLResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import secrets

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Print page storage (server-side print for Electron compatibility)
print_pages = {}

@api_router.post("/print")
async def create_print_page(request: Request):
    data = await request.json()
    import uuid as _uuid
    page_id = str(_uuid.uuid4())
    print_pages[page_id] = data.get('html', '')
    return {"id": page_id, "url": f"/api/print/{page_id}"}

@api_router.get("/print/{page_id}")
async def get_print_page(page_id: str):
    html = print_pages.pop(page_id, None)
    if not html:
        return HTMLResponse("<h1>Page expired. Please try again.</h1>", status_code=404)
    return HTMLResponse(html)

# Security
security = HTTPBasic()

# Default credentials (in production, store hashed in DB)
USERS = {
    "admin": {"password": "admin123", "role": "admin"},
    "staff": {"password": "staff123", "role": "staff"}
}


# User Models
class User(BaseModel):
    username: str
    role: str


class LoginRequest(BaseModel):
    username: str
    password: str


class LoginResponse(BaseModel):
    success: bool
    username: str
    role: str
    message: str


class PasswordChangeRequest(BaseModel):
    username: str
    current_password: str
    new_password: str


# Define Models
class MillEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    kms_year: str = ""  # e.g., "2025-2026"
    season: str = ""  # "Kharif" or "Rabi"
    truck_no: str = ""
    rst_no: str = ""  # RST Number
    tp_no: str = ""   # TP Number
    agent_name: str = ""
    mandi_name: str = ""
    kg: float = 0
    qntl: float = 0  # Auto calculated: kg / 100
    bag: int = 0
    g_deposite: float = 0
    gbw_cut: float = 0
    mill_w: float = 0  # Auto calculated: kg - gbw_cut
    plastic_bag: int = 0  # P.Pkt - Plastic packet count
    p_pkt_cut: float = 0  # Auto calculated: plastic_bag * 0.5
    moisture: float = 0  # Moisture percentage
    moisture_cut: float = 0  # Auto calculated moisture cut
    moisture_cut_percent: float = 0  # Moisture cut percentage (moisture - 17 if > 17)
    cutting_percent: float = 0  # Cutting percentage (5%, 5.26% etc)
    cutting: float = 0  # Auto calculated from percentage
    disc_dust_poll: float = 0
    final_w: float = 0  # Auto calculated
    g_issued: float = 0
    cash_paid: float = 0
    diesel_paid: float = 0
    remark: str = ""
    created_by: str = ""  # Username who created
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class MillEntryCreate(BaseModel):
    date: str
    kms_year: str = ""
    season: str = ""
    truck_no: str = ""
    rst_no: str = ""  # RST Number
    tp_no: str = ""   # TP Number
    agent_name: str = ""
    mandi_name: str = ""
    kg: float = 0
    bag: int = 0
    g_deposite: float = 0
    gbw_cut: float = 0
    plastic_bag: int = 0
    cutting_percent: float = 0
    disc_dust_poll: float = 0
    g_issued: float = 0
    moisture: float = 0
    cash_paid: float = 0
    diesel_paid: float = 0
    remark: str = ""


class MillEntryUpdate(BaseModel):
    date: Optional[str] = None
    kms_year: Optional[str] = None
    season: Optional[str] = None
    truck_no: Optional[str] = None
    rst_no: Optional[str] = None  # RST Number
    tp_no: Optional[str] = None   # TP Number
    agent_name: Optional[str] = None
    mandi_name: Optional[str] = None
    kg: Optional[float] = None
    bag: Optional[int] = None
    g_deposite: Optional[float] = None
    gbw_cut: Optional[float] = None
    plastic_bag: Optional[int] = None
    cutting_percent: Optional[float] = None
    disc_dust_poll: Optional[float] = None
    g_issued: Optional[float] = None
    moisture: Optional[float] = None
    cash_paid: Optional[float] = None
    diesel_paid: Optional[float] = None
    remark: Optional[str] = None


class TotalsResponse(BaseModel):
    total_kg: float = 0
    total_qntl: float = 0
    total_bag: int = 0
    total_g_deposite: float = 0
    total_gbw_cut: float = 0
    total_mill_w: float = 0
    total_p_pkt_cut: float = 0
    total_cutting: float = 0
    total_disc_dust_poll: float = 0
    total_final_w: float = 0
    total_g_issued: float = 0
    total_cash_paid: float = 0
    total_diesel_paid: float = 0


# Mandi Target Models
class MandiTarget(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    mandi_name: str
    target_qntl: float  # Base target in QNTL
    cutting_percent: float  # Expected cutting % (5%, 5.26% etc)
    expected_total: float = 0  # Auto: target_qntl + (target_qntl * cutting_percent / 100)
    base_rate: float = 10.0  # Rate per QNTL for target (e.g., ₹10)
    cutting_rate: float = 5.0  # Rate per QNTL for cutting excess (e.g., ₹5)
    kms_year: str  # e.g., "2025-2026"
    season: str  # "Kharif" or "Rabi"
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class MandiTargetCreate(BaseModel):
    mandi_name: str
    target_qntl: float
    cutting_percent: float = 5.0
    base_rate: float = 10.0
    cutting_rate: float = 5.0
    kms_year: str
    season: str


class MandiTargetUpdate(BaseModel):
    mandi_name: Optional[str] = None
    target_qntl: Optional[float] = None
    cutting_percent: Optional[float] = None
    base_rate: Optional[float] = None
    cutting_rate: Optional[float] = None
    kms_year: Optional[str] = None
    season: Optional[str] = None


# ============ MILLING ENTRY MODELS ============
class MillingEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    rice_type: str = "parboiled"  # "parboiled" or "raw"
    paddy_input_qntl: float = 0  # from available paddy stock (Mill W. QNTL)
    
    # Output percentages from paddy (user enters)
    rice_percent: float = 0
    bran_percent: float = 0
    kunda_percent: float = 0
    broken_percent: float = 0
    kanki_percent: float = 0
    husk_percent: float = 0  # auto-calculated as remainder
    
    # Auto-calculated QNTL from paddy
    rice_qntl: float = 0
    bran_qntl: float = 0
    kunda_qntl: float = 0
    broken_qntl: float = 0
    kanki_qntl: float = 0
    husk_qntl: float = 0
    
    # FRK used from stock for CMR
    frk_used_qntl: float = 0
    
    # CMR
    cmr_delivery_qntl: float = 0  # rice_qntl + frk_used_qntl
    outturn_ratio: float = 0      # cmr / paddy * 100
    
    # Meta
    kms_year: str = ""
    season: str = ""
    note: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class MillingEntryCreate(BaseModel):
    date: str
    rice_type: str = "parboiled"
    paddy_input_qntl: float = 0
    rice_percent: float = 0
    bran_percent: float = 0
    kunda_percent: float = 0
    broken_percent: float = 0
    kanki_percent: float = 0
    frk_used_qntl: float = 0
    kms_year: str = ""
    season: str = ""
    note: str = ""


# ============ FRK PURCHASE MODEL ============
class FrkPurchase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    party_name: str = ""
    quantity_qntl: float = 0
    rate_per_qntl: float = 0
    total_amount: float = 0  # auto: qty * rate
    note: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class FrkPurchaseCreate(BaseModel):
    date: str
    party_name: str = ""
    quantity_qntl: float = 0
    rate_per_qntl: float = 0
    note: str = ""
    kms_year: str = ""
    season: str = ""


# ============ BY-PRODUCT STOCK MODELS ============
class ByProductSale(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    product: str  # bran, kunda, broken, kanki, husk
    quantity_qntl: float = 0
    rate_per_qntl: float = 0
    total_amount: float = 0  # auto: qty * rate
    buyer_name: str = ""
    note: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class ByProductSaleCreate(BaseModel):
    date: str
    product: str
    quantity_qntl: float = 0
    rate_per_qntl: float = 0
    buyer_name: str = ""
    note: str = ""
    kms_year: str = ""
    season: str = ""


class MandiTargetSummary(BaseModel):
    id: str  # Target ID for edit/delete
    mandi_name: str
    target_qntl: float
    cutting_percent: float
    expected_total: float
    achieved_qntl: float  # Sum of final_w for this mandi
    pending_qntl: float  # expected_total - achieved_qntl
    progress_percent: float  # (achieved / expected) * 100
    base_rate: float
    cutting_rate: float
    target_amount: float  # target_qntl × base_rate
    cutting_qntl: float  # cutting excess QNTL
    cutting_amount: float  # cutting_qntl × cutting_rate
    total_agent_amount: float  # target_amount + cutting_amount
    kms_year: str
    season: str


# ============ PAYMENT MODELS ============

class PaymentRecord(BaseModel):
    amount: float
    date: str
    note: str = ""


class TruckPaymentStatus(BaseModel):
    entry_id: str
    truck_no: str
    date: str
    total_qntl: float
    total_bag: int
    final_qntl: float
    cash_taken: float
    diesel_taken: float
    rate_per_qntl: float
    gross_amount: float  # final_qntl × rate
    deductions: float  # cash + diesel
    net_amount: float  # gross - deductions
    paid_amount: float
    balance_amount: float
    status: str  # pending, partial, paid
    kms_year: str
    season: str
    agent_name: str
    mandi_name: str


class AgentPaymentStatus(BaseModel):
    """Agent payment based on Mandi Target (not achieved)"""
    mandi_name: str
    agent_name: str
    target_qntl: float
    cutting_percent: float
    cutting_qntl: float  # target × cutting%
    base_rate: float
    cutting_rate: float
    target_amount: float  # target_qntl × base_rate
    cutting_amount: float  # cutting_qntl × cutting_rate
    total_amount: float  # target_amount + cutting_amount
    achieved_qntl: float  # Actual achieved for reference
    is_target_complete: bool  # achieved >= expected_total
    paid_amount: float
    balance_amount: float
    status: str  # pending, partial, paid
    kms_year: str
    season: str


class SetRateRequest(BaseModel):
    rate_per_qntl: float


class MakePaymentRequest(BaseModel):
    amount: float
    note: str = ""


# Branding Settings Model
class BrandingSettings(BaseModel):
    company_name: str = "NAVKAR AGRO"
    tagline: str = "JOLKO, KESINGA - Mill Entry System"
    updated_by: str = ""
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class BrandingUpdateRequest(BaseModel):
    company_name: str
    tagline: str


def calculate_auto_fields(data: dict) -> dict:
    """Calculate automatic fields based on input data"""
    kg = data.get('kg', 0) or 0
    gbw_cut = data.get('gbw_cut', 0) or 0
    disc_dust_poll = data.get('disc_dust_poll', 0) or 0
    plastic_bag = data.get('plastic_bag', 0) or 0
    cutting_percent = data.get('cutting_percent', 0) or 0
    moisture = data.get('moisture', 0) or 0
    
    # P.Pkt cut calculation (0.5 kg per plastic bag)
    p_pkt_cut = round(plastic_bag * 0.5, 2)
    data['p_pkt_cut'] = p_pkt_cut
    
    # Mill W in KG and QNTL
    mill_w_kg = kg - gbw_cut
    mill_w_qntl = mill_w_kg / 100
    
    # Moisture cut: 17% tak no cut, uske upar (moisture - 17)% cut from Mill W QNTL
    moisture_cut_percent = max(0, moisture - 17)
    moisture_cut_qntl = round((mill_w_qntl * moisture_cut_percent) / 100, 2)
    moisture_cut_kg = round(moisture_cut_qntl * 100, 2)
    data['moisture_cut'] = moisture_cut_kg
    data['moisture_cut_qntl'] = moisture_cut_qntl
    data['moisture_cut_percent'] = moisture_cut_percent
    
    # Cutting from Mill W QNTL
    cutting_qntl = round((mill_w_qntl * cutting_percent) / 100, 2)
    cutting_kg = round(cutting_qntl * 100, 2)
    data['cutting'] = cutting_kg
    data['cutting_qntl'] = cutting_qntl
    
    # P.Pkt cut in QNTL
    p_pkt_cut_qntl = p_pkt_cut / 100
    
    # Disc/Dust/Poll in QNTL
    disc_dust_poll_qntl = disc_dust_poll / 100
    
    # Auto calculations
    data['qntl'] = round(kg / 100, 2)  # KG to Quintals
    data['mill_w'] = mill_w_kg  # Mill Weight in KG (stored)
    
    # Final W = Mill W QNTL - P.Pkt QNTL - Moisture Cut QNTL - Cutting QNTL - Disc/Dust QNTL
    final_w_qntl = mill_w_qntl - p_pkt_cut_qntl - moisture_cut_qntl - cutting_qntl - disc_dust_poll_qntl
    data['final_w'] = round(final_w_qntl * 100, 2)  # Store in KG for compatibility
    
    return data


def can_edit_entry(entry: dict, username: str, role: str) -> tuple:
    """Check if user can edit/delete entry"""
    if role == "admin":
        return True, "Admin access"
    
    # Staff can only edit their own entries within 5 minutes
    if entry.get('created_by') != username:
        return False, "Aap sirf apni entry edit kar sakte hain"
    
    created_at = entry.get('created_at', '')
    if created_at:
        try:
            created_time = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            now = datetime.now(timezone.utc)
            time_diff = now - created_time
            
            if time_diff > timedelta(minutes=5):
                return False, "5 minute se zyada ho gaye, ab edit nahi ho sakta"
        except:
            pass
    
    return True, "Edit allowed"


# ============ AUTH ENDPOINTS ============

@api_router.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    username = request.username
    password = request.password
    
    # Check from database first (for changed passwords)
    user_doc = await db.users.find_one({"username": username}, {"_id": 0})
    
    if user_doc:
        if user_doc.get("password") == password:
            return LoginResponse(
                success=True,
                username=username,
                role=user_doc.get("role"),
                message="Login successful"
            )
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    # Fallback to default users
    if username in USERS and USERS[username]["password"] == password:
        return LoginResponse(
            success=True,
            username=username,
            role=USERS[username]["role"],
            message="Login successful"
        )
    
    raise HTTPException(status_code=401, detail="Invalid username or password")


@api_router.get("/auth/verify")
async def verify_user(username: str, role: str):
    # Check from database first
    user_doc = await db.users.find_one({"username": username}, {"_id": 0})
    if user_doc and user_doc.get("role") == role:
        return {"valid": True, "username": username, "role": role}
    # Fallback to default users
    if username in USERS and USERS[username]["role"] == role:
        return {"valid": True, "username": username, "role": role}
    return {"valid": False}


@api_router.post("/auth/change-password")
async def change_password(request: PasswordChangeRequest):
    username = request.username
    current_password = request.current_password
    new_password = request.new_password
    
    # Check from database first
    user_doc = await db.users.find_one({"username": username}, {"_id": 0})
    
    if user_doc:
        # User exists in database
        if user_doc.get("password") != current_password:
            raise HTTPException(status_code=401, detail="Current password galat hai")
        
        await db.users.update_one(
            {"username": username},
            {"$set": {"password": new_password}}
        )
        return {"success": True, "message": "Password changed successfully"}
    
    # Check default users
    if username in USERS:
        if USERS[username]["password"] != current_password:
            raise HTTPException(status_code=401, detail="Current password galat hai")
        
        # Create user in database with new password
        await db.users.insert_one({
            "username": username,
            "password": new_password,
            "role": USERS[username]["role"]
        })
        return {"success": True, "message": "Password changed successfully"}
    
    raise HTTPException(status_code=404, detail="User not found")


# ============ BRANDING SETTINGS ============

@api_router.get("/branding")
async def get_branding():
    """Get current branding settings"""
    branding = await db.branding.find_one({}, {"_id": 0})
    if not branding:
        # Return default branding
        return {
            "company_name": "NAVKAR AGRO",
            "tagline": "JOLKO, KESINGA - Mill Entry System"
        }
    return branding


@api_router.put("/branding")
async def update_branding(request: BrandingUpdateRequest, username: str = "", role: str = ""):
    """Update branding settings (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf Admin branding update kar sakta hai")
    
    branding_data = {
        "company_name": request.company_name,
        "tagline": request.tagline,
        "updated_by": username,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.branding.update_one(
        {},
        {"$set": branding_data},
        upsert=True
    )
    
    return {"success": True, "message": "Branding update ho gaya", "branding": branding_data}


# Helper function to get branding for exports
async def get_company_name():
    branding = await db.branding.find_one({}, {"_id": 0})
    if branding:
        return branding.get("company_name", "NAVKAR AGRO"), branding.get("tagline", "")
    return "NAVKAR AGRO", "JOLKO, KESINGA"


# ============ MILL ENTRIES CRUD ============

@api_router.get("/")
async def root():
    return {"message": "Mill Entry API - Navkar Agro"}


@api_router.post("/entries", response_model=MillEntry)
async def create_entry(input: MillEntryCreate, username: str = "", role: str = ""):
    entry_dict = input.model_dump()
    entry_dict = calculate_auto_fields(entry_dict)
    entry_dict['created_by'] = username
    
    entry_obj = MillEntry(**entry_dict)
    doc = entry_obj.model_dump()
    
    await db.mill_entries.insert_one(doc)
    return entry_obj


@api_router.get("/entries", response_model=List[MillEntry])
async def get_entries(
    truck_no: Optional[str] = None,
    rst_no: Optional[str] = None,
    tp_no: Optional[str] = None,
    agent_name: Optional[str] = None,
    mandi_name: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    query = {}
    
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if rst_no:
        query["rst_no"] = {"$regex": rst_no, "$options": "i"}
    if tp_no:
        query["tp_no"] = {"$regex": tp_no, "$options": "i"}
    if agent_name:
        query["agent_name"] = {"$regex": agent_name, "$options": "i"}
    if mandi_name:
        query["mandi_name"] = {"$regex": mandi_name, "$options": "i"}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    # Date range filter
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        if date_query:
            query["date"] = date_query
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return entries


@api_router.get("/entries/{entry_id}", response_model=MillEntry)
async def get_entry(entry_id: str):
    entry = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    return entry


@api_router.put("/entries/{entry_id}", response_model=MillEntry)
async def update_entry(entry_id: str, input: MillEntryUpdate, username: str = "", role: str = ""):
    existing = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Check permission
    can_edit, message = can_edit_entry(existing, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    
    # Merge existing data with updates
    merged_data = {**existing, **update_data}
    merged_data = calculate_auto_fields(merged_data)
    merged_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.mill_entries.update_one(
        {"id": entry_id},
        {"$set": merged_data}
    )
    
    updated = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    return updated


@api_router.delete("/entries/{entry_id}")
async def delete_entry(entry_id: str, username: str = "", role: str = ""):
    existing = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Check permission
    can_edit, message = can_edit_entry(existing, username, role)
    if not can_edit:
        raise HTTPException(status_code=403, detail=message)
    
    result = await db.mill_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted successfully"}


@api_router.get("/totals", response_model=TotalsResponse)
async def get_totals(
    truck_no: Optional[str] = None,
    agent_name: Optional[str] = None,
    mandi_name: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    match_query = {}
    
    if truck_no:
        match_query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if agent_name:
        match_query["agent_name"] = {"$regex": agent_name, "$options": "i"}
    if mandi_name:
        match_query["mandi_name"] = {"$regex": mandi_name, "$options": "i"}
    if kms_year:
        match_query["kms_year"] = kms_year
    if season:
        match_query["season"] = season
    
    # Date range filter
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        if date_query:
            match_query["date"] = date_query
    
    pipeline = []
    if match_query:
        pipeline.append({"$match": match_query})
    
    pipeline.append({
        "$group": {
            "_id": None,
            "total_kg": {"$sum": "$kg"},
            "total_qntl": {"$sum": "$qntl"},
            "total_bag": {"$sum": "$bag"},
            "total_g_deposite": {"$sum": "$g_deposite"},
            "total_gbw_cut": {"$sum": "$gbw_cut"},
            "total_mill_w": {"$sum": "$mill_w"},
            "total_p_pkt_cut": {"$sum": "$p_pkt_cut"},
            "total_cutting": {"$sum": "$cutting"},
            "total_disc_dust_poll": {"$sum": "$disc_dust_poll"},
            "total_final_w": {"$sum": "$final_w"},
            "total_g_issued": {"$sum": "$g_issued"},
            "total_cash_paid": {"$sum": "$cash_paid"},
            "total_diesel_paid": {"$sum": "$diesel_paid"}
        }
    })
    
    result = await db.mill_entries.aggregate(pipeline).to_list(1)
    
    if result:
        totals = result[0]
        del totals['_id']
        return TotalsResponse(**totals)
    
    return TotalsResponse()


# ============ AUTO-SUGGEST ENDPOINTS ============

@api_router.get("/suggestions/trucks")
async def get_truck_suggestions(q: str = ""):
    if len(q) < 1:
        trucks = await db.mill_entries.distinct("truck_no")
    else:
        trucks = await db.mill_entries.distinct("truck_no", {"truck_no": {"$regex": q, "$options": "i"}})
    return {"suggestions": [t for t in trucks if t]}


@api_router.get("/suggestions/agents")
async def get_agent_suggestions(q: str = ""):
    if len(q) < 1:
        agents = await db.mill_entries.distinct("agent_name")
    else:
        agents = await db.mill_entries.distinct("agent_name", {"agent_name": {"$regex": q, "$options": "i"}})
    return {"suggestions": [a for a in agents if a]}


@api_router.get("/suggestions/mandis")
async def get_mandi_suggestions(q: str = "", agent_name: str = ""):
    query = {}
    if q:
        query["mandi_name"] = {"$regex": q, "$options": "i"}
    if agent_name:
        query["agent_name"] = agent_name
    
    mandis = await db.mill_entries.distinct("mandi_name", query if query else None)
    return {"suggestions": [m for m in mandis if m]}


@api_router.get("/suggestions/kms_years")
async def get_kms_year_suggestions():
    years = await db.mill_entries.distinct("kms_year")
    return {"suggestions": [y for y in years if y]}


# ============ EXPORT ENDPOINTS ============

@api_router.get("/export/excel")
async def export_excel(
    truck_no: Optional[str] = None,
    agent_name: Optional[str] = None,
    mandi_name: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export entries to styled Excel file"""
    query = {}
    
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if agent_name:
        query["agent_name"] = {"$regex": agent_name, "$options": "i"}
    if mandi_name:
        query["mandi_name"] = {"$regex": mandi_name, "$options": "i"}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mill Entries"
    
    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    
    title_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    
    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    total_font = Font(bold=True, size=9)
    
    qntl_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    final_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    gunny_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    cash_fill = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title
    ws.merge_cells('A1:Q1')
    company_name, tagline = await get_company_name()
    ws['A1'] = f"{company_name} - Mill Entries | KMS: {kms_year or 'All'} | {season or 'All Seasons'}"
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30
    
    # Date row
    ws.merge_cells('A2:Q2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws['A2'].alignment = center_align
    ws.row_dimensions[2].height = 20
    
    # Headers
    headers = [
        "Date", "Truck No", "Agent", "Mandi", "QNTL", "BAG", "G.Dep",
        "GBW Cut", "Mill W", "Moist%", "M.Cut", "Cut%", 
        "D/D/P", "Final W", "G.Issued", "Cash", "Diesel"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[3].height = 22
    
    # Data rows
    row_num = 4
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    for idx, entry in enumerate(entries):
        row_data = [
            entry.get('date', ''),
            entry.get('truck_no', ''),
            entry.get('agent_name', ''),
            entry.get('mandi_name', ''),
            round(entry.get('qntl', 0), 2),
            entry.get('bag', 0),
            entry.get('g_deposite', 0),
            round(entry.get('gbw_cut', 0), 2),
            round(entry.get('mill_w', 0) / 100, 2),
            entry.get('moisture', 0),
            round(entry.get('moisture_cut', 0) / 100, 2) if entry.get('moisture_cut') else 0,
            entry.get('cutting_percent', 0),
            entry.get('disc_dust_poll', 0),
            round(entry.get('final_w', 0) / 100, 2),
            entry.get('g_issued', 0),
            entry.get('cash_paid', 0),
            entry.get('diesel_paid', 0)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            if idx % 2 == 1:
                cell.fill = alt_fill
            
            # Special column colors
            if col == 5:  # QNTL
                cell.fill = qntl_fill
                cell.alignment = right_align
            elif col == 7:  # G.Deposite
                cell.fill = gunny_fill
                cell.alignment = right_align
            elif col == 14:  # Final W
                cell.fill = final_fill
                cell.font = Font(bold=True)
                cell.alignment = right_align
            elif col == 16:  # Cash
                cell.fill = cash_fill
                cell.alignment = right_align
            elif col == 17:  # Diesel
                cell.fill = cash_fill
                cell.alignment = right_align
            elif col in [6, 8, 9, 10, 11, 12, 13, 15]:
                cell.alignment = right_align
        
        row_num += 1
    
    # Totals row
    totals = await get_totals(truck_no, agent_name, mandi_name, kms_year, season)
    totals_data = [
        "TOTAL", "", "", "",
        round(totals.total_qntl, 2),
        totals.total_bag,
        totals.total_g_deposite,
        round(totals.total_gbw_cut, 2),
        round(totals.total_mill_w / 100, 2),
        "-",
        "-",
        "-",
        totals.total_disc_dust_poll,
        round(totals.total_final_w / 100, 2),
        totals.total_g_issued,
        totals.total_cash_paid,
        totals.total_diesel_paid
    ]
    
    for col, value in enumerate(totals_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
        if col >= 5:
            cell.alignment = right_align
    
    # Column widths
    col_widths = [10, 12, 12, 12, 8, 6, 6, 8, 8, 6, 6, 6, 6, 8, 8, 8, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"mill_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/export/pdf")
async def export_pdf(
    truck_no: Optional[str] = None,
    agent_name: Optional[str] = None,
    mandi_name: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export entries to styled PDF file (A4 Landscape)"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    
    query = {}
    
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    if agent_name:
        query["agent_name"] = {"$regex": agent_name, "$options": "i"}
    if mandi_name:
        query["mandi_name"] = {"$regex": mandi_name, "$options": "i"}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    totals = await get_totals(truck_no, agent_name, mandi_name, kms_year, season)
    
    # Create PDF buffer
    buffer = io.BytesIO()
    
    # A4 Landscape
    page_width, page_height = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=8*mm,
        rightMargin=8*mm,
        topMargin=8*mm,
        bottomMargin=8*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.white,
        alignment=TA_CENTER,
        spaceAfter=2*mm
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#475569'),
        alignment=TA_CENTER,
        spaceAfter=3*mm
    )
    
    # Title table with orange background
    company_name, tagline = await get_company_name()
    title_text = f"{company_name} - Mill Entries | KMS: {kms_year or 'All'} | {season or 'All Seasons'}"
    title_data = [[Paragraph(f"<b>{title_text}</b>", title_style)]]
    title_table = Table(title_data, colWidths=[page_width - 16*mm])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D97706')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(title_table)
    
    # Date
    date_text = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    elements.append(Paragraph(date_text, subtitle_style))
    
    # Table headers
    headers = [
        "Date", "Truck No", "Agent", "Mandi", "QNTL", "BAG", "G.Dep",
        "GBW Cut", "Mill W", "Moist%", "M.Cut", "Cut%", 
        "D/D/P", "Final W", "G.Issued", "Cash", "Diesel"
    ]
    
    # Build data rows
    table_data = [headers]
    
    for entry in entries:
        row = [
            entry.get('date', '')[:10] if entry.get('date') else '',
            entry.get('truck_no', '')[:10] if entry.get('truck_no') else '',
            entry.get('agent_name', '')[:10] if entry.get('agent_name') else '',
            entry.get('mandi_name', '')[:10] if entry.get('mandi_name') else '',
            f"{entry.get('qntl', 0):.2f}",
            str(entry.get('bag', 0)),
            str(entry.get('g_deposite', 0)),
            f"{entry.get('gbw_cut', 0):.1f}",
            f"{entry.get('mill_w', 0) / 100:.2f}",
            f"{entry.get('moisture', 0):.0f}",
            f"{(entry.get('moisture_cut', 0) / 100):.2f}" if entry.get('moisture_cut') else "0",
            f"{entry.get('cutting_percent', 0):.1f}",
            str(entry.get('disc_dust_poll', 0)),
            f"{entry.get('final_w', 0) / 100:.2f}",
            str(entry.get('g_issued', 0)),
            str(entry.get('cash_paid', 0)),
            str(entry.get('diesel_paid', 0))
        ]
        table_data.append(row)
    
    # Totals row
    totals_row = [
        "TOTAL", "", "", "",
        f"{totals.total_qntl:.2f}",
        str(totals.total_bag),
        str(int(totals.total_g_deposite)),
        f"{totals.total_gbw_cut:.1f}",
        f"{totals.total_mill_w / 100:.2f}",
        "-",
        "-",
        "-",
        str(int(totals.total_disc_dust_poll)),
        f"{totals.total_final_w / 100:.2f}",
        str(int(totals.total_g_issued)),
        str(int(totals.total_cash_paid)),
        str(int(totals.total_diesel_paid))
    ]
    table_data.append(totals_row)
    
    # Column widths (total ~265mm for A4 landscape with margins)
    col_widths = [14*mm, 16*mm, 16*mm, 16*mm, 12*mm, 10*mm, 10*mm, 12*mm, 12*mm, 
                  10*mm, 10*mm, 10*mm, 10*mm, 14*mm, 14*mm, 12*mm, 12*mm]
    
    # Create table
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Define colors
    header_bg = colors.HexColor('#1E293B')
    alt_row_bg = colors.HexColor('#F8FAFC')
    qntl_bg = colors.HexColor('#D1FAE5')
    gunny_bg = colors.HexColor('#DBEAFE')
    final_bg = colors.HexColor('#FDE68A')
    cash_bg = colors.HexColor('#FCE7F3')
    total_bg = colors.HexColor('#FEF3C7')
    
    # Table styles
    style_commands = [
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # All cells
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        
        # Right align numeric columns
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        
        # Totals row (last row)
        ('BACKGROUND', (0, -1), (-1, -1), total_bg),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 6),
    ]
    
    # Add alternating row colors for data rows
    for i in range(1, len(table_data) - 1):  # Exclude header and totals
        if i % 2 == 0:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), alt_row_bg))
    
    # Highlight special columns for all data rows
    for i in range(1, len(table_data) - 1):
        style_commands.append(('BACKGROUND', (4, i), (4, i), qntl_bg))  # QNTL
        style_commands.append(('BACKGROUND', (6, i), (6, i), gunny_bg))  # G.Dep
        style_commands.append(('BACKGROUND', (13, i), (13, i), final_bg))  # Final W
        style_commands.append(('BACKGROUND', (15, i), (16, i), cash_bg))  # Cash, Diesel
    
    # Bold Final W column
    style_commands.append(('FONTNAME', (13, 1), (13, -1), 'Helvetica-Bold'))
    
    main_table.setStyle(TableStyle(style_commands))
    elements.append(main_table)
    
    # Build PDF
    doc.build(elements)
    
    buffer.seek(0)
    filename = f"mill_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/export/truck-payments-excel")
async def export_truck_payments_excel(
    truck_no: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export truck payments to styled Excel file"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Build payment data
    payments_data = []
    total_net = 0
    total_paid = 0
    total_balance = 0
    
    for entry in entries:
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("final_w", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round(final_qntl * rate, 2)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions, 2)
        balance = round(max(0, net_amount - paid_amount), 2)
        status = "Paid" if balance < 0.10 else ("Partial" if paid_amount > 0 else "Pending")
        
        total_net += net_amount
        total_paid += paid_amount
        total_balance += balance
        
        payments_data.append({
            "date": entry.get("date", ""),
            "truck_no": entry.get("truck_no", ""),
            "mandi_name": entry.get("mandi_name", ""),
            "final_qntl": final_qntl,
            "rate": rate,
            "gross": gross_amount,
            "cash": cash_taken,
            "diesel": diesel_taken,
            "deductions": deductions,
            "net": net_amount,
            "paid": paid_amount,
            "balance": balance,
            "status": status
        })
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Truck Payments"
    
    # Styles
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    paid_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    pending_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:M1')
    company_name, tagline = await get_company_name()
    ws['A1'] = f"TRUCK PAYMENTS - {company_name} | KMS: {kms_year or 'All'} | {season or 'All'}"
    ws['A1'].font = Font(bold=True, size=14, color="D97706")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Date", "Truck No", "Mandi", "Final QNTL", "Rate", "Gross", "Cash", "Diesel", "Deductions", "Net Amount", "Paid", "Balance", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, p in enumerate(payments_data, 4):
        ws.cell(row=row_idx, column=1, value=p["date"])
        ws.cell(row=row_idx, column=2, value=p["truck_no"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value=p["mandi_name"])
        ws.cell(row=row_idx, column=4, value=p["final_qntl"])
        ws.cell(row=row_idx, column=5, value=f"₹{p['rate']}")
        ws.cell(row=row_idx, column=6, value=p["gross"])
        ws.cell(row=row_idx, column=7, value=p["cash"])
        ws.cell(row=row_idx, column=8, value=p["diesel"])
        ws.cell(row=row_idx, column=9, value=p["deductions"])
        ws.cell(row=row_idx, column=10, value=p["net"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=11, value=p["paid"])
        ws.cell(row=row_idx, column=12, value=p["balance"]).font = Font(bold=True, color="DC2626" if p["balance"] > 0 else "059669")
        status_cell = ws.cell(row=row_idx, column=13, value=p["status"])
        if p["status"] == "Paid":
            status_cell.fill = paid_fill
        elif p["status"] == "Pending":
            status_cell.fill = pending_fill
    
    # Totals row
    total_row = len(payments_data) + 4
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=10, value=round(total_net, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=11, value=round(total_paid, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=12, value=round(total_balance, 2)).font = Font(bold=True, color="DC2626")
    for col in range(1, 14):
        ws.cell(row=total_row, column=col).fill = total_fill
    
    # Column widths
    col_widths = [12, 14, 14, 12, 8, 10, 8, 8, 10, 12, 10, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"truck_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/export/truck-payments-pdf")
async def export_truck_payments_pdf(
    truck_no: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export truck payments to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.enums import TA_CENTER
    
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    if truck_no:
        query["truck_no"] = {"$regex": truck_no, "$options": "i"}
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Build payment data
    payments_data = []
    total_net = 0
    total_paid = 0
    total_balance = 0
    
    for entry in entries:
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("final_w", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round(final_qntl * rate, 2)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions, 2)
        balance = round(max(0, net_amount - paid_amount), 2)
        status = "Paid" if balance < 0.10 else ("Partial" if paid_amount > 0 else "Pending")
        
        total_net += net_amount
        total_paid += paid_amount
        total_balance += balance
        
        payments_data.append([
            entry.get("date", "")[:10],
            entry.get("truck_no", "")[:12],
            entry.get("mandi_name", "")[:12],
            f"{final_qntl}",
            f"Rs.{rate}",
            f"Rs.{gross_amount}",
            f"-Rs.{deductions}",
            f"Rs.{net_amount}",
            f"Rs.{paid_amount}",
            f"Rs.{balance}",
            status
        ])
    
    # Create PDF
    buffer = io.BytesIO()
    page_width, page_height = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.white, alignment=TA_CENTER)
    company_name, tagline = await get_company_name()
    title_data = [[Paragraph(f"<b>TRUCK PAYMENTS - {company_name} | KMS: {kms_year or 'All'} | {season or 'All'}</b>", title_style)]]
    title_table = Table(title_data, colWidths=[page_width - 20*mm])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D97706')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(title_table)
    elements.append(Table([[""]], colWidths=[page_width], rowHeights=[5*mm]))
    
    # Headers
    headers = ["Date", "Truck No", "Mandi", "QNTL", "Rate", "Gross", "Deduct", "Net", "Paid", "Balance", "Status"]
    table_data = [headers] + payments_data
    
    # Totals
    table_data.append(["TOTAL", "", "", "", "", "", "", f"Rs.{round(total_net, 2)}", f"Rs.{round(total_paid, 2)}", f"Rs.{round(total_balance, 2)}", ""])
    
    col_widths = [18*mm, 22*mm, 22*mm, 14*mm, 12*mm, 16*mm, 16*mm, 18*mm, 16*mm, 18*mm, 14*mm]
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]
    
    # Alternating rows and status colors
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
        if payments_data[i-1][-1] == "Paid":
            style_commands.append(('BACKGROUND', (-1, i), (-1, i), colors.HexColor('#D1FAE5')))
        elif payments_data[i-1][-1] == "Pending":
            style_commands.append(('BACKGROUND', (-1, i), (-1, i), colors.HexColor('#FEE2E2')))
    
    main_table.setStyle(TableStyle(style_commands))
    elements.append(main_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"truck_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.post("/entries/bulk-delete")
async def bulk_delete_entries(entry_ids: List[str], username: str = "", role: str = ""):
    """Bulk delete entries"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can bulk delete")
    
    result = await db.mill_entries.delete_many({"id": {"$in": entry_ids}})
    return {"message": f"{result.deleted_count} entries deleted successfully", "deleted_count": result.deleted_count}


# ============ MANDI TARGET ENDPOINTS ============

@api_router.get("/mandi-targets", response_model=List[MandiTarget])
async def get_mandi_targets(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get all mandi targets"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).sort("mandi_name", 1).to_list(100)
    return targets


@api_router.post("/mandi-targets", response_model=MandiTarget)
async def create_mandi_target(input: MandiTargetCreate, username: str = "", role: str = ""):
    """Create a new mandi target (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin target set kar sakta hai")
    
    # Check if target already exists for this mandi + kms_year + season
    existing = await db.mandi_targets.find_one({
        "mandi_name": input.mandi_name,
        "kms_year": input.kms_year,
        "season": input.season
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(status_code=400, detail=f"{input.mandi_name} ka target already set hai is KMS Year aur Season ke liye")
    
    # Calculate expected total
    expected_total = round(input.target_qntl + (input.target_qntl * input.cutting_percent / 100), 2)
    
    target_obj = MandiTarget(
        mandi_name=input.mandi_name,
        target_qntl=input.target_qntl,
        cutting_percent=input.cutting_percent,
        expected_total=expected_total,
        base_rate=input.base_rate,
        cutting_rate=input.cutting_rate,
        kms_year=input.kms_year,
        season=input.season,
        created_by=username
    )
    
    doc = target_obj.model_dump()
    await db.mandi_targets.insert_one(doc)
    return target_obj


@api_router.put("/mandi-targets/{target_id}", response_model=MandiTarget)
async def update_mandi_target(target_id: str, input: MandiTargetUpdate, username: str = "", role: str = ""):
    """Update a mandi target (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin target update kar sakta hai")
    
    existing = await db.mandi_targets.find_one({"id": target_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Target not found")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    merged = {**existing, **update_data}
    
    # Recalculate expected total
    merged["expected_total"] = round(merged["target_qntl"] + (merged["target_qntl"] * merged["cutting_percent"] / 100), 2)
    
    await db.mandi_targets.update_one({"id": target_id}, {"$set": merged})
    updated = await db.mandi_targets.find_one({"id": target_id}, {"_id": 0})
    return updated


@api_router.delete("/mandi-targets/{target_id}")
async def delete_mandi_target(target_id: str, username: str = "", role: str = ""):
    """Delete a mandi target (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin target delete kar sakta hai")
    
    result = await db.mandi_targets.delete_one({"id": target_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Target not found")
    return {"message": "Target deleted successfully"}


@api_router.get("/mandi-targets/summary", response_model=List[MandiTargetSummary])
async def get_mandi_target_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get mandi target vs achieved summary for dashboard"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    
    summaries = []
    for target in targets:
        # Get achieved sum for this mandi
        entry_query = {
            "mandi_name": target["mandi_name"],
            "kms_year": target["kms_year"],
            "season": target["season"]
        }
        
        pipeline = [
            {"$match": entry_query},
            {"$group": {"_id": None, "total_final_w": {"$sum": "$final_w"}}}
        ]
        
        result = await db.mill_entries.aggregate(pipeline).to_list(1)
        achieved_kg = result[0]["total_final_w"] if result else 0
        achieved_qntl = round(achieved_kg / 100, 2)
        
        expected_total = target["expected_total"]
        pending_qntl = round(max(0, expected_total - achieved_qntl), 2)
        progress_percent = round((achieved_qntl / expected_total * 100) if expected_total > 0 else 0, 1)
        
        # Calculate agent payment amounts
        target_qntl = target["target_qntl"]
        cutting_qntl = round(target_qntl * target["cutting_percent"] / 100, 2)
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        target_amount = round(target_qntl * base_rate, 2)
        cutting_amount = round(cutting_qntl * cutting_rate, 2)
        total_agent_amount = round(target_amount + cutting_amount, 2)
        
        summaries.append(MandiTargetSummary(
            id=target["id"],
            mandi_name=target["mandi_name"],
            target_qntl=target_qntl,
            cutting_percent=target["cutting_percent"],
            expected_total=expected_total,
            achieved_qntl=achieved_qntl,
            pending_qntl=pending_qntl,
            progress_percent=progress_percent,
            base_rate=base_rate,
            cutting_rate=cutting_rate,
            target_amount=target_amount,
            cutting_qntl=cutting_qntl,
            cutting_amount=cutting_amount,
            total_agent_amount=total_agent_amount,
            kms_year=target["kms_year"],
            season=target["season"]
        ))
    
    return summaries


# ============ DASHBOARD ENDPOINTS ============

@api_router.get("/dashboard/agent-totals")
async def get_agent_totals(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get agent-wise totals for bar chart"""
    match_query = {}
    if kms_year:
        match_query["kms_year"] = kms_year
    if season:
        match_query["season"] = season
    
    pipeline = []
    if match_query:
        pipeline.append({"$match": match_query})
    
    pipeline.extend([
        {
            "$group": {
                "_id": "$agent_name",
                "total_qntl": {"$sum": "$qntl"},
                "total_final_w_kg": {"$sum": "$final_w"},
                "total_entries": {"$sum": 1},
                "total_bag": {"$sum": "$bag"}
            }
        },
        {"$sort": {"total_final_w_kg": -1}}
    ])
    
    results = await db.mill_entries.aggregate(pipeline).to_list(50)
    
    agent_totals = []
    for r in results:
        if r["_id"]:  # Skip empty agent names
            agent_totals.append({
                "agent_name": r["_id"],
                "total_qntl": round(r["total_qntl"], 2),
                "total_final_w": round(r["total_final_w_kg"] / 100, 2),
                "total_entries": r["total_entries"],
                "total_bag": r["total_bag"]
            })
    
    return {"agent_totals": agent_totals}


@api_router.get("/dashboard/date-range-totals")
async def get_date_range_totals(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Get totals for a date range (for date filter reporting)"""
    match_query = {}
    
    if start_date and end_date:
        match_query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        match_query["date"] = {"$gte": start_date}
    elif end_date:
        match_query["date"] = {"$lte": end_date}
    
    if kms_year:
        match_query["kms_year"] = kms_year
    if season:
        match_query["season"] = season
    
    pipeline = []
    if match_query:
        pipeline.append({"$match": match_query})
    
    pipeline.append({
        "$group": {
            "_id": None,
            "total_kg": {"$sum": "$kg"},
            "total_qntl": {"$sum": "$qntl"},
            "total_bag": {"$sum": "$bag"},
            "total_final_w": {"$sum": "$final_w"},
            "total_entries": {"$sum": 1}
        }
    })
    
    result = await db.mill_entries.aggregate(pipeline).to_list(1)
    
    if result:
        data = result[0]
        return {
            "total_kg": round(data["total_kg"], 2),
            "total_qntl": round(data["total_qntl"], 2),
            "total_bag": data["total_bag"],
            "total_final_w": round(data["total_final_w"] / 100, 2),
            "total_entries": data["total_entries"],
            "start_date": start_date,
            "end_date": end_date
        }
    
    return {
        "total_kg": 0,
        "total_qntl": 0,
        "total_bag": 0,
        "total_final_w": 0,
        "total_entries": 0,
        "start_date": start_date,
        "end_date": end_date
    }


# ============ TRUCK PAYMENT ENDPOINTS ============

@api_router.get("/truck-payments", response_model=List[TruckPaymentStatus])
async def get_truck_payments(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get all truck payments with their status"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    payments = []
    for entry in entries:
        entry_id = entry.get("id")
        
        # Get payment record if exists
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        # Default rate 32, or from payment doc
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("final_w", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round(final_qntl * rate, 2)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions, 2)
        balance = round(net_amount - paid_amount, 2)
        
        # Use tolerance for floating-point precision (₹0.10 tolerance)
        status = "paid" if balance < 0.10 else ("partial" if paid_amount > 0 else "pending")
        
        payments.append(TruckPaymentStatus(
            entry_id=entry_id,
            truck_no=entry.get("truck_no", ""),
            date=entry.get("date", ""),
            total_qntl=round(entry.get("qntl", 0), 2),
            total_bag=entry.get("bag", 0),
            final_qntl=final_qntl,
            cash_taken=cash_taken,
            diesel_taken=diesel_taken,
            rate_per_qntl=rate,
            gross_amount=gross_amount,
            deductions=deductions,
            net_amount=net_amount,
            paid_amount=paid_amount,
            balance_amount=max(0, balance),
            status=status,
            kms_year=entry.get("kms_year", ""),
            season=entry.get("season", ""),
            agent_name=entry.get("agent_name", ""),
            mandi_name=entry.get("mandi_name", "")
        ))
    
    return payments


@api_router.put("/truck-payments/{entry_id}/rate")
async def set_truck_rate(entry_id: str, request: SetRateRequest, username: str = "", role: str = ""):
    """Set rate for a specific truck entry - auto-updates all entries with same truck_no + mandi_name"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin rate set kar sakta hai")
    
    entry = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    truck_no = entry.get("truck_no", "")
    mandi_name = entry.get("mandi_name", "")
    updated_count = 1
    
    if truck_no and mandi_name:
        # Find all entries with same truck_no + mandi_name
        matching = await db.mill_entries.find(
            {"truck_no": truck_no, "mandi_name": mandi_name}, {"_id": 0, "id": 1}
        ).to_list(None)
        for m in matching:
            await db.truck_payments.update_one(
                {"entry_id": m["id"]},
                {"$set": {"entry_id": m["id"], "rate_per_qntl": request.rate_per_qntl, "updated_at": datetime.now(timezone.utc).isoformat()}},
                upsert=True
            )
        updated_count = len(matching)
    else:
        await db.truck_payments.update_one(
            {"entry_id": entry_id},
            {"$set": {"entry_id": entry_id, "rate_per_qntl": request.rate_per_qntl, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
    
    return {"success": True, "message": f"Rate ₹{request.rate_per_qntl}/QNTL set for {updated_count} entries", "updated_count": updated_count, "truck_no": truck_no, "mandi_name": mandi_name}


@api_router.post("/truck-payments/{entry_id}/pay")
async def make_truck_payment(entry_id: str, request: MakePaymentRequest, username: str = "", role: str = ""):
    """Record a payment for truck (partial or full)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin payment kar sakta hai")
    
    # Check entry exists
    entry = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Get or create payment record
    payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
    current_paid = payment_doc.get("paid_amount", 0) if payment_doc else 0
    payments_history = payment_doc.get("payments_history", []) if payment_doc else []
    
    new_paid = current_paid + request.amount
    payments_history.append({
        "amount": request.amount,
        "date": datetime.now(timezone.utc).isoformat(),
        "note": request.note,
        "by": username
    })
    
    await db.truck_payments.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "entry_id": entry_id,
            "paid_amount": new_paid,
            "payments_history": payments_history,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"success": True, "message": f"₹{request.amount} payment recorded", "total_paid": new_paid}


@api_router.post("/truck-payments/{entry_id}/mark-paid")
async def mark_truck_paid(entry_id: str, username: str = "", role: str = ""):
    """Mark truck payment as fully paid"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin paid mark kar sakta hai")
    
    entry = await db.mill_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
    rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
    
    final_qntl = entry.get("final_w", 0) / 100
    cash_taken = entry.get("cash_paid", 0) or 0
    diesel_taken = entry.get("diesel_paid", 0) or 0
    net_amount = (final_qntl * rate) - cash_taken - diesel_taken
    
    payments_history = payment_doc.get("payments_history", []) if payment_doc else []
    payments_history.append({
        "amount": net_amount,
        "date": datetime.now(timezone.utc).isoformat(),
        "note": "Full payment - marked as paid",
        "by": username
    })
    
    await db.truck_payments.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "entry_id": entry_id,
            "paid_amount": net_amount,
            "payments_history": payments_history,
            "status": "paid",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"success": True, "message": "Truck payment cleared"}


# ============ AGENT PAYMENT ENDPOINTS ============

@api_router.get("/agent-payments", response_model=List[AgentPaymentStatus])
async def get_agent_payments(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get all agent payments based on mandi targets (not achieved)"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    # Get all mandi targets
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    
    payments = []
    for target in targets:
        mandi_name = target["mandi_name"]
        target_qntl = target["target_qntl"]
        cutting_percent = target["cutting_percent"]
        cutting_qntl = round(target_qntl * cutting_percent / 100, 2)
        expected_total = target["expected_total"]
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        
        # Calculate amounts
        target_amount = round(target_qntl * base_rate, 2)
        cutting_amount = round(cutting_qntl * cutting_rate, 2)
        total_amount = round(target_amount + cutting_amount, 2)
        
        # Get achieved for this mandi
        entry_query = {
            "mandi_name": mandi_name,
            "kms_year": target["kms_year"],
            "season": target["season"]
        }
        pipeline = [
            {"$match": entry_query},
            {"$group": {
                "_id": None, 
                "total_final_w": {"$sum": "$final_w"},
                "agent_name": {"$first": "$agent_name"}
            }}
        ]
        result = await db.mill_entries.aggregate(pipeline).to_list(1)
        achieved_kg = result[0]["total_final_w"] if result else 0
        achieved_qntl = round(achieved_kg / 100, 2)
        agent_name = result[0]["agent_name"] if result else mandi_name
        
        is_target_complete = achieved_qntl >= expected_total
        
        # Get payment record
        payment_doc = await db.agent_payments.find_one({
            "mandi_name": mandi_name,
            "kms_year": target["kms_year"],
            "season": target["season"]
        }, {"_id": 0})
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        balance = round(total_amount - paid_amount, 2)
        status = "paid" if balance <= 0 else ("partial" if paid_amount > 0 else "pending")
        
        payments.append(AgentPaymentStatus(
            mandi_name=mandi_name,
            agent_name=agent_name,
            target_qntl=target_qntl,
            cutting_percent=cutting_percent,
            cutting_qntl=cutting_qntl,
            base_rate=base_rate,
            cutting_rate=cutting_rate,
            target_amount=target_amount,
            cutting_amount=cutting_amount,
            total_amount=total_amount,
            achieved_qntl=achieved_qntl,
            is_target_complete=is_target_complete,
            paid_amount=paid_amount,
            balance_amount=max(0, balance),
            status=status,
            kms_year=target["kms_year"],
            season=target["season"]
        ))
    
    return payments


@api_router.post("/agent-payments/{mandi_name}/pay")
async def make_agent_payment(mandi_name: str, request: MakePaymentRequest, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Record a payment for agent/mandi (partial or full)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin payment kar sakta hai")
    
    # Get or create payment record
    payment_doc = await db.agent_payments.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    
    current_paid = payment_doc.get("paid_amount", 0) if payment_doc else 0
    payments_history = payment_doc.get("payments_history", []) if payment_doc else []
    
    new_paid = current_paid + request.amount
    payments_history.append({
        "amount": request.amount,
        "date": datetime.now(timezone.utc).isoformat(),
        "note": request.note,
        "by": username
    })
    
    await db.agent_payments.update_one(
        {"mandi_name": mandi_name, "kms_year": kms_year, "season": season},
        {"$set": {
            "mandi_name": mandi_name,
            "kms_year": kms_year,
            "season": season,
            "paid_amount": new_paid,
            "payments_history": payments_history,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"success": True, "message": f"₹{request.amount} payment recorded", "total_paid": new_paid}


@api_router.post("/agent-payments/{mandi_name}/mark-paid")
async def mark_agent_paid(mandi_name: str, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Mark agent/mandi payment as fully paid"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin paid mark kar sakta hai")
    
    # Get target for this mandi
    target = await db.mandi_targets.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    
    if not target:
        raise HTTPException(status_code=404, detail="Mandi target not found")
    
    # Calculate total amount based on target
    target_qntl = target["target_qntl"]
    cutting_qntl = target_qntl * target["cutting_percent"] / 100
    base_rate = target.get("base_rate", 10)
    cutting_rate = target.get("cutting_rate", 5)
    total_amount = (target_qntl * base_rate) + (cutting_qntl * cutting_rate)
    
    payment_doc = await db.agent_payments.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    payments_history = payment_doc.get("payments_history", []) if payment_doc else []
    payments_history.append({
        "amount": total_amount,
        "date": datetime.now(timezone.utc).isoformat(),
        "note": "Full payment - marked as paid",
        "by": username
    })
    
    await db.agent_payments.update_one(
        {"mandi_name": mandi_name, "kms_year": kms_year, "season": season},
        {"$set": {
            "mandi_name": mandi_name,
            "kms_year": kms_year,
            "season": season,
            "paid_amount": total_amount,
            "payments_history": payments_history,
            "status": "paid",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"success": True, "message": "Agent/Mandi payment cleared"}


@api_router.post("/truck-payments/{entry_id}/undo-paid")
async def undo_truck_paid(entry_id: str, username: str = "", role: str = ""):
    """Undo paid status - reset payment to 0 (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin undo kar sakta hai")
    
    payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
    if not payment_doc:
        raise HTTPException(status_code=404, detail="Payment record not found")
    
    payments_history = payment_doc.get("payments_history", [])
    payments_history.append({
        "amount": -payment_doc.get("paid_amount", 0),
        "date": datetime.now(timezone.utc).isoformat(),
        "note": "UNDO - Payment reversed",
        "by": username
    })
    
    await db.truck_payments.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "paid_amount": 0,
            "payments_history": payments_history,
            "status": "pending",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": "Payment undo ho gaya - status reset to pending"}


@api_router.post("/agent-payments/{mandi_name}/undo-paid")
async def undo_agent_paid(mandi_name: str, kms_year: str = "", season: str = "", username: str = "", role: str = ""):
    """Undo paid status - reset agent payment to 0 (Admin only)"""
    if role != "admin":
        raise HTTPException(status_code=403, detail="Sirf admin undo kar sakta hai")
    
    payment_doc = await db.agent_payments.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    
    if not payment_doc:
        raise HTTPException(status_code=404, detail="Payment record not found")
    
    payments_history = payment_doc.get("payments_history", [])
    payments_history.append({
        "amount": -payment_doc.get("paid_amount", 0),
        "date": datetime.now(timezone.utc).isoformat(),
        "note": "UNDO - Payment reversed",
        "by": username
    })
    
    await db.agent_payments.update_one(
        {"mandi_name": mandi_name, "kms_year": kms_year, "season": season},
        {"$set": {
            "paid_amount": 0,
            "payments_history": payments_history,
            "status": "pending",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": "Payment undo ho gaya - status reset to pending"}


@api_router.get("/truck-payments/{entry_id}/history")
async def get_truck_payment_history(entry_id: str):
    """Get payment history for a truck entry"""
    payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
    if not payment_doc:
        return {"history": [], "total_paid": 0}
    
    return {
        "history": payment_doc.get("payments_history", []),
        "total_paid": payment_doc.get("paid_amount", 0)
    }


@api_router.get("/agent-payments/{mandi_name}/history")
async def get_agent_payment_history(mandi_name: str, kms_year: str = "", season: str = ""):
    """Get payment history for an agent/mandi"""
    payment_doc = await db.agent_payments.find_one({
        "mandi_name": mandi_name,
        "kms_year": kms_year,
        "season": season
    }, {"_id": 0})
    
    if not payment_doc:
        return {"history": [], "total_paid": 0}
    
    return {
        "history": payment_doc.get("payments_history", []),
        "total_paid": payment_doc.get("paid_amount", 0)
    }


@api_router.get("/export/agent-payments-excel")
async def export_agent_payments_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Export agent/mandi payments to styled Excel file"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    
    payments_data = []
    total_amount_sum = 0
    total_paid_sum = 0
    total_balance_sum = 0
    
    for target in targets:
        mandi_name = target["mandi_name"]
        target_qntl = target["target_qntl"]
        cutting_qntl = round(target_qntl * target["cutting_percent"] / 100, 2)
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        
        target_amount = round(target_qntl * base_rate, 2)
        cutting_amount = round(cutting_qntl * cutting_rate, 2)
        total_amount = round(target_amount + cutting_amount, 2)
        
        # Get achieved
        entry_query = {"mandi_name": mandi_name, "kms_year": target["kms_year"], "season": target["season"]}
        pipeline = [{"$match": entry_query}, {"$group": {"_id": None, "total_final_w": {"$sum": "$final_w"}, "agent_name": {"$first": "$agent_name"}}}]
        result = await db.mill_entries.aggregate(pipeline).to_list(1)
        achieved_qntl = round(result[0]["total_final_w"] / 100, 2) if result else 0
        agent_name = result[0]["agent_name"] if result else mandi_name
        
        # Get payment
        payment_doc = await db.agent_payments.find_one({"mandi_name": mandi_name, "kms_year": target["kms_year"], "season": target["season"]}, {"_id": 0})
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        balance = round(max(0, total_amount - paid_amount), 2)
        status = "Paid" if balance <= 0 else ("Partial" if paid_amount > 0 else "Pending")
        
        total_amount_sum += total_amount
        total_paid_sum += paid_amount
        total_balance_sum += balance
        
        payments_data.append({
            "mandi_name": mandi_name,
            "agent_name": agent_name,
            "target_qntl": target_qntl,
            "cutting_qntl": cutting_qntl,
            "base_rate": base_rate,
            "cutting_rate": cutting_rate,
            "target_amount": target_amount,
            "cutting_amount": cutting_amount,
            "total_amount": total_amount,
            "achieved_qntl": achieved_qntl,
            "paid": paid_amount,
            "balance": balance,
            "status": status
        })
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Agent Payments"
    
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    paid_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    pending_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    ws.merge_cells('A1:M1')
    company_name, tagline = await get_company_name()
    ws['A1'] = f"AGENT/MANDI PAYMENTS - {company_name} | KMS: {kms_year or 'All'} | {season or 'All'}"
    ws['A1'].font = Font(bold=True, size=14, color="D97706")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ["Mandi", "Agent", "Target QNTL", "Cutting QNTL", "Base Rate", "Cut Rate", "Target Amt", "Cut Amt", "Total Amt", "Achieved", "Paid", "Balance", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, p in enumerate(payments_data, 4):
        ws.cell(row=row_idx, column=1, value=p["mandi_name"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=p["agent_name"])
        ws.cell(row=row_idx, column=3, value=p["target_qntl"])
        ws.cell(row=row_idx, column=4, value=p["cutting_qntl"])
        ws.cell(row=row_idx, column=5, value=f"₹{p['base_rate']}")
        ws.cell(row=row_idx, column=6, value=f"₹{p['cutting_rate']}")
        ws.cell(row=row_idx, column=7, value=p["target_amount"])
        ws.cell(row=row_idx, column=8, value=p["cutting_amount"])
        ws.cell(row=row_idx, column=9, value=p["total_amount"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=10, value=p["achieved_qntl"])
        ws.cell(row=row_idx, column=11, value=p["paid"])
        ws.cell(row=row_idx, column=12, value=p["balance"]).font = Font(bold=True, color="DC2626" if p["balance"] > 0 else "059669")
        status_cell = ws.cell(row=row_idx, column=13, value=p["status"])
        if p["status"] == "Paid":
            status_cell.fill = paid_fill
        elif p["status"] == "Pending":
            status_cell.fill = pending_fill
    
    total_row = len(payments_data) + 4
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=round(total_amount_sum, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=11, value=round(total_paid_sum, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=12, value=round(total_balance_sum, 2)).font = Font(bold=True, color="DC2626")
    for col in range(1, 14):
        ws.cell(row=total_row, column=col).fill = total_fill
    
    col_widths = [14, 12, 12, 12, 10, 10, 12, 10, 12, 10, 10, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"agent_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/export/agent-payments-pdf")
async def export_agent_payments_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Export agent/mandi payments to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.enums import TA_CENTER
    
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    
    payments_data = []
    total_amount_sum = 0
    total_paid_sum = 0
    total_balance_sum = 0
    
    for target in targets:
        mandi_name = target["mandi_name"]
        target_qntl = target["target_qntl"]
        cutting_qntl = round(target_qntl * target["cutting_percent"] / 100, 2)
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        
        target_amount = round(target_qntl * base_rate, 2)
        cutting_amount = round(cutting_qntl * cutting_rate, 2)
        total_amount = round(target_amount + cutting_amount, 2)
        
        entry_query = {"mandi_name": mandi_name, "kms_year": target["kms_year"], "season": target["season"]}
        pipeline = [{"$match": entry_query}, {"$group": {"_id": None, "total_final_w": {"$sum": "$final_w"}, "agent_name": {"$first": "$agent_name"}}}]
        result = await db.mill_entries.aggregate(pipeline).to_list(1)
        achieved_qntl = round(result[0]["total_final_w"] / 100, 2) if result else 0
        
        payment_doc = await db.agent_payments.find_one({"mandi_name": mandi_name, "kms_year": target["kms_year"], "season": target["season"]}, {"_id": 0})
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        balance = round(max(0, total_amount - paid_amount), 2)
        status = "Paid" if balance <= 0 else ("Partial" if paid_amount > 0 else "Pending")
        
        total_amount_sum += total_amount
        total_paid_sum += paid_amount
        total_balance_sum += balance
        
        payments_data.append([
            mandi_name[:12],
            f"{target_qntl}",
            f"{cutting_qntl}",
            f"Rs.{base_rate}+Rs.{cutting_rate}",
            f"Rs.{total_amount}",
            f"{achieved_qntl}",
            f"Rs.{paid_amount}",
            f"Rs.{balance}",
            status
        ])
    
    buffer = io.BytesIO()
    page_width, page_height = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.white, alignment=TA_CENTER)
    company_name, tagline = await get_company_name()
    title_data = [[Paragraph(f"<b>AGENT/MANDI PAYMENTS - {company_name} | KMS: {kms_year or 'All'} | {season or 'All'}</b>", title_style)]]
    title_table = Table(title_data, colWidths=[page_width - 20*mm])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D97706')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(title_table)
    elements.append(Table([[""]], colWidths=[page_width], rowHeights=[5*mm]))
    
    headers = ["Mandi", "Target", "Cutting", "Rates", "Total Amt", "Achieved", "Paid", "Balance", "Status"]
    table_data = [headers] + payments_data
    table_data.append(["TOTAL", "", "", "", f"Rs.{round(total_amount_sum, 2)}", "", f"Rs.{round(total_paid_sum, 2)}", f"Rs.{round(total_balance_sum, 2)}", ""])
    
    col_widths = [30*mm, 20*mm, 18*mm, 25*mm, 25*mm, 20*mm, 22*mm, 22*mm, 18*mm]
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]
    
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))
        if payments_data[i-1][-1] == "Paid":
            style_commands.append(('BACKGROUND', (-1, i), (-1, i), colors.HexColor('#D1FAE5')))
        elif payments_data[i-1][-1] == "Pending":
            style_commands.append(('BACKGROUND', (-1, i), (-1, i), colors.HexColor('#FEE2E2')))
    
    main_table.setStyle(TableStyle(style_commands))
    elements.append(main_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"agent_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/dashboard/monthly-trend")
async def get_monthly_trend(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Get monthly totals for trend chart"""
    match_query = {}
    if kms_year:
        match_query["kms_year"] = kms_year
    if season:
        match_query["season"] = season
    
    pipeline = []
    if match_query:
        pipeline.append({"$match": match_query})
    
    pipeline.extend([
        {
            "$addFields": {
                "month": {"$substr": ["$date", 0, 7]}  # Extract YYYY-MM
            }
        },
        {
            "$group": {
                "_id": "$month",
                "total_qntl": {"$sum": "$qntl"},
                "total_final_w_kg": {"$sum": "$final_w"},
                "total_entries": {"$sum": 1},
                "total_bag": {"$sum": "$bag"}
            }
        },
        {"$sort": {"_id": 1}}
    ])
    
    results = await db.mill_entries.aggregate(pipeline).to_list(12)
    
    monthly_data = []
    for r in results:
        if r["_id"]:
            monthly_data.append({
                "month": r["_id"],
                "total_qntl": round(r["total_qntl"], 2),
                "total_final_w": round(r["total_final_w_kg"] / 100, 2),
                "total_entries": r["total_entries"],
                "total_bag": r["total_bag"]
            })
    
    return {"monthly_data": monthly_data}


@api_router.get("/export/summary-report-pdf")
async def export_summary_report_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Export complete summary report - Truck Payments + Agent Payments + Targets"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    buffer = io.BytesIO()
    page_width, page_height = landscape(A4)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.white, alignment=TA_CENTER)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#D97706'), alignment=TA_LEFT, spaceBefore=10, spaceAfter=5)
    
    # Main Title
    title_data = [[Paragraph("<b>NAVKAR AGRO - COMPLETE SUMMARY REPORT</b>", title_style)]]
    title_table = Table(title_data, colWidths=[page_width - 20*mm])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D97706')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(title_table)
    
    # Subtitle
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#475569'), alignment=TA_CENTER)
    elements.append(Paragraph(f"KMS Year: {kms_year or 'All'} | Season: {season or 'All'} | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", sub_style))
    elements.append(Spacer(1, 10*mm))
    
    # ============ SECTION 1: MANDI TARGETS ============
    elements.append(Paragraph("MANDI TARGETS", section_style))
    
    targets = await db.mandi_targets.find(query, {"_id": 0}).to_list(100)
    if targets:
        target_headers = ["Mandi", "Target QNTL", "Cut %", "Expected", "Achieved", "Pending", "Progress"]
        target_data = [target_headers]
        
        for t in targets:
            entry_q = {"mandi_name": t["mandi_name"], "kms_year": t["kms_year"], "season": t["season"]}
            pipe = [{"$match": entry_q}, {"$group": {"_id": None, "total": {"$sum": "$final_w"}}}]
            res = await db.mill_entries.aggregate(pipe).to_list(1)
            achieved = round(res[0]["total"] / 100, 2) if res else 0
            expected = t["expected_total"]
            pending = round(max(0, expected - achieved), 2)
            progress = round((achieved / expected * 100) if expected > 0 else 0, 1)
            
            target_data.append([
                t["mandi_name"],
                str(t["target_qntl"]),
                f"{t['cutting_percent']}%",
                str(expected),
                str(achieved),
                str(pending),
                f"{progress}%"
            ])
        
        target_table = Table(target_data, colWidths=[35*mm, 25*mm, 18*mm, 25*mm, 25*mm, 25*mm, 22*mm])
        target_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ]))
        elements.append(target_table)
    else:
        elements.append(Paragraph("No targets set", styles['Normal']))
    
    elements.append(Spacer(1, 8*mm))
    
    # ============ SECTION 2: TRUCK PAYMENTS SUMMARY ============
    elements.append(Paragraph("TRUCK PAYMENTS (BHADA)", section_style))
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    truck_total_net = 0
    truck_total_paid = 0
    truck_total_balance = 0
    truck_data_rows = []
    
    for entry in entries:
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        paid = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("final_w", 0) / 100, 2)
        cash = entry.get("cash_paid", 0) or 0
        diesel = entry.get("diesel_paid", 0) or 0
        gross = round(final_qntl * rate, 2)
        net = round(gross - cash - diesel, 2)
        balance = round(max(0, net - paid), 2)
        status = "Paid" if balance < 0.10 else "Pending"
        
        truck_total_net += net
        truck_total_paid += paid
        truck_total_balance += balance
        
        truck_data_rows.append([
            entry.get("date", "")[:10],
            entry.get("truck_no", "")[:12],
            entry.get("mandi_name", "")[:12],
            f"{final_qntl}",
            f"Rs.{net}",
            f"Rs.{paid}",
            f"Rs.{balance}",
            status
        ])
    
    if truck_data_rows:
        truck_headers = ["Date", "Truck No", "Mandi", "QNTL", "Net Amt", "Paid", "Balance", "Status"]
        truck_table_data = [truck_headers] + truck_data_rows
        truck_table_data.append(["TOTAL", "", "", "", f"Rs.{round(truck_total_net, 2)}", f"Rs.{round(truck_total_paid, 2)}", f"Rs.{round(truck_total_balance, 2)}", ""])
        
        truck_table = Table(truck_table_data, colWidths=[20*mm, 25*mm, 25*mm, 18*mm, 25*mm, 22*mm, 22*mm, 18*mm])
        truck_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(truck_table)
    
    elements.append(Spacer(1, 8*mm))
    
    # ============ SECTION 3: AGENT PAYMENTS SUMMARY ============
    elements.append(Paragraph("AGENT/MANDI PAYMENTS", section_style))
    
    agent_total_amt = 0
    agent_total_paid = 0
    agent_total_balance = 0
    agent_data_rows = []
    
    for target in targets:
        mandi = target["mandi_name"]
        target_qntl = target["target_qntl"]
        cutting_qntl = round(target_qntl * target["cutting_percent"] / 100, 2)
        base_rate = target.get("base_rate", 10)
        cutting_rate = target.get("cutting_rate", 5)
        total_amt = round((target_qntl * base_rate) + (cutting_qntl * cutting_rate), 2)
        
        payment_doc = await db.agent_payments.find_one({"mandi_name": mandi, "kms_year": target["kms_year"], "season": target["season"]}, {"_id": 0})
        paid = payment_doc.get("paid_amount", 0) if payment_doc else 0
        balance = round(max(0, total_amt - paid), 2)
        status = "Paid" if balance <= 0 else "Pending"
        
        agent_total_amt += total_amt
        agent_total_paid += paid
        agent_total_balance += balance
        
        agent_data_rows.append([
            mandi,
            f"{target_qntl}",
            f"{cutting_qntl}",
            f"Rs.{base_rate}/Rs.{cutting_rate}",
            f"Rs.{total_amt}",
            f"Rs.{paid}",
            f"Rs.{balance}",
            status
        ])
    
    if agent_data_rows:
        agent_headers = ["Mandi", "Target", "Cutting", "Rates", "Total Amt", "Paid", "Balance", "Status"]
        agent_table_data = [agent_headers] + agent_data_rows
        agent_table_data.append(["TOTAL", "", "", "", f"Rs.{round(agent_total_amt, 2)}", f"Rs.{round(agent_total_paid, 2)}", f"Rs.{round(agent_total_balance, 2)}", ""])
        
        agent_table = Table(agent_table_data, colWidths=[30*mm, 20*mm, 20*mm, 30*mm, 25*mm, 22*mm, 22*mm, 18*mm])
        agent_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(agent_table)
    
    elements.append(Spacer(1, 8*mm))
    
    # ============ GRAND TOTAL ============
    elements.append(Paragraph("GRAND TOTAL", section_style))
    
    grand_total_amt = truck_total_net + agent_total_amt
    grand_total_paid = truck_total_paid + agent_total_paid
    grand_total_balance = truck_total_balance + agent_total_balance
    
    grand_data = [
        ["", "Total Amount", "Paid", "Balance"],
        ["Truck Payments", f"Rs.{round(truck_total_net, 2)}", f"Rs.{round(truck_total_paid, 2)}", f"Rs.{round(truck_total_balance, 2)}"],
        ["Agent Payments", f"Rs.{round(agent_total_amt, 2)}", f"Rs.{round(agent_total_paid, 2)}", f"Rs.{round(agent_total_balance, 2)}"],
        ["GRAND TOTAL", f"Rs.{round(grand_total_amt, 2)}", f"Rs.{round(grand_total_paid, 2)}", f"Rs.{round(grand_total_balance, 2)}"]
    ]
    
    grand_table = Table(grand_data, colWidths=[50*mm, 40*mm, 40*mm, 40*mm])
    grand_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D97706')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(grand_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"summary_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/export/truck-owner-excel")
async def export_truck_owner_excel(
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export truck owner consolidated payments to Excel"""
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Group by truck_no
    truck_data = {}
    for entry in entries:
        truck_no = entry.get("truck_no", "Unknown")
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("final_w", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round(final_qntl * rate, 2)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions, 2)
        balance = round(max(0, net_amount - paid_amount), 2)
        
        if truck_no not in truck_data:
            truck_data[truck_no] = {
                "truck_no": truck_no,
                "trips": 0,
                "total_qntl": 0,
                "total_gross": 0,
                "total_deductions": 0,
                "total_net": 0,
                "total_paid": 0,
                "total_balance": 0
            }
        
        truck_data[truck_no]["trips"] += 1
        truck_data[truck_no]["total_qntl"] += final_qntl
        truck_data[truck_no]["total_gross"] += gross_amount
        truck_data[truck_no]["total_deductions"] += deductions
        truck_data[truck_no]["total_net"] += net_amount
        truck_data[truck_no]["total_paid"] += paid_amount
        truck_data[truck_no]["total_balance"] += balance
    
    consolidated = list(truck_data.values())
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Truck Owner Payments"
    
    # Styles
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:I1')
    company_name, tagline = await get_company_name()
    ws['A1'] = f"TRUCK OWNER CONSOLIDATED PAYMENTS - {company_name} | KMS: {kms_year or 'All'} | {season or 'All'}"
    ws['A1'].font = Font(bold=True, size=14, color="0891B2")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:I2')
    ws['A2'] = "Ek truck ke saare trips ka combined payment"
    ws['A2'].font = Font(size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Truck No", "Total Trips", "Total QNTL", "Gross Amount", "Deductions", "Net Payable", "Paid", "Balance", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    grand_net = 0
    grand_paid = 0
    grand_balance = 0
    
    for row_idx, t in enumerate(consolidated, 5):
        status = "Paid" if t["total_balance"] < 0.10 else ("Partial" if t["total_paid"] > 0 else "Pending")
        
        ws.cell(row=row_idx, column=1, value=t["truck_no"]).font = Font(bold=True, size=11)
        ws.cell(row=row_idx, column=2, value=t["trips"]).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=3, value=round(t["total_qntl"], 2))
        ws.cell(row=row_idx, column=4, value=round(t["total_gross"], 2))
        ws.cell(row=row_idx, column=5, value=round(t["total_deductions"], 2))
        ws.cell(row=row_idx, column=6, value=round(t["total_net"], 2)).font = Font(bold=True)
        ws.cell(row=row_idx, column=7, value=round(t["total_paid"], 2))
        ws.cell(row=row_idx, column=8, value=round(t["total_balance"], 2)).font = Font(bold=True, color="DC2626" if t["total_balance"] > 0 else "059669")
        ws.cell(row=row_idx, column=9, value=status)
        
        grand_net += t["total_net"]
        grand_paid += t["total_paid"]
        grand_balance += t["total_balance"]
    
    # Grand Total row
    total_row = len(consolidated) + 5
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(consolidated)).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=6, value=round(grand_net, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=round(grand_paid, 2)).font = Font(bold=True, color="059669")
    ws.cell(row=total_row, column=8, value=round(grand_balance, 2)).font = Font(bold=True, color="DC2626")
    
    for col in range(1, 10):
        ws.cell(row=total_row, column=col).fill = total_fill
    
    # Column widths
    widths = [15, 12, 12, 14, 12, 14, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"truck_owner_consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/export/truck-owner-pdf")
async def export_truck_owner_pdf(
    kms_year: Optional[str] = None,
    season: Optional[str] = None
):
    """Export truck owner consolidated payments to PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    query = {}
    if kms_year:
        query["kms_year"] = kms_year
    if season:
        query["season"] = season
    
    entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Group by truck_no
    truck_data = {}
    for entry in entries:
        truck_no = entry.get("truck_no", "Unknown")
        entry_id = entry.get("id")
        payment_doc = await db.truck_payments.find_one({"entry_id": entry_id}, {"_id": 0})
        
        rate = payment_doc.get("rate_per_qntl", 32) if payment_doc else 32
        paid_amount = payment_doc.get("paid_amount", 0) if payment_doc else 0
        
        final_qntl = round(entry.get("final_w", 0) / 100, 2)
        cash_taken = entry.get("cash_paid", 0) or 0
        diesel_taken = entry.get("diesel_paid", 0) or 0
        
        gross_amount = round(final_qntl * rate, 2)
        deductions = cash_taken + diesel_taken
        net_amount = round(gross_amount - deductions, 2)
        balance = round(max(0, net_amount - paid_amount), 2)
        
        if truck_no not in truck_data:
            truck_data[truck_no] = {
                "truck_no": truck_no,
                "trips": 0,
                "total_qntl": 0,
                "total_gross": 0,
                "total_deductions": 0,
                "total_net": 0,
                "total_paid": 0,
                "total_balance": 0
            }
        
        truck_data[truck_no]["trips"] += 1
        truck_data[truck_no]["total_qntl"] += final_qntl
        truck_data[truck_no]["total_gross"] += gross_amount
        truck_data[truck_no]["total_deductions"] += deductions
        truck_data[truck_no]["total_net"] += net_amount
        truck_data[truck_no]["total_paid"] += paid_amount
        truck_data[truck_no]["total_balance"] += balance
    
    consolidated = list(truck_data.values())
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#0891B2'), alignment=1)
    company_name, tagline = await get_company_name()
    elements.append(Paragraph(f"TRUCK OWNER CONSOLIDATED PAYMENTS - {company_name}", title_style))
    elements.append(Paragraph(f"{company_name} - {tagline}", ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1)))
    elements.append(Paragraph(f"KMS Year: {kms_year or 'All'} | Season: {season or 'All'} | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", ParagraphStyle('Info', parent=styles['Normal'], fontSize=9, alignment=1, textColor=colors.gray)))
    elements.append(Spacer(1, 20))
    
    # Table
    table_data = [["Truck No", "Trips", "Total QNTL", "Gross", "Deductions", "Net Payable", "Paid", "Balance", "Status"]]
    
    grand_net = 0
    grand_paid = 0
    grand_balance = 0
    
    for t in consolidated:
        status = "PAID" if t["total_balance"] < 0.10 else ("PARTIAL" if t["total_paid"] > 0 else "PENDING")
        table_data.append([
            t["truck_no"],
            str(t["trips"]),
            f"{t['total_qntl']:.2f}",
            f"Rs.{t['total_gross']:.0f}",
            f"Rs.{t['total_deductions']:.0f}",
            f"Rs.{t['total_net']:.0f}",
            f"Rs.{t['total_paid']:.0f}",
            f"Rs.{t['total_balance']:.0f}",
            status
        ])
        grand_net += t["total_net"]
        grand_paid += t["total_paid"]
        grand_balance += t["total_balance"]
    
    # Grand Total
    table_data.append([
        "GRAND TOTAL",
        str(len(consolidated)),
        "",
        "",
        "",
        f"Rs.{grand_net:.0f}",
        f"Rs.{grand_paid:.0f}",
        f"Rs.{grand_balance:.0f}",
        ""
    ])
    
    table = Table(table_data, colWidths=[80, 50, 70, 70, 70, 80, 70, 70, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0891B2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"truck_owner_consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============ MILLING ENTRY CRUD APIs ============

def calculate_milling_fields(data: dict) -> dict:
    """Auto-calculate QNTL values from percentages and paddy input. FRK from stock."""
    paddy = data.get('paddy_input_qntl', 0) or 0
    
    rice_pct = data.get('rice_percent', 0) or 0
    bran_pct = data.get('bran_percent', 0) or 0
    kunda_pct = data.get('kunda_percent', 0) or 0
    broken_pct = data.get('broken_percent', 0) or 0
    kanki_pct = data.get('kanki_percent', 0) or 0
    
    used_pct = rice_pct + bran_pct + kunda_pct + broken_pct + kanki_pct
    husk_pct = max(0, round(100 - used_pct, 2))
    
    data['husk_percent'] = husk_pct
    data['rice_qntl'] = round(paddy * rice_pct / 100, 2)
    data['bran_qntl'] = round(paddy * bran_pct / 100, 2)
    data['kunda_qntl'] = round(paddy * kunda_pct / 100, 2)
    data['broken_qntl'] = round(paddy * broken_pct / 100, 2)
    data['kanki_qntl'] = round(paddy * kanki_pct / 100, 2)
    data['husk_qntl'] = round(paddy * husk_pct / 100, 2)
    
    frk_used = data.get('frk_used_qntl', 0) or 0
    data['cmr_delivery_qntl'] = round(data['rice_qntl'] + frk_used, 2)
    data['outturn_ratio'] = round(data['cmr_delivery_qntl'] / paddy * 100, 2) if paddy > 0 else 0
    
    return data


@api_router.post("/milling-entries")
async def create_milling_entry(input: MillingEntryCreate, username: str = "", role: str = ""):
    entry_dict = input.model_dump()
    entry_dict = calculate_milling_fields(entry_dict)
    entry_dict['created_by'] = username
    entry_obj = MillingEntry(**entry_dict)
    doc = entry_obj.model_dump()
    await db.milling_entries.insert_one(doc)
    doc.pop('_id', None)
    return doc


@api_router.get("/milling-entries")
async def get_milling_entries(
    rice_type: Optional[str] = None, kms_year: Optional[str] = None,
    season: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None
):
    query = {}
    if rice_type: query["rice_type"] = rice_type
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if date_from or date_to:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to: dq["$lte"] = date_to
        if dq: query["date"] = dq
    return await db.milling_entries.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.get("/milling-entries/{entry_id}")
async def get_milling_entry(entry_id: str):
    entry = await db.milling_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry: raise HTTPException(status_code=404, detail="Milling entry not found")
    return entry


@api_router.put("/milling-entries/{entry_id}")
async def update_milling_entry(entry_id: str, input: MillingEntryCreate, username: str = "", role: str = ""):
    existing = await db.milling_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Milling entry not found")
    update_dict = input.model_dump()
    update_dict = calculate_milling_fields(update_dict)
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.milling_entries.update_one({"id": entry_id}, {"$set": update_dict})
    return await db.milling_entries.find_one({"id": entry_id}, {"_id": 0})


@api_router.delete("/milling-entries/{entry_id}")
async def delete_milling_entry(entry_id: str, username: str = "", role: str = ""):
    existing = await db.milling_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Milling entry not found")
    await db.milling_entries.delete_one({"id": entry_id})
    return {"message": "Milling entry deleted", "id": entry_id}


@api_router.get("/paddy-stock")
async def get_paddy_stock(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    mill_entries = await db.mill_entries.find(query, {"mill_w": 1, "_id": 0}).to_list(10000)
    total_paddy_in = round(sum(e.get('mill_w', 0) for e in mill_entries) / 100, 2)
    milling_entries = await db.milling_entries.find(query, {"paddy_input_qntl": 1, "_id": 0}).to_list(10000)
    total_paddy_used = round(sum(e.get('paddy_input_qntl', 0) for e in milling_entries), 2)
    return {"total_paddy_in_qntl": total_paddy_in, "total_paddy_used_qntl": total_paddy_used, "available_paddy_qntl": round(total_paddy_in - total_paddy_used, 2)}


@api_router.get("/milling-summary")
async def get_milling_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.milling_entries.find(query, {"_id": 0}).to_list(1000)
    
    total_paddy = sum(e.get('paddy_input_qntl', 0) for e in entries)
    total_rice = sum(e.get('rice_qntl', 0) for e in entries)
    total_frk = sum(e.get('frk_used_qntl', 0) for e in entries)
    total_bran = sum(e.get('bran_qntl', 0) for e in entries)
    total_kunda = sum(e.get('kunda_qntl', 0) for e in entries)
    total_broken = sum(e.get('broken_qntl', 0) for e in entries)
    total_kanki = sum(e.get('kanki_qntl', 0) for e in entries)
    total_husk = sum(e.get('husk_qntl', 0) for e in entries)
    total_cmr = sum(e.get('cmr_delivery_qntl', 0) for e in entries)
    avg_outturn = round(total_cmr / total_paddy * 100, 2) if total_paddy > 0 else 0
    
    def type_summary(elist):
        tp = sum(e.get('paddy_input_qntl', 0) for e in elist)
        tr = sum(e.get('rice_qntl', 0) for e in elist)
        tf = sum(e.get('frk_used_qntl', 0) for e in elist)
        tc = sum(e.get('cmr_delivery_qntl', 0) for e in elist)
        return {"count": len(elist), "total_paddy_qntl": round(tp, 2), "total_rice_qntl": round(tr, 2),
            "total_frk_qntl": round(tf, 2), "total_cmr_qntl": round(tc, 2),
            "avg_outturn": round(tc / tp * 100, 2) if tp > 0 else 0}
    
    return {"total_entries": len(entries), "total_paddy_qntl": round(total_paddy, 2),
        "total_rice_qntl": round(total_rice, 2), "total_frk_qntl": round(total_frk, 2),
        "total_bran_qntl": round(total_bran, 2), "total_kunda_qntl": round(total_kunda, 2),
        "total_broken_qntl": round(total_broken, 2), "total_kanki_qntl": round(total_kanki, 2),
        "total_husk_qntl": round(total_husk, 2), "total_cmr_qntl": round(total_cmr, 2),
        "avg_outturn_ratio": avg_outturn,
        "parboiled": type_summary([e for e in entries if e.get('rice_type') == 'parboiled']),
        "raw": type_summary([e for e in entries if e.get('rice_type') == 'raw'])}


# ============ FRK PURCHASE APIs ============

@api_router.post("/frk-purchases")
async def create_frk_purchase(input: FrkPurchaseCreate, username: str = "", role: str = ""):
    d = input.model_dump()
    d['total_amount'] = round((d.get('quantity_qntl', 0) or 0) * (d.get('rate_per_qntl', 0) or 0), 2)
    d['created_by'] = username
    obj = FrkPurchase(**d)
    doc = obj.model_dump()
    await db.frk_purchases.insert_one(doc)
    doc.pop('_id', None)
    return doc


@api_router.get("/frk-purchases")
async def get_frk_purchases(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    return await db.frk_purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.delete("/frk-purchases/{purchase_id}")
async def delete_frk_purchase(purchase_id: str, username: str = "", role: str = ""):
    existing = await db.frk_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="FRK purchase not found")
    await db.frk_purchases.delete_one({"id": purchase_id})
    return {"message": "FRK purchase deleted", "id": purchase_id}


@api_router.get("/frk-stock")
async def get_frk_stock(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(1000)
    total_purchased = round(sum(p.get('quantity_qntl', 0) for p in purchases), 2)
    total_cost = round(sum(p.get('total_amount', 0) for p in purchases), 2)
    milling_entries = await db.milling_entries.find(query, {"frk_used_qntl": 1, "_id": 0}).to_list(1000)
    total_used = round(sum(e.get('frk_used_qntl', 0) for e in milling_entries), 2)
    return {"total_purchased_qntl": total_purchased, "total_used_qntl": total_used,
        "available_qntl": round(total_purchased - total_used, 2), "total_cost": total_cost}


# ============ BY-PRODUCT STOCK & SALE APIs ============

@api_router.get("/byproduct-stock")
async def get_byproduct_stock(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    milling_entries = await db.milling_entries.find(query, {"_id": 0}).to_list(1000)
    sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(1000)
    products = ["bran", "kunda", "broken", "kanki", "husk"]
    stock = {}
    for p in products:
        produced = round(sum(e.get(f'{p}_qntl', 0) for e in milling_entries), 2)
        sold = round(sum(s.get('quantity_qntl', 0) for s in sales if s.get('product') == p), 2)
        revenue = round(sum(s.get('total_amount', 0) for s in sales if s.get('product') == p), 2)
        stock[p] = {"produced_qntl": produced, "sold_qntl": sold, "available_qntl": round(produced - sold, 2), "total_revenue": revenue}
    return stock


@api_router.post("/byproduct-sales")
async def create_byproduct_sale(input: ByProductSaleCreate, username: str = "", role: str = ""):
    d = input.model_dump()
    d['total_amount'] = round((d.get('quantity_qntl', 0) or 0) * (d.get('rate_per_qntl', 0) or 0), 2)
    d['created_by'] = username
    obj = ByProductSale(**d)
    doc = obj.model_dump()
    await db.byproduct_sales.insert_one(doc)
    doc.pop('_id', None)
    return doc


@api_router.get("/byproduct-sales")
async def get_byproduct_sales(product: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if product: query["product"] = product
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    return await db.byproduct_sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.delete("/byproduct-sales/{sale_id}")
async def delete_byproduct_sale(sale_id: str, username: str = "", role: str = ""):
    existing = await db.byproduct_sales.find_one({"id": sale_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Sale entry not found")
    await db.byproduct_sales.delete_one({"id": sale_id})
    return {"message": "Sale entry deleted", "id": sale_id}


# ============ PADDY CUSTODY MAINTENANCE REGISTER ============

@api_router.get("/paddy-custody-register")
async def get_paddy_custody_register(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Paddy custody register - all movements: received (mill entries) and released (milling entries)"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    
    # Paddy received from mill entries
    mill_entries = await db.mill_entries.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    # Paddy released for milling
    milling_entries = await db.milling_entries.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    
    # Build register rows
    rows = []
    for e in mill_entries:
        rows.append({
            "date": e.get('date', ''),
            "type": "received",
            "description": f"Truck: {e.get('truck_no', '')} | Agent: {e.get('agent_name', '')} | Mandi: {e.get('mandi_name', '')}",
            "received_qntl": round(e.get('mill_w', 0) / 100, 2),
            "issued_qntl": 0,
            "source_id": e.get('id', '')
        })
    for e in milling_entries:
        rows.append({
            "date": e.get('date', ''),
            "type": "issued",
            "description": f"Milling ({e.get('rice_type', 'parboiled').title()}) | Rice: {e.get('rice_qntl', 0)}Q",
            "received_qntl": 0,
            "issued_qntl": e.get('paddy_input_qntl', 0),
            "source_id": e.get('id', '')
        })
    
    # Sort by date
    rows.sort(key=lambda x: x['date'])
    
    # Add running balance
    balance = 0
    for r in rows:
        balance += r['received_qntl'] - r['issued_qntl']
        r['balance_qntl'] = round(balance, 2)
    
    return {"rows": rows, "total_received": round(sum(r['received_qntl'] for r in rows), 2),
        "total_issued": round(sum(r['issued_qntl'] for r in rows), 2),
        "final_balance": round(balance, 2)}


# ============ MILLING REPORT EXPORT ============

@api_router.get("/milling-report/excel")
async def export_milling_report_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.milling_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Milling Report"
    
    # Header
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    title = f"Milling Report"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    ws.merge_cells('A1:L1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Date', 'Type', 'Paddy (Q)', 'Rice %', 'Rice (Q)', 'FRK Used (Q)', 'CMR (Q)', 'Outturn %', 'Bran (Q)', 'Kunda (Q)', 'Husk %', 'Note']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for i, e in enumerate(entries, 4):
        vals = [e.get('date',''), e.get('rice_type','').title(), e.get('paddy_input_qntl',0), e.get('rice_percent',0),
            e.get('rice_qntl',0), e.get('frk_used_qntl',0), e.get('cmr_delivery_qntl',0), e.get('outturn_ratio',0),
            e.get('bran_qntl',0), e.get('kunda_qntl',0), e.get('husk_percent',0), e.get('note','')]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            if col >= 3: cell.alignment = Alignment(horizontal='right')
    
    # Totals row
    tr = len(entries) + 4
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    if entries:
        for col, key in [(3,'paddy_input_qntl'),(5,'rice_qntl'),(6,'frk_used_qntl'),(7,'cmr_delivery_qntl'),(9,'bran_qntl'),(10,'kunda_qntl')]:
            ws.cell(row=tr, column=col, value=round(sum(e.get(key,0) for e in entries),2)).font = Font(bold=True)
    
    from openpyxl.utils import get_column_letter as gcl
    for i in range(1, 13):  # 12 columns
        ws.column_dimensions[gcl(i)].width = 14
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fn = f"milling_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@api_router.get("/milling-report/pdf")
async def export_milling_report_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.milling_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    title = "Milling Report"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    
    headers = ['Date', 'Type', 'Paddy(Q)', 'Rice%', 'Rice(Q)', 'FRK(Q)', 'CMR(Q)', 'Outturn%', 'Bran(Q)', 'Kunda(Q)', 'Husk%']
    data = [headers]
    tp = tr = tf = tc = tb = tk = 0
    for e in entries:
        tp += e.get('paddy_input_qntl',0); tr += e.get('rice_qntl',0); tf += e.get('frk_used_qntl',0)
        tc += e.get('cmr_delivery_qntl',0); tb += e.get('bran_qntl',0); tk += e.get('kunda_qntl',0)
        data.append([e.get('date',''), e.get('rice_type','').title()[:3], e.get('paddy_input_qntl',0),
            f"{e.get('rice_percent',0)}%", e.get('rice_qntl',0), e.get('frk_used_qntl',0),
            e.get('cmr_delivery_qntl',0), f"{e.get('outturn_ratio',0)}%", e.get('bran_qntl',0), e.get('kunda_qntl',0), f"{e.get('husk_percent',0)}%"])
    data.append(['TOTAL', '', round(tp,2), '', round(tr,2), round(tf,2), round(tc,2), '', round(tb,2), round(tk,2), ''])
    
    col_widths = [65, 35, 55, 40, 50, 45, 50, 55, 45, 50, 40]
    table = RLTable(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f8f8f8')]),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    fn = f"milling_report_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@api_router.get("/paddy-custody-register/excel")
async def export_paddy_custody_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    register = await get_paddy_custody_register(kms_year=kms_year, season=season)
    rows = register['rows']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Paddy Custody Register"
    
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    title = "Paddy Custody Maintenance Register"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Date', 'Description', 'Received (QNTL)', 'Released (QNTL)', 'Balance (QNTL)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for i, r in enumerate(rows, 4):
        vals = [r['date'], r['description'], r['received_qntl'] if r['received_qntl'] > 0 else '',
            r['issued_qntl'] if r['issued_qntl'] > 0 else '', r['balance_qntl']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            if col >= 3: cell.alignment = Alignment(horizontal='right')
            if r['type'] == 'received': cell.font = Font(color="006600")
            elif r['type'] == 'issued': cell.font = Font(color="CC0000")
    
    tr = len(rows) + 4
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tr, column=3, value=register['total_received']).font = Font(bold=True)
    ws.cell(row=tr, column=4, value=register['total_issued']).font = Font(bold=True)
    ws.cell(row=tr, column=5, value=register['final_balance']).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 16; ws.column_dimensions['D'].width = 16; ws.column_dimensions['E'].width = 16
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fn = f"paddy_custody_register_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})


@api_router.get("/paddy-custody-register/pdf")
async def export_paddy_custody_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    
    register = await get_paddy_custody_register(kms_year=kms_year, season=season)
    rows = register['rows']
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    title = "Paddy Custody Maintenance Register"
    if kms_year: title += f" - KMS {kms_year}"
    if season: title += f" ({season})"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    
    data = [['Date', 'Description', 'Received (Q)', 'Released (Q)', 'Balance (Q)']]
    for r in rows:
        data.append([r['date'], r['description'][:60], r['received_qntl'] if r['received_qntl'] > 0 else '-',
            r['issued_qntl'] if r['issued_qntl'] > 0 else '-', r['balance_qntl']])
    data.append(['TOTAL', '', register['total_received'], register['total_issued'], register['final_balance']])
    
    table = RLTable(data, colWidths=[65, 300, 70, 70, 70], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 7), ('FONTSIZE', (0,0), (-1,0), 8),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'), ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f8f8f8')]),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=paddy_custody_{datetime.now().strftime('%Y%m%d')}.pdf"})


@api_router.get("/frk-purchases/excel")
async def export_frk_purchases_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    purchases = await db.frk_purchases.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    wb = Workbook(); ws = wb.active; ws.title = "FRK Purchases"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    title = "FRK Purchase Register"
    if kms_year: title += f" - KMS {kms_year}"
    ws.merge_cells('A1:F1'); ws['A1'] = title; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    
    for col, h in enumerate(['Date', 'Party Name', 'Qty (QNTL)', 'Rate (₹/Q)', 'Amount (₹)', 'Note'], 1):
        c = ws.cell(row=3, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb; c.alignment = Alignment(horizontal='center')
    
    for i, p in enumerate(purchases, 4):
        for col, v in enumerate([p.get('date',''), p.get('party_name',''), p.get('quantity_qntl',0), p.get('rate_per_qntl',0), p.get('total_amount',0), p.get('note','')], 1):
            c = ws.cell(row=i, column=col, value=v); c.border = tb
            if col >= 3: c.alignment = Alignment(horizontal='right')
    
    tr = len(purchases) + 4
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tr, column=3, value=round(sum(p.get('quantity_qntl',0) for p in purchases),2)).font = Font(bold=True)
    ws.cell(row=tr, column=5, value=round(sum(p.get('total_amount',0) for p in purchases),2)).font = Font(bold=True)
    for letter in ['A','B','C','D','E','F']: ws.column_dimensions[letter].width = 16
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=frk_purchases_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/frk-purchases/pdf")
async def export_frk_purchases_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    purchases = await db.frk_purchases.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    title = "FRK Purchase Register"
    if kms_year: title += f" - KMS {kms_year}"
    elements.append(Paragraph(title, styles['Title'])); elements.append(Spacer(1, 12))
    
    data = [['Date', 'Party', 'Qty(Q)', 'Rate(₹)', 'Amount(₹)', 'Note']]
    tq = ta = 0
    for p in purchases:
        tq += p.get('quantity_qntl',0); ta += p.get('total_amount',0)
        data.append([p.get('date',''), p.get('party_name','')[:25], p.get('quantity_qntl',0), p.get('rate_per_qntl',0), p.get('total_amount',0), p.get('note','')[:20]])
    data.append(['TOTAL', '', round(tq,2), '', round(ta,2), ''])
    
    table = RLTable(data, colWidths=[60, 120, 55, 55, 70, 80], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 7), ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=frk_purchases_{datetime.now().strftime('%Y%m%d')}.pdf"})


@api_router.get("/byproduct-sales/excel")
async def export_byproduct_sales_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    sales = await db.byproduct_sales.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    stock_data = await get_byproduct_stock(kms_year=kms_year, season=season)
    
    wb = Workbook(); ws = wb.active; ws.title = "By-Product Sales"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    title = "By-Product Stock & Sales Report"
    if kms_year: title += f" - KMS {kms_year}"
    ws.merge_cells('A1:G1'); ws['A1'] = title; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    
    # Stock summary section
    ws.cell(row=3, column=1, value="Stock Summary").font = Font(bold=True, size=11)
    for col, h in enumerate(['Product', 'Produced (Q)', 'Sold (Q)', 'Available (Q)', 'Revenue (₹)'], 1):
        c = ws.cell(row=4, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    row = 5
    for prod, label in [('bran','Bran'), ('kunda','Kunda'), ('broken','Broken'), ('kanki','Kanki'), ('husk','Husk')]:
        s = stock_data.get(prod, {})
        for col, v in enumerate([label, s.get('produced_qntl',0), s.get('sold_qntl',0), s.get('available_qntl',0), s.get('total_revenue',0)], 1):
            c = ws.cell(row=row, column=col, value=v); c.border = tb
            if col >= 2: c.alignment = Alignment(horizontal='right')
        row += 1
    
    # Sales detail section
    row += 1
    ws.cell(row=row, column=1, value="Sales Detail").font = Font(bold=True, size=11)
    row += 1
    for col, h in enumerate(['Date', 'Product', 'Qty (Q)', 'Rate (₹/Q)', 'Amount (₹)', 'Buyer', 'Note'], 1):
        c = ws.cell(row=row, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    row += 1
    for s in sales:
        for col, v in enumerate([s.get('date',''), s.get('product','').title(), s.get('quantity_qntl',0), s.get('rate_per_qntl',0), s.get('total_amount',0), s.get('buyer_name',''), s.get('note','')], 1):
            c = ws.cell(row=row, column=col, value=v); c.border = tb
            if col >= 3 and col <= 5: c.alignment = Alignment(horizontal='right')
        row += 1
    
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(sum(s.get('quantity_qntl',0) for s in sales),2)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(sum(s.get('total_amount',0) for s in sales),2)).font = Font(bold=True)
    for letter in ['A','B','C','D','E','F','G']: ws.column_dimensions[letter].width = 16
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=byproduct_sales_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/byproduct-sales/pdf")
async def export_byproduct_sales_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    sales = await db.byproduct_sales.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    stock_data = await get_byproduct_stock(kms_year=kms_year, season=season)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    title = "By-Product Stock & Sales Report"
    if kms_year: title += f" - KMS {kms_year}"
    elements.append(Paragraph(title, styles['Title'])); elements.append(Spacer(1, 12))
    
    # Stock summary table
    elements.append(Paragraph("Stock Summary", styles['Heading2'])); elements.append(Spacer(1, 6))
    sdata = [['Product', 'Produced(Q)', 'Sold(Q)', 'Available(Q)', 'Revenue(₹)']]
    for prod, label in [('bran','Bran'), ('kunda','Kunda'), ('broken','Broken'), ('kanki','Kanki'), ('husk','Husk')]:
        s = stock_data.get(prod, {})
        sdata.append([label, s.get('produced_qntl',0), s.get('sold_qntl',0), s.get('available_qntl',0), s.get('total_revenue',0)])
    st = RLTable(sdata, colWidths=[70, 70, 60, 70, 70])
    st.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 8), ('ALIGN', (1,0), (-1,-1), 'RIGHT'), ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')]))
    elements.append(st); elements.append(Spacer(1, 15))
    
    # Sales table
    elements.append(Paragraph("Sales Detail", styles['Heading2'])); elements.append(Spacer(1, 6))
    data = [['Date', 'Product', 'Qty(Q)', 'Rate(₹)', 'Amount(₹)', 'Buyer']]
    tq = ta = 0
    for s in sales:
        tq += s.get('quantity_qntl',0); ta += s.get('total_amount',0)
        data.append([s.get('date',''), s.get('product','').title(), s.get('quantity_qntl',0), s.get('rate_per_qntl',0), s.get('total_amount',0), s.get('buyer_name','')[:20]])
    data.append(['TOTAL', '', round(tq,2), '', round(ta,2), ''])
    
    table = RLTable(data, colWidths=[55, 55, 45, 50, 60, 90], repeatRows=1)
    table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 7), ('ALIGN', (2,0), (-1,-1), 'RIGHT'), ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')]))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=byproduct_sales_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ CASH BOOK / DAILY CASH & BANK REGISTER ============

class CashTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    account: str  # "cash" or "bank"
    txn_type: str  # "jama" (credit/in) or "nikasi" (debit/out)
    category: str = ""
    description: str = ""
    amount: float = 0
    reference: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


@api_router.post("/cash-book")
async def add_cash_transaction(txn: CashTransaction, username: str = "", role: str = ""):
    txn_dict = txn.model_dump()
    txn_dict['created_by'] = username
    txn_dict['amount'] = round(txn_dict['amount'], 2)
    await db.cash_transactions.insert_one(txn_dict)
    txn_dict.pop('_id', None)
    return txn_dict


@api_router.get("/cash-book")
async def get_cash_transactions(kms_year: Optional[str] = None, season: Optional[str] = None,
                                 account: Optional[str] = None, date_from: Optional[str] = None,
                                 date_to: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if account: query["account"] = account
    if date_from or date_to:
        date_q = {}
        if date_from: date_q["$gte"] = date_from
        if date_to: date_q["$lte"] = date_to
        if date_q: query["date"] = date_q
    txns = await db.cash_transactions.find(query, {"_id": 0}).sort("date", -1).to_list(5000)
    return txns


@api_router.delete("/cash-book/{txn_id}")
async def delete_cash_transaction(txn_id: str):
    result = await db.cash_transactions.delete_one({"id": txn_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return {"message": "Transaction deleted", "id": txn_id}


@api_router.get("/cash-book/summary")
async def get_cash_book_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    txns = await db.cash_transactions.find(query, {"_id": 0}).to_list(10000)
    
    cash_in = sum(t['amount'] for t in txns if t.get('account') == 'cash' and t.get('txn_type') == 'jama')
    cash_out = sum(t['amount'] for t in txns if t.get('account') == 'cash' and t.get('txn_type') == 'nikasi')
    bank_in = sum(t['amount'] for t in txns if t.get('account') == 'bank' and t.get('txn_type') == 'jama')
    bank_out = sum(t['amount'] for t in txns if t.get('account') == 'bank' and t.get('txn_type') == 'nikasi')
    
    return {
        "cash_in": round(cash_in, 2),
        "cash_out": round(cash_out, 2),
        "cash_balance": round(cash_in - cash_out, 2),
        "bank_in": round(bank_in, 2),
        "bank_out": round(bank_out, 2),
        "bank_balance": round(bank_in - bank_out, 2),
        "total_balance": round((cash_in - cash_out) + (bank_in - bank_out), 2),
        "total_transactions": len(txns)
    }


@api_router.get("/cash-book/categories")
async def get_cash_book_categories():
    cats = await db.cash_book_categories.find({}, {"_id": 0}).to_list(500)
    return cats


@api_router.post("/cash-book/categories")
async def add_cash_book_category(request: Request):
    data = await request.json()
    name = (data.get("name") or "").strip()
    cat_type = data.get("type", "")  # cash_jama, cash_nikasi, bank_jama, bank_nikasi
    if not name or not cat_type:
        raise HTTPException(status_code=400, detail="Name and type required")
    existing = await db.cash_book_categories.find_one({"name": name, "type": cat_type})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    cat = {"id": str(uuid.uuid4()), "name": name, "type": cat_type, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.cash_book_categories.insert_one(cat)
    cat.pop('_id', None)
    return cat


@api_router.delete("/cash-book/categories/{cat_id}")
async def delete_cash_book_category(cat_id: str):
    result = await db.cash_book_categories.delete_one({"id": cat_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted", "id": cat_id}


@api_router.get("/cash-book/excel")
async def export_cash_book_excel(kms_year: Optional[str] = None, season: Optional[str] = None,
                                  account: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if account: query["account"] = account
    txns = await db.cash_transactions.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    summary = await get_cash_book_summary(kms_year=kms_year, season=season)
    
    wb = Workbook(); ws = wb.active; ws.title = "Cash Book"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    title = "Daily Cash Book / रोज़नामचा"
    if kms_year: title += f" - KMS {kms_year}"
    ws.merge_cells('A1:H1'); ws['A1'] = title; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    
    # Summary section
    ws.cell(row=3, column=1, value="Summary").font = Font(bold=True, size=11)
    for col, h in enumerate(['', 'Jama (In)', 'Nikasi (Out)', 'Balance'], 1):
        c = ws.cell(row=4, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    for col, v in enumerate(['Cash (नकद)', summary['cash_in'], summary['cash_out'], summary['cash_balance']], 1):
        c = ws.cell(row=5, column=col, value=v); c.border = tb
        if col >= 2: c.alignment = Alignment(horizontal='right'); c.number_format = '#,##0.00'
    for col, v in enumerate(['Bank (बैंक)', summary['bank_in'], summary['bank_out'], summary['bank_balance']], 1):
        c = ws.cell(row=6, column=col, value=v); c.border = tb
        if col >= 2: c.alignment = Alignment(horizontal='right'); c.number_format = '#,##0.00'
    ws.cell(row=7, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=7, column=4, value=summary['total_balance']).font = Font(bold=True)
    ws.cell(row=7, column=4).number_format = '#,##0.00'
    
    # Transactions
    row = 9
    ws.cell(row=row, column=1, value="Transactions").font = Font(bold=True, size=11)
    row += 1
    for col, h in enumerate(['Date', 'Account', 'Type', 'Category', 'Description', 'Jama (₹)', 'Nikasi (₹)', 'Reference'], 1):
        c = ws.cell(row=row, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb; c.alignment = Alignment(horizontal='center')
    row += 1
    for t in txns:
        jama = t['amount'] if t['txn_type'] == 'jama' else 0
        nikasi = t['amount'] if t['txn_type'] == 'nikasi' else 0
        for col, v in enumerate([t.get('date',''), 'Cash' if t.get('account')=='cash' else 'Bank',
            'Jama' if t.get('txn_type')=='jama' else 'Nikasi',
            t.get('category',''), t.get('description',''), jama, nikasi, t.get('reference','')], 1):
            c = ws.cell(row=row, column=col, value=v); c.border = tb
            if col in [6,7]: c.alignment = Alignment(horizontal='right'); c.number_format = '#,##0.00'
        row += 1
    
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(sum(t['amount'] for t in txns if t['txn_type']=='jama'),2)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(sum(t['amount'] for t in txns if t['txn_type']=='nikasi'),2)).font = Font(bold=True)
    for letter in ['A','B','C','D','E','F','G','H']: ws.column_dimensions[letter].width = 16
    
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cash_book_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/cash-book/pdf")
async def export_cash_book_pdf(kms_year: Optional[str] = None, season: Optional[str] = None,
                                account: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if account: query["account"] = account
    txns = await db.cash_transactions.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    summary = await get_cash_book_summary(kms_year=kms_year, season=season)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    title = "Daily Cash Book / रोज़नामचा"
    if kms_year: title += f" - KMS {kms_year}"
    elements.append(Paragraph(title, styles['Title'])); elements.append(Spacer(1, 12))
    
    # Summary table
    elements.append(Paragraph("Summary", styles['Heading2'])); elements.append(Spacer(1, 6))
    sdata = [['', 'Jama (In)', 'Nikasi (Out)', 'Balance'],
             ['Cash', summary['cash_in'], summary['cash_out'], summary['cash_balance']],
             ['Bank', summary['bank_in'], summary['bank_out'], summary['bank_balance']],
             ['Total', round(summary['cash_in']+summary['bank_in'],2), round(summary['cash_out']+summary['bank_out'],2), summary['total_balance']]]
    st = RLTable(sdata, colWidths=[80, 80, 80, 80])
    st.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 8), ('ALIGN', (1,0), (-1,-1), 'RIGHT'), ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f0f0f0'))]))
    elements.append(st); elements.append(Spacer(1, 15))
    
    # Transactions table
    elements.append(Paragraph("Transactions", styles['Heading2'])); elements.append(Spacer(1, 6))
    data = [['Date', 'Account', 'Type', 'Category', 'Description', 'Jama(₹)', 'Nikasi(₹)', 'Ref']]
    tj = tn = 0
    for t in txns:
        jama = t['amount'] if t['txn_type'] == 'jama' else 0
        nikasi = t['amount'] if t['txn_type'] == 'nikasi' else 0
        tj += jama; tn += nikasi
        data.append([t.get('date',''), 'Cash' if t.get('account')=='cash' else 'Bank',
            'Jama' if t.get('txn_type')=='jama' else 'Nikasi',
            t.get('category','')[:15], t.get('description','')[:20], jama if jama > 0 else '', nikasi if nikasi > 0 else '', t.get('reference','')[:12]])
    data.append(['TOTAL', '', '', '', '', round(tj,2), round(tn,2), ''])
    
    table = RLTable(data, colWidths=[55, 45, 40, 70, 100, 55, 55, 60], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 7), ('ALIGN', (5,0), (6,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cash_book_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ DC (DELIVERY CHALLAN) MANAGEMENT ============

class DCEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    dc_number: str
    date: str
    quantity_qntl: float = 0
    rice_type: str = "parboiled"  # parboiled / raw
    godown_name: str = ""
    deadline: str = ""
    notes: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class DCDelivery(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    dc_id: str
    date: str
    quantity_qntl: float = 0
    vehicle_no: str = ""
    driver_name: str = ""
    slip_no: str = ""
    godown_name: str = ""
    notes: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


@api_router.post("/dc-entries")
async def add_dc_entry(dc: DCEntry, username: str = ""):
    d = dc.model_dump()
    d['created_by'] = username
    d['quantity_qntl'] = round(d['quantity_qntl'], 2)
    await db.dc_entries.insert_one(d)
    d.pop('_id', None)
    return d


@api_router.get("/dc-entries")
async def get_dc_entries(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.dc_entries.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    # Attach delivery summary to each DC
    for e in entries:
        deliveries = await db.dc_deliveries.find({"dc_id": e["id"]}, {"_id": 0}).to_list(500)
        delivered = round(sum(d.get("quantity_qntl", 0) for d in deliveries), 2)
        e["delivered_qntl"] = delivered
        e["pending_qntl"] = round(e["quantity_qntl"] - delivered, 2)
        e["delivery_count"] = len(deliveries)
        e["status"] = "completed" if delivered >= e["quantity_qntl"] else ("partial" if delivered > 0 else "pending")
    return entries


@api_router.put("/dc-entries/{dc_id}")
async def update_dc_entry(dc_id: str, dc: DCEntry):
    d = dc.model_dump()
    d['quantity_qntl'] = round(d['quantity_qntl'], 2)
    d.pop('id', None)
    result = await db.dc_entries.update_one({"id": dc_id}, {"$set": d})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="DC not found")
    updated = await db.dc_entries.find_one({"id": dc_id}, {"_id": 0})
    return updated


@api_router.delete("/dc-entries/{dc_id}")
async def delete_dc_entry(dc_id: str):
    result = await db.dc_entries.delete_one({"id": dc_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="DC not found")
    await db.dc_deliveries.delete_many({"dc_id": dc_id})
    return {"message": "DC and its deliveries deleted", "id": dc_id}


@api_router.post("/dc-deliveries")
async def add_dc_delivery(delivery: DCDelivery, username: str = ""):
    d = delivery.model_dump()
    d['created_by'] = username
    d['quantity_qntl'] = round(d['quantity_qntl'], 2)
    dc = await db.dc_entries.find_one({"id": d["dc_id"]}, {"_id": 0})
    if not dc:
        raise HTTPException(status_code=404, detail="DC not found")
    await db.dc_deliveries.insert_one(d)
    d.pop('_id', None)
    return d


@api_router.get("/dc-deliveries")
async def get_dc_deliveries(dc_id: Optional[str] = None, kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if dc_id: query["dc_id"] = dc_id
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    return await db.dc_deliveries.find(query, {"_id": 0}).sort("date", -1).to_list(2000)


@api_router.delete("/dc-deliveries/{delivery_id}")
async def delete_dc_delivery(delivery_id: str):
    result = await db.dc_deliveries.delete_one({"id": delivery_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Delivery not found")
    return {"message": "Delivery deleted", "id": delivery_id}


@api_router.get("/dc-summary")
async def get_dc_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    dcs = await db.dc_entries.find(query, {"_id": 0}).to_list(1000)
    total_allotted = round(sum(d.get("quantity_qntl", 0) for d in dcs), 2)
    all_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    total_delivered = round(sum(d.get("quantity_qntl", 0) for d in all_deliveries), 2)
    completed = 0; partial = 0; pending_count = 0
    for dc in dcs:
        deld = sum(d.get("quantity_qntl", 0) for d in all_deliveries if d.get("dc_id") == dc["id"])
        if deld >= dc["quantity_qntl"]: completed += 1
        elif deld > 0: partial += 1
        else: pending_count += 1
    return {
        "total_dc": len(dcs), "total_allotted_qntl": total_allotted,
        "total_delivered_qntl": total_delivered, "total_pending_qntl": round(total_allotted - total_delivered, 2),
        "completed": completed, "partial": partial, "pending": pending_count,
        "total_deliveries": len(all_deliveries)
    }


@api_router.get("/dc-entries/excel")
async def export_dc_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    dcs = await db.dc_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    all_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    wb = Workbook(); ws = wb.active; ws.title = "DC Register"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A1:I1'); ws['A1'] = "DC (Delivery Challan) Register"; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    headers = ['DC No', 'Date', 'Rice Type', 'Allotted (Q)', 'Delivered (Q)', 'Pending (Q)', 'Status', 'Deadline', 'Godown']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb; c.alignment = Alignment(horizontal='center')
    row = 4
    for dc in dcs:
        deld = round(sum(d.get("quantity_qntl", 0) for d in all_deliveries if d.get("dc_id") == dc["id"]), 2)
        pend = round(dc["quantity_qntl"] - deld, 2)
        status = "Completed" if deld >= dc["quantity_qntl"] else ("Partial" if deld > 0 else "Pending")
        for col, v in enumerate([dc.get("dc_number",""), dc.get("date",""), (dc.get("rice_type","")).capitalize(), dc["quantity_qntl"], deld, pend, status, dc.get("deadline",""), dc.get("godown_name","")], 1):
            c = ws.cell(row=row, column=col, value=v); c.border = tb
            if col in [4,5,6]: c.alignment = Alignment(horizontal='right'); c.number_format = '#,##0.00'
        row += 1
    total_allot = round(sum(d["quantity_qntl"] for d in dcs), 2)
    total_del = round(sum(d.get("quantity_qntl",0) for d in all_deliveries), 2)
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_allot).font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_del).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_allot-total_del, 2)).font = Font(bold=True)
    # Delivery detail sheet
    ws2 = wb.create_sheet("Deliveries")
    dheaders = ['DC No', 'Date', 'Qty (Q)', 'Vehicle', 'Driver', 'Slip No', 'Godown', 'Note']
    for col, h in enumerate(dheaders, 1):
        c = ws2.cell(row=1, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    dc_map = {d["id"]: d.get("dc_number","") for d in dcs}
    for i, dl in enumerate(sorted(all_deliveries, key=lambda x: x.get("date","")), 2):
        for col, v in enumerate([dc_map.get(dl.get("dc_id",""),""), dl.get("date",""), dl.get("quantity_qntl",0), dl.get("vehicle_no",""), dl.get("driver_name",""), dl.get("slip_no",""), dl.get("godown_name",""), dl.get("notes","")], 1):
            ws2.cell(row=i, column=col, value=v).border = tb
    for letter in ['A','B','C','D','E','F','G','H','I']:
        ws.column_dimensions[letter].width = 15; ws2.column_dimensions[letter].width = 15
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dc_register_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/dc-entries/pdf")
async def export_dc_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    dcs = await db.dc_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    all_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("DC (Delivery Challan) Register", styles['Title'])); elements.append(Spacer(1, 12))
    data = [['DC No','Date','Type','Allotted(Q)','Delivered(Q)','Pending(Q)','Status','Deadline','Godown']]
    ta = td = 0
    for dc in dcs:
        deld = round(sum(d.get("quantity_qntl",0) for d in all_deliveries if d.get("dc_id")==dc["id"]),2)
        pend = round(dc["quantity_qntl"]-deld,2); ta += dc["quantity_qntl"]; td += deld
        status = "Done" if deld >= dc["quantity_qntl"] else ("Partial" if deld > 0 else "Pending")
        data.append([dc.get("dc_number",""), dc.get("date",""), (dc.get("rice_type","")).capitalize()[:5], dc["quantity_qntl"], deld, pend, status, dc.get("deadline",""), dc.get("godown_name","")[:12]])
    data.append(['TOTAL','','', round(ta,2), round(td,2), round(ta-td,2), '','',''])
    table = RLTable(data, colWidths=[55,55,40,55,55,50,40,55,60], repeatRows=1)
    table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a365d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,-1),7),('ALIGN',(3,0),(5,-1),'RIGHT'),('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#f0f0f0')),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')]))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=dc_register_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ MSP PAYMENT TRACKING ============

class MSPPayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    dc_id: str = ""  # optional link to DC
    amount: float = 0
    quantity_qntl: float = 0
    rate_per_qntl: float = 0
    payment_mode: str = ""  # NEFT/RTGS/Cheque/Cash
    reference: str = ""  # UTR/Cheque number
    bank_name: str = ""
    notes: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


@api_router.post("/msp-payments")
async def add_msp_payment(pay: MSPPayment, username: str = ""):
    d = pay.model_dump()
    d['created_by'] = username
    d['amount'] = round(d['amount'], 2)
    d['quantity_qntl'] = round(d['quantity_qntl'], 2)
    d['rate_per_qntl'] = round(d['rate_per_qntl'], 2)
    await db.msp_payments.insert_one(d)
    d.pop('_id', None)
    return d


@api_router.get("/msp-payments")
async def get_msp_payments(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.msp_payments.find(query, {"_id": 0}).sort("date", -1).to_list(2000)
    # Attach DC number
    dc_ids = list(set(p.get("dc_id","") for p in payments if p.get("dc_id")))
    dcs = {}
    if dc_ids:
        dc_docs = await db.dc_entries.find({"id": {"$in": dc_ids}}, {"_id": 0, "id": 1, "dc_number": 1}).to_list(500)
        dcs = {d["id"]: d.get("dc_number","") for d in dc_docs}
    for p in payments:
        p["dc_number"] = dcs.get(p.get("dc_id",""), "")
    return payments


@api_router.delete("/msp-payments/{payment_id}")
async def delete_msp_payment(payment_id: str):
    result = await db.msp_payments.delete_one({"id": payment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"message": "Payment deleted", "id": payment_id}


@api_router.get("/msp-payments/summary")
async def get_msp_payment_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.msp_payments.find(query, {"_id": 0}).to_list(5000)
    dcs = await db.dc_entries.find(query, {"_id": 0}).to_list(1000)
    all_deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    total_delivered = round(sum(d.get("quantity_qntl",0) for d in all_deliveries), 2)
    total_paid_amount = round(sum(p.get("amount",0) for p in payments), 2)
    total_paid_qty = round(sum(p.get("quantity_qntl",0) for p in payments), 2)
    avg_rate = round(total_paid_amount / total_paid_qty, 2) if total_paid_qty > 0 else 0
    return {
        "total_payments": len(payments), "total_paid_amount": total_paid_amount,
        "total_paid_qty": total_paid_qty, "avg_rate": avg_rate,
        "total_delivered_qntl": total_delivered,
        "pending_payment_qty": round(total_delivered - total_paid_qty, 2),
    }


@api_router.get("/msp-payments/excel")
async def export_msp_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.msp_payments.find(query, {"_id": 0}).sort("date", 1).to_list(5000)
    dc_ids = list(set(p.get("dc_id","") for p in payments if p.get("dc_id")))
    dcs = {}
    if dc_ids:
        dc_docs = await db.dc_entries.find({"id": {"$in": dc_ids}}, {"_id": 0}).to_list(500)
        dcs = {d["id"]: d.get("dc_number","") for d in dc_docs}
    wb = Workbook(); ws = wb.active; ws.title = "MSP Payments"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A1:H1'); ws['A1'] = "MSP Payment Register"; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    headers = ['Date','DC No','Qty (Q)','Rate (₹/Q)','Amount (₹)','Mode','Reference','Bank']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    row = 4
    for p in payments:
        for col, v in enumerate([p.get("date",""), dcs.get(p.get("dc_id",""),""), p.get("quantity_qntl",0), p.get("rate_per_qntl",0), p.get("amount",0), p.get("payment_mode",""), p.get("reference",""), p.get("bank_name","")], 1):
            c = ws.cell(row=row, column=col, value=v); c.border = tb
            if col in [3,4,5]: c.alignment = Alignment(horizontal='right'); c.number_format = '#,##0.00'
        row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(sum(p.get("quantity_qntl",0) for p in payments),2)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(sum(p.get("amount",0) for p in payments),2)).font = Font(bold=True)
    for letter in ['A','B','C','D','E','F','G','H']: ws.column_dimensions[letter].width = 15
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=msp_payments_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/msp-payments/pdf")
async def export_msp_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    payments = await db.msp_payments.find(query, {"_id": 0}).sort("date", 1).to_list(5000)
    dc_ids = list(set(p.get("dc_id","") for p in payments if p.get("dc_id")))
    dcs = {}
    if dc_ids:
        dc_docs = await db.dc_entries.find({"id": {"$in": dc_ids}}, {"_id": 0}).to_list(500)
        dcs = {d["id"]: d.get("dc_number","") for d in dc_docs}
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("MSP Payment Register", styles['Title'])); elements.append(Spacer(1, 12))
    data = [['Date','DC No','Qty(Q)','Rate(₹/Q)','Amount(₹)','Mode','Reference','Bank']]
    tq = ta = 0
    for p in payments:
        tq += p.get("quantity_qntl",0); ta += p.get("amount",0)
        data.append([p.get("date",""), dcs.get(p.get("dc_id",""),""), p.get("quantity_qntl",0), p.get("rate_per_qntl",0), p.get("amount",0), p.get("payment_mode",""), p.get("reference","")[:15], p.get("bank_name","")[:12]])
    data.append(['TOTAL','',round(tq,2),'',round(ta,2),'','',''])
    table = RLTable(data, colWidths=[55,55,50,50,60,45,70,60], repeatRows=1)
    table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a365d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,-1),7),('ALIGN',(2,0),(4,-1),'RIGHT'),('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#f0f0f0')),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')]))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=msp_payments_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ GUNNY BAG TRACKING ============

class GunnyBagEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    bag_type: str = "new"  # new (govt free) / old (market purchase)
    txn_type: str = "in"   # in / out
    quantity: int = 0
    source: str = ""       # where from / where to
    rate: float = 0        # rate per bag (for old/purchased)
    amount: float = 0
    reference: str = ""
    notes: str = ""
    kms_year: str = ""
    season: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


@api_router.post("/gunny-bags")
async def add_gunny_bag_entry(entry: GunnyBagEntry, username: str = ""):
    d = entry.model_dump()
    d['created_by'] = username
    d['amount'] = round(d.get('quantity', 0) * d.get('rate', 0), 2)
    await db.gunny_bags.insert_one(d)
    d.pop('_id', None)
    return d


@api_router.get("/gunny-bags")
async def get_gunny_bag_entries(kms_year: Optional[str] = None, season: Optional[str] = None, bag_type: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    if bag_type: query["bag_type"] = bag_type
    return await db.gunny_bags.find(query, {"_id": 0}).sort("date", -1).to_list(5000)


@api_router.delete("/gunny-bags/{entry_id}")
async def delete_gunny_bag_entry(entry_id: str):
    result = await db.gunny_bags.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Deleted", "id": entry_id}


@api_router.get("/gunny-bags/summary")
async def get_gunny_bag_summary(kms_year: Optional[str] = None, season: Optional[str] = None):
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.gunny_bags.find(query, {"_id": 0}).to_list(10000)
    # Manual gunny bag transactions (new=govt, old=market)
    result = {}
    for bt in ["new", "old"]:
        items = [e for e in entries if e.get("bag_type") == bt]
        total_in = sum(e.get("quantity",0) for e in items if e.get("txn_type") == "in")
        total_out = sum(e.get("quantity",0) for e in items if e.get("txn_type") == "out")
        total_cost = round(sum(e.get("amount",0) for e in items if e.get("txn_type") == "in"), 2)
        result[bt] = {"total_in": total_in, "total_out": total_out, "balance": total_in - total_out, "total_cost": total_cost}
    # Paddy-received bags from truck entries (auto-calculated)
    paddy_entries = await db.mill_entries.find(query, {"_id": 0, "bag": 1, "plastic_bag": 1, "g_issued": 1}).to_list(10000)
    paddy_bags = sum(e.get("bag", 0) for e in paddy_entries)
    paddy_ppkt = sum(e.get("plastic_bag", 0) for e in paddy_entries)
    paddy_g_issued = sum(e.get("g_issued", 0) for e in paddy_entries)
    result["paddy_bags"] = {"total": paddy_bags, "label": "Paddy Receive Bags"}
    result["ppkt"] = {"total": paddy_ppkt, "label": "P.Pkt (Plastic Bags)"}
    result["g_issued"] = {"total": paddy_g_issued, "label": "Govt Bags Issued (g)"}
    # Grand total: old bags + paddy bags + P.Pkt (govt bags NOT included)
    result["grand_total"] = result["old"]["balance"] + paddy_bags + paddy_ppkt
    return result


@api_router.get("/gunny-bags/excel")
async def export_gunny_bags_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.gunny_bags.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    summary = await get_gunny_bag_summary(kms_year=kms_year, season=season)
    wb = Workbook(); ws = wb.active; ws.title = "Gunny Bags"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A1:H1'); ws['A1'] = "Gunny Bag Register / बोरी रजिस्टर"; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    # Summary
    ws.cell(row=3, column=1, value="Summary").font = Font(bold=True, size=11)
    for col, h in enumerate(['Type', 'In', 'Out', 'Balance', 'Cost (₹)'], 1):
        c = ws.cell(row=4, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    for i, (bt, label) in enumerate([("new","New (Govt)"),("old","Old (Market)")], 5):
        s = summary.get(bt, {})
        for col, v in enumerate([label, s.get("total_in",0), s.get("total_out",0), s.get("balance",0), s.get("total_cost",0)], 1):
            c = ws.cell(row=i, column=col, value=v); c.border = tb
    # Paddy receive bags
    ws.cell(row=7, column=1, value="Paddy Receive Bags").border = tb
    ws.cell(row=7, column=4, value=summary.get("paddy_bags",{}).get("total",0)).border = tb
    ws.cell(row=8, column=1, value="P.Pkt (Plastic)").border = tb
    ws.cell(row=8, column=4, value=summary.get("ppkt",{}).get("total",0)).border = tb
    ws.cell(row=9, column=1, value="Govt Issued (g)").border = tb
    ws.cell(row=9, column=4, value=summary.get("g_issued",{}).get("total",0)).border = tb
    ws.cell(row=10, column=1, value="Total (Excl Govt)").font = Font(bold=True)
    ws.cell(row=10, column=4, value=summary.get("grand_total",0)).font = Font(bold=True)
    # Transactions
    row = 12
    ws.cell(row=row, column=1, value="Transactions").font = Font(bold=True, size=11); row += 1
    for col, h in enumerate(['Date','Type','In/Out','Qty','Source/To','Rate','Amount (₹)','Reference'], 1):
        c = ws.cell(row=row, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    row += 1
    for e in entries:
        for col, v in enumerate([e.get("date",""), "New" if e.get("bag_type")=="new" else "Old", "In" if e.get("txn_type")=="in" else "Out", e.get("quantity",0), e.get("source",""), e.get("rate",0), e.get("amount",0), e.get("reference","")], 1):
            ws.cell(row=row, column=col, value=v).border = tb
        row += 1
    for letter in ['A','B','C','D','E','F','G','H']: ws.column_dimensions[letter].width = 15
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gunny_bags_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/gunny-bags/pdf")
async def export_gunny_bags_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    entries = await db.gunny_bags.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    summary = await get_gunny_bag_summary(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("Gunny Bag Register", styles['Title'])); elements.append(Spacer(1, 10))
    # Summary
    sdata = [['Type','In','Out','Balance','Cost(₹)']]
    for bt, label in [("new","New(Govt)"),("old","Old(Market)")]:
        s = summary.get(bt, {})
        sdata.append([label, s.get("total_in",0), s.get("total_out",0), s.get("balance",0), s.get("total_cost",0)])
    st = RLTable(sdata, colWidths=[70,50,50,50,60])
    st.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a365d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,-1),8),('GRID',(0,0),(-1,-1),0.5,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')]))
    elements.append(st); elements.append(Spacer(1, 12))
    # Transactions
    elements.append(Paragraph("Transactions", styles['Heading2'])); elements.append(Spacer(1, 6))
    data = [['Date','Type','In/Out','Qty','Source/To','Rate','Amount(₹)','Ref']]
    for e in entries:
        data.append([e.get("date",""), "New" if e.get("bag_type")=="new" else "Old", "In" if e.get("txn_type")=="in" else "Out", e.get("quantity",0), e.get("source","")[:18], e.get("rate",0), e.get("amount",0), e.get("reference","")[:12]])
    table = RLTable(data, colWidths=[55,40,35,35,80,40,50,55], repeatRows=1)
    table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a365d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),0.5,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')]))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=gunny_bags_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ PHASE 4: REPORTING ============

@api_router.get("/reports/cmr-vs-dc")
async def report_cmr_vs_dc(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Compare milling output (CMR) vs DC allotment and deliveries"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    # Milling data
    milling = await db.milling_entries.find(query, {"_id": 0}).to_list(5000)
    total_paddy_milled = round(sum(e.get("paddy_input_qntl", 0) for e in milling), 2)
    total_rice_produced = round(sum(e.get("rice_qntl", 0) for e in milling), 2)
    total_frk_used = round(sum(e.get("frk_used_qntl", 0) for e in milling), 2)
    total_cmr = round(sum(e.get("cmr_delivery_qntl", 0) for e in milling), 2)
    avg_outturn = round(total_cmr / total_paddy_milled * 100, 2) if total_paddy_milled > 0 else 0
    # DC data
    dcs = await db.dc_entries.find(query, {"_id": 0}).to_list(1000)
    deliveries = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    total_dc_allotted = round(sum(d.get("quantity_qntl", 0) for d in dcs), 2)
    total_dc_delivered = round(sum(d.get("quantity_qntl", 0) for d in deliveries), 2)
    total_dc_pending = round(total_dc_allotted - total_dc_delivered, 2)
    # Comparison
    cmr_surplus = round(total_cmr - total_dc_allotted, 2)
    delivery_gap = round(total_cmr - total_dc_delivered, 2)  # CMR ready but not delivered
    # By-product revenue
    bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(5000)
    bp_revenue = round(sum(s.get("total_amount", 0) for s in bp_sales), 2)
    return {
        "milling": {"total_paddy_milled": total_paddy_milled, "total_rice_produced": total_rice_produced, "total_frk_used": total_frk_used, "total_cmr_ready": total_cmr, "avg_outturn_pct": avg_outturn, "milling_count": len(milling)},
        "dc": {"total_allotted": total_dc_allotted, "total_delivered": total_dc_delivered, "total_pending": total_dc_pending, "dc_count": len(dcs), "delivery_count": len(deliveries)},
        "comparison": {"cmr_vs_dc_allotted": cmr_surplus, "cmr_vs_dc_delivered": delivery_gap},
        "byproduct_revenue": bp_revenue
    }


@api_router.get("/reports/season-pnl")
async def report_season_pnl(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Season-wise Profit & Loss summary"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season
    # Income
    msp_payments = await db.msp_payments.find(query, {"_id": 0}).to_list(5000)
    msp_income = round(sum(p.get("amount", 0) for p in msp_payments), 2)
    bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(5000)
    bp_income = round(sum(s.get("total_amount", 0) for s in bp_sales), 2)
    # Expenses
    frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(5000)
    frk_cost = round(sum(p.get("total_amount", 0) for p in frk_purchases), 2)
    gunny_bags = await db.gunny_bags.find(query, {"_id": 0}).to_list(5000)
    gunny_cost = round(sum(g.get("amount", 0) for g in gunny_bags if g.get("txn_type") == "in"), 2)
    cash_txns = await db.cash_transactions.find(query, {"_id": 0}).to_list(10000)
    cash_expenses = round(sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "nikasi"), 2)
    cash_income_other = round(sum(t.get("amount", 0) for t in cash_txns if t.get("txn_type") == "jama"), 2)
    # Truck/Agent payments from mill entries
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(10000)
    truck_payments = round(sum(e.get("tp_paid", 0) for e in entries), 2)
    agent_payments = round(sum(e.get("agent_paid", 0) for e in entries), 2)
    total_income = round(msp_income + bp_income + cash_income_other, 2)
    total_expenses = round(frk_cost + gunny_cost + cash_expenses + truck_payments + agent_payments, 2)
    net_pnl = round(total_income - total_expenses, 2)
    return {
        "income": {"msp_payments": msp_income, "byproduct_sales": bp_income, "cash_book_jama": cash_income_other, "total": total_income},
        "expenses": {"frk_purchases": frk_cost, "gunny_bags": gunny_cost, "cash_book_nikasi": cash_expenses, "truck_payments": truck_payments, "agent_payments": agent_payments, "total": total_expenses},
        "net_pnl": net_pnl, "profit": net_pnl >= 0
    }


@api_router.get("/reports/cmr-vs-dc/excel")
async def export_cmr_vs_dc_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    data = await report_cmr_vs_dc(kms_year=kms_year, season=season)
    wb = Workbook(); ws = wb.active; ws.title = "CMR vs DC"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    title = "CMR vs DC Report"
    if kms_year: title += f" - KMS {kms_year}"
    ws.merge_cells('A1:D1'); ws['A1'] = title; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    # Milling
    ws.cell(row=3, column=1, value="MILLING OUTPUT").font = Font(bold=True, size=11, color="2563eb")
    items = [("Paddy Milled (Q)", data["milling"]["total_paddy_milled"]), ("Rice Produced (Q)", data["milling"]["total_rice_produced"]),
             ("FRK Used (Q)", data["milling"]["total_frk_used"]), ("CMR Ready (Q)", data["milling"]["total_cmr_ready"]),
             ("Avg Outturn %", data["milling"]["avg_outturn_pct"]), ("Milling Count", data["milling"]["milling_count"])]
    for i, (label, val) in enumerate(items, 4):
        ws.cell(row=i, column=1, value=label).border = tb
        c = ws.cell(row=i, column=2, value=val); c.border = tb; c.alignment = Alignment(horizontal='right')
    # DC
    row = 11
    ws.cell(row=row, column=1, value="DC ALLOTMENT & DELIVERY").font = Font(bold=True, size=11, color="16a34a")
    items2 = [("DC Allotted (Q)", data["dc"]["total_allotted"]), ("DC Delivered (Q)", data["dc"]["total_delivered"]),
              ("DC Pending (Q)", data["dc"]["total_pending"]), ("Total DCs", data["dc"]["dc_count"]), ("Total Deliveries", data["dc"]["delivery_count"])]
    for i, (label, val) in enumerate(items2, row+1):
        ws.cell(row=i, column=1, value=label).border = tb
        c = ws.cell(row=i, column=2, value=val); c.border = tb; c.alignment = Alignment(horizontal='right')
    # Comparison
    row = 18
    ws.cell(row=row, column=1, value="COMPARISON").font = Font(bold=True, size=11, color="d97706")
    ws.cell(row=row+1, column=1, value="CMR vs DC Allotted").border = tb
    ws.cell(row=row+1, column=2, value=data["comparison"]["cmr_vs_dc_allotted"]).border = tb
    ws.cell(row=row+2, column=1, value="CMR vs DC Delivered").border = tb
    ws.cell(row=row+2, column=2, value=data["comparison"]["cmr_vs_dc_delivered"]).border = tb
    ws.cell(row=row+3, column=1, value="By-Product Revenue (₹)").border = tb
    ws.cell(row=row+3, column=2, value=data["byproduct_revenue"]).border = tb
    for letter in ['A','B','C','D']: ws.column_dimensions[letter].width = 22
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cmr_vs_dc_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/reports/cmr-vs-dc/pdf")
async def export_cmr_vs_dc_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    data = await report_cmr_vs_dc(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("CMR vs DC Report", styles['Title'])); elements.append(Spacer(1, 12))
    rows = [['Metric', 'Value'],
        ['--- MILLING ---', ''], ['Paddy Milled (Q)', data['milling']['total_paddy_milled']], ['Rice Produced (Q)', data['milling']['total_rice_produced']],
        ['FRK Used (Q)', data['milling']['total_frk_used']], ['CMR Ready (Q)', data['milling']['total_cmr_ready']], ['Outturn %', data['milling']['avg_outturn_pct']],
        ['--- DC ---', ''], ['DC Allotted (Q)', data['dc']['total_allotted']], ['DC Delivered (Q)', data['dc']['total_delivered']], ['DC Pending (Q)', data['dc']['total_pending']],
        ['--- COMPARISON ---', ''], ['CMR vs DC Allotted', data['comparison']['cmr_vs_dc_allotted']], ['CMR vs DC Delivered', data['comparison']['cmr_vs_dc_delivered']],
        ['By-Product Revenue', f"₹{data['byproduct_revenue']}"]]
    table = RLTable(rows, colWidths=[200, 150])
    table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a365d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('ALIGN',(1,0),(1,-1),'RIGHT')]))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cmr_vs_dc_{datetime.now().strftime('%Y%m%d')}.pdf"})


@api_router.get("/reports/season-pnl/excel")
async def export_season_pnl_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    data = await report_season_pnl(kms_year=kms_year, season=season)
    wb = Workbook(); ws = wb.active; ws.title = "Season P&L"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    title = "Season P&L Report"
    if kms_year: title += f" - KMS {kms_year}"
    ws.merge_cells('A1:C1'); ws['A1'] = title; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    row = 3
    ws.cell(row=row, column=1, value="INCOME").font = Font(bold=True, size=11, color="16a34a")
    for label, val in [("MSP Payments", data["income"]["msp_payments"]), ("By-Product Sales", data["income"]["byproduct_sales"]),
                        ("Cash Book Jama", data["income"]["cash_book_jama"]), ("TOTAL INCOME", data["income"]["total"])]:
        row += 1
        ws.cell(row=row, column=1, value=label).border = tb
        c = ws.cell(row=row, column=2, value=val); c.border = tb; c.number_format = '#,##0.00'
        if label.startswith("TOTAL"): ws.cell(row=row, column=1).font = Font(bold=True); c.font = Font(bold=True)
    row += 2
    ws.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, size=11, color="dc2626")
    for label, val in [("FRK Purchases", data["expenses"]["frk_purchases"]), ("Gunny Bags", data["expenses"]["gunny_bags"]),
                        ("Cash Book Nikasi", data["expenses"]["cash_book_nikasi"]), ("Truck Payments", data["expenses"]["truck_payments"]),
                        ("Agent Payments", data["expenses"]["agent_payments"]), ("TOTAL EXPENSES", data["expenses"]["total"])]:
        row += 1
        ws.cell(row=row, column=1, value=label).border = tb
        c = ws.cell(row=row, column=2, value=val); c.border = tb; c.number_format = '#,##0.00'
        if label.startswith("TOTAL"): ws.cell(row=row, column=1).font = Font(bold=True); c.font = Font(bold=True)
    row += 2
    pnl_label = "NET PROFIT" if data["profit"] else "NET LOSS"
    ws.cell(row=row, column=1, value=pnl_label).font = Font(bold=True, size=12, color="16a34a" if data["profit"] else "dc2626")
    ws.cell(row=row, column=2, value=data["net_pnl"]).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    for letter in ['A','B','C']: ws.column_dimensions[letter].width = 22
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=season_pnl_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/reports/season-pnl/pdf")
async def export_season_pnl_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    data = await report_season_pnl(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("Season P&L Report", styles['Title'])); elements.append(Spacer(1, 12))
    # Income
    elements.append(Paragraph("INCOME", styles['Heading2'])); elements.append(Spacer(1, 4))
    idata = [['Source', 'Amount (₹)'], ['MSP Payments', data['income']['msp_payments']], ['By-Product Sales', data['income']['byproduct_sales']],
             ['Cash Book Jama', data['income']['cash_book_jama']], ['TOTAL', data['income']['total']]]
    it = RLTable(idata, colWidths=[200, 120])
    it.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#166534')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),('ALIGN',(1,0),(1,-1),'RIGHT')]))
    elements.append(it); elements.append(Spacer(1, 12))
    # Expenses
    elements.append(Paragraph("EXPENSES", styles['Heading2'])); elements.append(Spacer(1, 4))
    edata = [['Category', 'Amount (₹)'], ['FRK Purchases', data['expenses']['frk_purchases']], ['Gunny Bags', data['expenses']['gunny_bags']],
             ['Cash Book Nikasi', data['expenses']['cash_book_nikasi']], ['Truck Payments', data['expenses']['truck_payments']],
             ['Agent Payments', data['expenses']['agent_payments']], ['TOTAL', data['expenses']['total']]]
    et = RLTable(edata, colWidths=[200, 120])
    et.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#991b1b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),('ALIGN',(1,0),(1,-1),'RIGHT')]))
    elements.append(et); elements.append(Spacer(1, 15))
    # Net P&L
    pnl_color = colors.HexColor('#166534') if data['profit'] else colors.HexColor('#991b1b')
    pdata = [['NET ' + ('PROFIT' if data['profit'] else 'LOSS'), f"₹{data['net_pnl']}"]]
    pt = RLTable(pdata, colWidths=[200, 120])
    pt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),pnl_color),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,0),14),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('ALIGN',(1,0),(1,0),'RIGHT')]))
    elements.append(pt); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=season_pnl_{datetime.now().strftime('%Y%m%d')}.pdf"})


# ============ PHASE 5: CONSOLIDATED LEDGERS ============

@api_router.get("/reports/outstanding")
async def report_outstanding(kms_year: Optional[str] = None, season: Optional[str] = None):
    """Outstanding Report - all pending payments/deliveries across modules"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    # DC pending deliveries
    dcs = await db.dc_entries.find(query, {"_id": 0}).to_list(1000)
    all_dels = await db.dc_deliveries.find(query, {"_id": 0}).to_list(5000)
    dc_outstanding = []
    for dc in dcs:
        delivered = round(sum(d.get("quantity_qntl", 0) for d in all_dels if d.get("dc_id") == dc["id"]), 2)
        pending = round(dc["quantity_qntl"] - delivered, 2)
        if pending > 0:
            dc_outstanding.append({"dc_number": dc.get("dc_number", ""), "allotted": dc["quantity_qntl"], "delivered": delivered, "pending": pending, "deadline": dc.get("deadline", ""), "rice_type": dc.get("rice_type", "")})
    dc_pending_total = round(sum(d["pending"] for d in dc_outstanding), 2)

    # MSP payment pending
    msp_payments = await db.msp_payments.find(query, {"_id": 0}).to_list(5000)
    total_delivered_qntl = round(sum(d.get("quantity_qntl", 0) for d in all_dels), 2)
    total_msp_paid_qty = round(sum(p.get("quantity_qntl", 0) for p in msp_payments), 2)
    total_msp_paid_amt = round(sum(p.get("amount", 0) for p in msp_payments), 2)
    msp_pending_qty = round(total_delivered_qntl - total_msp_paid_qty, 2)

    # Truck payment pending (entries with cash_paid < expected)
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(10000)
    truck_map = {}
    for e in entries:
        truck = e.get("truck_no", "Unknown")
        if truck not in truck_map:
            truck_map[truck] = {"truck_no": truck, "total_trips": 0, "total_qty_qntl": 0, "total_cash_paid": 0, "total_diesel_paid": 0}
        truck_map[truck]["total_trips"] += 1
        truck_map[truck]["total_qty_qntl"] = round(truck_map[truck]["total_qty_qntl"] + (e.get("mill_w", 0) / 100), 2)
        truck_map[truck]["total_cash_paid"] = round(truck_map[truck]["total_cash_paid"] + (e.get("cash_paid", 0)), 2)
        truck_map[truck]["total_diesel_paid"] = round(truck_map[truck]["total_diesel_paid"] + (e.get("diesel_paid", 0)), 2)

    # Agent summary
    agent_map = {}
    for e in entries:
        agent = e.get("agent_name", "Unknown")
        if not agent: agent = "Unknown"
        if agent not in agent_map:
            agent_map[agent] = {"agent_name": agent, "total_entries": 0, "total_qty_qntl": 0}
        agent_map[agent]["total_entries"] += 1
        agent_map[agent]["total_qty_qntl"] = round(agent_map[agent]["total_qty_qntl"] + (e.get("mill_w", 0) / 100), 2)

    # FRK purchase outstanding
    frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(5000)
    frk_party_map = {}
    for p in frk_purchases:
        party = p.get("party_name", "Unknown")
        if party not in frk_party_map:
            frk_party_map[party] = {"party_name": party, "total_qty": 0, "total_amount": 0}
        frk_party_map[party]["total_qty"] = round(frk_party_map[party]["total_qty"] + (p.get("quantity_qntl", 0)), 2)
        frk_party_map[party]["total_amount"] = round(frk_party_map[party]["total_amount"] + (p.get("total_amount", 0)), 2)

    return {
        "dc_outstanding": {"items": dc_outstanding, "total_pending_qntl": dc_pending_total, "count": len(dc_outstanding)},
        "msp_outstanding": {"total_delivered_qntl": total_delivered_qntl, "total_paid_qty": total_msp_paid_qty, "total_paid_amount": total_msp_paid_amt, "pending_qty": msp_pending_qty},
        "trucks": list(truck_map.values()),
        "agents": list(agent_map.values()),
        "frk_parties": list(frk_party_map.values()),
    }


@api_router.get("/reports/party-ledger")
async def report_party_ledger(party_name: Optional[str] = None, party_type: Optional[str] = None,
                                kms_year: Optional[str] = None, season: Optional[str] = None):
    """Party Ledger - all transactions for a specific party or all parties"""
    query = {}
    if kms_year: query["kms_year"] = kms_year
    if season: query["season"] = season

    ledger = []

    # Paddy entries (Agent + Truck)
    entries = await db.mill_entries.find(query, {"_id": 0}).to_list(10000)
    if not party_type or party_type == "agent":
        for e in entries:
            agent = e.get("agent_name", "")
            if not agent: continue
            if party_name and agent.lower() != party_name.lower(): continue
            ledger.append({"date": e.get("date", ""), "party_name": agent, "party_type": "Agent",
                "description": f"Paddy: {round((e.get('mill_w',0))/100,2)}Q | Truck: {e.get('truck_no','')}",
                "debit": 0, "credit": round(e.get("cash_paid", 0) + e.get("diesel_paid", 0), 2), "ref": e.get("id", "")[:8]})
    if not party_type or party_type == "truck":
        for e in entries:
            truck = e.get("truck_no", "")
            if not truck: continue
            if party_name and truck.lower() != party_name.lower(): continue
            ledger.append({"date": e.get("date", ""), "party_name": truck, "party_type": "Truck",
                "description": f"Paddy: {round((e.get('mill_w',0))/100,2)}Q | Agent: {e.get('agent_name','')}",
                "debit": 0, "credit": round(e.get("cash_paid", 0) + e.get("diesel_paid", 0), 2), "ref": e.get("id", "")[:8]})

    # FRK purchases
    if not party_type or party_type == "frk_party":
        frk_purchases = await db.frk_purchases.find(query, {"_id": 0}).to_list(5000)
        for p in frk_purchases:
            party = p.get("party_name", "")
            if not party: continue
            if party_name and party.lower() != party_name.lower(): continue
            ledger.append({"date": p.get("date", ""), "party_name": party, "party_type": "FRK Seller",
                "description": f"FRK: {p.get('quantity_qntl',0)}Q @ ₹{p.get('rate_per_qntl',0)}/Q",
                "debit": round(p.get("total_amount", 0), 2), "credit": 0, "ref": p.get("id", "")[:8]})

    # By-product sales
    if not party_type or party_type == "buyer":
        bp_sales = await db.byproduct_sales.find(query, {"_id": 0}).to_list(5000)
        for s in bp_sales:
            buyer = s.get("buyer_name", "")
            if not buyer: continue
            if party_name and buyer.lower() != party_name.lower(): continue
            ledger.append({"date": s.get("date", ""), "party_name": buyer, "party_type": "Buyer",
                "description": f"{(s.get('product','')).capitalize()}: {s.get('quantity_qntl',0)}Q @ ₹{s.get('rate_per_qntl',0)}/Q",
                "debit": 0, "credit": round(s.get("total_amount", 0), 2), "ref": s.get("id", "")[:8]})

    ledger.sort(key=lambda x: x.get("date", ""), reverse=True)

    # Party list for filter
    parties = set()
    for item in ledger: parties.add((item["party_name"], item["party_type"]))
    party_list = [{"name": n, "type": t} for n, t in sorted(parties)]

    return {"ledger": ledger, "party_list": party_list, "total_debit": round(sum(l["debit"] for l in ledger), 2),
            "total_credit": round(sum(l["credit"] for l in ledger), 2)}


@api_router.get("/reports/outstanding/excel")
async def export_outstanding_excel(kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    data = await report_outstanding(kms_year=kms_year, season=season)
    wb = Workbook(); ws = wb.active; ws.title = "Outstanding Report"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A1:F1'); ws['A1'] = "Outstanding Report"; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    # DC Outstanding
    row = 3; ws.cell(row=row, column=1, value="DC PENDING DELIVERIES").font = Font(bold=True, size=11, color="dc2626"); row += 1
    for col, h in enumerate(['DC No', 'Allotted(Q)', 'Delivered(Q)', 'Pending(Q)', 'Deadline', 'Type'], 1):
        c = ws.cell(row=row, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    row += 1
    for d in data["dc_outstanding"]["items"]:
        for col, v in enumerate([d["dc_number"], d["allotted"], d["delivered"], d["pending"], d["deadline"], d["rice_type"]], 1):
            ws.cell(row=row, column=col, value=v).border = tb
        row += 1
    ws.cell(row=row, column=1, value="Total Pending").font = Font(bold=True)
    ws.cell(row=row, column=4, value=data["dc_outstanding"]["total_pending_qntl"]).font = Font(bold=True)
    # MSP Outstanding
    row += 2; ws.cell(row=row, column=1, value="MSP PAYMENT PENDING").font = Font(bold=True, size=11, color="d97706"); row += 1
    for label, val in [("Total Delivered (Q)", data["msp_outstanding"]["total_delivered_qntl"]), ("Paid Qty (Q)", data["msp_outstanding"]["total_paid_qty"]),
                        ("Paid Amount (₹)", data["msp_outstanding"]["total_paid_amount"]), ("Pending Qty (Q)", data["msp_outstanding"]["pending_qty"])]:
        ws.cell(row=row, column=1, value=label).border = tb; ws.cell(row=row, column=2, value=val).border = tb; row += 1
    # Trucks
    row += 1; ws.cell(row=row, column=1, value="TRUCK SUMMARY").font = Font(bold=True, size=11, color="2563eb"); row += 1
    for col, h in enumerate(['Truck No', 'Trips', 'Qty(Q)', 'Cash Paid', 'Diesel Paid'], 1):
        c = ws.cell(row=row, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    row += 1
    for t in data["trucks"]:
        for col, v in enumerate([t["truck_no"], t["total_trips"], t["total_qty_qntl"], t["total_cash_paid"], t["total_diesel_paid"]], 1):
            ws.cell(row=row, column=col, value=v).border = tb
        row += 1
    for letter in ['A','B','C','D','E','F']: ws.column_dimensions[letter].width = 18
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=outstanding_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/reports/outstanding/pdf")
async def export_outstanding_pdf(kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    data = await report_outstanding(kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("Outstanding Report", styles['Title'])); elements.append(Spacer(1, 12))
    # DC pending
    elements.append(Paragraph("DC Pending Deliveries", styles['Heading2'])); elements.append(Spacer(1, 4))
    ddata = [['DC No', 'Allotted(Q)', 'Delivered(Q)', 'Pending(Q)', 'Deadline', 'Type']]
    for d in data["dc_outstanding"]["items"]:
        ddata.append([d["dc_number"], d["allotted"], d["delivered"], d["pending"], d["deadline"], d["rice_type"]])
    ddata.append(['TOTAL', '', '', data["dc_outstanding"]["total_pending_qntl"], '', ''])
    dt = RLTable(ddata, colWidths=[60, 60, 60, 60, 60, 50])
    dt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#991b1b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTSIZE',(0,0),(-1,-1),8),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')]))
    elements.append(dt); elements.append(Spacer(1, 12))
    # MSP
    elements.append(Paragraph("MSP Payment Pending", styles['Heading2'])); elements.append(Spacer(1, 4))
    mdata = [['Metric', 'Value'], ['Delivered(Q)', data['msp_outstanding']['total_delivered_qntl']], ['Paid Qty(Q)', data['msp_outstanding']['total_paid_qty']],
             ['Paid Amount(₹)', data['msp_outstanding']['total_paid_amount']], ['Pending Qty(Q)', data['msp_outstanding']['pending_qty']]]
    mt = RLTable(mdata, colWidths=[150, 100])
    mt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#92400e')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTSIZE',(0,0),(-1,-1),8),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')]))
    elements.append(mt); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=outstanding_{datetime.now().strftime('%Y%m%d')}.pdf"})


@api_router.get("/reports/party-ledger/excel")
async def export_party_ledger_excel(party_name: Optional[str] = None, party_type: Optional[str] = None,
                                     kms_year: Optional[str] = None, season: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    data = await report_party_ledger(party_name=party_name, party_type=party_type, kms_year=kms_year, season=season)
    wb = Workbook(); ws = wb.active; ws.title = "Party Ledger"
    hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    title = "Party Ledger"
    if party_name: title += f" - {party_name}"
    ws.merge_cells('A1:G1'); ws['A1'] = title; ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    for col, h in enumerate(['Date', 'Party', 'Type', 'Description', 'Debit(₹)', 'Credit(₹)', 'Ref'], 1):
        c = ws.cell(row=3, column=col, value=h); c.fill = hf; c.font = hfont; c.border = tb
    for i, l in enumerate(data["ledger"], 4):
        for col, v in enumerate([l["date"], l["party_name"], l["party_type"], l["description"], l["debit"] if l["debit"] > 0 else "", l["credit"] if l["credit"] > 0 else "", l["ref"]], 1):
            c = ws.cell(row=i, column=col, value=v); c.border = tb
            if col in [5, 6]: c.number_format = '#,##0.00'
    row = len(data["ledger"]) + 4
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=5, value=data["total_debit"]).font = Font(bold=True)
    ws.cell(row=row, column=6, value=data["total_credit"]).font = Font(bold=True)
    for letter in ['A','B','C','D','E','F','G']: ws.column_dimensions[letter].width = 18
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=party_ledger_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@api_router.get("/reports/party-ledger/pdf")
async def export_party_ledger_pdf(party_name: Optional[str] = None, party_type: Optional[str] = None,
                                    kms_year: Optional[str] = None, season: Optional[str] = None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from io import BytesIO
    data = await report_party_ledger(party_name=party_name, party_type=party_type, kms_year=kms_year, season=season)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    elements = []; styles = getSampleStyleSheet()
    title = "Party Ledger"
    if party_name: title += f" - {party_name}"
    elements.append(Paragraph(title, styles['Title'])); elements.append(Spacer(1, 12))
    tdata = [['Date', 'Party', 'Type', 'Description', 'Debit(₹)', 'Credit(₹)', 'Ref']]
    for l in data["ledger"]:
        tdata.append([l["date"], l["party_name"][:18], l["party_type"], l["description"][:25],
            l["debit"] if l["debit"] > 0 else '-', l["credit"] if l["credit"] > 0 else '-', l["ref"]])
    tdata.append(['TOTAL', '', '', '', data["total_debit"], data["total_credit"], ''])
    table = RLTable(tdata, colWidths=[50, 80, 50, 120, 55, 55, 45], repeatRows=1)
    table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a365d')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),0.5,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),('ALIGN',(4,0),(5,-1),'RIGHT')]))
    elements.append(table); doc.build(elements); buffer.seek(0)
    return Response(content=buffer.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=party_ledger_{datetime.now().strftime('%Y%m%d')}.pdf"})


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
